"""파이프라인 시트 API — 한 장으로 보고하고 계획하는 영업 시트.

계정별로 그 주에 있었던 활동을 모아 보여준다(보고용). 위에서 아래로 읽으면
그대로 "저번주에 누구를 방문해서 이러이러했습니다"가 된다. 맨 위에는 이번 주
견적/납품 금액 합계를 실어, 그 주에 실제로 움직인 돈이 바로 보이게 한다.

기간 규칙: 이 회사의 한 주는 **월요일~금요일(5영업일)**이며, 주는 오직 그 주의
월요일 날짜로 식별한다(ISO 주차 개념을 쓰지 않는다). 이 규칙은 제거된 주간보고
기능에서 그대로 가져왔다.

계정 단위 집계(그룹핑·단계 판정·금액 산출)는 파이프라인 화면과 **같은 헬퍼**를
쓴다. 두 화면의 숫자가 갈라지면 보고 도구로서 신뢰를 잃기 때문이다.
"""

import json
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie

from reporting.models import DeliveryItem, History, Schedule
from reporting.funnel_views import (
    PIPELINE_STAGES,
    _pipeline_account_followup,
    _pipeline_account_groups,
    _pipeline_account_metadata,
    _pipeline_account_stage,
    _pipeline_contact_label,
    _pipeline_default_probability,
    _money_int,
    _select_pipeline_pricing,
    pipeline_followups_queryset,
)
from reporting.views import (
    _api_login_required_response,
    _dashboard_scope_users,
    _user_display_name,
    can_modify_user_data,
    get_user_profile,
)


# --------------------------------------------------------------------- 주 경계

WEEKDAY_LABELS = ['월', '화', '수', '목', '금', '토', '일']

STAGE_LABELS = {key: label for key, label, *_ in PIPELINE_STAGES}


def week_bounds(reference=None):
    """`reference`가 속한 주의 월~금 경계. 주 식별자는 월요일 날짜다."""
    today = reference or timezone.localdate()
    monday = today - timedelta(days=today.weekday())
    return monday, monday + timedelta(days=4)


def _parse_week_start(raw, today):
    """`?week=YYYY-MM-DD`를 그 주의 월요일로 정규화한다. 없으면 지난 주."""
    if raw:
        try:
            parsed = date.fromisoformat(str(raw).strip())
        except (TypeError, ValueError):
            parsed = None
        if parsed:
            return week_bounds(parsed)[0]
    # 기본값은 "저번 주" — 주간보고는 지난 주를 놓고 하는 것이라서.
    return week_bounds(today)[0] - timedelta(days=7)


def _week_options(today, count=8):
    """최근 `count`주 선택지(최신순)."""
    current_monday = week_bounds(today)[0]
    options = []
    for index in range(count):
        monday = current_monday - timedelta(days=7 * index)
        friday = monday + timedelta(days=4)
        if index == 0:
            suffix = ' (이번 주)'
        elif index == 1:
            suffix = ' (지난 주)'
        else:
            suffix = ''
        options.append({
            'value': monday.isoformat(),
            'label': f"{monday.strftime('%m/%d')} ~ {friday.strftime('%m/%d')}{suffix}",
            'weekStart': monday.isoformat(),
            'weekEnd': friday.isoformat(),
        })
    return options


# ------------------------------------------------------------------- 금액 해석
# 우선순위 체인은 실제 업무 규칙이다. 단순화하지 말 것.
#   납품품목 합계 -> History.delivery_amount -> Schedule.expected_revenue

def _decimal_or_zero(value):
    if value in (None, ''):
        return Decimal('0')
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _delivery_items_total(schedule):
    items = getattr(schedule, 'linked_delivery_items', []) or []
    return sum((_decimal_or_zero(i.total_price) for i in items), Decimal('0')), bool(items)


def _history_delivery_amount(schedule):
    histories = getattr(schedule, 'linked_histories', []) or []
    for history in histories:
        if history.action_type == 'delivery_schedule' and history.delivery_amount is not None:
            return _decimal_or_zero(history.delivery_amount), True
    for history in histories:
        if history.delivery_amount is not None:
            return _decimal_or_zero(history.delivery_amount), True
    return Decimal('0'), False


def _schedule_amount(schedule):
    item_total, has_items = _delivery_items_total(schedule)
    if has_items:
        return int(item_total)
    if schedule.activity_type == 'delivery':
        history_amount, found = _history_delivery_amount(schedule)
        if found:
            return int(history_amount)
    if schedule.expected_revenue is not None:
        return int(_decimal_or_zero(schedule.expected_revenue))
    return 0


# --------------------------------------------------------------- 활동 텍스트

def _activity_body(history):
    """활동 내용. `오늘 상황`이 있으면 그것, 없으면 `내용`.

    이 팀은 `연구원이 한 말`/`내가 확인한 사실` 필드를 쓰지 않으므로 참조하지
    않는다. 값이 없는 필드로 시트를 빈칸투성이로 만들지 않기 위함이다.
    """
    for value in (history.meeting_situation, history.content):
        text = (value or '').strip()
        if text:
            return text
    return ''


def _history_activity(history, viewer):
    when = history.meeting_date or timezone.localtime(history.created_at).date()
    return {
        'kind': 'history',
        'id': history.id,
        'date': when.isoformat(),
        'weekday': WEEKDAY_LABELS[when.weekday()],
        'type': history.get_action_type_display(),
        'body': _activity_body(history),
        'obstacle': (history.meeting_obstacles or '').strip(),
        'nextAction': (history.next_action or history.meeting_next_action or '').strip(),
        'nextActionDate': history.next_action_date.isoformat() if history.next_action_date else None,
        'amount': 0,
        'href': f'/notes/{history.id}/',
        'editable': can_modify_user_data(viewer, history.user),
    }


def _schedule_activity(schedule, viewer):
    when = schedule.visit_date
    return {
        'kind': 'schedule',
        'id': schedule.id,
        'date': when.isoformat(),
        'weekday': WEEKDAY_LABELS[when.weekday()],
        'type': schedule.get_activity_type_display(),
        'body': (schedule.notes or '').strip(),
        'obstacle': '',
        'nextAction': '',
        'nextActionDate': None,
        'amount': _schedule_amount(schedule),
        'href': f'/schedules/{schedule.id}/',
        # Schedule에는 장애물/다음액션 필드가 없다 — 메모(body)만 편집 가능하다.
        'editable': can_modify_user_data(viewer, schedule.user),
    }


# --------------------------------------------------------- 그리드 인라인 수정
# "그 자리에서" 고칠 수 있어야 시트가 보고서 초안이 아니라 진짜 보고 도구가 된다.

def _write_activity_body(history, text):
    """읽을 때의 우선순위(오늘 상황 → 내용)와 대칭이 되도록 같은 필드에 쓴다."""
    if (history.meeting_situation or '').strip():
        history.meeting_situation = text
    elif history.action_type == 'customer_meeting' and not (history.content or '').strip():
        history.meeting_situation = text
    else:
        history.content = text


@never_cache
@require_http_methods(["POST"])
def pipeline_sheet_activity_update_api(request, kind, activity_id):
    """그리드 셀 인라인 수정 — 영업노트 본문/장애물/다음액션/예정일, 일정 메모.

    화면이 쓰는 `_history_activity`/`_schedule_activity`를 그대로 재사용해
    응답을 만든다. 목록 재조회 없이도 행을 갱신할 수 있게 하기 위함이다.
    """
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    if kind not in ('history', 'schedule'):
        return JsonResponse({'success': False, 'error': '잘못된 요청입니다.'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': '잘못된 요청 형식입니다.'}, status=400)
    if not isinstance(payload, dict):
        return JsonResponse({'success': False, 'error': '잘못된 요청 형식입니다.'}, status=400)

    with transaction.atomic():
        if kind == 'history':
            activity = History.objects.select_for_update().filter(id=activity_id).first()
        else:
            activity = Schedule.objects.select_for_update().filter(id=activity_id).first()
        if not activity:
            return JsonResponse({'success': False, 'error': '기록을 찾을 수 없습니다.'}, status=404)

        if not can_modify_user_data(request.user, activity.user):
            return JsonResponse({'success': False, 'error': '본인의 기록만 수정할 수 있습니다.'}, status=403)

        if kind == 'history':
            if 'body' in payload:
                _write_activity_body(activity, (payload.get('body') or '').strip())
            if 'obstacle' in payload:
                activity.meeting_obstacles = (payload.get('obstacle') or '').strip()
            if 'nextAction' in payload:
                activity.next_action = (payload.get('nextAction') or '').strip()
            if 'nextActionDate' in payload:
                raw = (payload.get('nextActionDate') or '').strip()
                if raw:
                    try:
                        activity.next_action_date = date.fromisoformat(raw)
                    except ValueError:
                        return JsonResponse({'success': False, 'error': '날짜 형식이 올바르지 않습니다.'}, status=400)
                else:
                    activity.next_action_date = None
            activity.save()
            result = _history_activity(activity, request.user)
        else:
            if 'body' in payload:
                activity.notes = (payload.get('body') or '').strip()
                activity.save(update_fields=['notes'])
            result = _schedule_activity(activity, request.user)

    return JsonResponse({'success': True, 'activity': result})


# ------------------------------------------------------------ 계정 행 만들기

def _account_rows(request, today):
    """파이프라인 화면과 동일한 규칙으로 계정 단위 기본 행을 만든다."""
    followups = pipeline_followups_queryset(request, today)
    rows = []
    for group in _pipeline_account_groups(followups):
        stage = _pipeline_account_stage(group)
        fu = _pipeline_account_followup(group, stage)
        pricing = _select_pipeline_pricing(fu, stage)
        probability = pricing['probability']
        if probability is None:
            probability = _pipeline_default_probability(stage)
        amount = _money_int(pricing['amount'])
        metadata = _pipeline_account_metadata(fu)
        rows.append({
            **metadata,
            'company': str(fu.company) if fu.company else (fu.customer_name or '고객명 미정'),
            'department': str(fu.department) if fu.department else '',
            'contact': _pipeline_contact_label(fu),
            'owner': _user_display_name(fu.user),
            'ownerId': fu.user_id,
            'stage': stage,
            'stageLabel': STAGE_LABELS.get(stage, stage),
            'amount': amount,
            'probability': int(probability) if probability is not None else None,
            'weightedAmount': int(amount * (probability or 0) / 100) if amount else 0,
            'href': (
                f"/accounts/{metadata['accountId']}/"
                if metadata['accountType'] == 'department'
                else f"/customers/{metadata['accountId']}/"
            ),
            '_contactIds': metadata['contactIds'],
        })
    return rows


def _stage_definitions():
    return [
        {'id': key, 'label': label, 'color': color}
        for key, label, color, *_ in PIPELINE_STAGES
    ]


def _scope_payload(request, user_profile, scope_users, selected_user):
    can_select = bool(user_profile.can_view_all_users()) or user_profile.is_admin()
    return {
        'label': (
            _user_display_name(selected_user) if selected_user
            else (f'{user_profile.company.name} 팀' if user_profile.company else '전체')
        ),
        'canSelectUser': can_select,
        'selectedUserId': selected_user.id if selected_user else None,
        'users': [
            {'id': u.id, 'name': _user_display_name(u)}
            for u in scope_users.order_by('username')
        ] if can_select else [],
    }


# =============================================================== 주간 활동 집계

def _weekly_rows(request, week_start, week_end, base_rows):
    """계정별 행에 그 주의 활동을 붙인다."""
    contact_to_account = {}
    for row in base_rows:
        for contact_id in row['_contactIds']:
            contact_to_account[contact_id] = row['accountKey']

    contact_ids = list(contact_to_account.keys())
    buckets = {row['accountKey']: [] for row in base_rows}

    histories = History.objects.filter(
        followup_id__in=contact_ids,
        parent_history__isnull=True,
    ).filter(
        Q(meeting_date__gte=week_start, meeting_date__lte=week_end)
        | Q(meeting_date__isnull=True,
            created_at__date__gte=week_start,
            created_at__date__lte=week_end)
    ).select_related('followup').order_by('created_at')
    for history in histories:
        key = contact_to_account.get(history.followup_id)
        if key in buckets:
            buckets[key].append(_history_activity(history, request.user))

    schedules = Schedule.objects.filter(
        followup_id__in=contact_ids,
        visit_date__gte=week_start,
        visit_date__lte=week_end,
    ).exclude(status='cancelled').select_related('followup').prefetch_related(
        Prefetch('histories', queryset=History.objects.filter(
            parent_history__isnull=True,
        ).order_by('-created_at'), to_attr='linked_histories'),
        Prefetch('delivery_items_set', queryset=DeliveryItem.objects.order_by('id'),
                 to_attr='linked_delivery_items'),
    ).order_by('visit_date', 'visit_time')
    # 이번 주 견적/납품 금액 합계 — 요약 카드용. 손대지 못한 계정 대신 이걸 보여준다.
    quote_amount_total = 0
    delivery_amount_total = 0
    for schedule in schedules:
        key = contact_to_account.get(schedule.followup_id)
        if key not in buckets:
            continue
        activity = _schedule_activity(schedule, request.user)
        buckets[key].append(activity)
        if schedule.activity_type == 'quote':
            quote_amount_total += activity['amount']
        elif schedule.activity_type == 'delivery':
            delivery_amount_total += activity['amount']

    rows = []
    for row in base_rows:
        activities = sorted(buckets[row['accountKey']], key=lambda a: (a['date'], a['id']))
        next_actions = [a for a in activities if a['nextAction']]
        latest_next = next_actions[-1] if next_actions else None
        rows.append({
            **{k: v for k, v in row.items() if not k.startswith('_')},
            'activities': activities,
            'activityCount': len(activities),
            'weekAmount': sum(a['amount'] for a in activities),
            'nextAction': latest_next['nextAction'] if latest_next else '',
            'nextActionDate': latest_next['nextActionDate'] if latest_next else None,
            'hasObstacle': any(a['obstacle'] for a in activities),
        })
    return rows, quote_amount_total, delivery_amount_total


def _weekly_payload(request):
    user_profile = get_user_profile(request.user)
    scope_users, selected_user = _dashboard_scope_users(request, user_profile)
    today = timezone.localdate()
    week_start = _parse_week_start(request.GET.get('week'), today)
    week_end = week_start + timedelta(days=4)

    base_rows = _account_rows(request, today)
    rows, quote_amount_total, delivery_amount_total = _weekly_rows(request, week_start, week_end, base_rows)

    active = [r for r in rows if r['activityCount'] > 0]
    active.sort(key=lambda r: (-r['activityCount'], -r['amount'], r['company']))

    stage_totals = {key: {'count': 0, 'amount': 0} for key, *_ in PIPELINE_STAGES}
    for row in rows:
        bucket = stage_totals.get(row['stage'])
        if bucket is not None:
            bucket['count'] += 1
            bucket['amount'] += row['amount']

    return {
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'week': {
            'start': week_start.isoformat(),
            'end': week_end.isoformat(),
            'label': f"{week_start.strftime('%Y년 %m월 %d일')} 주",
            'isCurrent': week_start == week_bounds(today)[0],
        },
        'weekOptions': _week_options(today),
        'scope': _scope_payload(request, user_profile, scope_users, selected_user),
        'stages': _stage_definitions(),
        'stageTotals': stage_totals,
        'rows': active,
        'metrics': {
            'activeAccounts': len(active),
            'totalActivities': sum(r['activityCount'] for r in active),
            'quoteAmount': quote_amount_total,
            'deliveryAmount': delivery_amount_total,
        },
    }


# ===================================================================== 뷰

@never_cache
@ensure_csrf_cookie
@require_http_methods(["GET"])
def pipeline_sheet_weekly_api(request):
    """계정별 주간 활동."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response
    return JsonResponse(_weekly_payload(request))


# ================================================================ XLSX 내보내기

def _style_sheet(ws, widths, money_columns=(), rate_columns=()):
    """이 저장소의 다른 익스포트와 같은 서식을 적용한다."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(fill_type='solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = body_align
        for index in money_columns:
            row[index - 1].number_format = '#,##0'
        for index in rate_columns:
            row[index - 1].number_format = '0"%"'

    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _write_weekly_sheet(wb, weekly):
    """활동 1건이 한 행. 위에서 아래로 읽으면 그게 곧 보고다."""
    ws = wb.active
    ws.title = '주간 활동'
    ws.append([
        '날짜', '요일', '업체/학교', '부서/연구실', '담당자', '영업담당',
        '단계', '활동유형', '상황/내용', '장애물', '다음 액션', '예정일', '금액',
    ])
    for row in weekly['rows']:
        for activity in row['activities']:
            ws.append([
                activity['date'],
                activity['weekday'],
                row['company'],
                row['department'],
                row['contact'],
                row['owner'],
                row['stageLabel'],
                activity['type'],
                activity['body'],
                activity['obstacle'],
                activity['nextAction'],
                activity['nextActionDate'] or '',
                activity['amount'] or 0,
            ])
    _style_sheet(
        ws,
        widths=[12, 6, 22, 22, 18, 12, 12, 12, 52, 30, 30, 12, 14],
        money_columns=(13,),
    )


def _write_info_sheet(wb, request, weekly):
    ws = wb.create_sheet(title='다운로드 정보')
    week = weekly['week']
    wm = weekly['metrics']
    for row in [
        ('보고서', '파이프라인 시트'),
        ('주간 활동 기간', f"{week['start']} ~ {week['end']}"),
        ('범위', weekly['scope']['label']),
        ('활동 있는 계정', wm['activeAccounts']),
        ('총 활동 건수', wm['totalActivities']),
        ('이번 주 견적 금액', wm['quoteAmount']),
        ('이번 주 납품 금액', wm['deliveryAmount']),
        ('생성일시', timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')),
        ('생성자', _user_display_name(request.user)),
    ]:
        ws.append(row)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 42


@never_cache
@require_http_methods(["GET"])
def pipeline_sheet_export_api(request):
    """파이프라인 시트를 워크북 하나로 내려받는다.

    화면이 쓰는 payload 함수를 **그대로** 재사용한다. 엑셀과 화면의 숫자가
    갈라지면 보고 도구로서 신뢰를 잃기 때문이다.
    """
    from urllib.parse import quote as urlquote

    from django.http import HttpResponse
    from openpyxl import Workbook

    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    weekly = _weekly_payload(request)

    wb = Workbook()
    _write_weekly_sheet(wb, weekly)
    _write_info_sheet(wb, request, weekly)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"파이프라인시트_{weekly['week']['start']}_{weekly['week']['end']}.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{urlquote(filename)}"
    wb.save(response)
    return response
