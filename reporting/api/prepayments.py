"""Prepayment legacy pages, Excel exports, and JSON APIs.

This module was split out of ``reporting.views`` as part of the Django API
layer cleanup. A small transitional import of legacy view globals keeps shared
permission/helpers stable while the remaining backend service code is extracted.
"""

from reporting import views as _legacy_views

globals().update({
    name: getattr(_legacy_views, name)
    for name in dir(_legacy_views)
    if not name.startswith('__')
})

# ============================================
# 선결제 관리
# ============================================

@login_required
def prepayment_list_view(request):
    """선결제 목록 뷰 (같은 회사 직원 데이터 필터링)"""
    from reporting.models import Prepayment
    from django.db.models import Q, Sum
    from django.contrib.auth.models import User
    
    user_profile = get_user_profile(request.user)
    base_queryset = Prepayment.objects.select_related('department', 'department__company', 'customer', 'company', 'created_by')
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    requested_data_filter = request.GET.get('data_filter')
    data_filter = requested_data_filter or ('all' if user_profile and user_profile.is_manager() and user_profile.company else 'me')
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용)
    company_users = []
    if user_profile and user_profile.company:
        role_filter = ['salesman', 'manager'] if user_profile.is_manager() else ['salesman']
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role__in=role_filter,
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    is_viewing_others = False
    
    if data_filter == 'all' and user_profile and user_profile.company:
        # 같은 회사 전체
        role_filter = ['salesman', 'manager'] if user_profile.is_manager() else ['salesman']
        filter_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role__in=role_filter,
            is_active=True
        )
        base_queryset = base_queryset.filter(created_by__in=filter_users)
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        # 특정 직원 (같은 회사 확인)
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True
            )
            base_queryset = base_queryset.filter(created_by=selected_filter_user)
            is_viewing_others = True
        except User.DoesNotExist:
            base_queryset = base_queryset.filter(created_by=request.user)
            data_filter = 'me'
    else:
        # 'me' - 본인만
        base_queryset = base_queryset.filter(created_by=request.user)
    
    # 검색 필터
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        base_queryset = base_queryset.filter(
            Q(customer__customer_name__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(department__company__name__icontains=search_query) |
            Q(payer_name__icontains=search_query) |
            Q(memo__icontains=search_query)
        )
    
    if status_filter:
        base_queryset = base_queryset.filter(status=status_filter)
    
    # 정렬
    prepayments = base_queryset.order_by('-payment_date', '-created_at')
    
    # 통계
    stats = base_queryset.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance')
    )
    
    # 사용 금액 계산
    total_amount = stats['total_amount'] or 0
    total_balance = stats['total_balance'] or 0
    stats['total_used'] = total_amount - total_balance
    
    # 페이지네이션 추가
    from django.core.paginator import Paginator
    paginator = Paginator(prepayments, 30)  # 페이지당 30개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': '선결제 현황',
        'prepayments': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'stats': stats,
        # 새로운 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
        'can_create_prepayment': not user_profile.is_manager(),
    }
    
    return render(request, 'reporting/prepayment/list.html', context)


@login_required
def prepayment_create_view(request):
    """선결제 등록 뷰"""
    from reporting.models import Prepayment, FollowUp, Department
    from django import forms

    if get_user_profile(request.user).is_manager():
        messages.error(request, manager_core_readonly_message('선결제'))
        return redirect('reporting:prepayment_list')
    
    # Tailwind CSS 클래스
    input_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    select_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    textarea_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    
    class PrepaymentForm(forms.ModelForm):
        department = forms.ModelChoiceField(
            queryset=Department.objects.none(),
            label='계정/부서',
            widget=forms.Select(attrs={'class': select_class})
        )
        customer = forms.ModelChoiceField(
            queryset=FollowUp.objects.none(),
            label='담당자',
            widget=forms.Select(attrs={'class': select_class})
        )
        
        class Meta:
            model = Prepayment
            fields = ['department', 'customer', 'amount', 'payment_date', 'payment_method', 'payer_name', 'memo']
            widgets = {
                'amount': forms.NumberInput(attrs={'class': input_class, 'placeholder': '금액 입력'}),
                'payment_date': forms.DateInput(attrs={'class': input_class, 'type': 'date'}),
                'payment_method': forms.Select(attrs={'class': select_class}),
                'payer_name': forms.TextInput(attrs={'class': input_class, 'placeholder': '입금자명 (선택)'}),
                'memo': forms.Textarea(attrs={'class': textarea_class, 'rows': 3, 'placeholder': '메모 (선택)'}),
            }

        def clean(self):
            cleaned = super().clean()
            department = cleaned.get('department')
            customer = cleaned.get('customer')
            if department and customer and customer.department_id != department.id:
                raise forms.ValidationError('담당자는 선택한 계정에 속해야 합니다.')
            return cleaned

    def configure_form_querysets(form):
        user_profile = get_user_profile(request.user)
        if user_profile and user_profile.role != 'admin':
            same_company_users = UserProfile.objects.filter(
                company=user_profile.company
            ).values_list('user_id', flat=True)
            customer_queryset = FollowUp.objects.filter(user_id__in=same_company_users)
        else:
            customer_queryset = FollowUp.objects.all()
        form.fields['customer'].queryset = customer_queryset.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.filter(
            id__in=customer_queryset.values_list('department_id', flat=True)
        ).select_related('company')
    
    if request.method == 'POST':
        form = PrepaymentForm(request.POST)
        configure_form_querysets(form)
        if form.is_valid():
            prepayment = form.save(commit=False)
            prepayment.balance = prepayment.amount  # 초기 잔액 = 입금액
            prepayment.company = prepayment.department.company
            prepayment.created_by = request.user
            prepayment.save()
            _prepayment_create_ledger(
                prepayment,
                PrepaymentLedgerEntry.ENTRY_DEPOSIT,
                actor=request.user,
                amount=prepayment.amount,
                balance_before=0,
                balance_after=prepayment.balance,
                memo='Django 선결제 등록',
            )
            
            messages.success(request, f'{prepayment.department.name} 계정의 선결제 {prepayment.amount:,}원이 등록되었습니다.')
            return redirect('reporting:prepayment_detail', pk=prepayment.pk)
    else:
        # 한국 시간대의 오늘 날짜를 기본값으로 설정
        from django.utils import timezone
        import pytz
        korea_tz = pytz.timezone('Asia/Seoul')
        today_korea = timezone.now().astimezone(korea_tz).date()
        
        form = PrepaymentForm(initial={'payment_date': today_korea})
        configure_form_querysets(form)
    
    # 고객 목록 필터링 (회사별)
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.role != 'admin':
        # 같은 UserCompany 소속 사용자들이 등록한 고객만 표시
        from reporting.models import UserProfile
        same_company_users = UserProfile.objects.filter(
            company=user_profile.company
        ).values_list('user_id', flat=True)
        customer_queryset = FollowUp.objects.filter(user_id__in=same_company_users)
        form.fields['customer'].queryset = customer_queryset.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.filter(
            id__in=customer_queryset.values_list('department_id', flat=True)
        ).select_related('company')
    else:
        form.fields['customer'].queryset = FollowUp.objects.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.select_related('company')
    
    context = {
        'page_title': '선결제 등록',
        'form': form,
    }
    
    return render(request, 'reporting/prepayment/form.html', context)


@login_required
def prepayment_detail_view(request, pk):
    """선결제 상세 뷰"""
    from reporting.models import Prepayment, UserProfile
    
    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'department__company', 'customer', 'customer__department'),
        pk=pk,
    )
    
    # 권한 체크 - 같은 회사 사용자만 조회 가능
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.company:
        same_company_users = UserProfile.objects.filter(
            company=user_profile.company
        ).values_list('user_id', flat=True)
        
        if prepayment.created_by_id not in same_company_users:
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 본인 데이터 여부. Manager는 본인이 과거에 만든 데이터라도 핵심 선결제 수정은 불가합니다.
    is_owner = _prepayment_can_edit(request.user, prepayment)
    
    # 사용 내역
    usages = prepayment.usages.select_related(
        'schedule', 
        'schedule__followup'
    ).prefetch_related(
        'schedule__delivery_items_set'
    ).order_by('-used_at')
    
    # 각 usage에 delivery_items 첨부
    for usage in usages:
        if usage.schedule:
            usage.delivery_items = usage.schedule.delivery_items_set.all()
        else:
            usage.delivery_items = []
    
    # 금액 계산
    total_used = prepayment.amount - prepayment.balance
    usage_percent = 0
    if prepayment.amount > 0:
        from decimal import Decimal
        usage_percent = round(float(total_used / prepayment.amount) * 100, 1)
        balance_percent = round(float(prepayment.balance / prepayment.amount) * 100, 1)
    else:
        balance_percent = 0
    
    context = {
        'page_title': f'선결제 상세 - {(_prepayment_account_department(prepayment).name if _prepayment_account_department(prepayment) else prepayment.customer.customer_name)}',
        'prepayment': prepayment,
        'usages': usages,
        'total_used': total_used,
        'usage_percent': usage_percent,
        'balance_percent': balance_percent,
        'is_owner': is_owner,
    }
    
    return render(request, 'reporting/prepayment/detail.html', context)


@login_required
def prepayment_edit_view(request, pk):
    """선결제 수정 뷰"""
    from reporting.models import Prepayment, FollowUp, Department
    from django import forms
    
    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department'),
        pk=pk,
    )
    
    # 권한 체크 - 본인 데이터만 수정 가능. Manager는 핵심 선결제 수정 불가.
    if not _prepayment_can_edit(request.user, prepayment):
        messages.error(request, _prepayment_edit_denied_message(request.user, '수정'))
        return redirect('reporting:prepayment_list')
    
    # Tailwind CSS 클래스
    input_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    select_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    textarea_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    
    class PrepaymentEditForm(forms.ModelForm):
        department = forms.ModelChoiceField(
            queryset=Department.objects.none(),
            label='계정/부서',
            widget=forms.Select(attrs={'class': select_class})
        )
        customer = forms.ModelChoiceField(
            queryset=FollowUp.objects.none(),
            label='담당자',
            widget=forms.Select(attrs={'class': select_class})
        )
        
        class Meta:
            model = Prepayment
            fields = ['department', 'customer', 'amount', 'balance', 'payment_date', 'payment_method', 'payer_name', 'status', 'memo']
            widgets = {
                'amount': forms.NumberInput(attrs={'class': input_class, 'placeholder': '금액 입력'}),
                'balance': forms.NumberInput(attrs={'class': input_class, 'placeholder': '잔액'}),
                'payment_date': forms.DateInput(attrs={'class': input_class, 'type': 'date'}),
                'payment_method': forms.Select(attrs={'class': select_class}),
                'payer_name': forms.TextInput(attrs={'class': input_class, 'placeholder': '입금자명 (선택)'}),
                'status': forms.Select(attrs={'class': select_class}),
                'memo': forms.Textarea(attrs={'class': textarea_class, 'rows': 3, 'placeholder': '메모 (선택)'}),
            }

        def clean(self):
            cleaned = super().clean()
            department = cleaned.get('department')
            customer = cleaned.get('customer')
            if department and customer and customer.department_id != department.id:
                raise forms.ValidationError('담당자는 선택한 계정에 속해야 합니다.')
            return cleaned

    def configure_edit_form_querysets(form):
        user_profile = get_user_profile(request.user)
        if user_profile and user_profile.company:
            accessible_users = get_accessible_users(request.user, request)
            customer_queryset = FollowUp.objects.filter(user__in=accessible_users)
        else:
            customer_queryset = FollowUp.objects.all()
        form.fields['customer'].queryset = customer_queryset.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.filter(
            id__in=customer_queryset.values_list('department_id', flat=True)
        ).select_related('company')
    
    if request.method == 'POST':
        form = PrepaymentEditForm(request.POST, instance=prepayment)
        configure_edit_form_querysets(form)
        if form.is_valid():
            before_balance = prepayment.balance
            before_amount = prepayment.amount
            prepayment = form.save(commit=False)
            prepayment.company = prepayment.department.company
            prepayment.save()
            if before_balance != prepayment.balance or before_amount != prepayment.amount:
                _prepayment_create_ledger(
                    prepayment,
                    PrepaymentLedgerEntry.ENTRY_ADJUSTMENT,
                    actor=request.user,
                    amount=(prepayment.balance or 0) - (before_balance or 0),
                    balance_before=before_balance,
                    balance_after=prepayment.balance,
                    memo='Django 선결제 수정',
                )
            
            messages.success(request, '선결제 정보가 수정되었습니다.')
            return redirect('reporting:prepayment_detail', pk=prepayment.pk)
    else:
        form = PrepaymentEditForm(instance=prepayment)
        configure_edit_form_querysets(form)
    
    # 고객 목록 필터링 (회사별)
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.company:
        accessible_users = get_accessible_users(request.user, request)
        customer_queryset = FollowUp.objects.filter(user__in=accessible_users)
        form.fields['customer'].queryset = customer_queryset.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.filter(
            id__in=customer_queryset.values_list('department_id', flat=True)
        ).select_related('company')
    else:
        form.fields['customer'].queryset = FollowUp.objects.select_related('company', 'department')
        form.fields['department'].queryset = Department.objects.select_related('company')
    
    context = {
        'page_title': '선결제 수정',
        'form': form,
        'prepayment': prepayment,
    }
    
    return render(request, 'reporting/prepayment/form.html', context)


@login_required
@login_required
def prepayment_transfer_view(request, pk):
    """선결제 이관 뷰 - 같은 회사 내 다른 영업사원에게 이관"""
    from reporting.models import Prepayment, UserProfile
    from django.contrib.auth.models import User

    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department'),
        pk=pk,
    )

    # 본인 선결제만 이관 가능. Manager는 핵심 선결제 이관 불가.
    if not _prepayment_can_edit(request.user, prepayment):
        messages.error(request, _prepayment_edit_denied_message(request.user, '이관'))
        return redirect('reporting:prepayment_detail', pk=pk)

    # 같은 회사 내 다른 salesman 목록
    try:
        my_profile = UserProfile.objects.get(user=request.user)
        colleagues = UserProfile.objects.filter(
            company=my_profile.company,
            role='salesman'
        ).exclude(user=request.user).select_related('user')
    except UserProfile.DoesNotExist:
        colleagues = []

    if request.method == 'POST':
        target_user_id = request.POST.get('target_user')
        reason = request.POST.get('reason', '').strip()

        if not target_user_id:
            messages.error(request, '이관 대상을 선택해주세요.')
        else:
            target_user = get_object_or_404(User, pk=target_user_id)
            # 같은 회사 소속 확인
            if not UserProfile.objects.filter(user=target_user, company=my_profile.company).exists():
                messages.error(request, '같은 회사 소속이 아닙니다.')
            else:
                from_name = request.user.get_full_name() or request.user.username
                to_name = target_user.get_full_name() or target_user.username

                # 이관 메모 기록
                transfer_note = f"[이관] {from_name} → {to_name} ({prepayment.created_at.strftime('%Y-%m-%d')})"
                if reason:
                    transfer_note += f"\n사유: {reason}"

                prepayment.created_by = target_user
                if prepayment.memo:
                    prepayment.memo = prepayment.memo + '\n' + transfer_note
                else:
                    prepayment.memo = transfer_note
                prepayment.save()
                _prepayment_create_ledger(
                    prepayment,
                    PrepaymentLedgerEntry.ENTRY_TRANSFER,
                    actor=request.user,
                    target_user=target_user,
                    amount=0,
                    balance_before=prepayment.balance,
                    balance_after=prepayment.balance,
                    memo=transfer_note,
                )

                messages.success(request, f'선결제가 {to_name}님께 이관되었습니다.')
                return redirect('reporting:prepayment_detail', pk=pk)

    context = {
        'prepayment': prepayment,
        'colleagues': colleagues,
        'page_title': f'선결제 이관 - {(_prepayment_account_department(prepayment).name if _prepayment_account_department(prepayment) else prepayment.customer.customer_name)}',
    }
    return render(request, 'reporting/prepayment/transfer.html', context)


@login_required
def prepayment_delete_view(request, pk):
    """선결제 삭제 뷰"""
    from reporting.models import Prepayment
    from django.utils import timezone
    
    prepayment = get_object_or_404(Prepayment, pk=pk)
    
    # 권한 체크 - 본인 데이터만 삭제 가능. Manager는 핵심 선결제 삭제/취소 불가.
    if not _prepayment_can_edit(request.user, prepayment):
        messages.error(request, _prepayment_edit_denied_message(request.user, '삭제'))
        return redirect('reporting:prepayment_list')
    
    if request.method == 'POST':
        # 사용 내역 개수 확인
        usage_count = prepayment.usages.count()
        
        # 취소 요청인 경우
        if request.POST.get('action') == 'cancel':
            prepayment.status = 'cancelled'
            prepayment.cancelled_at = timezone.now()
            prepayment.cancel_reason = request.POST.get('cancel_reason', '사용자 요청으로 취소')
            prepayment.save()
            _prepayment_create_ledger(
                prepayment,
                PrepaymentLedgerEntry.ENTRY_CANCELLATION,
                actor=request.user,
                amount=prepayment.balance,
                balance_before=prepayment.balance,
                balance_after=prepayment.balance,
                memo=prepayment.cancel_reason,
            )
            messages.success(request, f'{prepayment.customer.customer_name}의 선결제가 취소되었습니다.')
            return redirect('reporting:prepayment_detail', pk=pk)
        
        # 삭제 요청인 경우
        if usage_count > 0:
            messages.error(request, f'이미 {usage_count}개의 사용 내역이 있는 선결제는 삭제할 수 없습니다.')
            return redirect('reporting:prepayment_detail', pk=pk)
        
        customer_name = prepayment.customer.customer_name
        _prepayment_create_ledger(
            prepayment,
            PrepaymentLedgerEntry.ENTRY_DELETION,
            actor=request.user,
            amount=prepayment.balance,
            balance_before=prepayment.balance,
            balance_after=prepayment.balance,
            memo='Django 선결제 삭제',
        )
        prepayment.delete()
        messages.success(request, f'{customer_name}의 선결제가 삭제되었습니다.')
        return redirect('reporting:prepayment_list')
    
    context = {
        'page_title': '선결제 삭제',
        'prepayment': prepayment,
    }
    
    return render(request, 'reporting/prepayment/delete_confirm.html', context)


@login_required
def prepayment_customer_view(request, customer_id):
    """부서별 선결제 관리 뷰 (부서 중심 - 동일 부서 내 모든 고객의 선결제 표시)"""
    from reporting.models import Prepayment, FollowUp, Department
    from django.db.models import Sum, Q, Count
    
    # 고객 정보 가져오기 (URL에서 전달받은 customer_id 기준)
    customer = get_object_or_404(FollowUp, pk=customer_id)
    
    # 부서 정보 가져오기
    department = customer.department
    company = customer.company
    
    # 권한 체크 - 고객의 담당자 또는 해당 고객에게 선결제를 등록한 사용자가 접근 가능
    user_profile = get_user_profile(request.user)
    
    # Admin과 Manager는 모든 고객에 접근 가능
    if not (user_profile.is_admin() or user_profile.is_manager()):
        # Salesman인 경우
        # 1. 고객의 담당자이거나
        # 2. 해당 고객에게 선결제를 등록한 적이 있는 경우 접근 가능
        is_customer_owner = (customer.user == request.user)
        has_prepayment = Prepayment.objects.filter(
            _prepayment_department_filter(customer.department) if customer.department_id else Q(customer=customer),
            created_by=request.user
        ).exists()
        
        if not (is_customer_owner or has_prepayment):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 현재 보고 있는 사용자 결정 (manager 기본값: 회사 전체)
    target_user = request.user
    target_users = User.objects.filter(id=request.user.id)
    
    if user_profile.can_view_all_users():
        user_filter = request.session.get('selected_user_id')
        if user_filter:
            accessible_users = get_accessible_users(request.user, request)
            try:
                target_user = accessible_users.get(id=user_filter)
                target_users = User.objects.filter(id=target_user.id)
            except User.DoesNotExist:
                target_user = request.user
                target_users = User.objects.filter(id=request.user.id)
        elif user_profile.is_manager() and user_profile.company:
            target_users = get_accessible_users(request.user, request).filter(is_active=True)
    
    # 부서 기준 조회: 동일 부서 내 모든 고객의 선결제 조회
    if department:
        # 부서 내 모든 고객(FollowUp) 조회
        department_followups = FollowUp.objects.filter(department=department).select_related('company', 'department')
        
        # 부서 내 모든 고객의 선결제 조회
        prepayments = Prepayment.objects.filter(
            _prepayment_department_filter(department),
            created_by__in=target_users
        ).select_related('department', 'company', 'customer', 'created_by').prefetch_related('usages').order_by('payment_date', 'id')
    else:
        # 부서 정보가 없는 경우 기존 고객 기준 조회
        department_followups = [customer]
        prepayments = Prepayment.objects.filter(
            customer=customer,
            created_by__in=target_users
        ).select_related('company', 'customer', 'created_by').prefetch_related('usages').order_by('payment_date', 'id')
    
    # 각 선결제의 사용금액 계산
    for prepayment in prepayments:
        prepayment.used_amount = prepayment.amount - prepayment.balance
    
    # 통계 계산
    stats = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance'),
        count=Count('id')
    )
    
    total_amount = stats['total_amount'] or 0
    total_balance = stats['total_balance'] or 0
    stats['total_used'] = total_amount - total_balance
    
    # 상태별 개수
    active_count = prepayments.filter(status='active').count()
    depleted_count = prepayments.filter(status='depleted').count()
    cancelled_count = prepayments.filter(status='cancelled').count()
    
    # 페이지 제목 구성
    if department and company:
        page_title = f'{company.name} - {department.name} 선결제 관리'
    elif company:
        page_title = f'{company.name} - 선결제 관리'
    else:
        page_title = f'{customer.customer_name} - 선결제 관리'
    
    context = {
        'page_title': page_title,
        'customer': customer,
        'company': company,
        'department': department,
        'department_followups': department_followups,
        'prepayments': prepayments,
        'stats': stats,
        'active_count': active_count,
        'depleted_count': depleted_count,
        'cancelled_count': cancelled_count,
    }
    
    return render(request, 'reporting/prepayment/customer.html', context)


@login_required
def prepayment_customer_excel(request, customer_id):
    """부서별 선결제 엑셀 다운로드 (부서 중심 - 동일 부서 내 모든 고객의 선결제 포함)"""
    from reporting.models import Prepayment, FollowUp, PrepaymentUsage, Department
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # 고객 정보 가져오기
    customer = get_object_or_404(FollowUp, pk=customer_id)
    
    # 부서 정보 가져오기
    department = customer.department
    company = customer.company
    
    # 권한 체크 - 고객의 담당자 또는 해당 고객에게 선결제를 등록한 사용자가 접근 가능
    user_profile = get_user_profile(request.user)
    
    # Admin과 Manager는 모든 고객에 접근 가능
    if not (user_profile.is_admin() or user_profile.is_manager()):
        # Salesman인 경우
        is_customer_owner = (customer.user == request.user)
        has_prepayment = Prepayment.objects.filter(
            _prepayment_department_filter(customer.department) if customer.department_id else Q(customer=customer),
            created_by=request.user
        ).exists()
        
        if not (is_customer_owner or has_prepayment):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 현재 보고 있는 사용자 결정 (manager 기본값: 회사 전체)
    target_user = request.user
    target_users = User.objects.filter(id=request.user.id)
    
    if user_profile.can_view_all_users():
        user_filter = request.session.get('selected_user_id')
        if user_filter:
            accessible_users = get_accessible_users(request.user, request)
            try:
                target_user = accessible_users.get(id=user_filter)
                target_users = User.objects.filter(id=target_user.id)
            except User.DoesNotExist:
                target_user = request.user
                target_users = User.objects.filter(id=request.user.id)
        elif user_profile.is_manager() and user_profile.company:
            target_users = get_accessible_users(request.user, request).filter(is_active=True)
    
    # 부서 기준 조회: 동일 부서 내 모든 고객의 선결제 조회
    if department:
        prepayments = Prepayment.objects.filter(
            _prepayment_department_filter(department),
            created_by__in=target_users
        ).select_related('department', 'company', 'customer', 'created_by').prefetch_related(
            'usages__schedule__delivery_items_set'
        ).order_by('payment_date', 'id')
    else:
        # 부서 정보가 없는 경우 기존 고객 기준 조회
        prepayments = Prepayment.objects.filter(
            customer=customer,
            created_by__in=target_users
        ).select_related('company', 'customer', 'created_by').prefetch_related(
            'usages__schedule__delivery_items_set'
        ).order_by('payment_date', 'id')
    
    # 엑셀 제목 구성
    if department and company:
        excel_title = f"{company.name} - {department.name} 선결제 요약"
        filename_prefix = f"{company.name}_{department.name}"
    elif company:
        excel_title = f"{company.name} 선결제 요약"
        filename_prefix = company.name
    else:
        excel_title = f"{customer.customer_name} 선결제 요약"
        filename_prefix = customer.customer_name
    
    # 엑셀 생성
    wb = Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    # ========================================
    # 시트 1: 선결제 요약
    # ========================================
    ws1 = wb.active
    ws1.title = "선결제 요약"
    
    # 제목 (부서 중심으로 변경)
    ws1.merge_cells('A1:I1')
    title_cell = ws1['A1']
    title_cell.value = excel_title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    ws1.row_dimensions[1].height = 30
    
    # 헤더 (고객명 컬럼 추가)
    headers1 = ['번호', '결제일', '고객명', '지불자', '결제방법', '선결제금액', '사용금액', '남은잔액', '상태']
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15   # 고객명
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 12
    
    # 데이터 행
    total_amount = 0
    total_used = 0
    total_balance = 0
    
    for idx, prepayment in enumerate(prepayments, 1):
        row = idx + 3
        used_amount = prepayment.amount - prepayment.balance
        
        total_amount += prepayment.amount
        total_used += used_amount
        total_balance += prepayment.balance
        
        # 데이터 (고객명 컬럼 추가)
        data = [
            idx,
            prepayment.payment_date.strftime('%Y-%m-%d'),
            prepayment.customer.customer_name if prepayment.customer else '-',  # 고객명 추가
            prepayment.payer_name or '-',
            prepayment.get_payment_method_display(),
            float(prepayment.amount),  # Decimal을 float으로 변환
            float(used_amount),
            float(prepayment.balance),
            prepayment.get_status_display(),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col_num)
            cell.value = value
            cell.border = border
            
            # 정렬 및 서식
            if col_num == 1 or col_num == 9:  # No, 상태
                cell.alignment = center_alignment
            elif col_num >= 6 and col_num <= 8:  # 금액 (컬럼 위치 조정)
                cell.alignment = right_alignment
                cell.number_format = '#,##0'
            
            # 상태별 배경색
            if col_num == 9:  # 상태 (컬럼 위치 조정)
                if prepayment.status == 'active':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif prepayment.status == 'depleted':
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
                elif prepayment.status == 'cancelled':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # 잔액에 따른 배경색
            if col_num == 8:  # 남은잔액 (컬럼 위치 조정)
                if prepayment.balance > 0:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    
    # 합계 행 (컬럼 위치 조정)
    summary_row = len(prepayments) + 4
    ws1.merge_cells(f'A{summary_row}:E{summary_row}')  # 5컬럼으로 확장
    summary_cell = ws1.cell(row=summary_row, column=1)
    summary_cell.value = "합계"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = center_alignment
    summary_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    summary_cell.border = border
    
    for col in range(2, 6):  # 2~5열까지
        ws1.cell(row=summary_row, column=col).border = border
        ws1.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 합계 금액 (컬럼 위치 조정: 6, 7, 8열)
    for col_num, value in [(6, float(total_amount)), (7, float(total_used)), (8, float(total_balance))]:
        cell = ws1.cell(row=summary_row, column=col_num)
        cell.value = value
        cell.font = Font(bold=True, size=11)
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
    
    ws1.cell(row=summary_row, column=9).border = border
    ws1.cell(row=summary_row, column=9).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ========================================
    # 시트 2: 품목별 집계
    # ========================================
    ws2 = wb.create_sheet(title="품목별 집계")
    
    # 제목
    ws2.merge_cells('A1:D1')
    title_cell2 = ws2['A1']
    title_cell2.value = "품목별 사용 집계"
    title_cell2.font = Font(bold=True, size=14)
    title_cell2.alignment = center_alignment
    ws2.row_dimensions[1].height = 30
    
    # 헤더
    headers2 = ['품목명', '총 수량', '단가', '총 사용금액']
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 18
    
    # 품목별 집계 데이터 수집
    # 중복 방지: 동일 schedule의 품목이 여러 선결제 usage에서 중복 집계되지 않도록 처리
    from collections import defaultdict
    item_stats = defaultdict(lambda: {'quantity': 0, 'amount': 0, 'count': 0, 'unit_prices': []})
    processed_schedules = set()  # 이미 처리한 schedule ID 추적
    
    for prepayment in prepayments:
        usages = prepayment.usages.all()
        for usage in usages:
            if usage.schedule and usage.schedule.id not in processed_schedules:
                processed_schedules.add(usage.schedule.id)  # 중복 방지
                delivery_items = usage.schedule.delivery_items_set.all()
                if delivery_items.exists():
                    for item in delivery_items:
                        # 금액 계산
                        item_amount = item.total_price if item.total_price else (float(item.quantity) * float(item.unit_price) * 1.1)
                        
                        # 통계 업데이트
                        item_stats[item.item_name]['quantity'] += float(item.quantity)
                        item_stats[item.item_name]['amount'] += float(item_amount)
                        item_stats[item.item_name]['count'] += 1
                        item_stats[item.item_name]['unit_prices'].append(float(item.unit_price))
    
    # 품목별 데이터 작성 (사용금액 기준 내림차순)
    sorted_items = sorted(item_stats.items(), key=lambda x: x[1]['amount'], reverse=True)
    row_num = 4
    total_summary_quantity = 0
    total_summary_amount = 0
    total_summary_unit_price = 0
    
    for item_name, stats in sorted_items:
        # 평균 단가 계산
        avg_unit_price = sum(stats['unit_prices']) / len(stats['unit_prices']) if stats['unit_prices'] else 0
        
        data = [
            item_name,
            stats['quantity'],
            avg_unit_price,
            stats['amount'],
        ]
        
        total_summary_quantity += stats['quantity']
        total_summary_amount += stats['amount']
        total_summary_unit_price += avg_unit_price
        
        for col_num, value in enumerate(data, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = right_alignment
                if col_num in [2, 3, 4]:  # 수량, 단가, 금액
                    cell.number_format = '#,##0'
        
        row_num += 1
    
    # 품목별 합계 행 추가
    if sorted_items:
        summary_row = row_num
        ws2.cell(row=summary_row, column=1).value = "합계"
        ws2.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=1).alignment = center_alignment
        ws2.cell(row=summary_row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=1).border = border
        
        ws2.cell(row=summary_row, column=2).value = total_summary_quantity
        ws2.cell(row=summary_row, column=2).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=2).alignment = right_alignment
        ws2.cell(row=summary_row, column=2).number_format = '#,##0'
        ws2.cell(row=summary_row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=2).border = border
        
        ws2.cell(row=summary_row, column=3).value = total_summary_unit_price
        ws2.cell(row=summary_row, column=3).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=3).alignment = right_alignment
        ws2.cell(row=summary_row, column=3).number_format = '#,##0'
        ws2.cell(row=summary_row, column=3).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=3).border = border
        
        ws2.cell(row=summary_row, column=4).value = total_summary_amount
        ws2.cell(row=summary_row, column=4).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=4).alignment = right_alignment
        ws2.cell(row=summary_row, column=4).number_format = '#,##0'
        ws2.cell(row=summary_row, column=4).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=4).border = border
        
        # 총 남은 선결제 잔액 정보 추가
        row_num += 2
        balance_row = row_num
        ws2.merge_cells(f'A{balance_row}:B{balance_row}')
        balance_label_cell = ws2.cell(row=balance_row, column=1)
        balance_label_cell.value = "총 남은 선결제 잔액"
        balance_label_cell.font = Font(bold=True, size=12, color="FFFFFF")
        balance_label_cell.alignment = center_alignment
        balance_label_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        balance_label_cell.border = border
        ws2.cell(row=balance_row, column=2).border = border
        ws2.cell(row=balance_row, column=2).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws2.merge_cells(f'C{balance_row}:D{balance_row}')
        balance_value_cell = ws2.cell(row=balance_row, column=3)
        balance_value_cell.value = float(total_balance)
        balance_value_cell.font = Font(bold=True, size=12)
        balance_value_cell.alignment = right_alignment
        balance_value_cell.number_format = '#,##0 "원"'
        balance_value_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        balance_value_cell.border = border
        ws2.cell(row=balance_row, column=4).border = border
        ws2.cell(row=balance_row, column=4).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # 품목 데이터가 없는 경우
    if not sorted_items:
        ws2.merge_cells('A4:D4')
        no_data_cell = ws2['A4']
        no_data_cell.value = "품목별 사용 내역이 없습니다."
        no_data_cell.alignment = center_alignment
        no_data_cell.font = Font(italic=True, color="999999")
    
    # HTTP 응답 (부서 중심 파일명 사용)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{filename_prefix}_선결제내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    
    wb.save(response)
    return response


@login_required
def prepayment_account_excel(request, department_id):
    """계정별 선결제 엑셀 다운로드."""
    department = get_object_or_404(Department.objects.select_related('company'), pk=department_id)
    representative = _prepayment_account_representative_for_request(request, department)
    if representative is None:
        messages.error(request, '접근 가능한 계정 선결제가 없습니다.')
        return redirect('reporting:prepayment_list')
    return prepayment_customer_excel(request, representative.id)


@login_required
def prepayment_list_excel(request):
    """전체 선결제 엑셀 다운로드"""
    from reporting.models import Prepayment
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    from django.db.models import Q
    
    # 권한 체크 및 데이터 필터링
    user_profile = get_user_profile(request.user)
    
    if user_profile.is_admin() or user_profile.is_manager():
        # Admin과 Manager는 모든 선결제 접근 가능
        prepayments = Prepayment.objects.all()
    else:
        # Salesman은 본인이 등록한 선결제 또는 본인이 등록한 고객의 선결제
        prepayments = Prepayment.objects.filter(
            Q(created_by=request.user) | Q(customer__user=request.user)
        )
    
    prepayments = prepayments.select_related(
        'department', 'department__company', 'customer', 'company', 'created_by'
    ).order_by('-payment_date', '-created_at')
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "전체 선결제"
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    # 제목
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"선결제 전체 내역 ({datetime.now().strftime('%Y-%m-%d')})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # 헤더
    headers = ['No', '계정', '담당자', '결제일', '지불자', '결제방법', '선결제금액', '사용금액', '남은잔액', '상태', '등록자', '등록일']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 16
    
    # 데이터 행
    total_amount = 0
    total_used = 0
    total_balance = 0
    
    for idx, prepayment in enumerate(prepayments, 1):
        row = idx + 3
        used_amount = prepayment.amount - prepayment.balance
        
        total_amount += prepayment.amount
        total_used += used_amount
        total_balance += prepayment.balance
        
        # 데이터
        department = _prepayment_account_department(prepayment)
        data = [
            idx,
            ' · '.join([value for value in [
                department.company.name if department and department.company else prepayment.company.name if prepayment.company else '',
                department.name if department else '',
            ] if value]) or '-',
            prepayment.customer.customer_name if prepayment.customer else '-',
            prepayment.payment_date.strftime('%Y-%m-%d'),
            prepayment.payer_name or '-',
            prepayment.get_payment_method_display(),
            prepayment.amount,
            used_amount,
            prepayment.balance,
            prepayment.get_status_display(),
            prepayment.created_by.get_full_name() or prepayment.created_by.username,
            prepayment.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = border
            
            # 정렬
            if col_num == 1 or col_num == 10:  # No, 상태
                cell.alignment = center_alignment
            elif col_num >= 7 and col_num <= 9:  # 금액
                cell.alignment = right_alignment
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            
            # 상태별 배경색
            if col_num == 10:
                if prepayment.status == 'active':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif prepayment.status == 'depleted':
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
                elif prepayment.status == 'cancelled':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # 잔액에 따른 배경색
            if col_num == 9:  # 남은잔액
                if prepayment.balance > 0:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    
    # 합계 행
    summary_row = len(prepayments) + 4
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "합계"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = center_alignment
    summary_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    summary_cell.border = border
    
    for col in range(2, 7):
        ws.cell(row=summary_row, column=col).border = border
        ws.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 합계 금액
    for col_num, value in [(7, total_amount), (8, total_used), (9, total_balance)]:
        cell = ws.cell(row=summary_row, column=col_num)
        cell.value = value
        cell.font = Font(bold=True, size=11)
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
    
    for col in range(10, 13):
        ws.cell(row=summary_row, column=col).border = border
        ws.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # HTTP 응답
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"선결제전체내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    
    wb.save(response)
    return response


def _prepayment_list_scope(request, user_profile):
    requested_data_filter = request.GET.get('data_filter')
    data_filter = requested_data_filter or ('all' if user_profile and user_profile.is_manager() and user_profile.company else 'me')
    filter_user_id = request.GET.get('filter_user') or request.GET.get('owner') or ''
    company_users = User.objects.none()

    if user_profile and user_profile.company:
        role_filter = ['salesman', 'manager'] if user_profile.is_manager() else ['salesman']
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role__in=role_filter,
            is_active=True,
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')

    selected_filter_user = None
    is_viewing_others = False

    if data_filter == 'all' and user_profile and user_profile.company:
        role_filter = ['salesman', 'manager'] if user_profile.is_manager() else ['salesman']
        users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role__in=role_filter,
            is_active=True,
        )
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True,
            )
            users = User.objects.filter(id=selected_filter_user.id)
            is_viewing_others = selected_filter_user.id != request.user.id
        except (User.DoesNotExist, ValueError):
            data_filter = 'me'
            filter_user_id = ''
            users = User.objects.filter(id=request.user.id)
    else:
        data_filter = 'me'
        filter_user_id = ''
        users = User.objects.filter(id=request.user.id)

    if selected_filter_user:
        label = _user_display_name(selected_filter_user)
    elif data_filter == 'all' and user_profile and user_profile.company:
        label = f'{user_profile.company.name} 영업팀'
    else:
        label = _user_display_name(request.user)

    return {
        'users': users,
        'data_filter': data_filter,
        'filter_user_id': str(filter_user_id or ''),
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
        'label': label,
    }


def _prepayment_account_department(prepayment):
    return prepayment_account_department(prepayment)


def _prepayment_account_company(prepayment):
    return prepayment_account_company(prepayment)


def _prepayment_department_filter(department):
    return prepayment_account_filter(department=department)


def _prepayment_create_ledger(
    prepayment,
    entry_type,
    *,
    actor=None,
    amount=0,
    balance_before=None,
    balance_after=None,
    memo='',
    schedule=None,
    usage=None,
    target_user=None,
    metadata=None,
):
    try:
        PrepaymentLedgerEntry.objects.create(
            prepayment=prepayment,
            department=_prepayment_account_department(prepayment),
            customer=prepayment.customer,
            schedule=schedule,
            usage=usage,
            entry_type=entry_type,
            amount=amount or 0,
            balance_before=balance_before,
            balance_after=balance_after,
            actor=actor,
            target_user=target_user,
            memo=memo or '',
            metadata=metadata or {},
        )
    except Exception:
        logger.exception('Failed to write prepayment ledger entry for prepayment_id=%s', getattr(prepayment, 'id', None))


def _prepayment_balance_row_payload(prepayment):
    item = _prepayment_item_payload(prepayment, prepayment.created_by)
    return {
        **item,
        'reactDetailHref': f'/prepayments/{prepayment.id}/',
        'accountHref': item['customerPrepaymentHref'],
    }


def _prepayment_usage_drilldown_payload(usage):
    return account_prepayment_usage_drilldown_payload(usage)


def _prepayment_ledger_payload(entry):
    prepayment = entry.prepayment
    schedule = entry.schedule
    return {
        'id': entry.id,
        'entryType': entry.entry_type,
        'entryTypeLabel': entry.get_entry_type_display(),
        'amount': _money_int(entry.amount),
        'balanceBefore': _money_int(entry.balance_before),
        'balanceAfter': _money_int(entry.balance_after),
        'memo': entry.memo or '',
        'createdAt': _datetime_or_none(entry.created_at),
        'actorName': _user_display_name(entry.actor) if entry.actor_id else '',
        'targetUserName': _user_display_name(entry.target_user) if entry.target_user_id else '',
        'prepaymentId': entry.prepayment_id,
        'prepaymentHref': f'/prepayments/{prepayment.id}/' if prepayment else '',
        'scheduleId': entry.schedule_id,
        'scheduleHref': f'/schedules/{schedule.id}/' if schedule else '',
        'djangoScheduleHref': reverse('reporting:schedule_detail', args=[schedule.id]) if schedule else '',
    }


def _prepayment_item_payload(prepayment, actor):
    return account_prepayment_item_payload(prepayment, actor)


def _prepayment_customer_payload(followup):
    return {
        'id': followup.id,
        'customerName': followup.customer_name or followup.manager or '고객명 미정',
        'companyId': followup.company_id,
        'companyName': followup.company.name if followup.company else '',
        'departmentId': followup.department_id,
        'departmentName': followup.department.name if followup.department else '',
        'ownerName': _user_display_name(followup.user),
        'label': ' · '.join([
            value
            for value in [
                followup.company.name if followup.company else '',
                followup.department.name if followup.department else '',
                followup.customer_name or followup.manager or '고객명 미정',
            ]
            if value
        ]),
    }


def _prepayment_customer_queryset_for_request(request):
    user_profile = get_user_profile(request.user)
    if user_profile.is_admin():
        return FollowUp.objects.all()
    elif user_profile.company:
        same_company_user_ids = UserProfile.objects.filter(
            company=user_profile.company,
        ).values_list('user_id', flat=True)
        return FollowUp.objects.filter(user_id__in=same_company_user_ids)
    return FollowUp.objects.filter(user=request.user)


def _prepayment_customer_options(request, current_customer=None):
    customer_queryset = _prepayment_customer_queryset_for_request(request)
    customers = list(
        customer_queryset.select_related('company', 'department', 'user')
        .order_by('company__name', 'department__name', 'customer_name')[:1000]
    )
    if current_customer and all(customer.id != current_customer.id for customer in customers):
        customers = [current_customer, *customers]
    return [_prepayment_customer_payload(customer) for customer in customers]


def _prepayment_account_options(request, current_department=None):
    customer_queryset = _prepayment_customer_queryset_for_request(request)
    departments = list(
        Department.objects.filter(
            id__in=customer_queryset.exclude(department__isnull=True).values_list('department_id', flat=True)
        ).select_related('company').order_by('company__name', 'name')[:1000]
    )
    if current_department and all(department.id != current_department.id for department in departments):
        departments = [current_department, *departments]

    customer_counts = {
        row['department_id']: row['count']
        for row in customer_queryset.exclude(department__isnull=True)
        .values('department_id')
        .annotate(count=Count('id'))
    }

    return [
        {
            'id': department.id,
            'companyId': department.company_id,
            'companyName': department.company.name if department.company else '',
            'departmentName': department.name,
            'name': department.name,
            'label': ' · '.join([
                value for value in [
                    department.company.name if department.company else '',
                    department.name,
                ] if value
            ]),
            'customerCount': customer_counts.get(department.id, 0),
        }
        for department in departments
    ]


def _prepayment_can_view(user, prepayment):
    return can_access_user_data(user, prepayment.created_by)


def _prepayment_can_create(user):
    return not get_user_profile(user).is_manager()


def _prepayment_can_edit(user, prepayment):
    if get_user_profile(user).is_manager():
        return False
    return prepayment.created_by_id == user.id


def _prepayment_edit_denied_message(user, action='수정'):
    if get_user_profile(user).is_manager():
        return manager_core_readonly_message('선결제')
    return f'본인이 등록한 선결제만 {action}할 수 있습니다.'


def _prepayment_form_options(request, prepayment=None):
    current_department = _prepayment_account_department(prepayment) if prepayment else None
    return {
        'accounts': _prepayment_account_options(
            request,
            current_department=current_department,
        ),
        'customers': _prepayment_customer_options(
            request,
            current_customer=prepayment.customer if prepayment else None,
        ),
        'paymentMethods': [
            {'value': value, 'label': label}
            for value, label in Prepayment.PAYMENT_METHOD_CHOICES
        ],
        'statuses': [
            {'value': value, 'label': label}
            for value, label in Prepayment.STATUS_CHOICES
        ],
    }


def _prepayment_transfer_options(request, prepayment):
    user_profile = get_user_profile(request.user)
    if not user_profile.company:
        return []

    profiles = UserProfile.objects.filter(
        company=user_profile.company,
        role='salesman',
        user__is_active=True,
    ).exclude(
        user_id=prepayment.created_by_id,
    ).select_related('user').order_by('user__first_name', 'user__username')

    return [
        {
            'id': profile.user_id,
            'name': _user_display_name(profile.user),
        }
        for profile in profiles
    ]


def _prepayment_usage_payload(usage):
    schedule = usage.schedule
    delivery_items = []
    if schedule:
        delivery_items = [
            {
                'id': item.id,
                'itemName': item.item_name,
                'quantity': item.quantity,
                'unit': item.unit or '',
                'unitPrice': _money_int(item.unit_price),
                'totalPrice': _money_int(item.total_price),
            }
            for item in schedule.delivery_items_set.all().order_by('id')
        ]

    return {
        'id': usage.id,
        'usedAt': _datetime_or_none(usage.used_at),
        'productName': usage.product_name or '',
        'quantity': int(usage.quantity or 0),
        'amount': _money_int(usage.amount),
        'remainingBalance': _money_int(usage.remaining_balance),
        'memo': usage.memo or '',
        'scheduleId': schedule.id if schedule else None,
        'scheduleDate': _date_or_none(schedule.visit_date) if schedule else None,
        'scheduleHref': f'/schedules/{schedule.id}/' if schedule else '',
        'djangoScheduleHref': reverse('reporting:schedule_detail', args=[schedule.id]) if schedule else '',
        'deliveryItems': delivery_items,
    }


def _prepayment_detail_payload(request, prepayment):
    usages = list(
        prepayment.usages.select_related(
            'schedule',
            'schedule__followup',
        ).prefetch_related(
            'schedule__delivery_items_set',
        ).order_by('-used_at')
    )
    ledger_entries = list(
        prepayment.ledger_entries.select_related('department', 'customer', 'schedule', 'actor', 'target_user')
        .order_by('-created_at', '-id')[:80]
    )
    amount = _money_int(prepayment.amount)
    balance = _money_int(prepayment.balance)
    total_used = max(amount - balance, 0)
    usage_percent = round((total_used / amount) * 100, 1) if amount > 0 else 0
    balance_percent = round((balance / amount) * 100, 1) if amount > 0 else 0
    can_manage = _prepayment_can_edit(request.user, prepayment)
    item_payload = _prepayment_item_payload(prepayment, request.user)
    usage_count = len(usages)
    delete_message = '' if usage_count == 0 else f'이미 {usage_count}개의 사용 내역이 있어 삭제할 수 없습니다.'

    return {
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'scope': {
            'label': _user_display_name(prepayment.created_by),
            'canManage': can_manage,
            'isOwner': can_manage,
        },
        'prepayment': {
            **item_payload,
            'memo': prepayment.memo or '',
            'usagePercent': usage_percent,
            'balancePercent': balance_percent,
        },
        'metrics': {
            'amount': amount,
            'balance': balance,
            'usedAmount': total_used,
            'usageCount': len(usages),
            'usagePercent': usage_percent,
            'balancePercent': balance_percent,
        },
        'links': {
            'prepayments': '/prepayments/',
            'reactDetail': f'/prepayments/{prepayment.id}/',
            'reactEdit': f'/prepayments/{prepayment.id}/edit/' if can_manage else '',
            'djangoList': reverse('reporting:prepayment_list'),
            'djangoDetail': reverse('reporting:prepayment_detail', args=[prepayment.id]),
            'djangoEdit': reverse('reporting:prepayment_edit', args=[prepayment.id]) if can_manage else '',
            'djangoDelete': reverse('reporting:prepayment_delete', args=[prepayment.id]) if can_manage else '',
            'djangoTransfer': reverse('reporting:prepayment_transfer', args=[prepayment.id]) if can_manage else '',
        },
        'actions': {
            'canCancel': bool(can_manage and prepayment.status != 'cancelled'),
            'cancelUrl': reverse('reporting:prepayment_cancel_api', args=[prepayment.id]) if can_manage else '',
            'canDelete': bool(can_manage and usage_count == 0),
            'deleteUrl': reverse('reporting:prepayment_delete_api', args=[prepayment.id]) if can_manage else '',
            'deleteMessage': delete_message,
            'canTransfer': bool(can_manage),
            'transferUrl': reverse('reporting:prepayment_transfer_api', args=[prepayment.id]) if can_manage else '',
            'transferUsers': _prepayment_transfer_options(request, prepayment) if can_manage else [],
        },
        'edit': {
            'canEdit': can_manage,
            'message': '' if can_manage else (
                manager_core_readonly_message('선결제')
                if get_user_profile(request.user).is_manager()
                else '본인이 등록한 선결제만 수정할 수 있습니다.'
            ),
            'submitUrl': reverse('reporting:prepayment_update_api', args=[prepayment.id]) if can_manage else '',
            'djangoUrl': reverse('reporting:prepayment_edit', args=[prepayment.id]) if can_manage else '',
            **_prepayment_form_options(request, prepayment=prepayment),
        },
        'usages': [
            _prepayment_usage_payload(usage)
            for usage in usages
        ],
        'ledgerEntries': [
            _prepayment_ledger_payload(entry)
            for entry in ledger_entries
        ],
    }


def _prepayment_create_payload(request):
    can_create = _prepayment_can_create(request.user)
    return {
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'create': {
            'canCreate': can_create,
            'message': '' if can_create else manager_core_readonly_message('선결제'),
            'submitUrl': reverse('reporting:prepayment_create_api') if can_create else '',
            'djangoUrl': reverse('reporting:prepayment_create') if can_create else '',
            **_prepayment_form_options(request),
        },
        'links': {
            'prepayments': '/prepayments/',
            'djangoList': reverse('reporting:prepayment_list'),
        },
    }


def _prepayment_customer_target_user(request, user_profile):
    target_user = request.user
    selected_user_id = request.GET.get('user') or request.GET.get('target_user') or request.session.get('selected_user_id')
    if user_profile.can_view_all_users() and selected_user_id:
        try:
            accessible_users = get_accessible_users(request.user, request)
            target_user = accessible_users.get(id=selected_user_id)
        except (User.DoesNotExist, ValueError, TypeError):
            target_user = request.user
    return target_user


def _prepayment_customer_target_scope(request, user_profile):
    selected_user_id = request.GET.get('user') or request.GET.get('target_user') or request.session.get('selected_user_id')
    if user_profile.can_view_all_users():
        accessible_users = get_accessible_users(request.user, request).filter(is_active=True)
        if selected_user_id:
            try:
                target_user = accessible_users.get(id=selected_user_id)
                return User.objects.filter(id=target_user.id), target_user, False
            except (User.DoesNotExist, ValueError, TypeError):
                pass
        if user_profile.is_manager() and user_profile.company:
            return accessible_users, None, True
    return User.objects.filter(id=request.user.id), request.user, False


def _prepayment_customer_access_allowed(request, customer):
    user_profile = get_user_profile(request.user)
    if user_profile.is_admin() or user_profile.is_manager():
        return True
    is_customer_owner = customer.user_id == request.user.id
    prepayment_filter = Q(customer=customer)
    if customer.department_id:
        prepayment_filter |= _prepayment_department_filter(customer.department)
    has_prepayment = Prepayment.objects.filter(
        prepayment_filter,
        created_by=request.user,
    ).exists()
    return bool(is_customer_owner or has_prepayment)


def _prepayment_customer_context_payload(request, customer):
    department = customer.department
    company = customer.company
    user_profile = get_user_profile(request.user)
    target_users, target_user, is_all_users = _prepayment_customer_target_scope(request, user_profile)

    if department:
        department_followups = list(
            FollowUp.objects.filter(department=department)
            .select_related('company', 'department', 'user')
            .order_by('customer_name', 'id')
        )
        prepayments = Prepayment.objects.filter(
            _prepayment_department_filter(department),
            created_by__in=target_users,
        )
        scope_name = ' · '.join([value for value in [
            company.name if company else '',
            department.name,
        ] if value])
        scope_mode = 'department'
    else:
        department_followups = [customer]
        prepayments = Prepayment.objects.filter(
            customer=customer,
            created_by__in=target_users,
        )
        scope_name = customer.customer_name or '고객명 미정'
        scope_mode = 'customer'

    prepayments = prepayments.select_related(
        'department',
        'department__company',
        'company',
        'customer',
        'customer__company',
        'customer__department',
        'created_by',
    ).annotate(
        usage_count=Count('usages', distinct=True),
    ).order_by('payment_date', 'id')

    stats = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance'),
        total_count=Count('id'),
        active_count=Count('id', filter=Q(status='active')),
        depleted_count=Count('id', filter=Q(status='depleted')),
        cancelled_count=Count('id', filter=Q(status='cancelled')),
    )
    total_amount = _money_int(stats['total_amount'])
    total_balance = _money_int(stats['total_balance'])
    rows = list(prepayments)
    prepayment_ids = [prepayment.id for prepayment in rows]

    usage_rows = list(
        PrepaymentUsage.objects.filter(prepayment_id__in=prepayment_ids)
        .select_related(
            'prepayment',
            'prepayment__department',
            'prepayment__customer',
            'prepayment__customer__department',
            'schedule',
            'schedule__followup',
        )
        .prefetch_related('schedule__delivery_items_set')
        .order_by('-used_at', '-id')[:120]
    )
    ledger_filter = Q(prepayment_id__in=prepayment_ids)
    if department:
        if is_all_users:
            ledger_filter |= Q(department=department, actor__in=target_users)
            ledger_filter |= Q(department=department, target_user__in=target_users)
        elif target_user:
            ledger_filter |= Q(department=department, actor=target_user)
            ledger_filter |= Q(department=department, target_user=target_user)
    ledger_rows = list(
        PrepaymentLedgerEntry.objects.filter(ledger_filter)
        .select_related('prepayment', 'department', 'customer', 'schedule', 'actor', 'target_user')
        .order_by('-created_at', '-id')[:160]
    )

    accessible_users = []
    if user_profile.can_view_all_users():
        accessible_users = [
            {
                'id': user.id,
                'name': _user_display_name(user),
            }
            for user in get_accessible_users(request.user, request).order_by('first_name', 'username')
        ]

    react_account = f'/prepayments/account/{department.id}/' if department else ''
    account_detail = f'/accounts/{department.id}/' if department else ''

    return {
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'scope': {
            'mode': scope_mode,
            'name': scope_name,
            'label': f'{scope_name} 선결제',
            'targetUserId': target_user.id if target_user else None,
            'targetUserName': _user_display_name(target_user) if target_user else '회사 전체',
            'canSelectUser': bool(user_profile.can_view_all_users()),
            'isAllUsers': is_all_users,
        },
        'customer': {
            'id': customer.id,
            'customerName': customer.customer_name or customer.manager or '고객명 미정',
            'companyId': company.id if company else None,
            'companyName': company.name if company else '',
            'departmentId': department.id if department else None,
            'departmentName': department.name if department else '',
            'ownerName': _user_display_name(customer.user),
            'href': f'/customers/{customer.id}/',
            'djangoHref': reverse('reporting:followup_detail', args=[customer.id]),
        },
        'departmentCustomers': [
            {
                'id': followup.id,
                'customerName': followup.customer_name or followup.manager or '고객명 미정',
                'ownerName': _user_display_name(followup.user),
                'href': f'/customers/{followup.id}/',
                'djangoHref': reverse('reporting:followup_detail', args=[followup.id]),
            }
            for followup in department_followups
        ],
        'metrics': {
            'totalAmount': total_amount,
            'totalBalance': total_balance,
            'totalUsed': max(total_amount - total_balance, 0),
            'totalCount': stats['total_count'] or 0,
            'activeCount': stats['active_count'] or 0,
            'depletedCount': stats['depleted_count'] or 0,
            'cancelledCount': stats['cancelled_count'] or 0,
            'deductionCount': len(usage_rows),
            'ledgerCount': len(ledger_rows),
        },
        'options': {
            'owners': accessible_users,
        },
        'links': {
            'prepayments': '/prepayments/',
            'reactAccount': react_account,
            'reactCustomer': f'/prepayments/customer/{customer.id}/',
            'djangoList': reverse('reporting:prepayment_list'),
            'djangoCustomer': reverse('reporting:prepayment_customer', args=[customer.id]),
            'djangoExcel': reverse('reporting:prepayment_customer_excel', args=[customer.id]),
            'accountExcel': reverse('reporting:prepayment_account_excel', args=[department.id]) if department else '',
            'accountDetail': account_detail,
            'customerDetail': f'/customers/{customer.id}/',
            'djangoCustomerDetail': reverse('reporting:followup_detail', args=[customer.id]),
        },
        'prepayments': [
            _prepayment_item_payload(prepayment, request.user)
            for prepayment in rows
        ],
        'balanceRows': [
            _prepayment_item_payload(prepayment, request.user)
            for prepayment in rows
        ],
        'deductionRows': [
            _prepayment_usage_drilldown_payload(usage)
            for usage in usage_rows
        ],
        'ledgerEntries': [
            _prepayment_ledger_payload(entry)
            for entry in ledger_rows
        ],
    }


def _prepayment_account_representative_for_request(request, department):
    user_profile = get_user_profile(request.user)
    scope_users, _selected_user = _dashboard_scope_users(request, user_profile)
    representative = account_representative_followup(department, scope_users)
    if representative is not None:
        return representative

    own_prepayment_customer_ids = Prepayment.objects.filter(
        _prepayment_department_filter(department),
        created_by__in=scope_users,
    ).values('customer_id')
    return (
        FollowUp.objects.filter(pk__in=own_prepayment_customer_ids)
        .select_related('company', 'department', 'user')
        .order_by('-updated_at', '-created_at', '-id')
        .first()
    )


def _prepayment_request_data(request):
    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except (TypeError, ValueError, UnicodeDecodeError):
            payload = {}
        return payload
    return request.POST


def _prepayment_field(data, *names):
    for name in names:
        value = data.get(name)
        if value not in (None, ''):
            return value
    return ''


def _prepayment_parse_form_data(request, *, existing=None):
    data = _prepayment_request_data(request)
    department_id = _prepayment_field(data, 'department', 'department_id', 'departmentId', 'account', 'account_id', 'accountId')
    customer_id = _prepayment_field(data, 'customer', 'customer_id', 'customerId')
    amount = _parse_optional_decimal(_prepayment_field(data, 'amount'))
    payment_date = _parse_iso_date_or_none(_prepayment_field(data, 'payment_date', 'paymentDate'))
    payment_method = str(_prepayment_field(data, 'payment_method', 'paymentMethod') or '').strip()
    payer_name = str(_prepayment_field(data, 'payer_name', 'payerName') or '').strip()[:100]
    memo = str(_prepayment_field(data, 'memo') or '').strip()

    current_customer = existing.customer if existing else None
    current_department = _prepayment_account_department(existing) if existing else None
    customer = None
    department = None

    if customer_id:
        try:
            customer = FollowUp.objects.select_related('company', 'department', 'user').get(id=customer_id)
        except (FollowUp.DoesNotExist, ValueError):
            raise ValueError('선택한 담당자를 찾을 수 없습니다.')
        department = customer.department

    if department_id:
        try:
            department = Department.objects.select_related('company').get(id=department_id)
        except (Department.DoesNotExist, ValueError):
            raise ValueError('선택한 계정을 찾을 수 없습니다.')

    if not department:
        raise ValueError('계정을 선택해주세요.')

    allowed_customer_ids = {
        option['id']
        for option in _prepayment_customer_options(request, current_customer=current_customer)
    }
    allowed_department_ids = {
        option['id']
        for option in _prepayment_account_options(request, current_department=current_department)
    }
    if department.id not in allowed_department_ids:
        raise PermissionError('접근 권한이 없는 계정입니다.')

    if customer is None and current_customer and current_customer.department_id == department.id:
        customer = current_customer

    if customer is None:
        customer = (
            _prepayment_customer_queryset_for_request(request)
            .filter(department=department)
            .select_related('company', 'department', 'user')
            .order_by('-is_active', '-updated_at', '-created_at', '-id')
            .first()
        )
        if customer is None:
            raise ValueError('계정에 연결된 담당자를 찾을 수 없습니다.')

    if customer.department_id != department.id:
        raise ValueError('담당자는 선택한 계정에 속해야 합니다.')
    if customer.id not in allowed_customer_ids or not can_access_followup(request.user, customer):
        raise PermissionError('접근 권한이 없는 담당자입니다.')
    if not department.company_id:
        raise ValueError('계정의 업체/학교 정보가 필요합니다.')
    if amount is None or amount <= 0:
        raise ValueError('선결제 금액은 1원 이상으로 입력해주세요.')
    if not payment_date:
        raise ValueError('입금 날짜를 선택해주세요.')
    if payment_method not in {value for value, _label in Prepayment.PAYMENT_METHOD_CHOICES}:
        raise ValueError('올바른 입금 방법을 선택해주세요.')

    cleaned = {
        'department': department,
        'customer': customer,
        'company': department.company,
        'amount': amount,
        'payment_date': payment_date,
        'payment_method': payment_method,
        'payer_name': payer_name,
        'memo': memo,
    }

    if existing:
        balance = _parse_optional_decimal(_prepayment_field(data, 'balance'))
        status = str(_prepayment_field(data, 'status') or '').strip()
        if balance is None:
            raise ValueError('잔액은 0원 이상으로 입력해주세요.')
        if balance > amount:
            raise ValueError('잔액은 선결제 금액보다 클 수 없습니다.')
        if status not in {value for value, _label in Prepayment.STATUS_CHOICES}:
            raise ValueError('올바른 상태를 선택해주세요.')
        cleaned['balance'] = balance
        cleaned['status'] = status

    return cleaned


def _prepayment_summary_payload(request):
    user_profile = get_user_profile(request.user)
    scope = _prepayment_list_scope(request, user_profile)
    search_query = (request.GET.get('search') or request.GET.get('q') or '').strip()
    status_filter = request.GET.get('status', '').strip()
    valid_statuses = {value for value, _label in Prepayment.STATUS_CHOICES}

    base_queryset = Prepayment.objects.select_related(
        'department',
        'department__company',
        'customer',
        'customer__company',
        'customer__department',
        'company',
        'created_by',
    ).filter(created_by__in=scope['users'])
    prepayments = base_queryset

    if search_query:
        prepayments = prepayments.filter(
            Q(customer__customer_name__icontains=search_query) |
            Q(customer__company__name__icontains=search_query) |
            Q(customer__department__name__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(department__company__name__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(payer_name__icontains=search_query) |
            Q(memo__icontains=search_query)
        )

    if status_filter in valid_statuses:
        prepayments = prepayments.filter(status=status_filter)
    else:
        status_filter = ''

    stats = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance'),
        active_count=Count('id', filter=Q(status='active')),
        depleted_count=Count('id', filter=Q(status='depleted')),
        cancelled_count=Count('id', filter=Q(status='cancelled')),
        total_count=Count('id'),
    )
    total_amount = _money_int(stats['total_amount'])
    total_balance = _money_int(stats['total_balance'])

    try:
        limit = int(request.GET.get('limit', 80))
    except (TypeError, ValueError):
        limit = 80
    limit = min(max(limit, 1), 200)

    ordered_prepayments = prepayments.annotate(
        usage_count=Count('usages', distinct=True),
    ).order_by('-payment_date', '-created_at')
    rows = list(ordered_prepayments[:limit])
    filtered_count = stats['total_count'] or 0

    query_string = request.GET.urlencode()
    django_list = reverse('reporting:prepayment_list')
    excel_href = reverse('reporting:prepayment_list_excel')
    if query_string:
        django_list = f'{django_list}?{query_string}'
        excel_href = f'{excel_href}?{query_string}'

    return {
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'scope': {
            'label': scope['label'],
            'dataFilter': scope['data_filter'],
            'filterUserId': int(scope['filter_user_id']) if scope['filter_user_id'] else None,
            'isViewingOthers': scope['is_viewing_others'],
            'canViewTeam': bool(user_profile and user_profile.company),
        },
        'filters': {
            'search': search_query,
            'status': status_filter,
            'dataFilter': scope['data_filter'],
            'filterUser': scope['filter_user_id'],
            'limit': limit,
        },
        'options': {
            'statuses': [
                {'value': value, 'label': label}
                for value, label in Prepayment.STATUS_CHOICES
            ],
            'owners': [
                {'id': user.id, 'name': _user_display_name(user)}
                for user in scope['company_users']
            ],
            'dataFilters': [
                {'value': 'me', 'label': '나'},
                {'value': 'all', 'label': '전체'},
                {'value': 'user', 'label': '직원 선택'},
            ],
        },
        'metrics': {
            'totalAmount': total_amount,
            'totalBalance': total_balance,
            'totalUsed': max(total_amount - total_balance, 0),
            'totalCount': filtered_count,
            'filteredPrepayments': filtered_count,
            'activeCount': stats['active_count'] or 0,
            'depletedCount': stats['depleted_count'] or 0,
            'cancelledCount': stats['cancelled_count'] or 0,
            'returnedCount': len(rows),
            'truncated': filtered_count > len(rows),
        },
        'links': {
            'djangoList': django_list,
            'create': reverse('reporting:prepayment_create') if _prepayment_can_create(request.user) and not scope['is_viewing_others'] else '',
            'excel': excel_href,
            'customers': '/customers/',
        },
        'prepayments': [
            _prepayment_item_payload(prepayment, request.user)
            for prepayment in rows
        ],
    }


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET"])
def prepayment_api_list(request):
    """선결제 API: 고객/일정 선택 목록 또는 React 선결제 현황 목록."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse(_prepayment_summary_payload(request))
    
    try:
        followup = FollowUp.objects.select_related('department', 'user').get(id=customer_id)
        if not can_access_followup(request.user, followup):
            return JsonResponse({'prepayments': [], 'error': '접근 권한이 없습니다.'}, status=403)

        schedule = None
        schedule_id = request.GET.get('schedule_id')
        if schedule_id:
            try:
                schedule_id = int(schedule_id)
            except (TypeError, ValueError):
                return JsonResponse({'prepayments': [], 'error': '일정 정보를 확인하세요.'}, status=400)
            schedule = Schedule.objects.filter(id=schedule_id).select_related('followup', 'user').first()
            if not schedule or not can_access_user_data(request.user, schedule.user):
                return JsonResponse({'prepayments': [], 'error': '접근 권한이 없습니다.'}, status=403)
            if schedule.followup_id != followup.id:
                schedule = None

        prepayments_data = _schedules_prepayment_options(request.user, followup, schedule)
        return JsonResponse({'prepayments': prepayments_data})
    except Exception as e:
        return JsonResponse({'prepayments': [], 'error': str(e)})


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET", "POST"])
def prepayment_create_api(request):
    """React 선결제 등록 화면용 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    if request.method == 'GET':
        return JsonResponse(_prepayment_create_payload(request))

    if not _prepayment_can_create(request.user):
        return JsonResponse({'success': False, 'error': manager_core_readonly_message('선결제')}, status=403)

    try:
        cleaned = _prepayment_parse_form_data(request)
    except PermissionError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=403)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    prepayment = Prepayment.objects.create(
        department=cleaned['department'],
        customer=cleaned['customer'],
        company=cleaned['company'],
        amount=cleaned['amount'],
        balance=cleaned['amount'],
        payment_date=cleaned['payment_date'],
        payment_method=cleaned['payment_method'],
        payer_name=cleaned['payer_name'],
        memo=cleaned['memo'],
        status='active',
        created_by=request.user,
    )
    _prepayment_create_ledger(
        prepayment,
        PrepaymentLedgerEntry.ENTRY_DEPOSIT,
        actor=request.user,
        amount=prepayment.amount,
        balance_before=0,
        balance_after=prepayment.balance,
        memo='선결제 등록',
    )

    return JsonResponse({
        'success': True,
        'message': '선결제를 등록했습니다.',
        'prepaymentId': prepayment.id,
        'href': f'/prepayments/{prepayment.id}/',
        'djangoHref': reverse('reporting:prepayment_detail', args=[prepayment.id]),
        'prepayment': _prepayment_item_payload(prepayment, request.user),
    }, status=201)


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET"])
def prepayment_customer_api(request, customer_id):
    """React 고객별/부서별 선결제 화면용 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    customer = get_object_or_404(
        FollowUp.objects.select_related('company', 'department', 'user'),
        pk=customer_id,
    )
    if not _prepayment_customer_access_allowed(request, customer):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)

    return JsonResponse(_prepayment_customer_context_payload(request, customer))


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET"])
def prepayment_account_api(request, department_id):
    """React 부서/연구실 계정 선결제 화면용 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    department = get_object_or_404(Department.objects.select_related('company'), pk=department_id)
    representative = _prepayment_account_representative_for_request(request, department)
    if representative is None:
        return JsonResponse({
            'success': False,
            'error': '접근 가능한 부서/연구실 계정 선결제가 없습니다.',
        }, status=403)

    return JsonResponse(_prepayment_customer_context_payload(request, representative))


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET"])
def prepayment_detail_api(request, pk):
    """React 선결제 상세 화면용 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    prepayment = get_object_or_404(
        Prepayment.objects.select_related(
            'department',
            'department__company',
            'customer',
            'customer__company',
            'customer__department',
            'company',
            'created_by',
        ).prefetch_related(
            'usages',
        ),
        pk=pk,
    )
    if not _prepayment_can_view(request.user, prepayment):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)

    return JsonResponse(_prepayment_detail_payload(request, prepayment))


@never_cache
@require_http_methods(["POST"])
def prepayment_update_api(request, pk):
    """React 선결제 수정 화면용 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department', 'company', 'created_by'),
        pk=pk,
    )
    if not _prepayment_can_view(request.user, prepayment):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    if not _prepayment_can_edit(request.user, prepayment):
        return JsonResponse({'success': False, 'error': _prepayment_edit_denied_message(request.user, '수정')}, status=403)

    try:
        cleaned = _prepayment_parse_form_data(request, existing=prepayment)
    except PermissionError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=403)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    before_amount = prepayment.amount
    before_balance = prepayment.balance
    before_department_id = prepayment.department_id
    before_customer_id = prepayment.customer_id
    before_status = prepayment.status

    prepayment.department = cleaned['department']
    prepayment.customer = cleaned['customer']
    prepayment.company = cleaned['company']
    prepayment.amount = cleaned['amount']
    prepayment.balance = cleaned['balance']
    prepayment.payment_date = cleaned['payment_date']
    prepayment.payment_method = cleaned['payment_method']
    prepayment.payer_name = cleaned['payer_name']
    prepayment.status = cleaned['status']
    prepayment.memo = cleaned['memo']
    if prepayment.status != 'cancelled':
        prepayment.cancelled_at = None
        prepayment.cancel_reason = ''
    prepayment.save(update_fields=[
        'customer',
        'department',
        'company',
        'amount',
        'balance',
        'payment_date',
        'payment_method',
        'payer_name',
        'status',
        'memo',
        'cancelled_at',
        'cancel_reason',
    ])
    if (
        before_amount != prepayment.amount
        or before_balance != prepayment.balance
        or before_department_id != prepayment.department_id
        or before_customer_id != prepayment.customer_id
        or before_status != prepayment.status
    ):
        _prepayment_create_ledger(
            prepayment,
            PrepaymentLedgerEntry.ENTRY_ADJUSTMENT,
            actor=request.user,
            amount=(prepayment.balance or 0) - (before_balance or 0),
            balance_before=before_balance,
            balance_after=prepayment.balance,
            memo='선결제 정보 수정',
            metadata={
                'before': {
                    'amount': _money_int(before_amount),
                    'balance': _money_int(before_balance),
                    'departmentId': before_department_id,
                    'customerId': before_customer_id,
                    'status': before_status,
                },
                'after': {
                    'amount': _money_int(prepayment.amount),
                    'balance': _money_int(prepayment.balance),
                    'departmentId': prepayment.department_id,
                    'customerId': prepayment.customer_id,
                    'status': prepayment.status,
                },
            },
        )

    return JsonResponse({
        'success': True,
        'message': '선결제 정보를 수정했습니다.',
        'prepaymentId': prepayment.id,
        'href': f'/prepayments/{prepayment.id}/',
        'djangoHref': reverse('reporting:prepayment_detail', args=[prepayment.id]),
        'prepayment': _prepayment_item_payload(prepayment, request.user),
    })


@never_cache
@require_http_methods(["POST"])
def prepayment_cancel_api(request, pk):
    """React 선결제 취소 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department', 'company', 'created_by'),
        pk=pk,
    )
    if not _prepayment_can_view(request.user, prepayment):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    if not _prepayment_can_edit(request.user, prepayment):
        return JsonResponse({'success': False, 'error': _prepayment_edit_denied_message(request.user, '취소')}, status=403)

    data = _prepayment_request_data(request)
    reason = str(_prepayment_field(data, 'cancel_reason', 'cancelReason', 'reason') or '사용자 요청으로 취소').strip()
    if not reason:
        reason = '사용자 요청으로 취소'

    prepayment.status = 'cancelled'
    prepayment.cancelled_at = timezone.now()
    prepayment.cancel_reason = reason[:1000]
    prepayment.save(update_fields=['status', 'cancelled_at', 'cancel_reason'])
    _prepayment_create_ledger(
        prepayment,
        PrepaymentLedgerEntry.ENTRY_CANCELLATION,
        actor=request.user,
        amount=prepayment.balance,
        balance_before=prepayment.balance,
        balance_after=prepayment.balance,
        memo=reason[:1000],
    )

    customer_name = prepayment.customer.customer_name if prepayment.customer else '고객'
    return JsonResponse({
        'success': True,
        'message': f'{customer_name}의 선결제가 취소되었습니다.',
        'prepaymentId': prepayment.id,
        'href': f'/prepayments/{prepayment.id}/',
        'djangoHref': reverse('reporting:prepayment_detail', args=[prepayment.id]),
        'prepayment': _prepayment_item_payload(prepayment, request.user),
    })


@never_cache
@require_http_methods(["POST"])
def prepayment_delete_api(request, pk):
    """React 선결제 삭제 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department', 'company', 'created_by'),
        pk=pk,
    )
    if not _prepayment_can_view(request.user, prepayment):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    if not _prepayment_can_edit(request.user, prepayment):
        return JsonResponse({'success': False, 'error': _prepayment_edit_denied_message(request.user, '삭제')}, status=403)

    usage_count = prepayment.usages.count()
    if usage_count > 0:
        return JsonResponse({
            'success': False,
            'error': f'이미 {usage_count}개의 사용 내역이 있는 선결제는 삭제할 수 없습니다.',
        }, status=400)

    customer_name = prepayment.customer.customer_name if prepayment.customer else '고객'
    _prepayment_create_ledger(
        prepayment,
        PrepaymentLedgerEntry.ENTRY_DELETION,
        actor=request.user,
        amount=prepayment.balance,
        balance_before=prepayment.balance,
        balance_after=prepayment.balance,
        memo='선결제 삭제',
    )
    prepayment.delete()
    return JsonResponse({
        'success': True,
        'message': f'{customer_name}의 선결제가 삭제되었습니다.',
        'href': '/prepayments/',
        'djangoHref': reverse('reporting:prepayment_list'),
    })


@never_cache
@require_http_methods(["POST"])
def prepayment_transfer_api(request, pk):
    """React 선결제 이관 API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    prepayment = get_object_or_404(
        Prepayment.objects.select_related('department', 'customer', 'customer__department', 'company', 'created_by'),
        pk=pk,
    )
    if not _prepayment_can_view(request.user, prepayment):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    if not _prepayment_can_edit(request.user, prepayment):
        return JsonResponse({'success': False, 'error': _prepayment_edit_denied_message(request.user, '이관')}, status=403)

    data = _prepayment_request_data(request)
    target_user_id = _prepayment_field(data, 'target_user', 'targetUser', 'targetUserId')
    reason = str(_prepayment_field(data, 'reason', 'transferReason') or '').strip()
    if not target_user_id:
        return JsonResponse({'success': False, 'error': '이관 대상을 선택해주세요.'}, status=400)

    try:
        target_user_id = int(target_user_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': '이관 대상 정보를 확인해주세요.'}, status=400)
    if target_user_id == request.user.id:
        return JsonResponse({'success': False, 'error': '본인에게는 이관할 수 없습니다.'}, status=400)

    user_profile = get_user_profile(request.user)
    if not user_profile.company:
        return JsonResponse({'success': False, 'error': '소속 회사 정보가 없어 이관할 수 없습니다.'}, status=400)

    target_user = User.objects.filter(
        pk=target_user_id,
        is_active=True,
        userprofile__company=user_profile.company,
        userprofile__role='salesman',
    ).first()
    if not target_user:
        return JsonResponse({'success': False, 'error': '같은 회사의 영업사원만 이관 대상으로 선택할 수 있습니다.'}, status=400)

    from_name = _user_display_name(request.user)
    to_name = _user_display_name(target_user)
    transfer_note = f"[이관] {from_name} → {to_name} ({prepayment.created_at.strftime('%Y-%m-%d')})"
    if reason:
        transfer_note += f"\n사유: {reason[:1000]}"

    prepayment.created_by = target_user
    prepayment.memo = f"{prepayment.memo}\n{transfer_note}" if prepayment.memo else transfer_note
    prepayment.save(update_fields=['created_by', 'memo'])
    _prepayment_create_ledger(
        prepayment,
        PrepaymentLedgerEntry.ENTRY_TRANSFER,
        actor=request.user,
        target_user=target_user,
        amount=0,
        balance_before=prepayment.balance,
        balance_after=prepayment.balance,
        memo=transfer_note,
    )

    return JsonResponse({
        'success': True,
        'message': f'선결제가 {to_name}님께 이관되었습니다.',
        'prepaymentId': prepayment.id,
        'href': f'/prepayments/{prepayment.id}/',
        'djangoHref': reverse('reporting:prepayment_detail', args=[prepayment.id]),
        'prepayment': _prepayment_item_payload(prepayment, request.user),
    })


