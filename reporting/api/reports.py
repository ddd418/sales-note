"""Reports and data-quality JSON APIs split out of reporting.views.

This module intentionally imports a few legacy helpers from reporting.views while the
large view module is being decomposed in phases. URL names and response payloads stay
unchanged.
"""

import html
import json
import re
import unicodedata
from datetime import date, timedelta
from urllib.parse import urlencode

from django.contrib.auth.models import User
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Count, Max, Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reporting.services.account_ledger import (
    account_key_for_followup,
    account_operational_ledgers_for_followups,
    account_representative_followup,
)
from reporting.models import (
    AccountCleanupAuditLog,
    AccountCleanupDecision,
    CustomerAsset,
    Department,
    FollowUp,
    History,
    Quote,
    Schedule,
    ServiceCase,
)
from reporting.views import (
    _analytics_api_date_range,
    _analytics_api_scope_users,
    _analytics_user_payload,
    _api_login_required_response,
    _customers_edit_config,
    _date_or_none,
    _user_display_name,
    can_access_followup,
    can_modify_user_data,
    get_user_profile,
)

REPORTS_DELIVERY_FILTERS = {'any', 'with', 'without'}
REPORTS_PREPAYMENT_BALANCE_FILTERS = {'any', 'with', 'without'}
REPORTS_EXPORT_SCOPES = {'filtered', 'all', 'deliveries', 'prepayment_balance', 'cleanup_candidates'}
REPORTS_SORT_OPTIONS = {'recent', 'quote_items'}


def _account_cleanup_request_payload(request):
    if request.body:
        try:
            return json.loads(request.body.decode('utf-8') or '{}'), None
        except (json.JSONDecodeError, UnicodeDecodeError):
            return {}, JsonResponse({
                'success': False,
                'error': 'invalid_json',
                'message': '요청 JSON 형식이 올바르지 않습니다.',
            }, status=400)
    return request.POST.dict(), None


def _account_cleanup_payload_text(payload, *keys):
    for key in keys:
        value = payload.get(key)
        if value not in (None, ''):
            return str(value).strip()
    return ''


def _account_cleanup_scope_users(request, user_profile):
    if request.user.is_superuser or (user_profile and user_profile.is_admin()):
        return User.objects.all()
    scope_users, _user_list, _selected_user = _analytics_api_scope_users(request, user_profile)
    return scope_users


def _account_cleanup_followup_ref(followup):
    return {
        'id': followup.id,
        'customerName': followup.customer_name or '',
        'manager': followup.manager or '',
        'email': followup.email or '',
        'status': followup.status,
        'companyId': followup.company_id,
        'companyName': followup.company.name if followup.company else '',
        'departmentId': followup.department_id,
        'departmentName': followup.department.name if followup.department else '',
        'href': f'/customers/{followup.id}/',
    }


def _reports_request_int(value):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _reports_filter_params(request):
    delivery_filter = (
        request.GET.get('delivery_filter')
        or request.GET.get('delivery')
        or 'any'
    ).strip()
    if delivery_filter not in REPORTS_DELIVERY_FILTERS:
        delivery_filter = 'any'

    prepayment_balance_filter = (
        request.GET.get('prepayment_balance_filter')
        or request.GET.get('prepayment_balance')
        or 'any'
    ).strip()
    if prepayment_balance_filter not in REPORTS_PREPAYMENT_BALANCE_FILTERS:
        prepayment_balance_filter = 'any'

    export_scope = (request.GET.get('export_scope') or 'filtered').strip()
    if export_scope not in REPORTS_EXPORT_SCOPES:
        export_scope = 'filtered'

    sort = (
        request.GET.get('sort')
        or request.GET.get('ordering')
        or 'recent'
    ).strip()
    sort_aliases = {
        'quoteItems': 'quote_items',
        'quote_item_count': 'quote_items',
        'has_quote_items': 'quote_items',
    }
    sort = sort_aliases.get(sort, sort)
    if sort not in REPORTS_SORT_OPTIONS:
        sort = 'recent'

    return {
        'query': (request.GET.get('q') or request.GET.get('query') or '').strip(),
        'companyId': _reports_request_int(request.GET.get('company_id') or request.GET.get('company')),
        'departmentId': _reports_request_int(request.GET.get('department_id') or request.GET.get('department')),
        'deliveryFilter': delivery_filter,
        'prepaymentBalanceFilter': prepayment_balance_filter,
        'exportScope': export_scope,
        'sort': sort,
    }


def _reports_product_match_account_filter(query, filter_users=None, date_from=None, date_to=None):
    if not query or filter_users is None:
        return None

    item_match = (
        Q(delivery_items_set__item_name__icontains=query)
        | Q(delivery_items_set__notes__icontains=query)
        | Q(delivery_items_set__option_description__icontains=query)
        | Q(delivery_items_set__product__product_code__icontains=query)
        | Q(delivery_items_set__product__description__icontains=query)
        | Q(delivery_items_set__product__specification__icontains=query)
    )
    schedule_rows = (
        Schedule.objects.filter(
            user__in=filter_users,
            followup__isnull=False,
            activity_type__in=['delivery', 'quote'],
            **_reports_date_filter_kwargs('visit_date', date_from, date_to),
        )
        .filter(item_match)
        .exclude(status='cancelled')
        .values('followup_id', 'followup__department_id')
        .distinct()
    )

    quote_item_match = (
        Q(items__product__product_code__icontains=query)
        | Q(items__product__description__icontains=query)
        | Q(items__product__specification__icontains=query)
        | Q(items__description__icontains=query)
    )
    quote_rows = (
        Quote.objects.filter(
            user__in=filter_users,
            followup__isnull=False,
            **_reports_date_filter_kwargs('quote_date', date_from, date_to),
        )
        .filter(quote_item_match)
        .values('followup_id', 'followup__department_id')
        .distinct()
    )

    followup_ids = set()
    department_ids = set()
    for row in list(schedule_rows) + list(quote_rows):
        followup_id = row.get('followup_id')
        department_id = row.get('followup__department_id')
        if followup_id:
            followup_ids.add(followup_id)
        if department_id:
            department_ids.add(department_id)

    if not followup_ids and not department_ids:
        return None

    account_filter = Q()
    if followup_ids:
        account_filter |= Q(id__in=followup_ids)
    if department_ids:
        account_filter |= Q(department_id__in=department_ids)
    return account_filter


def _reports_apply_account_filters(followups_qs, filters, filter_users=None, date_from=None, date_to=None):
    query = filters.get('query') or ''
    if query:
        account_filter = (
            Q(company__name__icontains=query)
            | Q(department__name__icontains=query)
            | Q(customer_name__icontains=query)
            | Q(manager__icontains=query)
            | Q(email__icontains=query)
            | Q(phone_number__icontains=query)
        )
        product_filter = _reports_product_match_account_filter(
            query,
            filter_users=filter_users,
            date_from=date_from,
            date_to=date_to,
        )
        if product_filter is not None:
            account_filter |= product_filter
        followups_qs = followups_qs.filter(
            account_filter
        ).distinct()
    company_id = filters.get('companyId')
    if company_id:
        followups_qs = followups_qs.filter(company_id=company_id)
    department_id = filters.get('departmentId')
    if department_id:
        followups_qs = followups_qs.filter(department_id=department_id)
    return followups_qs


def _reports_filter_options(followups_qs):
    company_rows = (
        followups_qs.exclude(company__isnull=True)
        .values('company_id', 'company__name')
        .distinct()
        .order_by('company__name')
    )
    department_rows = (
        followups_qs.exclude(department__isnull=True)
        .values('department_id', 'department__name', 'company_id', 'company__name')
        .distinct()
        .order_by('company__name', 'department__name')
    )
    return {
        'companies': [
            {
                'id': row['company_id'],
                'name': row['company__name'] or '',
            }
            for row in company_rows
        ],
        'departments': [
            {
                'id': row['department_id'],
                'name': row['department__name'] or '',
                'companyId': row['company_id'],
                'companyName': row['company__name'] or '',
            }
            for row in department_rows
        ],
    }


def _reports_date_filter_kwargs(field_name, date_from=None, date_to=None):
    filters = {}
    if date_from:
        filters[f'{field_name}__gte'] = date_from
    if date_to:
        filters[f'{field_name}__lte'] = date_to
    return filters


def _reports_previous_date_range(date_from, date_to):
    span_days = max((date_to - date_from).days + 1, 1)
    previous_to = date_from - timedelta(days=1)
    previous_from = previous_to - timedelta(days=span_days - 1)
    return previous_from, previous_to


def _reports_latest_date(current_value, candidate_value):
    candidate = _date_or_none(candidate_value)
    if not candidate:
        return current_value
    if not current_value or candidate > current_value:
        return candidate
    return current_value


def _reports_customer_row_base(followup):
    account_key = account_key_for_followup(followup)
    account_id = followup.department_id or followup.id
    account_href = f'/accounts/{followup.department_id}/' if followup.department_id else f'/customers/{followup.id}/'
    account_label = followup.department.name if followup.department else (followup.customer_name or followup.manager or '고객명 미정')
    contact_label = followup.customer_name or followup.manager or '담당자 미정'
    prepayment_href = (
        f'/prepayments/account/{followup.department_id}/'
        if followup.department_id else f'/prepayments/customer/{followup.id}/'
    )
    return {
        'accountKey': account_key,
        'id': account_id,
        'accountId': followup.department_id,
        'representativeCustomerId': followup.id,
        'customer': account_label,
        'company': followup.company.name if followup.company else '',
        'department': followup.department.name if followup.department else '',
        'manager': contact_label,
        'contactCount': 1,
        'contactPreview': [contact_label],
        'owner': _user_display_name(followup.user),
        'status': followup.status,
        'statusLabel': followup.get_status_display(),
        'priority': followup.priority,
        'priorityLabel': followup.get_priority_display(),
        'pipelineStage': followup.pipeline_stage,
        'pipelineStageLabel': followup.get_pipeline_stage_display(),
        'deliveryCount': 0,
        'deliveryAmount': 0,
        'normalDeliveryCount': 0,
        'normalDeliveryAmount': 0,
        'prepaymentDeliveryCount': 0,
        'prepaymentDeliveryAmount': 0,
        'prepaymentUsedAmount': 0,
        'lastDeliveryDate': None,
        'quoteCount': 0,
        'quoteAmount': 0,
        'quoteItemCount': 0,
        'lastQuoteDate': None,
        'serviceCount': 0,
        'openServiceCount': 0,
        'lastServiceDate': None,
        'prepaymentCount': 0,
        'prepaymentAmount': 0,
        'prepaymentBalance': 0,
        'prepaymentUsedTotal': 0,
        'lastPrepaymentDate': None,
        'lastActivityDate': None,
        'recentDeliveryItems': [],
        'recentQuoteItems': [],
        'cleanupCandidateCount': 0,
        'cleanupRiskLabel': '',
        'cleanupTypes': [],
        'links': {
            'account': account_href,
            'prepayments': prepayment_href,
            'customer': f'/customers/{followup.id}/',
        },
        'drilldown': {
            'contacts': [],
            'deliveries': [],
            'quotes': [],
            'prepayments': [],
            'services': [],
        },
        'href': account_href,
        'customerHref': f'/customers/{followup.id}/',
        'djangoHref': reverse('reporting:followup_detail', args=[followup.id]),
    }


def _reports_contact_drilldown_payload(followup):
    return {
        'id': followup.id,
        'name': followup.customer_name or followup.manager or '담당자 미정',
        'manager': followup.manager or '',
        'role': followup.contact_role,
        'roleLabel': followup.get_contact_role_display(),
        'email': followup.email or '',
        'phone': followup.phone_number or '',
        'ownerName': _user_display_name(followup.user),
        'href': f'/customers/{followup.id}/',
    }


def _reports_append_drilldown(row, section, item, limit=6):
    items = row.get('drilldown', {}).get(section)
    if items is None:
        return
    if len(items) < limit:
        items.append(item)


def _reports_record_item_summary(record):
    labels = []
    for item in record.get('items', [])[:3]:
        name = (
            item.get('itemName')
            or item.get('productCode')
            or item.get('productDescription')
            or ''
        )
        if not name:
            continue
        quantity = item.get('quantity') or 0
        unit = item.get('unit') or ''
        if quantity:
            labels.append(f'{name} {quantity}{unit}')
        else:
            labels.append(name)
    if not labels:
        notes = (record.get('notes') or '').strip().replace('\n', ' ')
        if notes:
            return notes[:80]
        return '품목 미기재'
    hidden_count = max((record.get('itemCount') or 0) - len(labels), 0)
    summary = ', '.join(labels)
    if hidden_count:
        summary += f' 외 {hidden_count}개'
    return summary


def _reports_quote_record_label(record):
    return record.get('quoteNumber') or ('견적 일정' if record.get('recordType') == 'quote_schedule' else '견적')


def _reports_customer_operations_payload(followups_qs, filter_users, date_from, date_to, actor):
    followups = list(
        followups_qs.select_related(
            'user', 'company', 'department',
        ).order_by('company__name', 'department__name', 'customer_name', 'id')
    )
    followup_ids = [followup.id for followup in followups]
    account_key_by_followup_id = {}
    rows_by_key = {}
    for followup in followups:
        account_key = account_key_for_followup(followup)
        account_key_by_followup_id[followup.id] = account_key
        contact_label = followup.customer_name or followup.manager or '담당자 미정'
        if account_key not in rows_by_key:
            rows_by_key[account_key] = _reports_customer_row_base(followup)
            _reports_append_drilldown(rows_by_key[account_key], 'contacts', _reports_contact_drilldown_payload(followup), limit=10)
            continue
        row = rows_by_key[account_key]
        row['contactCount'] += 1
        if contact_label and contact_label not in row['contactPreview'] and len(row['contactPreview']) < 4:
            row['contactPreview'].append(contact_label)
        _reports_append_drilldown(row, 'contacts', _reports_contact_drilldown_payload(followup), limit=10)

    if not followup_ids:
        return {
            'metrics': {
                'totalCustomers': 0,
                'customersWithDeliveries': 0,
                'deliveryCount': 0,
                'deliveryAmount': 0,
                'normalDeliveryCount': 0,
                'normalDeliveryAmount': 0,
                'prepaymentDeliveryCount': 0,
                'prepaymentDeliveryAmount': 0,
                'prepaymentUsedAmount': 0,
                'quoteCount': 0,
                'quoteAmount': 0,
                'quoteItemCount': 0,
                'serviceCount': 0,
                'openServiceCount': 0,
                'prepaymentCount': 0,
                'prepaymentAmount': 0,
                'prepaymentBalance': 0,
                'prepaymentUsedTotal': 0,
            },
            'rows': [],
        }

    account_ledgers = account_operational_ledgers_for_followups(
        followups,
        filter_users,
        date_from=date_from,
        date_to=date_to,
        actor=actor,
        record_limit=50,
    )
    for account_key, ledger in account_ledgers.items():
        row = rows_by_key.get(account_key)
        if not row:
            continue
        ledger_metrics = ledger.get('metrics') or {}
        for field in [
            'deliveryCount',
            'deliveryAmount',
            'normalDeliveryCount',
            'normalDeliveryAmount',
            'prepaymentDeliveryCount',
            'prepaymentDeliveryAmount',
            'prepaymentUsedAmount',
            'quoteCount',
            'quoteAmount',
            'prepaymentCount',
            'prepaymentAmount',
            'prepaymentBalance',
            'prepaymentUsedTotal',
        ]:
            row[field] = ledger_metrics.get(field) or 0
        for field in ['lastDeliveryDate', 'lastQuoteDate', 'lastPrepaymentDate', 'lastActivityDate']:
            row[field] = ledger_metrics.get(field)
        row['quoteItemCount'] = int(ledger_metrics.get('quoteItemCount') or 0)

        for record in (ledger.get('deliveryRecords') or [])[:5]:
            total_amount = record.get('totalAmount') or 0
            row['recentDeliveryItems'].append({
                'id': record.get('id'),
                'date': record.get('date'),
                'label': _reports_record_item_summary(record),
                'amount': total_amount,
                'paymentSource': record.get('paymentSource') or 'normal',
                'paymentSourceLabel': record.get('paymentSourceLabel') or '일반 납품',
                'paymentStatus': record.get('paymentStatus') or 'normal',
                'paymentStatusLabel': record.get('paymentStatusLabel') or record.get('paymentSourceLabel') or '일반 납품',
                'href': record.get('href') or '',
            })
        for record in (ledger.get('deliveryRecords') or [])[:6]:
            total_amount = record.get('totalAmount') or 0
            _reports_append_drilldown(row, 'deliveries', {
                'id': record.get('id'),
                'date': record.get('date'),
                'label': _reports_record_item_summary(record),
                'amount': total_amount,
                'customerName': record.get('customerName') or '',
                'ownerName': record.get('ownerName') or '',
                'paymentSource': record.get('paymentSource') or 'normal',
                'paymentSourceLabel': record.get('paymentSourceLabel') or '일반 납품',
                'paymentStatusLabel': record.get('paymentStatusLabel') or record.get('paymentSourceLabel') or '일반 납품',
                'href': record.get('href') or '',
            })
        for record in (ledger.get('quoteRecords') or [])[:5]:
            item_count = int(record.get('itemCount') or 0)
            row['recentQuoteItems'].append({
                'id': record.get('id'),
                'date': record.get('date'),
                'label': _reports_record_item_summary(record),
                'amount': record.get('totalAmount') or 0,
                'itemCount': item_count,
                'quoteLabel': _reports_quote_record_label(record),
                'statusLabel': record.get('statusLabel') or '',
                'source': record.get('source') or ('견적 일정' if record.get('recordType') == 'quote_schedule' else '견적'),
                'href': record.get('href') or '',
            })
        for record in (ledger.get('quoteRecords') or [])[:6]:
            _reports_append_drilldown(row, 'quotes', {
                'id': record.get('id'),
                'date': record.get('date'),
                'label': _reports_quote_record_label(record),
                'amount': record.get('totalAmount') or 0,
                'itemCount': int(record.get('itemCount') or 0),
                'customerName': record.get('customerName') or '',
                'ownerName': record.get('ownerName') or '',
                'statusLabel': record.get('statusLabel') or '',
                'href': record.get('href') or '',
            })
        for record in (ledger.get('prepaymentRecords') or [])[:6]:
            prepayment_id = record.get('id')
            _reports_append_drilldown(row, 'prepayments', {
                'id': prepayment_id,
                'date': record.get('paymentDate'),
                'label': record.get('payerName') or record.get('customerName') or '입금자 미정',
                'amount': record.get('amount') or 0,
                'balance': record.get('balance') or 0,
                'statusLabel': record.get('statusLabel') or '',
                'customerName': record.get('customerName') or '',
                'ownerName': record.get('ownerName') or '',
                'href': f'/prepayments/{prepayment_id}/' if prepayment_id else '',
            })

    open_service_statuses = ['received', 'in_progress', 'waiting']
    service_cases = list(
        ServiceCase.objects.filter(
            followup_id__in=followup_ids,
            received_date__gte=date_from,
            received_date__lte=date_to,
        ).filter(
            Q(created_by__in=filter_users) | Q(assigned_to__in=filter_users) | Q(asset__created_by__in=filter_users),
        ).select_related('followup', 'asset', 'created_by', 'assigned_to').distinct().order_by('-received_date', '-created_at')
    )
    for case in service_cases:
        row = rows_by_key.get(account_key_by_followup_id.get(case.followup_id))
        if not row:
            continue
        row['serviceCount'] += 1
        if case.status in open_service_statuses:
            row['openServiceCount'] += 1
        row['lastServiceDate'] = _reports_latest_date(row['lastServiceDate'], case.received_date)
        row['lastActivityDate'] = _reports_latest_date(row['lastActivityDate'], case.received_date)
        _reports_append_drilldown(row, 'services', {
            'id': case.id,
            'date': _date_or_none(case.received_date),
            'label': case.asset.asset_name if case.asset_id else case.get_case_type_display(),
            'statusLabel': case.get_status_display(),
            'customerName': case.followup.customer_name or case.followup.manager or '',
            'ownerName': _user_display_name(case.created_by) if case.created_by else '',
            'href': f'/assets/?asset={case.asset_id}' if case.asset_id else '',
        })

    service_schedules = list(
        Schedule.objects.filter(
            followup_id__in=followup_ids,
            user__in=filter_users,
            activity_type='service',
            visit_date__gte=date_from,
            visit_date__lte=date_to,
        ).select_related('followup', 'user').order_by('-visit_date', '-visit_time', '-id')
    )
    for schedule in service_schedules:
        row = rows_by_key.get(account_key_by_followup_id.get(schedule.followup_id))
        if not row:
            continue
        row['serviceCount'] += 1
        if schedule.status != 'completed':
            row['openServiceCount'] += 1
        row['lastServiceDate'] = _reports_latest_date(row['lastServiceDate'], schedule.visit_date)
        row['lastActivityDate'] = _reports_latest_date(row['lastActivityDate'], schedule.visit_date)
        _reports_append_drilldown(row, 'services', {
            'id': schedule.id,
            'date': _date_or_none(schedule.visit_date),
            'label': schedule.notes[:80] if schedule.notes else '서비스 일정',
            'statusLabel': schedule.get_status_display(),
            'customerName': schedule.followup.customer_name or schedule.followup.manager or '',
            'ownerName': _user_display_name(schedule.user),
            'href': f'/schedules/{schedule.id}/',
        })

    rows = sorted(
        rows_by_key.values(),
        key=lambda row: (
            row['lastActivityDate'] or row['lastDeliveryDate'] or row['lastQuoteDate'] or row['lastPrepaymentDate'] or '',
            row['company'],
            row['department'],
            row['customer'],
        ),
        reverse=True,
    )
    metrics = {
        'totalCustomers': len(rows),
        'customersWithDeliveries': sum(1 for row in rows if row['deliveryCount'] > 0),
        'deliveryCount': sum(row['deliveryCount'] for row in rows),
        'deliveryAmount': sum(row['deliveryAmount'] for row in rows),
        'normalDeliveryCount': sum(row['normalDeliveryCount'] for row in rows),
        'normalDeliveryAmount': sum(row['normalDeliveryAmount'] for row in rows),
        'prepaymentDeliveryCount': sum(row['prepaymentDeliveryCount'] for row in rows),
        'prepaymentDeliveryAmount': sum(row['prepaymentDeliveryAmount'] for row in rows),
        'prepaymentUsedAmount': sum(row['prepaymentUsedAmount'] for row in rows),
        'quoteCount': sum(row['quoteCount'] for row in rows),
        'quoteAmount': sum(row['quoteAmount'] for row in rows),
        'quoteItemCount': sum(row['quoteItemCount'] for row in rows),
        'serviceCount': sum(row['serviceCount'] for row in rows),
        'openServiceCount': sum(row['openServiceCount'] for row in rows),
        'prepaymentCount': sum(row['prepaymentCount'] for row in rows),
        'prepaymentAmount': sum(row['prepaymentAmount'] for row in rows),
        'prepaymentBalance': sum(row['prepaymentBalance'] for row in rows),
        'prepaymentUsedTotal': sum(row['prepaymentUsedTotal'] for row in rows),
    }
    return {
        'metrics': metrics,
        'rows': rows,
    }


def _reports_operation_metrics(rows):
    return {
        'totalCustomers': len(rows),
        'customersWithDeliveries': sum(1 for row in rows if row.get('deliveryCount', 0) > 0),
        'deliveryCount': sum(row.get('deliveryCount', 0) for row in rows),
        'deliveryAmount': sum(row.get('deliveryAmount', 0) for row in rows),
        'normalDeliveryCount': sum(row.get('normalDeliveryCount', 0) for row in rows),
        'normalDeliveryAmount': sum(row.get('normalDeliveryAmount', 0) for row in rows),
        'prepaymentDeliveryCount': sum(row.get('prepaymentDeliveryCount', 0) for row in rows),
        'prepaymentDeliveryAmount': sum(row.get('prepaymentDeliveryAmount', 0) for row in rows),
        'prepaymentUsedAmount': sum(row.get('prepaymentUsedAmount', 0) for row in rows),
        'quoteCount': sum(row.get('quoteCount', 0) for row in rows),
        'quoteAmount': sum(row.get('quoteAmount', 0) for row in rows),
        'quoteItemCount': sum(row.get('quoteItemCount', 0) for row in rows),
        'serviceCount': sum(row.get('serviceCount', 0) for row in rows),
        'openServiceCount': sum(row.get('openServiceCount', 0) for row in rows),
        'prepaymentCount': sum(row.get('prepaymentCount', 0) for row in rows),
        'prepaymentAmount': sum(row.get('prepaymentAmount', 0) for row in rows),
        'prepaymentBalance': sum(row.get('prepaymentBalance', 0) for row in rows),
        'prepaymentUsedTotal': sum(row.get('prepaymentUsedTotal', 0) for row in rows),
    }


def _reports_cleanup_marker_map(data_quality):
    markers = {'department': {}, 'contact': {}}

    def add_marker(scope, key, marker_type, label):
        if not key:
            return
        bucket = markers[scope].setdefault(key, {
            'count': 0,
            'types': [],
            'labels': [],
        })
        bucket['count'] += 1
        if marker_type not in bucket['types']:
            bucket['types'].append(marker_type)
        if label not in bucket['labels']:
            bucket['labels'].append(label)

    for group in data_quality.get('duplicateAccounts') or []:
        for department in group.get('departments') or []:
            add_marker('department', department.get('id'), 'duplicate_account', '계정명 유사')
        for department_id in group.get('departmentIds') or []:
            add_marker('department', department_id, 'duplicate_account', '계정명 유사')

    for group in data_quality.get('duplicateContacts') or []:
        for contact in group.get('contacts') or []:
            add_marker('contact', contact.get('id'), 'duplicate_contact', '담당자 중복')
            if contact.get('departmentId'):
                add_marker('department', contact.get('departmentId'), 'duplicate_contact', '담당자 중복')

    for contact in data_quality.get('contactsWithoutDepartment') or []:
        add_marker('contact', contact.get('id'), 'missing_department', '부서 미지정')

    for contact in data_quality.get('contactsWithoutCompany') or []:
        add_marker('contact', contact.get('id'), 'missing_company', '업체 미지정')
        if contact.get('departmentId'):
            add_marker('department', contact.get('departmentId'), 'missing_company', '업체 미지정')

    return markers


def _reports_attach_cleanup_markers(operations, cleanup_markers):
    for row in operations.get('rows') or []:
        marker = {'count': 0, 'types': [], 'labels': []}
        department_id = row.get('accountId')
        if department_id:
            marker = cleanup_markers.get('department', {}).get(department_id, marker)
        else:
            marker = cleanup_markers.get('contact', {}).get(row.get('representativeCustomerId'), marker)
        row['cleanupCandidateCount'] = marker.get('count') or 0
        row['cleanupTypes'] = marker.get('types') or []
        row['cleanupRiskLabel'] = '정리 후보' if marker.get('count') else ''
    return operations


def _reports_row_matches_filters(row, filters):
    delivery_filter = filters.get('deliveryFilter') or 'any'
    if delivery_filter == 'with' and row.get('deliveryCount', 0) <= 0:
        return False
    if delivery_filter == 'without' and row.get('deliveryCount', 0) > 0:
        return False

    balance_filter = filters.get('prepaymentBalanceFilter') or 'any'
    if balance_filter == 'with' and row.get('prepaymentBalance', 0) <= 0:
        return False
    if balance_filter == 'without' and row.get('prepaymentBalance', 0) > 0:
        return False
    return True


def _reports_desc_date_key(value):
    try:
        return -date.fromisoformat(str(value or '')[:10]).toordinal()
    except (TypeError, ValueError):
        return 0


def _reports_sort_rows(rows, sort):
    rows = list(rows)
    if sort == 'quote_items':
        return sorted(
            rows,
            key=lambda row: (
                0 if int(row.get('quoteItemCount') or 0) > 0 else 1,
                -int(row.get('quoteItemCount') or 0),
                _reports_desc_date_key(row.get('lastQuoteDate')),
                _reports_desc_date_key(row.get('lastActivityDate')),
                row.get('company') or '',
                row.get('department') or '',
                row.get('customer') or '',
            ),
        )
    return rows


def _reports_apply_row_filters(operations, filters, export_scope='filtered'):
    rows = list(operations.get('rows') or [])
    if export_scope == 'all':
        filtered_rows = rows
    elif export_scope == 'deliveries':
        filtered_rows = [row for row in rows if row.get('deliveryCount', 0) > 0]
    elif export_scope == 'prepayment_balance':
        filtered_rows = [row for row in rows if row.get('prepaymentBalance', 0) > 0]
    elif export_scope == 'cleanup_candidates':
        filtered_rows = [row for row in rows if row.get('cleanupCandidateCount', 0) > 0]
    else:
        filtered_rows = [row for row in rows if _reports_row_matches_filters(row, filters)]
    filtered_rows = _reports_sort_rows(filtered_rows, filters.get('sort') or 'recent')
    return {
        **operations,
        'metrics': _reports_operation_metrics(filtered_rows),
        'rows': filtered_rows,
    }


def _reports_operations_comparison(current_metrics, previous_metrics, previous_from, previous_to):
    metric_keys = [
        'totalCustomers',
        'customersWithDeliveries',
        'deliveryCount',
        'deliveryAmount',
        'normalDeliveryCount',
        'normalDeliveryAmount',
        'prepaymentDeliveryCount',
        'prepaymentUsedAmount',
        'quoteCount',
        'quoteAmount',
        'quoteItemCount',
        'serviceCount',
        'openServiceCount',
        'prepaymentCount',
        'prepaymentAmount',
        'prepaymentBalance',
        'prepaymentUsedTotal',
    ]
    deltas = {}
    change_rates = {}
    for key in metric_keys:
        current_value = current_metrics.get(key, 0) or 0
        previous_value = previous_metrics.get(key, 0) or 0
        deltas[key] = current_value - previous_value
        change_rates[key] = round((current_value - previous_value) / previous_value * 100, 1) if previous_value else None
    return {
        'dateFrom': previous_from.isoformat(),
        'dateTo': previous_to.isoformat(),
        'metrics': {key: previous_metrics.get(key, 0) or 0 for key in metric_keys},
        'deltas': deltas,
        'changeRates': change_rates,
    }


_REPORTS_CLEANUP_BRACKET_RE = re.compile(r'[\(\[\{（【〈《].*?[\)\]\}）】〉》]')
_REPORTS_CLEANUP_ALIAS_RULES = [
    (re.compile(r'\bseoul\s*national\s*(?:univ|university)\b|\bsnu\b|서울\s*대(?:학교)?', re.IGNORECASE), '서울대학교'),
    (re.compile(r'\b서울\s*대\s*병원\b|\bsnuh\b', re.IGNORECASE), '서울대학교병원'),
    (re.compile(r'\byonsei\s*(?:univ|university)?\b|연세\s*대(?:학교)?|세브란스', re.IGNORECASE), '연세대학교'),
    (re.compile(r'\bkorea\s*(?:univ|university)?\b|고려\s*대(?:학교)?|고대', re.IGNORECASE), '고려대학교'),
    (re.compile(r'\bkaist\b|카이스트|한국\s*과학\s*기술원', re.IGNORECASE), '한국과학기술원'),
    (re.compile(r'\bpostech\b|포항\s*공대|포항\s*공과\s*대(?:학교)?', re.IGNORECASE), '포항공과대학교'),
    (re.compile(r'\bkist\b|한국\s*과학\s*기술\s*연구원', re.IGNORECASE), '한국과학기술연구원'),
    (re.compile(r'성균관\s*대(?:학교)?|성대', re.IGNORECASE), '성균관대학교'),
    (re.compile(r'한양\s*대(?:학교)?', re.IGNORECASE), '한양대학교'),
    (re.compile(r'중앙\s*대(?:학교)?', re.IGNORECASE), '중앙대학교'),
    (re.compile(r'경희\s*대(?:학교)?', re.IGNORECASE), '경희대학교'),
    (re.compile(r'부산\s*대(?:학교)?', re.IGNORECASE), '부산대학교'),
    (re.compile(r'전남\s*대(?:학교)?', re.IGNORECASE), '전남대학교'),
    (re.compile(r'전북\s*대(?:학교)?', re.IGNORECASE), '전북대학교'),
    (re.compile(r'충남\s*대(?:학교)?', re.IGNORECASE), '충남대학교'),
    (re.compile(r'충북\s*대(?:학교)?', re.IGNORECASE), '충북대학교'),
    (re.compile(r'가톨릭\s*대(?:학교)?', re.IGNORECASE), '가톨릭대학교'),
]
_REPORTS_CLEANUP_LAB_SUFFIX_RE = re.compile(
    r'(?:연구\s*실|실험\s*실|랩|lab(?:oratory)?|labo)\s*$',
    re.IGNORECASE,
)
_REPORTS_CLEANUP_LAB_WORD_RE = re.compile(
    r'(?:연구\s*실|실험\s*실|랩|lab(?:oratory)?|labo)',
    re.IGNORECASE,
)
_REPORTS_UNASSIGNED_COMPANY_KEYS = {
    '', '미지정', '업체미지정', '회사미지정', '고객사미지정', '무소속', 'unknown', 'none',
}
_REPORTS_UNASSIGNED_DEPARTMENT_KEYS = {
    '', '미지정', '부서미지정', '연구실미지정', '계정미지정', '미정', 'unknown', 'none',
}


def _reports_cleanup_compact_key(value):
    return re.sub(r'[\s\-_./,()（）\[\]{}<>:;·ㆍ]+', '', str(value or '').strip().lower())


def _reports_cleanup_key(value):
    text = html.unescape(str(value or ''))
    text = unicodedata.normalize('NFKC', text).strip().lower()
    text = _REPORTS_CLEANUP_BRACKET_RE.sub(' ', text)
    text = re.sub(r'[\"\'`“”‘’「」『』]', ' ', text)
    for pattern, replacement in _REPORTS_CLEANUP_ALIAS_RULES:
        text = pattern.sub(replacement, text)
    text = _REPORTS_CLEANUP_LAB_WORD_RE.sub('연구실', text)
    text = _REPORTS_CLEANUP_LAB_SUFFIX_RE.sub('', text)
    text = re.sub(r'\b(?:co|corp|corporation|inc|ltd)\b', ' ', text, flags=re.IGNORECASE)
    return _reports_cleanup_compact_key(text)


def _reports_is_unassigned_company(company):
    return company is None or _reports_cleanup_compact_key(company.name) in _REPORTS_UNASSIGNED_COMPANY_KEYS


def _reports_is_unassigned_department(department):
    return department is None or _reports_cleanup_compact_key(department.name) in _REPORTS_UNASSIGNED_DEPARTMENT_KEYS


def _reports_cleanup_candidate_key(candidate_type, *ids):
    stable_ids = [str(item) for item in ids if item not in (None, '')]
    return f'{candidate_type}:{"-".join(stable_ids)}'


def _reports_cleanup_candidate_base(candidate_type, candidate_key, **refs):
    return {
        'candidateType': candidate_type,
        'candidateKey': candidate_key,
        'decisionUrl': reverse('reporting:account_cleanup_decision_api'),
        'reviewStatus': 'new',
        'reviewStatusLabel': '신규',
        'decisionReason': '',
        'decisionUpdatedAt': None,
        'decisionUpdatedBy': '',
        **refs,
    }


def _reports_cleanup_decision_payload(decision):
    return {
        'reviewStatus': decision.decision,
        'reviewStatusLabel': decision.get_decision_display(),
        'decisionReason': decision.reason or '',
        'decisionUpdatedAt': decision.updated_at.isoformat() if decision.updated_at else None,
        'decisionUpdatedBy': _user_display_name(decision.updated_by or decision.created_by) if (decision.updated_by or decision.created_by) else '',
    }


def _reports_cleanup_visible_candidates(candidates, decisions):
    visible = []
    held_count = 0
    dismissed_count = 0
    for candidate in candidates:
        decision = decisions.get(candidate.get('candidateKey'))
        if decision and decision.decision == AccountCleanupDecision.DECISION_DISMISSED:
            dismissed_count += 1
            continue
        if decision:
            candidate = {**candidate, **_reports_cleanup_decision_payload(decision)}
            if decision.decision == AccountCleanupDecision.DECISION_HOLD:
                held_count += 1
        visible.append(candidate)
    return visible, held_count, dismissed_count


def _reports_cleanup_department_label(department):
    if not department:
        return ''
    return ' · '.join([item for item in [department.company.name if department.company else '', department.name] if item])


def _reports_cleanup_contact_label(followup):
    if not followup:
        return ''
    account = ' · '.join([
        item for item in [
            followup.company.name if followup.company else '',
            followup.department.name if followup.department else '',
        ] if item
    ])
    name = followup.customer_name or followup.manager or f'담당자 #{followup.id}'
    return f'{account} · {name}' if account else name


def _reports_cleanup_history_payload(filter_users, limit=10):
    audit_logs = list(
        AccountCleanupAuditLog.objects.filter(
            Q(created_by__in=filter_users) | Q(created_by__isnull=True)
        ).select_related(
            'created_by',
            'source_department__company',
            'target_department__company',
            'source_followup__company',
            'source_followup__department',
            'target_followup__company',
            'target_followup__department',
        )[:limit]
    )
    decisions = list(
        AccountCleanupDecision.objects.filter(
            Q(created_by__in=filter_users) | Q(updated_by__in=filter_users)
        ).select_related(
            'created_by',
            'updated_by',
            'source_department__company',
            'target_department__company',
            'source_followup__company',
            'source_followup__department',
            'target_followup__company',
            'target_followup__department',
        )[:limit]
    )

    items = []
    for log in audit_logs:
        source_label = _reports_cleanup_department_label(log.source_department) or _reports_cleanup_contact_label(log.source_followup)
        target_label = _reports_cleanup_department_label(log.target_department) or _reports_cleanup_contact_label(log.target_followup)
        detail = ' → '.join([item for item in [source_label, target_label] if item])
        items.append({
            'id': f'audit-{log.id}',
            'kind': 'audit',
            'title': log.get_action_type_display(),
            'statusLabel': log.get_mode_display(),
            'detail': detail,
            'actorName': _user_display_name(log.created_by) if log.created_by else '',
            'createdAt': log.created_at.isoformat() if log.created_at else None,
            'auditLogId': log.id,
            'actionType': log.action_type,
            'mode': log.mode,
        })

    for decision in decisions:
        source_label = _reports_cleanup_department_label(decision.source_department) or _reports_cleanup_contact_label(decision.source_followup)
        target_label = _reports_cleanup_department_label(decision.target_department) or _reports_cleanup_contact_label(decision.target_followup)
        detail = decision.label or ' → '.join([item for item in [source_label, target_label] if item])
        items.append({
            'id': f'decision-{decision.id}',
            'kind': 'decision',
            'title': decision.get_candidate_type_display(),
            'statusLabel': decision.get_decision_display(),
            'detail': detail,
            'actorName': _user_display_name(decision.updated_by or decision.created_by) if (decision.updated_by or decision.created_by) else '',
            'createdAt': decision.updated_at.isoformat() if decision.updated_at else None,
            'candidateKey': decision.candidate_key,
            'candidateType': decision.candidate_type,
            'decision': decision.decision,
            'decisionUrl': reverse('reporting:account_cleanup_decision_api'),
            'reason': decision.reason or '',
            'sourceDepartmentId': decision.source_department_id,
            'targetDepartmentId': decision.target_department_id,
            'sourceFollowupId': decision.source_followup_id,
            'targetFollowupId': decision.target_followup_id,
        })

    items.sort(key=lambda item: item.get('createdAt') or '', reverse=True)
    return items[:limit]


def _reports_cleanup_contact_payload(followup):
    record_count = (
        int(getattr(followup, 'dq_schedule_count', 0) or 0)
        + int(getattr(followup, 'dq_history_count', 0) or 0)
        + int(getattr(followup, 'dq_quote_count', 0) or 0)
        + int(getattr(followup, 'dq_prepayment_count', 0) or 0)
    )
    return {
        'id': followup.id,
        'name': followup.customer_name or followup.manager or '담당자 미정',
        'manager': followup.manager or '',
        'email': followup.email or '',
        'phone': followup.phone_number or '',
        'companyName': followup.company.name if followup.company else '',
        'departmentName': followup.department.name if followup.department else '',
        'departmentId': followup.department_id,
        'ownerName': _user_display_name(followup.user),
        'scheduleCount': int(getattr(followup, 'dq_schedule_count', 0) or 0),
        'historyCount': int(getattr(followup, 'dq_history_count', 0) or 0),
        'quoteCount': int(getattr(followup, 'dq_quote_count', 0) or 0),
        'prepaymentCount': int(getattr(followup, 'dq_prepayment_count', 0) or 0),
        'recordCount': record_count,
        'recordSummary': (
            f"일정 {int(getattr(followup, 'dq_schedule_count', 0) or 0)}"
            f" · 노트 {int(getattr(followup, 'dq_history_count', 0) or 0)}"
            f" · 견적 {int(getattr(followup, 'dq_quote_count', 0) or 0)}"
            f" · 선결제 {int(getattr(followup, 'dq_prepayment_count', 0) or 0)}"
        ),
        'href': f'/customers/{followup.id}/',
        'accountHref': f'/accounts/{followup.department_id}/' if followup.department_id else f'/customers/{followup.id}/',
    }


def _reports_cleanup_peer_target_id(source_department_id, department_ids):
    for department_id in department_ids:
        if department_id and department_id != source_department_id:
            return department_id
    return None


def _reports_cleanup_department_payload(department_id, grouped_followups, peer_department_ids=None):
    department_followups = [
        item for item in grouped_followups
        if item.department_id == department_id
    ]
    if not department_followups:
        return None
    first = department_followups[0]
    contacts = [_reports_cleanup_contact_payload(item) for item in department_followups]
    return {
        'id': department_id,
        'name': first.department.name if first.department else '부서명 없음',
        'companyName': first.company.name if first.company else '',
        'accountHref': f'/accounts/{department_id}/',
        'contactCount': len(department_followups),
        'recordCount': sum(item['recordCount'] for item in contacts),
        'scheduleCount': sum(item['scheduleCount'] for item in contacts),
        'historyCount': sum(item['historyCount'] for item in contacts),
        'quoteCount': sum(item['quoteCount'] for item in contacts),
        'prepaymentCount': sum(item['prepaymentCount'] for item in contacts),
        'contacts': contacts[:6],
    }


def _reports_data_quality_payload(followups_qs, filter_users, limit=12, history_limit=10):
    followups = list(
        followups_qs.select_related('user', 'company', 'department')
        .annotate(
            dq_schedule_count=Count('schedules', distinct=True),
            dq_history_count=Count('histories', filter=Q(histories__parent_history__isnull=True), distinct=True),
            dq_quote_count=Count('quotes', distinct=True),
            dq_prepayment_count=Count('prepayments', distinct=True),
        )
        .order_by('company__name', 'department__name', 'customer_name', 'id')
    )

    account_name_groups = {}
    duplicate_contact_groups = {}
    contacts_without_department = []
    contacts_without_company = []
    for followup in followups:
        company_name = followup.company.name if followup.company else ''
        department_name = followup.department.name if followup.department else ''
        company_unassigned = _reports_is_unassigned_company(followup.company)
        department_unassigned = _reports_is_unassigned_department(followup.department)
        if followup.company_id and followup.department_id and not company_unassigned and not department_unassigned:
            account_key = (
                _reports_cleanup_key(company_name),
                _reports_cleanup_key(department_name),
            )
            if account_key[0] and account_key[1]:
                account_name_groups.setdefault(account_key, []).append(followup)

        if followup.department_id and not department_unassigned:
            contact_scope = f'department:{followup.department_id}'
        elif followup.company_id and not company_unassigned:
            contact_scope = f'company:{followup.company_id}'
        else:
            contact_scope = 'unassigned'
        identity = (followup.email or '').strip().lower()
        if not identity:
            identity = _reports_cleanup_key(followup.customer_name or followup.manager)
        if identity:
            duplicate_contact_groups.setdefault((contact_scope, identity), []).append(followup)

        if department_unassigned:
            contacts_without_department.append(followup)
        if company_unassigned:
            contacts_without_company.append(followup)

    duplicate_accounts = []
    for grouped_followups in account_name_groups.values():
        department_ids = sorted({item.department_id for item in grouped_followups if item.department_id})
        department_names = sorted({item.department.name for item in grouped_followups if item.department})
        if len(department_ids) < 2:
            continue
        first = grouped_followups[0]
        department_payloads = [
            item for item in [
                _reports_cleanup_department_payload(department_id, grouped_followups, department_ids)
                for department_id in department_ids
            ] if item
        ]
        contact_payloads = [_reports_cleanup_contact_payload(item) for item in grouped_followups]
        source_department_id = department_ids[0] if department_ids else None
        target_department_id = _reports_cleanup_peer_target_id(source_department_id, department_ids)
        duplicate_accounts.append({
            **_reports_cleanup_candidate_base(
                AccountCleanupDecision.CANDIDATE_DUPLICATE_ACCOUNT,
                _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_DUPLICATE_ACCOUNT, *department_ids),
                sourceDepartmentId=source_department_id,
                targetDepartmentId=target_department_id,
            ),
            'companyName': first.company.name if first.company else '',
            'normalizedDepartmentName': _reports_cleanup_key(first.department.name if first.department else ''),
            'departmentNames': department_names,
            'departmentIds': department_ids,
            'contactCount': len(grouped_followups),
            'recordCount': sum(item['recordCount'] for item in contact_payloads),
            'riskLevel': 'review',
            'riskLabel': '검토 필요',
            'suggestedAction': '같은 업체 안에서 부서/연구실명이 매우 유사합니다. 실제 같은 계정인지 확인 후 병합 후보로 검토하세요.',
            'departments': department_payloads,
            'contacts': contact_payloads[:6],
        })
    duplicate_accounts.sort(key=lambda item: (-item['contactCount'], item['companyName'], item['normalizedDepartmentName']))

    duplicate_contacts = []
    for grouped_followups in duplicate_contact_groups.values():
        if len(grouped_followups) < 2:
            continue
        first = grouped_followups[0]
        contact_payloads = [_reports_cleanup_contact_payload(item) for item in grouped_followups]
        contact_ids = sorted(item.id for item in grouped_followups)
        duplicate_contacts.append({
            **_reports_cleanup_candidate_base(
                AccountCleanupDecision.CANDIDATE_DUPLICATE_CONTACT,
                _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_DUPLICATE_CONTACT, *contact_ids),
                sourceFollowupId=contact_ids[0] if contact_ids else None,
                targetFollowupId=contact_ids[1] if len(contact_ids) > 1 else None,
            ),
            'companyName': first.company.name if first.company else '',
            'departmentName': first.department.name if first.department else '',
            'identity': (first.email or first.customer_name or first.manager or '담당자 미정'),
            'contactCount': len(grouped_followups),
            'recordCount': sum(item['recordCount'] for item in contact_payloads),
            'contactIds': contact_ids,
            'riskLevel': 'review',
            'riskLabel': '검토 필요',
            'suggestedAction': '같은 계정 범위에서 이메일 또는 이름 키가 같습니다. 실제 동일 담당자인지 확인 후 기록 이관/병합을 검토하세요.',
            'contacts': contact_payloads[:6],
        })
    duplicate_contacts.sort(key=lambda item: (-item['contactCount'], item['companyName'], item['departmentName'], item['identity']))

    contacts_without_department.sort(
        key=lambda item: (
            -(
                int(getattr(item, 'dq_schedule_count', 0) or 0)
                + int(getattr(item, 'dq_history_count', 0) or 0)
                + int(getattr(item, 'dq_quote_count', 0) or 0)
                + int(getattr(item, 'dq_prepayment_count', 0) or 0)
            ),
            item.company.name if item.company else '',
            item.customer_name or item.manager or '',
        )
    )
    contacts_without_company.sort(key=lambda item: (item.customer_name or item.manager or '', item.id))
    unassigned_department_candidates = [
        {
            **_reports_cleanup_contact_payload(item),
            **_reports_cleanup_candidate_base(
                AccountCleanupDecision.CANDIDATE_UNASSIGNED_DEPARTMENT,
                _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_UNASSIGNED_DEPARTMENT, item.id),
                sourceFollowupId=item.id,
            ),
        }
        for item in contacts_without_department
    ]
    unassigned_company_candidates = [
        {
            **_reports_cleanup_contact_payload(item),
            **_reports_cleanup_candidate_base(
                AccountCleanupDecision.CANDIDATE_UNASSIGNED_COMPANY,
                _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_UNASSIGNED_COMPANY, item.id),
                sourceFollowupId=item.id,
            ),
        }
        for item in contacts_without_company
    ]

    candidate_keys = [
        item['candidateKey']
        for item in (
            duplicate_accounts
            + duplicate_contacts
            + unassigned_department_candidates
            + unassigned_company_candidates
        )
        if item.get('candidateKey')
    ]
    decisions = {
        item.candidate_key: item
        for item in AccountCleanupDecision.objects.filter(candidate_key__in=candidate_keys).select_related(
            'created_by', 'updated_by',
        )
    }
    duplicate_accounts, held_accounts, dismissed_accounts = _reports_cleanup_visible_candidates(
        duplicate_accounts,
        decisions,
    )
    duplicate_contacts, held_contacts, dismissed_contacts = _reports_cleanup_visible_candidates(
        duplicate_contacts,
        decisions,
    )
    unassigned_department_candidates, held_departments, dismissed_departments = _reports_cleanup_visible_candidates(
        unassigned_department_candidates,
        decisions,
    )
    unassigned_company_candidates, held_companies, dismissed_companies = _reports_cleanup_visible_candidates(
        unassigned_company_candidates,
        decisions,
    )

    held_candidate_count = held_accounts + held_contacts + held_departments + held_companies
    dismissed_candidate_count = dismissed_accounts + dismissed_contacts + dismissed_departments + dismissed_companies
    cleanup_candidate_count = (
        len(duplicate_accounts)
        + len(duplicate_contacts)
        + len(unassigned_department_candidates)
        + len(unassigned_company_candidates)
    )
    return {
        'metrics': {
            'duplicateAccountGroups': len(duplicate_accounts),
            'duplicateContactGroups': len(duplicate_contacts),
            'contactsWithoutDepartment': len(unassigned_department_candidates),
            'contactsWithoutCompany': len(unassigned_company_candidates),
            'cleanupCandidateCount': cleanup_candidate_count,
            'heldCandidateCount': held_candidate_count,
            'dismissedCandidateCount': dismissed_candidate_count,
        },
        'normalizationRule': '업체/부서/담당자명은 괄호 표기, 대학/병원/기관 약칭, 연구실/Lab 표기 차이를 정규화한 값으로 후보를 찾습니다.',
        'duplicateAccounts': duplicate_accounts[:limit],
        'duplicateContacts': duplicate_contacts[:limit],
        'contactsWithoutDepartment': unassigned_department_candidates[:limit],
        'contactsWithoutCompany': unassigned_company_candidates[:limit],
        'history': _reports_cleanup_history_payload(filter_users, limit=history_limit),
    }


def _account_cleanup_payload_ref(payload, key):
    value = payload.get(key)
    if value in (None, ''):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _account_cleanup_payload_label(payload, source_department, target_department, source_followup, target_followup):
    label = _account_cleanup_payload_text(payload, 'label', 'candidateLabel')
    if label:
        return label[:255]
    source_label = _reports_cleanup_department_label(source_department) or _reports_cleanup_contact_label(source_followup)
    target_label = _reports_cleanup_department_label(target_department) or _reports_cleanup_contact_label(target_followup)
    return ' → '.join([item for item in [source_label, target_label] if item])[:255]


def _account_cleanup_accessible_department(department_id, scope_users):
    if not department_id:
        return None, None
    try:
        department = Department.objects.select_related('company').get(id=department_id)
    except Department.DoesNotExist:
        return None, JsonResponse({'success': False, 'error': 'department_not_found'}, status=404)
    if account_representative_followup(department, scope_users) is None:
        return None, JsonResponse({'success': False, 'error': 'department_forbidden'}, status=403)
    return department, None


def _account_cleanup_accessible_followup(followup_id, scope_users):
    if not followup_id:
        return None, None
    try:
        followup = FollowUp.objects.select_related('user', 'company', 'department').get(id=followup_id)
    except FollowUp.DoesNotExist:
        return None, JsonResponse({'success': False, 'error': 'followup_not_found'}, status=404)
    if not FollowUp.objects.filter(pk=followup.id, user__in=scope_users).exists():
        return None, JsonResponse({'success': False, 'error': 'followup_forbidden'}, status=403)
    return followup, None


@ensure_csrf_cookie
@never_cache
@require_http_methods(["POST"])
def account_cleanup_decision_api(request):
    """Hold, dismiss, or restore a data-quality cleanup candidate."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    payload, error_response = _account_cleanup_request_payload(request)
    if error_response:
        return error_response

    candidate_type = _account_cleanup_payload_text(payload, 'candidateType', 'candidate_type')
    candidate_key = _account_cleanup_payload_text(payload, 'candidateKey', 'candidate_key')
    decision_value = _account_cleanup_payload_text(payload, 'decision', 'reviewStatus', 'status').lower()
    if candidate_type not in {value for value, _label in AccountCleanupDecision.CANDIDATE_TYPE_CHOICES}:
        return JsonResponse({'success': False, 'error': 'invalid_candidate_type'}, status=400)
    if not candidate_key or not candidate_key.startswith(f'{candidate_type}:'):
        return JsonResponse({'success': False, 'error': 'invalid_candidate_key'}, status=400)

    restore_values = {'active', 'new', 'restore', 'reset', 'clear'}
    valid_decisions = {value for value, _label in AccountCleanupDecision.DECISION_CHOICES}
    if decision_value not in valid_decisions and decision_value not in restore_values:
        return JsonResponse({'success': False, 'error': 'invalid_decision'}, status=400)

    user_profile = get_user_profile(request.user)
    scope_users = _account_cleanup_scope_users(request, user_profile)
    source_department, error_response = _account_cleanup_accessible_department(
        _account_cleanup_payload_ref(payload, 'sourceDepartmentId') or _account_cleanup_payload_ref(payload, 'source_department_id'),
        scope_users,
    )
    if error_response:
        return error_response
    target_department, error_response = _account_cleanup_accessible_department(
        _account_cleanup_payload_ref(payload, 'targetDepartmentId') or _account_cleanup_payload_ref(payload, 'target_department_id'),
        scope_users,
    )
    if error_response:
        return error_response
    source_followup, error_response = _account_cleanup_accessible_followup(
        _account_cleanup_payload_ref(payload, 'sourceFollowupId') or _account_cleanup_payload_ref(payload, 'source_followup_id'),
        scope_users,
    )
    if error_response:
        return error_response
    target_followup, error_response = _account_cleanup_accessible_followup(
        _account_cleanup_payload_ref(payload, 'targetFollowupId') or _account_cleanup_payload_ref(payload, 'target_followup_id'),
        scope_users,
    )
    if error_response:
        return error_response

    if decision_value in restore_values:
        existing_decision = AccountCleanupDecision.objects.filter(candidate_key=candidate_key).select_related(
            'source_department',
            'target_department',
            'source_followup',
            'target_followup',
        ).first()
        if existing_decision and not any([source_department, target_department, source_followup, target_followup]):
            source_department, error_response = _account_cleanup_accessible_department(
                existing_decision.source_department_id,
                scope_users,
            )
            if error_response:
                return error_response
            target_department, error_response = _account_cleanup_accessible_department(
                existing_decision.target_department_id,
                scope_users,
            )
            if error_response:
                return error_response
            source_followup, error_response = _account_cleanup_accessible_followup(
                existing_decision.source_followup_id,
                scope_users,
            )
            if error_response:
                return error_response
            target_followup, error_response = _account_cleanup_accessible_followup(
                existing_decision.target_followup_id,
                scope_users,
            )
            if error_response:
                return error_response
        deleted_count, _ = AccountCleanupDecision.objects.filter(candidate_key=candidate_key).delete()
        return JsonResponse({
            'success': True,
            'message': '정리 후보 판단을 복구했습니다.',
            'restored': True,
            'deletedCount': deleted_count,
            'candidateKey': candidate_key,
        })

    if not any([source_department, target_department, source_followup, target_followup]):
        return JsonResponse({'success': False, 'error': 'candidate_refs_required'}, status=400)

    label = _account_cleanup_payload_label(payload, source_department, target_department, source_followup, target_followup)
    reason = _account_cleanup_payload_text(payload, 'reason', 'memo')
    decision, created = AccountCleanupDecision.objects.get_or_create(
        candidate_key=candidate_key,
        defaults={
            'candidate_type': candidate_type,
            'decision': decision_value,
            'label': label,
            'reason': reason,
            'source_department': source_department,
            'target_department': target_department,
            'source_followup': source_followup,
            'target_followup': target_followup,
            'created_by': request.user,
            'updated_by': request.user,
        },
    )
    if not created:
        decision.candidate_type = candidate_type
        decision.decision = decision_value
        decision.label = label
        decision.reason = reason
        decision.source_department = source_department
        decision.target_department = target_department
        decision.source_followup = source_followup
        decision.target_followup = target_followup
        decision.updated_by = request.user
        decision.save(update_fields=[
            'candidate_type',
            'decision',
            'label',
            'reason',
            'source_department',
            'target_department',
            'source_followup',
            'target_followup',
            'updated_by',
            'updated_at',
        ])
    return JsonResponse({
        'success': True,
        'message': f'정리 후보를 {decision.get_decision_display()} 처리했습니다.',
        'candidateKey': decision.candidate_key,
        'candidateType': decision.candidate_type,
        'decision': decision.decision,
        'reviewStatusLabel': decision.get_decision_display(),
    })


@ensure_csrf_cookie
@never_cache
@require_http_methods(["POST"])
def data_quality_contact_assign_account_api(request, followup_id):
    """Assign an unassigned or placeholder contact to a concrete Department account."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    payload, error_response = _account_cleanup_request_payload(request)
    if error_response:
        return error_response
    target_department_id = _account_cleanup_payload_ref(payload, 'departmentId') or _account_cleanup_payload_ref(payload, 'department_id')
    if not target_department_id:
        return JsonResponse({'success': False, 'error': 'department_required'}, status=400)

    followup = get_object_or_404(
        FollowUp.objects.select_related('user', 'company', 'department'),
        pk=followup_id,
    )
    if not can_access_followup(request.user, followup):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    if not can_modify_user_data(request.user, followup.user):
        return JsonResponse({'success': False, 'error': '수정 권한이 없습니다.'}, status=403)

    target_department = get_object_or_404(Department.objects.select_related('company'), pk=target_department_id)
    user_profile = get_user_profile(request.user)
    edit_config = _customers_edit_config(request, user_profile, followup, True)
    allowed_department_ids = {item['id'] for item in edit_config['departments']}
    if target_department.id not in allowed_department_ids:
        return JsonResponse({'success': False, 'error': '접근 권한이 없는 부서/연구실입니다.'}, status=403)

    before_snapshot = {
        'contact': _account_cleanup_followup_ref(followup),
        'wasCompanyUnassigned': _reports_is_unassigned_company(followup.company),
        'wasDepartmentUnassigned': _reports_is_unassigned_department(followup.department),
    }
    with transaction.atomic():
        audit_log = AccountCleanupAuditLog.objects.create(
            created_by=request.user,
            action_type=AccountCleanupAuditLog.ACTION_CONTACT_ACCOUNT_ASSIGN,
            mode=AccountCleanupAuditLog.MODE_EXECUTE,
            source_department=followup.department,
            target_department=target_department,
            source_followup=followup,
            before_snapshot=before_snapshot,
            result={'status': 'started'},
        )
        followup.company = target_department.company
        followup.department = target_department
        followup.save(update_fields=['company', 'department', 'updated_at'])
        followup.refresh_from_db()
        AccountCleanupDecision.objects.filter(candidate_key__in=[
            _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_UNASSIGNED_COMPANY, followup.id),
            _reports_cleanup_candidate_key(AccountCleanupDecision.CANDIDATE_UNASSIGNED_DEPARTMENT, followup.id),
        ]).delete()
        audit_log.after_snapshot = {'contact': _account_cleanup_followup_ref(followup)}
        audit_log.result = {
            'status': 'completed',
            'targetDepartmentId': target_department.id,
            'targetCompanyId': target_department.company_id,
        }
        audit_log.save(update_fields=['after_snapshot', 'result', 'updated_at'])

    return JsonResponse({
        'success': True,
        'message': '담당자를 계정에 연결했습니다.',
        'auditLogId': audit_log.id,
        'contact': _reports_cleanup_contact_payload(followup),
    })


@ensure_csrf_cookie
@never_cache
@require_http_methods(["GET"])
def reports_summary_api(request):
    """React 보고서/분석 화면용 JSON API."""
    auth_response = _api_login_required_response(request)
    if auth_response:
        return auth_response

    user_profile = get_user_profile(request.user)
    today = timezone.localdate()
    date_from, date_to = _analytics_api_date_range(request, today)
    filter_users, salesperson_list, selected_user = _analytics_api_scope_users(request, user_profile)
    report_filters = _reports_filter_params(request)
    previous_from, previous_to = _reports_previous_date_range(date_from, date_to)

    histories_qs = History.objects.filter(
        user__in=filter_users,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        parent_history__isnull=True,
    )
    base_followups_qs = FollowUp.objects.filter(user__in=filter_users)
    filter_options = _reports_filter_options(base_followups_qs)
    followups_qs = _reports_apply_account_filters(
        base_followups_qs,
        report_filters,
        filter_users=filter_users,
        date_from=date_from,
        date_to=date_to,
    )
    customer_operations = _reports_customer_operations_payload(
        followups_qs,
        filter_users,
        date_from,
        date_to,
        request.user,
    )
    cleanup_limit = 12
    cleanup_history_limit = 10
    try:
        cleanup_limit = max(1, min(int(request.GET.get('cleanup_limit') or cleanup_limit), 100))
    except (TypeError, ValueError):
        cleanup_limit = 12
    try:
        cleanup_history_limit = max(1, min(int(request.GET.get('cleanup_history_limit') or cleanup_history_limit), 100))
    except (TypeError, ValueError):
        cleanup_history_limit = 10
    data_quality = _reports_data_quality_payload(
        followups_qs,
        filter_users,
        limit=cleanup_limit,
        history_limit=cleanup_history_limit,
    )
    cleanup_markers = _reports_cleanup_marker_map(data_quality)
    customer_operations = _reports_attach_cleanup_markers(customer_operations, cleanup_markers)
    customer_operations = _reports_apply_row_filters(customer_operations, report_filters)
    previous_operations = _reports_customer_operations_payload(
        followups_qs,
        filter_users,
        previous_from,
        previous_to,
        request.user,
    )
    previous_operations = _reports_attach_cleanup_markers(previous_operations, cleanup_markers)
    previous_operations = _reports_apply_row_filters(previous_operations, report_filters)
    operations_comparison = _reports_operations_comparison(
        customer_operations.get('metrics') or {},
        previous_operations.get('metrics') or {},
        previous_from,
        previous_to,
    )

    total_histories = histories_qs.count()
    completed_followups = histories_qs.filter(
        next_action_date__isnull=False,
        reviewed_at__isnull=False,
    ).count()
    overdue_followups = History.objects.filter(
        user__in=filter_users,
        next_action_date__lt=today,
        next_action_date__isnull=False,
        reviewed_at__isnull=True,
        parent_history__isnull=True,
    ).count()
    upcoming_followups = History.objects.filter(
        user__in=filter_users,
        next_action_date__gte=today,
        next_action_date__isnull=False,
        reviewed_at__isnull=True,
        parent_history__isnull=True,
    ).count()
    active_pipeline = followups_qs.filter(
        status='active',
        pipeline_stage__in=['potential', 'contact', 'quote', 'negotiation'],
    ).count()

    activity_by_user = {
        row['user_id']: row
        for row in histories_qs.values('user_id').annotate(
            history_count=Count('id'),
            last=Max('created_at'),
        )
    }
    active_followups_by_user = {
        row['user_id']: row['count']
        for row in FollowUp.objects.filter(
            user__in=filter_users,
            status='active',
        ).values('user_id').annotate(count=Count('id'))
    }
    overdue_by_user = {
        row['user_id']: row['count']
        for row in History.objects.filter(
            user__in=filter_users,
            next_action_date__lt=today,
            next_action_date__isnull=False,
            reviewed_at__isnull=True,
            parent_history__isnull=True,
        ).values('user_id').annotate(count=Count('id'))
    }
    activity_report = []
    for user in filter_users.select_related('userprofile'):
        activity = activity_by_user.get(user.id, {})
        activity_report.append({
            'user': _analytics_user_payload(user),
            'historyCount': activity.get('history_count', 0),
            'followupCount': active_followups_by_user.get(user.id, 0),
            'overdueCount': overdue_by_user.get(user.id, 0),
            'lastActivityAt': activity.get('last'),
        })
    activity_report.sort(key=lambda item: item['historyCount'], reverse=True)

    active_followup_ids = histories_qs.values_list('followup_id', flat=True).distinct()
    customer_report = []
    customer_followups = list(FollowUp.objects.filter(
        pk__in=active_followup_ids,
    ).select_related('company', 'department', 'user').order_by('-updated_at')[:50])
    customer_history_by_followup = {
        row['followup_id']: row
        for row in History.objects.filter(
            followup_id__in=[followup.id for followup in customer_followups],
            parent_history__isnull=True,
        ).values('followup_id').annotate(
            last=Max('created_at'),
            next_date=Max('next_action_date'),
        )
    }
    for followup in customer_followups:
        last_hist = customer_history_by_followup.get(followup.id, {})
        customer_report.append({
            'id': followup.id,
            'customer': followup.customer_name or '',
            'company': followup.company.name if followup.company else '',
            'department': followup.department.name if followup.department else '',
            'owner': _user_display_name(followup.user),
            'pipelineStage': followup.pipeline_stage,
            'pipelineStageLabel': followup.get_pipeline_stage_display(),
            'lastActivityAt': last_hist['last'],
            'nextActionDate': last_hist['next_date'],
            'href': f'/customers/{followup.id}/',
            'djangoHref': reverse('reporting:followup_detail', args=[followup.id]),
        })

    stage_order = ['potential', 'contact', 'quote', 'negotiation', 'won', 'lost']
    stage_counts = {
        row['pipeline_stage']: row['count']
        for row in followups_qs.values('pipeline_stage').annotate(count=Count('id'))
    }
    pipeline_summary = []
    for stage in stage_order:
        pipeline_summary.append({
            'stage': stage,
            'label': dict(FollowUp.PIPELINE_STAGE_CHOICES).get(stage, stage),
            'count': stage_counts.get(stage, 0),
        })

    open_service_statuses = ['received', 'in_progress', 'waiting']
    asset_scope = CustomerAsset.objects.filter(created_by__in=filter_users)
    asset_metrics = {
        'totalAssets': asset_scope.count(),
        'activeAssets': asset_scope.filter(status='active').count(),
        'openServiceAssets': asset_scope.filter(service_cases__status__in=open_service_statuses).distinct().count(),
        'overdueServiceAssets': asset_scope.filter(
            service_cases__status__in=open_service_statuses,
            service_cases__due_date__lt=today,
        ).distinct().count(),
        'dueCalibrationAssets': asset_scope.filter(
            calibration_records__next_due_date__gte=today,
            calibration_records__next_due_date__lte=today + timedelta(days=30),
        ).distinct().count(),
        'overdueCalibrationAssets': asset_scope.filter(
            calibration_records__next_due_date__lt=today,
        ).distinct().count(),
    }

    from urllib.parse import urlencode
    query_payload = {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
    }
    if selected_user and (user_profile.is_admin() or user_profile.is_manager()):
        query_payload['user_id'] = selected_user.id
    if report_filters.get('query'):
        query_payload['q'] = report_filters['query']
    if report_filters.get('companyId'):
        query_payload['company_id'] = report_filters['companyId']
    if report_filters.get('departmentId'):
        query_payload['department_id'] = report_filters['departmentId']
    if report_filters.get('deliveryFilter') and report_filters['deliveryFilter'] != 'any':
        query_payload['delivery_filter'] = report_filters['deliveryFilter']
    if report_filters.get('prepaymentBalanceFilter') and report_filters['prepaymentBalanceFilter'] != 'any':
        query_payload['prepayment_balance_filter'] = report_filters['prepaymentBalanceFilter']
    if report_filters.get('exportScope') and report_filters['exportScope'] != 'filtered':
        query_payload['export_scope'] = report_filters['exportScope']
    if report_filters.get('sort') and report_filters['sort'] != 'recent':
        query_payload['sort'] = report_filters['sort']
    query = '?' + urlencode(query_payload)
    can_export = bool(user_profile.is_admin() or user_profile.is_manager())
    return JsonResponse({
        'success': True,
        'source': 'django',
        'generatedAt': timezone.now().isoformat(),
        'filters': {
            'dateFrom': date_from.isoformat(),
            'dateTo': date_to.isoformat(),
            'selectedUserId': selected_user.id if selected_user and (user_profile.is_admin() or user_profile.is_manager()) else None,
            'query': report_filters.get('query') or '',
            'companyId': report_filters.get('companyId'),
            'departmentId': report_filters.get('departmentId'),
            'deliveryFilter': report_filters.get('deliveryFilter') or 'any',
            'prepaymentBalanceFilter': report_filters.get('prepaymentBalanceFilter') or 'any',
            'exportScope': report_filters.get('exportScope') or 'filtered',
            'sort': report_filters.get('sort') or 'recent',
        },
        'scope': {
            'canFilterUsers': bool(user_profile.is_admin() or user_profile.is_manager()),
            'canExport': can_export,
            'label': _user_display_name(selected_user) if selected_user else (
                f'{user_profile.company.name} 팀' if user_profile.company else '전체'
            ),
            'salespeople': [_analytics_user_payload(user) for user in salesperson_list],
            **filter_options,
        },
        'metrics': {
            'totalHistories': total_histories,
            'completedFollowups': completed_followups,
            'overdueFollowups': overdue_followups,
            'upcomingFollowups': upcoming_followups,
            'activePipeline': active_pipeline,
            **asset_metrics,
        },
        'activityReport': activity_report,
        'customerReport': customer_report,
        'customerOperations': customer_operations,
        'comparison': {
            'customerOperations': operations_comparison,
        },
        'dataQuality': data_quality,
        'pipelineSummary': pipeline_summary,
        'links': {
            'activityCsv': reverse('reporting:analytics_activity_csv') + query,
            'pipelineCsv': reverse('reporting:analytics_pipeline_csv') + query,
            'activityXlsx': reverse('reporting:analytics_activity_xlsx') + query,
            'pipelineXlsx': reverse('reporting:analytics_pipeline_xlsx') + query,
            'customerOperationsXlsx': reverse('reporting:reports_customer_operations_xlsx') + query,
            'assets': '/assets/',
            'legacy': reverse('reporting:analytics_dashboard'),
        },
    }, encoder=DjangoJSONEncoder)


@never_cache
@login_required
@require_http_methods(["GET"])
def reports_customer_operations_xlsx_export_api(request):
    """React 보고서의 부서/연구실 계정별 운영 현황표를 XLSX로 내보낸다."""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote

    user_profile = get_user_profile(request.user)
    if not (user_profile.is_admin() or user_profile.is_manager()):
        return HttpResponseForbidden('접근 권한이 없습니다.')

    today = timezone.localdate()
    date_from, date_to = _analytics_api_date_range(request, today)
    filter_users, _salesperson_list, selected_user = _analytics_api_scope_users(request, user_profile)
    report_filters = _reports_filter_params(request)
    base_followups_qs = FollowUp.objects.filter(user__in=filter_users)
    followups_qs = _reports_apply_account_filters(
        base_followups_qs,
        report_filters,
        filter_users=filter_users,
        date_from=date_from,
        date_to=date_to,
    )
    data_quality = _reports_data_quality_payload(followups_qs, filter_users)
    cleanup_markers = _reports_cleanup_marker_map(data_quality)
    operations = _reports_customer_operations_payload(
        followups_qs,
        filter_users,
        date_from,
        date_to,
        request.user,
    )
    operations = _reports_attach_cleanup_markers(operations, cleanup_markers)
    operations = _reports_apply_row_filters(
        operations,
        report_filters,
        export_scope=report_filters.get('exportScope') or 'filtered',
    )
    rows = operations.get('rows') or []
    metrics = operations.get('metrics') or {}

    wb = Workbook()
    ws = wb.active
    ws.title = '계정별 운영 현황'

    headers = [
        '계정', '업체/학교', '부서/연구실', '담당자수', '담당자 미리보기',
        '영업 담당자', '파이프라인', '상태', '우선순위',
        '납품건수', '납품금액', '선결제차감건수', '선결제차감액',
        '일반납품건수', '일반납품금액', '견적건수', '견적금액',
        '선결제건수', '선결제총액', '선결제잔액', '선결제사용액',
        '서비스건수', '진행서비스', '최근납품일', '최근견적일', '최근선결제일',
        '최근서비스일', '최근활동일', '정리후보', '정리유형', '최근견적품목', '최근납품품목', '계정링크',
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF')
    thin_side = Side(style='thin', color='D1D5DB')
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    money_format = '#,##0'

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row in rows:
        recent_quote_items = []
        for item in row.get('recentQuoteItems') or []:
            parts = [
                str(item.get('date') or ''),
                item.get('quoteLabel') or item.get('source') or '',
                item.get('statusLabel') or '',
                item.get('label') or '',
            ]
            amount = item.get('amount') or 0
            if amount:
                parts.append(f'{int(amount):,}원')
            recent_quote_items.append(' · '.join([part for part in parts if part]))
        recent_items = []
        for item in row.get('recentDeliveryItems') or []:
            parts = [
                str(item.get('date') or ''),
                item.get('paymentSourceLabel') or '',
                item.get('paymentStatusLabel') or '',
                item.get('label') or '',
            ]
            amount = item.get('amount') or 0
            if amount:
                parts.append(f'{int(amount):,}원')
            recent_items.append(' · '.join([part for part in parts if part]))
        ws.append([
            row.get('customer') or '',
            row.get('company') or '',
            row.get('department') or '',
            row.get('contactCount') or 0,
            ', '.join(row.get('contactPreview') or []),
            row.get('owner') or '',
            row.get('pipelineStageLabel') or '',
            row.get('statusLabel') or '',
            row.get('priorityLabel') or '',
            row.get('deliveryCount') or 0,
            row.get('deliveryAmount') or 0,
            row.get('prepaymentDeliveryCount') or 0,
            row.get('prepaymentUsedAmount') or 0,
            row.get('normalDeliveryCount') or 0,
            row.get('normalDeliveryAmount') or 0,
            row.get('quoteCount') or 0,
            row.get('quoteAmount') or 0,
            row.get('prepaymentCount') or 0,
            row.get('prepaymentAmount') or 0,
            row.get('prepaymentBalance') or 0,
            row.get('prepaymentUsedTotal') or 0,
            row.get('serviceCount') or 0,
            row.get('openServiceCount') or 0,
            row.get('lastDeliveryDate') or '',
            row.get('lastQuoteDate') or '',
            row.get('lastPrepaymentDate') or '',
            row.get('lastServiceDate') or '',
            row.get('lastActivityDate') or '',
            row.get('cleanupCandidateCount') or 0,
            ', '.join(row.get('cleanupTypes') or []),
            '\n'.join(recent_quote_items),
            '\n'.join(recent_items),
            request.build_absolute_uri(row.get('href') or '') if row.get('href') else '',
        ])

    for sheet_row in ws.iter_rows(min_row=2):
        for cell in sheet_row:
            cell.border = border
            cell.alignment = body_alignment
        for col_idx in [11, 13, 15, 17, 19, 20, 21]:
            sheet_row[col_idx - 1].number_format = money_format

    widths = [
        24, 24, 22, 10, 28,
        16, 16, 14, 12,
        10, 14, 14, 16,
        14, 16, 10, 14,
        12, 14, 14, 14,
        12, 12, 12, 12, 12,
        12, 12, 10, 20, 46, 46, 46,
    ]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    info = wb.create_sheet(title='다운로드 정보')
    scope_label = _user_display_name(selected_user) if selected_user else (
        f'{user_profile.company.name} 팀' if user_profile.company else '전체'
    )
    info_rows = [
        ('보고서', '부서/연구실 계정별 운영 현황'),
        ('기간', f'{date_from.isoformat()} ~ {date_to.isoformat()}'),
        ('범위', scope_label),
        ('검색어', report_filters.get('query') or '전체'),
        ('업체 ID', report_filters.get('companyId') or '전체'),
        ('부서 ID', report_filters.get('departmentId') or '전체'),
        ('납품 필터', report_filters.get('deliveryFilter') or 'any'),
        ('선결제 잔액 필터', report_filters.get('prepaymentBalanceFilter') or 'any'),
        ('엑셀 범위', report_filters.get('exportScope') or 'filtered'),
        ('표시 계정', metrics.get('totalCustomers') or len(rows)),
        ('납품 건수', metrics.get('deliveryCount') or 0),
        ('납품 금액', metrics.get('deliveryAmount') or 0),
        ('선결제 차감 납품', metrics.get('prepaymentDeliveryCount') or 0),
        ('일반 납품', metrics.get('normalDeliveryCount') or 0),
        ('견적 건수', metrics.get('quoteCount') or 0),
        ('선결제 잔액', metrics.get('prepaymentBalance') or 0),
        ('생성일시', timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')),
        ('생성자', _user_display_name(request.user)),
    ]
    for row in info_rows:
        info.append(row)
    info.column_dimensions['A'].width = 20
    info.column_dimensions['B'].width = 42

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'계정별_운영현황_{date_from.isoformat()}_{date_to.isoformat()}.xlsx'
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response
