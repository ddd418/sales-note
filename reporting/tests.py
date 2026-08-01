import json
import os
import tempfile
from datetime import time, timedelta
from pathlib import Path
from urllib.parse import urljoin
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, Client, override_settings
from django.urls import reverse
from django.contrib.auth.models import User
from django.template.loader import get_template
from django.utils import timezone
from reporting.models import (
    AccountCleanupAuditLog,
    AccountCleanupDecision,
    Company,
    Department,
    DepartmentMemo,
    DemoRecord,
    DocumentGenerationLog,
    DocumentTemplate,
    FollowUp,
    History,
    PersonalSchedule,
    CustomerAsset,
    ServiceCase,
    CalibrationRecord,
    DeliveryItem,
    Prepayment,
    PrepaymentLedgerEntry,
    PrepaymentUsage,
    Product,
    Quote,
    QuoteItem,
    Schedule,
    ScheduleQuoteGroupNote,
    UserProfile,
    UserCompany,
)
from reporting.services.test_fixtures import create_account_ledger_fixture


FRONTEND_BASE_URL = 'https://sales-note-frontend-production.up.railway.app/'


def frontend_url(path):
    return urljoin(FRONTEND_BASE_URL, path.lstrip('/'))


# ─────────────────────────────────────────────────────────────────────────────
# 헬퍼: 역할이 있는 사용자 생성
# ─────────────────────────────────────────────────────────────────────────────

def make_user(username, password='TestPass123!', role='salesman',
              can_use_ai=False, can_download_excel=False, company=None):
    """테스트용 사용자 생성 헬퍼"""
    user = User.objects.create_user(username=username, password=password)
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.role = role
    profile.can_use_ai = can_use_ai
    profile.can_download_excel = can_download_excel
    if company:
        profile.company = company
    profile.save()
    return user


class AuthenticationSmoke(TestCase):
    """인증 기본 smoke 테스트"""

    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='testuser',
            password='TestPass123!',
        )

    def test_login_page_returns_200(self):
        """로그인 페이지 접근 가능"""
        response = self.client.get(reverse('reporting:login'))
        self.assertEqual(response.status_code, 200)

    def test_unauthenticated_followup_list_redirects(self):
        """미인증 상태에서 거래처 목록 접근 시 로그인으로 리다이렉트"""
        response = self.client.get(reverse('reporting:followup_list'))
        self.assertIn(response.status_code, [302, 301])
        self.assertIn('/login', response.get('Location', ''))

    def test_login_success(self):
        """올바른 자격 증명으로 로그인하면 프론트 CRM 대시보드로 이동"""
        response = self.client.post(
            reverse('reporting:login'),
            {'username': 'testuser', 'password': 'TestPass123!'},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response['Location'],
            'https://sales-note-frontend-production.up.railway.app/dashboard/',
        )
        self.assertEqual(str(self.client.session.get('_auth_user_id')), str(self.user.id))

    def test_login_page_preserves_next_hidden_field(self):
        """React 직접 URL에서 온 next 값은 로그인 form POST까지 유지"""
        response = self.client.get(reverse('reporting:login'), {'next': '/customers/42/'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="next"', html=False)
        self.assertContains(response, 'value="/customers/42/"', html=False)

    def test_login_success_redirects_to_next_react_path(self):
        """로그인 성공 후 상대 React next 경로로 복귀"""
        response = self.client.post(
            reverse('reporting:login'),
            {
                'username': 'testuser',
                'password': 'TestPass123!',
                'next': '/customers/42/',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], '/customers/42/')
        self.assertEqual(str(self.client.session.get('_auth_user_id')), str(self.user.id))

    @override_settings(FRONTEND_PIPELINE_URL=FRONTEND_BASE_URL)
    def test_login_success_allows_configured_frontend_absolute_next(self):
        """프론트 운영 도메인의 absolute next URL은 허용"""
        next_url = frontend_url('customers/?date_from=2026-05-01')

        response = self.client.post(
            reverse('reporting:login'),
            {
                'username': 'testuser',
                'password': 'TestPass123!',
                'next': next_url,
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], next_url)

    def test_login_success_rejects_external_next(self):
        """외부 도메인 next는 Django safe redirect 검증으로 차단"""
        response = self.client.post(
            reverse('reporting:login'),
            {
                'username': 'testuser',
                'password': 'TestPass123!',
                'next': 'https://example.com/steal',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response['Location'],
            'https://sales-note-frontend-production.up.railway.app/dashboard/',
        )

    def test_login_fail_wrong_password(self):
        """잘못된 비밀번호로 로그인 실패"""
        response = self.client.post(
            reverse('reporting:login'),
            {'username': 'testuser', 'password': 'wrongpassword'},
        )
        self.assertEqual(response.status_code, 200)  # 로그인 페이지 재표시

    def test_followup_list_authenticated(self):
        """인증 후 거래처 목록은 React 고객 화면으로 이동"""
        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:followup_list'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], frontend_url('customers/'))

    def test_opportunity_list_url_removed(self):
        """별도 영업기회 목록 URL은 제거되어야 함"""
        self.client.force_login(self.user)
        response = self.client.get('/reporting/opportunities/')
        self.assertEqual(response.status_code, 404)

    def test_schedule_list_authenticated(self):
        """인증 후 일정 목록은 React 일정 화면으로 이동"""
        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:schedule_list'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], frontend_url('schedules/'))

    def test_schedule_calendar_authenticated(self):
        """인증 후 일정 캘린더는 React 캘린더로 이동"""
        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:schedule_calendar'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], frontend_url('schedules/calendar/'))

    def test_history_list_authenticated(self):
        """인증 후 영업 활동 목록은 React 영업노트 화면으로 이동"""
        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:history_list'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], frontend_url('notes/'))


class BackendReactFrontendServingTests(TestCase):
    """Django web service can serve the built React CRM shell directly."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.dist_dir = Path(self.temp_dir.name)
        (self.dist_dir / 'assets').mkdir()
        (self.dist_dir / 'index.html').write_text(
            '<!doctype html><div id="root"></div><script type="module" src="/assets/app.js"></script>',
            encoding='utf-8',
        )
        (self.dist_dir / 'assets' / 'app.js').write_text('console.log("crm");', encoding='utf-8')
        self.override = override_settings(FRONTEND_DIST_DIR=self.dist_dir)
        self.override.enable()

    def tearDown(self):
        self.override.disable()
        self.temp_dir.cleanup()

    def test_react_route_serves_index(self):
        response = self.client.get('/dashboard/')

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/html', response['Content-Type'])
        self.assertIn('no-cache', response['Cache-Control'])
        self.assertContains(response, '<div id="root"></div>', html=False)

    def test_react_asset_serves_immutable_gzip_static(self):
        response = self.client.get('/assets/app.js', HTTP_ACCEPT_ENCODING='br, gzip')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Encoding'], 'gzip')
        self.assertIn('max-age=31536000', response['Cache-Control'])
        self.assertIn('immutable', response['Cache-Control'])
        self.assertIn('text/javascript', response['Content-Type'])

    def test_removed_frontend_routes_stay_removed(self):
        response = self.client.get('/downloads/')

        self.assertEqual(response.status_code, 404)
        self.assertIn('text/plain', response['Content-Type'])

        for path in ('/mailbox/', '/business-cards/'):
            removed_response = self.client.get(path)
            self.assertEqual(removed_response.status_code, 404, path)

    def test_backend_api_is_not_intercepted_by_react_shell(self):
        response = self.client.get('/reporting/api/customers/')

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')


class CoreCrmLegacyRedirectTests(TestCase):
    """Core Django template pages should hand users to React during migration."""

    def setUp(self):
        self.client = Client()
        self.company_profile = UserCompany.objects.create(name='React전환회사')
        self.user = make_user('react_redirect_user', role='salesman', company=self.company_profile)
        self.company = Company.objects.create(name='React전환고객사', created_by=self.user)
        self.department = Department.objects.create(name='React전환부서', company=self.company, created_by=self.user)
        self.followup = FollowUp.objects.create(
            user=self.user,
            company=self.company,
            department=self.department,
            customer_name='React전환담당자',
        )
        self.schedule = Schedule.objects.create(
            user=self.user,
            company=self.company_profile,
            followup=self.followup,
            visit_date=timezone.localdate(),
            visit_time=time(9, 0),
            activity_type='customer_meeting',
        )
        self.history = History.objects.create(
            user=self.user,
            company=self.company_profile,
            followup=self.followup,
            action_type='customer_meeting',
            content='React 전환 테스트',
        )
        self.prepayment = Prepayment.objects.create(
            department=self.department,
            customer=self.followup,
            company=self.company,
            amount=100000,
            balance=80000,
            payment_date=timezone.localdate(),
            payment_method='transfer',
            payer_name='React전환입금자',
            created_by=self.user,
        )
        from reporting.models import Product
        self.product = Product.objects.create(
            product_code='REACT-PRODUCT-001',
            description='React 전환 제품',
            specification='전환규격',
            unit='EA',
            standard_price=1000,
            created_by=self.user,
        )
        self.personal_schedule = PersonalSchedule.objects.create(
            user=self.user,
            company=self.company_profile,
            title='React 개인 일정',
            content='React 개인 일정 내용',
            schedule_date=timezone.localdate(),
            schedule_time=time(10, 0),
        )
        self.document_template = DocumentTemplate.objects.create(
            company=self.company_profile,
            document_type='quotation',
            name='React 전환 견적서',
            file=SimpleUploadedFile(
                'react-template.xlsx',
                b'react document template',
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            file_type='xlsx',
            created_by=self.user,
        )
        self.addCleanup(self.document_template.file.delete, False)
        self.client.force_login(self.user)

    def assertReactRedirect(self, url, expected):
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], frontend_url(expected))

    def test_core_list_page_redirects(self):
        self.assertReactRedirect(reverse('reporting:dashboard'), 'dashboard/')
        self.assertReactRedirect(reverse('reporting:followup_list'), 'customers/')
        self.assertReactRedirect(reverse('reporting:history_list'), 'notes/')
        self.assertReactRedirect(reverse('reporting:schedule_list'), 'schedules/')
        self.assertReactRedirect(reverse('reporting:schedule_calendar'), 'schedules/calendar/')
        self.assertReactRedirect(reverse('reporting:funnel_pipeline'), 'pipeline/')
        self.assertReactRedirect(reverse('reporting:prepayment_list'), 'prepayments/')
        self.assertReactRedirect(reverse('reporting:product_list'), 'products/')
        self.assertReactRedirect(reverse('reporting:profile'), 'profile/')
        self.assertReactRedirect(reverse('reporting:profile_edit'), 'profile/?edit=1')

    def test_core_detail_page_redirects(self):
        self.assertReactRedirect(reverse('reporting:followup_detail', args=[self.followup.id]), f'customers/{self.followup.id}/')
        self.assertReactRedirect(reverse('reporting:customer_detail_report', args=[self.followup.id]), f'customers/{self.followup.id}/')
        self.assertReactRedirect(reverse('reporting:history_detail', args=[self.history.id]), f'notes/{self.history.id}/')
        self.assertReactRedirect(reverse('reporting:schedule_detail', args=[self.schedule.id]), f'schedules/{self.schedule.id}/')
        self.assertReactRedirect(reverse('reporting:prepayment_detail', args=[self.prepayment.id]), f'prepayments/{self.prepayment.id}/')
        self.assertReactRedirect(reverse('reporting:prepayment_edit', args=[self.prepayment.id]), f'prepayments/{self.prepayment.id}/edit/')
        self.assertReactRedirect(reverse('reporting:prepayment_delete', args=[self.prepayment.id]), f'prepayments/{self.prepayment.id}/?delete=1')
        self.assertReactRedirect(reverse('reporting:prepayment_transfer', args=[self.prepayment.id]), f'prepayments/{self.prepayment.id}/?transfer=1')
        self.assertReactRedirect(reverse('reporting:prepayment_customer', args=[self.followup.id]), f'prepayments/customer/{self.followup.id}/')
        self.assertReactRedirect(reverse('reporting:history_delete', args=[self.history.id]), f'notes/{self.history.id}/?delete=1')
        self.assertReactRedirect(reverse('reporting:schedule_delete', args=[self.schedule.id]), f'schedules/{self.schedule.id}/?delete=1')
        self.assertReactRedirect(
            reverse('reporting:personal_schedule_detail', args=[self.personal_schedule.id]),
            f'schedules/calendar/?personal={self.personal_schedule.id}',
        )
        self.assertReactRedirect(
            reverse('reporting:personal_schedule_edit', args=[self.personal_schedule.id]),
            f'schedules/calendar/?personal={self.personal_schedule.id}&edit=1',
        )
        self.assertReactRedirect(
            reverse('reporting:personal_schedule_delete', args=[self.personal_schedule.id]),
            f'schedules/calendar/?personal={self.personal_schedule.id}&delete=1',
        )
        self.assertReactRedirect(reverse('reporting:document_template_list'), 'documents/')
        self.assertReactRedirect(
            reverse('reporting:document_template_edit', args=[self.document_template.id]),
            f'documents/?template_id={self.document_template.id}&edit=1',
        )
        self.assertReactRedirect(
            reverse('reporting:document_template_delete', args=[self.document_template.id]),
            f'documents/?template_id={self.document_template.id}&delete=1',
        )
        self.assertReactRedirect(
            reverse('reporting:document_template_toggle_default', args=[self.document_template.id]),
            f'documents/?template_id={self.document_template.id}',
        )
        self.assertReactRedirect(
            reverse('reporting:product_edit', args=[self.product.id]),
            f'products/?product={self.product.id}&edit=1',
        )
        self.assertReactRedirect(
            reverse('reporting:product_delete', args=[self.product.id]),
            f'products/?product={self.product.id}&delete=1',
        )

    def test_core_create_page_redirects_preserve_relevant_query(self):
        self.assertReactRedirect(reverse('reporting:followup_create'), 'customers/?create=1')
        self.assertReactRedirect(
            f"{reverse('reporting:schedule_create')}?followup={self.followup.id}&date=2026-05-20",
            f'schedules/?customer={self.followup.id}&date=2026-05-20&create=1',
        )
        self.assertReactRedirect(
            reverse('reporting:history_create_from_schedule', args=[self.schedule.id]),
            f'notes/?create=1&schedule={self.schedule.id}',
        )
        self.assertReactRedirect(reverse('reporting:prepayment_create'), 'prepayments/new/')
        self.assertReactRedirect(
            f"{reverse('reporting:personal_schedule_create')}?date=2026-05-20&time=10:30",
            'schedules/calendar/?date=2026-05-20&time=10%3A30&create=personal',
        )
        self.assertReactRedirect(reverse('reporting:document_template_create'), 'documents/?create=1')
        self.assertReactRedirect(reverse('reporting:product_create'), 'products/?create=1')
        self.assertReactRedirect(reverse('reporting:product_bulk_create'), 'products/?import=1')

    def test_core_filter_query_is_translated(self):
        self.assertReactRedirect(
            f"{reverse('reporting:followup_list')}?pipeline_stage=quote&q=Kim",
            'customers/?stage=quote&q=Kim',
        )
        self.assertReactRedirect(
            f"{reverse('reporting:prepayment_list')}?search=입금자&status=active&data_filter=all",
            'prepayments/?q=%EC%9E%85%EA%B8%88%EC%9E%90&status=active&data_filter=all',
        )
        self.assertReactRedirect(
            f"{reverse('reporting:product_list')}?search=PCR&is_active=true&sort=delivery_count",
            'products/?q=PCR&status=active&sort=deliveryCount',
        )

    def test_user_management_legacy_routes_redirect_to_react_employees(self):
        self.assertReactRedirect(
            f"{reverse('reporting:user_list')}?search=kim&role=salesman",
            'employees/?q=kim&role=salesman',
        )
        self.assertReactRedirect(reverse('reporting:user_create'), 'employees/?create=1')
        self.assertReactRedirect(
            reverse('reporting:user_edit', args=[self.user.id]),
            f'employees/?employee={self.user.id}&edit=1',
        )
        self.assertReactRedirect(
            reverse('reporting:manager_user_list'),
            'employees/',
        )
        self.assertReactRedirect(
            reverse('reporting:manager_user_edit', args=[self.user.id]),
            f'employees/?employee={self.user.id}&edit=1',
        )

    def test_non_get_legacy_create_action_is_not_redirected_to_react(self):
        response = self.client.post(reverse('reporting:schedule_create'), {
            'followup': str(self.followup.id),
            'visit_date': '2026-05-20',
            'activity_type': 'customer_meeting',
        })
        self.assertNotEqual(response.get('Location', ''), frontend_url('schedules/?create=1'))


class ReactNavigationApiTests(TestCase):
    """React navigation API regression tests."""

    def setUp(self):
        self.client = Client()
        self.user = make_user('nav-user')

    def test_navigation_api_requires_login_json(self):
        response = self.client.get(reverse('reporting:navigation_api'))

        self.assertEqual(response.status_code, 401)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertEqual(payload['error'], 'login_required')

    def test_navigation_api_excludes_forbidden_menu_entries(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:navigation_api'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        items_by_id = {item['id']: item for item in payload['items']}
        self.assertNotIn('analytics', items_by_id)
        self.assertNotIn('dataCleanup', items_by_id)
        self.assertNotIn('downloads', items_by_id)
        self.assertNotIn('mail', items_by_id)
        self.assertNotIn('businessCards', items_by_id)
        self.assertNotIn('assets', items_by_id)
        self.assertNotIn('services', items_by_id)
        self.assertEqual(items_by_id['demos']['href'], '/demos/')
        self.assertEqual(items_by_id['demos']['label'], '데모관리')
        self.assertEqual(items_by_id['receivables']['href'], '/receivables/')
        self.assertEqual(items_by_id['receivables']['label'], '외상고객')
        self.assertEqual(items_by_id['pipelineSheet']['href'], '/pipeline-sheet/')
        self.assertEqual(items_by_id['pipelineSheet']['label'], '파이프라인 시트')
        self.assertEqual(items_by_id['profile']['href'], '/profile/')
        self.assertEqual(items_by_id['profile']['label'], '프로필')
        self.assertNotIn('employees', items_by_id)
        self.assertNotIn('userAdmin', items_by_id)
        self.assertFalse(payload['capabilities']['canManageUsers'])

    def test_navigation_api_includes_employee_management_for_manager_only(self):
        company = UserCompany.objects.create(name='직원관리메뉴회사')
        manager = make_user('nav-manager', role='manager', company=company)
        admin = make_user('nav-admin', role='admin', company=company)

        self.client.force_login(manager)
        manager_response = self.client.get(reverse('reporting:navigation_api'))
        self.assertEqual(manager_response.status_code, 200)
        manager_items = {item['id']: item for item in manager_response.json()['items']}
        self.assertIn('employees', manager_items)
        self.assertEqual(manager_items['employees']['label'], '직원관리')
        self.assertEqual(manager_items['employees']['href'], '/employees/')
        self.assertTrue(manager_response.json()['capabilities']['canManageEmployees'])
        self.assertNotIn('userAdmin', manager_items)
        self.assertFalse(manager_response.json()['capabilities']['canManageUsers'])

        self.client.force_login(admin)
        admin_response = self.client.get(reverse('reporting:navigation_api'))
        admin_items = {item['id']: item for item in admin_response.json()['items']}
        self.assertNotIn('employees', admin_items)
        self.assertTrue(admin_response.json()['capabilities']['canManageEmployees'])
        self.assertIn('userAdmin', admin_items)
        self.assertEqual(admin_items['userAdmin']['label'], '사용자관리')
        self.assertEqual(admin_items['userAdmin']['href'], '/employees/')
        self.assertTrue(admin_response.json()['capabilities']['canManageUsers'])

    def test_navigation_api_role_menu_differences(self):
        company = UserCompany.objects.create(name='권한별메뉴회사')
        salesman = make_user('nav-role-salesman', role='salesman', company=company)
        manager = make_user('nav-role-manager', role='manager', company=company)
        admin = make_user('nav-role-admin', role='admin', company=company)

        self.client.force_login(salesman)
        salesman_payload = self.client.get(reverse('reporting:navigation_api')).json()
        salesman_ids = {item['id'] for item in salesman_payload['items']}
        self.assertNotIn('dataCleanup', salesman_ids)
        self.assertNotIn('downloads', salesman_ids)
        self.assertIn('demos', salesman_ids)
        self.assertIn('receivables', salesman_ids)
        self.assertNotIn('mail', salesman_ids)
        self.assertNotIn('employees', salesman_ids)
        self.assertNotIn('userAdmin', salesman_ids)
        self.assertFalse(salesman_payload['capabilities']['canManageUsers'])

        self.client.force_login(manager)
        manager_payload = self.client.get(reverse('reporting:navigation_api')).json()
        manager_ids = {item['id'] for item in manager_payload['items']}
        self.assertNotIn('dataCleanup', manager_ids)
        self.assertNotIn('downloads', manager_ids)
        self.assertIn('demos', manager_ids)
        self.assertIn('receivables', manager_ids)
        self.assertIn('employees', manager_ids)
        self.assertNotIn('mail', manager_ids)
        self.assertNotIn('userAdmin', manager_ids)
        self.assertTrue(manager_payload['capabilities']['canManageEmployees'])

        self.client.force_login(admin)
        admin_payload = self.client.get(reverse('reporting:navigation_api')).json()
        admin_ids = {item['id'] for item in admin_payload['items']}
        self.assertNotIn('dataCleanup', admin_ids)
        self.assertNotIn('downloads', admin_ids)
        self.assertNotIn('mail', admin_ids)
        self.assertIn('demos', admin_ids)
        self.assertIn('receivables', admin_ids)
        self.assertIn('userAdmin', admin_ids)
        self.assertNotIn('employees', admin_ids)
        self.assertTrue(admin_payload['capabilities']['canManageEmployees'])
        self.assertTrue(admin_payload['capabilities']['canManageUsers'])


class DemoRecordsApiTests(TestCase):
    """Demo management API regression tests."""

    def setUp(self):
        self.client = Client()
        self.user_company = UserCompany.objects.create(name='데모테스트회사')
        self.user = make_user('demo-user', company=self.user_company)
        self.manager = make_user('demo-manager', role='manager', company=self.user_company)
        self.company = Company.objects.create(name='데모고객사', created_by=self.user)
        self.department = Department.objects.create(company=self.company, name='데모연구실', created_by=self.user)
        self.followup = FollowUp.objects.create(
            user=self.user,
            user_company=self.user_company,
            customer_name='데모담당자',
            company=self.company,
            department=self.department,
        )
        self.product = Product.objects.create(
            product_code='DEMO-PRODUCT-1',
            unit='EA',
            standard_price=1000000,
            created_by=self.user,
        )

    def test_demo_records_api_requires_login_json(self):
        response = self.client.get(reverse('reporting:demo_records_api'))

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_create_and_list_product_linked_demo_record(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:demo_record_create_api'),
            data=json.dumps({
                'departmentId': self.department.id,
                'customerId': self.followup.id,
                'productId': self.product.id,
                'quantity': 2,
                'status': 'active',
                'startDate': '2026-06-01',
                'expectedReturnDate': '2026-06-30',
                'notes': '현장 데모',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        created = DemoRecord.objects.get()
        self.assertEqual(created.product, self.product)
        self.assertEqual(created.product_name, self.product.product_code)
        self.assertEqual(created.followup, self.followup)

        list_response = self.client.get(reverse('reporting:demo_records_api'), {'status': 'all'})
        payload = list_response.json()
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(payload['summary']['total'], 1)
        self.assertEqual(payload['demos'][0]['productName'], self.product.product_code)
        self.assertEqual(payload['demos'][0]['customerName'], '데모담당자')
        self.assertEqual(payload['options']['accounts'][0]['departmentId'], self.department.id)
        self.assertEqual(payload['options']['products'][0]['id'], self.product.id)

    def test_customer_detail_includes_demo_summary(self):
        DemoRecord.objects.create(
            company=self.company,
            department=self.department,
            followup=self.followup,
            product=self.product,
            product_name=self.product.product_code,
            quantity=1,
            status='scheduled',
            owner=self.user,
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[self.followup.id]))

        self.assertEqual(response.status_code, 200)
        demo_summary = response.json()['demoSummary']
        self.assertEqual(demo_summary['metrics']['total'], 1)
        self.assertEqual(demo_summary['demos'][0]['productName'], self.product.product_code)
        self.assertEqual(demo_summary['links']['demos'], f'/demos/?department={self.department.id}')

    def test_manager_cannot_create_demo_record(self):
        self.client.force_login(self.manager)

        response = self.client.post(
            reverse('reporting:demo_record_create_api'),
            data=json.dumps({
                'departmentId': self.department.id,
                'productId': self.product.id,
                'quantity': 1,
                'status': 'active',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(DemoRecord.objects.count(), 0)


class RemovedStandaloneMenuRouteTests(TestCase):
    """Forbidden standalone menu routes should not resolve in Django."""

    def test_removed_backend_menu_routes_return_404(self):
        self.assertEqual(self.client.get('/reporting/data-cleanup/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/downloads/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/api/downloads/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/api/accounts/10/cleanup-preview/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/mailbox/inbox/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/api/mailbox/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/business-cards/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/api/business-cards/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/gmail/connect/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/imap/connect/').status_code, 404)
        self.assertEqual(self.client.get('/reporting/api/profile/imap/connect/').status_code, 404)

    def test_removed_account_cleanup_preview_frontend_route_returns_404(self):
        self.assertEqual(self.client.get('/accounts/10/cleanup-preview/').status_code, 404)


class ReceivablesApiTests(TestCase):
    """외상고객 React API tests."""

    def setUp(self):
        self.client = Client()
        self.user_company = UserCompany.objects.create(name='외상API소속')
        self.company = Company.objects.create(name='외상API업체', created_by=None)
        self.department = Department.objects.create(company=self.company, name='외상API부서')
        self.user = make_user('receivable-owner', company=self.user_company)
        self.manager = make_user('receivable-manager', role='manager', company=self.user_company)
        self.other_company = UserCompany.objects.create(name='외상API타사')
        self.other_user = make_user('receivable-other', company=self.other_company)
        self.followup = FollowUp.objects.create(
            user=self.user,
            user_company=self.user_company,
            company=self.company,
            department=self.department,
            customer_name='외상 담당자',
            manager='외상 PI',
        )
        self.schedule = Schedule.objects.create(
            user=self.user,
            company=self.user_company,
            followup=self.followup,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='delivery',
            status='completed',
        )

    def test_receivables_api_lists_open_items_and_summary(self):
        DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Open Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=True,
        )
        DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Default Credit Kit',
            quantity=1,
            unit='EA',
            unit_price=50000,
            tax_invoice_issued=False,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:receivables_api'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['summary']['itemCount'], 2)
        self.assertEqual(payload['summary']['openItemCount'], 2)
        self.assertEqual(payload['summary']['totalOutstanding'], 165000)
        self.assertEqual(payload['summary']['totalCreditAmount'], 165000)
        self.assertNotIn('unregistered', [status['value'] for status in payload['filters']['statuses']])
        item_names = [item['itemName'] for item in payload['items']]
        self.assertIn('Open Kit', item_names)
        self.assertIn('Default Credit Kit', item_names)
        default_item = next(item for item in payload['items'] if item['itemName'] == 'Default Credit Kit')
        self.assertTrue(default_item['taxInvoiceIssued'])
        self.assertEqual(default_item['statusLabel'], '외상 진행중')
        self.assertEqual(default_item['outstandingAmount'], 55000)
        self.assertEqual(payload['customers'][0]['outstandingAmount'], 165000)

    def test_receivables_api_excludes_prepayment_delivery_items(self):
        DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Normal Credit Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=True,
        )
        prepayment_schedule = Schedule.objects.create(
            user=self.user,
            company=self.user_company,
            followup=self.followup,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(11, 0),
            activity_type='delivery',
            status='completed',
            use_prepayment=True,
            prepayment_amount=250000,
            delivery_payment_type=Schedule.DELIVERY_PAYMENT_TYPE_PREPAYMENT,
            delivery_payment_status=Schedule.DELIVERY_PAYMENT_STATUS_PREPAYMENT,
        )
        DeliveryItem.objects.create(
            schedule=prepayment_schedule,
            item_name='Prepaid Kit',
            quantity=1,
            unit='EA',
            unit_price=250000,
            tax_invoice_issued=True,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:receivables_api'), {'status': 'all'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        item_names = [item['itemName'] for item in payload['items']]
        self.assertIn('Normal Credit Kit', item_names)
        self.assertNotIn('Prepaid Kit', item_names)
        self.assertEqual(payload['summary']['itemCount'], 1)
        self.assertEqual(payload['summary']['totalOutstanding'], 110000)

    def test_receivable_item_status_api_blocks_prepayment_usage_item(self):
        item = DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Prepayment Usage Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=False,
        )
        prepayment = Prepayment.objects.create(
            department=self.department,
            customer=self.followup,
            company=self.company,
            amount=200000,
            balance=90000,
            payment_date=timezone.localdate(),
            payment_method='transfer',
            payer_name='외상API 선결제',
            created_by=self.user,
        )
        PrepaymentUsage.objects.create(
            prepayment=prepayment,
            schedule=self.schedule,
            schedule_item=item,
            product_name=item.item_name,
            quantity=item.quantity,
            amount=110000,
            remaining_balance=90000,
        )
        url = reverse('reporting:receivable_item_status_api', args=[item.id])
        self.client.force_login(self.user)

        response = self.client.post(
            url,
            data=json.dumps({'taxInvoiceIssued': True}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 409)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertIn('선결제', payload['error'])
        item.refresh_from_db()
        self.assertFalse(item.tax_invoice_issued)

    def test_receivable_item_status_api_checks_card_and_cancels(self):
        item = DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Status Kit',
            quantity=2,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=False,
        )
        history = History.objects.create(
            user=self.user,
            company=self.user_company,
            followup=self.followup,
            schedule=self.schedule,
            action_type='delivery_schedule',
            tax_invoice_issued=False,
        )
        url = reverse('reporting:receivable_item_status_api', args=[item.id])
        self.client.force_login(self.user)

        response = self.client.post(
            url,
            data=json.dumps({'cardPaymentReceived': True}),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        item.refresh_from_db()
        history.refresh_from_db()
        self.assertTrue(item.tax_invoice_issued)
        self.assertTrue(item.card_payment_received)
        self.assertTrue(item.receivable_settled)
        self.assertTrue(history.tax_invoice_issued)
        self.assertEqual(response.json()['item']['outstandingAmount'], 0)
        self.assertEqual(response.json()['item']['statusLabel'], '카드결제 완료')

        card_response = self.client.post(
            url,
            data=json.dumps({'cardPaymentReceived': False}),
            content_type='application/json',
        )
        self.assertEqual(card_response.status_code, 200)
        item.refresh_from_db()
        self.assertTrue(item.tax_invoice_issued)
        self.assertFalse(item.card_payment_received)
        self.assertFalse(item.receivable_settled)
        self.assertEqual(card_response.json()['item']['outstandingAmount'], 220000)
        self.assertEqual(card_response.json()['item']['statusLabel'], '외상 진행중')

        cancel_response = self.client.post(
            url,
            data=json.dumps({'taxInvoiceIssued': False}),
            content_type='application/json',
        )
        self.assertEqual(cancel_response.status_code, 200)
        item.refresh_from_db()
        self.assertTrue(item.tax_invoice_issued)
        self.assertFalse(item.card_payment_received)
        self.assertFalse(item.receivable_settled)
        self.assertEqual(cancel_response.json()['item']['outstandingAmount'], 220000)

    def test_receivable_item_status_api_locks_item_without_nullable_outer_joins(self):
        item = DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Join Safe Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=True,
        )
        url = reverse('reporting:receivable_item_status_api', args=[item.id])
        self.client.force_login(self.user)

        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        with CaptureQueriesContext(connection) as captured:
            response = self.client.post(
                url,
                data=json.dumps({'cardPaymentReceived': True}),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 200)
        item.refresh_from_db()
        self.assertTrue(item.card_payment_received)
        delivery_item_queries = [
            query['sql']
            for query in captured.captured_queries
            if 'FROM "reporting_deliveryitem"' in query['sql']
        ]
        self.assertTrue(delivery_item_queries)
        self.assertNotIn('JOIN', delivery_item_queries[0].upper())

    def test_receivable_item_status_api_blocks_manager_and_other_company(self):
        item = DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='Blocked Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
        )
        url = reverse('reporting:receivable_item_status_api', args=[item.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(
            url,
            data=json.dumps({'taxInvoiceIssued': True}),
            content_type='application/json',
        )
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.other_user)
        other_response = self.client.post(
            url,
            data=json.dumps({'taxInvoiceIssued': True}),
            content_type='application/json',
        )
        self.assertEqual(other_response.status_code, 403)


class EmployeeManagementApiTests(TestCase):
    """React user/employee management API tests for admin and manager roles."""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='직원관리API회사')
        self.other_company = UserCompany.objects.create(name='직원관리API타사회사')
        self.manager = make_user('employee-api-manager', role='manager', company=self.company)
        self.salesman = make_user('employee-api-sales', role='salesman', company=self.company)
        self.coworker = make_user('employee-api-coworker', role='salesman', company=self.company)
        self.other_user = make_user('employee-api-other', role='salesman', company=self.other_company)
        self.url = reverse('reporting:employees_management_api')

    def test_employee_management_api_requires_manager(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 401)

        self.client.force_login(self.salesman)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['error'], 'management_required')

    def test_employee_management_api_lists_same_company_people(self):
        self.coworker.first_name = '길동'
        self.coworker.last_name = '홍'
        self.coworker.email = 'coworker@example.com'
        self.coworker.save(update_fields=['first_name', 'last_name', 'email'])
        profile = self.coworker.userprofile
        profile.can_download_excel = True
        profile.can_use_ai = True
        profile.created_by = self.manager
        profile.save(update_fields=['can_download_excel', 'can_use_ai', 'created_by'])
        self.salesman.is_active = False
        self.salesman.save(update_fields=['is_active'])
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['scope']['canManage'])
        self.assertEqual(payload['scope']['companyName'], self.company.name)
        ids = {item['id'] for item in payload['employees']}
        self.assertIn(self.manager.id, ids)
        self.assertIn(self.salesman.id, ids)
        self.assertIn(self.coworker.id, ids)
        self.assertNotIn(self.other_user.id, ids)
        self.assertEqual(payload['metrics']['totalEmployees'], 3)
        self.assertEqual(payload['metrics']['inactiveEmployees'], 1)
        coworker_payload = next(item for item in payload['employees'] if item['id'] == self.coworker.id)
        self.assertEqual(coworker_payload['name'], '길동 홍')
        self.assertTrue(coworker_payload['canDownloadExcel'])
        self.assertTrue(coworker_payload['canUseAi'])
        self.assertEqual(coworker_payload['createdByName'], self.manager.username)
        self.assertEqual(coworker_payload['editHref'], f'/employees/?employee={self.coworker.id}&edit=1')
        self.assertEqual(coworker_payload['updateHref'], reverse('reporting:employees_update_api', args=[self.coworker.id]))
        self.assertEqual(coworker_payload['toggleActiveHref'], reverse('reporting:employees_toggle_active_api', args=[self.coworker.id]))
        manager_payload = next(item for item in payload['employees'] if item['id'] == self.manager.id)
        self.assertEqual(manager_payload['editHref'], '')
        self.assertTrue(manager_payload['isCurrentUser'])

    def test_employee_management_api_filters_by_search_and_role(self):
        self.coworker.first_name = '필터'
        self.coworker.last_name = '대상'
        self.coworker.save(update_fields=['first_name', 'last_name'])
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {'q': '필터', 'role': 'salesman'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['filters']['q'], '필터')
        self.assertEqual(payload['filters']['role'], 'salesman')
        self.assertEqual([item['id'] for item in payload['employees']], [self.coworker.id])

    def test_admin_employee_management_api_lists_all_and_filters_company(self):
        admin = make_user('employee-api-admin', role='admin')
        self.client.force_login(admin)

        response = self.client.get(self.url, {'company': str(self.other_company.id)})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['scope']['mode'], 'admin')
        self.assertTrue(payload['scope']['canChangeRole'])
        self.assertEqual(payload['filters']['company'], str(self.other_company.id))
        ids = {item['id'] for item in payload['employees']}
        self.assertIn(self.other_user.id, ids)
        self.assertNotIn(self.salesman.id, ids)
        self.assertGreaterEqual(len(payload['options']['companies']), 2)

    def test_manager_can_create_salesman_only_in_own_company(self):
        self.client.force_login(self.manager)
        response = self.client.post(
            reverse('reporting:employees_create_api'),
            data=json.dumps({
                'username': 'employee-api-created',
                'password': 'secret-password',
                'passwordConfirm': 'secret-password',
                'role': 'manager',
                'companyName': self.other_company.name,
                'canUseAi': True,
                'canDownloadExcel': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = User.objects.get(username='employee-api-created')
        profile = created.userprofile
        self.assertEqual(profile.role, 'salesman')
        self.assertEqual(profile.company_id, self.company.id)
        self.assertFalse(profile.can_use_ai)
        self.assertTrue(profile.can_download_excel)
        self.assertEqual(payload['employee']['companyId'], self.company.id)

    def test_manager_cannot_update_other_company_employee(self):
        self.client.force_login(self.manager)
        response = self.client.post(
            reverse('reporting:employees_update_api', args=[self.other_user.id]),
            data=json.dumps({'username': 'blocked-other-company'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 404)

    def test_manager_update_keeps_role_company_and_ai_fixed(self):
        self.client.force_login(self.manager)
        response = self.client.post(
            reverse('reporting:employees_update_api', args=[self.coworker.id]),
            data=json.dumps({
                'username': 'employee-api-coworker-updated',
                'role': 'manager',
                'companyName': self.other_company.name,
                'canUseAi': True,
                'canDownloadExcel': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.coworker.refresh_from_db()
        profile = self.coworker.userprofile
        self.assertEqual(self.coworker.username, 'employee-api-coworker-updated')
        self.assertEqual(profile.role, 'salesman')
        self.assertEqual(profile.company_id, self.company.id)
        self.assertFalse(profile.can_use_ai)
        self.assertTrue(profile.can_download_excel)

    def test_admin_can_create_update_and_toggle_user(self):
        admin = make_user('employee-api-admin-mutate', role='admin')
        self.client.force_login(admin)
        create_response = self.client.post(
            reverse('reporting:employees_create_api'),
            data=json.dumps({
                'username': 'employee-api-admin-created',
                'password': 'secret-password',
                'passwordConfirm': 'secret-password',
                'role': 'manager',
                'companyId': str(self.company.id),
                'canUseAi': True,
                'canDownloadExcel': True,
            }),
            content_type='application/json',
        )
        self.assertEqual(create_response.status_code, 201)
        created = User.objects.get(username='employee-api-admin-created')
        self.assertEqual(created.userprofile.role, 'manager')
        self.assertTrue(created.userprofile.can_use_ai)

        update_response = self.client.post(
            reverse('reporting:employees_update_api', args=[created.id]),
            data=json.dumps({
                'username': 'employee-api-admin-updated',
                'role': 'salesman',
                'companyName': self.other_company.name,
                'canUseAi': False,
                'canDownloadExcel': False,
            }),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        created.refresh_from_db()
        created.userprofile.refresh_from_db()
        self.assertEqual(created.username, 'employee-api-admin-updated')
        self.assertEqual(created.userprofile.role, 'salesman')
        self.assertEqual(created.userprofile.company_id, self.other_company.id)
        self.assertFalse(created.userprofile.can_use_ai)

        toggle_response = self.client.post(
            reverse('reporting:employees_toggle_active_api', args=[created.id]),
            data=json.dumps({'isActive': False}),
            content_type='application/json',
        )
        self.assertEqual(toggle_response.status_code, 200)
        created.refresh_from_db()
        self.assertFalse(created.is_active)


class SalesNoteReadonlyBearerApiTests(TestCase):
    """Readonly MCP bearer token access should cover safe GET API surfaces only."""

    token = 'readonly-test-token'

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='Readonly MCP 회사')
        self.readonly_user = make_user(
            'readonly-mcp-admin',
            role='admin',
            company=self.company,
            can_use_ai=True,
        )

    def _auth_headers(self):
        return {'HTTP_AUTHORIZATION': f'Bearer {self.token}'}

    @patch.dict(os.environ, {
        'SALES_NOTE_READONLY_TOKEN': token,
        'SALES_NOTE_READONLY_USERNAME': 'readonly-mcp-admin',
    })
    def test_readonly_bearer_can_read_expanded_get_apis(self):
        endpoints = [
            reverse('reporting:navigation_api'),
            reverse('reporting:dashboard_summary_api'),
            reverse('reporting:customers_summary_api'),
            reverse('reporting:notes_summary_api'),
            reverse('reporting:schedules_summary_api'),
            reverse('reporting:prepayment_api_list'),
            reverse('reporting:product_api_list'),
            reverse('reporting:products_management_api'),
            reverse('reporting:document_templates_api'),
        ]

        for endpoint in endpoints:
            with self.subTest(endpoint=endpoint):
                response = self.client.get(endpoint, **self._auth_headers())
                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertTrue(payload.get('success', True), payload)

    @patch.dict(os.environ, {
        'SALES_NOTE_READONLY_TOKEN': token,
        'SALES_NOTE_READONLY_USERNAME': 'readonly-mcp-admin',
    })
    def test_readonly_bearer_does_not_allow_writes(self):
        response = self.client.post(reverse('reporting:notes_create_api'), **self._auth_headers())

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')


class ReactReportsProfileBusinessCardApiTests(TestCase):
    """Reports/profile/business card React API regression tests."""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='Hana CRM')
        self.user = make_user(
            'react-api-user',
            password='TestPass123!',
            company=self.company,
            can_download_excel=True,
        )
        self.manager = make_user('react-api-manager', role='manager', company=self.company)
        self.other = make_user('react-api-other', company=self.company)
        self.customer_company = Company.objects.create(name='고객사 A', created_by=self.user)
        self.department = Department.objects.create(
            company=self.customer_company,
            name='연구실 A',
            created_by=self.user,
        )
        self.followup = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            company=self.customer_company,
            department=self.department,
            customer_name='김고객',
            email='customer@example.com',
            pipeline_stage='quote',
        )

    def test_common_account_ledger_feeds_reports_customer_detail_and_ai(self):
        from ai_chat.services import gather_prepayment_data, gather_quote_delivery_data
        from reporting.services.account_ledger import account_operational_ledger_for_followups

        today = timezone.localdate()
        create_account_ledger_fixture(
            self.user,
            user_company=self.company,
            company=self.customer_company,
            department=self.department,
            today=today,
            prefix='ledgercommon',
        )
        shared_followups = FollowUp.objects.filter(user=self.user, department=self.department)
        service_ledger = account_operational_ledger_for_followups(
            shared_followups,
            [self.user],
            actor=self.user,
            record_limit=None,
        )
        service_metrics = service_ledger['metrics']
        self.assertEqual(service_metrics['deliveryRecords'], 2)
        self.assertEqual(service_metrics['deliveryAmount'], 90000)
        self.assertEqual(service_metrics['prepaymentDeliveryRecords'], 1)
        self.assertEqual(service_metrics['prepaymentUsedAmount'], 60000)
        self.assertEqual(service_metrics['quoteRecords'], 1)
        self.assertEqual(service_metrics['quoteAmount'], 110000)
        self.assertEqual(service_metrics['prepaymentRecords'], 1)
        self.assertEqual(service_metrics['prepaymentBalance'], 40000)

        self.client.force_login(self.user)
        account_response = self.client.get(reverse('reporting:account_detail_summary_api', args=[self.department.id]))
        self.assertEqual(account_response.status_code, 200)
        account_metrics = account_response.json()['operationalRecords']['metrics']
        self.assertEqual(account_metrics['deliveryRecords'], service_metrics['deliveryRecords'])
        self.assertEqual(account_metrics['quoteRecords'], service_metrics['quoteRecords'])
        self.assertEqual(account_metrics['prepaymentRecords'], service_metrics['prepaymentRecords'])
        self.assertEqual(account_metrics['prepaymentUsedAmount'], service_metrics['prepaymentUsedAmount'])

        ai_quote_delivery = gather_quote_delivery_data(self.department, self.user)
        self.assertEqual(ai_quote_delivery['summary']['total_deliveries'], service_metrics['deliveryRecords'])
        self.assertEqual(ai_quote_delivery['summary']['total_delivery_amount'], service_metrics['deliveryAmount'])
        self.assertEqual(ai_quote_delivery['summary']['total_quotes'], service_metrics['quoteRecords'])
        self.assertEqual(ai_quote_delivery['summary']['total_quote_amount'], service_metrics['quoteAmount'])
        self.assertIn('common_account_ledger', {row['ledgerSource'] for row in ai_quote_delivery['deliveries']})

        ai_prepayments = gather_prepayment_data(shared_followups)
        self.assertEqual(ai_prepayments['summary']['total_count'], service_metrics['prepaymentRecords'])
        self.assertEqual(ai_prepayments['summary']['total_remaining_balance'], service_metrics['prepaymentBalance'])

    def test_profile_api_update_and_password_change(self):
        self.client.force_login(self.user)

        update_response = self.client.post(
            reverse('reporting:profile_api_update'),
            data=json.dumps({
                'username': 'react-profile-user',
                'firstName': '길동',
                'lastName': '홍',
                'email': 'profile@example.com',
            }),
            content_type='application/json',
        )
        self.assertEqual(update_response.status_code, 200)
        self.user.refresh_from_db()
        self.assertEqual(self.user.username, 'react-profile-user')
        self.assertEqual(self.user.email, 'profile@example.com')

        password_response = self.client.post(
            reverse('reporting:profile_api_password'),
            data=json.dumps({
                'oldPassword': 'TestPass123!',
                'newPassword1': 'NewPass12345!',
                'newPassword2': 'NewPass12345!',
            }),
            content_type='application/json',
        )
        self.assertEqual(password_response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('NewPass12345!'))


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: 익명 사용자 URL 차단 테스트
# ─────────────────────────────────────────────────────────────────────────────

class AnonymousAccessTests(TestCase):
    """익명 사용자가 모든 내부 CRM 페이지에 접근할 수 없음을 검증"""

    def setUp(self):
        self.client = Client()

    def _assert_redirects_to_login(self, url):
        response = self.client.get(url)
        self.assertIn(
            response.status_code, [301, 302],
            msg=f"Expected redirect for anonymous access to {url}, got {response.status_code}"
        )
        location = response.get('Location', '')
        self.assertIn('login', location, msg=f"Redirect target should be login for {url}, got {location}")

    def test_dashboard_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:dashboard'))

    def test_followup_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:followup_list'))

    def test_history_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:history_list'))

    def test_schedule_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:schedule_list'))

    def test_schedule_calendar_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:schedule_calendar'))

    def test_funnel_pipeline_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:funnel_pipeline'))

    def test_document_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:document_template_list'))

    def test_followup_excel_download_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:followup_excel_download'))

    def test_followup_basic_excel_download_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:followup_basic_excel_download'))

    def test_prepayment_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:prepayment_list'))

    def test_user_list_blocked(self):
        self._assert_redirects_to_login(reverse('reporting:user_list'))


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: export 권한 테스트 (salesman=403, manager=200, admin=200)
# ─────────────────────────────────────────────────────────────────────────────

class ExportPermissionTests(TestCase):
    """CSV/XLSX export 뷰의 역할별 권한 차단 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='테스트회사')
        self.salesman = make_user('salesman1', role='salesman', company=self.company)
        self.manager = make_user('manager1', role='manager', company=self.company)
        self.admin = make_user('admin1', role='admin', company=self.company)

    def test_followup_excel_salesman_blocked(self):
        """followup excel download: can_download_excel=False salesman 차단"""
        self.client.force_login(self.salesman)
        r = self.client.get(reverse('reporting:followup_excel_download'))
        # 권한 없으면 리다이렉트
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Salesman without excel perm should be blocked, got {r.status_code}")

    def test_followup_basic_excel_salesman_blocked(self):
        """followup basic excel download: can_download_excel=False salesman 차단"""
        self.client.force_login(self.salesman)
        r = self.client.get(reverse('reporting:followup_basic_excel_download'))
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Salesman without excel perm should be blocked, got {r.status_code}")

    def test_followup_excel_admin_allowed(self):
        """followup excel download: admin 허용"""
        self.client.force_login(self.admin)
        r = self.client.get(reverse('reporting:followup_excel_download'))
        self.assertEqual(r.status_code, 200)

    def test_followup_basic_excel_admin_allowed(self):
        """followup basic excel download: admin 허용"""
        self.client.force_login(self.admin)
        r = self.client.get(reverse('reporting:followup_basic_excel_download'))
        self.assertEqual(r.status_code, 200)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: AI 권한 테스트
# ─────────────────────────────────────────────────────────────────────────────

class AIPermissionTests(TestCase):
    """AI 기능 접근 권한 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='테스트AI회사')
        self.user_no_ai = make_user('no_ai_user', role='salesman',
                                    can_use_ai=False, company=self.company)
        self.user_with_ai = make_user('ai_user', role='salesman',
                                      can_use_ai=True, company=self.company)

    def test_ai_departments_blocked_without_permission(self):
        """can_use_ai=False 사용자는 AI 분석 페이지에서 리다이렉트"""
        self.client.force_login(self.user_no_ai)
        r = self.client.get('/ai/')
        # ai_permission_required 데코레이터가 대시보드로 리다이렉트
        self.assertIn(r.status_code, [302, 403],
                      msg=f"User without AI perm should be blocked, got {r.status_code}")

    def test_ai_departments_accessible_with_permission(self):
        """can_use_ai=True 사용자는 React AI 브리핑 화면으로 이동"""
        self.client.force_login(self.user_with_ai)
        r = self.client.get('/ai/')
        self.assertEqual(r.status_code, 302,
                         msg=f"User with AI perm should redirect, got {r.status_code}")
        self.assertEqual(r['Location'], '/ai-workspace/')


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: 대시보드 smoke 테스트
# ─────────────────────────────────────────────────────────────────────────────

class DashboardSmokeTests(TestCase):
    """대시보드 legacy URL 전환 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='대시보드테스트회사')
        self.user = make_user('dash_user', role='salesman', company=self.company)

    def test_dashboard_returns_200(self):
        """인증 후 대시보드는 React 대시보드로 이동"""
        self.client.force_login(self.user)
        r = self.client.get(reverse('reporting:dashboard'))
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r['Location'], frontend_url('dashboard/'))

    def test_dashboard_head_redirects_to_react(self):
        """HEAD 요청도 템플릿 렌더링 대신 React로 이동"""
        self.client.force_login(self.user)
        r = self.client.head(reverse('reporting:dashboard'))
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r['Location'], frontend_url('dashboard/'))

    def test_dashboard_unauthenticated_redirects(self):
        """미인증 대시보드 접근 → 로그인 리다이렉트"""
        r = self.client.get(reverse('reporting:dashboard'))
        self.assertIn(r.status_code, [301, 302])
        self.assertIn('login', r.get('Location', ''))

    def test_dashboard_api_still_returns_key_sections(self):
        """대시보드 데이터는 React API에서 계속 제공"""
        self.client.force_login(self.user)
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        self.assertEqual(r.status_code, 200)
        payload = r.json()
        self.assertEqual(payload['source'], 'django')
        self.assertIn('metrics', payload)
        self.assertEqual(payload['links']['operationalDashboard'], '/dashboard/')


class DashboardSummaryApiTests(TestCase):
    """React 대시보드 읽기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='대시보드API회사')
        self.other_company = UserCompany.objects.create(name='대시보드API타사회사')
        self.user = make_user('dash_api_me', role='salesman', company=self.company)
        self.coworker = make_user('dash_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('dash_api_manager', role='manager', company=self.company)
        self.other_user = make_user('dash_api_other', role='salesman', company=self.other_company)
        self.url = reverse('reporting:dashboard_summary_api')

    def _create_customer(self, owner, name, overdue=True, today_schedule=True):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import Company, Department, FollowUp, History, Schedule

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            company=customer_company,
            department=department,
            priority='urgent',
            pipeline_stage='quote',
            customer_grade='A',
            ai_score=82,
        )
        if today_schedule:
            Schedule.objects.create(
                user=owner,
                company=owner.userprofile.company,
                followup=followup,
                visit_date=timezone.localdate(),
                visit_time=time(10, 0),
                status='scheduled',
                activity_type='customer_meeting',
                notes='오늘 미팅',
            )
        History.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            action_type='customer_meeting',
            content=f'{name} 미팅 기록',
            next_action='후속 전화',
            next_action_date=timezone.localdate() - timedelta(days=1) if overdue else timezone.localdate(),
        )
        return followup

    def test_dashboard_summary_api_requires_login_json(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_dashboard_summary_api_uses_salesman_own_scope(self):
        own = self._create_customer(self.user, '내고객')
        coworker = self._create_customer(self.coworker, '동료고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['source'], 'django')
        self.assertEqual(payload['metrics']['totalCustomers'], 1)
        priority_ids = {item['id'] for item in payload['priorityCustomers']}
        self.assertIn(own.id, priority_ids)
        self.assertNotIn(coworker.id, priority_ids)

    def test_dashboard_summary_api_includes_dashboard_sections(self):
        followup = self._create_customer(self.user, '요약고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['metrics']['todaySchedules'], 1)
        self.assertEqual(payload['metrics']['overdueActions'], 1)
        self.assertEqual(payload['today']['items'][0]['customer'], '요약고객 담당자')
        self.assertEqual(payload['overdueActions'][0]['nextAction'], '후속 전화')
        self.assertEqual(payload['recentActivities'][0]['customer'], '요약고객 담당자')
        self.assertTrue(any(item['stage'] == followup.pipeline_stage for item in payload['pipelineSummary']))
        self.assertEqual(payload['links']['createNote'], '/notes/?create=1')

    def test_dashboard_summary_api_excludes_stale_quote_submission_followups(self):
        from datetime import time, timedelta
        from decimal import Decimal
        from reporting.models import History

        today = timezone.localdate()
        followup = self._create_customer(
            self.user,
            '견적완료',
            overdue=False,
            today_schedule=False,
        )
        History.objects.filter(followup=followup).update(reviewed_at=timezone.now())
        stale_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=followup,
            action_type='customer_meeting',
            content='견적서 제출 예정',
            next_action='견적서 및 비교표 제출',
            next_action_date=today - timedelta(days=1),
        )
        active_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=followup,
            action_type='customer_meeting',
            content='제출된 견적 검토 상황 확인',
            next_action='견적 검토 여부 확인',
            next_action_date=today - timedelta(days=1),
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=followup,
            visit_date=today,
            visit_time=time(10, 0),
            status='scheduled',
            activity_type='quote',
            expected_revenue=Decimal('1200000'),
        )
        DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=quote_schedule,
            user=self.user,
            transaction_number='DASH-ST-Q-001',
            output_format='pdf',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        overdue_ids = {item['id'] for item in payload['overdueActions']}
        self.assertNotIn(stale_history.id, overdue_ids)
        self.assertIn(active_history.id, overdue_ids)
        self.assertEqual(payload['metrics']['overdueActions'], 1)

    def test_dashboard_summary_api_manager_sees_same_company_only(self):
        own = self._create_customer(self.user, '회사내고객')
        coworker = self._create_customer(self.coworker, '회사내동료')
        other = self._create_customer(self.other_user, '타사고객')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['metrics']['totalCustomers'], 2)
        priority_ids = {item['id'] for item in payload['priorityCustomers']}
        self.assertIn(own.id, priority_ids)
        self.assertIn(coworker.id, priority_ids)
        self.assertNotIn(other.id, priority_ids)
        self.assertTrue(payload['scope']['canViewAll'])

    def test_dashboard_summary_api_includes_year_and_quarter_revenue(self):
        from datetime import date, time
        from decimal import Decimal
        from reporting.models import DeliveryItem, Prepayment, Schedule

        today = timezone.localdate()
        quarter = ((today.month - 1) // 3) + 1
        quarter_start_month = ((quarter - 1) * 3) + 1
        followup = self._create_customer(
            self.user,
            '매출고객',
            overdue=False,
            today_schedule=False,
        )
        coworker_followup = self._create_customer(
            self.coworker,
            '동료매출',
            overdue=False,
            today_schedule=False,
        )

        def create_delivery(owner, target_followup, visit_date, unit_price):
            schedule = Schedule.objects.create(
                user=owner,
                company=owner.userprofile.company,
                followup=target_followup,
                visit_date=visit_date,
                visit_time=time(11, 0),
                status='completed',
                activity_type='delivery',
            )
            DeliveryItem.objects.create(
                schedule=schedule,
                item_name='납품품목',
                quantity=1,
                unit_price=Decimal(str(unit_price)),
            )
            return int(Decimal(str(unit_price)) * Decimal('1.1'))

        def create_prepayment(owner, target_followup, payment_date, amount, status='active'):
            Prepayment.objects.create(
                created_by=owner,
                customer=target_followup,
                department=target_followup.department,
                company=target_followup.company,
                amount=Decimal(str(amount)),
                balance=Decimal(str(amount)),
                payment_date=payment_date,
                status=status,
            )
            return int(Decimal(str(amount)))

        expected_year = 0
        expected_quarter = 0
        expected_month = create_delivery(self.user, followup, today, 100000)
        expected_year += expected_month
        expected_quarter += expected_month

        quarter_delivery = create_delivery(
            self.user,
            followup,
            date(today.year, quarter_start_month, 1),
            200000,
        )
        expected_year += quarter_delivery
        expected_quarter += quarter_delivery
        if today.month == quarter_start_month:
            expected_month += quarter_delivery

        if quarter_start_month > 1:
            expected_year += create_delivery(
                self.user,
                followup,
                date(today.year, 1, 15),
                300000,
            )

        create_delivery(self.user, followup, date(today.year - 1, 12, 15), 400000)
        create_delivery(self.coworker, coworker_followup, today, 500000)

        monthly_prepayment = create_prepayment(self.user, followup, today, 70000)
        expected_year += monthly_prepayment
        expected_quarter += monthly_prepayment
        expected_month += monthly_prepayment

        quarter_prepayment = create_prepayment(
            self.user,
            followup,
            date(today.year, quarter_start_month, 1),
            80000,
            status='depleted',
        )
        expected_year += quarter_prepayment
        expected_quarter += quarter_prepayment
        if today.month == quarter_start_month:
            expected_month += quarter_prepayment

        create_prepayment(self.user, followup, today, 90000, status='cancelled')
        create_prepayment(self.user, followup, date(today.year - 1, 12, 15), 100000)
        create_prepayment(self.coworker, coworker_followup, today, 500000)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['metrics']['yearRevenue'], expected_year)
        self.assertEqual(payload['metrics']['quarterRevenue'], expected_quarter)
        self.assertEqual(payload['metrics']['monthlyRevenue'], expected_month)
        self.assertEqual(payload['revenuePeriod']['year'], today.year)
        self.assertEqual(payload['revenuePeriod']['quarter'], quarter)


class CustomersSummaryApiTests(TestCase):
    """React 고객 화면 읽기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='고객API회사')
        self.other_company = UserCompany.objects.create(name='고객API타사회사')
        self.user = make_user('customers_api_me', role='salesman', company=self.company)
        self.coworker = make_user('customers_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('customers_api_manager', role='manager', company=self.company)
        self.admin = make_user('customers_api_admin', role='admin', company=self.company)
        self.other_user = make_user('customers_api_other', role='salesman', company=self.other_company)
        self.url = reverse('reporting:customers_summary_api')

    def _create_customer(self, owner, name, priority='urgent', stage='quote'):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import Company, Department, FollowUp, History

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            manager=f'{name} 책임',
            company=customer_company,
            department=department,
            priority=priority,
            pipeline_stage=stage,
            customer_grade='A',
            ai_score=80,
        )
        History.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            action_type='customer_meeting',
            content=f'{name} 고객 상담',
            next_action='다음 연락',
            next_action_date=timezone.localdate() - timedelta(days=1),
        )
        return followup

    def test_customers_summary_api_requires_login_json(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_customers_summary_api_uses_salesman_own_scope(self):
        own = self._create_customer(self.user, '내고객')
        coworker = self._create_customer(self.coworker, '동료고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['customers']}
        self.assertIn(own.id, ids)
        self.assertNotIn(coworker.id, ids)
        self.assertEqual(payload['metrics']['totalCustomers'], 1)
        self.assertTrue(payload['create']['canCreate'])
        self.assertEqual(payload['links']['createCustomer'], '/customers/?create=1')
        self.assertEqual(payload['create']['submitUrl'], reverse('reporting:followup_create_ajax'))
        self.assertEqual(payload['create']['companySubmitUrl'], reverse('reporting:company_create_api'))
        self.assertEqual(payload['create']['departmentSubmitUrl'], reverse('reporting:department_create_api'))
        company_option = next(option for option in payload['create']['companies'] if option['id'] == own.company_id)
        self.assertTrue(company_option['canManage'])
        self.assertFalse(company_option['canDelete'])
        self.assertEqual(company_option['updateUrl'], reverse('reporting:company_update_api', args=[own.company_id]))
        self.assertEqual(company_option['deleteUrl'], reverse('reporting:company_delete_api', args=[own.company_id]))
        self.assertIn('부서', company_option['deleteMessage'])
        department_option = next(option for option in payload['create']['departments'] if option['id'] == own.department_id)
        self.assertTrue(department_option['canManage'])
        self.assertFalse(department_option['canDelete'])
        self.assertEqual(department_option['updateUrl'], reverse('reporting:department_update_api', args=[own.department_id]))
        self.assertEqual(department_option['deleteUrl'], reverse('reporting:department_delete_api', args=[own.department_id]))
        self.assertIn('담당자', department_option['deleteMessage'])
        self.assertIn('내고객 책임', department_option['searchText'])

    def test_customers_summary_api_returns_department_account_rows(self):
        from datetime import time, timedelta
        from reporting.models import FollowUp, Schedule

        target = self._create_customer(self.user, '계정고객')
        sibling = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='계정고객 추가담당',
            manager='계정고객 추가책임',
            company=target.company,
            department=target.department,
            priority='scheduled',
            pipeline_stage='contact',
        )
        Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=sibling,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(11, 0),
            activity_type='customer_meeting',
            status='scheduled',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        customer_ids = {item['id'] for item in payload['customers']}
        self.assertIn(target.id, customer_ids)
        self.assertIn(sibling.id, customer_ids)
        self.assertEqual(payload['metrics']['totalCustomers'], 2)
        self.assertEqual(payload['metrics']['totalAccounts'], 1)
        self.assertEqual(payload['metrics']['filteredAccounts'], 1)
        self.assertEqual(len(payload['accounts']), 1)
        account = payload['accounts'][0]
        self.assertEqual(account['id'], target.department_id)
        self.assertEqual(account['accountId'], target.department_id)
        self.assertEqual(account['accountType'], 'department')
        self.assertIn(account['representativeCustomerId'], {target.id, sibling.id})
        self.assertEqual(account['customer'], target.department.name)
        self.assertEqual(account['href'], f'/accounts/{target.department_id}/')
        self.assertIn(account['customerHref'], {
            reverse('reporting:followup_detail', args=[target.id]),
            reverse('reporting:followup_detail', args=[sibling.id]),
        })
        self.assertEqual(account['contactCount'], 2)
        self.assertIn('계정고객 담당자', account['contactPreview'])
        self.assertIn('계정고객 추가담당', account['contactPreview'])
        self.assertEqual(account['activityCount'], 1)
        self.assertEqual(account['scheduleCount'], 1)
        self.assertEqual(account['upcomingScheduleCount'], 1)

    def test_customers_summary_api_search_includes_empty_department_accounts(self):
        from reporting.models import Company, Department

        company = Company.objects.create(name='서울시 보건환경연구원', created_by=self.user)
        department = Department.objects.create(
            company=company,
            name='감염병 검사 연구실',
            notes='담당자 등록 전 계정',
            created_by=self.user,
        )
        self._create_customer(self.user, '검색제외고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url, {'q': '서울시 보건'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['filters']['q'], '서울시 보건')
        self.assertEqual(payload['metrics']['filteredCustomers'], 0)
        self.assertEqual(payload['metrics']['filteredAccounts'], 1)
        self.assertEqual(payload['pagination']['totalRows'], 1)
        self.assertEqual(payload['pagination']['accountRows'], 1)
        self.assertEqual(payload['customers'], [])
        self.assertEqual(len(payload['accounts']), 1)
        account = payload['accounts'][0]
        self.assertEqual(account['id'], department.id)
        self.assertEqual(account['accountId'], department.id)
        self.assertEqual(account['accountType'], 'department')
        self.assertIsNone(account['representativeCustomerId'])
        self.assertEqual(account['customer'], '담당자 없음')
        self.assertEqual(account['company'], '서울시 보건환경연구원')
        self.assertEqual(account['department'], '감염병 검사 연구실')
        self.assertEqual(account['contactCount'], 0)
        self.assertEqual(account['href'], f'/accounts/{department.id}/')
        self.assertEqual(account['createScheduleHref'], f'/schedules/?create=1&department={department.id}')

    def test_department_autocomplete_finds_department_by_pi_manager_name(self):
        target = self._create_customer(self.user, 'PI검색고객')
        target.manager = '김PI교수'
        target.save(update_fields=['manager'])
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:department_autocomplete'), {
            'q': '김PI',
            'company': target.company_id,
        })

        self.assertEqual(response.status_code, 200)
        department_ids = {item['id'] for item in response.json()['results']}
        self.assertIn(target.department_id, department_ids)

    def test_customers_summary_api_defaults_to_latest_updated_first(self):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import FollowUp

        older = self._create_customer(self.user, '오래된고객')
        newer = self._create_customer(self.user, '최근고객')
        FollowUp.objects.filter(pk=older.pk).update(
            updated_at=timezone.now() - timedelta(days=5),
            created_at=timezone.now() - timedelta(days=5),
        )
        FollowUp.objects.filter(pk=newer.pk).update(
            updated_at=timezone.now(),
            created_at=timezone.now(),
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['customers'][0]['id'], newer.id)

    def test_customers_summary_api_filters_search_owner_and_priority(self):
        target = self._create_customer(self.user, 'PCR핵심', priority='urgent')
        self._create_customer(self.user, '일반', priority='scheduled')
        self._create_customer(self.coworker, '동료PCR', priority='urgent')
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {
            'q': 'PCR',
            'owner': str(self.user.id),
            'priority': 'urgent',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = [item['id'] for item in payload['customers']]
        self.assertEqual(ids, [target.id])
        self.assertEqual(payload['filters']['q'], 'PCR')
        self.assertTrue(any(option['id'] == self.user.id for option in payload['options']['owners']))

    def test_customers_summary_api_filters_company_grade_stage_and_score_level(self):
        target = self._create_customer(self.user, '정밀필터', priority='urgent', stage='quote')
        target.customer_grade = 'VIP'
        target.ai_score = 95
        target.save(update_fields=['customer_grade', 'ai_score'])
        other = self._create_customer(self.user, '정밀제외', priority='scheduled', stage='potential')
        other.customer_grade = 'C'
        other.ai_score = 15
        other.save(update_fields=['customer_grade', 'ai_score'])
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {
            'company': str(target.company_id),
            'grade': 'VIP',
            'stage': 'quote',
            'level': 'critical',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = [item['id'] for item in payload['customers']]
        self.assertEqual(ids, [target.id])
        self.assertEqual(payload['filters']['company'], str(target.company_id))
        self.assertEqual(payload['filters']['grade'], 'VIP')
        self.assertEqual(payload['filters']['stage'], 'quote')
        self.assertEqual(payload['filters']['level'], 'critical')
        company_option = next(option for option in payload['options']['companies'] if option['id'] == target.company_id)
        self.assertEqual(company_option['name'], target.company.name)
        self.assertTrue(any(option['value'] == 'VIP' for option in payload['options']['grades']))
        self.assertTrue(any(option['value'] == 'critical' for option in payload['options']['scoreLevels']))

    def test_customers_summary_api_returns_row_policy_pagination_and_export_links(self):
        from urllib.parse import parse_qs, urlparse

        profile = self.user.userprofile
        profile.can_download_excel = True
        profile.save(update_fields=['can_download_excel'])
        for index in range(12):
            self._create_customer(self.user, f'페이징{index:02d}', priority='scheduled')
        self.client.force_login(self.user)

        response = self.client.get(self.url, {
            'mode': 'contact',
            'page': '2',
            'page_size': '10',
            'q': '페이징',
            'priority': 'scheduled',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['filters']['mode'], 'contact')
        self.assertEqual(payload['pagination']['rowMode'], 'contact')
        self.assertEqual(payload['pagination']['page'], 2)
        self.assertEqual(payload['pagination']['pageSize'], 10)
        self.assertEqual(payload['pagination']['totalRows'], 12)
        self.assertEqual(payload['pagination']['totalPages'], 2)
        self.assertEqual(len(payload['customers']), 2)
        self.assertTrue(any(option['value'] == 'account' for option in payload['options']['rowModes']))
        self.assertTrue(any(option['value'] == 'contact' for option in payload['options']['rowModes']))
        self.assertTrue(payload['export']['canDownload'])
        full_query = parse_qs(urlparse(payload['export']['fullUrl']).query)
        basic_query = parse_qs(urlparse(payload['export']['basicUrl']).query)
        self.assertEqual(full_query['search'], ['페이징'])
        self.assertEqual(full_query['priority'], ['scheduled'])
        self.assertEqual(basic_query['search'], ['페이징'])

    def test_customers_summary_api_manager_sees_same_company_only(self):
        own = self._create_customer(self.user, '회사내고객')
        coworker = self._create_customer(self.coworker, '회사내동료')
        other = self._create_customer(self.other_user, '타사고객')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['customers']}
        self.assertIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        priority_ids = {item['id'] for item in payload['priorityCustomers']}
        self.assertIn(own.id, priority_ids)
        self.assertTrue(payload['scope']['canViewAll'])
        self.assertFalse(payload['create']['canCreate'])

    def test_customers_summary_api_admin_sees_all_company_data(self):
        own = self._create_customer(self.user, '관리자회사내')
        other = self._create_customer(self.other_user, '관리자타사')
        self.client.force_login(self.admin)

        response = self.client.get(self.url, {'mode': 'contact'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['customers']}
        self.assertIn(own.id, ids)
        self.assertIn(other.id, ids)
        self.assertTrue(payload['scope']['canViewAll'])
        self.assertTrue(payload['create']['canCreate'])

    def test_customers_summary_api_includes_activity_and_schedule_snapshot(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import History, Schedule

        target = self._create_customer(self.user, '일정있는고객', priority='urgent')
        History.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            action_type='quote',
            content='견적 재확인',
            next_action='견적 후속',
            next_action_date=timezone.localdate() + timedelta(days=2),
        )
        upcoming = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 30),
            activity_type='quote',
            status='scheduled',
            location='고객 연구실',
        )
        Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() - timedelta(days=3),
            visit_time=time(9, 0),
            activity_type='customer_meeting',
            status='completed',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        customer = next(item for item in payload['customers'] if item['id'] == target.id)
        self.assertEqual(customer['activityCount'], 2)
        self.assertEqual(customer['scheduleCount'], 2)
        self.assertEqual(customer['upcomingScheduleCount'], 1)
        self.assertEqual(customer['overdueActionCount'], 1)
        self.assertEqual(customer['upcomingSchedule']['id'], upcoming.id)
        self.assertEqual(customer['upcomingSchedule']['activityLabel'], '견적 제출')
        self.assertEqual(customer['upcomingSchedule']['time'], '10:30')
        self.assertEqual(
            customer['upcomingSchedule']['createHistoryHref'],
            f'/notes/?create=1&customer={target.id}&schedule={upcoming.id}',
        )
        self.assertEqual(
            customer['upcomingSchedule']['djangoCreateHistoryHref'],
            reverse('reporting:history_create_from_schedule', args=[upcoming.id]),
        )
        self.assertEqual(payload['metrics']['scheduledCustomers'], 1)

    def test_customers_apis_exclude_stale_quote_submission_overdue_count(self):
        from datetime import time, timedelta
        from decimal import Decimal
        from reporting.models import History, Schedule

        today = timezone.localdate()
        target = self._create_customer(self.user, '견적완료고객')
        History.objects.filter(followup=target).update(reviewed_at=timezone.now())
        stale_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            action_type='customer_meeting',
            content='견적서 제출 예정',
            next_action='견적서 및 비교표 제출',
            next_action_date=today - timedelta(days=1),
        )
        active_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            action_type='customer_meeting',
            content='견적 검토 상황 확인',
            next_action='견적 검토 여부 확인',
            next_action_date=today - timedelta(days=1),
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today,
            visit_time=time(10, 30),
            activity_type='quote',
            status='scheduled',
            expected_revenue=Decimal('770000'),
        )
        DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=quote_schedule,
            user=self.user,
            transaction_number='CUSTOMER-ST-Q-001',
            output_format='pdf',
        )
        self.client.force_login(self.user)

        list_response = self.client.get(self.url)

        self.assertEqual(list_response.status_code, 200)
        customer = next(item for item in list_response.json()['customers'] if item['id'] == target.id)
        self.assertEqual(customer['overdueActionCount'], 1)
        self.assertEqual(customer['nextAction'], '견적 검토 여부 확인')
        self.assertNotEqual(customer['nextAction'], stale_history.next_action)

        detail_response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        overdue_ids = {item['id'] for item in detail_payload['overdueActions']}
        self.assertNotIn(stale_history.id, overdue_ids)
        self.assertIn(active_history.id, overdue_ids)
        self.assertEqual(detail_payload['metrics']['overdueActions'], 1)

    def test_followup_create_ajax_creates_customer_for_salesman(self):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name='빠른등록 회사', created_by=self.user)
        department = Department.objects.create(
            company=customer_company,
            name='빠른등록 연구실',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:followup_create_ajax'), {
            'customer_name': '빠른등록 담당자',
            'company': str(customer_company.id),
            'department': str(department.id),
            'priority': 'urgent',
            'manager': '빠른 책임',
            'phone_number': '010-0000-0000',
            'email': 'quick@example.com',
            'notes': 'React 빠른 등록',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['href'], f"/customers/{payload['followup_id']}/")
        followup = FollowUp.objects.get(id=payload['followup_id'])
        self.assertEqual(followup.customer_name, '빠른등록 담당자')
        self.assertEqual(followup.user, self.user)
        self.assertEqual(followup.user_company, self.company)
        self.assertEqual(followup.priority, 'scheduled')

    def test_followup_create_ajax_blocks_manager(self):
        from reporting.models import Company, Department

        customer_company = Company.objects.create(name='매니저차단 회사', created_by=self.user)
        department = Department.objects.create(
            company=customer_company,
            name='매니저차단 연구실',
            created_by=self.user,
        )
        self.client.force_login(self.manager)

        response = self.client.post(reverse('reporting:followup_create_ajax'), {
            'customer_name': '매니저 생성 시도',
            'company': str(customer_company.id),
            'department': str(department.id),
            'priority': 'urgent',
        })

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])

    def test_company_and_department_create_apis_create_records_for_salesman(self):
        from reporting.models import Company, Department

        self.client.force_login(self.user)

        company_response = self.client.post(reverse('reporting:company_create_api'), {
            'name': 'React인라인 회사',
        })

        self.assertEqual(company_response.status_code, 200)
        company_payload = company_response.json()
        self.assertTrue(company_payload['success'])
        company = Company.objects.get(id=company_payload['company']['id'])
        self.assertEqual(company.name, 'React인라인 회사')
        self.assertEqual(company.created_by, self.user)

        department_response = self.client.post(reverse('reporting:department_create_api'), {
            'company_id': str(company.id),
            'name': 'React인라인 연구실',
        })

        self.assertEqual(department_response.status_code, 200)
        department_payload = department_response.json()
        self.assertTrue(department_payload['success'])
        department = Department.objects.get(id=department_payload['department']['id'])
        self.assertEqual(department.company, company)
        self.assertEqual(department.name, 'React인라인 연구실')
        self.assertEqual(department.created_by, self.user)

    def test_company_and_department_create_apis_block_manager(self):
        from reporting.models import Company

        customer_company = Company.objects.create(name='매니저업체차단 회사', created_by=self.user)
        self.client.force_login(self.manager)

        company_response = self.client.post(reverse('reporting:company_create_api'), {
            'name': '매니저 신규 업체',
        })
        department_response = self.client.post(reverse('reporting:department_create_api'), {
            'company_id': str(customer_company.id),
            'name': '매니저 신규 부서',
        })

        self.assertEqual(company_response.status_code, 403)
        self.assertFalse(company_response.json()['success'])
        self.assertEqual(department_response.status_code, 403)
        self.assertFalse(department_response.json()['success'])

    def test_department_create_api_blocks_other_company(self):
        from reporting.models import Company

        other_company = Company.objects.create(name='타사부서차단 회사', created_by=self.other_user)
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:department_create_api'), {
            'company_id': str(other_company.id),
            'name': '타사 신규 부서',
        })

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['error'], '접근 권한이 없는 업체입니다.')

    def test_company_and_department_manage_apis_update_and_delete_owner_records(self):
        from reporting.models import Company, Department

        company = Company.objects.create(name='수정전 업체', created_by=self.user)
        department_parent = Company.objects.create(name='부서수정 부모업체', created_by=self.user)
        department = Department.objects.create(
            company=department_parent,
            name='수정전 부서',
            created_by=self.user,
        )
        coworker_company = Company.objects.create(name='동료수정전 업체', created_by=self.coworker)
        coworker_department = Department.objects.create(
            company=coworker_company,
            name='동료수정전 부서',
            created_by=self.coworker,
        )
        self.client.force_login(self.user)

        company_update = self.client.post(reverse('reporting:company_update_api', args=[company.id]), {
            'name': '수정후 업체',
        })
        department_update = self.client.post(reverse('reporting:department_update_api', args=[department.id]), {
            'name': '수정후 부서',
        })
        coworker_company_update = self.client.post(reverse('reporting:company_update_api', args=[coworker_company.id]), {
            'name': '동료수정후 업체',
        })
        coworker_department_update = self.client.post(reverse('reporting:department_update_api', args=[coworker_department.id]), {
            'name': '동료수정후 부서',
        })

        self.assertEqual(company_update.status_code, 200)
        self.assertTrue(company_update.json()['success'])
        self.assertEqual(Company.objects.get(id=company.id).name, '수정후 업체')
        self.assertEqual(department_update.status_code, 200)
        self.assertTrue(department_update.json()['success'])
        self.assertEqual(Department.objects.get(id=department.id).name, '수정후 부서')
        self.assertEqual(coworker_company_update.status_code, 200)
        self.assertTrue(coworker_company_update.json()['success'])
        self.assertEqual(Company.objects.get(id=coworker_company.id).name, '동료수정후 업체')
        self.assertEqual(coworker_department_update.status_code, 200)
        self.assertTrue(coworker_department_update.json()['success'])
        self.assertEqual(Department.objects.get(id=coworker_department.id).name, '동료수정후 부서')

        department_delete = self.client.post(reverse('reporting:department_delete_api', args=[department.id]))
        company_delete = self.client.post(reverse('reporting:company_delete_api', args=[company.id]))

        self.assertEqual(department_delete.status_code, 200)
        self.assertTrue(department_delete.json()['success'])
        self.assertFalse(Department.objects.filter(id=department.id).exists())
        self.assertEqual(company_delete.status_code, 200)
        self.assertTrue(company_delete.json()['success'])
        self.assertFalse(Company.objects.filter(id=company.id).exists())

    def test_department_update_api_moves_department_to_same_scope_company_and_updates_contacts(self):
        from django.utils import timezone
        from reporting.models import Company, CustomerAsset, Prepayment

        followup = self._create_customer(self.user, '소속이동')
        department = followup.department
        target_company = Company.objects.create(name='소속이동 경희대학교', created_by=self.coworker)
        asset = CustomerAsset.objects.create(
            company=followup.company,
            department=department,
            primary_followup=followup,
            asset_name='소속이동 장비',
            created_by=self.user,
        )
        prepayment = Prepayment.objects.create(
            department=department,
            customer=followup,
            company=followup.company,
            amount=100000,
            balance=100000,
            payment_date=timezone.localdate(),
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:department_update_api', args=[department.id]), {
            'name': department.name,
            'company_id': str(target_company.id),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertTrue(payload['department']['moved'])
        self.assertEqual(payload['department']['company_id'], target_company.id)
        self.assertEqual(payload['department']['updated_counts']['followups'], 1)
        self.assertEqual(payload['department']['updated_counts']['assets'], 1)
        self.assertEqual(payload['department']['updated_counts']['prepayments'], 1)
        department.refresh_from_db()
        followup.refresh_from_db()
        asset.refresh_from_db()
        prepayment.refresh_from_db()
        self.assertEqual(department.company_id, target_company.id)
        self.assertEqual(followup.company_id, target_company.id)
        self.assertEqual(asset.company_id, target_company.id)
        self.assertEqual(prepayment.company_id, target_company.id)

    def test_department_update_api_blocks_move_to_other_company_scope(self):
        from reporting.models import Company

        followup = self._create_customer(self.user, '타사소속이동차단')
        department = followup.department
        original_company_id = department.company_id
        other_company = Company.objects.create(name='타사소속이동 대상', created_by=self.other_user)
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:department_update_api', args=[department.id]), {
            'name': department.name,
            'company_id': str(other_company.id),
        })

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])
        department.refresh_from_db()
        followup.refresh_from_db()
        self.assertEqual(department.company_id, original_company_id)
        self.assertEqual(followup.company_id, original_company_id)

    def test_company_and_department_manage_apis_block_manager_and_other_user(self):
        from reporting.models import Company, Department

        company = Company.objects.create(name='수정권한차단 업체', created_by=self.user)
        department = Department.objects.create(
            company=company,
            name='수정권한차단 부서',
            created_by=self.user,
        )

        self.client.force_login(self.manager)
        manager_company_update = self.client.post(reverse('reporting:company_update_api', args=[company.id]), {
            'name': '매니저수정',
        })
        manager_department_delete = self.client.post(reverse('reporting:department_delete_api', args=[department.id]))

        self.assertEqual(manager_company_update.status_code, 403)
        self.assertFalse(manager_company_update.json()['success'])
        self.assertEqual(manager_department_delete.status_code, 403)
        self.assertFalse(manager_department_delete.json()['success'])

        self.client.force_login(self.other_user)
        other_company_delete = self.client.post(reverse('reporting:company_delete_api', args=[company.id]))
        other_department_update = self.client.post(reverse('reporting:department_update_api', args=[department.id]), {
            'name': '타사수정',
        })

        self.assertEqual(other_company_delete.status_code, 403)
        self.assertFalse(other_company_delete.json()['success'])
        self.assertEqual(other_department_update.status_code, 403)
        self.assertFalse(other_department_update.json()['success'])

    def test_company_and_department_delete_apis_block_records_in_use(self):
        target = self._create_customer(self.user, '삭제차단')
        self.client.force_login(self.user)

        company_response = self.client.post(reverse('reporting:company_delete_api', args=[target.company_id]))
        department_response = self.client.post(reverse('reporting:department_delete_api', args=[target.department_id]))

        self.assertEqual(company_response.status_code, 400)
        self.assertFalse(company_response.json()['success'])
        self.assertIn('삭제할 수 없습니다', company_response.json()['error'])
        self.assertEqual(department_response.status_code, 400)
        self.assertFalse(department_response.json()['success'])
        self.assertIn('담당자', department_response.json()['error'])

    def test_companies_management_api_requires_login_json(self):
        response = self.client.get(reverse('reporting:companies_management_api'))

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_companies_management_api_salesman_owner_permissions_and_search(self):
        from reporting.models import Company

        own = self._create_customer(self.user, '업체관리내고객')
        coworker = self._create_customer(self.coworker, '업체관리동료고객')
        other = self._create_customer(self.other_user, '업체관리타사고객')
        free_company = Company.objects.create(name='업체관리 삭제가능', created_by=self.user)
        hidden_move_target = Company.objects.create(name='숨은 이동 대상', created_by=self.coworker)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:companies_management_api'), {'q': '업체관리'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['permissions']['canCreateCompany'])
        self.assertFalse(payload['permissions']['readOnly'])
        self.assertNotIn('companyId', payload['filters'])
        company_ids = {company['id'] for company in payload['companies']}
        self.assertIn(own.company_id, company_ids)
        self.assertIn(coworker.company_id, company_ids)
        self.assertIn(free_company.id, company_ids)
        self.assertNotIn(hidden_move_target.id, company_ids)
        self.assertNotIn(other.company_id, company_ids)
        move_company_ids = {company['id'] for company in payload['departmentMoveCompanies']}
        self.assertIn(hidden_move_target.id, move_company_ids)

        own_company = next(company for company in payload['companies'] if company['id'] == own.company_id)
        coworker_company = next(company for company in payload['companies'] if company['id'] == coworker.company_id)
        free_company_payload = next(company for company in payload['companies'] if company['id'] == free_company.id)
        self.assertTrue(own_company['canManage'])
        self.assertFalse(own_company['canDelete'])
        self.assertIn('부서', own_company['deleteMessage'])
        self.assertNotIn('href', own_company)
        self.assertNotIn('djangoHref', own_company)
        self.assertNotIn('departmentsUrl', own_company)
        self.assertNotIn('customersUrl', own_company)
        self.assertTrue(own_company['departments'][0]['canManage'])
        self.assertIn('담당자', own_company['departments'][0]['deleteMessage'])
        self.assertNotIn('cleanupPreviewHref', own_company['departments'][0])
        self.assertTrue(coworker_company['canManage'])
        self.assertTrue(coworker_company['departments'][0]['canManage'])
        self.assertTrue(free_company_payload['canDelete'])
        self.assertEqual(free_company_payload['deleteMessage'], '')

    def test_companies_management_api_search_avoids_join_explosion_query(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from reporting.models import Company

        self._create_customer(self.user, '서울검색내고객')
        Company.objects.create(name='서울검색빈업체', created_by=self.user)
        self.client.force_login(self.user)

        with CaptureQueriesContext(connection) as captured:
            response = self.client.get(reverse('reporting:companies_management_api'), {'q': '서울'})

        self.assertEqual(response.status_code, 200)
        heavy_company_queries = [
            query['sql']
            for query in captured.captured_queries
            if (
                'FROM "reporting_company"' in query['sql']
                and '"reporting_customerasset"' in query['sql']
                and '"reporting_prepayment"' in query['sql']
                and 'GROUP BY "reporting_company"' in query['sql']
            )
        ]
        self.assertEqual(
            heavy_company_queries,
            [],
            '업체 검색은 장비/선결제/담당자 count를 한 SQL에 조인해 임시 파일을 키우면 안 됩니다.',
        )

    def test_companies_management_api_manager_readonly_same_company(self):
        own = self._create_customer(self.user, '매니저업체조회내고객')
        coworker = self._create_customer(self.coworker, '매니저업체조회동료고객')
        other = self._create_customer(self.other_user, '매니저업체조회타사고객')
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:companies_management_api'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload['permissions']['canCreateCompany'])
        self.assertTrue(payload['permissions']['readOnly'])
        self.assertTrue(payload['permissions']['readOnlyMessage'])
        company_ids = {company['id'] for company in payload['companies']}
        self.assertIn(own.company_id, company_ids)
        self.assertIn(coworker.company_id, company_ids)
        self.assertNotIn(other.company_id, company_ids)
        own_company = next(company for company in payload['companies'] if company['id'] == own.company_id)
        self.assertFalse(own_company['canManage'])
        self.assertFalse(own_company['departments'][0]['canManage'])
        self.assertTrue(any(salesman['username'] == self.user.username for salesman in own_company['salesmen']))

    def test_companies_management_api_admin_sees_all_and_can_manage(self):
        own = self._create_customer(self.user, '관리자업체조회내고객')
        other = self._create_customer(self.other_user, '관리자업체조회타사고객')
        self.client.force_login(self.admin)

        response = self.client.get(reverse('reporting:companies_management_api'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        company_ids = {company['id'] for company in payload['companies']}
        self.assertIn(own.company_id, company_ids)
        self.assertIn(other.company_id, company_ids)
        self.assertTrue(payload['scope']['canViewAll'])
        self.assertTrue(payload['permissions']['canCreateCompany'])
        self.assertTrue(all(company['canManage'] for company in payload['companies']))
        self.assertTrue(all(department['canManage'] for department in payload['departments']))

    def test_company_legacy_get_routes_redirect_to_react_management(self):
        target = self._create_customer(self.user, '업체레거시리다이렉트')
        self.client.force_login(self.user)

        list_response = self.client.get(reverse('reporting:company_list'))
        detail_response = self.client.get(reverse('reporting:company_detail', args=[target.company_id]))
        department_response = self.client.get(reverse('reporting:department_edit', args=[target.department_id]))

        self.assertEqual(list_response.status_code, 302)
        self.assertIn('/companies/', list_response['Location'])
        self.assertEqual(detail_response.status_code, 302)
        self.assertIn('/companies/', detail_response['Location'])
        self.assertIn(f'company_id={target.company_id}', detail_response['Location'])
        self.assertEqual(department_response.status_code, 302)
        self.assertIn('/companies/', department_response['Location'])
        self.assertIn(f'department_id={target.department_id}', department_response['Location'])

    def test_customer_detail_summary_api_requires_login_json(self):
        target = self._create_customer(self.user, '상세로그인')
        url = reverse('reporting:customer_detail_summary_api', args=[target.id])

        response = self.client.get(url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_customer_detail_summary_api_returns_notes_and_schedules(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import History, HistoryFile, Schedule, ScheduleFile

        target = self._create_customer(self.user, '상세고객', priority='urgent')
        upcoming = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(11, 0),
            activity_type='quote',
            status='scheduled',
            location='상세 회의실',
        )
        note = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            action_type='quote',
            content='상세 견적 메모',
            next_action='상세 후속',
            next_action_date=timezone.localdate() + timedelta(days=1),
        )
        note_file_body = b'customer-note'
        schedule_file_body = b'customer-schedule'
        note_file = HistoryFile.objects.create(
            history=note,
            file=SimpleUploadedFile('customer-note.txt', note_file_body, content_type='text/plain'),
            original_filename='customer-note.txt',
            file_size=len(note_file_body),
            uploaded_by=self.user,
        )
        schedule_file = ScheduleFile.objects.create(
            schedule=upcoming,
            file=SimpleUploadedFile('customer-schedule.txt', schedule_file_body, content_type='text/plain'),
            original_filename='customer-schedule.txt',
            file_size=len(schedule_file_body),
            uploaded_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['customer']['id'], target.id)
        self.assertGreaterEqual(payload['metrics']['recentNotes'], 2)
        self.assertEqual(payload['upcomingSchedules'][0]['id'], upcoming.id)
        self.assertEqual(payload['recentSchedules'][0]['id'], upcoming.id)
        self.assertTrue(payload['links']['djangoDetail'].endswith(f'/followups/{target.id}/'))
        self.assertTrue(payload['links']['djangoEdit'].endswith(f'/followups/{target.id}/edit/'))
        self.assertEqual(payload['links']['createNote'], f'/notes/?create=1&customer={target.id}')
        self.assertEqual(payload['links']['pipeline'], '/pipeline/')
        self.assertEqual(payload['attachments']['metrics']['totalFiles'], 2)
        self.assertEqual(payload['attachments']['metrics']['noteFiles'], 1)
        self.assertEqual(payload['attachments']['metrics']['scheduleFiles'], 1)
        attachment_types = {item['fileType'] for item in payload['attachments']['recentFiles']}
        self.assertIn('note', attachment_types)
        self.assertIn('schedule', attachment_types)
        note_attachment = next(item for item in payload['attachments']['recentFiles'] if item['fileType'] == 'note')
        schedule_attachment = next(item for item in payload['attachments']['recentFiles'] if item['fileType'] == 'schedule')
        self.assertEqual(note_attachment['sourceHref'], f'/notes/{note.id}/')
        self.assertEqual(note_attachment['downloadHref'], reverse('reporting:file_download', args=[note_file.id]))
        self.assertEqual(schedule_attachment['sourceHref'], f'/schedules/{upcoming.id}/')
        self.assertEqual(schedule_attachment['downloadHref'], reverse('reporting:schedule_file_download', args=[schedule_file.id]))
        self.assertTrue(payload['edit']['canEdit'])
        self.assertTrue(payload['edit']['canDelete'])
        self.assertEqual(payload['edit']['submitUrl'], reverse('reporting:customer_update_api', args=[target.id]))
        self.assertEqual(payload['edit']['deleteUrl'], reverse('reporting:customer_delete_api', args=[target.id]))
        self.assertTrue(any(option['id'] == target.company_id for option in payload['edit']['companies']))
        self.assertTrue(any(option['id'] == target.department_id for option in payload['edit']['departments']))

    def test_customer_detail_summary_api_includes_scoped_prepayment_summary(self):
        from django.utils import timezone
        from reporting.models import Prepayment

        target = self._create_customer(self.user, '상세선결제', priority='urgent')
        first = Prepayment.objects.create(
            customer=target,
            company=target.company,
            amount=100000,
            balance=80000,
            payment_date=timezone.localdate(),
            payer_name='상세입금자',
            status='active',
            created_by=self.user,
        )
        second = Prepayment.objects.create(
            customer=target,
            company=target.company,
            amount=50000,
            balance=0,
            payment_date=timezone.localdate(),
            payer_name='상세소진',
            status='depleted',
            created_by=self.user,
        )
        coworker_prepayment = Prepayment.objects.create(
            customer=target,
            company=target.company,
            amount=999000,
            balance=999000,
            payment_date=timezone.localdate(),
            payer_name='동료입금자',
            status='active',
            created_by=self.coworker,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        summary = response.json()['prepaymentSummary']
        self.assertEqual(summary['metrics']['totalAmount'], 150000)
        self.assertEqual(summary['metrics']['totalBalance'], 80000)
        self.assertEqual(summary['metrics']['totalUsed'], 70000)
        self.assertEqual(summary['metrics']['totalCount'], 2)
        self.assertEqual(summary['metrics']['activeCount'], 1)
        self.assertEqual(summary['metrics']['depletedCount'], 1)
        self.assertEqual(summary['links']['prepayments'], '/prepayments/')
        self.assertEqual(summary['links']['accountPrepayments'], f'/prepayments/account/{target.department_id}/')
        self.assertEqual(summary['links']['customerPrepayments'], f'/prepayments/customer/{target.id}/')
        self.assertTrue(summary['links']['djangoCustomerPrepayments'].endswith(f'/prepayment/customer/{target.id}/'))
        prepayment_ids = {item['id'] for item in summary['recentPrepayments']}
        self.assertIn(first.id, prepayment_ids)
        self.assertIn(second.id, prepayment_ids)
        self.assertNotIn(coworker_prepayment.id, prepayment_ids)

    def test_customer_detail_summary_api_includes_operational_records_with_payment_source(self):
        from datetime import time, timedelta
        from decimal import Decimal
        from django.utils import timezone
        from reporting.models import (
            DeliveryItem,
            FollowUp,
            Prepayment,
            PrepaymentUsage,
            Product,
            Quote,
            QuoteItem,
            Schedule,
            History,
        )

        today = timezone.localdate()
        target = self._create_customer(self.user, '운영기록고객', priority='urgent')
        sibling = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='운영기록 같은부서 담당자',
            manager='운영기록 같은부서 책임',
            company=target.company,
            department=target.department,
            priority='scheduled',
            pipeline_stage='quote',
        )
        product = Product.objects.create(
            product_code='OP-QUOTE-001',
            standard_price=Decimal('120000'),
            created_by=self.user,
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=10),
            visit_time=time(10, 0),
            activity_type='quote',
            status='completed',
            notes='운영 견적 일정',
        )
        quote = Quote.objects.create(
            quote_number='OP-Q-001',
            schedule=quote_schedule,
            followup=target,
            user=self.user,
            valid_until=today + timedelta(days=20),
            stage='sent',
            notes='운영 견적 메모',
        )
        QuoteItem.objects.create(
            quote=quote,
            product=product,
            quantity=2,
            unit_price=Decimal('120000'),
        )
        sibling_quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=sibling,
            visit_date=today - timedelta(days=8),
            visit_time=time(13, 0),
            activity_type='quote',
            status='completed',
            notes='같은 부서 견적 일정',
        )
        Quote.objects.create(
            quote_number='OP-Q-SAME-DEPT',
            schedule=sibling_quote_schedule,
            followup=sibling,
            user=self.user,
            valid_until=today + timedelta(days=20),
            stage='sent',
            notes='같은 부서 견적 메모',
        )
        service_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=4),
            visit_time=time(9, 30),
            activity_type='service',
            status='scheduled',
            notes='장비 점검 서비스 일정',
        )
        service_note = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            schedule=service_schedule,
            action_type='service',
            service_status='received',
            content='서비스 일정에 연결된 영업노트 메모',
        )
        prepayment = Prepayment.objects.create(
            customer=target,
            company=target.company,
            amount=100000,
            balance=40000,
            payment_date=today - timedelta(days=7),
            payer_name='운영입금자',
            status='active',
            created_by=self.user,
        )
        sibling_prepayment = Prepayment.objects.create(
            customer=sibling,
            company=sibling.company,
            amount=50000,
            balance=50000,
            payment_date=today - timedelta(days=6),
            payer_name='같은부서입금자',
            status='active',
            created_by=self.user,
        )
        prepaid_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=2),
            visit_time=time(11, 0),
            activity_type='delivery',
            status='completed',
            notes='구조화된 선결제 차감',
            use_prepayment=True,
            prepayment=prepayment,
            prepayment_amount=Decimal('60000'),
        )
        prepaid_item = DeliveryItem.objects.create(
            schedule=prepaid_schedule,
            item_name='선결제 납품품목',
            quantity=1,
            unit_price=Decimal('60000'),
        )
        PrepaymentUsage.objects.create(
            prepayment=prepayment,
            schedule=prepaid_schedule,
            schedule_item=prepaid_item,
            product_name='선결제 납품품목',
            quantity=1,
            amount=Decimal('60000'),
            remaining_balance=Decimal('40000'),
        )
        normal_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=1),
            visit_time=time(14, 0),
            activity_type='delivery',
            status='completed',
            notes='메모에 선결제라고 써도 구조화 차감 없음',
        )
        DeliveryItem.objects.create(
            schedule=normal_schedule,
            item_name='일반 납품품목',
            quantity=1,
            unit_price=Decimal('30000'),
        )
        sibling_delivery_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=sibling,
            visit_date=today,
            visit_time=time(15, 0),
            activity_type='delivery',
            status='completed',
            notes='같은 부서 납품',
        )
        DeliveryItem.objects.create(
            schedule=sibling_delivery_schedule,
            item_name='같은 부서 납품품목',
            quantity=1,
            unit_price=Decimal('70000'),
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        records = payload['operationalRecords']
        self.assertEqual(payload['account']['id'], target.department_id)
        self.assertEqual(payload['account']['type'], 'department')
        self.assertEqual(payload['account']['contactCount'], 2)
        self.assertEqual(payload['account']['ledgerScopeLabel'], '부서/연구실 계정 공유 원장')
        self.assertIn('납품, 견적, 선결제, 서비스', payload['account']['ledgerScopeDescription'])
        contact_names = {contact['name'] for contact in payload['account']['contacts']}
        self.assertIn('운영기록고객 담당자', contact_names)
        self.assertIn('운영기록 같은부서 담당자', contact_names)
        self.assertEqual(records['metrics']['serviceRecords'], 1)
        self.assertEqual(records['metrics']['quoteRecords'], 2)
        self.assertEqual(records['metrics']['deliveryRecords'], 3)
        self.assertEqual(records['metrics']['prepaymentDeliveryRecords'], 1)
        self.assertEqual(records['metrics']['normalDeliveryRecords'], 2)
        self.assertEqual(records['metrics']['prepaymentRecords'], 2)
        self.assertEqual(payload['prepaymentSummary']['metrics']['totalCount'], 2)
        service_record_types = {record['recordType'] for record in records['serviceRecords']}
        self.assertIn('service_schedule', service_record_types)
        self.assertNotIn('service_case', service_record_types)
        self.assertNotIn('service_history', service_record_types)
        self.assertFalse(any(record['summary'] == service_note.content for record in records['serviceRecords']))
        self.assertTrue(any(record['id'] == service_schedule.id and record['recordType'] == 'service_schedule' for record in records['serviceRecords']))
        self.assertIn(service_note.id, {note['id'] for note in payload['recentNotes']})
        quote_numbers = {record['quoteNumber'] for record in records['quoteRecords']}
        self.assertIn('OP-Q-001', quote_numbers)
        self.assertIn('OP-Q-SAME-DEPT', quote_numbers)
        delivery_by_id = {item['id']: item for item in records['deliveryRecords']}
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['paymentSource'], 'prepayment')
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['paymentSourceLabel'], '선결제 차감 납품')
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['paymentStatus'], 'prepayment_deduction')
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['paymentStatusLabel'], '선결제 차감 납품')
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['prepaymentAmount'], 60000)
        self.assertEqual(delivery_by_id[prepaid_schedule.id]['prepaymentUsages'][0]['amount'], 60000)
        self.assertEqual(delivery_by_id[normal_schedule.id]['paymentSource'], 'normal')
        self.assertEqual(delivery_by_id[normal_schedule.id]['paymentSourceLabel'], '일반 납품')
        self.assertEqual(delivery_by_id[normal_schedule.id]['paymentStatus'], 'normal')
        self.assertEqual(delivery_by_id[normal_schedule.id]['paymentStatusLabel'], '일반 납품')
        self.assertEqual(delivery_by_id[normal_schedule.id]['prepaymentAmount'], 0)
        self.assertIn('선결제 사용 필드와 PrepaymentUsage 기록이 없습니다.', delivery_by_id[normal_schedule.id]['paymentEvidence'])
        self.assertEqual(delivery_by_id[sibling_delivery_schedule.id]['customerName'], '운영기록 같은부서 담당자')
        payer_names = {record['payerName'] for record in records['prepaymentRecords']}
        self.assertIn('운영입금자', payer_names)
        self.assertIn('같은부서입금자', payer_names)

        account_response = self.client.get(reverse('reporting:account_detail_summary_api', args=[target.department_id]))
        self.assertEqual(account_response.status_code, 200)
        account_payload = account_response.json()
        self.assertEqual(account_payload['links']['accountDetail'], f'/accounts/{target.department_id}/')
        self.assertEqual(account_payload['account']['contactCount'], 2)
        self.assertEqual(account_payload['account']['ledgerScopeLabel'], '부서/연구실 계정 공유 원장')
        self.assertEqual(account_payload['operationalRecords']['metrics']['deliveryRecords'], 3)
        self.assertEqual(account_payload['operationalRecords']['metrics']['prepaymentDeliveryRecords'], 1)
        self.assertEqual(account_payload['operationalRecords']['metrics']['normalDeliveryRecords'], 2)
        self.assertEqual(account_payload['operationalRecords']['metrics']['quoteRecords'], 2)
        self.assertEqual(account_payload['operationalRecords']['metrics']['prepaymentRecords'], 2)
        self.assertEqual(account_payload['operationalRecords']['metrics']['serviceRecords'], 1)
        self.assertNotIn('service_history', {
            record['recordType']
            for record in account_payload['operationalRecords']['serviceRecords']
        })

    def test_customer_delivery_records_xlsx_export_downloads_department_shared_deliveries(self):
        from datetime import time, timedelta
        from decimal import Decimal
        from io import BytesIO
        from openpyxl import load_workbook
        from django.utils import timezone
        from reporting.models import DeliveryItem, FollowUp, Prepayment, PrepaymentUsage, Schedule

        today = timezone.localdate()
        target = self._create_customer(self.user, '납품엑셀고객', priority='urgent')
        sibling = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='납품엑셀 같은부서 담당자',
            manager='납품엑셀 같은부서 책임',
            company=target.company,
            department=target.department,
            priority='scheduled',
            pipeline_stage='quote',
        )
        other_target = self._create_customer(self.user, '다른납품고객', priority='urgent')
        prepayment = Prepayment.objects.create(
            customer=target,
            company=target.company,
            amount=Decimal('200000'),
            balance=Decimal('150000'),
            payment_date=today - timedelta(days=3),
            payer_name='엑셀입금자',
            created_by=self.user,
        )
        prepaid_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=2),
            visit_time=time(10, 0),
            activity_type='delivery',
            status='completed',
            use_prepayment=True,
            prepayment=prepayment,
            prepayment_amount=Decimal('50000'),
            notes='엑셀 선결제 납품',
        )
        prepaid_item = DeliveryItem.objects.create(
            schedule=prepaid_schedule,
            item_name='엑셀 선결제 품목',
            quantity=2,
            unit='EA',
            unit_price=Decimal('25000'),
        )
        PrepaymentUsage.objects.create(
            prepayment=prepayment,
            schedule=prepaid_schedule,
            schedule_item=prepaid_item,
            product_name='엑셀 선결제 품목',
            quantity=2,
            amount=Decimal('50000'),
            remaining_balance=Decimal('150000'),
        )
        normal_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=today - timedelta(days=1),
            visit_time=time(14, 0),
            activity_type='delivery',
            status='completed',
            notes='엑셀 일반 납품',
        )
        DeliveryItem.objects.create(
            schedule=normal_schedule,
            item_name='엑셀 일반 품목',
            quantity=1,
            unit='EA',
            unit_price=Decimal('30000'),
        )
        sibling_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=sibling,
            visit_date=today,
            visit_time=time(13, 0),
            activity_type='delivery',
            status='completed',
            notes='같은 부서 납품',
        )
        DeliveryItem.objects.create(
            schedule=sibling_schedule,
            item_name='엑셀 같은부서 품목',
            quantity=1,
            unit='EA',
            unit_price=Decimal('40000'),
        )
        other_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=other_target,
            visit_date=today,
            visit_time=time(15, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=other_schedule,
            item_name='다른 고객 품목',
            quantity=1,
            unit_price=Decimal('99999'),
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_delivery_records_xlsx_export_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment;', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook['납품 기록']
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], '납품일')
        self.assertEqual(rows[0][8], '결제상태')
        item_names = {row[10] for row in rows[1:]}
        self.assertIn('엑셀 선결제 품목', item_names)
        self.assertIn('엑셀 일반 품목', item_names)
        self.assertIn('엑셀 같은부서 품목', item_names)
        self.assertNotIn('다른 고객 품목', item_names)
        customer_names = {row[2] for row in rows[1:]}
        self.assertIn('납품엑셀 같은부서 담당자', customer_names)
        payment_labels = {row[7] for row in rows[1:]}
        self.assertIn('선결제 차감 납품', payment_labels)
        self.assertIn('일반 납품', payment_labels)
        payment_status_labels = {row[8] for row in rows[1:]}
        self.assertIn('선결제 차감 납품', payment_status_labels)
        self.assertIn('일반 납품', payment_status_labels)
        prepaid_row = next(row for row in rows[1:] if row[10] == '엑셀 선결제 품목')
        normal_row = next(row for row in rows[1:] if row[10] == '엑셀 일반 품목')
        self.assertEqual(prepaid_row[9], 50000)
        self.assertIn('PrepaymentUsage 합계=50,000원', prepaid_row[17])
        self.assertEqual(normal_row[9], 0)
        self.assertIn('선결제 사용 필드와 PrepaymentUsage 기록이 없습니다.', normal_row[17])

    def test_customer_delivery_records_xlsx_export_blocks_out_of_scope_customer(self):
        target = self._create_customer(self.other_user, '타사납품엑셀고객')
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_delivery_records_xlsx_export_api', args=[target.id]))

        self.assertEqual(response.status_code, 403)

    def test_customer_delivery_records_xlsx_export_requires_login(self):
        target = self._create_customer(self.user, '익명납품엑셀고객')

        response = self.client.get(reverse('reporting:customer_delivery_records_xlsx_export_api', args=[target.id]))

        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response['Location'])

    def test_customer_detail_summary_api_excludes_customer_ai_payload(self):
        target = self._create_customer(self.user, 'AI제거고객', priority='urgent')
        profile = self.user.userprofile
        profile.can_use_ai = True
        profile.save(update_fields=['can_use_ai'])
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertNotIn('aiDepartment', payload)
        self.assertIn('customer', payload)
        self.assertIn('recentNotes', payload)

    def test_account_detail_summary_api_includes_management_fields_and_contact_roles(self):
        from reporting.models import FollowUp

        target = self._create_customer(self.user, '계정관리필드', priority='urgent')
        target.department.address = '공용 계정 주소'
        target.department.notes = '공용 계정 메모'
        target.department.save(update_fields=['address', 'notes'])
        target.contact_role = FollowUp.CONTACT_ROLE_PI
        target.save(update_fields=['contact_role'])
        sibling = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='계정관리필드 세금담당',
            manager='세금 책임',
            company=target.company,
            department=target.department,
            contact_role=FollowUp.CONTACT_ROLE_TAX_INVOICE,
            is_active=False,
            priority='scheduled',
            pipeline_stage='contact',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:account_detail_summary_api', args=[target.department_id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        account = payload['account']
        self.assertEqual(account['address'], '공용 계정 주소')
        self.assertEqual(account['notes'], '공용 계정 메모')
        self.assertEqual(account['piContactName'], target.customer_name)
        self.assertEqual(account['contactCount'], 2)
        self.assertEqual(account['activeContactCount'], 1)
        self.assertEqual(account['inactiveContactCount'], 1)
        self.assertEqual(account['href'], f'/accounts/{target.department_id}/')
        self.assertNotIn('cleanupPreviewHref', account)
        self.assertTrue(account['management']['canManage'])
        self.assertEqual(account['management']['accountSubmitUrl'], reverse('reporting:account_update_api', args=[target.department_id]))
        self.assertEqual(account['management']['contactCreateUrl'], reverse('reporting:account_contact_create_api', args=[target.department_id]))
        self.assertEqual(payload['links']['accountDetail'], f'/accounts/{target.department_id}/')
        self.assertNotIn('accountCleanupPreview', payload['links'])
        self.assertTrue(payload['permissions']['canManageAccount'])
        self.assertTrue(payload['permissions']['canCreateNote'])
        self.assertTrue(payload['permissions']['canCreateSchedule'])
        role_values = {option['value'] for option in account['management']['contactRoles']}
        self.assertEqual(role_values, {'pi', 'practitioner', 'purchasing', 'tax_invoice'})
        contacts = {contact['id']: contact for contact in account['contacts']}
        self.assertEqual(contacts[target.id]['contactRoleLabel'], 'PI')
        self.assertTrue(contacts[target.id]['isActive'])
        self.assertEqual(contacts[sibling.id]['contactRoleLabel'], '세금계산서 담당자')
        self.assertFalse(contacts[sibling.id]['isActive'])
        self.assertTrue(contacts[sibling.id]['updateUrl'].endswith(f'/contacts/{sibling.id}/update/'))

    def test_empty_department_account_detail_allows_first_contact_creation(self):
        from reporting.models import Company, Department, FollowUp, History

        company = Company.objects.create(name='빈계정 회사', created_by=self.user)
        department = Department.objects.create(
            company=company,
            name='빈계정 연구실',
            address='빈 계정 주소',
            notes='담당자 등록 전 메모',
            created_by=self.user,
        )
        note = History.objects.create(
            user=self.user,
            company=self.company,
            department=department,
            action_type='customer_meeting',
            content='담당자 없이 부서에 먼저 남긴 영업노트',
            next_action='담당자 확인',
        )
        self.client.force_login(self.user)

        detail_response = self.client.get(reverse('reporting:account_detail_summary_api', args=[department.id]))

        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertTrue(detail_payload['success'])
        self.assertEqual(detail_payload['customer']['customer'], '담당자 없음')
        self.assertEqual(detail_payload['customer']['accountId'], department.id)
        self.assertEqual(detail_payload['account']['contactCount'], 0)
        self.assertEqual(detail_payload['account']['address'], '빈 계정 주소')
        self.assertEqual(detail_payload['account']['notes'], '담당자 등록 전 메모')
        self.assertEqual(detail_payload['metrics']['recentNotes'], 1)
        self.assertEqual(len(detail_payload['recentNotes']), 1)
        self.assertEqual(detail_payload['recentNotes'][0]['id'], note.id)
        self.assertEqual(detail_payload['recentNotes'][0]['customer'], '담당자 미등록')
        self.assertEqual(detail_payload['recentNotes'][0]['customerHref'], f'/accounts/{department.id}/')
        self.assertTrue(detail_payload['account']['management']['canManage'])
        self.assertEqual(
            detail_payload['account']['management']['contactCreateUrl'],
            reverse('reporting:account_contact_create_api', args=[department.id]),
        )
        self.assertTrue(detail_payload['permissions']['canManageAccount'])
        self.assertFalse(detail_payload['permissions']['canCreateNote'])
        self.assertFalse(detail_payload['permissions']['canCreateSchedule'])

        update_response = self.client.post(reverse('reporting:account_update_api', args=[department.id]), {
            'company': str(company.id),
            'department_name': '빈계정 연구실 수정',
            'address': '수정된 빈 계정 주소',
            'notes': '수정된 담당자 등록 전 메모',
        })
        self.assertEqual(update_response.status_code, 200)
        department.refresh_from_db()
        self.assertEqual(department.name, '빈계정 연구실 수정')
        self.assertEqual(department.address, '수정된 빈 계정 주소')
        self.assertEqual(department.notes, '수정된 담당자 등록 전 메모')

        create_response = self.client.post(reverse('reporting:account_contact_create_api', args=[department.id]), {
            'customer_name': '첫 담당자',
            'contact_role': FollowUp.CONTACT_ROLE_PRACTITIONER,
            'department': str(department.id),
            'priority': 'scheduled',
            'status': 'active',
            'pipeline_stage': 'contact',
            'phone_number': '010-0000-0000',
            'email': 'first@example.com',
            'is_active': 'true',
        })

        self.assertEqual(create_response.status_code, 200)
        created = FollowUp.objects.get(id=create_response.json()['followup_id'])
        self.assertEqual(created.user, self.user)
        self.assertEqual(created.company, company)
        self.assertEqual(created.department, department)
        self.assertEqual(created.customer_name, '첫 담당자')
        self.assertEqual(create_response.json()['accountHref'], f'/accounts/{department.id}/')

    def test_empty_department_account_detail_allows_same_company_coworker_management(self):
        from reporting.models import Company, Department, FollowUp

        company = Company.objects.create(name='동료빈계정 회사', created_by=self.coworker)
        department = Department.objects.create(
            company=company,
            name='동료빈계정 연구실',
            address='동료가 등록한 주소',
            notes='동료가 등록한 메모',
            created_by=self.coworker,
        )
        self.client.force_login(self.user)

        detail_response = self.client.get(reverse('reporting:account_detail_summary_api', args=[department.id]))

        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertTrue(detail_payload['account']['management']['canManage'])
        self.assertTrue(detail_payload['permissions']['canManageAccount'])

        update_response = self.client.post(reverse('reporting:account_update_api', args=[department.id]), {
            'company': str(company.id),
            'department_name': '동료빈계정 연구실 수정',
            'address': '내가 수정한 주소',
            'notes': '공동 관리 메모',
        })
        self.assertEqual(update_response.status_code, 200)
        department.refresh_from_db()
        self.assertEqual(department.name, '동료빈계정 연구실 수정')
        self.assertEqual(department.address, '내가 수정한 주소')
        self.assertEqual(department.notes, '공동 관리 메모')

        create_response = self.client.post(reverse('reporting:account_contact_create_api', args=[department.id]), {
            'customer_name': '공동관리 담당자',
            'contact_role': FollowUp.CONTACT_ROLE_PRACTITIONER,
            'department': str(department.id),
            'status': 'active',
            'pipeline_stage': 'contact',
            'is_active': 'true',
        })
        self.assertEqual(create_response.status_code, 200)
        created = FollowUp.objects.get(id=create_response.json()['followup_id'])
        self.assertEqual(created.user, self.user)
        self.assertEqual(created.company, company)
        self.assertEqual(created.department, department)

    def test_empty_department_account_keeps_manager_read_only_and_blocks_other_company(self):
        from reporting.models import Company, Department

        company = Company.objects.create(name='빈계정 권한 회사', created_by=self.user)
        department = Department.objects.create(
            company=company,
            name='빈계정 권한 연구실',
            created_by=self.user,
        )

        self.client.force_login(self.manager)
        manager_detail = self.client.get(reverse('reporting:account_detail_summary_api', args=[department.id]))
        self.assertEqual(manager_detail.status_code, 200)
        manager_payload = manager_detail.json()
        self.assertFalse(manager_payload['account']['management']['canManage'])
        self.assertFalse(manager_payload['permissions']['canManageAccount'])
        self.assertTrue(manager_payload['permissions']['readOnlyMessage'])

        manager_create = self.client.post(reverse('reporting:account_contact_create_api', args=[department.id]), {
            'customer_name': '차단 담당자',
            'department': str(department.id),
        })
        self.assertEqual(manager_create.status_code, 403)

        self.client.force_login(self.other_user)
        other_detail = self.client.get(reverse('reporting:account_detail_summary_api', args=[department.id]))
        self.assertEqual(other_detail.status_code, 403)

    def test_account_update_api_updates_department_info_and_contact_company(self):
        from reporting.models import Company, Department, FollowUp

        target = self._create_customer(self.user, '계정정보수정', priority='scheduled')
        next_company = Company.objects.create(name='계정정보수정 새회사', created_by=self.user)
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:account_update_api', args=[target.department_id]), {
            'company': str(next_company.id),
            'department_name': '계정정보수정 새연구실',
            'address': '새 계정 주소',
            'notes': '새 계정 메모',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        department = Department.objects.get(id=target.department_id)
        self.assertEqual(department.company, next_company)
        self.assertEqual(department.name, '계정정보수정 새연구실')
        self.assertEqual(department.address, '새 계정 주소')
        self.assertEqual(department.notes, '새 계정 메모')
        self.assertEqual(FollowUp.objects.get(id=target.id).company, next_company)

    def test_account_contact_api_creates_moves_and_inactivates_contact(self):
        from reporting.models import Department, FollowUp

        target = self._create_customer(self.user, '담당자관리', priority='urgent')
        other_department = Department.objects.create(
            company=target.company,
            name='담당자관리 이동연구실',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        create_response = self.client.post(reverse('reporting:account_contact_create_api', args=[target.department_id]), {
            'customer_name': '구매 담당자',
            'contact_role': FollowUp.CONTACT_ROLE_PURCHASING,
            'department': str(target.department_id),
            'priority': 'scheduled',
            'status': 'active',
            'pipeline_stage': 'contact',
            'phone_number': '010-2222-3333',
            'email': 'buyer@example.com',
            'is_active': 'true',
        })

        self.assertEqual(create_response.status_code, 200)
        created = FollowUp.objects.get(id=create_response.json()['followup_id'])
        self.assertEqual(created.department, target.department)
        self.assertEqual(created.contact_role, FollowUp.CONTACT_ROLE_PURCHASING)
        self.assertTrue(created.is_active)

        update_response = self.client.post(reverse('reporting:account_contact_update_api', args=[target.department_id, created.id]), {
            'customer_name': '세금 담당자',
            'contact_role': FollowUp.CONTACT_ROLE_TAX_INVOICE,
            'department': str(other_department.id),
            'priority': 'followup',
            'status': 'paused',
            'pipeline_stage': 'quote',
            'phone_number': '010-4444-5555',
            'email': 'tax@example.com',
            'is_active': 'false',
            'notes': '이동 및 비활성화',
        })

        self.assertEqual(update_response.status_code, 200)
        moved = FollowUp.objects.get(id=created.id)
        self.assertEqual(moved.customer_name, '세금 담당자')
        self.assertEqual(moved.department, other_department)
        self.assertEqual(moved.company, other_department.company)
        self.assertEqual(moved.contact_role, FollowUp.CONTACT_ROLE_TAX_INVOICE)
        self.assertFalse(moved.is_active)
        self.assertEqual(moved.priority, 'scheduled')
        self.assertEqual(moved.status, 'paused')
        self.assertEqual(moved.pipeline_stage, 'quote')

    def test_account_management_api_blocks_manager(self):
        target = self._create_customer(self.user, '계정관리권한차단', priority='scheduled')
        self.client.force_login(self.manager)

        detail_response = self.client.get(reverse('reporting:account_detail_summary_api', args=[target.department_id]))
        response = self.client.post(reverse('reporting:account_update_api', args=[target.department_id]), {
            'company': str(target.company_id),
            'department_name': target.department.name,
        })

        self.assertEqual(detail_response.status_code, 200)
        detail_payload = detail_response.json()
        self.assertFalse(detail_payload['account']['management']['canManage'])
        self.assertFalse(detail_payload['edit']['canEdit'])
        self.assertFalse(detail_payload['edit']['canDelete'])
        self.assertFalse(detail_payload['permissions']['canManageAccount'])
        self.assertFalse(detail_payload['permissions']['canCreateNote'])
        self.assertFalse(detail_payload['permissions']['canCreateSchedule'])
        self.assertTrue(detail_payload['permissions']['readOnlyMessage'])
        self.assertTrue(all(not contact['canManage'] for contact in detail_payload['account']['contacts']))
        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])

    def test_customer_update_api_updates_customer_for_owner(self):
        from reporting.models import Company, Department, FollowUp

        target = self._create_customer(self.user, '수정대상', priority='scheduled', stage='potential')
        next_company = Company.objects.create(name='수정가능 회사', created_by=self.user)
        next_department = Department.objects.create(
            company=next_company,
            name='수정가능 연구실',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:customer_update_api', args=[target.id]), {
            'customer_name': '수정완료 담당자',
            'company': str(next_company.id),
            'department': str(next_department.id),
            'priority': 'urgent',
            'status': 'paused',
            'pipeline_stage': 'quote',
            'manager': '수정 책임',
            'phone_number': '010-1111-2222',
            'email': 'edited@example.com',
            'address': '수정 주소',
            'notes': 'React 상세 수정',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['href'], f'/customers/{target.id}/')
        updated = FollowUp.objects.get(id=target.id)
        self.assertEqual(updated.customer_name, '수정완료 담당자')
        self.assertEqual(updated.company, next_company)
        self.assertEqual(updated.department, next_department)
        self.assertEqual(updated.priority, 'scheduled')
        self.assertEqual(updated.status, 'paused')
        self.assertEqual(updated.pipeline_stage, 'quote')
        self.assertTrue(updated.pipeline_manually_set)
        self.assertEqual(updated.email, 'edited@example.com')

    def test_customer_priority_update_api_is_deprecated(self):
        from reporting.models import FollowUp

        target = self._create_customer(self.user, '우선순위폐기', priority='scheduled', stage='potential')
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:customer_priority_update', args=[target.id]), {
            'priority': 'urgent',
        })

        self.assertEqual(response.status_code, 410)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertIn('파이프라인', payload['error'])
        target.refresh_from_db()
        self.assertEqual(target.priority, 'scheduled')
        self.assertEqual(
            FollowUp.objects.filter(priority__in=['urgent', 'followup'], id=target.id).count(),
            0,
        )

    def test_customer_update_api_blocks_manager_and_coworker(self):
        target = self._create_customer(self.user, '수정권한차단')
        payload = {
            'customer_name': '권한없는수정',
            'company': str(target.company_id),
            'department': str(target.department_id),
            'priority': target.priority,
            'status': target.status,
            'pipeline_stage': target.pipeline_stage,
        }

        self.client.force_login(self.manager)
        manager_response = self.client.post(reverse('reporting:customer_update_api', args=[target.id]), payload)
        self.assertEqual(manager_response.status_code, 403)
        self.assertFalse(manager_response.json()['success'])

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(reverse('reporting:customer_update_api', args=[target.id]), payload)
        self.assertEqual(coworker_response.status_code, 403)
        self.assertFalse(coworker_response.json()['success'])

    def test_customer_delete_api_deletes_owner_customer_and_blocks_readonly_users(self):
        target = self._create_customer(self.user, '삭제대상')
        blocked = self._create_customer(self.user, '삭제차단대상')

        self.client.force_login(self.manager)
        manager_response = self.client.post(reverse('reporting:customer_delete_api', args=[blocked.id]))
        self.assertEqual(manager_response.status_code, 403)
        self.assertFalse(manager_response.json()['success'])

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(reverse('reporting:customer_delete_api', args=[blocked.id]))
        self.assertEqual(coworker_response.status_code, 403)
        self.assertFalse(coworker_response.json()['success'])

        self.client.force_login(self.user)
        response = self.client.post(reverse('reporting:customer_delete_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['href'], '/customers/')
        self.assertFalse(FollowUp.objects.filter(id=target.id).exists())
        self.assertTrue(FollowUp.objects.filter(id=blocked.id).exists())

    def test_customer_update_api_blocks_other_company_selection(self):
        from reporting.models import Company, Department

        target = self._create_customer(self.user, '타사업체수정차단')
        other_company = Company.objects.create(name='타사업체수정 회사', created_by=self.other_user)
        other_department = Department.objects.create(
            company=other_company,
            name='타사업체수정 연구실',
            created_by=self.other_user,
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:customer_update_api', args=[target.id]), {
            'customer_name': '타사변경시도',
            'company': str(other_company.id),
            'department': str(other_department.id),
            'priority': target.priority,
            'status': target.status,
            'pipeline_stage': target.pipeline_stage,
        })

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])

    def test_customer_detail_summary_api_blocks_other_company_customer(self):
        target = self._create_customer(self.other_user, '타사상세')
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_detail_summary_api', args=[target.id]))

        self.assertEqual(response.status_code, 403)


class QuoteItemsApiTests(TestCase):
    """부서 기준 견적 품목 불러오기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='견적품목API회사')
        self.user = make_user('quote_items_me', role='salesman', company=self.company)
        self.coworker = make_user('quote_items_coworker', role='salesman', company=self.company)

        from reporting.models import Company, Department

        self.customer_company = Company.objects.create(name='견적품목 고객사', created_by=self.user)
        self.department = Department.objects.create(
            company=self.customer_company,
            name='공동 연구실',
            created_by=self.user,
        )
        self.other_department = Department.objects.create(
            company=self.customer_company,
            name='다른 연구실',
            created_by=self.user,
        )

    def _create_followup(self, owner, name, department=None):
        from reporting.models import FollowUp

        return FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=name,
            company=self.customer_company,
            department=department or self.department,
            priority='urgent',
            pipeline_stage='quote',
        )

    def _create_quote_schedule(self, followup, owner, item_name, unit_price, quote_group=''):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = Schedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name=item_name,
            quantity=1,
            unit_price=unit_price,
            quote_group=quote_group,
        )
        return schedule

    def test_quote_items_api_returns_all_own_quotes_in_same_department(self):
        target = self._create_followup(self.user, '대표 고객')
        same_department = self._create_followup(self.user, '같은 부서 고객')
        other_department = self._create_followup(self.user, '다른 부서 고객', self.other_department)
        coworker_customer = self._create_followup(self.coworker, '동료 고객')
        first = self._create_quote_schedule(target, self.user, 'PCR 장비', 1000000)
        second = self._create_quote_schedule(same_department, self.user, '원심분리기', 2000000)
        self._create_quote_schedule(other_department, self.user, '다른 부서 품목', 3000000)
        self._create_quote_schedule(coworker_customer, self.coworker, '동료 품목', 4000000)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['count'], 2)
        schedule_ids = {item['schedule_id'] for item in payload['quotes']}
        self.assertEqual(schedule_ids, {first.id, second.id})
        customer_names = {item['customer_name'] for item in payload['quotes']}
        self.assertEqual(customer_names, {'대표 고객', '같은 부서 고객'})

    def test_quote_items_api_splits_same_schedule_by_quote_group(self):
        from reporting.models import DeliveryItem

        target = self._create_followup(self.user, '구분 선택 고객')
        quote_schedule = self._create_quote_schedule(
            target,
            self.user,
            '보상판매 품목',
            1000000,
            quote_group='보상판매',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='수리 품목',
            quantity=1,
            unit_price=2000000,
            quote_group='수리',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['count'], 2)
        options_by_group = {item['quoteGroup']: item for item in payload['quotes']}
        self.assertEqual(set(options_by_group), {'보상판매', '수리'})
        self.assertEqual(options_by_group['보상판매']['optionId'], f'{quote_schedule.id}:보상판매')
        self.assertEqual(options_by_group['수리']['optionId'], f'{quote_schedule.id}:수리')
        self.assertEqual(options_by_group['보상판매']['quoteGroupLabel'], '보상판매')
        self.assertEqual(options_by_group['수리']['quoteGroupLabel'], '수리')
        self.assertEqual(options_by_group['보상판매']['scheduleId'], quote_schedule.id)
        self.assertEqual(options_by_group['수리']['scheduleId'], quote_schedule.id)
        self.assertEqual([item['itemName'] for item in options_by_group['보상판매']['items']], ['보상판매 품목'])
        self.assertEqual([item['itemName'] for item in options_by_group['수리']['items']], ['수리 품목'])

    def test_quote_items_api_returns_react_delivery_import_fields(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import DeliveryItem, Product, Schedule

        target = self._create_followup(self.user, 'React 납품')
        product = Product.objects.create(
            product_code='PIP-1000',
            unit='SET',
            specification='1000ul',
            description='피펫',
            standard_price=150000,
            created_by=self.user,
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            product=product,
            item_name='PIP-1000',
            quantity=3,
            unit_price=120000,
            discount_rate=5,
            tax_invoice_issued=True,
            quote_group='수리',
            notes='오링 교체',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        quote = payload['quotes'][0]
        quote_item = quote['items'][0]
        self.assertEqual(quote['id'], quote_schedule.id)
        self.assertEqual(quote['optionId'], f'{quote_schedule.id}:수리')
        self.assertEqual(quote['scheduleId'], quote_schedule.id)
        self.assertEqual(quote['quoteGroup'], '수리')
        self.assertEqual(quote['quoteGroupLabel'], '수리')
        self.assertEqual(quote['href'], f'/schedules/{quote_schedule.id}/')
        self.assertEqual(quote['djangoHref'], reverse('reporting:schedule_detail', args=[quote_schedule.id]))
        self.assertEqual(quote_item['id'], item.id)
        self.assertEqual(quote_item['itemName'], 'PIP-1000')
        self.assertEqual(quote_item['unit'], 'SET')
        self.assertEqual(quote_item['unitPrice'], 120000.0)
        self.assertEqual(quote_item['discountRate'], 5.0)
        self.assertEqual(quote_item['discountUnitPrice'], 114000.0)
        self.assertEqual(quote_item['effectiveUnitPrice'], 114000.0)
        self.assertEqual(quote_item['productId'], product.id)
        self.assertEqual(quote_item['productCode'], 'PIP-1000')
        self.assertEqual(quote_item['productDescription'], '피펫')
        self.assertEqual(quote_item['sourceQuoteScheduleId'], quote_schedule.id)
        self.assertEqual(quote_item['sourceQuoteItemId'], item.id)
        self.assertTrue(quote_item['taxInvoiceIssued'])
        self.assertEqual(quote_item['quoteGroup'], '수리')
        self.assertEqual(quote_item['quoteGroupLabel'], '수리')
        self.assertEqual(quote_item['notes'], '오링 교체')

    def test_quote_items_api_recovers_unit_price_from_legacy_total_price(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        target = self._create_followup(self.user, '레거시 총액 견적')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='총액만 있는 견적 품목',
            quantity=2,
            unit='EA',
            unit_price=10000,
            quote_group='수리',
        )
        DeliveryItem.objects.filter(pk=item.pk).update(unit_price=None, total_price=110000)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        quote = payload['quotes'][0]
        quote_item = quote['items'][0]
        self.assertEqual(quote['remainingAmount'], 110000.0)
        self.assertEqual(quote_item['totalPrice'], 110000.0)
        self.assertEqual(quote_item['remainingAmount'], 110000.0)
        self.assertEqual(quote_item['unitPrice'], 50000.0)
        self.assertEqual(quote_item['effectiveUnitPrice'], 50000.0)

    def test_quote_items_api_treats_legacy_zero_discount_unit_price_as_blank(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        target = self._create_followup(self.user, '할인단가0 견적')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='SO825.0002',
            quantity=1,
            unit='EA',
            unit_price=379950,
        )
        DeliveryItem.objects.filter(pk=item.pk).update(discount_rate=0, discount_unit_price=0, total_price=0)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        quote_item = payload['quotes'][0]['items'][0]
        self.assertIsNone(quote_item['discountUnitPrice'])
        self.assertEqual(quote_item['discountRate'], 0.0)
        self.assertEqual(quote_item['unitPrice'], 379950.0)
        self.assertEqual(quote_item['effectiveUnitPrice'], 379950.0)
        self.assertEqual(quote_item['totalPrice'], 417945.0)

    def test_quote_items_api_returns_remaining_items_after_partial_delivery_import(self):
        from datetime import time
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        target = self._create_followup(self.user, '부분 납품 고객')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='quote',
            status='completed',
        )
        sold_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='판매된 견적 품목',
            quantity=1,
            unit='EA',
            unit_price=30000,
            quote_group='보상판매',
        )
        remaining_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='남은 견적 품목',
            quantity=1,
            unit='EA',
            unit_price=90000,
            quote_group='수리',
        )
        delivery_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate(),
            visit_time=time(11, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=delivery_schedule,
            source_quote_schedule=quote_schedule,
            source_quote_item=sold_item,
            item_name='판매된 견적 품목',
            quantity=1,
            unit='EA',
            unit_price=30000,
            quote_group='보상판매',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['count'], 1)
        quote = payload['quotes'][0]
        self.assertEqual(quote['quoteGroup'], '수리')
        self.assertEqual(quote['items'][0]['id'], remaining_item.id)
        self.assertEqual(quote['items'][0]['sourceQuoteItemId'], remaining_item.id)

    def test_quote_items_api_exposes_partial_delivery_remaining_quantities(self):
        from datetime import time
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        target = self._create_followup(self.user, '동일 품목 부분 납품 고객')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='quote',
            status='completed',
        )
        quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='부분 납품 견적 품목',
            quantity=3,
            unit='EA',
            unit_price=10000,
            quote_group='보상판매',
        )
        delivery_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target,
            visit_date=timezone.localdate(),
            visit_time=time(11, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=delivery_schedule,
            source_quote_schedule=quote_schedule,
            source_quote_item=quote_item,
            item_name='부분 납품 견적 품목',
            quantity=1,
            unit='EA',
            unit_price=10000,
            quote_group='보상판매',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        quote = payload['quotes'][0]
        item = quote['items'][0]
        self.assertEqual(quote['deliveryStatus'], 'partial')
        self.assertEqual(quote['deliveryStatusLabel'], '부분 납품 잔여')
        self.assertTrue(quote['hasPartialDelivery'])
        self.assertEqual(quote['quotedAmount'], 33000.0)
        self.assertEqual(quote['deliveredAmount'], 11000.0)
        self.assertEqual(quote['remainingAmount'], 22000.0)
        self.assertEqual(item['originalQuantity'], 3.0)
        self.assertEqual(item['deliveredQuantity'], 1.0)
        self.assertEqual(item['remainingQuantity'], 2.0)
        self.assertEqual(item['quantity'], 2)

    def test_quote_items_api_bulk_progress_avoids_per_quote_queries(self):
        from datetime import time
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        target = self._create_followup(self.user, '대량 견적 고객')
        for index in range(8):
            quote_schedule = self._create_quote_schedule(
                target,
                self.user,
                f'완료 견적 품목 {index}',
                10000 + index,
                quote_group='수리',
            )
            quote_schedule.status = 'completed'
            quote_schedule.save(update_fields=['status'])
            quote_item = quote_schedule.delivery_items_set.first()
            quote_item.quantity = 2
            quote_item.save(update_fields=['quantity', 'total_price', 'updated_at'])
            if index % 2 == 0:
                delivery_schedule = Schedule.objects.create(
                    user=self.user,
                    company=self.company,
                    followup=target,
                    visit_date=timezone.localdate(),
                    visit_time=time(11, 0),
                    activity_type='delivery',
                    status='completed',
                )
                DeliveryItem.objects.create(
                    schedule=delivery_schedule,
                    source_quote_schedule=quote_schedule,
                    source_quote_item=quote_item,
                    item_name=quote_item.item_name,
                    quantity=1,
                    unit='EA',
                    unit_price=quote_item.unit_price,
                    quote_group='수리',
                )
        for index in range(4):
            self._create_quote_schedule(
                target,
                self.user,
                f'진행 견적 품목 {index}',
                20000 + index,
                quote_group='보상판매',
            )
        self.client.force_login(self.user)

        with CaptureQueriesContext(connection) as captured:
            response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['count'], 8)
        self.assertLessEqual(len(captured), 20)

    def test_quote_items_api_excludes_completed_quote_schedules(self):
        target = self._create_followup(self.user, '완료 제외 고객')
        completed = self._create_quote_schedule(target, self.user, '완료된 견적 품목', 1000000)
        completed.status = 'completed'
        completed.save(update_fields=['status'])
        scheduled = self._create_quote_schedule(target, self.user, '진행 중 견적 품목', 2000000)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:followup_quote_items_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        schedule_ids = {item['scheduleId'] for item in payload['quotes']}
        self.assertEqual(schedule_ids, {scheduled.id})

    def test_customer_records_api_includes_quote_schedules_without_quote_model(self):
        target = self._create_followup(self.user, '기록 대표')
        same_department = self._create_followup(self.user, '기록 같은 부서')
        first = self._create_quote_schedule(target, self.user, '견적A', 1000000)
        second = self._create_quote_schedule(same_department, self.user, '견적B', 2000000)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:customer_records_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['quote_count'], 2)
        quote_ids = {item['id'] for item in payload['quotes']}
        self.assertEqual(quote_ids, {first.id, second.id})
        self.assertEqual(payload['total_quote_amount'], 3300000.0)


class NotesSummaryApiTests(TestCase):
    """React 영업노트 화면 읽기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='노트API회사')
        self.other_company = UserCompany.objects.create(name='노트API타사회사')
        self.user = make_user('notes_api_me', role='salesman', company=self.company)
        self.coworker = make_user('notes_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('notes_api_manager', role='manager', company=self.company)
        self.admin = make_user('notes_api_admin', role='admin', company=self.company)
        self.other_user = make_user('notes_api_other', role='salesman', company=self.other_company)
        self.other_manager = make_user('notes_api_other_manager', role='manager', company=self.other_company)
        self.url = reverse('reporting:notes_summary_api')
        self.create_url = reverse('reporting:notes_create_api')

    def _create_note(
        self,
        owner,
        name,
        action_type='customer_meeting',
        content='고객 상담 기록',
        next_action='후속 연락',
        next_action_date=None,
        reviewed=False,
    ):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import Company, Department, FollowUp, History

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            manager=f'{name} 책임',
            company=customer_company,
            department=department,
            priority='urgent',
            pipeline_stage='quote',
        )
        history = History.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            action_type=action_type,
            content=content,
            next_action=next_action,
            next_action_date=next_action_date or timezone.localdate() - timedelta(days=1),
            reviewed_at=timezone.now() if reviewed else None,
            reviewer=self.manager if reviewed else None,
        )
        return history

    def _create_department_only(self, owner, name):
        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        return Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )

    def test_notes_summary_api_requires_login_json(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_notes_summary_api_uses_salesman_own_scope(self):
        own = self._create_note(self.user, '내노트')
        coworker = self._create_note(self.coworker, '동료노트')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['notes']}
        self.assertIn(own.id, ids)
        self.assertNotIn(coworker.id, ids)
        self.assertEqual(payload['metrics']['totalNotes'], 1)
        note = payload['notes'][0]
        self.assertFalse(payload['scope']['canReview'])
        self.assertFalse(note['canReview'])
        self.assertEqual(note['reviewToggleHref'], '')
        self.assertEqual(note['href'], f'/notes/{own.id}/')
        self.assertIn(f'/reporting/histories/{own.id}/', note['djangoHref'])

    def test_notes_summary_api_filters_search_owner_action_review_and_next_action(self):
        target = self._create_note(
            self.user,
            'PCR핵심',
            action_type='quote',
            content='PCR 견적 후속 필요',
            reviewed=False,
        )
        self._create_note(self.user, 'PCR완료', action_type='quote', content='PCR 견적 완료', reviewed=True)
        self._create_note(self.user, '서비스', action_type='service', content='PCR 서비스', reviewed=False)
        self._create_note(self.coworker, 'PCR동료', action_type='quote', content='PCR 동료 건', reviewed=False)
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {
            'q': 'PCR',
            'owner': str(self.user.id),
            'actionType': 'quote',
            'review': 'unreviewed',
            'nextAction': 'overdue',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = [item['id'] for item in payload['notes']]
        self.assertEqual(ids, [target.id])
        self.assertEqual(payload['filters']['q'], 'PCR')
        self.assertTrue(any(option['value'] == 'quote' for option in payload['options']['actionTypes']))
        self.assertEqual(payload['metrics']['filteredNotes'], 1)

    def test_notes_summary_api_defaults_to_recent_one_month_range(self):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import History

        recent = self._create_note(self.user, '최근노트')
        old = self._create_note(self.user, '오래된노트')
        old_created_at = timezone.now() - timedelta(days=45)
        History.objects.filter(pk=old.id).update(created_at=old_created_at)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['notes']}
        self.assertIn(recent.id, ids)
        self.assertNotIn(old.id, ids)
        self.assertEqual(payload['filters']['dateTo'], timezone.localdate().isoformat())

    def test_notes_summary_api_searches_department_contact_names(self):
        from reporting.models import FollowUp, History

        department = self._create_department_only(self.user, '부서메모검색')
        contact = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='홍길동',
            manager='검색 책임',
            company=department.company,
            department=department,
            priority='normal',
            pipeline_stage='potential',
        )
        note = History.objects.create(
            user=self.user,
            company=self.company,
            department=department,
            action_type='customer_meeting',
            content='고객 없는 부서에 남긴 방문 메모',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url, {'q': contact.customer_name})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item['id'] for item in payload['notes']], [note.id])

    def test_notes_summary_api_manager_sees_same_company_only(self):
        own = self._create_note(self.user, '회사내노트')
        coworker = self._create_note(self.coworker, '회사내동료노트')
        other = self._create_note(self.other_user, '타사노트')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['notes']}
        self.assertIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        self.assertEqual(payload['metrics']['totalNotes'], 2)
        self.assertTrue(payload['scope']['canViewAll'])

    def test_notes_summary_api_exposes_review_metadata_for_manager(self):
        from reporting.models import History

        target = self._create_note(
            self.user,
            '검토대상',
            action_type='customer_meeting',
            content='검토가 필요한 고객 미팅',
            reviewed=False,
        )
        History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='관리자 확인 메모',
            created_by=self.manager,
        )
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {'owner': str(self.user.id)})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        note = next(item for item in payload['notes'] if item['id'] == target.id)
        self.assertTrue(payload['scope']['canReview'])
        self.assertTrue(note['canReview'])
        self.assertIn(f'/reporting/histories/{target.id}/toggle-reviewed/', note['reviewToggleHref'])
        self.assertIsNone(note['reviewedAt'])
        self.assertEqual(note['reviewer'], '')
        self.assertEqual(note['replyCount'], 1)
        self.assertEqual(note['fileCount'], 0)

    def test_notes_summary_api_does_not_expose_review_action_for_admin(self):
        target = self._create_note(
            self.user,
            '어드민검토제외',
            action_type='customer_meeting',
            content='회사 매니저만 검토 처리 가능',
            reviewed=False,
        )
        self.client.force_login(self.admin)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        note = next(item for item in payload['notes'] if item['id'] == target.id)
        self.assertFalse(payload['scope']['canReview'])
        self.assertFalse(note['canReview'])
        self.assertEqual(note['reviewToggleHref'], '')

    def test_notes_summary_api_includes_react_create_options_for_salesman(self):
        target = self._create_note(self.user, '작성대상')
        department_only = self._create_department_only(self.user, '고객없는작성대상')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['links']['createNote'], '/notes/?create=1')
        self.assertTrue(payload['create']['canCreate'])
        self.assertEqual(payload['create']['submitUrl'], self.create_url)
        customer_ids = {item['id'] for item in payload['create']['customers']}
        department_ids = {item['id'] for item in payload['create']['departments']}
        self.assertIn(target.followup_id, customer_ids)
        self.assertIn(target.followup.department_id, department_ids)
        self.assertIn(department_only.id, department_ids)
        self.assertTrue(any(item['value'] == 'customer_meeting' for item in payload['create']['actionTypes']))
        self.assertFalse(any(item['value'] == 'memo' for item in payload['create']['actionTypes']))

    def test_notes_summary_api_department_create_options_search_contact_names(self):
        from reporting.models import FollowUp

        target = self._create_note(self.user, '부서검색대상')
        FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='홍연구원',
            manager='김책임교수',
            email='researcher@example.com',
            phone_number='010-1234-5678',
            company=target.followup.company,
            department=target.followup.department,
            priority='scheduled',
            pipeline_stage='potential',
        )
        FollowUp.objects.create(
            user=self.coworker,
            user_company=self.company,
            customer_name='동료전용연구원',
            company=target.followup.company,
            department=target.followup.department,
            priority='scheduled',
            pipeline_stage='potential',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        department_option = next(
            item for item in payload['create']['departments']
            if item['id'] == target.followup.department_id
        )
        self.assertIn('홍연구원', department_option['searchText'])
        self.assertIn('김책임교수', department_option['searchText'])
        self.assertIn('researcher@example.com', department_option['searchText'])
        self.assertIn('010-1234-5678', department_option['searchText'])
        self.assertNotIn('동료전용연구원', department_option['searchText'])

    def test_notes_summary_api_preloads_more_than_legacy_department_select_limit(self):
        customer_company = Company.objects.create(name='대량부서선택회사', created_by=self.user)
        target_department = None
        for index in range(205):
            department = Department.objects.create(
                company=customer_company,
                name=f'연구실 {index:03d}',
                created_by=self.user,
            )
            if index == 204:
                target_department = department
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        department_ids = {item['id'] for item in payload['create']['departments']}
        self.assertGreater(len(payload['create']['departments']), 180)
        self.assertIn(target_department.id, department_ids)

    def test_notes_summary_api_includes_schedule_create_options_without_overdue_filter(self):
        from datetime import time, timedelta
        from reporting.models import DeliveryItem, Schedule

        target = self._create_note(self.user, '일정옵션')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            visit_date=timezone.localdate() + timedelta(days=2),
            visit_time=time(14, 0),
            activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='PCR kit',
            quantity=2,
            unit='EA',
            unit_price=1000,
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        schedule_option = next(item for item in payload['create']['schedules'] if item['id'] == schedule.id)
        self.assertEqual(schedule_option['followupId'], target.followup_id)
        self.assertEqual(schedule_option['suggestedActionType'], 'delivery_schedule')
        self.assertIn('PCR kit', schedule_option['deliveryItems'])
        self.assertEqual(schedule_option['deliveryAmount'], 2200)
        self.assertFalse(any(option['value'] == 'overdue' for option in payload['options']['nextActionStates']))

    def test_notes_summary_api_labels_service_activity_as_memo(self):
        service_note = self._create_note(self.user, '서비스라벨', action_type='service')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        service_option = next(item for item in payload['options']['actionTypes'] if item['value'] == 'service')
        service_count = next(item for item in payload['actionCounts'] if item['value'] == 'service')
        note = next(item for item in payload['notes'] if item['id'] == service_note.id)
        self.assertEqual(service_option['label'], '메모')
        self.assertEqual(service_count['label'], '메모')
        self.assertEqual(note['actionLabel'], '메모')

    def test_notes_summary_api_marks_date_only_next_action_as_scheduled_display(self):
        next_date = timezone.localdate() + timedelta(days=31)
        target = self._create_note(
            self.user,
            '날짜만후속',
            next_action='',
            next_action_date=next_date,
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        note = next(item for item in payload['notes'] if item['id'] == target.id)
        self.assertEqual(note['nextAction'], '')
        self.assertEqual(note['nextActionDisplay'], '후속 예정')
        self.assertEqual(note['nextActionDate'], next_date.isoformat())

    def test_notes_create_api_requires_login_json(self):
        response = self.client.post(
            self.create_url,
            data=json.dumps({}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_notes_create_api_creates_own_customer_note(self):
        from django.utils import timezone
        from reporting.models import History, Schedule

        target = self._create_note(self.user, '빠른작성기준')
        self.client.force_login(self.user)
        followup_date = timezone.localdate()

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'customer_meeting',
                'content': 'React에서 바로 작성한 영업노트',
                'nextAction': '다음 주 견적 확인',
                'nextActionDate': followup_date.isoformat(),
                'activityDate': timezone.localdate().isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertTrue(payload['success'])
        created = History.objects.get(pk=payload['historyId'])
        self.assertEqual(created.user, self.user)
        self.assertEqual(created.followup_id, target.followup_id)
        self.assertEqual(created.content, 'React에서 바로 작성한 영업노트')
        self.assertEqual(created.next_action, '다음 주 견적 확인')
        self.assertEqual(created.meeting_date, timezone.localdate())
        followup_schedule = Schedule.objects.get(
            user=self.user,
            followup=target.followup,
            visit_date=followup_date,
            activity_type='customer_meeting',
        )
        self.assertEqual(followup_schedule.visit_time.strftime('%H:%M'), '09:00')
        self.assertEqual(followup_schedule.status, 'scheduled')
        self.assertIn('자동 생성: 영업노트 후속 미팅', followup_schedule.notes)
        self.assertIn(f'/notes/{created.id}/', followup_schedule.notes)
        self.assertEqual(created.schedule_id, followup_schedule.id)
        self.assertTrue(payload['followupScheduleCreated'])
        self.assertEqual(payload['followupSchedule']['id'], followup_schedule.id)
        self.assertIn('후속 미팅 일정을 생성', payload['message'])

    def test_notes_create_api_creates_department_only_note(self):
        from django.utils import timezone
        from reporting.models import History, Schedule

        department = self._create_department_only(self.user, '고객없는노트')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': department.id,
                'actionType': 'customer_meeting',
                'content': '고객 등록 전 부서에 남긴 영업노트',
                'nextAction': '담당자 확인 후 연결',
                'activityDate': timezone.localdate().isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = History.objects.get(pk=payload['historyId'])
        self.assertIsNone(created.followup_id)
        self.assertEqual(created.department_id, department.id)
        self.assertEqual(created.content, '고객 등록 전 부서에 남긴 영업노트')
        self.assertEqual(payload['note']['customer'], '담당자 미등록')
        self.assertEqual(payload['note']['departmentId'], department.id)
        self.assertEqual(payload['note']['customerHref'], f'/accounts/{department.id}/')
        self.assertEqual(Schedule.objects.count(), 0)

    def test_notes_create_api_creates_department_only_followup_meeting_schedule(self):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import History, Schedule

        department = self._create_department_only(self.user, '부서후속일정')
        followup_date = timezone.localdate() + timedelta(days=3)
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': department.id,
                'actionType': 'customer_meeting',
                'content': '부서만 연결한 영업노트',
                'nextAction': '담당자 만나서 요구사항 확인',
                'nextActionDate': followup_date.isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = History.objects.get(pk=payload['historyId'])
        followup_schedule = Schedule.objects.get(
            user=self.user,
            followup__isnull=True,
            department=department,
            visit_date=followup_date,
            activity_type='customer_meeting',
        )
        self.assertEqual(created.schedule_id, followup_schedule.id)
        self.assertEqual(followup_schedule.company, self.company)
        self.assertIn('담당자 만나서 요구사항 확인', followup_schedule.notes)
        self.assertTrue(payload['followupScheduleCreated'])
        self.assertEqual(payload['followupSchedule']['departmentId'], department.id)

    def test_notes_create_api_reuses_existing_followup_meeting_schedule(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import History, Schedule

        target = self._create_note(self.user, '기존후속일정')
        followup_date = timezone.localdate() + timedelta(days=5)
        existing_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            department=target.followup.department,
            visit_date=followup_date,
            visit_time=time(14, 0),
            activity_type='customer_meeting',
            status='scheduled',
            notes='이미 등록한 후속 미팅',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'customer_meeting',
                'content': '기존 후속 미팅이 있는 영업노트',
                'nextAction': '미팅에서 예산 확인',
                'nextActionDate': followup_date.isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = History.objects.get(pk=payload['historyId'])
        self.assertEqual(
            Schedule.objects.filter(
                user=self.user,
                followup=target.followup,
                visit_date=followup_date,
                activity_type='customer_meeting',
            ).count(),
            1,
        )
        self.assertEqual(created.schedule_id, existing_schedule.id)
        self.assertFalse(payload['followupScheduleCreated'])
        self.assertEqual(payload['followupSchedule']['id'], existing_schedule.id)
        self.assertIn('기존 후속 미팅 일정', payload['message'])

    def test_notes_create_api_requires_customer_when_department_has_contacts(self):
        target = self._create_note(self.user, '고객있는부서')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': target.followup.department_id,
                'actionType': 'customer_meeting',
                'content': '고객 있는 부서를 고객 없이 작성 시도',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('고객을 선택하세요', response.json()['error'])

    def test_notes_create_api_links_schedule_and_uses_schedule_date(self):
        from datetime import time, timedelta
        from reporting.models import History, Schedule

        target = self._create_note(self.user, '일정연결작성')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 30),
            activity_type='customer_meeting',
        )
        other_target = self._create_note(self.user, '다른일정')
        other_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=other_target.followup,
            visit_date=timezone.localdate(),
            visit_time=time(11, 0),
            activity_type='customer_meeting',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'scheduleId': schedule.id,
                'actionType': 'customer_meeting',
                'content': '일정 상세에서 작성한 영업노트',
                'nextAction': '샘플 반응 확인',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = History.objects.get(pk=payload['historyId'])
        self.assertEqual(created.schedule_id, schedule.id)
        self.assertEqual(created.followup_id, target.followup_id)
        self.assertEqual(created.meeting_date, schedule.visit_date)
        self.assertEqual(payload['reactHref'], f'/notes/{created.id}/')

        mismatch_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'scheduleId': other_schedule.id,
                'actionType': 'customer_meeting',
                'content': '잘못된 일정 연결',
            }),
            content_type='application/json',
        )
        self.assertEqual(mismatch_response.status_code, 400)
        self.assertIn('고객이 일치하지 않습니다', mismatch_response.json()['error'])

    def test_notes_create_api_links_department_only_schedule(self):
        from datetime import time, timedelta
        from reporting.models import History, Schedule

        department = self._create_department_only(self.user, '부서일정노트')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            department=department,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 30),
            activity_type='customer_meeting',
        )
        other_department = self._create_department_only(self.user, '다른부서일정')
        other_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            department=other_department,
            visit_date=timezone.localdate(),
            visit_time=time(11, 0),
            activity_type='customer_meeting',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': department.id,
                'scheduleId': schedule.id,
                'actionType': 'customer_meeting',
                'content': '부서만 있는 일정 상세에서 작성한 영업노트',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        created = History.objects.get(pk=response.json()['historyId'])
        self.assertIsNone(created.followup_id)
        self.assertEqual(created.department_id, department.id)
        self.assertEqual(created.schedule_id, schedule.id)
        self.assertEqual(created.meeting_date, schedule.visit_date)

        mismatch_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': department.id,
                'scheduleId': other_schedule.id,
                'actionType': 'customer_meeting',
                'content': '잘못된 부서 일정 연결',
            }),
            content_type='application/json',
        )
        self.assertEqual(mismatch_response.status_code, 400)
        self.assertIn('부서/연구실과 일정이 일치하지 않습니다', mismatch_response.json()['error'])

    def test_notes_create_api_links_delivery_schedule_and_copies_delivery_summary(self):
        from datetime import time, timedelta
        from reporting.models import DeliveryItem, History, Schedule

        target = self._create_note(self.user, '납품일정연결')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            visit_date=timezone.localdate() + timedelta(days=3),
            visit_time=time(15, 0),
            activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Centrifuge tube',
            quantity=3,
            unit='BOX',
            unit_price=5000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'scheduleId': schedule.id,
                'actionType': 'delivery_schedule',
                'content': '납품 일정 완료 보고',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        created = History.objects.get(pk=response.json()['historyId'])
        self.assertEqual(created.schedule_id, schedule.id)
        self.assertEqual(created.delivery_date, schedule.visit_date)
        self.assertIn('Centrifuge tube', created.delivery_items)
        self.assertEqual(created.delivery_amount, 16500)

    def test_notes_create_api_blocks_manager_and_other_owner_customer(self):
        target = self._create_note(self.coworker, '동료작성차단')
        coworker_department = self._create_department_only(self.coworker, '동료부서작성차단')

        self.client.force_login(self.manager)
        manager_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'customer_meeting',
                'content': '매니저 작성 시도',
            }),
            content_type='application/json',
        )
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.user)
        other_owner_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'customer_meeting',
                'content': '동료 고객 작성 시도',
            }),
            content_type='application/json',
        )
        self.assertEqual(other_owner_response.status_code, 403)

        other_department_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': coworker_department.id,
                'actionType': 'customer_meeting',
                'content': '동료 부서 작성 시도',
            }),
            content_type='application/json',
        )
        self.assertEqual(other_department_response.status_code, 403)

    def test_notes_detail_api_returns_detail_and_edit_config(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from django.utils import timezone
        from reporting.models import History, HistoryFile

        target = self._create_note(
            self.user,
            '상세대상',
            action_type='customer_meeting',
            content='초기 상담',
        )
        target.meeting_date = timezone.localdate()
        target.meeting_situation = '예산 검토 중'
        target.meeting_next_action = '견적서 발송'
        target.save()
        history_file = HistoryFile.objects.create(
            history=target,
            file=SimpleUploadedFile('note-detail.txt', b'note detail memo', content_type='text/plain'),
            original_filename='note-detail.txt',
            file_size=16,
            uploaded_by=self.user,
        )
        self.addCleanup(history_file.file.delete, False)
        owner_reply = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='실무자 댓글',
        )
        manager_reply = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='관리자 확인 메모',
            created_by=self.manager,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:notes_detail_api', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['note']['id'], target.id)
        self.assertEqual(payload['note']['href'], f'/notes/{target.id}/')
        self.assertEqual(payload['note']['content'], '초기 상담')
        self.assertEqual(payload['note']['meetingSituation'], '예산 검토 중')
        self.assertEqual(payload['links']['notes'], '/notes/')
        self.assertIn(f'/reporting/histories/{target.id}/', payload['links']['djangoDetail'])
        self.assertTrue(payload['edit']['canEdit'])
        self.assertEqual(payload['edit']['submitUrl'], reverse('reporting:notes_update_api', args=[target.id]))
        self.assertEqual(payload['links']['uploadFiles'], reverse('reporting:note_file_upload', args=[target.id]))
        self.assertTrue(payload['note']['canDelete'])
        self.assertEqual(payload['note']['deleteHref'], reverse('reporting:notes_delete_api', args=[target.id]))
        self.assertEqual(payload['links']['deleteNote'], reverse('reporting:notes_delete_api', args=[target.id]))
        self.assertEqual(payload['note']['files'][0]['id'], history_file.id)
        self.assertEqual(payload['note']['files'][0]['deleteHref'], reverse('reporting:file_delete', args=[history_file.id]))
        self.assertTrue(payload['comments']['canCreate'])
        self.assertEqual(payload['comments']['submitUrl'], reverse('reporting:add_manager_memo_to_history_api', args=[target.id]))
        owner_reply_payload = next(reply for reply in payload['note']['replies'] if reply['id'] == owner_reply.id)
        manager_reply_payload = next(reply for reply in payload['note']['replies'] if reply['id'] == manager_reply.id)
        self.assertEqual(owner_reply_payload['authorRole'], '댓글')
        self.assertTrue(owner_reply_payload['canDelete'])
        self.assertEqual(owner_reply_payload['deleteHref'], reverse('reporting:delete_manager_memo_api', args=[owner_reply.id]))
        self.assertEqual(manager_reply_payload['authorRole'], '매니저 메모')
        self.assertFalse(manager_reply_payload['canDelete'])
        customer_ids = {item['id'] for item in payload['edit']['customers']}
        self.assertIn(target.followup_id, customer_ids)

    def test_notes_detail_api_manager_read_only_and_other_company_blocked(self):
        target = self._create_note(self.user, '매니저상세', action_type='quote', content='견적 확인')

        self.client.force_login(self.manager)
        manager_response = self.client.get(reverse('reporting:notes_detail_api', args=[target.id]))
        self.assertEqual(manager_response.status_code, 200)
        manager_payload = manager_response.json()
        self.assertFalse(manager_payload['edit']['canEdit'])
        self.assertFalse(manager_payload['note']['canDelete'])
        self.assertTrue(manager_payload['scope']['canReview'])

        self.client.force_login(self.other_manager)
        other_response = self.client.get(reverse('reporting:notes_detail_api', args=[target.id]))
        self.assertEqual(other_response.status_code, 403)

    def test_notes_update_api_updates_owned_note(self):
        from django.utils import timezone

        target = self._create_note(self.user, '수정대상', action_type='customer_meeting', content='수정 전')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:notes_update_api', args=[target.id]),
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'customer_meeting',
                'activityDate': timezone.localdate().isoformat(),
                'content': 'React 상세에서 수정',
                'meetingSituation': '도입 검토',
                'meetingResearcherQuote': '다음 주에 다시 확인하겠습니다',
                'meetingConfirmedFacts': '예산은 6월 배정',
                'meetingObstacles': '내부 승인 필요',
                'meetingNextAction': '승인자 연락',
                'nextAction': '견적서 재발송',
                'nextActionDate': timezone.localdate().isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['message'], '영업노트를 수정했습니다.')
        target.refresh_from_db()
        self.assertEqual(target.content, 'React 상세에서 수정')
        self.assertEqual(target.meeting_situation, '')
        self.assertEqual(target.meeting_researcher_quote, '')
        self.assertEqual(target.meeting_confirmed_facts, '')
        self.assertEqual(target.meeting_obstacles, '')
        self.assertEqual(target.meeting_next_action, '')
        self.assertEqual(target.next_action, '견적서 재발송')
        self.assertEqual(target.meeting_date, timezone.localdate())

    def test_notes_update_api_updates_department_only_note_without_customer(self):
        from django.utils import timezone
        from reporting.models import FollowUp, History

        department = self._create_department_only(self.user, '부서수정대상')
        FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='나중에 등록된 담당자',
            manager='담당자',
            company=department.company,
            department=department,
            priority='scheduled',
            pipeline_stage='contact',
        )
        target = History.objects.create(
            user=self.user,
            company=self.company,
            department=department,
            action_type='customer_meeting',
            content='부서-only 수정 전',
            next_action='기존 액션',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:notes_update_api', args=[target.id]),
            data=json.dumps({
                'departmentId': department.id,
                'actionType': 'customer_meeting',
                'activityDate': timezone.localdate().isoformat(),
                'content': '부서-only React 수정',
                'nextAction': '담당자 확인',
                'nextActionDate': timezone.localdate().isoformat(),
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        target.refresh_from_db()
        self.assertIsNone(target.followup_id)
        self.assertEqual(target.department_id, department.id)
        self.assertEqual(target.content, '부서-only React 수정')
        self.assertEqual(target.next_action, '담당자 확인')
        self.assertEqual(target.meeting_date, timezone.localdate())
        self.assertEqual(payload['note']['customer'], '담당자 미등록')
        self.assertEqual(payload['note']['departmentId'], department.id)

    def test_notes_update_api_updates_service_note_without_status(self):
        target = self._create_note(self.user, '서비스수정대상', action_type='service', content='서비스 전')
        target.service_status = 'received'
        target.save(update_fields=['service_status'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:notes_update_api', args=[target.id]),
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'service',
                'content': '상태 없이 서비스 노트 수정',
                'nextAction': '처리 내용 확인',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        target.refresh_from_db()
        self.assertEqual(target.action_type, 'service')
        self.assertEqual(target.content, '상태 없이 서비스 노트 수정')
        self.assertIsNone(target.service_status)
        self.assertEqual(payload['note']['serviceStatus'], '')

    def test_notes_update_api_blocks_manager_and_other_company_customer(self):
        target = self._create_note(self.user, '수정차단', action_type='quote', content='견적 전')
        other_target = self._create_note(self.other_user, '타사고객', action_type='quote', content='타사')

        self.client.force_login(self.manager)
        manager_response = self.client.post(
            reverse('reporting:notes_update_api', args=[target.id]),
            data=json.dumps({
                'followupId': target.followup_id,
                'actionType': 'quote',
                'content': '매니저 수정 시도',
            }),
            content_type='application/json',
        )
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.user)
        other_company_response = self.client.post(
            reverse('reporting:notes_update_api', args=[target.id]),
            data=json.dumps({
                'followupId': other_target.followup_id,
                'actionType': 'quote',
                'content': '타사 고객으로 변경 시도',
            }),
            content_type='application/json',
        )
        self.assertEqual(other_company_response.status_code, 403)

    def test_notes_delete_api_allows_owner_only_and_blocks_comment_delete(self):
        from reporting.models import History

        target = self._create_note(self.user, '삭제대상', action_type='quote', content='삭제할 노트')
        delete_url = reverse('reporting:notes_delete_api', args=[target.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(delete_url)
        self.assertEqual(manager_response.status_code, 403)
        self.assertTrue(History.objects.filter(pk=target.id).exists())

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(delete_url)
        self.assertEqual(coworker_response.status_code, 403)
        self.assertTrue(History.objects.filter(pk=target.id).exists())

        reply = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='댓글은 노트 삭제 API 대상 아님',
        )
        self.client.force_login(self.user)
        reply_response = self.client.post(reverse('reporting:notes_delete_api', args=[reply.id]))
        self.assertEqual(reply_response.status_code, 403)
        self.assertTrue(History.objects.filter(pk=reply.id).exists())

        response = self.client.post(delete_url)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['redirect'], '/notes/')
        self.assertFalse(History.objects.filter(pk=target.id).exists())

    def test_note_file_upload_api_allows_owner_only(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import HistoryFile

        target = self._create_note(self.user, '파일업로드', action_type='quote', content='견적 파일')
        upload_url = reverse('reporting:note_file_upload', args=[target.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('manager.txt', b'manager memo', content_type='text/plain'),
        })
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('coworker.txt', b'coworker memo', content_type='text/plain'),
        })
        self.assertEqual(coworker_response.status_code, 403)

        self.client.force_login(self.user)
        response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('owner.txt', b'owner memo', content_type='text/plain'),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        history_file = HistoryFile.objects.get(history=target)
        self.addCleanup(history_file.file.delete, False)
        self.assertEqual(history_file.original_filename, 'owner.txt')
        self.assertEqual(payload['files'][0]['id'], history_file.id)
        self.assertEqual(payload['files'][0]['downloadHref'], reverse('reporting:file_download', args=[history_file.id]))
        self.assertEqual(payload['files'][0]['deleteHref'], reverse('reporting:file_delete', args=[history_file.id]))

    def test_note_file_delete_api_allows_owner_only(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import History, HistoryFile

        target = self._create_note(self.user, '파일삭제', action_type='quote', content='삭제 파일')
        history_file = HistoryFile.objects.create(
            history=target,
            file=SimpleUploadedFile('delete-me.txt', b'delete memo', content_type='text/plain'),
            original_filename='delete-me.txt',
            file_size=11,
            uploaded_by=self.user,
        )
        delete_url = reverse('reporting:file_delete', args=[history_file.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(delete_url)
        self.assertEqual(manager_response.status_code, 403)
        self.assertTrue(HistoryFile.objects.filter(pk=history_file.id).exists())

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(delete_url)
        self.assertEqual(coworker_response.status_code, 403)
        self.assertTrue(HistoryFile.objects.filter(pk=history_file.id).exists())

        memo = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='첨부 삭제 차단 댓글',
        )
        memo_file = HistoryFile.objects.create(
            history=memo,
            file=SimpleUploadedFile('memo-file.txt', b'memo file', content_type='text/plain'),
            original_filename='memo-file.txt',
            file_size=9,
            uploaded_by=self.user,
        )
        self.addCleanup(memo_file.file.delete, False)
        self.client.force_login(self.user)
        memo_response = self.client.post(reverse('reporting:file_delete', args=[memo_file.id]))
        self.assertEqual(memo_response.status_code, 403)
        self.assertTrue(HistoryFile.objects.filter(pk=memo_file.id).exists())

        response = self.client.post(delete_url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertFalse(HistoryFile.objects.filter(pk=history_file.id).exists())

    def test_note_file_download_blocks_anonymous_and_out_of_scope_users(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import HistoryFile

        target = self._create_note(self.user, '파일다운로드', action_type='quote', content='다운로드 파일')
        history_file = HistoryFile.objects.create(
            history=target,
            file=SimpleUploadedFile('download-note.txt', b'download memo', content_type='text/plain'),
            original_filename='download-note.txt',
            file_size=13,
            uploaded_by=self.user,
        )
        self.addCleanup(history_file.file.delete, False)
        download_url = reverse('reporting:file_download', args=[history_file.id])

        anonymous_response = self.client.get(download_url)
        self.assertEqual(anonymous_response.status_code, 302)
        self.assertIn('/reporting/login/', anonymous_response['Location'])

        self.client.force_login(self.other_user)
        other_response = self.client.get(download_url)
        self.assertNotEqual(other_response.status_code, 200)
        self.assertNotIn('attachment', other_response.get('Content-Disposition', ''))

        self.client.force_login(self.user)
        owner_response = self.client.get(download_url)
        self.assertEqual(owner_response.status_code, 200)
        self.assertIn('attachment', owner_response.get('Content-Disposition', ''))
        owner_response.close()

    def test_note_reply_create_api_allows_owner_and_same_company_manager(self):
        from reporting.models import History

        target = self._create_note(self.user, '댓글작성', action_type='quote', content='견적 댓글')
        reply_url = reverse('reporting:add_manager_memo_to_history_api', args=[target.id])

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(reply_url, {'memo': '동료 댓글 시도'})
        self.assertEqual(coworker_response.status_code, 403)

        self.client.force_login(self.other_manager)
        other_manager_response = self.client.post(reply_url, {'memo': '타사 매니저 댓글 시도'})
        self.assertEqual(other_manager_response.status_code, 403)

        self.client.force_login(self.user)
        owner_response = self.client.post(reply_url, {'memo': '실무자 댓글'})
        self.assertEqual(owner_response.status_code, 200)
        owner_reply = History.objects.get(parent_history=target, content='실무자 댓글')
        self.assertEqual(owner_reply.user, self.user)
        self.assertIsNone(owner_reply.created_by)

        self.client.force_login(self.manager)
        manager_response = self.client.post(reply_url, {'memo': '관리자 메모'})
        self.assertEqual(manager_response.status_code, 200)
        manager_reply = History.objects.get(parent_history=target, content='관리자 메모')
        self.assertEqual(manager_reply.user, self.user)
        self.assertEqual(manager_reply.created_by, self.manager)

    def test_note_reply_delete_api_allows_author_only(self):
        from reporting.models import History

        target = self._create_note(self.user, '댓글삭제', action_type='quote', content='삭제 댓글')
        owner_reply = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='실무자 삭제 댓글',
        )
        manager_reply = History.objects.create(
            user=self.user,
            company=self.company,
            followup=target.followup,
            parent_history=target,
            action_type='memo',
            content='매니저 삭제 메모',
            created_by=self.manager,
        )

        owner_delete_url = reverse('reporting:delete_manager_memo_api', args=[owner_reply.id])
        manager_delete_url = reverse('reporting:delete_manager_memo_api', args=[manager_reply.id])

        self.client.force_login(self.manager)
        manager_denied = self.client.delete(owner_delete_url)
        self.assertEqual(manager_denied.status_code, 403)
        self.assertTrue(History.objects.filter(pk=owner_reply.id).exists())

        self.client.force_login(self.user)
        owner_denied = self.client.delete(manager_delete_url)
        self.assertEqual(owner_denied.status_code, 403)
        self.assertTrue(History.objects.filter(pk=manager_reply.id).exists())

        owner_response = self.client.delete(owner_delete_url)
        self.assertEqual(owner_response.status_code, 200)
        self.assertFalse(History.objects.filter(pk=owner_reply.id).exists())

        self.client.force_login(self.manager)
        manager_response = self.client.delete(manager_delete_url)
        self.assertEqual(manager_response.status_code, 200)
        self.assertFalse(History.objects.filter(pk=manager_reply.id).exists())

    def test_history_toggle_reviewed_allows_manager_only(self):
        target = self._create_note(
            self.user,
            '토글대상',
            action_type='quote',
            content='견적 보고 검토',
            reviewed=False,
        )
        toggle_url = reverse('reporting:history_toggle_reviewed', args=[target.id])

        self.client.force_login(self.user)
        denied = self.client.post(toggle_url)
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.admin)
        admin_denied = self.client.post(toggle_url)
        self.assertEqual(admin_denied.status_code, 403)

        self.client.force_login(self.other_manager)
        other_manager_denied = self.client.post(toggle_url)
        self.assertEqual(other_manager_denied.status_code, 403)

        self.client.force_login(self.manager)
        response = self.client.post(toggle_url)
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertTrue(payload['is_reviewed'])
        target.refresh_from_db()
        self.assertIsNotNone(target.reviewed_at)
        self.assertEqual(target.reviewer, self.manager)


class PrepaymentsSummaryApiTests(TestCase):
    """React 선결제 현황 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='선결제API회사')
        self.other_company = UserCompany.objects.create(name='선결제API타사회사')
        self.user = make_user('prepayment_api_me', role='salesman', company=self.company)
        self.coworker = make_user('prepayment_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('prepayment_api_manager', role='manager', company=self.company)
        self.other_user = make_user('prepayment_api_other', role='salesman', company=self.other_company)
        self.url = reverse('reporting:prepayment_api_list')

    def _create_customer(self, owner, name):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        return FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            manager=f'{name} 책임자',
            company=customer_company,
            department=department,
        )

    def _create_prepayment(self, owner, name, amount=100000, balance=70000, status='active', payer='입금자'):
        from django.utils import timezone
        from reporting.models import Prepayment

        customer = self._create_customer(owner, name)
        return Prepayment.objects.create(
            department=customer.department,
            customer=customer,
            company=customer.company,
            amount=amount,
            balance=balance,
            payment_date=timezone.localdate(),
            payer_name=payer,
            status=status,
            created_by=owner,
        )

    def test_prepayment_summary_api_requires_login_json(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_prepayment_summary_api_defaults_to_current_user(self):
        own = self._create_prepayment(self.user, '내선결제', amount=100000, balance=70000, payer='내입금자')
        coworker = self._create_prepayment(self.coworker, '동료선결제', amount=200000, balance=200000, payer='동료입금자')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['scope']['dataFilter'], 'me')
        self.assertFalse(payload['scope']['isViewingOthers'])
        ids = {item['id'] for item in payload['prepayments']}
        self.assertIn(own.id, ids)
        self.assertNotIn(coworker.id, ids)
        self.assertEqual(payload['metrics']['totalAmount'], 100000)
        self.assertEqual(payload['metrics']['totalBalance'], 70000)
        self.assertEqual(payload['metrics']['totalUsed'], 30000)
        self.assertEqual(payload['links']['create'], reverse('reporting:prepayment_create'))
        own_payload = payload['prepayments'][0]
        self.assertEqual(own_payload['payerName'], '내입금자')
        self.assertEqual(own_payload['customerHref'], f'/customers/{own.customer_id}/')
        self.assertTrue(own_payload['canManage'])

    def test_prepayment_summary_api_filters_team_scope_search_and_status(self):
        own = self._create_prepayment(self.user, '내활성', amount=100000, balance=90000, status='active', payer='내입금')
        coworker = self._create_prepayment(
            self.coworker,
            '동료소진',
            amount=150000,
            balance=0,
            status='depleted',
            payer='동료입금자',
        )
        other = self._create_prepayment(
            self.other_user,
            '타사소진',
            amount=999000,
            balance=0,
            status='depleted',
            payer='동료입금자',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url, {
            'data_filter': 'all',
            'status': 'depleted',
            'search': '동료입금자',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['prepayments']}
        self.assertNotIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        self.assertTrue(payload['scope']['isViewingOthers'])
        self.assertEqual(payload['filters']['status'], 'depleted')
        self.assertEqual(payload['metrics']['depletedCount'], 1)
        self.assertEqual(payload['metrics']['totalAmount'], 150000)
        self.assertEqual(payload['links']['create'], '')

    def test_manager_prepayment_summary_defaults_to_company_scope_read_only(self):
        own = self._create_prepayment(self.user, '매니저회사내선결제', amount=100000, balance=70000, payer='내입금')
        coworker = self._create_prepayment(self.coworker, '매니저동료선결제', amount=200000, balance=50000, payer='동료입금')
        manager_owned = self._create_prepayment(self.manager, '매니저과거선결제', amount=300000, balance=300000, payer='매니저입금')
        other = self._create_prepayment(self.other_user, '매니저타사선결제', amount=999000, balance=999000, payer='타사입금')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['scope']['dataFilter'], 'all')
        self.assertTrue(payload['scope']['isViewingOthers'])
        ids = {item['id'] for item in payload['prepayments']}
        self.assertIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertIn(manager_owned.id, ids)
        self.assertNotIn(other.id, ids)
        self.assertEqual(payload['links']['create'], '')
        self.assertEqual(payload['metrics']['totalAmount'], 600000)
        owners = {item['ownerId']: item['ownerName'] for item in payload['prepayments']}
        self.assertEqual(owners[self.user.id], self.user.username)
        self.assertEqual(owners[self.coworker.id], self.coworker.username)
        self.assertEqual(owners[self.manager.id], self.manager.username)
        self.assertTrue(all(not item['canManage'] for item in payload['prepayments']))


class PrepaymentDetailApiTests(TestCase):
    """React 선결제 상세/등록/수정 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='선결제상세API회사')
        self.other_company = UserCompany.objects.create(name='선결제상세API타사회사')
        self.user = make_user('prepayment_detail_me', role='salesman', company=self.company)
        self.coworker = make_user('prepayment_detail_coworker', role='salesman', company=self.company)
        self.manager = make_user('prepayment_detail_manager', role='manager', company=self.company)
        self.other_user = make_user('prepayment_detail_other', role='salesman', company=self.other_company)

    def _create_customer(self, owner, name):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        return FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            company=customer_company,
            department=department,
        )

    def _create_prepayment(self, owner, name='선결제', amount=100000, balance=70000):
        from django.utils import timezone
        from reporting.models import Prepayment

        customer = self._create_customer(owner, name)
        return Prepayment.objects.create(
            department=customer.department,
            customer=customer,
            company=customer.company,
            amount=amount,
            balance=balance,
            payment_date=timezone.localdate(),
            payment_method='transfer',
            payer_name=f'{name} 입금자',
            memo='초기 메모',
            created_by=owner,
        )

    def test_prepayment_detail_api_returns_usage_and_edit_config(self):
        from datetime import time
        from django.utils import timezone
        from reporting.models import DeliveryItem, PrepaymentUsage, Schedule

        prepayment = self._create_prepayment(self.user, amount=120000, balance=90000)
        schedule = Schedule.objects.create(
            user=self.user,
            followup=prepayment.customer,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='테스트 품목',
            quantity=2,
            unit='EA',
            unit_price=15000,
            total_price=30000,
        )
        PrepaymentUsage.objects.create(
            prepayment=prepayment,
            schedule=schedule,
            product_name='테스트 품목',
            quantity=2,
            amount=30000,
            remaining_balance=90000,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:prepayment_detail_api', args=[prepayment.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['prepayment']['id'], prepayment.id)
        self.assertTrue(payload['edit']['canEdit'])
        self.assertEqual(payload['metrics']['usedAmount'], 30000)
        self.assertEqual(payload['usages'][0]['amount'], 30000)
        self.assertEqual(payload['usages'][0]['deliveryItems'][0]['itemName'], '테스트 품목')
        self.assertEqual(payload['links']['reactEdit'], f'/prepayments/{prepayment.id}/edit/')
        self.assertTrue(payload['actions']['canCancel'])
        self.assertFalse(payload['actions']['canDelete'])
        self.assertIn('1개의 사용 내역', payload['actions']['deleteMessage'])
        self.assertTrue(payload['actions']['canTransfer'])
        self.assertEqual(payload['actions']['cancelUrl'], reverse('reporting:prepayment_cancel_api', args=[prepayment.id]))
        self.assertEqual(payload['actions']['deleteUrl'], reverse('reporting:prepayment_delete_api', args=[prepayment.id]))
        self.assertEqual(payload['actions']['transferUrl'], reverse('reporting:prepayment_transfer_api', args=[prepayment.id]))
        self.assertIn(self.coworker.id, [user['id'] for user in payload['actions']['transferUsers']])

    def test_prepayment_detail_api_blocks_other_company(self):
        prepayment = self._create_prepayment(self.other_user)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:prepayment_detail_api', args=[prepayment.id]))

        self.assertEqual(response.status_code, 403)

    def test_prepayment_create_api_creates_with_initial_balance(self):
        from reporting.models import Prepayment

        customer = self._create_customer(self.user, '등록고객')
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:prepayment_create_api'), {
            'department': str(customer.department_id),
            'customer': str(customer.id),
            'amount': '250000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'payer_name': '등록 입금자',
            'memo': 'React 등록',
        })

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        created = Prepayment.objects.get(id=payload['prepaymentId'])
        self.assertEqual(created.created_by, self.user)
        self.assertEqual(int(created.amount), 250000)
        self.assertEqual(int(created.balance), 250000)
        self.assertEqual(created.company, customer.company)
        self.assertEqual(created.department, customer.department)
        self.assertEqual(payload['href'], f'/prepayments/{created.id}/')
        self.assertTrue(
            PrepaymentLedgerEntry.objects.filter(
                prepayment=created,
                department=customer.department,
                entry_type=PrepaymentLedgerEntry.ENTRY_DEPOSIT,
                amount=250000,
            ).exists()
        )

    def test_prepayment_create_api_allows_account_first_without_contact_payload(self):
        from reporting.models import Prepayment

        customer = self._create_customer(self.user, '계정우선등록')
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:prepayment_create_api'), {
            'department': str(customer.department_id),
            'amount': '180000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'payer_name': '계정 입금자',
        })

        self.assertEqual(response.status_code, 201)
        created = Prepayment.objects.get(id=response.json()['prepaymentId'])
        self.assertEqual(created.department, customer.department)
        self.assertEqual(created.customer, customer)
        self.assertEqual(response.json()['prepayment']['departmentId'], customer.department_id)

    def test_manager_cannot_create_or_manage_prepayments_even_when_owner(self):
        from reporting.models import Prepayment

        customer = self._create_customer(self.user, '매니저등록차단')
        self.client.force_login(self.manager)

        create_payload = self.client.get(reverse('reporting:prepayment_create_api')).json()
        self.assertFalse(create_payload['create']['canCreate'])
        self.assertEqual(create_payload['create']['submitUrl'], '')

        response = self.client.post(reverse('reporting:prepayment_create_api'), {
            'department': str(customer.department_id),
            'customer': str(customer.id),
            'amount': '250000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'payer_name': '매니저 입금자',
        })
        self.assertEqual(response.status_code, 403)
        self.assertFalse(Prepayment.objects.filter(payer_name='매니저 입금자').exists())

        manager_owned = self._create_prepayment(self.manager, name='매니저과거소유')
        detail = self.client.get(reverse('reporting:prepayment_detail_api', args=[manager_owned.id]))
        self.assertEqual(detail.status_code, 200)
        detail_payload = detail.json()
        self.assertFalse(detail_payload['edit']['canEdit'])
        self.assertFalse(detail_payload['actions']['canCancel'])
        self.assertFalse(detail_payload['actions']['canDelete'])
        self.assertFalse(detail_payload['actions']['canTransfer'])
        self.assertIn('Manager 계정', detail_payload['edit']['message'])

        denied_update = self.client.post(reverse('reporting:prepayment_update_api', args=[manager_owned.id]), {
            'department': str(manager_owned.customer.department_id),
            'customer': str(manager_owned.customer_id),
            'amount': '100000',
            'balance': '70000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'status': 'active',
        })
        self.assertEqual(denied_update.status_code, 403)

        denied_cancel = self.client.post(reverse('reporting:prepayment_cancel_api', args=[manager_owned.id]), {
            'cancel_reason': '매니저 취소 시도',
        })
        self.assertEqual(denied_cancel.status_code, 403)

        denied_delete = self.client.post(reverse('reporting:prepayment_delete_api', args=[manager_owned.id]))
        self.assertEqual(denied_delete.status_code, 403)

        denied_transfer = self.client.post(reverse('reporting:prepayment_transfer_api', args=[manager_owned.id]), {
            'target_user': str(self.coworker.id),
            'reason': '매니저 이관 시도',
        })
        self.assertEqual(denied_transfer.status_code, 403)
        manager_owned.refresh_from_db()
        self.assertEqual(manager_owned.created_by, self.manager)

    def test_prepayment_update_api_only_owner_and_validates_balance(self):
        prepayment = self._create_prepayment(self.user, amount=100000, balance=80000)

        self.client.force_login(self.coworker)
        denied = self.client.post(reverse('reporting:prepayment_update_api', args=[prepayment.id]), {
            'department': str(prepayment.customer.department_id),
            'customer': str(prepayment.customer_id),
            'amount': '100000',
            'balance': '70000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'status': 'active',
        })
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.user)
        invalid = self.client.post(reverse('reporting:prepayment_update_api', args=[prepayment.id]), {
            'department': str(prepayment.customer.department_id),
            'customer': str(prepayment.customer_id),
            'amount': '100000',
            'balance': '120000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'status': 'active',
        })
        self.assertEqual(invalid.status_code, 400)

        response = self.client.post(reverse('reporting:prepayment_update_api', args=[prepayment.id]), {
            'department': str(prepayment.customer.department_id),
            'customer': str(prepayment.customer_id),
            'amount': '110000',
            'balance': '70000',
            'payment_date': '2026-05-10',
            'payment_method': 'card',
            'payer_name': '수정 입금자',
            'status': 'active',
            'memo': 'React 수정',
        })
        self.assertEqual(response.status_code, 200)
        prepayment.refresh_from_db()
        self.assertEqual(int(prepayment.amount), 110000)
        self.assertEqual(int(prepayment.balance), 70000)
        self.assertEqual(prepayment.payment_method, 'card')
        self.assertEqual(prepayment.memo, 'React 수정')
        self.assertTrue(
            PrepaymentLedgerEntry.objects.filter(
                prepayment=prepayment,
                entry_type=PrepaymentLedgerEntry.ENTRY_ADJUSTMENT,
                balance_before=80000,
                balance_after=70000,
            ).exists()
        )

    def test_prepayment_update_api_keeps_contact_auxiliary_when_omitted(self):
        prepayment = self._create_prepayment(self.user, name='보조담당자', amount=100000, balance=80000)
        original_customer = prepayment.customer
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:prepayment_update_api', args=[prepayment.id]), {
            'department': str(original_customer.department_id),
            'amount': '100000',
            'balance': '75000',
            'payment_date': '2026-05-10',
            'payment_method': 'transfer',
            'status': 'active',
            'memo': '담당자 미지정 수정',
        })

        self.assertEqual(response.status_code, 200)
        prepayment.refresh_from_db()
        self.assertEqual(prepayment.department_id, original_customer.department_id)
        self.assertEqual(prepayment.customer_id, original_customer.id)
        self.assertEqual(int(prepayment.balance), 75000)

    def test_prepayment_cancel_api_only_owner_and_records_reason(self):
        prepayment = self._create_prepayment(self.user)

        self.client.force_login(self.coworker)
        denied = self.client.post(reverse('reporting:prepayment_cancel_api', args=[prepayment.id]), {
            'cancel_reason': '동료 취소 시도',
        })
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.user)
        response = self.client.post(reverse('reporting:prepayment_cancel_api', args=[prepayment.id]), {
            'cancel_reason': 'React 취소',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        prepayment.refresh_from_db()
        self.assertEqual(prepayment.status, 'cancelled')
        self.assertEqual(prepayment.cancel_reason, 'React 취소')
        self.assertIsNotNone(prepayment.cancelled_at)
        self.assertEqual(payload['prepayment']['status'], 'cancelled')
        self.assertTrue(
            PrepaymentLedgerEntry.objects.filter(
                prepayment=prepayment,
                entry_type=PrepaymentLedgerEntry.ENTRY_CANCELLATION,
                memo='React 취소',
            ).exists()
        )

    def test_prepayment_delete_api_blocks_used_records_and_deletes_unused(self):
        from reporting.models import Prepayment, PrepaymentUsage

        used_prepayment = self._create_prepayment(self.user)
        PrepaymentUsage.objects.create(
            prepayment=used_prepayment,
            product_name='삭제 차단 품목',
            quantity=1,
            amount=10000,
            remaining_balance=60000,
        )
        self.client.force_login(self.user)

        blocked = self.client.post(reverse('reporting:prepayment_delete_api', args=[used_prepayment.id]))
        self.assertEqual(blocked.status_code, 400)
        self.assertTrue(Prepayment.objects.filter(id=used_prepayment.id).exists())

        unused_prepayment = self._create_prepayment(self.user, name='삭제가능')
        deleted = self.client.post(reverse('reporting:prepayment_delete_api', args=[unused_prepayment.id]))

        self.assertEqual(deleted.status_code, 200)
        self.assertTrue(deleted.json()['success'])
        self.assertEqual(deleted.json()['href'], '/prepayments/')
        self.assertFalse(Prepayment.objects.filter(id=unused_prepayment.id).exists())

    def test_prepayment_transfer_api_moves_owner_and_appends_memo(self):
        prepayment = self._create_prepayment(self.user)
        self.client.force_login(self.user)

        other_company_response = self.client.post(reverse('reporting:prepayment_transfer_api', args=[prepayment.id]), {
            'target_user': str(self.other_user.id),
            'reason': '타사 이관 시도',
        })
        self.assertEqual(other_company_response.status_code, 400)

        response = self.client.post(reverse('reporting:prepayment_transfer_api', args=[prepayment.id]), {
            'target_user': str(self.coworker.id),
            'reason': '담당자 변경',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['prepayment']['ownerId'], self.coworker.id)
        self.assertFalse(payload['prepayment']['canManage'])

        prepayment.refresh_from_db()
        self.assertEqual(prepayment.created_by, self.coworker)
        self.assertIn('[이관]', prepayment.memo)
        self.assertIn('담당자 변경', prepayment.memo)
        self.assertTrue(
            PrepaymentLedgerEntry.objects.filter(
                prepayment=prepayment,
                entry_type=PrepaymentLedgerEntry.ENTRY_TRANSFER,
                actor=self.user,
                target_user=self.coworker,
            ).exists()
        )

        detail = self.client.get(reverse('reporting:prepayment_detail_api', args=[prepayment.id]))
        self.assertEqual(detail.status_code, 200)
        detail_payload = detail.json()
        self.assertFalse(detail_payload['edit']['canEdit'])
        self.assertFalse(detail_payload['actions']['canTransfer'])


class PrepaymentCustomerApiTests(TestCase):
    """React 고객별/부서별 선결제 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='고객별선결제API회사')
        self.other_company = UserCompany.objects.create(name='고객별선결제API타사회사')
        self.user = make_user('prepayment_customer_me', role='salesman', company=self.company)
        self.coworker = make_user('prepayment_customer_coworker', role='salesman', company=self.company)
        self.manager = make_user('prepayment_customer_manager', role='manager', company=self.company)
        self.other_user = make_user('prepayment_customer_other', role='salesman', company=self.other_company)

    def _create_department_customers(self, owner=None):
        from reporting.models import Company, Department, FollowUp

        owner = owner or self.user
        customer_company = Company.objects.create(name=f'고객별선결제 고객사 {owner.username}', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name='공동 연구실',
            created_by=owner,
        )
        first = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name='1번 담당자',
            company=customer_company,
            department=department,
        )
        second = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name='2번 담당자',
            company=customer_company,
            department=department,
        )
        return customer_company, department, first, second

    def _create_prepayment(self, owner, customer, amount=100000, balance=70000, status='active', payer='입금자'):
        from django.utils import timezone
        from reporting.models import Prepayment

        return Prepayment.objects.create(
            department=customer.department,
            customer=customer,
            company=customer.company,
            amount=amount,
            balance=balance,
            payment_date=timezone.localdate(),
            payment_method='transfer',
            payer_name=payer,
            memo='고객별 메모',
            status=status,
            created_by=owner,
        )

    def test_customer_prepayment_api_returns_department_scope_and_metrics(self):
        _company, department, first, second = self._create_department_customers()
        self._create_prepayment(self.user, first, amount=100000, balance=70000, status='active', payer='첫 입금')
        self._create_prepayment(self.user, second, amount=200000, balance=0, status='depleted', payer='둘째 입금')
        self._create_prepayment(self.coworker, first, amount=500000, balance=500000, payer='동료 입금')
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['scope']['mode'], 'department')
        self.assertEqual(payload['customer']['departmentId'], department.id)
        self.assertEqual(payload['scope']['targetUserId'], self.user.id)
        self.assertEqual(len(payload['departmentCustomers']), 2)
        self.assertEqual(payload['metrics']['totalAmount'], 300000)
        self.assertEqual(payload['metrics']['totalBalance'], 70000)
        self.assertEqual(payload['metrics']['totalUsed'], 230000)
        self.assertEqual(payload['metrics']['totalCount'], 2)
        self.assertEqual(payload['metrics']['activeCount'], 1)
        self.assertEqual(payload['metrics']['depletedCount'], 1)
        self.assertEqual([item['customerId'] for item in payload['prepayments']], [first.id, second.id])
        self.assertEqual(payload['links']['reactAccount'], f'/prepayments/account/{department.id}/')
        self.assertEqual(payload['links']['reactCustomer'], f'/prepayments/customer/{first.id}/')
        self.assertEqual(payload['links']['accountDetail'], f'/accounts/{department.id}/')
        self.assertEqual(payload['links']['djangoExcel'], reverse('reporting:prepayment_customer_excel', args=[first.id]))
        self.assertEqual(payload['prepayments'][0]['customerPrepaymentHref'], f'/prepayments/account/{department.id}/')

    def test_account_prepayment_api_returns_department_scope_and_metrics(self):
        from datetime import time
        from reporting.models import DeliveryItem, PrepaymentLedgerEntry, PrepaymentUsage, Schedule

        _company, department, first, second = self._create_department_customers()
        first_prepayment = self._create_prepayment(self.user, first, amount=110000, balance=90000, status='active', payer='계정 첫 입금')
        self._create_prepayment(self.user, second, amount=220000, balance=20000, status='active', payer='계정 둘째 입금')
        self._create_prepayment(self.coworker, first, amount=500000, balance=500000, payer='동료 입금')
        schedule = Schedule.objects.create(
            user=self.user,
            followup=first,
            visit_date=timezone.localdate(),
            visit_time=time(10, 30),
            activity_type='delivery',
            status='completed',
            use_prepayment=True,
            prepayment=first_prepayment,
            prepayment_amount=20000,
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='계정 차감 품목',
            quantity=1,
            unit='EA',
            unit_price=20000,
            total_price=20000,
        )
        usage = PrepaymentUsage.objects.create(
            prepayment=first_prepayment,
            schedule=schedule,
            product_name='계정 차감 품목',
            quantity=1,
            amount=20000,
            remaining_balance=90000,
        )
        PrepaymentLedgerEntry.objects.create(
            prepayment=first_prepayment,
            department=department,
            customer=first,
            schedule=schedule,
            usage=usage,
            entry_type=PrepaymentLedgerEntry.ENTRY_DELIVERY_DEDUCTION,
            amount=20000,
            balance_before=110000,
            balance_after=90000,
            actor=self.user,
            memo='계정 차감 테스트',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:prepayment_account_api', args=[department.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['scope']['mode'], 'department')
        self.assertEqual(payload['customer']['departmentId'], department.id)
        self.assertEqual(payload['scope']['targetUserId'], self.user.id)
        self.assertEqual(payload['metrics']['totalAmount'], 330000)
        self.assertEqual(payload['metrics']['totalBalance'], 110000)
        self.assertEqual(payload['metrics']['totalUsed'], 220000)
        self.assertEqual(payload['metrics']['deductionCount'], 1)
        self.assertGreaterEqual(payload['metrics']['ledgerCount'], 1)
        self.assertEqual([item['customerId'] for item in payload['prepayments']], [first.id, second.id])
        self.assertEqual(payload['balanceRows'][0]['departmentId'], department.id)
        self.assertEqual(payload['deductionRows'][0]['amount'], 20000)
        self.assertEqual(payload['deductionRows'][0]['deliveryItems'][0]['itemName'], '계정 차감 품목')
        self.assertEqual(payload['ledgerEntries'][0]['entryType'], PrepaymentLedgerEntry.ENTRY_DELIVERY_DEDUCTION)
        self.assertEqual(payload['links']['reactAccount'], f'/prepayments/account/{department.id}/')
        self.assertEqual(payload['links']['accountDetail'], f'/accounts/{department.id}/')
        self.assertEqual(payload['links']['accountExcel'], reverse('reporting:prepayment_account_excel', args=[department.id]))

        excel_response = self.client.get(reverse('reporting:prepayment_account_excel', args=[department.id]))
        self.assertEqual(excel_response.status_code, 200)
        self.assertIn('spreadsheetml.sheet', excel_response['Content-Type'])

    def test_account_prepayment_api_allows_salesman_with_own_prepayment_and_blocks_unrelated(self):
        _company, department, first, _second = self._create_department_customers(owner=self.coworker)
        self._create_prepayment(self.user, first, amount=90000, balance=50000, payer='계정 접근 허용 입금')

        self.client.force_login(self.user)
        allowed = self.client.get(reverse('reporting:prepayment_account_api', args=[department.id]))
        self.assertEqual(allowed.status_code, 200)
        self.assertEqual(allowed.json()['metrics']['totalAmount'], 90000)

        self.client.force_login(self.other_user)
        blocked = self.client.get(reverse('reporting:prepayment_account_api', args=[department.id]))
        self.assertEqual(blocked.status_code, 403)

    def test_customer_prepayment_api_uses_selected_accessible_user_for_manager(self):
        _company, _department, first, _second = self._create_department_customers()
        self._create_prepayment(self.user, first, amount=100000, balance=100000, payer='내 입금')
        self._create_prepayment(self.coworker, first, amount=250000, balance=150000, payer='동료 입금')
        session = self.client.session
        session['selected_user_id'] = str(self.coworker.id)
        session.save()
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['scope']['canSelectUser'])
        self.assertEqual(payload['scope']['targetUserId'], self.coworker.id)
        self.assertEqual(payload['metrics']['totalAmount'], 250000)
        self.assertEqual(len(payload['prepayments']), 1)
        self.assertEqual(payload['prepayments'][0]['ownerId'], self.coworker.id)

    def test_customer_prepayment_api_manager_defaults_to_company_all_users(self):
        _company, _department, first, _second = self._create_department_customers()
        self._create_prepayment(self.user, first, amount=100000, balance=100000, payer='내 입금')
        self._create_prepayment(self.coworker, first, amount=250000, balance=150000, payer='동료 입금')
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['scope']['canSelectUser'])
        self.assertTrue(payload['scope']['isAllUsers'])
        self.assertIsNone(payload['scope']['targetUserId'])
        self.assertEqual(payload['metrics']['totalAmount'], 350000)
        owner_ids = {item['ownerId'] for item in payload['prepayments']}
        self.assertEqual(owner_ids, {self.user.id, self.coworker.id})
        self.assertTrue(all(not item['canManage'] for item in payload['prepayments']))

    def test_customer_prepayment_api_allows_salesman_with_own_prepayment_and_blocks_unrelated(self):
        _company, _department, first, _second = self._create_department_customers(owner=self.coworker)
        self._create_prepayment(self.user, first, amount=90000, balance=50000, payer='접근 허용 입금')

        self.client.force_login(self.user)
        allowed = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))
        self.assertEqual(allowed.status_code, 200)
        self.assertEqual(allowed.json()['metrics']['totalAmount'], 90000)

        self.client.force_login(self.other_user)
        blocked = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))
        self.assertEqual(blocked.status_code, 403)

    def test_customer_prepayment_api_requires_login(self):
        _company, _department, first, _second = self._create_department_customers()

        response = self.client.get(reverse('reporting:prepayment_customer_api', args=[first.id]))

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')


class SchedulesSummaryApiTests(TestCase):
    """React 일정 화면 읽기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='일정API회사')
        self.other_company = UserCompany.objects.create(name='일정API타사회사')
        self.user = make_user('schedules_api_me', role='salesman', company=self.company)
        self.coworker = make_user('schedules_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('schedules_api_manager', role='manager', company=self.company)
        self.other_user = make_user('schedules_api_other', role='salesman', company=self.other_company)
        self.url = reverse('reporting:schedules_summary_api')
        self.create_url = reverse('reporting:schedules_create_api')
        self.calendar_url = reverse('reporting:schedules_calendar_api')
        self.personal_create_url = reverse('reporting:personal_schedules_create_api')

    def _create_customer(self, owner, name):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        return FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            manager=f'{name} 책임',
            company=customer_company,
            department=department,
            priority='urgent',
            pipeline_stage='quote',
        )

    def _create_department_only(self, owner, name):
        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        return Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )

    def _create_schedule(
        self,
        owner,
        name,
        activity_type='customer_meeting',
        status='scheduled',
        visit_date=None,
    ):
        import datetime
        from django.utils import timezone
        from reporting.models import Schedule

        followup = self._create_customer(owner, name)
        return Schedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            visit_date=visit_date or timezone.localdate(),
            visit_time=datetime.time(9, 0),
            activity_type=activity_type,
            status=status,
            location=f'{name} 회의실',
            notes=f'{name} 일정 메모',
        )

    def _create_personal_schedule(self, owner, title, schedule_date=None):
        import datetime
        from django.utils import timezone
        from reporting.models import PersonalSchedule

        return PersonalSchedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            title=title,
            content=f'{title} 내용',
            schedule_date=schedule_date or timezone.localdate(),
            schedule_time=datetime.time(14, 0),
        )

    def test_schedules_summary_api_requires_login_json(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_schedules_summary_api_uses_salesman_own_scope(self):
        own = self._create_schedule(self.user, '내일정')
        personal = self._create_personal_schedule(self.user, '내 개인 일정')
        coworker = self._create_schedule(self.coworker, '동료일정')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {(item['type'], item['id']) for item in payload['schedules']}
        self.assertIn(('customer', own.id), ids)
        self.assertIn(('personal', personal.id), ids)
        self.assertNotIn(('customer', coworker.id), ids)
        self.assertEqual(payload['metrics']['totalSchedules'], 2)
        self.assertTrue(payload['create']['canCreate'])
        self.assertEqual(payload['create']['submitUrl'], self.create_url)
        self.assertTrue(any(customer['id'] == own.followup_id for customer in payload['create']['customers']))
        self.assertTrue(any(department['id'] == own.followup.department_id for department in payload['create']['departments']))
        own_item = next(item for item in payload['schedules'] if item['type'] == 'customer' and item['id'] == own.id)
        self.assertEqual(own_item['href'], f'/schedules/{own.id}/')
        self.assertEqual(own_item['djangoHref'], reverse('reporting:schedule_detail', args=[own.id]))

    def test_schedules_summary_api_includes_customerless_departments_for_create(self):
        department = self._create_department_only(self.user, '고객없는일정대상')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        department_ids = {item['id'] for item in payload['create']['departments']}
        customer_department_ids = {item.get('departmentId') for item in payload['create']['customers']}
        self.assertIn(department.id, department_ids)
        self.assertNotIn(department.id, customer_department_ids)

    def test_schedules_summary_api_defaults_to_latest_schedule_first(self):
        from datetime import timedelta
        from django.utils import timezone

        older = self._create_schedule(
            self.user,
            '오래된일정',
            visit_date=timezone.localdate() - timedelta(days=4),
        )
        newer = self._create_schedule(
            self.user,
            '최근일정',
            visit_date=timezone.localdate() + timedelta(days=2),
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['schedules'][0]['type'], 'customer')
        self.assertEqual(payload['schedules'][0]['id'], newer.id)
        self.assertNotEqual(payload['schedules'][0]['id'], older.id)

    def test_schedules_summary_api_filters_search_owner_status_activity_and_range(self):
        from datetime import timedelta
        from django.utils import timezone

        target = self._create_schedule(
            self.user,
            'PCR핵심',
            activity_type='quote',
            status='scheduled',
            visit_date=timezone.localdate() + timedelta(days=1),
        )
        self._create_schedule(
            self.user,
            'PCR완료',
            activity_type='quote',
            status='completed',
            visit_date=timezone.localdate() + timedelta(days=1),
        )
        self._create_schedule(
            self.user,
            'PCR서비스',
            activity_type='service',
            status='scheduled',
            visit_date=timezone.localdate() + timedelta(days=1),
        )
        self._create_schedule(
            self.coworker,
            'PCR동료',
            activity_type='quote',
            status='scheduled',
            visit_date=timezone.localdate() + timedelta(days=1),
        )
        self.client.force_login(self.manager)

        response = self.client.get(self.url, {
            'q': 'PCR',
            'owner': str(self.user.id),
            'status': 'scheduled',
            'activityType': 'quote',
            'range': 'week',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = [(item['type'], item['id']) for item in payload['schedules']]
        self.assertEqual(ids, [('customer', target.id)])
        self.assertEqual(payload['filters']['q'], 'PCR')
        self.assertTrue(any(option['value'] == 'quote' for option in payload['options']['activityTypes']))
        self.assertFalse(any(option['value'] == 'service' for option in payload['options']['activityTypes']))
        self.assertEqual(payload['metrics']['filteredSchedules'], 1)

    def test_schedules_summary_api_excludes_service_schedule_type(self):
        self._create_schedule(self.user, '미팅일정', activity_type='customer_meeting')
        service_schedule = self._create_schedule(self.user, '서비스일정', activity_type='service')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['schedules'] if item['type'] == 'customer'}
        self.assertNotIn(service_schedule.id, ids)
        self.assertFalse(any(option['value'] == 'service' for option in payload['create']['activityTypes']))
        self.assertFalse(any(item['value'] == 'service' for item in payload['activityCounts']))
        self.assertEqual(payload['metrics']['customerSchedules'], 1)

    def test_schedules_summary_api_manager_sees_same_company_only(self):
        own = self._create_schedule(self.user, '회사내일정')
        coworker = self._create_schedule(self.coworker, '회사내동료일정')
        other = self._create_schedule(self.other_user, '타사일정')
        self.client.force_login(self.manager)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['schedules'] if item['type'] == 'customer'}
        self.assertIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        self.assertEqual(payload['metrics']['totalSchedules'], 2)
        self.assertTrue(payload['scope']['canViewAll'])
        self.assertFalse(payload['create']['canCreate'])

    def test_schedules_calendar_api_requires_login_json(self):
        response = self.client.get(self.calendar_url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_schedules_calendar_api_returns_month_range_items(self):
        import datetime
        from reporting.models import History

        target_date = datetime.date(2026, 5, 10)
        outside_date = datetime.date(2026, 6, 1)
        own = self._create_schedule(self.user, '월간일정', visit_date=target_date)
        personal = self._create_personal_schedule(self.user, '월간 개인 일정', schedule_date=target_date)
        outside = self._create_schedule(self.user, '범위밖일정', visit_date=outside_date)
        report = History.objects.create(
            user=self.user,
            company=self.company,
            followup=own.followup,
            schedule=own,
            action_type='customer_meeting',
            content='캘린더에서 보여줄 미팅 보고 본문',
            meeting_situation='PCR 장비 도입 검토 중',
            meeting_confirmed_facts='예산 담당자 확인',
            meeting_next_action='견적서 송부',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.calendar_url, {
            'start': '2026-05-01',
            'end': '2026-05-31',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {(item['type'], item['id']) for item in payload['schedules']}
        self.assertIn(('customer', own.id), ids)
        self.assertIn(('personal', personal.id), ids)
        self.assertNotIn(('customer', outside.id), ids)
        own_item = next(item for item in payload['schedules'] if item['type'] == 'customer' and item['id'] == own.id)
        personal_item = next(item for item in payload['schedules'] if item['type'] == 'personal' and item['id'] == personal.id)
        self.assertTrue(own_item['canEdit'])
        self.assertEqual(own_item['statusUpdateHref'], reverse('reporting:schedule_status_update', args=[own.id]))
        self.assertEqual(own_item['djangoEditHref'], reverse('reporting:schedule_edit', args=[own.id]))
        self.assertEqual(own_item['deleteHref'], reverse('reporting:schedule_delete', args=[own.id]))
        self.assertEqual(
            {option['value'] for option in own_item['statusOptions']},
            {'scheduled', 'completed', 'cancelled'},
        )
        self.assertEqual(own_item['reports'][0]['id'], report.id)
        self.assertEqual(own_item['reports'][0]['content'], '캘린더에서 보여줄 미팅 보고 본문')
        self.assertEqual(own_item['reports'][0]['meetingSituation'], 'PCR 장비 도입 검토 중')
        self.assertEqual(own_item['reports'][0]['meetingConfirmedFacts'], '예산 담당자 확인')
        self.assertEqual(own_item['reports'][0]['nextAction'], '견적서 송부')
        self.assertTrue(personal_item['canEdit'])
        self.assertEqual(personal_item['href'], f'/schedules/calendar/?personal={personal.id}&month=2026-05')
        self.assertEqual(personal_item['deleteHref'], reverse('reporting:personal_schedules_delete_api', args=[personal.id]))
        self.assertEqual(personal_item['djangoEditHref'], reverse('reporting:personal_schedule_edit', args=[personal.id]))
        self.assertEqual(personal_item['statusOptions'], [])
        self.assertEqual(personal_item['reports'], [])
        self.assertEqual(payload['filters']['start'], '2026-05-01')
        self.assertEqual(payload['filters']['end'], '2026-05-31')
        self.assertEqual(payload['metrics']['totalSchedules'], 2)
        self.assertEqual(payload['links']['calendar'], '/schedules/calendar/')
        self.assertEqual(payload['links']['djangoCalendar'], reverse('reporting:schedule_calendar'))
        self.assertTrue(payload['create']['canCreate'])
        self.assertEqual(payload['create']['submitUrl'], self.create_url)
        self.assertFalse(any(option['value'] == 'service' for option in payload['create']['activityTypes']))
        self.assertEqual(payload['create']['personalSchedule']['submitUrl'], self.personal_create_url)
        self.assertTrue(any(customer['id'] == own.followup_id for customer in payload['create']['customers']))

    def test_schedules_calendar_api_excludes_service_schedule_type(self):
        import datetime

        target_date = datetime.date(2026, 5, 10)
        meeting = self._create_schedule(self.user, '월간미팅', activity_type='customer_meeting', visit_date=target_date)
        service = self._create_schedule(self.user, '월간서비스', activity_type='service', visit_date=target_date)
        self.client.force_login(self.user)

        response = self.client.get(self.calendar_url, {
            'start': '2026-05-01',
            'end': '2026-05-31',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['schedules'] if item['type'] == 'customer'}
        self.assertIn(meeting.id, ids)
        self.assertNotIn(service.id, ids)
        self.assertEqual(payload['metrics']['customerSchedules'], 1)

    def test_schedules_calendar_api_all_filter_uses_same_company_only(self):
        import datetime

        target_date = datetime.date(2026, 5, 10)
        own = self._create_schedule(self.user, '회사내월간일정', visit_date=target_date)
        coworker = self._create_schedule(self.coworker, '동료월간일정', visit_date=target_date)
        own_personal = self._create_personal_schedule(self.user, '회사내개인월간일정', schedule_date=target_date)
        coworker_personal = self._create_personal_schedule(self.coworker, '동료개인월간일정', schedule_date=target_date)
        other = self._create_schedule(self.other_user, '타사회사월간일정', visit_date=target_date)
        self.client.force_login(self.user)

        response = self.client.get(self.calendar_url, {
            'start': '2026-05-01',
            'end': '2026-05-31',
            'data_filter': 'all',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['schedules'] if item['type'] == 'customer'}
        self.assertIn(own.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        own_item = next(item for item in payload['schedules'] if item['type'] == 'customer' and item['id'] == own.id)
        coworker_item = next(item for item in payload['schedules'] if item['type'] == 'customer' and item['id'] == coworker.id)
        self.assertTrue(own_item['canEdit'])
        self.assertFalse(coworker_item['canEdit'])
        self.assertEqual(coworker_item['statusUpdateHref'], '')
        self.assertEqual(coworker_item['deleteHref'], '')
        own_personal_item = next(item for item in payload['schedules'] if item['type'] == 'personal' and item['id'] == own_personal.id)
        coworker_personal_item = next(item for item in payload['schedules'] if item['type'] == 'personal' and item['id'] == coworker_personal.id)
        self.assertTrue(own_personal_item['canEdit'])
        self.assertEqual(own_personal_item['deleteHref'], reverse('reporting:personal_schedules_delete_api', args=[own_personal.id]))
        self.assertFalse(coworker_personal_item['canEdit'])
        self.assertEqual(coworker_personal_item['deleteHref'], '')
        self.assertEqual(payload['scope']['dataFilter'], 'all')
        self.assertTrue(any(option['id'] == self.coworker.id for option in payload['options']['users']))

    def test_schedules_calendar_api_manager_defaults_to_company_scope_without_me_filter(self):
        import datetime

        target_date = datetime.date(2026, 5, 10)
        manager_schedule = self._create_schedule(self.manager, '매니저월간일정', visit_date=target_date)
        coworker = self._create_schedule(self.coworker, '직원월간일정', visit_date=target_date)
        other = self._create_schedule(self.other_user, '타사회사월간일정', visit_date=target_date)
        self.client.force_login(self.manager)

        response = self.client.get(self.calendar_url, {
            'start': '2026-05-01',
            'end': '2026-05-31',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {item['id'] for item in payload['schedules'] if item['type'] == 'customer'}
        self.assertIn(manager_schedule.id, ids)
        self.assertIn(coworker.id, ids)
        self.assertNotIn(other.id, ids)
        self.assertEqual(payload['scope']['dataFilter'], 'all')
        self.assertEqual(payload['filters']['dataFilter'], 'all')
        filter_options = payload['options']['dataFilters']
        self.assertEqual(filter_options[0], {'value': 'all', 'label': '직원전체'})
        self.assertFalse(any(option['value'] == 'me' for option in filter_options))
        self.assertFalse(payload['create']['canCreate'])

    def test_schedules_calendar_api_user_filter_limits_to_selected_company_user(self):
        import datetime

        target_date = datetime.date(2026, 5, 10)
        self._create_schedule(self.user, '내월간일정', visit_date=target_date)
        coworker = self._create_schedule(self.coworker, '선택직원월간일정', visit_date=target_date)
        self.client.force_login(self.manager)

        response = self.client.get(self.calendar_url, {
            'start': '2026-05-01',
            'end': '2026-05-31',
            'data_filter': 'user',
            'filter_user': str(self.coworker.id),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = [(item['type'], item['id']) for item in payload['schedules']]
        self.assertEqual(ids, [('customer', coworker.id)])
        self.assertEqual(payload['scope']['dataFilter'], 'user')
        self.assertEqual(payload['scope']['filterUserId'], self.coworker.id)

    def test_schedules_create_api_requires_login_json(self):
        import json

        response = self.client.post(
            self.create_url,
            data=json.dumps({}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_schedules_create_api_blocks_manager(self):
        import json

        followup = self._create_customer(self.user, '매니저차단')
        self.client.force_login(self.manager)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)

    def test_personal_schedules_create_api_salesman_creates_own_schedule(self):
        import json
        from reporting.models import History, PersonalSchedule

        self.client.force_login(self.user)

        response = self.client.post(
            self.personal_create_url,
            data=json.dumps({
                'title': 'React 개인 일정',
                'content': 'React 개인 일정 내용',
                'scheduleDate': '2026-05-10',
                'scheduleTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertTrue(payload['success'])
        personal_schedule = PersonalSchedule.objects.get(pk=payload['scheduleId'])
        self.assertEqual(personal_schedule.user, self.user)
        self.assertEqual(personal_schedule.company, self.company)
        self.assertEqual(personal_schedule.title, 'React 개인 일정')
        self.assertEqual(personal_schedule.schedule_date.isoformat(), '2026-05-10')
        self.assertEqual(personal_schedule.schedule_time.strftime('%H:%M'), '10:30')
        self.assertEqual(payload['href'], f'/schedules/calendar/?personal={personal_schedule.id}&month=2026-05')
        self.assertEqual(payload['schedule']['id'], personal_schedule.id)
        self.assertEqual(payload['schedule']['href'], f'/schedules/calendar/?personal={personal_schedule.id}&month=2026-05')
        self.assertTrue(payload['edit']['canEdit'])
        self.assertEqual(
            payload['edit']['submitUrl'],
            reverse('reporting:personal_schedules_update_api', args=[personal_schedule.id]),
        )
        self.assertTrue(History.objects.filter(
            personal_schedule=personal_schedule,
            parent_history__isnull=True,
            content='개인 일정: React 개인 일정',
        ).exists())

    def test_personal_schedules_create_api_blocks_manager(self):
        import json
        from reporting.models import PersonalSchedule

        self.client.force_login(self.manager)

        response = self.client.post(
            self.personal_create_url,
            data=json.dumps({
                'title': '매니저 개인 일정',
                'scheduleDate': '2026-05-10',
                'scheduleTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(PersonalSchedule.objects.filter(title='매니저 개인 일정').exists())

    def test_personal_schedules_detail_api_manager_reads_same_company_without_edit(self):
        personal_schedule = self._create_personal_schedule(self.user, '매니저조회 개인 일정')
        other_schedule = self._create_personal_schedule(self.other_user, '타사 개인 일정')
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:personal_schedules_detail_api', args=[personal_schedule.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['schedule']['id'], personal_schedule.id)
        self.assertEqual(payload['schedule']['href'], f'/schedules/calendar/?personal={personal_schedule.id}&month={personal_schedule.schedule_date:%Y-%m}')
        self.assertFalse(payload['edit']['canEdit'])
        self.assertEqual(payload['links']['deleteSchedule'], '')

        denied = self.client.get(reverse('reporting:personal_schedules_detail_api', args=[other_schedule.id]))
        self.assertEqual(denied.status_code, 403)

    def test_personal_schedules_update_and_delete_api_are_owner_only(self):
        import json
        from reporting.models import History, PersonalSchedule

        personal_schedule = self._create_personal_schedule(self.user, '수정전 개인 일정')
        History.objects.create(
            user=self.user,
            company=self.company,
            personal_schedule=personal_schedule,
            action_type='memo',
            content='개인 일정: 수정전 개인 일정',
            created_by=self.user,
        )
        update_url = reverse('reporting:personal_schedules_update_api', args=[personal_schedule.id])
        delete_url = reverse('reporting:personal_schedules_delete_api', args=[personal_schedule.id])

        self.client.force_login(self.coworker)
        blocked_update = self.client.post(
            update_url,
            data=json.dumps({
                'title': '동료 수정 시도',
                'scheduleDate': '2026-05-11',
                'scheduleTime': '11:30',
            }),
            content_type='application/json',
        )
        self.assertEqual(blocked_update.status_code, 403)
        blocked_delete = self.client.post(delete_url)
        self.assertEqual(blocked_delete.status_code, 403)
        self.assertTrue(PersonalSchedule.objects.filter(pk=personal_schedule.id).exists())

        self.client.force_login(self.user)
        response = self.client.post(
            update_url,
            data=json.dumps({
                'title': '수정후 개인 일정',
                'content': '수정된 개인 일정 내용',
                'scheduleDate': '2026-05-11',
                'scheduleTime': '11:30',
            }),
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        personal_schedule.refresh_from_db()
        self.assertEqual(personal_schedule.title, '수정후 개인 일정')
        self.assertEqual(personal_schedule.content, '수정된 개인 일정 내용')
        self.assertEqual(personal_schedule.schedule_date.isoformat(), '2026-05-11')
        self.assertEqual(personal_schedule.schedule_time.strftime('%H:%M'), '11:30')
        self.assertEqual(
            History.objects.get(personal_schedule=personal_schedule, parent_history__isnull=True).content,
            '개인 일정: 수정후 개인 일정',
        )

        delete_response = self.client.post(delete_url)
        self.assertEqual(delete_response.status_code, 200)
        self.assertTrue(delete_response.json()['success'])
        self.assertFalse(PersonalSchedule.objects.filter(pk=personal_schedule.id).exists())

    def test_schedules_create_api_salesman_creates_own_schedule(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.user, '빠른등록')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'quote',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'location': '고객 회의실',
                'notes': '견적 일정 등록',
                'expectedRevenue': '1200000',
                'probability': '63',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertTrue(payload['success'])
        schedule = Schedule.objects.get(pk=payload['scheduleId'])
        self.assertEqual(schedule.user, self.user)
        self.assertEqual(schedule.followup, followup)
        self.assertEqual(schedule.activity_type, 'quote')
        self.assertEqual(schedule.location, '고객 회의실')
        self.assertEqual(int(schedule.expected_revenue), 1200000)
        self.assertEqual(schedule.probability, 65)
        self.assertEqual(payload['schedule']['probability'], 65)
        self.assertEqual(payload['schedule']['id'], schedule.id)
        self.assertEqual(payload['href'], f'/schedules/{schedule.id}/')

    def test_schedules_create_api_quote_advances_pipeline_card(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.user, '자동견적카드')
        followup.pipeline_stage = 'contact'
        followup.save(update_fields=['pipeline_stage'])
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'quote',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'expectedRevenue': '1200000',
                'probability': '63',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        schedule = Schedule.objects.get(pk=response.json()['scheduleId'])
        followup.refresh_from_db()
        self.assertEqual(schedule.activity_type, 'quote')
        self.assertEqual(followup.pipeline_stage, 'quote')

        pipeline_response = self.client.get(reverse('reporting:pipeline_command_center_api'))
        self.assertEqual(pipeline_response.status_code, 200)
        deal = next(item for item in pipeline_response.json()['deals'] if item['id'] == followup.id)
        self.assertEqual(deal['stage'], 'quote')
        self.assertEqual(deal['stageLabel'], '견적 제출')
        self.assertEqual(deal['value'], 1200000)
        self.assertEqual(deal['latestQuote']['basisType'], 'schedule')
        self.assertEqual(deal['latestQuote']['source'], '견적 일정')
        self.assertEqual(deal['latestQuote']['quoteDate'], '2026-05-10')

    def test_schedules_create_api_requires_quote_probability(self):
        import json

        followup = self._create_customer(self.user, '견적확률필수')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'quote',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'expectedRevenue': '1200000',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('견적 성공 확률은 필수입니다.', response.json()['error'])

    def test_schedules_create_api_meeting_probability_is_optional_null(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.user, '미팅확률선택')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'notes': '미팅 확률 미입력',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        schedule = Schedule.objects.get(pk=payload['scheduleId'])
        self.assertIsNone(schedule.probability)
        self.assertIsNone(payload['schedule']['probability'])

    def test_schedules_create_api_meeting_probability_is_normalized(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.user, '미팅확률보정')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'probability': '62',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        schedule = Schedule.objects.get(pk=payload['scheduleId'])
        self.assertEqual(schedule.probability, 60)
        self.assertEqual(payload['schedule']['probability'], 60)

    def test_schedules_create_api_salesman_creates_department_only_schedule(self):
        import json
        from reporting.models import Schedule

        department = self._create_department_only(self.user, '고객없는일정')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': department.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'location': '부서 방문',
                'notes': '담당자 등록 전 방문 일정',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        schedule = Schedule.objects.get(pk=payload['scheduleId'])
        self.assertEqual(schedule.user, self.user)
        self.assertIsNone(schedule.followup_id)
        self.assertEqual(schedule.department_id, department.id)
        self.assertEqual(schedule.location, '부서 방문')
        self.assertIsNone(schedule.probability)
        self.assertIsNone(payload['schedule']['probability'])
        self.assertEqual(payload['schedule']['customer'], '담당자 미등록')
        self.assertEqual(payload['schedule']['departmentId'], department.id)
        self.assertEqual(payload['schedule']['customerHref'], f'/accounts/{department.id}/')

    def test_schedules_create_api_requires_customer_when_department_has_contacts(self):
        import json

        followup = self._create_customer(self.user, '고객있는일정부서')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': followup.department_id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('고객을 선택하세요', response.json()['error'])

    def test_schedules_create_api_blocks_other_salesman_customer(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.coworker, '동료고객')
        coworker_department = self._create_department_only(self.coworker, '동료부서')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Schedule.objects.filter(followup=followup, user=self.user).exists())

        department_response = self.client.post(
            self.create_url,
            data=json.dumps({
                'departmentId': coworker_department.id,
                'activityType': 'customer_meeting',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )
        self.assertEqual(department_response.status_code, 403)
        self.assertFalse(Schedule.objects.filter(department=coworker_department, user=self.user).exists())

    def test_schedules_create_api_rejects_service_activity_type(self):
        import json
        from reporting.models import Schedule

        followup = self._create_customer(self.user, '서비스차단')
        self.client.force_login(self.user)

        response = self.client.post(
            self.create_url,
            data=json.dumps({
                'followupId': followup.id,
                'activityType': 'service',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(Schedule.objects.filter(followup=followup, activity_type='service').exists())

    def test_schedules_detail_api_returns_detail_and_edit_config(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import DeliveryItem, History, ScheduleFile

        schedule = self._create_schedule(self.user, '상세일정', activity_type='delivery')
        schedule.quote_extra_notes = '전체 견적 기타사항'
        schedule.save(update_fields=['quote_extra_notes'])
        History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            content='납품 보고',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='PCR Kit',
            quantity=2,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
            notes='PCR 적요',
            option_description='PCR 옵션 설명',
        )
        schedule_file = ScheduleFile.objects.create(
            schedule=schedule,
            file=SimpleUploadedFile('schedule-note.txt', b'schedule file note', content_type='text/plain'),
            original_filename='schedule-note.txt',
            file_size=18,
            uploaded_by=self.user,
        )
        self.addCleanup(schedule_file.file.delete, False)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['schedule']['id'], schedule.id)
        self.assertEqual(payload['schedule']['href'], f'/schedules/{schedule.id}/')
        self.assertEqual(payload['schedule']['customerHref'], f'/customers/{schedule.followup_id}/')
        self.assertEqual(payload['schedule']['quoteExtraNotes'], '전체 견적 기타사항')
        self.assertEqual(payload['schedule']['quoteGroupNotes'][0]['notes'], '전체 견적 기타사항')
        self.assertTrue(payload['edit']['canEdit'])
        self.assertEqual(payload['edit']['submitUrl'], reverse('reporting:schedules_update_api', args=[schedule.id]))
        self.assertEqual(payload['relatedNotes'][0]['id'], schedule.histories.first().id)
        self.assertEqual(payload['deliveryItems'][0]['itemName'], 'PCR Kit')
        self.assertEqual(payload['deliveryItems'][0]['discountRate'], 10.0)
        self.assertEqual(payload['deliveryItems'][0]['discountUnitPrice'], 90000)
        self.assertEqual(payload['deliveryItems'][0]['effectiveUnitPrice'], 90000)
        self.assertEqual(payload['deliveryItems'][0]['totalPrice'], 198000)
        self.assertEqual(payload['deliveryItems'][0]['notes'], 'PCR 적요')
        self.assertEqual(payload['deliveryItems'][0]['optionDescription'], 'PCR 옵션 설명')
        self.assertEqual(payload['links']['uploadFiles'], reverse('reporting:schedule_file_upload', args=[schedule.id]))
        self.assertEqual(payload['links']['updateDeliveryItems'], reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]))
        self.assertEqual(payload['links']['toggleTaxInvoice'], '')
        self.assertEqual(payload['links']['prepayments'], reverse('reporting:prepayment_api_list'))
        self.assertEqual(payload['links']['deleteSchedule'], reverse('reporting:schedule_delete', args=[schedule.id]))
        self.assertTrue(payload['taxInvoice']['applies'])
        self.assertEqual(payload['taxInvoice']['status'], 'pending')
        self.assertEqual(payload['taxInvoice']['totalCount'], 1)
        self.assertEqual(payload['taxInvoice']['issuedCount'], 0)
        self.assertEqual(payload['taxInvoice']['pendingCount'], 1)
        self.assertFalse(payload['taxInvoice']['canToggle'])
        self.assertEqual(payload['taxInvoice']['toggleUrl'], '')
        self.assertEqual(payload['schedule']['files'][0]['id'], schedule_file.id)
        self.assertEqual(payload['schedule']['files'][0]['deleteHref'], reverse('reporting:schedule_file_delete', args=[schedule_file.id]))
        document_types = [item['type'] for item in payload['documents']['items']]
        self.assertEqual(document_types, ['transaction_statement', 'delivery_note'])
        first_document = payload['documents']['items'][0]
        self.assertEqual(first_document['previewHref'], reverse('reporting:get_document_template_data', args=['transaction_statement', schedule.id]))
        self.assertEqual(
            first_document['formats'][0]['href'],
            reverse('reporting:generate_document_pdf_format', args=['transaction_statement', schedule.id, 'pdf']),
        )
        self.assertEqual(
            first_document['formats'][1]['href'],
            reverse('reporting:generate_document_pdf_format', args=['transaction_statement', schedule.id, 'xlsx']),
        )
        self.assertEqual(payload['documents']['templateManagerHref'], '/documents/')
        self.assertEqual(payload['documents']['djangoTemplateManagerHref'], reverse('reporting:document_template_list'))
        self.assertIn('거래명세서 PDF', payload['documents']['autoAttachLabel'])

    def test_schedules_detail_api_document_actions_match_activity_type(self):
        quote_schedule = self._create_schedule(self.user, '견적서류', activity_type='quote')
        meeting_schedule = self._create_schedule(self.user, '미팅서류없음', activity_type='customer_meeting')
        self.client.force_login(self.user)

        quote_response = self.client.get(reverse('reporting:schedules_detail_api', args=[quote_schedule.id]))
        meeting_response = self.client.get(reverse('reporting:schedules_detail_api', args=[meeting_schedule.id]))

        self.assertEqual(quote_response.status_code, 200)
        self.assertEqual(meeting_response.status_code, 200)
        quote_payload = quote_response.json()
        meeting_payload = meeting_response.json()
        self.assertTrue(quote_payload['documents']['canGenerate'])
        self.assertEqual([item['type'] for item in quote_payload['documents']['items']], ['quotation'])
        self.assertIn('견적서 PDF', quote_payload['documents']['autoAttachLabel'])
        self.assertEqual(
            quote_payload['documents']['items'][0]['formats'][1]['href'],
            reverse('reporting:generate_document_pdf_format', args=['quotation', quote_schedule.id, 'xlsx']),
        )
        self.assertFalse(meeting_payload['documents']['canGenerate'])
        self.assertEqual(meeting_payload['documents']['items'], [])
        self.assertEqual(meeting_payload['documents']['autoAttachLabel'], '')

    def test_schedules_detail_api_splits_quotation_documents_by_quote_group(self):
        from reporting.models import DeliveryItem

        quote_schedule = self._create_schedule(self.user, '복수견적서류', activity_type='quote')
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Trade In Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            quote_group='보상판매',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Repair Service',
            quantity=1,
            unit='EA',
            unit_price=50000,
            quote_group='수리',
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[quote_schedule.id]))

        self.assertEqual(response.status_code, 200)
        actions = response.json()['documents']['items']
        self.assertEqual([action['quoteGroup'] for action in actions], ['보상판매', '수리'])
        self.assertEqual([action['label'] for action in actions], ['보상판매 견적서', '수리 견적서'])
        self.assertIn('quote_group=', actions[0]['previewHref'])
        self.assertIn('quote_group=', actions[1]['formats'][0]['href'])
        self.assertEqual(actions[0]['itemCount'], 1)

    def test_schedules_detail_api_includes_quote_commercial_checks(self):
        import datetime
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        quote_schedule = self._create_schedule(
            self.user,
            '견적정합성',
            activity_type='quote',
            status='completed',
        )
        delivered_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Trade In PCR',
            quantity=2,
            unit='EA',
            unit_price=100000,
            quote_group='보상판매',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Repair Buffer',
            quantity=1,
            unit='EA',
            unit_price=50000,
            quote_group='수리',
        )
        delivery_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=quote_schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(11, 0),
            activity_type='delivery',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=delivery_schedule,
            source_quote_schedule=quote_schedule,
            source_quote_item=delivered_item,
            item_name='Trade In PCR',
            quantity=1,
            unit='EA',
            unit_price=100000,
            quote_group='보상판매',
        )
        log = DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=quote_schedule,
            user=self.user,
            transaction_number='Q-CHECK-001',
            output_format='pdf',
            file=SimpleUploadedFile('trade-in-quote.pdf', b'%PDF quote', content_type='application/pdf'),
            filename='trade-in-quote.pdf',
            file_size=10,
            quote_group='보상판매',
        )
        self.addCleanup(log.file.delete, False)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[quote_schedule.id]))

        self.assertEqual(response.status_code, 200)
        checks = response.json()['commercialChecks']
        self.assertTrue(checks['applies'])
        self.assertEqual(checks['kind'], 'quote')
        self.assertEqual(checks['summary']['quoteGroupCount'], 2)
        self.assertEqual(checks['summary']['quoteItemCount'], 2)
        self.assertEqual(checks['summary']['quoteAmount'], 275000)
        self.assertEqual(checks['summary']['deliveredAmount'], 110000)
        self.assertEqual(checks['summary']['remainingAmount'], 165000)
        groups = {group['quoteGroup']: group for group in checks['quoteGroups']}
        self.assertEqual(groups['보상판매']['registeredQuotationCount'], 1)
        self.assertEqual(groups['보상판매']['fulfillmentStatus'], 'partial')
        self.assertEqual(groups['보상판매']['deliveredAmount'], 110000)
        self.assertEqual(groups['보상판매']['remainingAmount'], 110000)
        self.assertEqual(groups['수리']['registeredQuotationCount'], 0)
        codes = [warning['code'] for warning in checks['warnings']]
        self.assertIn('missing_registered_quotation', codes)
        self.assertIn('missing_auto_attach_candidate', codes)
        self.assertIn('completed_quote_still_importable', codes)

    def test_schedules_detail_api_includes_registered_generated_documents(self):
        schedule = self._create_schedule(self.user, '등록서류목록', activity_type='delivery')
        log = DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='transaction_statement',
            schedule=schedule,
            user=self.user,
            transaction_number='TS-20260512-001',
            output_format='pdf',
            file=SimpleUploadedFile('statement.pdf', b'%PDF statement', content_type='application/pdf'),
            filename='statement.pdf',
            file_size=14,
        )
        self.addCleanup(log.file.delete, False)
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 200)
        documents = response.json()['documents']
        self.assertEqual(documents['registeredDocumentCount'], 1)
        self.assertEqual(documents['registeredQuotationCount'], 0)
        registered = documents['registeredDocuments'][0]
        self.assertEqual(registered['id'], log.id)
        self.assertEqual(registered['documentType'], 'transaction_statement')
        self.assertEqual(registered['documentTypeLabel'], '거래명세서')
        self.assertEqual(registered['downloadHref'], reverse('reporting:generated_document_download', args=[log.id]))
        self.assertEqual(registered['deleteHref'], reverse('reporting:generated_document_delete', args=[log.id]))
        self.assertTrue(registered['canDelete'])

    def test_schedules_detail_api_includes_delivery_commercial_checks(self):
        from reporting.models import DeliveryItem, History

        schedule = self._create_schedule(self.user, '납품정합성', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='PCR Kit',
            quantity=2,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
        )
        History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            delivery_items='PCR Kit 2EA',
            delivery_amount=1000,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 200)
        checks = response.json()['commercialChecks']
        self.assertTrue(checks['applies'])
        self.assertEqual(checks['kind'], 'delivery')
        self.assertEqual(checks['summary']['deliveryItemCount'], 1)
        self.assertEqual(checks['summary']['deliveryAmount'], 198000)
        self.assertFalse(checks['summary']['autoAttachReady'])
        self.assertEqual(checks['delivery']['autoAttachStatus'], 'missing')
        self.assertEqual(checks['delivery']['historyAmountMismatches'][0]['noteAmount'], 1000)
        self.assertEqual(checks['delivery']['historyAmountMismatches'][0]['itemAmount'], 198000)
        codes = [warning['code'] for warning in checks['warnings']]
        self.assertIn('missing_auto_attach_candidate', codes)
        self.assertIn('delivery_note_amount_mismatch', codes)

    def test_generated_document_delete_api_allows_owner_only(self):
        schedule = self._create_schedule(self.user, '등록서류삭제', activity_type='quote')
        log = DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=schedule,
            user=self.user,
            transaction_number='Q-20260512-001',
            output_format='pdf',
            file=SimpleUploadedFile('quote.pdf', b'%PDF quote', content_type='application/pdf'),
            filename='quote.pdf',
            file_size=10,
            quote_group='수리',
        )
        self.addCleanup(log.file.delete, False)
        delete_url = reverse('reporting:generated_document_delete', args=[log.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(delete_url)
        self.assertEqual(manager_response.status_code, 403)
        self.assertTrue(DocumentGenerationLog.objects.filter(pk=log.id).exists())

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(delete_url)
        self.assertEqual(coworker_response.status_code, 403)
        self.assertTrue(DocumentGenerationLog.objects.filter(pk=log.id).exists())

        self.client.force_login(self.user)
        response = self.client.post(delete_url)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertFalse(DocumentGenerationLog.objects.filter(pk=log.id).exists())

    def test_schedules_detail_api_manager_read_only_and_other_company_blocked(self):
        from reporting.models import DeliveryItem

        schedule = self._create_schedule(self.user, '읽기전용', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Read Only Kit',
            quantity=1,
            unit='EA',
            unit_price=1000,
        )
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:schedules_detail_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()['edit']['canEdit'])
        self.assertEqual(response.json()['links']['updateDeliveryItems'], '')
        self.assertEqual(response.json()['links']['deleteSchedule'], '')
        self.assertEqual(response.json()['links']['toggleTaxInvoice'], '')
        self.assertTrue(response.json()['taxInvoice']['applies'])
        self.assertFalse(response.json()['taxInvoice']['canToggle'])

        self.client.force_login(self.other_user)
        denied = self.client.get(reverse('reporting:schedules_detail_api', args=[schedule.id]))
        self.assertEqual(denied.status_code, 403)

    def test_schedule_tax_invoice_toggle_is_removed_in_favor_of_receivables_menu(self):
        from reporting.models import DeliveryItem, History

        schedule = self._create_schedule(self.user, '세금계산서토글', activity_type='delivery')
        first = DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Tax Kit A',
            quantity=1,
            unit='EA',
            unit_price=1000,
            tax_invoice_issued=False,
        )
        second = DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Tax Kit B',
            quantity=1,
            unit='EA',
            unit_price=1000,
            tax_invoice_issued=True,
        )
        history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            tax_invoice_issued=False,
        )
        toggle_url = reverse('reporting:toggle_schedule_delivery_tax_invoice', args=[schedule.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(toggle_url)
        self.assertEqual(manager_response.status_code, 410)
        self.assertIn('/receivables/', manager_response.json()['redirect'])

        self.client.force_login(self.user)
        response = self.client.post(toggle_url)
        self.assertEqual(response.status_code, 410)
        self.assertFalse(response.json()['success'])
        first.refresh_from_db()
        second.refresh_from_db()
        history.refresh_from_db()
        self.assertFalse(first.tax_invoice_issued)
        self.assertTrue(second.tax_invoice_issued)
        self.assertFalse(history.tax_invoice_issued)

    def test_schedule_delete_ajax_allows_owner_and_removes_related_history(self):
        from reporting.models import History, Schedule

        schedule = self._create_schedule(self.user, '삭제일정', activity_type='delivery')
        related_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            content='삭제될 납품 기록',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedule_delete', args=[schedule.id]),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertFalse(Schedule.objects.filter(pk=schedule.id).exists())
        self.assertFalse(History.objects.filter(pk=related_history.id).exists())

    def test_schedule_delete_ajax_blocks_non_owner(self):
        from reporting.models import Schedule

        schedule = self._create_schedule(self.user, '타인삭제차단')
        self.client.force_login(self.coworker)

        response = self.client.post(
            reverse('reporting:schedule_delete', args=[schedule.id]),
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])
        self.assertTrue(Schedule.objects.filter(pk=schedule.id).exists())

    def test_product_api_list_returns_accessible_product_master_data(self):
        from reporting.models import Product

        global_product = Product.objects.create(
            product_code='MASTER-GLOBAL-PCR',
            unit='BOX',
            specification='96 reactions',
            standard_price=1000,
            created_by=None,
        )
        own_product = Product.objects.create(
            product_code='MASTER-OWN-PCR',
            unit='EA',
            standard_price=2000,
            created_by=self.user,
        )
        coworker_product = Product.objects.create(
            product_code='MASTER-COWORKER-PCR',
            unit='SET',
            standard_price=3000,
            created_by=self.coworker,
        )
        other_product = Product.objects.create(
            product_code='MASTER-OTHER-PCR',
            unit='EA',
            standard_price=4000,
            created_by=self.other_user,
        )
        inactive_product = Product.objects.create(
            product_code='MASTER-INACTIVE-PCR',
            unit='EA',
            standard_price=5000,
            is_active=False,
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:product_api_list'), {'search': 'MASTER-'})

        self.assertEqual(response.status_code, 200)
        products = response.json()['products']
        product_codes = {product['product_code'] for product in products}
        self.assertIn(global_product.product_code, product_codes)
        self.assertIn(own_product.product_code, product_codes)
        self.assertIn(coworker_product.product_code, product_codes)
        self.assertNotIn(other_product.product_code, product_codes)
        self.assertNotIn(inactive_product.product_code, product_codes)
        global_payload = next(product for product in products if product['product_code'] == global_product.product_code)
        self.assertEqual(global_payload['unit'], 'BOX')
        self.assertEqual(global_payload['specification'], '96 reactions')
        self.assertEqual(global_payload['current_price'], 1000.0)

    def test_schedules_update_api_updates_owned_schedule(self):
        import json

        schedule = self._create_schedule(self.user, '수정전')
        target_followup = self._create_customer(self.coworker, '수정후')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': target_followup.id,
                'activityType': 'delivery',
                'status': 'completed',
                'visitDate': '2026-05-11',
                'visitTime': '15:45',
                'location': '수정 회의실',
                'notes': '일정 수정 메모',
                'expectedRevenue': '2500000',
                'probability': '82',
                'expectedCloseDate': '2026-06-01',
                'purchaseConfirmed': True,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        schedule.refresh_from_db()
        self.assertEqual(schedule.followup, target_followup)
        self.assertEqual(schedule.activity_type, 'delivery')
        self.assertEqual(schedule.status, 'completed')
        self.assertEqual(schedule.visit_date.isoformat(), '2026-05-11')
        self.assertEqual(schedule.visit_time.strftime('%H:%M'), '15:45')
        self.assertEqual(schedule.location, '수정 회의실')
        self.assertEqual(schedule.notes, '일정 수정 메모')
        self.assertEqual(int(schedule.expected_revenue), 2500000)
        self.assertEqual(schedule.probability, 80)
        self.assertEqual(payload['schedule']['probability'], 80)
        self.assertEqual(schedule.expected_close_date.isoformat(), '2026-06-01')
        self.assertTrue(schedule.purchase_confirmed)
        self.assertEqual(payload['schedule']['id'], schedule.id)
        self.assertEqual(payload['message'], '일정을 수정했습니다.')

    def test_schedules_update_api_delivery_completed_forces_pipeline_card_to_won(self):
        import json

        schedule = self._create_schedule(self.user, '납품완료수주', activity_type='quote')
        schedule.followup.pipeline_stage = 'quote'
        schedule.followup.pipeline_manually_set = True
        schedule.followup.save(update_fields=['pipeline_stage', 'pipeline_manually_set'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'delivery',
                'status': 'completed',
                'visitDate': '2026-05-11',
                'visitTime': '15:45',
                'location': '납품 회의실',
                'notes': '납품 완료',
                'expectedRevenue': '2500000',
                'probability': '100',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        schedule.refresh_from_db()
        schedule.followup.refresh_from_db()
        self.assertEqual(schedule.activity_type, 'delivery')
        self.assertEqual(schedule.status, 'completed')
        self.assertEqual(schedule.followup.pipeline_stage, 'won')
        self.assertFalse(schedule.followup.pipeline_manually_set)

        pipeline_response = self.client.get(reverse('reporting:pipeline_command_center_api'))
        self.assertEqual(pipeline_response.status_code, 200)
        deal = next(item for item in pipeline_response.json()['deals'] if item['id'] == schedule.followup_id)
        self.assertEqual(deal['stage'], 'won')
        self.assertEqual(deal['stageLabel'], '수주')

    def test_schedules_update_api_requires_quote_probability(self):
        import json

        schedule = self._create_schedule(self.user, '견적수정확률필수', activity_type='quote')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'quote',
                'status': 'scheduled',
                'visitDate': '2026-05-11',
                'visitTime': '15:45',
                'expectedRevenue': '2500000',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('견적 성공 확률은 필수입니다.', response.json()['error'])

    def test_schedules_update_api_quote_advances_pipeline_card(self):
        import json

        schedule = self._create_schedule(self.user, '견적수정카드', activity_type='customer_meeting')
        schedule.followup.pipeline_stage = 'contact'
        schedule.followup.save(update_fields=['pipeline_stage'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'quote',
                'status': 'scheduled',
                'visitDate': '2026-05-12',
                'visitTime': '16:15',
                'expectedRevenue': '2500000',
                'probability': '82',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        schedule.refresh_from_db()
        schedule.followup.refresh_from_db()
        self.assertEqual(schedule.activity_type, 'quote')
        self.assertEqual(schedule.probability, 80)
        self.assertEqual(schedule.followup.pipeline_stage, 'quote')

        pipeline_response = self.client.get(reverse('reporting:pipeline_command_center_api'))
        self.assertEqual(pipeline_response.status_code, 200)
        deal = next(item for item in pipeline_response.json()['deals'] if item['id'] == schedule.followup_id)
        self.assertEqual(deal['stage'], 'quote')
        self.assertEqual(deal['value'], 2500000)
        self.assertEqual(deal['probability'], 80)

    def test_schedules_update_api_quote_cancel_moves_pipeline_card_to_lost(self):
        import json

        schedule = self._create_schedule(self.user, '견적취소실주', activity_type='quote')
        schedule.followup.pipeline_stage = 'quote'
        schedule.followup.save(update_fields=['pipeline_stage'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'quote',
                'status': 'cancelled',
                'visitDate': '2026-05-12',
                'visitTime': '16:15',
                'expectedRevenue': '2500000',
                'probability': '75',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        schedule.refresh_from_db()
        schedule.followup.refresh_from_db()
        self.assertEqual(schedule.status, 'cancelled')
        self.assertEqual(schedule.followup.pipeline_stage, 'lost')

        pipeline_response = self.client.get(reverse('reporting:pipeline_command_center_api'))
        self.assertEqual(pipeline_response.status_code, 200)
        deal = next(item for item in pipeline_response.json()['deals'] if item['id'] == schedule.followup_id)
        self.assertEqual(deal['stage'], 'lost')
        self.assertEqual(deal['stageLabel'], '실주')
        self.assertEqual(deal['value'], 2500000)

    def test_schedule_status_update_api_quote_cancel_moves_pipeline_card_to_lost(self):
        schedule = self._create_schedule(self.user, '상태버튼견적취소', activity_type='quote')
        schedule.followup.pipeline_stage = 'quote'
        schedule.followup.save(update_fields=['pipeline_stage'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedule_status_update', args=[schedule.id]),
            data={'status': 'cancelled'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        schedule.refresh_from_db()
        schedule.followup.refresh_from_db()
        self.assertEqual(schedule.status, 'cancelled')
        self.assertEqual(schedule.followup.pipeline_stage, 'lost')

    def test_schedule_status_update_api_delivery_completed_moves_pipeline_card_to_won(self):
        schedule = self._create_schedule(self.user, '상태버튼납품완료', activity_type='delivery')
        schedule.followup.pipeline_stage = 'quote'
        schedule.followup.pipeline_manually_set = True
        schedule.followup.save(update_fields=['pipeline_stage', 'pipeline_manually_set'])
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedule_status_update', args=[schedule.id]),
            data={'status': 'completed'},
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        schedule.refresh_from_db()
        schedule.followup.refresh_from_db()
        self.assertEqual(schedule.status, 'completed')
        self.assertEqual(schedule.followup.pipeline_stage, 'won')
        self.assertFalse(schedule.followup.pipeline_manually_set)

    def test_prepayment_api_list_includes_same_department_and_existing_usage(self):
        from django.utils import timezone
        from reporting.models import FollowUp, Prepayment, PrepaymentUsage

        schedule = self._create_schedule(self.user, '선결제조회', activity_type='delivery')
        same_department_customer = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            customer_name='같은부서 고객',
            manager='같은부서 담당',
            company=schedule.followup.company,
            department=schedule.followup.department,
        )
        other_department_customer = self._create_customer(self.user, '다른부서')
        active_prepayment = Prepayment.objects.create(
            customer=same_department_customer,
            company=same_department_customer.company,
            amount=100000,
            balance=80000,
            payment_date=timezone.localdate(),
            payer_name='같은부서입금',
            created_by=self.user,
        )
        selected_depleted_prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=40000,
            balance=0,
            payment_date=timezone.localdate(),
            payer_name='기존차감',
            status='depleted',
            created_by=self.user,
        )
        PrepaymentUsage.objects.create(
            prepayment=selected_depleted_prepayment,
            schedule=schedule,
            product_name='기존 납품',
            quantity=1,
            amount=40000,
            remaining_balance=0,
        )
        Prepayment.objects.create(
            customer=other_department_customer,
            company=other_department_customer.company,
            amount=50000,
            balance=50000,
            payment_date=timezone.localdate(),
            payer_name='다른부서입금',
            created_by=self.user,
        )
        schedule.use_prepayment = True
        schedule.prepayment = selected_depleted_prepayment
        schedule.prepayment_amount = 40000
        schedule.save(update_fields=['use_prepayment', 'prepayment', 'prepayment_amount'])
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:prepayment_api_list'), {
            'customer_id': schedule.followup_id,
            'schedule_id': schedule.id,
        })

        self.assertEqual(response.status_code, 200)
        prepayments = response.json()['prepayments']
        ids = {item['id'] for item in prepayments}
        self.assertIn(active_prepayment.id, ids)
        self.assertIn(selected_depleted_prepayment.id, ids)
        selected_payload = next(item for item in prepayments if item['id'] == selected_depleted_prepayment.id)
        self.assertEqual(selected_payload['balance'], 0)
        self.assertEqual(selected_payload['selectedAmount'], 40000)
        self.assertEqual(selected_payload['availableBalance'], 40000)

    def test_schedules_update_api_applies_and_restores_prepayments(self):
        import json
        from django.utils import timezone
        from reporting.models import Prepayment, PrepaymentUsage

        schedule = self._create_schedule(self.user, '선결제수정', activity_type='delivery')
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=100000,
            balance=100000,
            payment_date=timezone.localdate(),
            payer_name='선결제고객',
            created_by=self.user,
        )
        update_url = reverse('reporting:schedules_update_api', args=[schedule.id])
        self.client.force_login(self.user)

        apply_response = self.client.post(
            update_url,
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'delivery',
                'status': 'scheduled',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'usePrepayment': True,
                'prepayments': [
                    {'id': prepayment.id, 'amount': '60000'},
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(apply_response.status_code, 200)
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 40000)
        self.assertTrue(schedule.use_prepayment)
        self.assertEqual(schedule.prepayment, prepayment)
        self.assertEqual(int(schedule.prepayment_amount), 60000)
        usage = PrepaymentUsage.objects.get(schedule=schedule)
        self.assertEqual(int(usage.amount), 60000)
        self.assertEqual(apply_response.json()['schedule']['prepaymentUsages'][0]['amount'], 60000)

        restore_response = self.client.post(
            update_url,
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'delivery',
                'status': 'scheduled',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'usePrepayment': False,
                'prepayments': [],
            }),
            content_type='application/json',
        )

        self.assertEqual(restore_response.status_code, 200)
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 100000)
        self.assertFalse(schedule.use_prepayment)
        self.assertIsNone(schedule.prepayment)
        self.assertEqual(int(schedule.prepayment_amount), 0)
        self.assertFalse(PrepaymentUsage.objects.filter(schedule=schedule).exists())

    def test_schedules_update_api_blocks_over_balance_prepayment(self):
        import json
        from django.utils import timezone
        from reporting.models import Prepayment, PrepaymentUsage

        schedule = self._create_schedule(self.user, '선결제잔액차단', activity_type='delivery')
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=1000,
            balance=1000,
            payment_date=timezone.localdate(),
            payer_name='잔액부족',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_update_api', args=[schedule.id]),
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'delivery',
                'status': 'scheduled',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
                'usePrepayment': True,
                'prepayments': [
                    {'id': prepayment.id, 'amount': '2000'},
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('잔액이 부족', response.json()['error'])
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 1000)
        self.assertFalse(schedule.use_prepayment)
        self.assertFalse(PrepaymentUsage.objects.filter(schedule=schedule).exists())

    def test_schedules_update_api_blocks_manager_and_other_company_customer(self):
        import json

        schedule = self._create_schedule(self.user, '수정차단')
        other_followup = self._create_customer(self.other_user, '타사고객')
        update_url = reverse('reporting:schedules_update_api', args=[schedule.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(
            update_url,
            data=json.dumps({
                'followupId': schedule.followup_id,
                'activityType': 'customer_meeting',
                'status': 'scheduled',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.user)
        other_company_response = self.client.post(
            update_url,
            data=json.dumps({
                'followupId': other_followup.id,
                'activityType': 'customer_meeting',
                'status': 'scheduled',
                'visitDate': '2026-05-10',
                'visitTime': '10:30',
            }),
            content_type='application/json',
        )
        self.assertEqual(other_company_response.status_code, 403)
        schedule.refresh_from_db()
        self.assertNotEqual(schedule.followup, other_followup)

    def test_schedule_delivery_items_update_api_updates_owned_items_and_history(self):
        import json
        from reporting.models import DeliveryItem, History

        schedule = self._create_schedule(self.user, '납품품목', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Old Kit',
            quantity=1,
            unit='EA',
            unit_price=1000,
        )
        history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            delivery_items='old',
            delivery_amount=1000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'quoteGroupNotes': {
                    '보상판매': '보상판매 기타사항',
                    '수리': '수리 기타사항',
                },
                'items': [
                    {
                        'itemName': 'PCR Kit',
                        'quantity': 2,
                        'unit': 'EA',
                        'unitPrice': '100000',
                        'discountRate': '10',
                        'taxInvoiceIssued': True,
                        'quoteGroup': '보상판매',
                        'notes': 'PCR 적요',
                        'optionDescription': 'PCR 옵션 설명',
                    },
                    {
                        'itemName': 'Buffer',
                        'quantity': 3,
                        'unit': 'BOX',
                        'unitPrice': '',
                        'taxInvoiceIssued': False,
                        'quoteGroup': '수리',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['message'], '납품 품목을 저장했습니다.')
        items = list(DeliveryItem.objects.filter(schedule=schedule).order_by('id'))
        self.assertEqual([item.item_name for item in items], ['PCR Kit', 'Buffer'])
        self.assertEqual(items[0].quantity, 2)
        self.assertEqual(items[0].unit, 'EA')
        self.assertEqual(int(items[0].unit_price), 100000)
        self.assertEqual(float(items[0].discount_rate), 10.0)
        self.assertEqual(int(items[0].discount_unit_price), 90000)
        self.assertEqual(int(items[0].get_effective_unit_price()), 90000)
        self.assertEqual(int(items[0].total_price), 198000)
        self.assertFalse(items[0].tax_invoice_issued)
        self.assertEqual(items[0].quote_group, '보상판매')
        self.assertEqual(items[0].notes, 'PCR 적요')
        self.assertEqual(items[0].option_description, 'PCR 옵션 설명')
        self.assertIsNone(items[1].unit_price)
        self.assertEqual(items[1].quote_group, '수리')
        schedule.refresh_from_db()
        self.assertEqual(schedule.quote_extra_notes, '')
        self.assertEqual(
            {
                note.quote_group: note.notes
                for note in ScheduleQuoteGroupNote.objects.filter(schedule=schedule)
            },
            {'보상판매': '보상판매 기타사항', '수리': '수리 기타사항'},
        )
        history.refresh_from_db()
        self.assertIn('PCR Kit', history.delivery_items)
        self.assertIn('Buffer', history.delivery_items)
        self.assertEqual(int(history.delivery_amount), 198000)
        self.assertEqual(payload['deliveryItems'][0]['itemName'], 'PCR Kit')
        self.assertEqual(payload['deliveryItems'][0]['discountRate'], 10.0)
        self.assertEqual(payload['deliveryItems'][0]['discountUnitPrice'], 90000)
        self.assertEqual(payload['deliveryItems'][0]['effectiveUnitPrice'], 90000)
        self.assertEqual(payload['deliveryItems'][0]['totalPrice'], 198000)
        self.assertEqual(payload['deliveryItems'][0]['quoteGroup'], '보상판매')
        self.assertEqual(payload['deliveryItems'][0]['notes'], 'PCR 적요')
        self.assertEqual(payload['deliveryItems'][0]['optionDescription'], 'PCR 옵션 설명')
        self.assertIsNone(payload['deliveryItems'][1]['unitPrice'])
        self.assertIsNone(payload['deliveryItems'][1]['discountUnitPrice'])
        self.assertEqual(
            {note['quoteGroup']: note['notes'] for note in payload['schedule']['quoteGroupNotes']},
            {'보상판매': '보상판매 기타사항', '수리': '수리 기타사항'},
        )

    def test_schedule_delivery_items_update_api_preserves_existing_receivable_status(self):
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem

        schedule = self._create_schedule(self.user, '외상상태보존', activity_type='delivery')
        item = DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Receivable Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            tax_invoice_issued=True,
            card_payment_received=True,
            receivable_settled=True,
            receivable_settled_at=timezone.now(),
            receivable_settled_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'id': item.id,
                        'itemName': 'Receivable Kit Updated',
                        'quantity': 2,
                        'unit': 'EA',
                        'unitPrice': '100000',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        updated = DeliveryItem.objects.get(schedule=schedule)
        self.assertEqual(updated.item_name, 'Receivable Kit Updated')
        self.assertTrue(updated.tax_invoice_issued)
        self.assertTrue(updated.card_payment_received)
        self.assertTrue(updated.receivable_settled)
        self.assertEqual(updated.receivable_settled_by, self.user)

    def test_schedule_delivery_items_update_api_applies_reapplies_and_restores_prepayment(self):
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Prepayment, PrepaymentUsage

        schedule = self._create_schedule(self.user, '납품품목선결제', activity_type='delivery')
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=100000,
            balance=100000,
            payment_date=timezone.localdate(),
            payer_name='선결제입금자',
            created_by=self.user,
        )
        update_url = reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id])
        self.client.force_login(self.user)

        apply_response = self.client.post(
            update_url,
            data=json.dumps({
                'usePrepayment': True,
                'prepayments': [{'id': prepayment.id, 'amount': '60000'}],
                'items': [
                    {
                        'itemName': 'Prepaid Kit',
                        'quantity': 2,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(apply_response.status_code, 200)
        self.assertIn('선결제를 차감', apply_response.json()['message'])
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 40000)
        self.assertTrue(schedule.use_prepayment)
        self.assertEqual(schedule.prepayment, prepayment)
        self.assertEqual(int(schedule.prepayment_amount), 60000)
        usage = PrepaymentUsage.objects.get(schedule=schedule)
        self.assertEqual(int(usage.amount), 60000)
        self.assertEqual(usage.schedule_item.item_name, 'Prepaid Kit')
        self.assertEqual(apply_response.json()['schedule']['prepaymentUsages'][0]['amount'], 60000)

        reapply_response = self.client.post(
            update_url,
            data=json.dumps({
                'usePrepayment': True,
                'prepayments': [{'id': prepayment.id, 'amount': '30000'}],
                'items': [
                    {
                        'itemName': 'Prepaid Kit Updated',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(reapply_response.status_code, 200)
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 70000)
        self.assertEqual(int(schedule.prepayment_amount), 30000)
        self.assertEqual(PrepaymentUsage.objects.filter(schedule=schedule).count(), 1)
        self.assertEqual(int(PrepaymentUsage.objects.get(schedule=schedule).amount), 30000)

        restore_response = self.client.post(
            update_url,
            data=json.dumps({
                'usePrepayment': False,
                'prepayments': [],
                'items': [
                    {
                        'itemName': 'Prepaid Kit Updated',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(restore_response.status_code, 200)
        self.assertIn('선결제 차감을 해제', restore_response.json()['message'])
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 100000)
        self.assertFalse(schedule.use_prepayment)
        self.assertIsNone(schedule.prepayment)
        self.assertEqual(int(schedule.prepayment_amount), 0)
        self.assertFalse(PrepaymentUsage.objects.filter(schedule=schedule).exists())

    def test_schedule_delivery_items_update_api_locks_only_prepayment_rows_when_applying_prepayment(self):
        import json
        from unittest.mock import patch
        from django.db.models.query import QuerySet
        from django.utils import timezone
        from reporting.models import Prepayment

        schedule = self._create_schedule(self.user, '납품선결제잠금', activity_type='delivery')
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=100000,
            balance=100000,
            payment_date=timezone.localdate(),
            payer_name='잠금입금자',
            created_by=self.user,
        )
        update_url = reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id])
        prepayment_lock_calls = []
        original_select_for_update = QuerySet.select_for_update

        def spy_select_for_update(queryset, *args, **kwargs):
            if queryset.model is Prepayment:
                prepayment_lock_calls.append(dict(kwargs))
            return original_select_for_update(queryset, *args, **kwargs)

        self.client.force_login(self.user)
        with patch.object(QuerySet, 'select_for_update', spy_select_for_update):
            response = self.client.post(
                update_url,
                data=json.dumps({
                    'usePrepayment': True,
                    'prepayments': [{'id': prepayment.id, 'amount': '60000'}],
                    'items': [
                        {
                            'itemName': 'Lock Kit',
                            'quantity': 2,
                            'unit': 'EA',
                            'unitPrice': '50000',
                        },
                    ],
                }),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 200, response.content.decode())
        self.assertTrue(response.json()['success'])
        self.assertTrue(
            any(call.get('of') == ('self',) for call in prepayment_lock_calls),
            prepayment_lock_calls,
        )

    def test_schedule_delivery_items_update_api_treats_zero_discount_unit_price_without_rate_as_blank(self):
        import json
        from reporting.models import DeliveryItem

        schedule = self._create_schedule(self.user, '납품할인단가0', activity_type='delivery')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'itemName': 'SO825.0002',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '379950',
                        'discountRate': '',
                        'discountUnitPrice': '0',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        item = DeliveryItem.objects.get(schedule=schedule)
        self.assertIsNone(item.discount_unit_price)
        self.assertEqual(float(item.discount_rate), 0.0)
        self.assertEqual(int(item.get_effective_unit_price()), 379950)
        self.assertEqual(int(item.total_price), 417945)
        self.assertIsNone(payload['deliveryItems'][0]['discountUnitPrice'])
        self.assertEqual(payload['deliveryItems'][0]['effectiveUnitPrice'], 379950)
        self.assertEqual(payload['deliveryItems'][0]['totalPrice'], 417945)

    def test_schedule_delivery_items_update_api_blocks_prepayment_above_delivery_total(self):
        import json
        from django.utils import timezone
        from reporting.models import Prepayment

        schedule = self._create_schedule(self.user, '납품선결제상한서버', activity_type='delivery')
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=1000000,
            balance=1000000,
            payment_date=timezone.localdate(),
            payer_name='상한입금자',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'usePrepayment': True,
                'prepayments': [{'id': prepayment.id, 'amount': '60000'}],
                'items': [
                    {
                        'itemName': 'Limit Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('납품 품목 합계', response.json()['error'])
        prepayment.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 1000000)

    def test_schedule_delivery_items_update_api_blocks_over_balance_prepayment_without_saving_items(self):
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Prepayment, PrepaymentUsage

        schedule = self._create_schedule(self.user, '납품품목선결제잔액차단', activity_type='delivery')
        original_item = DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Original Kit',
            quantity=1,
            unit='EA',
            unit_price=1000,
        )
        prepayment = Prepayment.objects.create(
            customer=schedule.followup,
            company=schedule.followup.company,
            amount=1000,
            balance=1000,
            payment_date=timezone.localdate(),
            payer_name='잔액부족입금자',
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'usePrepayment': True,
                'prepayments': [{'id': prepayment.id, 'amount': '2000'}],
                'items': [
                    {
                        'itemName': 'Blocked Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '2000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('잔액이 부족', response.json()['error'])
        prepayment.refresh_from_db()
        schedule.refresh_from_db()
        self.assertEqual(int(prepayment.balance), 1000)
        self.assertFalse(schedule.use_prepayment)
        self.assertFalse(PrepaymentUsage.objects.filter(schedule=schedule).exists())
        items = list(DeliveryItem.objects.filter(schedule=schedule))
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].id, original_item.id)
        self.assertEqual(items[0].item_name, 'Original Kit')

    def test_schedule_delivery_items_update_api_marks_imported_quote_completed(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '견적불러오기납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Quoted PCR Kit',
            quantity=2,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'sourceQuoteScheduleIds': [quote_schedule.id],
                'items': [
                    {
                        'sourceQuoteScheduleId': quote_schedule.id,
                        'itemName': 'Quoted PCR Kit',
                        'quantity': 2,
                        'unit': 'EA',
                        'unitPrice': '50000',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['completedQuoteScheduleIds'], [quote_schedule.id])
        self.assertIn('원본 견적 일정', payload['message'])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'completed')
        self.assertEqual(DeliveryItem.objects.get(schedule=schedule).item_name, 'Quoted PCR Kit')

    def test_schedule_delivery_items_update_api_marks_checked_quote_completed_without_imported_items(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '견적체크수동납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Checked Quote Kit',
            quantity=5,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'checkedQuoteScheduleIds': [quote_schedule.id],
                'items': [
                    {
                        'itemName': 'Manual Delivered Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '70000',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['completedQuoteScheduleIds'], [quote_schedule.id])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'completed')
        delivery_item = DeliveryItem.objects.get(schedule=schedule)
        self.assertEqual(delivery_item.item_name, 'Manual Delivered Kit')
        self.assertIsNone(delivery_item.source_quote_schedule_id)
        self.assertIsNone(delivery_item.source_quote_item_id)

    def test_schedule_delivery_items_update_api_marks_checked_department_only_quote_completed(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        department = self._create_department_only(self.user, '부서전용견적납품')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=None,
            department=department,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(11, 0),
            activity_type='delivery',
            status='scheduled',
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=None,
            department=department,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Department Only Quote Kit',
            quantity=3,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'checkedQuoteScheduleIds': [quote_schedule.id],
                'items': [
                    {
                        'itemName': 'Manual Department Delivery Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '70000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['completedQuoteScheduleIds'], [quote_schedule.id])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'completed')

    def test_schedule_delivery_items_update_api_rejects_checked_quote_from_different_department(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        delivery_department = self._create_department_only(self.user, '납품부서')
        quote_department = self._create_department_only(self.user, '다른견적부서')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=None,
            department=delivery_department,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(11, 0),
            activity_type='delivery',
            status='scheduled',
        )
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=None,
            department=quote_department,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Wrong Department Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'checkedQuoteScheduleIds': [quote_schedule.id],
                'items': [
                    {
                        'itemName': 'Manual Delivery Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '70000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertIn('다른 고객/부서', response.json()['error'])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'scheduled')

    def test_schedule_delivery_items_update_api_keeps_partial_imported_quote_scheduled(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, History, Schedule

        schedule = self._create_schedule(self.user, '부분견적불러오기납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='completed',
        )
        sold_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Thirty Thousand Kit',
            quantity=1,
            unit='EA',
            unit_price=30000,
            quote_group='보상판매',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Unsold Kit',
            quantity=1,
            unit='EA',
            unit_price=70000,
            quote_group='수리',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'sourceQuoteScheduleId': quote_schedule.id,
                        'sourceQuoteItemId': sold_item.id,
                        'itemName': 'Thirty Thousand Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '30000',
                        'quoteGroup': '보상판매',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['completedQuoteScheduleIds'], [])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'scheduled')
        delivery_item = DeliveryItem.objects.get(schedule=schedule)
        self.assertEqual(delivery_item.source_quote_schedule_id, quote_schedule.id)
        self.assertEqual(delivery_item.source_quote_item_id, sold_item.id)
        history = History.objects.get(schedule=schedule, action_type='delivery_schedule')
        self.assertIn('Thirty Thousand Kit', history.delivery_items)
        self.assertNotIn('Unsold Kit', history.delivery_items)
        self.assertEqual(int(history.delivery_amount), 33000)

    def test_schedule_delivery_items_update_api_rejects_over_imported_quote_quantity(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '초과견적불러오기납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Limited Quote Kit',
            quantity=2,
            unit='EA',
            unit_price=30000,
        )
        previous_delivery = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(11, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=previous_delivery,
            source_quote_schedule=quote_schedule,
            source_quote_item=quote_item,
            item_name='Limited Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=30000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'sourceQuoteScheduleId': quote_schedule.id,
                        'sourceQuoteItemId': quote_item.id,
                        'itemName': 'Limited Quote Kit',
                        'quantity': 2,
                        'unit': 'EA',
                        'unitPrice': '30000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('남은 견적 수량은 1EA', response.json()['error'])
        self.assertFalse(DeliveryItem.objects.filter(schedule=schedule).exists())

    def test_schedule_delivery_items_update_api_rejects_duplicate_source_quote_item_rows(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '중복견적품목납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Duplicate Guard Kit',
            quantity=2,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'sourceQuoteScheduleId': quote_schedule.id,
                        'sourceQuoteItemId': quote_item.id,
                        'itemName': 'Duplicate Guard Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                    {
                        'sourceQuoteScheduleId': quote_schedule.id,
                        'sourceQuoteItemId': quote_item.id,
                        'itemName': 'Duplicate Guard Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('견적 품목이 중복되었습니다', response.json()['error'])
        self.assertFalse(DeliveryItem.objects.filter(schedule=schedule).exists())

    def test_schedule_delivery_items_update_api_reopens_quote_when_import_link_removed(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '견적연결해제납품', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='completed',
        )
        quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Linked Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=40000,
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            source_quote_schedule=quote_schedule,
            source_quote_item=quote_item,
            item_name='Linked Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=40000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'itemName': 'Manual Replacement Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '10000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'scheduled')
        delivery_item = DeliveryItem.objects.get(schedule=schedule)
        self.assertEqual(delivery_item.item_name, 'Manual Replacement Kit')
        self.assertIsNone(delivery_item.source_quote_schedule_id)
        self.assertIsNone(delivery_item.source_quote_item_id)

    def test_schedule_delivery_items_update_api_collects_existing_source_quotes_without_distinct_lock(self):
        import datetime
        import json
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '견적원본중복잠금회피', activity_type='delivery')
        quote_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='completed',
        )
        quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Existing Linked Kit',
            quantity=2,
            unit='EA',
            unit_price=40000,
        )
        for index in range(2):
            DeliveryItem.objects.create(
                schedule=schedule,
                source_quote_schedule=quote_schedule,
                source_quote_item=quote_item if index == 0 else None,
                item_name=f'Existing Linked Kit {index}',
                quantity=1,
                unit='EA',
                unit_price=40000,
            )
        self.client.force_login(self.user)

        with CaptureQueriesContext(connection) as captured:
            response = self.client.post(
                reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
                data=json.dumps({
                    'items': [
                        {
                            'itemName': 'Manual Replacement Kit',
                            'quantity': 1,
                            'unit': 'EA',
                            'unitPrice': '10000',
                        },
                    ],
                }),
                content_type='application/json',
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        source_quote_queries = [
            query['sql'].upper()
            for query in captured.captured_queries
            if 'SOURCE_QUOTE_SCHEDULE_ID' in query['sql'].upper()
        ]
        self.assertTrue(source_quote_queries)
        self.assertFalse(any('DISTINCT' in query for query in source_quote_queries))
        quote_schedule.refresh_from_db()
        self.assertEqual(quote_schedule.status, 'scheduled')

    def test_completed_quote_items_do_not_increment_product_sold_count(self):
        from reporting.models import DeliveryItem, Product

        quote_schedule = self._create_schedule(self.user, '완료견적판매수량제외', activity_type='quote', status='completed')
        product = Product.objects.create(
            product_code='QUOTE-NOT-SOLD',
            unit='EA',
            standard_price=30000,
            created_by=self.user,
        )

        DeliveryItem.objects.create(
            schedule=quote_schedule,
            product=product,
            item_name='QUOTE-NOT-SOLD',
            quantity=2,
            unit='EA',
            unit_price=30000,
        )

        product.refresh_from_db()
        self.assertEqual(product.total_sold, 0)

    def test_notes_detail_uses_actual_delivery_schedule_items_over_stale_history_text(self):
        from reporting.models import DeliveryItem, History

        schedule = self._create_schedule(self.user, '실제납품보고', activity_type='delivery', status='completed')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Actually Sold Kit',
            quantity=1,
            unit='EA',
            unit_price=30000,
        )
        history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=schedule.followup,
            schedule=schedule,
            action_type='delivery_schedule',
            delivery_items='Actually Sold Kit: 1EA (33,000원)\nUnsold Kit: 1EA (77,000원)',
            delivery_amount=110000,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:notes_detail_api', args=[history.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['note']['deliveryAmount'], 33000)
        self.assertIn('Actually Sold Kit', payload['note']['deliveryItems'])
        self.assertNotIn('Unsold Kit', payload['note']['deliveryItems'])

    def test_notes_detail_uses_actual_delivery_items_when_delivery_note_is_linked_to_quote_schedule(self):
        import datetime
        from django.utils import timezone
        from reporting.models import DeliveryItem, History, Schedule

        quote_schedule = self._create_schedule(self.user, '견적연결납품노트', activity_type='quote', status='completed')
        sold_quote_item = DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='56722',
            quantity=1,
            unit='EA',
            unit_price=30000,
            discount_unit_price=30000,
            quote_group='수리',
        )
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='SO447.100E',
            quantity=1,
            unit='EA',
            unit_price=480000,
            discount_unit_price=336000,
            quote_group='보상판매',
        )
        delivery_schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=quote_schedule.followup,
            visit_date=timezone.localdate() + datetime.timedelta(days=1),
            visit_time=datetime.time(9, 0),
            activity_type='delivery',
            status='completed',
        )
        DeliveryItem.objects.create(
            schedule=delivery_schedule,
            item_name=sold_quote_item.item_name,
            quantity=1,
            unit=sold_quote_item.unit,
            unit_price=sold_quote_item.unit_price,
            discount_unit_price=sold_quote_item.discount_unit_price,
            quote_group=sold_quote_item.quote_group,
        )
        stale_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=quote_schedule.followup,
            schedule=quote_schedule,
            action_type='delivery_schedule',
            delivery_items='56722: 1EA (33,000원)\nSO447.100E: 1EA (369,600원)',
            delivery_amount=402600,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:notes_detail_api', args=[stale_history.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['note']['deliveryAmount'], 33000)
        self.assertIn('56722', payload['note']['deliveryItems'])
        self.assertNotIn('SO447.100E', payload['note']['deliveryItems'])

    def test_notes_detail_does_not_count_quote_items_as_delivery_without_actual_delivery(self):
        from reporting.models import DeliveryItem, History

        quote_schedule = self._create_schedule(self.user, '견적만있는납품노트', activity_type='quote', status='completed')
        DeliveryItem.objects.create(
            schedule=quote_schedule,
            item_name='Unsold Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=70000,
        )
        stale_history = History.objects.create(
            user=self.user,
            company=self.company,
            followup=quote_schedule.followup,
            schedule=quote_schedule,
            action_type='delivery_schedule',
            delivery_items='Unsold Quote Kit: 1EA (77,000원)',
            delivery_amount=77000,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse('reporting:notes_detail_api', args=[stale_history.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['note']['deliveryAmount'], 0)
        self.assertEqual(payload['note']['deliveryItems'], '')

    def test_schedule_delivery_items_update_api_does_not_create_delivery_history_for_quote_schedule(self):
        import json
        from reporting.models import DeliveryItem, History

        quote_schedule = self._create_schedule(self.user, '견적품목저장노트방지', activity_type='quote')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[quote_schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'itemName': 'Quote Only Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                        'quoteGroup': '수리',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(DeliveryItem.objects.filter(schedule=quote_schedule).count(), 1)
        self.assertFalse(
            History.objects.filter(schedule=quote_schedule, action_type='delivery_schedule').exists()
        )

    def test_schedule_update_delivery_items_legacy_does_not_create_delivery_history_for_quote_schedule(self):
        from reporting.models import DeliveryItem, History

        quote_schedule = self._create_schedule(self.user, '레거시견적품목저장노트방지', activity_type='quote')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedule_update_delivery_items', args=[quote_schedule.id]),
            data={
                'delivery_items[0][name]': 'Legacy Quote Only Kit',
                'delivery_items[0][quantity]': '1',
                'delivery_items[0][unit_price]': '50000',
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(DeliveryItem.objects.filter(schedule=quote_schedule).count(), 1)
        self.assertFalse(
            History.objects.filter(schedule=quote_schedule, action_type='delivery_schedule').exists()
        )

    def test_schedule_delivery_items_update_api_rejects_coworker_source_quote_completion(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import DeliveryItem, Schedule

        schedule = self._create_schedule(self.user, '동료견적차단납품', activity_type='delivery')
        coworker_quote = Schedule.objects.create(
            user=self.coworker,
            company=self.company,
            followup=schedule.followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        DeliveryItem.objects.create(
            schedule=coworker_quote,
            item_name='Coworker Quote Kit',
            quantity=1,
            unit='EA',
            unit_price=10000,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'sourceQuoteScheduleIds': [coworker_quote.id],
                'items': [
                    {
                        'sourceQuoteScheduleId': coworker_quote.id,
                        'itemName': 'Coworker Quote Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '10000',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertIn('본인이 작성한 견적 일정', response.json()['error'])
        coworker_quote.refresh_from_db()
        self.assertEqual(coworker_quote.status, 'scheduled')
        self.assertFalse(DeliveryItem.objects.filter(schedule=schedule).exists())

    def test_schedule_delivery_items_update_api_accepts_product_master_selection(self):
        import json
        from reporting.models import DeliveryItem, Product

        schedule = self._create_schedule(self.user, '제품선택납품', activity_type='delivery')
        product = Product.objects.create(
            product_code='MASTER-DELIVERY-PCR',
            unit='BOX',
            specification='100 tests',
            standard_price=12345,
            created_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'productId': product.id,
                        'itemName': '',
                        'quantity': 2,
                        'unit': '',
                        'unitPrice': '',
                        'taxInvoiceIssued': True,
                        'notes': '제품 마스터 선택',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        item = DeliveryItem.objects.get(schedule=schedule)
        self.assertEqual(item.product, product)
        self.assertEqual(item.item_name, product.product_code)
        self.assertEqual(item.unit, 'BOX')
        self.assertEqual(int(item.unit_price), 12345)
        self.assertEqual(int(item.total_price), 27159)
        payload_item = response.json()['deliveryItems'][0]
        self.assertEqual(payload_item['productId'], product.id)
        self.assertEqual(payload_item['productCode'], product.product_code)
        self.assertEqual(payload_item['productDescription'], '')
        self.assertEqual(payload_item['unit'], 'BOX')
        self.assertEqual(payload_item['unitPrice'], 12345)
        self.assertEqual(payload_item['totalPrice'], 27159)

    def test_schedule_delivery_items_update_api_saves_quote_groups(self):
        import json
        from reporting.models import DeliveryItem

        schedule = self._create_schedule(self.user, '견적구분저장', activity_type='quote')
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'itemName': 'Trade In Kit',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '100000',
                        'quoteGroup': '보상판매',
                        'taxInvoiceIssued': False,
                    },
                    {
                        'itemName': 'Repair Service',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '50000',
                        'quoteGroup': '수리',
                        'taxInvoiceIssued': False,
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        items = list(DeliveryItem.objects.filter(schedule=schedule).order_by('id'))
        self.assertEqual([item.quote_group for item in items], ['보상판매', '수리'])
        payload_items = response.json()['deliveryItems']
        self.assertEqual([item['quoteGroup'] for item in payload_items], ['보상판매', '수리'])
        self.assertEqual([item['quoteGroupLabel'] for item in payload_items], ['보상판매', '수리'])

    def test_schedule_delivery_items_update_api_blocks_inaccessible_product(self):
        import json
        from reporting.models import DeliveryItem, Product

        schedule = self._create_schedule(self.user, '타사제품차단', activity_type='delivery')
        other_product = Product.objects.create(
            product_code='MASTER-OTHER-PRIVATE',
            unit='EA',
            standard_price=5000,
            created_by=self.other_user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id]),
            data=json.dumps({
                'items': [
                    {
                        'productId': other_product.id,
                        'itemName': '허용되지 않은 제품',
                        'quantity': 1,
                        'unit': 'EA',
                        'unitPrice': '5000',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('선택한 제품을 찾을 수 없습니다', response.json()['error'])
        self.assertFalse(DeliveryItem.objects.filter(schedule=schedule).exists())

    def test_schedule_delivery_items_update_api_blocks_manager_and_coworker(self):
        import json
        from reporting.models import DeliveryItem

        schedule = self._create_schedule(self.user, '납품차단', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Protected Kit',
            quantity=1,
            unit='EA',
            unit_price=5000,
        )
        update_url = reverse('reporting:schedules_delivery_items_update_api', args=[schedule.id])
        payload = {
            'items': [
                {
                    'itemName': 'Changed Kit',
                    'quantity': 2,
                    'unit': 'EA',
                    'unitPrice': '10000',
                },
            ],
        }

        self.client.force_login(self.manager)
        manager_response = self.client.post(update_url, data=json.dumps(payload), content_type='application/json')
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(update_url, data=json.dumps(payload), content_type='application/json')
        self.assertEqual(coworker_response.status_code, 403)

        self.assertEqual(DeliveryItem.objects.get(schedule=schedule).item_name, 'Protected Kit')

    def test_schedule_file_upload_api_allows_owner_only(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import ScheduleFile

        schedule = self._create_schedule(self.user, '파일업로드')
        upload_url = reverse('reporting:schedule_file_upload', args=[schedule.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('manager.txt', b'manager memo', content_type='text/plain'),
        })
        self.assertEqual(manager_response.status_code, 403)

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('coworker.txt', b'coworker memo', content_type='text/plain'),
        })
        self.assertEqual(coworker_response.status_code, 403)

        self.client.force_login(self.user)
        response = self.client.post(upload_url, {
            'files': SimpleUploadedFile('owner.txt', b'owner memo', content_type='text/plain'),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        schedule_file = ScheduleFile.objects.get(schedule=schedule)
        self.addCleanup(schedule_file.file.delete, False)
        self.assertEqual(schedule_file.original_filename, 'owner.txt')
        self.assertEqual(payload['files'][0]['id'], schedule_file.id)
        self.assertEqual(payload['files'][0]['downloadHref'], reverse('reporting:schedule_file_download', args=[schedule_file.id]))
        self.assertEqual(payload['files'][0]['deleteHref'], reverse('reporting:schedule_file_delete', args=[schedule_file.id]))

    def test_schedule_file_delete_api_allows_owner_only(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import ScheduleFile

        schedule = self._create_schedule(self.user, '파일삭제')
        schedule_file = ScheduleFile.objects.create(
            schedule=schedule,
            file=SimpleUploadedFile('delete-me.txt', b'delete memo', content_type='text/plain'),
            original_filename='delete-me.txt',
            file_size=11,
            uploaded_by=self.user,
        )
        delete_url = reverse('reporting:schedule_file_delete', args=[schedule_file.id])

        self.client.force_login(self.manager)
        manager_response = self.client.post(delete_url)
        self.assertEqual(manager_response.status_code, 403)
        self.assertTrue(ScheduleFile.objects.filter(pk=schedule_file.id).exists())

        self.client.force_login(self.coworker)
        coworker_response = self.client.post(delete_url)
        self.assertEqual(coworker_response.status_code, 403)
        self.assertTrue(ScheduleFile.objects.filter(pk=schedule_file.id).exists())

        self.client.force_login(self.user)
        response = self.client.post(delete_url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertFalse(ScheduleFile.objects.filter(pk=schedule_file.id).exists())

    def test_schedule_file_download_blocks_anonymous_and_out_of_scope_users(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from reporting.models import ScheduleFile

        schedule = self._create_schedule(self.user, '파일다운로드')
        schedule_file = ScheduleFile.objects.create(
            schedule=schedule,
            file=SimpleUploadedFile('schedule-download.txt', b'schedule memo', content_type='text/plain'),
            original_filename='schedule-download.txt',
            file_size=13,
            uploaded_by=self.user,
        )
        self.addCleanup(schedule_file.file.delete, False)
        download_url = reverse('reporting:schedule_file_download', args=[schedule_file.id])

        anonymous_response = self.client.get(download_url)
        self.assertEqual(anonymous_response.status_code, 302)
        self.assertIn('/reporting/login/', anonymous_response['Location'])

        self.client.force_login(self.other_user)
        other_response = self.client.get(download_url)
        self.assertEqual(other_response.status_code, 403)
        self.assertNotIn('attachment', other_response.get('Content-Disposition', ''))

        self.client.force_login(self.user)
        owner_response = self.client.get(download_url)
        self.assertEqual(owner_response.status_code, 200)
        self.assertIn('attachment', owner_response.get('Content-Disposition', ''))
        owner_response.close()


class DocumentTemplatesReactApiTests(TestCase):
    """React 서류 템플릿 관리 API 회귀 테스트"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='서류API회사')
        self.other_company = UserCompany.objects.create(name='서류API타사회사')
        self.admin = make_user('doc-admin', role='admin', company=self.company)
        self.manager = make_user('doc-manager', role='manager', company=self.company)
        self.salesman = make_user('doc-sales', role='salesman', company=self.company)
        self.other_manager = make_user('doc-other-manager', role='manager', company=self.other_company)
        self.list_url = reverse('reporting:document_templates_api')
        self.create_url = reverse('reporting:document_template_api_create')

    def _uploaded_xlsx(self, name='template.xlsx'):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile(
            name,
            b'fake xlsx content',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _create_template(self, company, name, document_type='quotation', is_default=False, created_by=None):
        template = DocumentTemplate.objects.create(
            company=company,
            document_type=document_type,
            name=name,
            description=f'{name} 설명',
            file=self._uploaded_xlsx(f'{name}.xlsx'),
            file_type='xlsx',
            is_default=is_default,
            created_by=created_by or self.manager,
        )
        self.addCleanup(template.file.delete, False)
        return template

    def _create_schedule(self, owner, name='서류고객', activity_type='quote'):
        from reporting.models import Schedule

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            company=customer_company,
            department=department,
            manager=f'{name} 책임',
        )
        return Schedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate(),
            visit_time=timezone.now().time(),
            activity_type=activity_type,
        )

    def test_document_templates_api_requires_login(self):
        response = self.client.get(self.list_url)

        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_schedule_document_preview_generate_and_download_are_protected(self):
        schedule = self._create_schedule(self.manager, name='보호서류', activity_type='quote')
        self._create_template(self.company, '보호견적서', is_default=True)
        log = DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=schedule,
            user=self.manager,
            transaction_number='PROTECT-001',
            output_format='pdf',
            file=SimpleUploadedFile('protected-quote.pdf', b'%PDF protected', content_type='application/pdf'),
            filename='protected-quote.pdf',
            file_size=14,
        )
        self.addCleanup(log.file.delete, False)

        preview_url = reverse('reporting:get_document_template_data', args=['quotation', schedule.id])
        generate_url = reverse('reporting:generate_document_pdf_format', args=['quotation', schedule.id, 'pdf'])
        download_url = reverse('reporting:generated_document_download', args=[log.id])

        anonymous_preview = self.client.get(preview_url)
        anonymous_generate = self.client.post(generate_url)
        anonymous_download = self.client.get(download_url)
        self.assertEqual(anonymous_preview.status_code, 302)
        self.assertIn('/reporting/login/', anonymous_preview['Location'])
        self.assertEqual(anonymous_generate.status_code, 302)
        self.assertIn('/reporting/login/', anonymous_generate['Location'])
        self.assertEqual(anonymous_download.status_code, 302)
        self.assertIn('/reporting/login/', anonymous_download['Location'])

        self.client.force_login(self.other_manager)
        other_preview = self.client.get(preview_url)
        other_generate = self.client.post(generate_url)
        other_download = self.client.get(download_url)
        self.assertEqual(other_preview.status_code, 403)
        self.assertEqual(other_generate.status_code, 403)
        self.assertEqual(other_download.status_code, 403)

    def test_document_templates_api_lists_same_company_and_summary(self):
        default_template = self._create_template(self.company, '기본견적서', is_default=True)
        delivery_template = self._create_template(self.company, '납품서', document_type='delivery_note')
        self._create_template(self.other_company, '타사견적서', created_by=self.other_manager)
        self.client.force_login(self.salesman)

        response = self.client.get(self.list_url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['source'], 'django')
        self.assertFalse(payload['create']['canCreate'])
        self.assertEqual({item['id'] for item in payload['templates']}, {default_template.id, delivery_template.id})
        self.assertNotIn('타사견적서', [item['name'] for item in payload['templates']])
        quotation_summary = next(item for item in payload['summary']['byType'] if item['type'] == 'quotation')
        self.assertEqual(quotation_summary['count'], 1)
        self.assertEqual(quotation_summary['defaultCount'], 1)
        variable_tokens = {
            variable['token']
            for group in payload['templateVariableGroups']
            for variable in group['variables']
        }
        self.assertIn('{{견적기타사항}}', variable_tokens)
        self.assertIn('{{견적구분}}', variable_tokens)
        self.assertIn('{{견적명}}', variable_tokens)
        self.assertIn('{{견적제목}}', variable_tokens)
        self.assertIn('{{품목1_옵션}}', variable_tokens)
        self.assertIn('{{품목1_옵션설명}}', variable_tokens)
        self.assertIn('{{품목1_적요}}', variable_tokens)
        self.assertIn('{{품목1_기준단가}}', variable_tokens)
        self.assertIn('{{품목1_할인율}}', variable_tokens)
        self.assertIn('{{품목1_할인단가}}', variable_tokens)
        self.assertEqual(payload['links']['djangoList'], reverse('reporting:document_template_list'))

    def test_document_templates_api_includes_recent_generation_logs_scoped_by_company(self):
        quote_schedule = self._create_schedule(self.manager, name='견적이력', activity_type='quote')
        delivery_schedule = self._create_schedule(self.manager, name='납품이력', activity_type='delivery')
        other_schedule = self._create_schedule(self.other_manager, name='타사이력', activity_type='quote')
        quote_log = DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='quotation',
            schedule=quote_schedule,
            user=self.manager,
            transaction_number='Q-20260511-001',
            output_format='xlsx',
        )
        DocumentGenerationLog.objects.create(
            company=self.company,
            document_type='delivery_note',
            schedule=delivery_schedule,
            user=self.manager,
            transaction_number='D-20260511-001',
            output_format='pdf',
        )
        DocumentGenerationLog.objects.create(
            company=self.other_company,
            document_type='quotation',
            schedule=other_schedule,
            user=self.other_manager,
            transaction_number='OTHER-001',
            output_format='xlsx',
        )
        self.client.force_login(self.salesman)

        response = self.client.get(self.list_url, {'type': 'quotation'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['summary']['generatedToday'], 1)
        self.assertEqual(payload['summary']['recentGenerationCount'], 1)
        self.assertEqual([item['id'] for item in payload['recentGenerations']], [quote_log.id])
        generation = payload['recentGenerations'][0]
        self.assertEqual(generation['transactionNumber'], 'Q-20260511-001')
        self.assertEqual(generation['documentTypeLabel'], '견적서')
        self.assertEqual(generation['outputFormatLabel'], 'Excel')
        self.assertEqual(generation['createdBy'], self.manager.get_full_name() or self.manager.username)
        self.assertEqual(generation['customerName'], '견적이력 담당자')
        self.assertEqual(generation['customerCompany'], '견적이력 회사')
        self.assertEqual(generation['departmentName'], '견적이력 연구실')
        self.assertEqual(generation['schedule']['href'], f'/schedules/{quote_schedule.id}/')
        self.assertEqual(generation['schedule']['djangoHref'], reverse('reporting:schedule_detail', args=[quote_schedule.id]))
        self.assertNotIn('OTHER-001', [item['transactionNumber'] for item in payload['recentGenerations']])

    def test_document_templates_api_filters_by_document_type(self):
        quotation = self._create_template(self.company, '견적서')
        self._create_template(self.company, '거래명세서', document_type='transaction_statement')
        self.client.force_login(self.manager)

        response = self.client.get(self.list_url, {'type': 'quotation'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item['id'] for item in payload['templates']], [quotation.id])
        self.assertTrue(payload['create']['canCreate'])

    def test_document_template_create_update_delete_api(self):
        old_default = self._create_template(self.company, '기존기본', is_default=True)
        self.client.force_login(self.manager)

        create_response = self.client.post(self.create_url, {
            'documentType': 'quotation',
            'name': '신규견적서',
            'description': 'React 업로드',
            'isDefault': 'true',
            'file': self._uploaded_xlsx('new-template.xlsx'),
        })

        self.assertEqual(create_response.status_code, 200)
        created_payload = create_response.json()
        self.assertTrue(created_payload['success'])
        created = DocumentTemplate.objects.get(pk=created_payload['template']['id'])
        self.addCleanup(created.file.delete, False)
        self.assertEqual(created.company, self.company)
        self.assertEqual(created.created_by, self.manager)
        self.assertTrue(created.is_default)
        old_default.refresh_from_db()
        self.assertFalse(old_default.is_default)

        update_response = self.client.post(
            reverse('reporting:document_template_api_update', args=[created.id]),
            {
                'documentType': 'transaction_statement',
                'name': '수정거래명세서',
                'description': '수정 설명',
                'isDefault': 'false',
            },
        )

        self.assertEqual(update_response.status_code, 200)
        created.refresh_from_db()
        self.assertEqual(created.document_type, 'transaction_statement')
        self.assertEqual(created.name, '수정거래명세서')
        self.assertFalse(created.is_default)

        delete_response = self.client.post(reverse('reporting:document_template_api_delete', args=[created.id]))

        self.assertEqual(delete_response.status_code, 200)
        created.refresh_from_db()
        self.assertFalse(created.is_active)

    def test_document_template_api_blocks_salesman_mutations(self):
        template = self._create_template(self.company, '수정불가')
        self.client.force_login(self.salesman)

        create_response = self.client.post(self.create_url, {
            'documentType': 'quotation',
            'name': '권한없음',
            'file': self._uploaded_xlsx('blocked.xlsx'),
        })
        update_response = self.client.post(
            reverse('reporting:document_template_api_update', args=[template.id]),
            {'documentType': 'quotation', 'name': '수정시도'},
        )
        delete_response = self.client.post(reverse('reporting:document_template_api_delete', args=[template.id]))

        self.assertEqual(create_response.status_code, 403)
        self.assertEqual(update_response.status_code, 403)
        self.assertEqual(delete_response.status_code, 403)

    def test_document_template_api_blocks_other_company(self):
        other_template = self._create_template(self.other_company, '타사서류', created_by=self.other_manager)
        self.client.force_login(self.manager)

        update_response = self.client.post(
            reverse('reporting:document_template_api_update', args=[other_template.id]),
            {'documentType': 'quotation', 'name': '타사수정'},
        )
        toggle_response = self.client.post(reverse('reporting:document_template_api_toggle_default', args=[other_template.id]))

        self.assertEqual(update_response.status_code, 403)
        self.assertEqual(toggle_response.status_code, 403)

    def test_document_template_toggle_default_api_uses_existing_single_default_rule(self):
        old_default = self._create_template(self.company, '기존기본', is_default=True)
        new_template = self._create_template(self.company, '새기본')
        self.client.force_login(self.salesman)

        response = self.client.post(reverse('reporting:document_template_api_toggle_default', args=[new_template.id]))

        self.assertEqual(response.status_code, 200)
        new_template.refresh_from_db()
        old_default.refresh_from_db()
        self.assertTrue(new_template.is_default)
        self.assertFalse(old_default.is_default)

    def test_document_template_data_includes_quote_discount_and_note_variables(self):
        from reporting.models import DeliveryItem

        self.manager.first_name = '재현'
        self.manager.last_name = '안'
        self.manager.save(update_fields=['first_name', 'last_name'])
        self._create_template(self.company, '견적기본', is_default=True)
        schedule = self._create_schedule(self.manager, name='견적변수', activity_type='quote')
        schedule.notes = '견적 메모'
        schedule.quote_extra_notes = '전체 견적 기타사항'
        schedule.save(update_fields=['notes', 'quote_extra_notes'])
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='PCR Kit',
            quantity=2,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
            notes='품목 적요',
            option_description='품목 옵션 설명',
        )
        self.client.force_login(self.salesman)

        response = self.client.get(reverse('reporting:get_document_template_data', args=['quotation', schedule.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        variables = payload['variables']
        self.assertEqual(variables['실무자'], '안재현')
        self.assertEqual(variables['영업담당자'], '안재현')
        self.assertEqual(variables['담당영업'], '안재현')
        self.assertEqual(variables['메모'], '견적 메모')
        self.assertEqual(variables['기타사항'], '전체 견적 기타사항')
        self.assertEqual(variables['견적기타사항'], '전체 견적 기타사항')
        self.assertEqual(variables['품목1_옵션'], '품목 옵션 설명')
        self.assertEqual(variables['품목1_옵션설명'], '품목 옵션 설명')
        self.assertEqual(variables['품목1_적요'], '품목 적요')
        self.assertEqual(variables['품목1_비고'], '품목 적요')
        self.assertEqual(variables['품목1_기준단가'], '100,000')
        self.assertEqual(variables['품목1_할인율'], '10%')
        self.assertEqual(variables['품목1_할인단가'], '90,000')
        self.assertEqual(variables['품목1_단가'], '90,000')
        self.assertEqual(variables['공급가액'], '180,000')
        self.assertEqual(variables['부가세액'], '18,000')
        self.assertEqual(variables['총액'], '198,000')
        self.assertEqual(payload['items'][0]['unitPrice'], 90000)
        self.assertEqual(payload['items'][0]['baseUnitPrice'], 100000)
        self.assertEqual(payload['items'][0]['discountUnitPrice'], 90000)
        self.assertEqual(payload['items'][0]['discountRate'], 10.0)
        self.assertEqual(payload['items'][0]['notes'], '품목 적요')
        self.assertEqual(payload['items'][0]['optionDescription'], '품목 옵션 설명')

    def test_document_template_data_hides_quotation_base_unit_price_when_requested(self):
        from reporting.models import DeliveryItem

        self._create_template(self.company, '견적기본', is_default=True)
        schedule = self._create_schedule(self.manager, name='기준단가숨김', activity_type='quote')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Hidden Base Kit',
            quantity=2,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
        )
        self.client.force_login(self.salesman)

        response = self.client.get(
            reverse('reporting:get_document_template_data', args=['quotation', schedule.id]),
            {'hide_base_unit_price': '1'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertTrue(payload['hide_base_unit_price'])
        variables = payload['variables']
        self.assertEqual(variables['품목1_기준단가'], '')
        self.assertEqual(variables['품목1_할인단가'], '90,000')
        self.assertEqual(variables['품목1_단가'], '90,000')
        self.assertEqual(variables['공급가액'], '180,000')
        self.assertIsNone(payload['items'][0]['baseUnitPrice'])
        self.assertTrue(payload['items'][0]['baseUnitPriceHidden'])
        self.assertEqual(payload['items'][0]['unitPrice'], 90000)
        self.assertEqual(payload['items'][0]['discountUnitPrice'], 90000)

    def test_document_template_data_uses_billable_unit_price_for_transaction_statement_base_token(self):
        from reporting.models import DeliveryItem

        self._create_template(self.company, '거래명세서기본', document_type='transaction_statement', is_default=True)
        schedule = self._create_schedule(self.manager, name='거래기준단가', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Visible Base Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
        )
        self.client.force_login(self.salesman)

        response = self.client.get(
            reverse('reporting:get_document_template_data', args=['transaction_statement', schedule.id]),
            {'hide_base_unit_price': '1'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload['hide_base_unit_price'])
        self.assertEqual(payload['variables']['품목1_기준단가'], '90,000')
        self.assertEqual(payload['variables']['품목1_단가'], '90,000')
        self.assertEqual(payload['items'][0]['baseUnitPrice'], 100000)
        self.assertFalse(payload['items'][0]['baseUnitPriceHidden'])

    def test_document_template_data_normalizes_legacy_reversed_korean_salesperson_name(self):
        self.manager.first_name = '안'
        self.manager.last_name = '재현'
        self.manager.save(update_fields=['first_name', 'last_name'])
        self._create_template(self.company, '견적담당자', is_default=True)
        schedule = self._create_schedule(self.manager, name='담당자변수', activity_type='quote')
        self.client.force_login(self.salesman)

        response = self.client.get(reverse('reporting:get_document_template_data', args=['quotation', schedule.id]))

        self.assertEqual(response.status_code, 200)
        variables = response.json()['variables']
        self.assertEqual(variables['실무자'], '안재현')
        self.assertEqual(variables['영업담당자'], '안재현')
        self.assertEqual(variables['담당영업'], '안재현')

    def test_document_template_data_filters_quotation_items_by_quote_group(self):
        from reporting.models import DeliveryItem

        self._create_template(self.company, '견적구분기본', is_default=True)
        schedule = self._create_schedule(self.manager, name='견적구분변수', activity_type='quote')
        ScheduleQuoteGroupNote.objects.create(schedule=schedule, quote_group='보상판매', notes='보상판매 조건')
        ScheduleQuoteGroupNote.objects.create(schedule=schedule, quote_group='수리', notes='수리 조건')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Trade In Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            quote_group='보상판매',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Repair Service',
            quantity=1,
            unit='EA',
            unit_price=50000,
            quote_group='수리',
        )
        self.client.force_login(self.salesman)

        response = self.client.get(
            reverse('reporting:get_document_template_data', args=['quotation', schedule.id]),
            {'quote_group': '수리'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['quote_group'], '수리')
        self.assertEqual(payload['quote_group_label'], '수리')
        self.assertEqual(payload['variables']['견적구분'], '수리')
        self.assertEqual(payload['variables']['견적제목'], '수리 견적서')
        self.assertEqual(payload['variables']['견적기타사항'], '수리 조건')
        self.assertEqual(payload['variables']['기타사항'], '수리 조건')
        self.assertEqual(payload['item_count'], 1)
        self.assertEqual(payload['items'][0]['name'], 'Repair Service')
        self.assertEqual(payload['items'][0]['quoteGroup'], '수리')
        self.assertEqual(payload['variables']['품목1_이름'], 'Repair Service')
        self.assertNotIn('품목2_이름', payload['variables'])

    def test_document_pdf_layout_helper_sets_a4_fit_to_page(self):
        import os
        import tempfile
        import zipfile
        from openpyxl import Workbook
        from reporting.views import _ensure_xlsx_a4_print_layout

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '견적서'
        sheet['J40'] = 'A4 자동 맞춤 테스트'
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        self.addCleanup(lambda: os.path.exists(temp_path) and os.unlink(temp_path))
        workbook.save(temp_path)

        changed = _ensure_xlsx_a4_print_layout(temp_path)

        self.assertTrue(changed)
        with zipfile.ZipFile(temp_path, 'r') as archive:
            sheet_xml = archive.read('xl/worksheets/sheet1.xml').decode('utf-8')
        self.assertIn('fitToPage="1"', sheet_xml)
        self.assertIn('paperSize="9"', sheet_xml)
        self.assertIn('fitToWidth="1"', sheet_xml)
        self.assertIn('fitToHeight="0"', sheet_xml)
        self.assertIn('left="0.25"', sheet_xml)

    def test_document_base_unit_price_column_helper_hides_token_column(self):
        import os
        import tempfile
        import zipfile
        from xml.etree import ElementTree as ET
        from openpyxl import Workbook
        from reporting.views import _hide_xlsx_base_unit_price_columns

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '품목'
        sheet['B1'] = '기준단가'
        sheet['C1'] = '단가'
        sheet['A2'] = '{{품목1_이름}}'
        sheet['B2'] = '{{품목1_기준단가}}'
        sheet['C2'] = '{{품목1_단가}}'
        sheet.column_dimensions['B'].width = 14
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        self.addCleanup(lambda: os.path.exists(temp_path) and os.unlink(temp_path))
        workbook.save(temp_path)

        changed = _hide_xlsx_base_unit_price_columns(temp_path)

        self.assertTrue(changed)
        with zipfile.ZipFile(temp_path, 'r') as archive:
            sheet_root = ET.fromstring(archive.read('xl/worksheets/sheet1.xml'))

        namespace = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        hidden_cols = [
            col for col in sheet_root.findall('.//s:cols/s:col', namespace)
            if col.get('hidden') == '1'
        ]
        self.assertEqual([(col.get('min'), col.get('max')) for col in hidden_cols], [('2', '2')])

    def test_document_generate_xlsx_hides_base_unit_price_column_when_requested(self):
        import io
        import zipfile
        from xml.etree import ElementTree as ET
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook
        from reporting.models import DeliveryItem

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '품목'
        sheet['B1'] = '기준단가'
        sheet['C1'] = '단가'
        sheet['A2'] = '{{품목1_이름}}'
        sheet['B2'] = '{{품목1_기준단가}}'
        sheet['C2'] = '{{품목1_단가}}'
        output = io.BytesIO()
        workbook.save(output)
        template = DocumentTemplate.objects.create(
            company=self.company,
            document_type='quotation',
            name='열숨김견적서',
            file=SimpleUploadedFile(
                'column-hide-quote.xlsx',
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            file_type='xlsx',
            is_default=True,
            created_by=self.manager,
        )
        self.addCleanup(template.file.delete, False)
        schedule = self._create_schedule(self.manager, name='열숨김', activity_type='quote')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Column Hidden Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            discount_rate=10,
        )
        self.client.force_login(self.salesman)

        response = self.client.post(
            reverse('reporting:generate_document_pdf_format', args=['quotation', schedule.id, 'xlsx']),
            {'hide_base_unit_price': '1'},
        )

        self.assertEqual(response.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(response.content), 'r') as archive:
            sheet_root = ET.fromstring(archive.read('xl/worksheets/sheet1.xml'))

        namespace = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        hidden_cols = [
            col for col in sheet_root.findall('.//s:cols/s:col', namespace)
            if col.get('hidden') == '1'
        ]
        self.assertEqual([(col.get('min'), col.get('max')) for col in hidden_cols], [('2', '2')])

    def test_document_generate_transaction_statement_uses_discount_unit_price_for_base_token(self):
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook, load_workbook
        from reporting.models import DeliveryItem

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '품목'
        sheet['B1'] = '단가'
        sheet['C1'] = '공급가액'
        sheet['A2'] = '{{품목1_이름}}'
        sheet['B2'] = '{{품목1_기준단가}}'
        sheet['C2'] = '{{품목1_금액}}'
        output = io.BytesIO()
        workbook.save(output)
        template = DocumentTemplate.objects.create(
            company=self.company,
            document_type='transaction_statement',
            name='거래명세서단가',
            file=SimpleUploadedFile(
                'transaction-statement-unit-price.xlsx',
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            file_type='xlsx',
            is_default=True,
            created_by=self.manager,
        )
        self.addCleanup(template.file.delete, False)
        schedule = self._create_schedule(self.manager, name='거래명세서단가', activity_type='delivery')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Discounted Statement Kit',
            quantity=2,
            unit='EA',
            unit_price=110000,
            discount_unit_price=84000,
        )
        self.client.force_login(self.salesman)

        response = self.client.post(
            reverse('reporting:generate_document_pdf_format', args=['transaction_statement', schedule.id, 'xlsx']),
        )

        self.assertEqual(response.status_code, 200)
        generated = load_workbook(io.BytesIO(response.content))
        generated_sheet = generated.active
        self.assertEqual(generated_sheet['A2'].value, 'Discounted Statement Kit')
        self.assertEqual(generated_sheet['B2'].value, '84,000')
        self.assertEqual(generated_sheet['C2'].value, '168,000')

    def test_document_generate_xlsx_inserts_quote_item_option_rows(self):
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook, load_workbook
        from reporting.models import DeliveryItem

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '품목코드'
        sheet['B1'] = '품목명(규격)'
        sheet['C1'] = '수량'
        sheet['D1'] = '단가'
        sheet['E1'] = '적요'
        sheet['A2'] = '{{품목1_이름}}'
        sheet['B2'] = '{{품목1_설명}}'
        sheet['C2'] = '{{품목1_수량}}'
        sheet['D2'] = '{{품목1_단가}}'
        sheet['E2'] = '{{품목1_적요}}'
        sheet['A3'] = '{{품목2_이름}}'
        sheet['B3'] = '{{품목2_설명}}'
        sheet['C3'] = '{{품목2_수량}}'
        sheet['D3'] = '{{품목2_단가}}'
        sheet['E3'] = '{{품목2_적요}}'
        sheet['A4'] = '합계'
        output = io.BytesIO()
        workbook.save(output)
        template = DocumentTemplate.objects.create(
            company=self.company,
            document_type='quotation',
            name='옵션행견적서',
            file=SimpleUploadedFile(
                'quote-option-row.xlsx',
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            file_type='xlsx',
            is_default=True,
            created_by=self.manager,
        )
        self.addCleanup(template.file.delete, False)
        schedule = self._create_schedule(self.manager, name='옵션행견적', activity_type='quote')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Option Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            notes='적요 A',
            option_description='옵션 A 포함, 설치 조건 별도',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Plain Kit',
            quantity=2,
            unit='EA',
            unit_price=50000,
        )
        self.client.force_login(self.salesman)

        response = self.client.post(
            reverse('reporting:generate_document_pdf_format', args=['quotation', schedule.id, 'xlsx']),
        )

        self.assertEqual(response.status_code, 200)
        generated = load_workbook(io.BytesIO(response.content))
        generated_sheet = generated.active
        self.assertEqual(generated_sheet['A2'].value, 'Option Kit')
        self.assertEqual(generated_sheet['E2'].value, '적요 A')
        self.assertEqual(generated_sheet['B3'].value, '옵션: 옵션 A 포함, 설치 조건 별도')
        self.assertIn('B3:E3', [str(merge_range) for merge_range in generated_sheet.merged_cells.ranges])
        self.assertEqual(generated_sheet['A4'].value, 'Plain Kit')
        self.assertEqual(generated_sheet['A5'].value, '합계')

    def test_document_generate_xlsx_replaces_quote_item_option_variables(self):
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook, load_workbook
        from reporting.models import DeliveryItem

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '{{품목1_이름}}'
        sheet['B1'] = '{{품목1_옵션설명}}'
        output = io.BytesIO()
        workbook.save(output)
        template = DocumentTemplate.objects.create(
            company=self.company,
            document_type='quotation',
            name='옵션견적서',
            file=SimpleUploadedFile(
                'quote-option.xlsx',
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
            file_type='xlsx',
            is_default=True,
            created_by=self.manager,
        )
        self.addCleanup(template.file.delete, False)
        schedule = self._create_schedule(self.manager, name='옵션견적', activity_type='quote')
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='Option Kit',
            quantity=1,
            unit='EA',
            unit_price=100000,
            option_description='옵션 A 포함, 설치 조건 별도',
        )
        self.client.force_login(self.salesman)

        response = self.client.post(
            reverse('reporting:generate_document_pdf_format', args=['quotation', schedule.id, 'xlsx']),
        )

        self.assertEqual(response.status_code, 200)
        generated = load_workbook(io.BytesIO(response.content))
        generated_sheet = generated.active
        self.assertEqual(generated_sheet['A1'].value, 'Option Kit')
        self.assertEqual(generated_sheet['B1'].value, '옵션 A 포함, 설치 조건 별도')

    def test_document_item_note_layout_helper_wraps_and_expands_note_rows(self):
        import os
        import tempfile
        import zipfile
        from xml.etree import ElementTree as ET
        from openpyxl import Workbook
        from reporting.views import _expand_xlsx_item_note_rows

        workbook = Workbook()
        sheet = workbook.active
        sheet.column_dimensions['B'].width = 12
        sheet['B5'] = '{{품목1_적요}}'
        long_note = (
            '내부 세척 및 오염 제거 후 정상 볼륨 확인. '
            '오링 마모가 심해 교체가 필요하며 수리 진행 여부 회신 요청.'
        )
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        self.addCleanup(lambda: os.path.exists(temp_path) and os.unlink(temp_path))
        workbook.save(temp_path)

        changed = _expand_xlsx_item_note_rows(temp_path, {'품목1_적요': long_note})

        self.assertTrue(changed)
        with zipfile.ZipFile(temp_path, 'r') as archive:
            styles_xml = archive.read('xl/styles.xml').decode('utf-8')
            sheet_root = ET.fromstring(archive.read('xl/worksheets/sheet1.xml'))

        namespace = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        row = sheet_root.find(".//s:row[@r='5']", namespace)
        cell = sheet_root.find(".//s:c[@r='B5']", namespace)
        self.assertIsNotNone(row)
        self.assertIsNotNone(cell)
        self.assertEqual(row.get('customHeight'), '1')
        self.assertGreater(float(row.get('ht')), 15.0)
        self.assertNotEqual(cell.get('s'), '0')
        self.assertIn('wrapText="1"', styles_xml)

    def test_document_template_text_layout_helper_wraps_long_replaced_text(self):
        import os
        import tempfile
        import zipfile
        from xml.etree import ElementTree as ET
        from openpyxl import Workbook
        from reporting.views import _expand_xlsx_template_text_rows

        workbook = Workbook()
        sheet = workbook.active
        sheet.column_dimensions['B'].width = 10
        sheet['B4'] = '{{업체명}}'
        long_company_name = '아주 긴 학교명 및 산학협력단 공동연구센터 세포분석실'
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        self.addCleanup(lambda: os.path.exists(temp_path) and os.unlink(temp_path))
        workbook.save(temp_path)

        changed = _expand_xlsx_template_text_rows(temp_path, {'업체명': long_company_name})

        self.assertTrue(changed)
        with zipfile.ZipFile(temp_path, 'r') as archive:
            styles_xml = archive.read('xl/styles.xml').decode('utf-8')
            sheet_root = ET.fromstring(archive.read('xl/worksheets/sheet1.xml'))

        namespace = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        row = sheet_root.find(".//s:row[@r='4']", namespace)
        cell = sheet_root.find(".//s:c[@r='B4']", namespace)
        self.assertIsNotNone(row)
        self.assertIsNotNone(cell)
        self.assertEqual(row.get('customHeight'), '1')
        self.assertGreater(float(row.get('ht')), 15.0)
        self.assertNotEqual(cell.get('s'), '0')
        self.assertIn('wrapText="1"', styles_xml)

    def test_document_bold_strip_helper_removes_bold_styles(self):
        import os
        import tempfile
        import zipfile
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from reporting.views import _strip_xlsx_bold_formatting

        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = '견적서'
        sheet['A1'].font = Font(bold=True)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_path = temp_file.name
        self.addCleanup(lambda: os.path.exists(temp_path) and os.unlink(temp_path))
        workbook.save(temp_path)

        changed = _strip_xlsx_bold_formatting(temp_path)

        self.assertTrue(changed)
        with zipfile.ZipFile(temp_path, 'r') as archive:
            styles_xml = archive.read('xl/styles.xml').decode('utf-8')
        self.assertNotIn('<b ', styles_xml)
        self.assertNotIn('<b/>', styles_xml)
        self.assertNotIn('<b>', styles_xml)


class ScheduleAiCoachApiTests(TestCase):
    """일정 AI 코치 API 검증 (AI 워크스페이스 메뉴 제거와 무관한 별도 기능, 보존)"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='AI코치회사')
        self.other_company = UserCompany.objects.create(name='AI코치타사회사')
        self.user = make_user('ai_coach_me', role='salesman', can_use_ai=True, company=self.company)
        self.no_ai_user = make_user('ai_coach_no_permission', role='salesman', can_use_ai=False, company=self.company)

    def _create_customer(self, owner, name):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            manager=f'{name} 책임',
            company=customer_company,
            department=department,
            priority='urgent',
            customer_grade='A',
            ai_score=88,
        )
        return followup, department

    def test_schedule_ai_coach_requires_ai_permission(self):
        from datetime import time

        followup, _department = self._create_customer(self.no_ai_user, '일정코치권한없음')
        schedule = Schedule.objects.create(
            user=self.no_ai_user,
            company=self.company,
            followup=followup,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        self.client.force_login(self.no_ai_user)

        response = self.client.post(reverse('reporting:schedule_ai_coach_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['error'], 'permission_denied')

    def test_schedule_ai_coach_blocks_inaccessible_schedule(self):
        from datetime import time

        outsider = make_user('ai_coach_schedule_outsider', role='salesman', can_use_ai=True, company=self.other_company)
        followup, _department = self._create_customer(outsider, '일정코치타사회사')
        schedule = Schedule.objects.create(
            user=outsider,
            company=self.other_company,
            followup=followup,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            activity_type='customer_meeting',
            status='scheduled',
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:schedule_ai_coach_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()['error'], 'permission_denied')

    def test_schedule_ai_coach_returns_unsaved_fallback_for_accessible_schedule(self):
        from datetime import time
        from decimal import Decimal
        from reporting.models import AIWorkspaceQuestionLog, DeliveryItem

        followup, _department = self._create_customer(self.user, '일정코치')
        schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=followup,
            visit_date=timezone.localdate(),
            visit_time=time(14, 30),
            activity_type='delivery',
            status='scheduled',
            expected_revenue=Decimal('330000'),
            notes='납품 품목과 세금계산서 확인 필요',
        )
        DeliveryItem.objects.create(
            schedule=schedule,
            item_name='P1000',
            quantity=3,
            unit='EA',
            unit_price=Decimal('100000'),
        )
        self.client.force_login(self.user)

        response = self.client.post(reverse('reporting:schedule_ai_coach_api', args=[schedule.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['context']['scheduleId'], schedule.id)
        self.assertFalse(payload['context']['stored'])
        self.assertIn('납품', payload['coach']['summary'])
        self.assertEqual(payload['coach']['afterMeetingNoteDraft']['content'], '')
        self.assertEqual(payload['coach']['afterMeetingNoteDraft']['nextAction'], '')
        self.assertEqual(payload['coach']['afterMeetingNoteDraft']['actionType'], 'delivery_schedule')
        self.assertEqual(payload['coach']['mailDraft']['subject'], '')
        self.assertEqual(payload['coach']['mailDraft']['body'], '')
        evidence_hrefs = {item.get('href') for item in payload['coach']['evidence']}
        self.assertIn(f'/schedules/{schedule.id}/', evidence_hrefs)
        self.assertIn(f'/customers/{followup.id}/', evidence_hrefs)
        self.assertEqual(AIWorkspaceQuestionLog.objects.count(), 0)


class PipelineApiTests(TestCase):
    """React 파일럿용 파이프라인 읽기 API 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='파이프라인API회사')
        self.user = make_user('pipeline_api_me', role='salesman', company=self.company)
        self.coworker = make_user('pipeline_api_coworker', role='salesman', company=self.company)
        self.manager = make_user('pipeline_api_manager', role='manager', company=self.company)
        self.url = reverse('reporting:pipeline_command_center_api')
        self.move_url = reverse('reporting:funnel_pipeline_move')

    def _create_pipeline_customer(self, owner, name, stage='quote'):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import Company, Department, FollowUp, History, Quote, Schedule

        customer_company = Company.objects.create(name=f'{name} 회사', created_by=owner)
        department = Department.objects.create(
            company=customer_company,
            name=f'{name} 연구실',
            created_by=owner,
        )
        followup = FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=f'{name} 담당자',
            company=customer_company,
            department=department,
            pipeline_stage=stage,
            customer_grade='A',
        )
        schedule = Schedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate() + timedelta(days=1),
            visit_time=time(10, 0),
            status='scheduled',
            activity_type='quote',
        )
        Quote.objects.create(
            quote_number=f'Q-{name}',
            schedule=schedule,
            followup=followup,
            user=owner,
            valid_until=timezone.localdate() + timedelta(days=30),
            subtotal=1000000,
            probability=65,
            stage='sent',
        )
        History.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            action_type='quote',
            content='견적 후속 필요',
            next_action='견적서 확인 전화',
            next_action_date=timezone.localdate() - timedelta(days=1),
        )
        return followup

    def _create_quote_for_followup(self, followup, owner, suffix, stage, subtotal, converted=False):
        from datetime import timedelta
        from django.utils import timezone
        from reporting.models import Quote

        schedule = followup.schedules.first()
        return Quote.objects.create(
            quote_number=f'Q-{suffix}',
            schedule=schedule,
            followup=followup,
            user=owner,
            valid_until=timezone.localdate() + timedelta(days=30),
            subtotal=subtotal,
            probability=100 if stage in ('approved', 'converted') else 65,
            stage=stage,
            converted_to_delivery=converted,
        )

    def _create_delivery_item(self, schedule, item_name, unit_price, quantity=1):
        from reporting.models import DeliveryItem

        return DeliveryItem.objects.create(
            schedule=schedule,
            item_name=item_name,
            quantity=quantity,
            unit_price=unit_price,
        )

    def _create_history_item(self, history, item_name, unit_price, quantity=1):
        from reporting.models import DeliveryItem

        return DeliveryItem.objects.create(
            history=history,
            item_name=item_name,
            quantity=quantity,
            unit_price=unit_price,
        )

    def _create_delivery_schedule(self, followup, owner, name, unit_price, quantity=1, days_delta=-1):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import Schedule

        schedule = Schedule.objects.create(
            user=owner,
            company=owner.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate() + timedelta(days=days_delta),
            visit_time=time(11, 0),
            status='completed',
            activity_type='delivery',
        )
        self._create_delivery_item(schedule, name, unit_price, quantity)
        return schedule

    def test_pipeline_api_requires_login(self):
        response = self.client.get(self.url)
        self.assertIn(response.status_code, [301, 302, 401])
        if response.status_code in [301, 302]:
            self.assertIn('login', response.get('Location', ''))

    def test_pipeline_api_returns_current_user_scope(self):
        own = self._create_pipeline_customer(self.user, '내고객')
        coworker = self._create_pipeline_customer(self.coworker, '동료고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['source'], 'django')
        deal_ids = {deal['id'] for deal in payload['deals']}
        self.assertIn(own.id, deal_ids)
        self.assertNotIn(coworker.id, deal_ids)

    def test_pipeline_api_includes_metrics_stages_and_tasks(self):
        from django.utils import timezone

        own = self._create_pipeline_customer(self.user, '지표고객')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertGreaterEqual(payload['metrics']['totalPipelineValue'], 1000000)
        self.assertEqual(payload['metrics']['activeCount'], 1)
        self.assertEqual(payload['metrics']['overdueCount'], 1)
        self.assertTrue(any(stage['id'] == own.pipeline_stage for stage in payload['stages']))
        self.assertTrue(any(task['title'] == '견적 후속 지연 고객' for task in payload['priorityTasks']))
        deal = payload['deals'][0]
        self.assertEqual(deal['stageLabel'], '견적 제출')
        self.assertIn('recentActivities', deal)
        self.assertEqual(deal['latestQuote']['amount'], 1100000)
        self.assertEqual(deal['latestQuote']['quoteDate'], timezone.localdate().isoformat())
        self.assertEqual(deal['nextSchedule']['type'], '견적 제출')
        self.assertIn('csrftoken', response.cookies)

    def test_pipeline_api_leaves_unentered_contact_and_potential_probability_null(self):
        from reporting.models import Company, Department, FollowUp

        customer_company = Company.objects.create(name='확률미입력회사', created_by=self.user)
        department = Department.objects.create(
            company=customer_company,
            name='확률미입력연구실',
            created_by=self.user,
        )
        potential = FollowUp.objects.create(
            user=self.user,
            user_company=self.user.userprofile.company,
            customer_name='잠재 고객',
            company=customer_company,
            department=department,
            pipeline_stage='potential',
        )
        contact_company = Company.objects.create(name='접촉확률미입력회사', created_by=self.user)
        contact_department = Department.objects.create(
            company=contact_company,
            name='접촉확률미입력연구실',
            created_by=self.user,
        )
        contact = FollowUp.objects.create(
            user=self.user,
            user_company=self.user.userprofile.company,
            customer_name='접촉 고객',
            company=contact_company,
            department=contact_department,
            pipeline_stage='contact',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        deals = {deal['id']: deal for deal in payload['deals']}
        self.assertIsNone(deals[potential.id]['probability'])
        self.assertIsNone(deals[contact.id]['probability'])
        self.assertEqual(payload['metrics']['weightedPipelineValue'], 0)

    def test_pipeline_api_uses_stage_relevant_quote_amount(self):
        from datetime import timedelta
        from django.utils import timezone

        quote_followup = self._create_pipeline_customer(self.user, '견적가격', stage='quote')
        self._create_delivery_item(quote_followup.schedules.first(), '견적품목', 2000000)
        self._create_quote_for_followup(
            quote_followup, self.user, 'quote-latest-rejected', 'rejected', 3000000
        )
        negotiation_followup = self._create_pipeline_customer(self.user, '협상가격', stage='negotiation')
        self._create_delivery_item(negotiation_followup.schedules.first(), '협상품목', 2000000)
        self._create_quote_for_followup(
            negotiation_followup, self.user, 'negotiation-active', 'negotiation', 2000000
        )
        self._create_quote_for_followup(
            negotiation_followup, self.user, 'negotiation-latest-expired', 'expired', 5000000
        )
        won_followup = self._create_pipeline_customer(self.user, '수주가격', stage='won')
        self._create_delivery_schedule(won_followup, self.user, '납품품목', 4000000)
        self._create_quote_for_followup(
            won_followup, self.user, 'won-approved', 'approved', 4000000
        )
        self._create_quote_for_followup(
            won_followup, self.user, 'won-latest-draft', 'draft', 9000000
        )
        lost_followup = self._create_pipeline_customer(self.user, '실주가격', stage='lost')
        self._create_delivery_item(lost_followup.schedules.first(), '실주견적품목', 1500000)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        deals = {deal['id']: deal for deal in payload['deals']}
        self.assertEqual(deals[quote_followup.id]['value'], 2200000)
        self.assertEqual(deals[quote_followup.id]['latestQuote']['source'], '견적 일정')
        self.assertEqual(deals[quote_followup.id]['latestQuote']['basisType'], 'schedule')
        self.assertEqual(deals[quote_followup.id]['latestQuote']['items'][0]['itemName'], '견적품목')
        self.assertEqual(deals[quote_followup.id]['latestQuote']['items'][0]['totalPrice'], 2200000)
        self.assertEqual(deals[negotiation_followup.id]['value'], 2200000)
        self.assertEqual(deals[negotiation_followup.id]['latestQuote']['source'], '견적 일정')
        self.assertEqual(deals[negotiation_followup.id]['latestQuote']['items'][0]['itemName'], '협상품목')
        self.assertEqual(deals[won_followup.id]['value'], 4400000)
        self.assertEqual(deals[won_followup.id]['latestQuote']['source'], '실제 납품 매출')
        self.assertEqual(deals[won_followup.id]['latestQuote']['basisType'], 'delivery')
        self.assertEqual(deals[won_followup.id]['quoteComparison']['quotedAmount'], 4400000)
        self.assertEqual(deals[won_followup.id]['quoteComparison']['actualAmount'], 4400000)
        self.assertEqual(deals[won_followup.id]['quoteComparison']['deltaAmount'], 0)
        self.assertEqual(deals[won_followup.id]['quoteComparison']['status'], 'match')
        self.assertEqual(deals[lost_followup.id]['value'], 1650000)
        self.assertEqual(deals[lost_followup.id]['latestQuote']['source'], '견적 일정')
        self.assertEqual(deals[lost_followup.id]['latestQuote']['basisType'], 'schedule')
        self.assertEqual(deals[lost_followup.id]['latestQuote']['quoteDate'], (timezone.localdate() + timedelta(days=1)).isoformat())
        stages = {stage['id']: stage for stage in payload['stages']}
        self.assertEqual(stages['won']['totalValue'], 4400000)
        self.assertEqual(stages['lost']['totalValue'], 1650000)

    def test_pipeline_api_uses_latest_quote_schedule_date_amount(self):
        from datetime import time, timedelta
        from django.utils import timezone
        from reporting.models import Schedule

        followup = self._create_pipeline_customer(self.user, '날짜별견적', stage='quote')
        first_schedule = followup.schedules.first()
        self._create_delivery_item(first_schedule, '과거 견적품목', 1000000)
        second_schedule = Schedule.objects.create(
            user=self.user,
            company=self.user.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate() + timedelta(days=2),
            visit_time=time(14, 0),
            status='scheduled',
            activity_type='quote',
        )
        self._create_delivery_item(second_schedule, '최신 견적품목', 2000000)
        same_day_schedule = Schedule.objects.create(
            user=self.user,
            company=self.user.userprofile.company,
            followup=followup,
            visit_date=timezone.localdate() + timedelta(days=2),
            visit_time=time(15, 0),
            status='scheduled',
            activity_type='quote',
        )
        self._create_delivery_item(same_day_schedule, '같은날 견적품목', 500000)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        deal = next(deal for deal in response.json()['deals'] if deal['id'] == followup.id)
        self.assertEqual(deal['value'], 2750000)
        self.assertEqual(deal['latestQuote']['source'], '견적 일정 2건')
        self.assertIn('외 1건', deal['latestQuote']['number'])
        self.assertEqual(deal['latestQuote']['basisDate'], (timezone.localdate() + timedelta(days=2)).isoformat())
        self.assertEqual(deal['latestQuote']['quoteDate'], (timezone.localdate() + timedelta(days=2)).isoformat())

    def test_pipeline_api_sums_all_delivery_dates_amount_for_won(self):
        """계정이 여러 번 납품했으면 최근 1건이 아니라 전부 합산해야 한다."""
        from datetime import timedelta
        from django.utils import timezone

        followup = self._create_pipeline_customer(self.user, '날짜별수주', stage='won')
        self._create_delivery_schedule(followup, self.user, '과거납품', 1000000, days_delta=-20)
        self._create_delivery_schedule(followup, self.user, '최신납품', 2000000, days_delta=-1)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        deal = next(deal for deal in payload['deals'] if deal['id'] == followup.id)
        self.assertEqual(deal['value'], 3300000)
        self.assertEqual(deal['latestQuote']['source'], '실제 납품 매출 2건')
        self.assertEqual(deal['latestQuote']['basisType'], 'delivery')
        self.assertEqual(deal['latestQuote']['basisDate'], (timezone.localdate() - timedelta(days=1)).isoformat())
        self.assertIsNone(deal['latestQuote']['quoteDate'])
        stages = {stage['id']: stage for stage in payload['stages']}
        self.assertEqual(stages['won']['totalValue'], 3300000)

    def test_pipeline_api_uses_quote_history_items_before_quote_model_fallback(self):
        from django.utils import timezone

        followup = self._create_pipeline_customer(self.user, '견적히스토리', stage='quote')
        quote_history = followup.histories.filter(action_type='quote').first()
        self._create_history_item(quote_history, '히스토리견적품목', 3000000)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        deal = next(deal for deal in response.json()['deals'] if deal['id'] == followup.id)
        self.assertEqual(deal['value'], 3300000)
        self.assertEqual(deal['latestQuote']['source'], '견적 활동')
        self.assertEqual(deal['latestQuote']['basisType'], 'history')
        self.assertEqual(deal['latestQuote']['quoteDate'], timezone.localdate().isoformat())
        self.assertEqual(deal['latestQuote']['items'][0]['itemName'], '히스토리견적품목')
        self.assertEqual(deal['latestQuote']['items'][0]['totalPrice'], 3300000)

    def test_pipeline_api_includes_quote_model_items_for_quote_fallback(self):
        from reporting.models import Product, QuoteItem

        followup = self._create_pipeline_customer(self.user, '견적서품목', stage='quote')
        quote = followup.quotes.first()
        product = Product.objects.create(
            product_code='PIPELINE-QUOTE-ITEM',
            unit='BOX',
            standard_price=100000,
            created_by=self.user,
        )
        QuoteItem.objects.create(
            quote=quote,
            product=product,
            quantity=2,
            unit_price=100000,
        )
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        deal = next(deal for deal in response.json()['deals'] if deal['id'] == followup.id)
        self.assertEqual(deal['latestQuote']['basisType'], 'quote')
        self.assertEqual(deal['latestQuote']['items'][0]['itemName'], 'PIPELINE-QUOTE-ITEM')
        self.assertEqual(deal['latestQuote']['items'][0]['quantity'], 2)
        self.assertEqual(deal['latestQuote']['items'][0]['unit'], 'BOX')
        self.assertEqual(deal['latestQuote']['items'][0]['totalPrice'], 200000)

    def test_pipeline_api_uses_delivery_history_items_for_won_revenue(self):
        from django.utils import timezone
        from reporting.models import History

        followup = self._create_pipeline_customer(self.user, '수주히스토리', stage='won')
        delivery_history = History.objects.create(
            user=self.user,
            company=self.user.userprofile.company,
            followup=followup,
            action_type='delivery_schedule',
            content='실제 납품 완료',
            delivery_date=timezone.localdate(),
        )
        self._create_history_item(delivery_history, '히스토리납품품목', 5000000)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        deal = next(deal for deal in response.json()['deals'] if deal['id'] == followup.id)
        self.assertEqual(deal['value'], 5500000)
        self.assertEqual(deal['latestQuote']['source'], '실제 납품 매출')
        self.assertEqual(deal['latestQuote']['basisType'], 'delivery')
        self.assertEqual(deal['quoteComparison']['quotedAmount'], 1100000)
        self.assertEqual(deal['quoteComparison']['actualAmount'], 5500000)
        self.assertEqual(deal['quoteComparison']['deltaAmount'], 4400000)
        self.assertEqual(deal['quoteComparison']['deltaRate'], 400.0)
        self.assertEqual(deal['quoteComparison']['status'], 'over')

    def test_pipeline_api_groups_same_department_contacts_into_account_deal(self):
        from reporting.models import FollowUp

        professor = self._create_pipeline_customer(self.user, '같은연구실교수', stage='quote')
        self._create_delivery_item(professor.schedules.first(), '교수견적품목', 84000, 2)
        researcher = FollowUp.objects.create(
            user=self.user,
            user_company=self.user.userprofile.company,
            customer_name='김종환 연구원',
            company=professor.company,
            department=professor.department,
            pipeline_stage='won',
            customer_grade='A',
        )
        self._create_delivery_schedule(researcher, self.user, '연구원납품품목', 84000, 2)
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        account_deals = [
            deal
            for deal in payload['deals']
            if deal['accountType'] == 'department' and deal['accountId'] == professor.department_id
        ]
        self.assertEqual(len(account_deals), 1)
        deal = account_deals[0]
        self.assertEqual(deal['id'], researcher.id)
        self.assertEqual(deal['stage'], 'won')
        self.assertEqual(deal['value'], 184800)
        self.assertEqual(deal['latestQuote']['source'], '실제 납품 매출')
        self.assertEqual(deal['latestQuote']['basisType'], 'delivery')
        self.assertEqual(deal['quoteComparison']['quotedAmount'], 184800)
        self.assertEqual(deal['quoteComparison']['actualAmount'], 184800)
        self.assertEqual(deal['contactCount'], 2)
        self.assertCountEqual(deal['contactIds'], [professor.id, researcher.id])
        self.assertIn('김종환 연구원', deal['contact'])
        self.assertIn('외 1명', deal['contact'])
        stages = {stage['id']: stage for stage in payload['stages']}
        self.assertEqual(stages['won']['count'], 1)
        self.assertEqual(payload['metrics']['activeCount'], 1)

    def test_pipeline_api_marks_potential_overflow_after_top_ten(self):
        for index in range(12):
            self._create_pipeline_customer(self.user, f'잠재{index}', stage='potential')
        self.client.force_login(self.user)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        potential_deals = [deal for deal in response.json()['deals'] if deal['stage'] == 'potential']
        self.assertEqual(len(potential_deals), 12)
        self.assertEqual(sum(1 for deal in potential_deals if deal['isPotentialOverflow']), 2)
        self.assertTrue(all('attentionScore' in deal for deal in potential_deals))
        self.assertTrue(all('attentionReason' in deal for deal in potential_deals))

    def test_pipeline_move_updates_accessible_followup_stage(self):
        followup = self._create_pipeline_customer(self.user, '이동고객', stage='potential')
        schedule = followup.schedules.first()
        self.client.force_login(self.user)

        response = self.client.post(
            self.move_url,
            data=json.dumps({'followup_id': followup.id, 'stage': 'quote'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'quote')
        self.assertTrue(followup.pipeline_manually_set)
        schedule.refresh_from_db()
        self.assertEqual(schedule.activity_type, 'quote')
        self.assertEqual(schedule.status, 'scheduled')

    def test_pipeline_move_updates_same_department_followup_stages(self):
        from reporting.models import FollowUp

        followup = self._create_pipeline_customer(self.user, '계정이동고객', stage='potential')
        related = FollowUp.objects.create(
            user=self.user,
            user_company=self.user.userprofile.company,
            customer_name='계정이동실무자',
            company=followup.company,
            department=followup.department,
            pipeline_stage='contact',
        )
        self.client.force_login(self.user)

        response = self.client.post(
            self.move_url,
            data=json.dumps({'followup_id': followup.id, 'stage': 'quote'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertEqual(response.json()['updatedCount'], 2)
        followup.refresh_from_db()
        related.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'quote')
        self.assertEqual(related.pipeline_stage, 'quote')
        self.assertTrue(followup.pipeline_manually_set)
        self.assertTrue(related.pipeline_manually_set)

    def test_pipeline_move_rejects_invalid_stage(self):
        followup = self._create_pipeline_customer(self.user, '잘못된단계', stage='potential')
        self.client.force_login(self.user)

        response = self.client.post(
            self.move_url,
            data=json.dumps({'followup_id': followup.id, 'stage': 'invalid'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['success'])
        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'potential')

    def test_pipeline_move_rejects_manager(self):
        followup = self._create_pipeline_customer(self.user, '매니저차단', stage='potential')
        self.client.force_login(self.manager)

        response = self.client.post(
            self.move_url,
            data=json.dumps({'followup_id': followup.id, 'stage': 'quote'}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(response.json()['success'])
        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'potential')


class SchedulePipelineBackfillCommandTests(TestCase):
    """Existing schedule rows can be synced into the pipeline after deployment."""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='일정파이프라인백필회사')
        self.user = make_user('schedule_pipeline_backfill', role='salesman', company=self.company)
        customer_company = Company.objects.create(name='백필거래처', created_by=self.user)
        department = Department.objects.create(
            company=customer_company,
            name='백필연구실',
            created_by=self.user,
        )
        self.followup = FollowUp.objects.create(
            user=self.user,
            user_company=self.company,
            company=customer_company,
            department=department,
            customer_name='백필담당자',
            pipeline_stage='potential',
            pipeline_manually_set=True,
        )
        self.schedule = Schedule.objects.create(
            user=self.user,
            company=self.company,
            followup=self.followup,
            visit_date=timezone.localdate(),
            visit_time=time(14, 0),
            status='completed',
            activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=self.schedule,
            item_name='BACKFILL-WON',
            quantity=1,
            unit='EA',
            unit_price=168000,
        )

    def test_sync_schedule_pipeline_command_dry_run_reports_delivery_won_without_saving(self):
        from io import StringIO

        from django.core.management import call_command

        output = StringIO()

        call_command(
            'sync_schedule_pipeline',
            '--schedule-id',
            str(self.schedule.id),
            '--json',
            stdout=output,
        )

        payload = json.loads(output.getvalue())
        self.assertEqual(payload['mode'], 'dry_run')
        self.assertEqual(payload['changed'], 1)
        self.assertEqual(payload['changes'][0]['fromStage'], 'potential')
        self.assertEqual(payload['changes'][0]['toStage'], 'won')
        self.assertFalse(payload['changes'][0]['applied'])
        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'potential')
        self.assertTrue(self.followup.pipeline_manually_set)

    def test_sync_schedule_pipeline_command_apply_updates_won_value_for_pipeline_api(self):
        from io import StringIO

        from django.core.management import call_command

        output = StringIO()

        call_command(
            'sync_schedule_pipeline',
            '--schedule-id',
            str(self.schedule.id),
            '--apply',
            '--json',
            stdout=output,
        )

        payload = json.loads(output.getvalue())
        self.assertEqual(payload['mode'], 'apply')
        self.assertEqual(payload['changed'], 1)
        self.assertTrue(payload['changes'][0]['applied'])
        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'won')
        self.assertFalse(self.followup.pipeline_manually_set)

        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        self.assertEqual(response.status_code, 200)
        deal = next(item for item in response.json()['deals'] if item['id'] == self.followup.id)
        self.assertEqual(deal['stage'], 'won')
        self.assertEqual(deal['stageLabel'], '수주')
        self.assertEqual(deal['value'], 184800)
        self.assertEqual(deal['latestQuote']['basisType'], 'delivery')


# ─────────────────────────────────────────────────────────────────────────────
# Phase 7: 권한 격리 테스트 (can_access_user_data)
# ─────────────────────────────────────────────────────────────────────────────

class PermissionIsolationTests(TestCase):
    """역할 기반 데이터 격리 검증"""

    def test_can_access_user_data_same_company(self):
        """같은 회사 사용자끼리 접근 가능"""
        from reporting.views import can_access_user_data
        company = UserCompany.objects.create(name='같은회사')
        u1 = make_user('user_a', company=company)
        u2 = make_user('user_b', company=company)
        self.assertTrue(can_access_user_data(u1, u2))

    def test_can_access_user_data_different_company(self):
        """다른 회사 사용자 접근 차단"""
        from reporting.views import can_access_user_data
        c1 = UserCompany.objects.create(name='회사A')
        c2 = UserCompany.objects.create(name='회사B')
        u1 = make_user('user_c', company=c1)
        u2 = make_user('user_d', company=c2)
        self.assertFalse(can_access_user_data(u1, u2))

    def test_can_access_user_data_no_company(self):
        """company=None 사용자 간 상호 접근 차단 (None == None 버그 없음)"""
        from reporting.views import can_access_user_data
        u1 = make_user('user_e')  # company=None
        u2 = make_user('user_f')  # company=None
        # 서로 다른 사용자이고 company가 None → 접근 불가
        self.assertFalse(can_access_user_data(u1, u2))

    def test_can_access_user_data_self(self):
        """자기 자신의 데이터 항상 접근 가능"""
        from reporting.views import can_access_user_data
        u = make_user('user_g')
        self.assertTrue(can_access_user_data(u, u))

    def test_admin_can_access_all(self):
        """admin은 모든 회사 사용자 데이터 접근 가능"""
        from reporting.views import can_access_user_data
        c1 = UserCompany.objects.create(name='회사C')
        c2 = UserCompany.objects.create(name='회사D')
        admin_user = make_user('admin_x', role='admin', company=c1)
        other_user = make_user('other_x', role='salesman', company=c2)
        self.assertTrue(can_access_user_data(admin_user, other_user))

    def test_can_modify_user_data_manager_blocked(self):
        """manager는 타인 데이터 수정 불가"""
        from reporting.views import can_modify_user_data
        company = UserCompany.objects.create(name='수정테스트회사')
        mgr = make_user('mgr_x', role='manager', company=company)
        sales = make_user('sales_x', role='salesman', company=company)
        self.assertFalse(can_modify_user_data(mgr, sales))

    def test_can_modify_user_data_salesman_own(self):
        """salesman은 자기 자신 데이터 수정 가능"""
        from reporting.views import can_modify_user_data
        u = make_user('sales_own')
        self.assertTrue(can_modify_user_data(u, u))

    def test_can_modify_user_data_salesman_other_blocked(self):
        """salesman은 타인 데이터 수정 불가"""
        from reporting.views import can_modify_user_data
        company = UserCompany.objects.create(name='수정테스트회사2')
        u1 = make_user('sales_p', role='salesman', company=company)
        u2 = make_user('sales_q', role='salesman', company=company)
        self.assertFalse(can_modify_user_data(u1, u2))


# ─────────────────────────────────────────────────────────────────────────────
# Manager 역할 권한 검증 테스트
# ─────────────────────────────────────────────────────────────────────────────

class ManagerRolePermissionTests(TestCase):
    """Manager(뷰어)는 영업노트/일정/고객 데이터를 생성·수정할 수 없음을 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='테스트회사')
        self.manager = make_user('mgr_test', role='manager', company=self.company)
        self.salesman = make_user('slm_test', role='salesman', company=self.company)

    # ── 히스토리 생성 차단 (일정 기반) ────────────────────────────────────

    def test_manager_cannot_access_history_create_from_schedule(self):
        """Manager: 일정 기반 히스토리 생성 → 리다이렉트/403 차단"""
        from reporting.models import Company, Department, FollowUp, Schedule
        import datetime
        # 최소 필요 객체 생성
        company = Company.objects.create(name='테스트업체', created_by=self.salesman)
        dept = Department.objects.create(name='테스트부서', company=company, created_by=self.salesman)
        followup = FollowUp.objects.create(
            user=self.salesman, customer_name='테스트고객',
            company=company, department=dept
        )
        schedule = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=datetime.date.today(),
            visit_time=datetime.time(9, 0),
            activity_type='customer_meeting'
        )
        self.client.force_login(self.manager)
        url = reverse('reporting:history_create_from_schedule', args=[schedule.pk])
        r = self.client.get(url)
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager GET history_create_from_schedule: expected redirect/403, got {r.status_code}")

    def test_manager_cannot_post_history_create_from_schedule(self):
        """Manager: 일정 기반 히스토리 생성 POST → 리다이렉트/403 차단"""
        from reporting.models import Company, Department, FollowUp, Schedule
        import datetime
        company = Company.objects.create(name='테스트업체2', created_by=self.salesman)
        dept = Department.objects.create(name='테스트부서2', company=company, created_by=self.salesman)
        followup = FollowUp.objects.create(
            user=self.salesman, customer_name='테스트고객2',
            company=company, department=dept
        )
        schedule = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=datetime.date.today(),
            visit_time=datetime.time(9, 0),
            activity_type='customer_meeting'
        )
        self.client.force_login(self.manager)
        url = reverse('reporting:history_create_from_schedule', args=[schedule.pk])
        r = self.client.post(url, {'action_type': 'customer_meeting'})
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager POST history_create_from_schedule: expected redirect/403, got {r.status_code}")

    # ── 일정 생성 차단 ──────────────────────────────────────────────────────

    def test_manager_cannot_get_schedule_create(self):
        """Manager: 일정 생성 폼 GET → 리다이렉트(차단)"""
        self.client.force_login(self.manager)
        r = self.client.get(reverse('reporting:schedule_create'))
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager GET schedule_create: expected redirect/403, got {r.status_code}")

    def test_manager_cannot_post_schedule_create(self):
        """Manager: 일정 생성 POST → 리다이렉트(차단)"""
        self.client.force_login(self.manager)
        r = self.client.post(reverse('reporting:schedule_create'), {
            'visit_date': '2026-05-01',
            'activity_type': 'customer_meeting',
        })
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager POST schedule_create: expected redirect/403, got {r.status_code}")

    # ── 고객(팔로우업) 생성 차단 ────────────────────────────────────────────

    def test_manager_cannot_get_followup_create(self):
        """Manager: 고객 생성 폼 GET → 리다이렉트(차단)"""
        self.client.force_login(self.manager)
        r = self.client.get(reverse('reporting:followup_create'))
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager GET followup_create: expected redirect/403, got {r.status_code}")

    def test_manager_cannot_post_followup_create(self):
        """Manager: 고객 생성 POST → 리다이렉트(차단)"""
        self.client.force_login(self.manager)
        r = self.client.post(reverse('reporting:followup_create'), {
            'customer_name': '홍길동',
        })
        self.assertIn(r.status_code, [302, 403],
                      msg=f"Manager POST followup_create: expected redirect/403, got {r.status_code}")

    # ── Salesman은 정상 접근 가능 (form 렌더링) ─────────────────────────────

    def test_salesman_can_get_schedule_create(self):
        """Salesman: 일정 생성 폼 GET → React 생성 화면"""
        self.client.force_login(self.salesman)
        r = self.client.get(reverse('reporting:schedule_create'))
        self.assertEqual(r.status_code, 302,
                         msg=f"Salesman GET schedule_create: expected 302, got {r.status_code}")
        self.assertEqual(r['Location'], frontend_url('schedules/?create=1'))

    def test_salesman_can_get_followup_create(self):
        """Salesman: 고객 생성 폼 GET → React 생성 화면"""
        self.client.force_login(self.salesman)
        r = self.client.get(reverse('reporting:followup_create'))
        self.assertEqual(r.status_code, 302,
                         msg=f"Salesman GET followup_create: expected 302, got {r.status_code}")
        self.assertEqual(r['Location'], frontend_url('customers/?create=1'))

    # ── 조회는 허용 ─────────────────────────────────────────────────────────

    def test_manager_can_view_history_list(self):
        """Manager: 히스토리 목록 조회 → React 영업노트"""
        self.client.force_login(self.manager)
        r = self.client.get(reverse('reporting:history_list'))
        self.assertEqual(r.status_code, 302,
                         msg=f"Manager GET history_list: expected 302, got {r.status_code}")
        self.assertEqual(r['Location'], frontend_url('notes/'))

    def test_manager_can_view_schedule_list(self):
        """Manager: 일정 목록 조회 → React 일정"""
        self.client.force_login(self.manager)
        r = self.client.get(reverse('reporting:schedule_list'))
        self.assertEqual(r.status_code, 302,
                         msg=f"Manager GET schedule_list: expected 302, got {r.status_code}")
        self.assertEqual(r['Location'], frontend_url('schedules/'))

    def test_manager_can_view_followup_list(self):
        """Manager: 고객 목록 조회 → React 고객"""
        self.client.force_login(self.manager)
        r = self.client.get(reverse('reporting:followup_list'))
        self.assertEqual(r.status_code, 302,
                         msg=f"Manager GET followup_list: expected 302, got {r.status_code}")
        self.assertEqual(r['Location'], frontend_url('customers/'))


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8: 디버그 엔드포인트 제거 확인 테스트
# ─────────────────────────────────────────────────────────────────────────────

class DebugEndpointTests(TestCase):
    """Phase 8: debug/user-company/ 엔드포인트가 제거되었는지 확인"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='디버그테스트회사')
        self.superuser = User.objects.create_superuser(
            username='superuser_debug', password='TestPass123!'
        )
        self.regular_user = make_user('regular_debug', role='salesman', company=self.company)

    def test_debug_endpoint_does_not_exist(self):
        """debug/user-company/ URL이 URL 설정에 존재하지 않음"""
        from django.urls import NoReverseMatch
        with self.assertRaises(NoReverseMatch):
            reverse('reporting:debug_user_company_info')

    def test_debug_url_returns_404(self):
        """debug/user-company/ 직접 접근 시 404 반환"""
        self.client.force_login(self.superuser)
        r = self.client.get('/reporting/debug/user-company/')
        self.assertEqual(r.status_code, 404,
                         msg=f"debug URL should be 404, got {r.status_code}")

    def test_debug_url_anonymous_returns_404(self):
        """미인증 사용자 debug URL 접근 시 404 반환"""
        r = self.client.get('/reporting/debug/user-company/')
        self.assertEqual(r.status_code, 404,
                         msg=f"anonymous debug URL should be 404, got {r.status_code}")


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8: 파일 업로드 MIME 검증 테스트
# ─────────────────────────────────────────────────────────────────────────────

class FileUploadValidationTests(TestCase):
    """Phase 8: 파일 업로드 MIME 검증 및 확장자 화이트리스트 테스트"""

    def _make_file(self, name, content):
        """테스트용 가짜 InMemoryUploadedFile 생성"""
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile(name, content)

    def test_valid_pdf_accepted(self):
        """올바른 PDF 파일 (매직 바이트 + 확장자 일치) 허용"""
        from reporting.views import validate_file_upload
        f = self._make_file('test.pdf', b'%PDF-1.4 valid pdf content')
        ok, msg = validate_file_upload(f)
        self.assertTrue(ok, msg=f"Valid PDF should be accepted: {msg}")

    def test_valid_jpeg_accepted(self):
        """올바른 JPEG 파일 허용"""
        from reporting.views import validate_file_upload
        f = self._make_file('photo.jpg', b'\xff\xd8\xff\xe0' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertTrue(ok, msg=f"Valid JPEG should be accepted: {msg}")

    def test_valid_png_accepted(self):
        """올바른 PNG 파일 허용"""
        from reporting.views import validate_file_upload
        f = self._make_file('image.png', b'\x89PNG\r\n\x1a\n' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertTrue(ok, msg=f"Valid PNG should be accepted: {msg}")

    def test_invalid_extension_rejected(self):
        """허용되지 않은 확장자 차단"""
        from reporting.views import validate_file_upload
        f = self._make_file('malware.exe', b'MZ\x90\x00' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertFalse(ok, msg="EXE file should be rejected")

    def test_disguised_exe_as_pdf_rejected(self):
        """EXE 파일을 PDF로 위장한 경우 차단 (MIME 스푸핑 방지)"""
        from reporting.views import validate_file_upload
        # .pdf 확장자지만 실제로는 EXE 매직 바이트 MZ
        f = self._make_file('fake.pdf', b'MZ\x90\x00' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertFalse(ok, msg=f"EXE disguised as PDF should be rejected: {msg}")

    def test_disguised_exe_as_jpg_rejected(self):
        """EXE 파일을 JPG로 위장한 경우 차단"""
        from reporting.views import validate_file_upload
        f = self._make_file('photo.jpg', b'MZ\x90\x00' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertFalse(ok, msg=f"EXE disguised as JPG should be rejected: {msg}")

    def test_oversized_file_rejected(self):
        """10MB 초과 파일 차단"""
        from reporting.views import validate_file_upload
        import io
        from django.core.files.uploadedfile import InMemoryUploadedFile
        content = b'%PDF' + b'\x00' * (10 * 1024 * 1024 + 1)
        buf = io.BytesIO(content)
        f = InMemoryUploadedFile(buf, 'file', 'big.pdf', 'application/pdf', len(content), None)
        ok, msg = validate_file_upload(f)
        self.assertFalse(ok, msg="Oversized file should be rejected")

    def test_valid_docx_accepted(self):
        """올바른 DOCX 파일 (ZIP 기반) 허용"""
        from reporting.views import validate_file_upload
        # DOCX는 ZIP 포맷 (PK\x03\x04 시그니처)
        f = self._make_file('report.docx', b'PK\x03\x04' + b'\x00' * 100)
        ok, msg = validate_file_upload(f)
        self.assertTrue(ok, msg=f"Valid DOCX should be accepted: {msg}")


# ─────────────────────────────────────────────────────────────────────────────
# Phase 9: 프로덕션 설정 보안 검증 테스트
# ─────────────────────────────────────────────────────────────────────────────

class ProductionSettingsTests(TestCase):
    """Phase 9: settings_production.py 보안 설정 유효성 검증"""

    def test_session_save_every_request_disabled_by_default(self):
        """React API 조회 요청마다 세션 row를 저장하지 않도록 기본값을 비활성화"""
        from django.conf import settings as django_settings
        self.assertFalse(
            getattr(django_settings, 'SESSION_SAVE_EVERY_REQUEST', False),
            "SESSION_SAVE_EVERY_REQUEST 기본값은 DB write volume을 줄이기 위해 False여야 합니다.",
        )

    def test_allowed_hosts_no_invalid_wildcards(self):
        """ALLOWED_HOSTS에 Django 미지원 와일드카드(*.xxx)가 없음을 확인"""
        from django.conf import settings as django_settings
        for host in django_settings.ALLOWED_HOSTS:
            self.assertFalse(
                host.startswith('*.'),
                f"ALLOWED_HOSTS에 미지원 와일드카드 발견: {host}"
            )

    def test_email_encryption_key_is_bytes_or_none(self):
        """EMAIL_ENCRYPTION_KEY가 bytes 또는 None인지 확인 (문자열 금지)"""
        from django.conf import settings as django_settings
        key = getattr(django_settings, 'EMAIL_ENCRYPTION_KEY', 'NOT_SET')
        if key != 'NOT_SET' and key is not None:
            self.assertIsInstance(
                key, bytes,
                f"EMAIL_ENCRYPTION_KEY는 bytes여야 합니다. 현재 타입: {type(key)}"
            )

    def test_email_encryption_key_not_hardcoded_default(self):
        """EMAIL_ENCRYPTION_KEY가 알려진 하드코딩 기본값이 아님을 확인"""
        from django.conf import settings as django_settings
        key = getattr(django_settings, 'EMAIL_ENCRYPTION_KEY', None)
        # 이전에 사용된 안전하지 않은 공개 기본값
        UNSAFE_FALLBACK = b'YXNkZmFzZGZhc2RmYXNkZmFzZGZhc2RmYXNkZmFzZGY='
        if key is not None:
            self.assertNotEqual(
                key, UNSAFE_FALLBACK,
                "EMAIL_ENCRYPTION_KEY가 알려진 안전하지 않은 기본값으로 설정되어 있습니다."
            )

    def test_hsts_seconds_env_non_negative(self):
        """HSTS_SECONDS 환경변수가 있으면 0 이상인지 확인"""
        import os
        val_str = os.environ.get('HSTS_SECONDS', '0')
        val = int(val_str)
        self.assertGreaterEqual(val, 0, "HSTS_SECONDS는 0 이상이어야 합니다")

    def test_secure_content_type_nosniff(self):
        """프로덕션 환경(not DEBUG)에서 MIME 스니핑 방지 헤더가 활성화됨"""
        from django.conf import settings as django_settings
        if not django_settings.DEBUG:
            self.assertTrue(
                getattr(django_settings, 'SECURE_CONTENT_TYPE_NOSNIFF', False),
                "프로덕션에서 SECURE_CONTENT_TYPE_NOSNIFF가 활성화되어야 합니다"
            )

    def test_secret_key_not_insecure_prefix_in_production(self):
        """RAILWAY_ENVIRONMENT가 설정된 실제 프로덕션에서 django-insecure- 접두어 금지"""
        import os
        from django.conf import settings as django_settings
        # RAILWAY_ENVIRONMENT가 실제로 설정된 경우에만 검증 (로컬 개발 환경 제외)
        if os.environ.get('RAILWAY_ENVIRONMENT'):
            self.assertFalse(
                django_settings.SECRET_KEY.startswith('django-insecure-'),
                "Railway 프로덕션에서 insecure SECRET_KEY(django-insecure- 접두어)를 사용하면 안 됩니다."
            )


class OperationsHealthTests(TestCase):
    """운영 health/readiness endpoint smoke tests."""

    def test_healthz_returns_public_liveness_without_data(self):
        response = self.client.get('/healthz/')

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['service'], 'sales-note-backend')
        self.assertNotIn('customers', payload)
        self.assertIn('no-store', response.headers.get('Cache-Control', ''))

    @override_settings(ALLOWED_HOSTS=['healthcheck.railway.app'])
    def test_healthz_accepts_railway_healthcheck_host(self):
        response = self.client.get('/healthz/', HTTP_HOST='healthcheck.railway.app')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')

    def test_readyz_returns_database_and_migration_status(self):
        response = self.client.get('/readyz/')

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'ok')
        self.assertEqual(payload['checks']['database']['status'], 'ok')
        self.assertEqual(payload['checks']['migrations']['pending'], 0)

    def test_backup_status_does_not_require_email_host_setting(self):
        with patch.object(__import__('django.conf', fromlist=['settings']).settings, 'EMAIL_HOST', None, create=True):
            response = self.client.get('/reporting/backup/status/')

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])


class OperationsCommandTests(TestCase):
    """운영 자동화 management command smoke tests."""

    @override_settings(DEBUG=True)
    def test_audit_runtime_config_outputs_json_without_secret_values(self):
        from io import StringIO
        from django.conf import settings as django_settings
        from django.core.management import call_command

        output = StringIO()
        call_command('audit_runtime_config', '--json', stdout=output)

        payload = json.loads(output.getvalue())
        self.assertIn(payload['status'], ('ok', 'warning'))
        self.assertNotIn(getattr(django_settings, 'SECRET_KEY', ''), output.getvalue())

    def test_simple_backup_json_writes_retained_artifact(self):
        from tempfile import TemporaryDirectory
        from django.core.management import call_command

        with TemporaryDirectory() as temp_dir:
            call_command('simple_backup', '--format=json', f'--output-dir={temp_dir}', '--keep=1')
            files = os.listdir(temp_dir)

        self.assertEqual(len(files), 1)
        self.assertTrue(files[0].startswith('sales_note_backup_'))
        self.assertTrue(files[0].endswith('.json'))


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8.5: 제품 규격/단위 저장 테스트 (Bug 1)
# ─────────────────────────────────────────────────────────────────────────────

class ProductSpecificationSaveTests(TestCase):
    """제품 생성/수정 시 specification 및 unit 필드가 올바르게 저장되는지 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='제품테스트회사')
        self.salesman = make_user('prod_salesman', role='salesman', company=self.company)
        self.client.force_login(self.salesman)

    def test_product_create_saves_specification(self):
        """일반 폼 제출로 제품 생성 시 specification 저장"""
        from reporting.models import Product
        response = self.client.post(
            reverse('reporting:product_create'),
            {
                'product_code': 'TEST-SPEC-001',
                'standard_price': '10000',
                'specification': '100x200mm',
                'unit': 'EA',
                'is_active': 'on',
            },
        )
        # 성공 시 목록으로 리다이렉트
        self.assertIn(response.status_code, [200, 302])
        product = Product.objects.filter(product_code='TEST-SPEC-001').first()
        self.assertIsNotNone(product, '제품이 생성되어야 합니다')
        self.assertEqual(product.specification, '100x200mm',
                         '규격(specification)이 저장되어야 합니다')

    def test_product_create_saves_unit(self):
        """일반 폼 제출로 제품 생성 시 unit 저장"""
        from reporting.models import Product
        self.client.post(
            reverse('reporting:product_create'),
            {
                'product_code': 'TEST-UNIT-001',
                'standard_price': '5000',
                'specification': '',
                'unit': 'SET',
                'is_active': 'on',
            },
        )
        product = Product.objects.filter(product_code='TEST-UNIT-001').first()
        self.assertIsNotNone(product, '제품이 생성되어야 합니다')
        self.assertEqual(product.unit, 'SET', '단위(unit)가 저장되어야 합니다')

    def test_product_edit_saves_specification(self):
        """제품 수정 시 specification 저장"""
        from reporting.models import Product
        product = Product.objects.create(
            product_code='EDIT-SPEC-001',
            standard_price=1000,
            specification='',
            unit='EA',
            created_by=self.salesman,
        )
        self.client.post(
            reverse('reporting:product_edit', args=[product.pk]),
            {
                'product_code': 'EDIT-SPEC-001',
                'standard_price': '1000',
                'specification': '200x300mm',
                'unit': 'EA',
                'is_active': 'on',
            },
        )
        product.refresh_from_db()
        self.assertEqual(product.specification, '200x300mm',
                         '수정된 규격(specification)이 저장되어야 합니다')

    def test_product_edit_saves_unit(self):
        """제품 수정 시 unit 저장"""
        from reporting.models import Product
        product = Product.objects.create(
            product_code='EDIT-UNIT-001',
            standard_price=1000,
            specification='',
            unit='EA',
            created_by=self.salesman,
        )
        self.client.post(
            reverse('reporting:product_edit', args=[product.pk]),
            {
                'product_code': 'EDIT-UNIT-001',
                'standard_price': '1000',
                'specification': '',
                'unit': 'BOX',
                'is_active': 'on',
            },
        )
        product.refresh_from_db()
        self.assertEqual(product.unit, 'BOX',
                         '수정된 단위(unit)가 저장되어야 합니다')

    def test_product_edit_existing_data_preserved(self):
        """제품 수정 시 기존 데이터(가격 등)가 보존됨"""
        from reporting.models import Product
        from decimal import Decimal
        product = Product.objects.create(
            product_code='PRES-001',
            standard_price=Decimal('9999'),
            specification='기존규격',
            unit='EA',
            created_by=self.salesman,
        )
        self.client.post(
            reverse('reporting:product_edit', args=[product.pk]),
            {
                'product_code': 'PRES-001',
                'standard_price': '9999',
                'specification': '새규격',
                'unit': 'EA',
                'is_active': 'on',
            },
        )
        product.refresh_from_db()
        self.assertEqual(product.standard_price, Decimal('9999'),
                         '수정 후 기존 가격이 보존되어야 합니다')
        self.assertEqual(product.specification, '새규격')


class ProductManagementReactApiTests(TestCase):
    """React 제품관리 API 회귀 테스트"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='제품React회사')
        self.salesman = make_user('product-react-sales', role='salesman', company=self.company)
        self.manager = make_user('product-react-manager', role='manager', company=self.company)
        self.client.force_login(self.salesman)

    def _uploaded_products_xlsx(self, rows, filename='products-upload.xlsx'):
        from io import BytesIO
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)
        return SimpleUploadedFile(
            filename,
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_product_bulk_upsert_updates_existing_ecount_overlap_and_disables_promo(self):
        import json
        from decimal import Decimal
        from reporting.models import Product

        existing = Product.objects.create(
            product_code='ECOUNT-001',
            description='기존 설명',
            specification='OLD',
            unit='EA',
            standard_price=Decimal('1000'),
            is_promo=True,
            promo_price=Decimal('800'),
            created_by=self.salesman,
        )
        no_description_row = Product.objects.create(
            product_code='ECOUNT-003',
            description='보존할 설명',
            specification='OLD-SPEC',
            unit='EA',
            standard_price=Decimal('3000'),
            created_by=self.salesman,
        )

        response = self.client.post(
            reverse('reporting:products_bulk_upsert_api'),
            data=json.dumps({
                'products': [
                    {
                        'productCode': 'ECOUNT-001',
                        'description': '새 설명',
                        'specification': 'NEW',
                        'unit': 'BOX',
                        'standardPrice': '1500',
                    },
                    {
                        'productCode': 'ECOUNT-002',
                        'description': '신규 설명',
                        'specification': 'SPEC',
                        'unit': 'SET',
                        'standardPrice': '2500',
                    },
                    {
                        'productCode': 'ECOUNT-003',
                        'specification': 'NEW-SPEC',
                        'unit': 'EA',
                        'standardPrice': '3300',
                    },
                ],
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['createdCount'], 1)
        self.assertEqual(payload['updatedCount'], 2)
        existing.refresh_from_db()
        self.assertEqual(existing.description, '새 설명')
        self.assertEqual(existing.specification, 'NEW')
        self.assertEqual(existing.unit, 'BOX')
        self.assertEqual(existing.standard_price, Decimal('1500'))
        self.assertFalse(existing.is_promo)
        self.assertIsNone(existing.promo_price)
        no_description_row.refresh_from_db()
        self.assertEqual(no_description_row.description, '보존할 설명')
        self.assertEqual(no_description_row.specification, 'NEW-SPEC')
        self.assertEqual(no_description_row.standard_price, Decimal('3300'))
        created = Product.objects.get(product_code='ECOUNT-002')
        self.assertEqual(created.created_by, self.salesman)

    def test_product_current_price_ignores_legacy_promotion(self):
        from decimal import Decimal
        from reporting.models import Product

        product = Product.objects.create(
            product_code='PROMO-OFF-001',
            standard_price=Decimal('1000'),
            is_promo=True,
            promo_price=Decimal('700'),
            created_by=self.salesman,
        )

        response = self.client.get(reverse('reporting:product_api_list'), {'search': 'PROMO-OFF'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()['products'][0]
        self.assertEqual(payload['current_price'], 1000.0)
        self.assertFalse(payload['is_promo'])
        self.assertEqual(product.get_current_price(), Decimal('1000'))

    def test_product_api_list_supports_limit_and_specification_search(self):
        from reporting.models import Product

        for index in range(5):
            Product.objects.create(
                product_code=f'LIMIT-SEARCH-{index}',
                specification='공통규격',
                standard_price=1000 + index,
                created_by=self.salesman,
            )
        Product.objects.create(
            product_code='LIMIT-SPEC-ONLY',
            specification='특수규격',
            standard_price=2000,
            created_by=self.salesman,
        )

        response = self.client.get(reverse('reporting:product_api_list'), {'search': '특수규격', 'limit': '2'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['count'], 1)
        self.assertEqual(payload['totalCount'], 1)
        self.assertFalse(payload['hasMore'])
        self.assertEqual(payload['products'][0]['product_code'], 'LIMIT-SPEC-ONLY')

        response = self.client.get(reverse('reporting:product_api_list'), {'search': 'LIMIT-SEARCH', 'limit': '2'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['count'], 2)
        self.assertEqual(payload['totalCount'], 5)
        self.assertTrue(payload['hasMore'])

    def test_product_management_manager_is_read_only_for_company_products(self):
        import json
        from reporting.models import Product

        product = Product.objects.create(
            product_code='MANAGER-READONLY-001',
            description='manager visible product',
            standard_price=1000,
            created_by=self.salesman,
        )
        other_company = UserCompany.objects.create(name='제품React타사회사')
        other_user = make_user('product-react-other-sales', role='salesman', company=other_company)
        Product.objects.create(
            product_code='MANAGER-READONLY-OTHER',
            standard_price=2000,
            created_by=other_user,
        )
        self.client.force_login(self.manager)

        response = self.client.get(reverse('reporting:products_management_api'), {'q': 'MANAGER-READONLY'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload['scope']['canManage'])
        self.assertEqual(payload['links']['save'], '')
        self.assertEqual(payload['links']['bulkUpsert'], '')
        self.assertEqual(payload['links']['bulkDelete'], '')
        self.assertEqual(payload['links']['excelImport'], '')
        product_codes = {item['productCode'] for item in payload['products']}
        self.assertIn(product.product_code, product_codes)
        self.assertNotIn('MANAGER-READONLY-OTHER', product_codes)
        visible = next(item for item in payload['products'] if item['productCode'] == product.product_code)
        self.assertEqual(visible['createdBy'], self.salesman.username)

        save_response = self.client.post(
            reverse('reporting:product_save_api'),
            data=json.dumps({
                'productCode': 'MANAGER-READONLY-NEW',
                'standardPrice': '3000',
            }),
            content_type='application/json',
        )
        self.assertEqual(save_response.status_code, 403)

        edit_response = self.client.post(
            reverse('reporting:product_update_api', args=[product.id]),
            data=json.dumps({
                'productCode': product.product_code,
                'standardPrice': '9999',
            }),
            content_type='application/json',
        )
        self.assertEqual(edit_response.status_code, 403)

        bulk_response = self.client.post(
            reverse('reporting:products_bulk_upsert_api'),
            data=json.dumps({'products': [{'productCode': 'MANAGER-BULK', 'standardPrice': '100'}]}),
            content_type='application/json',
        )
        self.assertEqual(bulk_response.status_code, 403)

        delete_response = self.client.post(
            reverse('reporting:products_bulk_delete_api'),
            data=json.dumps({'productCodes': [product.product_code]}),
            content_type='application/json',
        )
        self.assertEqual(delete_response.status_code, 403)
        product.refresh_from_db()
        self.assertEqual(product.standard_price, 1000)

        import_response = self.client.post(
            reverse('reporting:products_excel_import_api'),
            {'file': self._uploaded_products_xlsx([['품번', '기준단가'], ['MANAGER-XLSX', 1000]])},
        )
        self.assertEqual(import_response.status_code, 403)

    def test_product_management_includes_selected_product_outside_first_page(self):
        from reporting.models import Product

        for index in range(55):
            Product.objects.create(
                product_code=f'PAGE-PRODUCT-{index:03d}',
                standard_price=1000 + index,
                created_by=self.salesman,
            )
        selected = Product.objects.create(
            product_code='ZZZ-SELECTED-LEGACY',
            standard_price=9999,
            created_by=self.salesman,
        )

        response = self.client.get(reverse('reporting:products_management_api'), {
            'page_size': '10',
            'selected_product_id': selected.id,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['products'][0]['id'], selected.id)
        self.assertIn('excelImport', payload['links'])

    def test_products_excel_import_creates_and_updates_rows(self):
        from decimal import Decimal
        from reporting.models import Product

        existing = Product.objects.create(
            product_code='XLSX-UP-001',
            description='기존',
            specification='OLD',
            unit='EA',
            standard_price=Decimal('1000'),
            created_by=self.salesman,
        )
        upload = self._uploaded_products_xlsx([
            ['품번', '제품설명', '규격', '단위', '기준단가', '상태'],
            ['XLSX-UP-001', '수정 설명', 'NEW', 'BOX', 1500, '활성'],
            ['XLSX-UP-002', '신규 설명', 'SPEC', 'SET', 2500, '비활성'],
        ])

        response = self.client.post(reverse('reporting:products_excel_import_api'), {'file': upload})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['createdCount'], 1)
        self.assertEqual(payload['updatedCount'], 1)
        self.assertEqual(payload['errorCount'], 0)
        existing.refresh_from_db()
        self.assertEqual(existing.description, '수정 설명')
        self.assertEqual(existing.specification, 'NEW')
        self.assertEqual(existing.unit, 'BOX')
        self.assertEqual(existing.standard_price, Decimal('1500'))
        created = Product.objects.get(product_code='XLSX-UP-002')
        self.assertEqual(created.created_by, self.salesman)
        self.assertFalse(created.is_active)

    def test_product_bulk_delete_deletes_unused_and_blocks_used_product(self):
        import json
        from reporting.models import DeliveryItem, Product

        unused = Product.objects.create(product_code='DELETE-UNUSED', standard_price=1000, created_by=self.salesman)
        used = Product.objects.create(product_code='DELETE-USED', standard_price=2000, created_by=self.salesman)
        DeliveryItem.objects.create(product=used, item_name='사용 제품', quantity=1, unit_price=2000)

        response = self.client.post(
            reverse('reporting:products_bulk_delete_api'),
            data=json.dumps({'productCodes': [unused.product_code, used.product_code, 'DELETE-MISSING']}),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['deletedCount'], 1)
        self.assertEqual(payload['blockedCount'], 1)
        self.assertEqual(payload['missingCount'], 1)
        blocked = next(item for item in payload['results'] if item['productCode'] == used.product_code)
        self.assertTrue(blocked['canReplace'])
        self.assertEqual(blocked['deliveryItemCount'], 1)
        self.assertEqual(blocked['referenceCount'], 1)
        self.assertEqual(blocked['references'][0]['referenceType'], 'deliveryItem')
        self.assertEqual(blocked['references'][0]['itemName'], used.product_code)
        self.assertFalse(Product.objects.filter(product_code=unused.product_code).exists())
        self.assertTrue(Product.objects.filter(product_code=used.product_code).exists())

    def test_product_delete_replaces_used_items_one_reference_at_a_time(self):
        import datetime
        import json
        from django.utils import timezone
        from reporting.models import Company, DeliveryItem, Department, FollowUp, Product, Quote, QuoteItem, Schedule

        old = Product.objects.create(
            product_code='DELETE-REPLACE-OLD',
            unit='EA',
            standard_price=2000,
            created_by=self.salesman,
        )
        replacement = Product.objects.create(
            product_code='DELETE-REPLACE-NEW',
            unit='SET',
            standard_price=3000,
            created_by=self.salesman,
        )
        company = Company.objects.create(name='제품대체 고객사', created_by=self.salesman)
        department = Department.objects.create(company=company, name='제품대체 연구실', created_by=self.salesman)
        followup = FollowUp.objects.create(
            user=self.salesman,
            user_company=self.company,
            customer_name='제품대체 담당자',
            company=company,
            department=department,
        )
        schedule = Schedule.objects.create(
            user=self.salesman,
            company=self.company,
            followup=followup,
            visit_date=timezone.localdate(),
            visit_time=datetime.time(10, 0),
            activity_type='quote',
            status='scheduled',
        )
        delivery_item = DeliveryItem.objects.create(
            schedule=schedule,
            product=old,
            item_name=old.product_code,
            quantity=2,
            unit_price=2000,
        )
        quote = Quote.objects.create(
            quote_number='Q-REPLACE-001',
            schedule=schedule,
            followup=followup,
            user=self.salesman,
            valid_until=timezone.localdate() + datetime.timedelta(days=30),
        )
        quote_item = QuoteItem.objects.create(
            quote=quote,
            product=old,
            quantity=1,
            unit_price=2000,
        )

        blocked_response = self.client.post(
            reverse('reporting:products_bulk_delete_api'),
            data=json.dumps({
                'productCodes': [old.product_code],
            }),
            content_type='application/json',
        )

        self.assertEqual(blocked_response.status_code, 200)
        blocked_payload = blocked_response.json()
        self.assertEqual(blocked_payload['deletedCount'], 0)
        self.assertEqual(blocked_payload['blockedCount'], 1)
        blocked = blocked_payload['results'][0]
        self.assertEqual(blocked['referenceCount'], 2)
        reference_ids = {(item['referenceType'], item['referenceId']) for item in blocked['references']}
        self.assertEqual(reference_ids, {('deliveryItem', delivery_item.id), ('quoteItem', quote_item.id)})

        response = self.client.post(
            reverse('reporting:product_replace_reference_api'),
            data=json.dumps({
                'productCode': old.product_code,
                'referenceType': 'deliveryItem',
                'referenceId': delivery_item.id,
                'replacementProductId': replacement.id,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertFalse(payload['deletedOriginal'])
        self.assertEqual(payload['replacementProductCode'], replacement.product_code)
        self.assertEqual(payload['result']['status'], 'blocked')
        self.assertEqual(payload['result']['referenceCount'], 1)
        delivery_item.refresh_from_db()
        self.assertEqual(delivery_item.product, replacement)
        self.assertEqual(delivery_item.item_name, replacement.product_code)
        self.assertEqual(delivery_item.unit, 'SET')

        response = self.client.post(
            reverse('reporting:product_replace_reference_api'),
            data=json.dumps({
                'productCode': old.product_code,
                'referenceType': 'quoteItem',
                'referenceId': quote_item.id,
                'replacementProductId': replacement.id,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertTrue(payload['deletedOriginal'])
        self.assertEqual(payload['result']['status'], 'deleted')
        self.assertFalse(Product.objects.filter(product_code=old.product_code).exists())
        quote_item.refresh_from_db()
        self.assertEqual(quote_item.product, replacement)

    def test_product_replace_history_delivery_item_updates_history_summary(self):
        import json
        from reporting.models import Company, DeliveryItem, Department, FollowUp, History, Product

        old = Product.objects.create(
            product_code='DELETE-HISTORY-OLD',
            unit='EA',
            standard_price=2000,
            created_by=self.salesman,
        )
        replacement = Product.objects.create(
            product_code='DELETE-HISTORY-NEW',
            unit='BOX',
            standard_price=3000,
            created_by=self.salesman,
        )
        company = Company.objects.create(name='히스토리대체 고객사', created_by=self.salesman)
        department = Department.objects.create(company=company, name='히스토리대체 부서', created_by=self.salesman)
        followup = FollowUp.objects.create(
            user=self.salesman,
            user_company=self.company,
            customer_name='히스토리대체 담당자',
            company=company,
            department=department,
        )
        history = History.objects.create(
            user=self.salesman,
            company=self.company,
            followup=followup,
            action_type='delivery_schedule',
            content='납품 기록',
        )
        delivery_item = DeliveryItem.objects.create(
            history=history,
            product=old,
            item_name=old.product_code,
            quantity=3,
            unit_price=2000,
        )

        blocked_response = self.client.post(
            reverse('reporting:products_bulk_delete_api'),
            data=json.dumps({'productCodes': [old.product_code]}),
            content_type='application/json',
        )

        self.assertEqual(blocked_response.status_code, 200)
        blocked = blocked_response.json()['results'][0]
        self.assertEqual(blocked['references'][0]['historyId'], history.id)

        response = self.client.post(
            reverse('reporting:product_replace_reference_api'),
            data=json.dumps({
                'productCode': old.product_code,
                'referenceType': 'deliveryItem',
                'referenceId': delivery_item.id,
                'replacementProductId': replacement.id,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['deletedOriginal'])
        self.assertFalse(Product.objects.filter(product_code=old.product_code).exists())
        delivery_item.refresh_from_db()
        history.refresh_from_db()
        self.assertEqual(delivery_item.product, replacement)
        self.assertEqual(delivery_item.item_name, replacement.product_code)
        self.assertEqual(delivery_item.unit, 'BOX')
        self.assertIn(replacement.product_code, history.delivery_items)
        self.assertEqual(int(history.delivery_amount), 6600)

    def test_products_excel_export_returns_xlsx(self):
        from reporting.models import Product

        Product.objects.create(product_code='XLSX-001', description='다운로드', standard_price=1000, created_by=self.salesman)

        response = self.client.get(reverse('reporting:products_excel_export_api'))

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertIn('products-', response['Content-Disposition'])


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8.5: 대시보드 일정 표시 테스트 (Bug 2 & 3)
# ─────────────────────────────────────────────────────────────────────────────

class DashboardScheduleDisplayTests(TestCase):
    """React dashboard API 일정 표시 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='대시보드일정테스트회사')
        self.salesman = make_user('dash_sched_user', role='salesman', company=self.company)
        self.client.force_login(self.salesman)

        from reporting.models import Company, Department, FollowUp, Schedule
        from django.utils import timezone
        from datetime import timedelta
        import datetime

        company = Company.objects.create(name='테스트고객사', created_by=self.salesman)
        dept = Department.objects.create(name='테스트부서', company=company, created_by=self.salesman)
        followup = FollowUp.objects.create(
            user=self.salesman, customer_name='테스트담당자',
            company=company, department=dept,
        )
        today = timezone.localdate()

        # 오늘 예정 일정
        self.today_scheduled = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today, visit_time=datetime.time(9, 0),
            activity_type='customer_meeting', status='scheduled',
        )
        # 오늘 완료된 일정
        self.today_completed = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today, visit_time=datetime.time(14, 0),
            activity_type='customer_meeting', status='completed',
        )
        # 내일 예정 일정
        self.tomorrow_scheduled = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today + timedelta(days=1), visit_time=datetime.time(10, 0),
            activity_type='customer_meeting', status='scheduled',
        )
        # 3일 후 완료 일정
        self.upcoming_completed = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today + timedelta(days=3), visit_time=datetime.time(11, 0),
            activity_type='customer_meeting', status='completed',
        )
        # 8일 후 (범위 밖) 일정
        self.out_of_range = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today + timedelta(days=8), visit_time=datetime.time(9, 0),
            activity_type='customer_meeting', status='scheduled',
        )
        # 어제 일정 (과거 - upcoming에 포함 안 됨)
        self.yesterday = Schedule.objects.create(
            user=self.salesman, followup=followup,
            visit_date=today - timedelta(days=1), visit_time=datetime.time(9, 0),
            activity_type='customer_meeting', status='scheduled',
        )

    def test_dashboard_returns_200(self):
        """대시보드 API 200 응답"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        self.assertEqual(r.status_code, 200)

    def test_today_schedules_includes_scheduled(self):
        """today.items에 오늘 예정 일정 포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['today']['items']]
        self.assertIn(self.today_scheduled.pk, ids,
                      '오늘 예정 일정이 today.items에 포함되어야 합니다')

    def test_today_schedules_includes_completed(self):
        """today.items에 오늘 완료된 일정도 포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['today']['items']]
        self.assertIn(self.today_completed.pk, ids,
                      '오늘 완료된 일정도 today.items에 포함되어야 합니다')

    def test_upcoming_includes_tomorrow_scheduled(self):
        """upcomingSchedules에 내일 예정 일정 포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['upcomingSchedules']]
        self.assertIn(self.tomorrow_scheduled.pk, ids,
                      '내일 예정 일정이 upcomingSchedules에 포함되어야 합니다')

    def test_upcoming_includes_completed_within_range(self):
        """upcomingSchedules에 이번 주 완료된 일정도 포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['upcomingSchedules']]
        self.assertIn(self.upcoming_completed.pk, ids,
                      '이번 주 완료된 일정도 upcomingSchedules에 포함되어야 합니다')

    def test_upcoming_excludes_out_of_range(self):
        """upcomingSchedules에 6일 초과 일정은 미포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['upcomingSchedules']]
        self.assertNotIn(self.out_of_range.pk, ids,
                         '6일 초과 일정은 upcomingSchedules에 포함되지 않아야 합니다')

    def test_upcoming_excludes_past_schedules(self):
        """upcomingSchedules에 과거 일정 미포함"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        ids = [s['id'] for s in r.json()['upcomingSchedules']]
        self.assertNotIn(self.yesterday.pk, ids,
                         '어제 일정은 upcomingSchedules에 포함되지 않아야 합니다')

    def test_schedule_count_nonzero_when_schedules_exist(self):
        """일정이 있을 때 API 일정 지표가 0이 아님"""
        r = self.client.get(reverse('reporting:dashboard_summary_api'))
        metrics = r.json()['metrics']
        self.assertGreater(int(metrics['todaySchedules']) + int(metrics['weeklySchedules']), 0)


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8.6-1: 세금계산서 요청 API 테스트
# ─────────────────────────────────────────────────────────────────────────────

class TaxInvoiceRequestAPITests(TestCase):
    """세금계산서 요청 API (followup_tax_invoices_api / tax_invoice_update_status_api) 테스트"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='세금계산서테스트회사')
        self.salesman = make_user('taxinv_salesman', role='salesman', company=self.company)
        self.manager = make_user('taxinv_manager', role='manager', company=self.company)
        self.other_salesman = make_user('taxinv_other', role='salesman', company=self.company)
        self.client.force_login(self.salesman)

        from reporting.models import Company, Department, FollowUp, Schedule
        import datetime

        cust_company = Company.objects.create(name='세금계산서고객사', created_by=self.salesman)
        dept = Department.objects.create(name='세금계산서부서', company=cust_company, created_by=self.salesman)
        self.followup = FollowUp.objects.create(
            user=self.salesman, customer_name='테스트담당자',
            company=cust_company, department=dept,
        )
        self.delivery_schedule = Schedule.objects.create(
            user=self.salesman, followup=self.followup,
            visit_date=datetime.date(2026, 6, 1),
            visit_time=datetime.time(10, 0),
            activity_type='delivery', status='completed',
        )

    def _url_list(self):
        return reverse('reporting:followup_tax_invoices_api',
                       kwargs={'followup_id': self.followup.pk})

    def _url_status(self, req_id):
        return reverse('reporting:tax_invoice_update_status_api',
                       kwargs={'request_id': req_id})

    # ── GET: 목록 조회 ─────────────────────────────────────────────────────

    def test_get_list_success(self):
        """로그인한 영업사원이 GET 요청 시 200 + success=True 반환"""
        r = self.client.get(self._url_list())
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data.get('success'), 'success 필드가 True여야 합니다')
        self.assertIn('tax_invoices', data)
        self.assertIn('delivery_schedules', data)

    def test_get_list_requires_login(self):
        """비로그인 상태에서 GET 요청 시 리다이렉트 또는 403"""
        self.client.logout()
        r = self.client.get(self._url_list())
        self.assertIn(r.status_code, [302, 403],
                      '비로그인 요청은 302 또는 403이어야 합니다')

    def test_get_list_other_company_blocked(self):
        """다른 회사 사용자는 403 반환"""
        other_company = UserCompany.objects.create(name='다른회사')
        outsider = make_user('taxinv_outsider', role='salesman', company=other_company)
        self.client.force_login(outsider)
        r = self.client.get(self._url_list())
        self.assertEqual(r.status_code, 403)

    # ── POST: 요청 생성 ────────────────────────────────────────────────────

    def test_post_create_request_removed(self):
        """세금계산서 요청 생성은 외상고객 메뉴 전환 후 차단"""
        r = self.client.post(self._url_list(), {
            'schedule_id': self.delivery_schedule.pk,
            'memo': '발행 부탁드립니다',
        })
        self.assertEqual(r.status_code, 410)
        data = r.json()
        self.assertFalse(data.get('success'))
        self.assertEqual(data.get('redirect'), '/receivables/')

    def test_post_create_duplicate_blocked_by_removed_flow(self):
        """기존 요청 여부와 관계없이 신규 요청 생성은 차단"""
        from reporting.models import TaxInvoiceRequest
        TaxInvoiceRequest.objects.create(
            followup=self.followup,
            schedule=self.delivery_schedule,
            status='requested',
            requested_by=self.salesman,
        )
        r = self.client.post(self._url_list(), {
            'schedule_id': self.delivery_schedule.pk,
        })
        self.assertEqual(r.status_code, 410)

    def test_post_without_schedule_removed(self):
        """일정 없는 세금계산서 요청 생성도 차단"""
        r = self.client.post(self._url_list(), {'memo': '일정 없는 요청'})
        self.assertEqual(r.status_code, 410)

    # ── 상태 변경: 발행완료 ────────────────────────────────────────────────

    def test_salesman_cannot_issue(self):
        """구 세금계산서 상태 변경 API는 차단"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        r = self.client.post(self._url_status(req.pk), {'status': 'issued'})
        self.assertEqual(r.status_code, 410)

    def test_manager_issue_removed(self):
        """매니저도 구 세금계산서 상태 변경 API를 사용할 수 없음"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        self.client.force_login(self.manager)
        r = self.client.post(self._url_status(req.pk), {
            'status': 'issued', 'memo': '발행 처리함'
        })
        self.assertEqual(r.status_code, 410)

    # ── 상태 변경: 취소 ───────────────────────────────────────────────────

    def test_requester_cancel_removed(self):
        """요청자 본인도 구 세금계산서 요청 취소 API를 사용할 수 없음"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        r = self.client.post(self._url_status(req.pk), {
            'status': 'cancelled', 'memo': '취소 사유'
        })
        self.assertEqual(r.status_code, 410)

    def test_other_salesman_cannot_cancel_others_request(self):
        """다른 영업사원은 타인의 요청을 취소 불가 → 403"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        self.client.force_login(self.other_salesman)
        r = self.client.post(self._url_status(req.pk), {
            'status': 'cancelled', 'memo': '취소'
        })
        self.assertEqual(r.status_code, 410)

    # ── 보류 처리 ─────────────────────────────────────────────────────────

    def test_manager_set_on_hold_removed(self):
        """구 세금계산서 보류 API는 차단"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        self.client.force_login(self.manager)
        r = self.client.post(self._url_status(req.pk), {
            'status': 'on_hold', 'memo': '검토 중'
        })
        self.assertEqual(r.status_code, 410)

    # ── 잘못된 상태값 ─────────────────────────────────────────────────────

    def test_invalid_status_value_returns_400(self):
        """올바르지 않은 status 값은 400 반환"""
        from reporting.models import TaxInvoiceRequest
        req = TaxInvoiceRequest.objects.create(
            followup=self.followup, status='requested',
            requested_by=self.salesman,
        )
        self.client.force_login(self.manager)
        r = self.client.post(self._url_status(req.pk), {'status': 'INVALID'})
        self.assertEqual(r.status_code, 410)

    # ── 404 처리 ─────────────────────────────────────────────────────────

    def test_nonexistent_followup_returns_404(self):
        """존재하지 않는 followup_id는 404 반환"""
        r = self.client.get(
            reverse('reporting:followup_tax_invoices_api',
                    kwargs={'followup_id': 99999})
        )
        self.assertEqual(r.status_code, 404)

    def test_nonexistent_request_id_returns_404(self):
        """존재하지 않는 request_id는 404 반환"""
        self.client.force_login(self.manager)
        r = self.client.post(
            reverse('reporting:tax_invoice_update_status_api',
                    kwargs={'request_id': 99999}),
            {'status': 'issued'},
        )
        self.assertEqual(r.status_code, 410)


# ─────────────────────────────────────────────────────────────────────────────
# [재현] 대시보드 통합 검색 API 테스트
# ─────────────────────────────────────────────────────────────────────────────

class DashboardSearchAPITests(TestCase):
    """dashboard_search_api 통합 검색 API 테스트"""

    def setUp(self):
        import datetime
        self.client = Client()
        # 자사 (검색 허용 범위)
        self.company_uc = UserCompany.objects.create(name='검색테스트회사')
        self.salesman = make_user('ds_salesman', role='salesman', company=self.company_uc)
        self.client.force_login(self.salesman)

        # 타사 (검색 제외 범위)
        self.other_uc = UserCompany.objects.create(name='타사회사')
        self.other_user = make_user('ds_other', role='salesman', company=self.other_uc)

        from reporting.models import Company, Department, FollowUp, Schedule, History, DeliveryItem

        # ── 자사 거래처/연구실/담당자 ──────────────────────────────────────
        cust = Company.objects.create(name='검색한국대학교', created_by=self.salesman)
        self.dept = Department.objects.create(name='PCR연구실', company=cust, created_by=self.salesman)
        self.followup = FollowUp.objects.create(
            user=self.salesman,
            customer_name='김연구원',
            company=cust,
            department=self.dept,
            user_company=self.company_uc,
        )

        # 납품 품목 (DeliveryItem)
        sched = Schedule.objects.create(
            user=self.salesman, followup=self.followup,
            visit_date=datetime.date(2026, 5, 1),
            visit_time=datetime.time(10, 0),
            activity_type='delivery', status='completed',
            notes='PCR 실험 관련 방문',
        )
        DeliveryItem.objects.create(
            schedule=sched,
            item_name='PCR 시약 키트',
            quantity=10, unit_price=5000, total_price=50000,
        )
        # 활동 내역
        History.objects.create(
            user=self.salesman, followup=self.followup,
            action_type='customer_meeting',
            content='PCR 장비 데모 진행',
        )

        # ── 타사 거래처/연구실 (검색 제외 확인용) ──────────────────────────
        other_cust = Company.objects.create(name='타사학교', created_by=self.other_user)
        other_dept = Department.objects.create(name='PCR타부서', company=other_cust, created_by=self.other_user)
        other_fu = FollowUp.objects.create(
            user=self.other_user,
            customer_name='박타사',
            company=other_cust,
            department=other_dept,
            user_company=self.other_uc,
        )

    def _url(self):
        return reverse('reporting:dashboard_search_api')

    def test_requires_login(self):
        """비로그인 시 로그인 페이지로 리다이렉트."""
        self.client.logout()
        r = self.client.get(self._url(), {'q': 'PCR'})
        self.assertIn(r.status_code, [302, 401, 403])

    def test_short_query_returns_400(self):
        """1자 검색어는 400 에러를 반환."""
        r = self.client.get(self._url(), {'q': 'P'})
        self.assertEqual(r.status_code, 400)
        data = r.json()
        self.assertIn('error', data)

    def test_empty_query_returns_400(self):
        """빈 검색어는 400 에러를 반환."""
        r = self.client.get(self._url(), {'q': ''})
        self.assertEqual(r.status_code, 400)

    def test_keyword_finds_delivery_item(self):
        """납품 품목명 키워드로 연구실을 찾는다."""
        r = self.client.get(self._url(), {'q': 'PCR 시약'})
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data['success'])
        dept_names = [d['department_name'] for d in data['departments']]
        self.assertIn('PCR연구실', dept_names)

    def test_keyword_finds_history_content(self):
        """활동 내용 키워드로 연구실을 찾는다."""
        r = self.client.get(self._url(), {'q': 'PCR 장비'})
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data['success'])
        dept_names = [d['department_name'] for d in data['departments']]
        self.assertIn('PCR연구실', dept_names)

    def test_other_company_excluded(self):
        """타사 연구실은 검색 결과에 포함되지 않는다."""
        r = self.client.get(self._url(), {'q': 'PCR타부서'})
        self.assertEqual(r.status_code, 200)
        data = r.json()
        dept_names = [d['department_name'] for d in data['departments']]
        self.assertNotIn('PCR타부서', dept_names)

    def test_no_match_returns_empty_list(self):
        """매칭 없으면 빈 리스트 반환."""
        r = self.client.get(self._url(), {'q': '없는키워드XYZ9999'})
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertTrue(data['success'])
        self.assertEqual(data['result_count'], 0)
        self.assertEqual(data['departments'], [])


# ─────────────────────────────────────────────────────────────────────────────
# Phase 8.6-2: 부가세 모드 (VAT Mode) 테스트
# ─────────────────────────────────────────────────────────────────────────────

class ScheduleVatModeTests(TestCase):
    """Schedule.vat_mode 필드 및 Quote.save() 부가세 계산 테스트"""

    def setUp(self):
        import datetime
        from decimal import Decimal
        from reporting.models import UserCompany, Company, Department, FollowUp, Schedule, Quote

        self.company_uc = UserCompany.objects.create(name='VAT테스트회사')
        self.salesman = make_user('vat_salesman', role='salesman', company=self.company_uc)
        cust = Company.objects.create(name='VAT테스트거래처', created_by=self.salesman)
        dept = Department.objects.create(name='VAT테스트연구실', company=cust, created_by=self.salesman)
        self.followup = FollowUp.objects.create(
            user=self.salesman,
            customer_name='VAT담당자',
            company=cust,
            department=dept,
            user_company=self.company_uc,
        )
        self.base_schedule_date = datetime.date(2026, 6, 1)
        self.base_schedule_time = datetime.time(10, 0)

    def _make_schedule(self, vat_mode='excluded'):
        from reporting.models import Schedule
        import datetime
        return Schedule.objects.create(
            user=self.salesman,
            followup=self.followup,
            visit_date=self.base_schedule_date,
            visit_time=self.base_schedule_time,
            activity_type='quote',
            status='scheduled',
            vat_mode=vat_mode,
        )

    def _make_quote(self, schedule, subtotal_val, probability=50):
        import datetime
        from decimal import Decimal
        from reporting.models import Quote
        return Quote.objects.create(
            quote_number=f'Q-TEST-{schedule.pk}-{subtotal_val}',
            schedule=schedule,
            followup=self.followup,
            user=self.salesman,
            valid_until=self.base_schedule_date,
            subtotal=subtotal_val,
            probability=probability,
        )

    def test_default_vat_mode_is_excluded(self):
        """vat_mode 기본값은 'excluded'이어야 한다."""
        schedule = self._make_schedule()
        self.assertEqual(schedule.vat_mode, 'excluded')

    def test_vat_excluded_calculation(self):
        """부가세 별도: tax = subtotal * 10%, total = subtotal + tax."""
        from decimal import Decimal
        schedule = self._make_schedule(vat_mode='excluded')
        quote = self._make_quote(schedule, subtotal_val=100000)
        self.assertEqual(quote.tax_amount, Decimal('10000'))
        self.assertEqual(quote.total_amount, Decimal('110000'))

    def test_vat_included_calculation(self):
        """부가세 포함: total = subtotal(입력값), tax = total - total/1.1."""
        from decimal import Decimal, ROUND_HALF_UP
        schedule = self._make_schedule(vat_mode='included')
        quote = self._make_quote(schedule, subtotal_val=110000)
        supply = (Decimal('110000') / Decimal('1.1')).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        expected_tax = Decimal('110000') - supply
        self.assertEqual(quote.total_amount, Decimal('110000'))
        self.assertEqual(quote.tax_amount, expected_tax)

    def test_vat_none_calculation(self):
        """부가세 없음: tax = 0, total = subtotal."""
        from decimal import Decimal
        schedule = self._make_schedule(vat_mode='none')
        quote = self._make_quote(schedule, subtotal_val=100000)
        self.assertEqual(quote.tax_amount, Decimal('0'))
        self.assertEqual(quote.total_amount, Decimal('100000'))

    def test_vat_excluded_weighted_revenue(self):
        """부가세 별도: 가중매출 = total * probability / 100."""
        from decimal import Decimal
        schedule = self._make_schedule(vat_mode='excluded')
        quote = self._make_quote(schedule, subtotal_val=100000, probability=50)
        self.assertEqual(quote.weighted_revenue, Decimal('55000'))  # 110000 * 0.5

    def test_quote_probability_is_normalized_to_five_percent_step(self):
        """견적 성공 확률은 저장 시 가까운 5% 단위로 정규화된다."""
        from decimal import Decimal

        schedule = self._make_schedule(vat_mode='excluded')
        quote = self._make_quote(schedule, subtotal_val=100000, probability=63)

        self.assertEqual(quote.probability, 65)
        self.assertEqual(quote.weighted_revenue, Decimal('71500'))  # 110000 * 0.65

    def test_vat_none_weighted_revenue(self):
        """부가세 없음: 가중매출 = total * probability / 100."""
        from decimal import Decimal
        schedule = self._make_schedule(vat_mode='none')
        quote = self._make_quote(schedule, subtotal_val=100000, probability=50)
        self.assertEqual(quote.weighted_revenue, Decimal('50000'))  # 100000 * 0.5

    def test_schedule_form_includes_vat_mode(self):
        """스케줄 생성 시 vat_mode가 POST 데이터로 저장된다."""
        import datetime
        from reporting.models import Schedule
        self.client = Client()
        self.client.force_login(self.salesman)

        post_data = {
            'followup': self.followup.pk,
            'visit_date': '2026-06-10',
            'visit_time': '10:00',
            'activity_type': 'quote',
            'location': '',
            'status': 'scheduled',
            'notes': 'VAT 모드 테스트',
            'vat_mode': 'none',
        }
        r = self.client.post(reverse('reporting:schedule_create'), post_data, follow=False)
        # 성공 시 리다이렉트
        self.assertIn(r.status_code, [200, 302])
        created = Schedule.objects.filter(
            followup=self.followup, vat_mode='none'
        ).first()
        self.assertIsNotNone(created, 'vat_mode=none인 스케줄이 생성되어야 합니다.')


class WriteProxyPhase0HardeningTests(TestCase):
    """쓰기 프록시 노출 전 하드닝(Phase 0) 회귀 테스트.

    - won 보호: 견적 일정 취소 시그널이 수주(won) 기회를 강등하지 않음
    - department_memo_api: 교차 테넌트 접근 차단
    - department_assign_category: 연락처 없는 부서의 무검사 갭 차단
    - personal_schedule_add_comment: 비소유자 댓글 차단
    """

    def setUp(self):
        self.client = Client()
        self.user_company = UserCompany.objects.create(name='하드닝회사')
        self.owner = make_user('phase0_owner', role='salesman', company=self.user_company)
        self.other = make_user('phase0_other', role='salesman', company=self.user_company)
        self.other_company = UserCompany.objects.create(name='다른회사')
        self.outsider = make_user('phase0_outsider', role='salesman', company=self.other_company)

        self.company = Company.objects.create(name='하드닝고객사', created_by=self.owner)
        self.department = Department.objects.create(
            name='하드닝부서', company=self.company, created_by=self.owner,
        )
        self.followup = FollowUp.objects.create(
            user=self.owner, customer_name='하드닝고객',
            company=self.company, department=self.department,
        )

    def _make_quote_schedule_with_opportunity(self, stage):
        from reporting.models import OpportunityTracking
        opportunity = OpportunityTracking.objects.create(
            followup=self.followup, title='하드닝 기회', current_stage=stage,
        )
        schedule = Schedule.objects.create(
            user=self.owner,
            followup=self.followup,
            visit_date=timezone.localdate(),
            visit_time=time(10, 0),
            status='scheduled',
            activity_type='quote',
            opportunity=opportunity,
        )
        return schedule, opportunity

    # ── #1 won 보호 시그널 ──────────────────────────────────────────
    def test_won_opportunity_not_demoted_when_quote_schedule_cancelled(self):
        schedule, opportunity = self._make_quote_schedule_with_opportunity('won')
        schedule.status = 'cancelled'
        schedule.save()
        opportunity.refresh_from_db()
        self.assertEqual(
            opportunity.current_stage, 'won',
            '수주(won) 기회는 견적 취소로 강등되면 안 된다.',
        )

    def test_open_quote_opportunity_still_demoted_on_cancel(self):
        # 양성 대조: 수주 전 견적 기회는 취소 시 정상적으로 quote_lost로 전환
        schedule, opportunity = self._make_quote_schedule_with_opportunity('quote')
        schedule.status = 'cancelled'
        schedule.save()
        opportunity.refresh_from_db()
        self.assertEqual(opportunity.current_stage, 'quote_lost')

    # ── #2 부서 메모 교차 테넌트 차단 ───────────────────────────────
    def test_department_memo_api_blocks_outsider(self):
        from reporting.models import DepartmentMemo
        self.client.force_login(self.outsider)
        url = reverse('reporting:department_memo_api', args=[self.department.id])
        r = self.client.post(url, {'content': '침입 메모'})
        self.assertEqual(r.status_code, 403)
        self.assertFalse(
            DepartmentMemo.objects.filter(department=self.department).exists()
        )

    def test_department_memo_api_allows_owner(self):
        from reporting.models import DepartmentMemo
        self.client.force_login(self.owner)
        url = reverse('reporting:department_memo_api', args=[self.department.id])
        r = self.client.post(url, {'content': '정상 메모'})
        self.assertEqual(r.status_code, 200)
        self.assertTrue(
            DepartmentMemo.objects.filter(
                department=self.department, content='정상 메모',
            ).exists()
        )

    # ── #3 연락처 없는 부서 카테고리 할당 갭 ─────────────────────────
    def test_department_assign_category_blocks_outsider_on_empty_department(self):
        empty_dept = Department.objects.create(
            name='빈부서', company=self.company, created_by=self.owner,
        )
        self.client.force_login(self.outsider)
        url = reverse('reporting:department_assign_category', args=[empty_dept.id])
        r = self.client.post(url, {'category_id': ''})
        self.assertFalse(r.json().get('success'))

    # ── #4 개인 일정 댓글 권한 ──────────────────────────────────────
    def test_personal_schedule_add_comment_blocks_non_owner(self):
        from reporting.models import PersonalSchedule
        ps = PersonalSchedule.objects.create(
            user=self.owner, title='비공개 일정',
            schedule_date=timezone.localdate(), schedule_time=time(9, 0),
        )
        self.client.force_login(self.other)  # 같은 회사지만 salesman(전체 열람 불가)
        url = reverse('reporting:personal_schedule_add_comment', args=[ps.id])
        r = self.client.post(url, {'content': '남의 일정 댓글'})
        self.assertEqual(r.status_code, 403)

    def test_personal_schedule_add_comment_allows_owner(self):
        from reporting.models import PersonalSchedule
        ps = PersonalSchedule.objects.create(
            user=self.owner, title='내 일정',
            schedule_date=timezone.localdate(), schedule_time=time(9, 0),
        )
        self.client.force_login(self.owner)
        url = reverse('reporting:personal_schedule_add_comment', args=[ps.id])
        r = self.client.post(url, {'content': '내 일정 댓글'})
        self.assertEqual(r.status_code, 200)


class WriteBearerAuthTests(TestCase):
    """Phase 1 쓰기 토큰 인증 인프라(`reporting/write_api.py` + WriteBearerMiddleware).

    - 유효 토큰 + 화이트리스트 POST → acting 유저로 인증, CSRF 우회(이 요청만)
    - 세션 요청은 CSRF 그대로 강제, 토큰 미설정/오류/비허용 url은 차단
    - acting 유저는 비-staff·비-admin 이어야 함 (권한 붕괴 방지)
    """

    TOKEN = 'sekret-write-123'

    def setUp(self):
        self.user_company = UserCompany.objects.create(name='쓰기회사')
        self.write_user = make_user('write_actor', role='salesman', company=self.user_company)
        self.admin_user = make_user('write_admin', role='admin', company=self.user_company)
        self.staff_user = make_user('write_staff', role='salesman', company=self.user_company)
        self.staff_user.is_staff = True
        self.staff_user.save(update_fields=['is_staff'])

        self.company = Company.objects.create(name='쓰기고객사', created_by=self.write_user)
        self.department = Department.objects.create(
            name='쓰기부서', company=self.company, created_by=self.write_user,
        )
        self.followup = FollowUp.objects.create(
            user=self.write_user, customer_name='쓰기고객',
            company=self.company, department=self.department,
        )

    def _env(self, token=None, user_id=None):
        return {
            'SALES_NOTE_WRITE_TOKEN': self.TOKEN if token is None else token,
            'SALES_NOTE_WRITE_USER_ID': str(self.write_user.id) if user_id is None else str(user_id),
        }

    def _make_schedule(self):
        return Schedule.objects.create(
            user=self.write_user, followup=self.followup,
            visit_date=timezone.localdate(), visit_time=time(9, 0),
            activity_type='customer_meeting',
        )

    def _post_request(self, url, token=None):
        from django.test import RequestFactory
        bearer = self.TOKEN if token is None else token
        return RequestFactory().post(url, HTTP_AUTHORIZATION=f'Bearer {bearer}')

    # ── authenticate_write_bearer 단위 검증 ────────────────────────────
    def test_valid_token_on_allowlisted_url_authenticates(self):
        from reporting.write_api import authenticate_write_bearer
        req = self._post_request(reverse('reporting:schedules_create_api'))
        with patch.dict(os.environ, self._env()):
            self.assertTrue(authenticate_write_bearer(req))
        self.assertEqual(req.user, self.write_user)
        self.assertTrue(getattr(req, 'salesnote_write_api', False))

    def test_denied_url_rejected(self):
        # deny 세트(로그인)는 토큰으로도 인증되지 않는다.
        from reporting.write_api import authenticate_write_bearer
        req = self._post_request(reverse('reporting:login'))
        with patch.dict(os.environ, self._env()):
            self.assertFalse(authenticate_write_bearer(req))

    def test_allowed_non_denied_url_authenticates(self):
        # deny 가 아닌 임의 쓰기(notes_update_api)는 "전부 쓰기"라 인증된다.
        from reporting.write_api import authenticate_write_bearer
        req = self._post_request(reverse('reporting:notes_update_api', args=[999999]))
        with patch.dict(os.environ, self._env()):
            self.assertTrue(authenticate_write_bearer(req))

    def test_get_method_rejected(self):
        from django.test import RequestFactory
        from reporting.write_api import authenticate_write_bearer
        req = RequestFactory().get(
            reverse('reporting:schedules_create_api'),
            HTTP_AUTHORIZATION=f'Bearer {self.TOKEN}',
        )
        with patch.dict(os.environ, self._env()):
            self.assertFalse(authenticate_write_bearer(req))

    def test_wrong_token_rejected(self):
        from reporting.write_api import authenticate_write_bearer
        req = self._post_request(reverse('reporting:schedules_create_api'), token='WRONG')
        with patch.dict(os.environ, self._env()):
            self.assertFalse(authenticate_write_bearer(req))

    def test_unset_env_token_rejected(self):
        from reporting.write_api import authenticate_write_bearer
        req = self._post_request(reverse('reporting:schedules_create_api'))
        with patch.dict(os.environ, self._env(token='')):
            self.assertFalse(authenticate_write_bearer(req))

    # ── get_write_api_user 안전 검증 ───────────────────────────────────
    def test_write_user_resolves_salesman(self):
        from reporting.write_api import get_write_api_user
        with patch.dict(os.environ, self._env()):
            self.assertEqual(get_write_api_user(), self.write_user)

    def test_write_user_refuses_admin_role(self):
        from reporting.write_api import get_write_api_user
        with patch.dict(os.environ, self._env(user_id=self.admin_user.id)):
            self.assertIsNone(get_write_api_user())

    def test_write_user_refuses_staff(self):
        from reporting.write_api import get_write_api_user
        with patch.dict(os.environ, self._env(user_id=self.staff_user.id)):
            self.assertIsNone(get_write_api_user())

    # ── 미들웨어 + CSRF 통합 검증 ──────────────────────────────────────
    def test_write_token_moves_schedule_without_session_or_csrf(self):
        schedule = self._make_schedule()
        client = Client(enforce_csrf_checks=True)
        url = reverse('reporting:schedule_move_api', args=[schedule.pk])
        with patch.dict(os.environ, self._env()):
            r = client.post(url, {'new_date': '2026-08-15'},
                            HTTP_AUTHORIZATION=f'Bearer {self.TOKEN}')
        self.assertEqual(r.status_code, 200)
        schedule.refresh_from_db()
        self.assertEqual(schedule.visit_date.isoformat(), '2026-08-15')

    def test_session_post_still_requires_csrf(self):
        # 세션 요청은 쓰기 토큰 우회를 받지 않고 CSRF 가 그대로 강제된다.
        schedule = self._make_schedule()
        client = Client(enforce_csrf_checks=True)
        client.force_login(self.write_user)
        url = reverse('reporting:schedule_move_api', args=[schedule.pk])
        r = client.post(url, {'new_date': '2026-08-15'})  # CSRF 토큰 없음, bearer 없음
        self.assertEqual(r.status_code, 403)
        schedule.refresh_from_db()
        self.assertNotEqual(schedule.visit_date.isoformat(), '2026-08-15')

    def test_unset_env_token_leaves_write_path_inert(self):
        schedule = self._make_schedule()
        client = Client(enforce_csrf_checks=True)
        url = reverse('reporting:schedule_move_api', args=[schedule.pk])
        with patch.dict(os.environ, self._env(token='')):
            r = client.post(url, {'new_date': '2026-08-15'},
                            HTTP_AUTHORIZATION='Bearer anything')
        self.assertIn(r.status_code, (302, 401, 403))
        schedule.refresh_from_db()
        self.assertNotEqual(schedule.visit_date.isoformat(), '2026-08-15')

    def test_write_token_cannot_reach_denied_endpoint(self):
        # deny 세트(로그인 POST)는 토큰이 무시되어 정상 처리되지 않는다.
        client = Client(enforce_csrf_checks=True)
        url = reverse('reporting:login')
        with patch.dict(os.environ, self._env()):
            r = client.post(url, {'username': 'x', 'password': 'y'},
                            HTTP_AUTHORIZATION=f'Bearer {self.TOKEN}')
        self.assertNotEqual(r.status_code, 200)

    def test_destructive_action_requires_confirm_header(self):
        # 확인 필요 액션(납품품목 교체)에 확인 헤더 없이 토큰 → 428, 뷰 미실행.
        client = Client(enforce_csrf_checks=True)
        url = reverse('reporting:schedules_delivery_items_update_api', args=[999999])
        with patch.dict(os.environ, self._env()):
            r = client.post(url, '{}', content_type='application/json',
                            HTTP_AUTHORIZATION=f'Bearer {self.TOKEN}')
        self.assertEqual(r.status_code, 428)

    def test_destructive_action_passes_gate_with_confirm_header(self):
        # 확인 헤더가 있으면 게이트 통과 → 뷰 실행(없는 일정이라 428 은 아님).
        client = Client(enforce_csrf_checks=True)
        url = reverse('reporting:schedules_delivery_items_update_api', args=[999999])
        with patch.dict(os.environ, self._env()):
            r = client.post(url, '{}', content_type='application/json',
                            HTTP_AUTHORIZATION=f'Bearer {self.TOKEN}',
                            HTTP_X_SALESNOTE_WRITE_CONFIRM='yes')
        self.assertNotEqual(r.status_code, 428)


class PipelineHideCardTests(TestCase):
    """파이프라인 카드 숨김/복원(보드에서만 제거, 데이터 보존)."""

    def setUp(self):
        self.client = Client()
        self.company_uc = UserCompany.objects.create(name='파이프회사')
        self.owner = make_user('pipe_owner', role='salesman', company=self.company_uc)
        self.manager = make_user('pipe_mgr', role='manager', company=self.company_uc)
        self.company = Company.objects.create(name='파이프고객사', created_by=self.owner)
        self.department = Department.objects.create(
            name='파이프부서', company=self.company, created_by=self.owner,
        )
        self.followup = FollowUp.objects.create(
            user=self.owner, customer_name='파이프고객',
            company=self.company, department=self.department,
        )

    def _pipeline(self):
        self.client.force_login(self.owner)
        data = self.client.get(reverse('reporting:pipeline_command_center_api')).json()
        return (
            [d['id'] for d in data.get('deals', [])],
            [h['id'] for h in data.get('hiddenDeals', [])],
        )

    def _post(self, name):
        return self.client.post(
            reverse(name),
            data=json.dumps({'followup_id': self.followup.id}),
            content_type='application/json',
        )

    def test_hide_removes_from_board_and_unhide_restores(self):
        deals, hidden = self._pipeline()
        self.assertIn(self.followup.id, deals)
        self.assertNotIn(self.followup.id, hidden)

        self.client.force_login(self.owner)
        self.assertEqual(self._post('reporting:funnel_pipeline_hide').status_code, 200)
        self.followup.refresh_from_db()
        self.assertTrue(self.followup.pipeline_hidden)
        deals, hidden = self._pipeline()
        self.assertNotIn(self.followup.id, deals)
        self.assertIn(self.followup.id, hidden)

        self.client.force_login(self.owner)
        self.assertEqual(self._post('reporting:funnel_pipeline_unhide').status_code, 200)
        self.followup.refresh_from_db()
        self.assertFalse(self.followup.pipeline_hidden)
        deals, hidden = self._pipeline()
        self.assertIn(self.followup.id, deals)
        self.assertNotIn(self.followup.id, hidden)

    def test_hide_preserves_customer_record(self):
        self.client.force_login(self.owner)
        self._post('reporting:funnel_pipeline_hide')
        # 숨겨도 고객 레코드는 그대로 존재
        self.assertTrue(FollowUp.objects.filter(pk=self.followup.id).exists())

    def test_manager_cannot_hide(self):
        self.client.force_login(self.manager)
        self.assertEqual(self._post('reporting:funnel_pipeline_hide').status_code, 403)

    def test_inactive_contact_does_not_dominate_account_stage(self):
        # 같은 계정: 활성 견적 연락처 + 비활성 수주 연락처 → 계정 단계는 '견적'(수주가 지배 X).
        from datetime import time
        dept = Department.objects.create(
            name='단계부서', company=self.company, created_by=self.owner,
        )
        active_quote = FollowUp.objects.create(
            user=self.owner, customer_name='활성견적', company=self.company, department=dept,
            pipeline_stage='quote', is_active=True,
        )
        # 견적 단계 카드가 근거 없이 0원이면 이제 보드에서 빠지므로, 이 테스트의
        # 실제 목적(비활성 연락처가 단계를 지배하지 않는지)을 검증하려면 근거를 심어야 한다.
        Schedule.objects.create(
            user=self.owner, company=self.company_uc, followup=active_quote,
            visit_date=timezone.localdate(), visit_time=time(10, 0),
            status='completed', activity_type='quote', expected_revenue=1000000,
        )
        FollowUp.objects.create(
            user=self.owner, customer_name='비활성수주', company=self.company, department=dept,
            pipeline_stage='won', is_active=False,
        )
        self.client.force_login(self.owner)
        data = self.client.get(reverse('reporting:pipeline_command_center_api')).json()
        account_deal = next((d for d in data['deals'] if d['id'] == active_quote.id), None)
        self.assertIsNotNone(account_deal, '활성 견적 연락처가 대표 카드여야 한다')
        self.assertEqual(account_deal['stage'], 'quote')



# ─────────────────────────────────────────────────────────────────────────────
# 파이프라인 시트 API 테스트 (주간 활동)
# ─────────────────────────────────────────────────────────────────────────────

class PipelineSheetApiTests(TestCase):
    """파이프라인 시트: 계정별 주간 활동 검증"""

    def setUp(self):
        from datetime import timedelta
        self.client = Client()
        self.company = UserCompany.objects.create(name='시트API회사')
        self.user = make_user('sheet_api_me', role='salesman', company=self.company)
        self.other = make_user('sheet_api_other', role='salesman', company=self.company)
        self.weekly_url = reverse('reporting:pipeline_sheet_weekly_api')
        # 지난 주(월~금)를 기준 주로 삼는다 — 시트 기본값과 동일.
        today = timezone.localdate()
        this_monday = today - timedelta(days=today.weekday())
        self.week_start = this_monday - timedelta(days=7)
        self.week_end = self.week_start + timedelta(days=4)

    def _account(self, owner, name, stage='quote'):
        from reporting.models import Company, Department, FollowUp
        customer_company = Company.objects.create(name=name + '업체', created_by=owner)
        department = Department.objects.create(
            company=customer_company, name=name + '연구실', created_by=owner,
        )
        return FollowUp.objects.create(
            user=owner,
            user_company=owner.userprofile.company,
            customer_name=name + '담당자',
            company=customer_company,
            department=department,
            pipeline_stage=stage,
        )

    def test_weekly_api_requires_login_json(self):
        response = self.client.get(self.weekly_url)
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json()['error'], 'login_required')

    def test_weekly_api_groups_activities_by_account_within_the_week(self):
        from datetime import timedelta
        from reporting.models import History

        followup = self._account(self.user, '주간활동')
        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='customer_meeting', meeting_date=self.week_start,
            meeting_situation='신규 장비 예산 확보 중',
            meeting_obstacles='예산 3월 확정',
            next_action='견적서 재발행',
            next_action_date=self.week_end + timedelta(days=3),
        )
        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_end, content='경쟁사 방문했다고 함',
        )
        # 기준 주 밖의 활동 — 잡히면 안 됨
        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_start - timedelta(days=10),
            content='범위 밖 메모',
        )
        self.client.force_login(self.user)

        response = self.client.get(self.weekly_url, {'week': self.week_start.isoformat()})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['week']['start'], self.week_start.isoformat())
        self.assertEqual(payload['week']['end'], self.week_end.isoformat())
        row = next(r for r in payload['rows'] if r['department'].startswith('주간활동'))
        self.assertEqual(row['activityCount'], 2)
        bodies = [a['body'] for a in row['activities']]
        self.assertIn('신규 장비 예산 확보 중', bodies)
        self.assertNotIn('범위 밖 메모', bodies)
        # 오늘 상황이 있으면 그것을, 없으면 content를 본문으로 쓴다.
        self.assertIn('경쟁사 방문했다고 함', bodies)
        self.assertTrue(row['hasObstacle'])
        self.assertEqual(row['nextAction'], '견적서 재발행')

    def test_weekly_api_reports_quote_and_delivery_amount_for_the_week(self):
        """미접촉 계정 자리를 대신하는 이번 주 견적/납품 금액 합계."""
        from datetime import time
        from reporting.models import DeliveryItem, Schedule

        followup = self._account(self.user, '주간금액')
        Schedule.objects.create(
            user=self.user, company=self.company, followup=followup,
            visit_date=self.week_start, visit_time=time(10, 0),
            status='completed', activity_type='quote', expected_revenue=3000000,
        )
        delivery_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=followup,
            visit_date=self.week_end, visit_time=time(11, 0),
            status='completed', activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=delivery_schedule, item_name='장비', quantity=1, unit_price=1000000,
        )
        self.client.force_login(self.user)

        response = self.client.get(self.weekly_url, {'week': self.week_start.isoformat()})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['metrics']['quoteAmount'], 3000000)
        self.assertEqual(payload['metrics']['deliveryAmount'], 1100000)
        self.assertNotIn('untouchedCount', payload['metrics'])
        self.assertNotIn('untouchedAmount', payload['metrics'])
        self.assertNotIn('untouchedAccounts', payload)

    def test_weekly_api_scopes_to_own_accounts_for_salesman(self):
        from reporting.models import History
        mine = self._account(self.user, '내계정')
        theirs = self._account(self.other, '남의계정')
        for followup, owner in ((mine, self.user), (theirs, self.other)):
            History.objects.create(
                user=owner, company=self.company, followup=followup,
                action_type='memo', meeting_date=self.week_start, content='주간 메모',
            )
        self.client.force_login(self.user)

        payload = self.client.get(
            self.weekly_url, {'week': self.week_start.isoformat()}
        ).json()

        departments = [r['department'] for r in payload['rows']]
        self.assertTrue(any(d.startswith('내계정') for d in departments))
        self.assertFalse(any(d.startswith('남의계정') for d in departments))

    # ------------------------------------------------------------ 엑셀

    def test_export_requires_login(self):
        response = self.client.get(reverse('reporting:pipeline_sheet_export_api'))
        self.assertEqual(response.status_code, 401)

    def test_export_writes_one_row_per_activity_and_all_sheets(self):
        """엑셀은 화면 payload를 그대로 쓴다 — 활동 1건이 한 행."""
        import io as _io
        from openpyxl import load_workbook
        from reporting.models import History

        followup = self._account(self.user, '엑셀')
        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='customer_meeting', meeting_date=self.week_start,
            meeting_situation='데모 요청', meeting_obstacles='예산 미확정',
        )
        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_end, content='재방문 약속',
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse('reporting:pipeline_sheet_export_api'),
            {'week': self.week_start.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('attachment', response['Content-Disposition'])
        wb = load_workbook(_io.BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ['주간 활동', '다운로드 정보'])
        weekly_ws = wb['주간 활동']
        bodies = [row[8] for row in weekly_ws.iter_rows(min_row=2, values_only=True)]
        self.assertIn('데모 요청', bodies)
        self.assertIn('재방문 약속', bodies)

    # ------------------------------------------------------------ 인라인 수정

    def _update_activity(self, kind, activity_id, patch):
        return self.client.post(
            reverse('reporting:pipeline_sheet_activity_update_api', args=[kind, activity_id]),
            data=json.dumps(patch),
            content_type='application/json',
        )

    def test_activity_update_requires_login(self):
        response = self._update_activity('history', 1, {'body': '내용'})
        self.assertEqual(response.status_code, 401)

    def test_activity_update_rejects_invalid_kind(self):
        self.client.force_login(self.user)
        response = self._update_activity('quote', 1, {'body': '내용'})
        self.assertEqual(response.status_code, 400)

    def test_activity_update_edits_history_body_obstacle_and_next_action(self):
        from reporting.models import History

        followup = self._account(self.user, '인라인수정')
        history = History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_start, content='원래 메모',
        )
        self.client.force_login(self.user)

        response = self._update_activity('history', history.id, {
            'body': '수정된 메모',
            'obstacle': '예산 미확정',
            'nextAction': '견적 재발행',
            'nextActionDate': (self.week_end + timedelta(days=2)).isoformat(),
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['activity']['body'], '수정된 메모')
        self.assertEqual(payload['activity']['obstacle'], '예산 미확정')
        self.assertEqual(payload['activity']['nextAction'], '견적 재발행')
        history.refresh_from_db()
        self.assertEqual(history.content, '수정된 메모')
        self.assertEqual(history.meeting_obstacles, '예산 미확정')
        self.assertEqual(history.next_action, '견적 재발행')
        self.assertEqual(history.next_action_date, self.week_end + timedelta(days=2))

    def test_activity_update_writes_meeting_situation_for_customer_meeting(self):
        """`오늘 상황`이 현재 표시 필드면, 같은 필드에 계속 써야 내용이 갈라지지 않는다."""
        from reporting.models import History

        followup = self._account(self.user, '미팅수정')
        history = History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='customer_meeting', meeting_date=self.week_start,
            meeting_situation='기존 상황',
        )
        self.client.force_login(self.user)

        response = self._update_activity('history', history.id, {'body': '새 상황'})

        self.assertEqual(response.status_code, 200)
        history.refresh_from_db()
        self.assertEqual(history.meeting_situation, '새 상황')
        self.assertEqual(history.content, None)

    def test_activity_update_clears_next_action_date_when_blank(self):
        from reporting.models import History

        followup = self._account(self.user, '예정일삭제')
        history = History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_start, content='메모',
            next_action='할 일', next_action_date=self.week_end,
        )
        self.client.force_login(self.user)

        response = self._update_activity('history', history.id, {
            'nextAction': '', 'nextActionDate': '',
        })

        self.assertEqual(response.status_code, 200)
        history.refresh_from_db()
        self.assertEqual(history.next_action, '')
        self.assertIsNone(history.next_action_date)

    def test_activity_update_edits_schedule_notes(self):
        from datetime import time
        from reporting.models import Schedule

        followup = self._account(self.user, '일정메모수정')
        schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=followup,
            visit_date=self.week_start, visit_time=time(10, 0),
            status='scheduled', activity_type='customer_meeting',
            notes='자동 생성: 영업노트 후속 미팅',
        )
        self.client.force_login(self.user)

        response = self._update_activity('schedule', schedule.id, {'body': '방문 전 준비물 확인'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['activity']['body'], '방문 전 준비물 확인')
        schedule.refresh_from_db()
        self.assertEqual(schedule.notes, '방문 전 준비물 확인')

    def test_activity_update_blocks_editing_other_users_history(self):
        from reporting.models import History

        followup = self._account(self.other, '남의기록')
        history = History.objects.create(
            user=self.other, company=self.company, followup=followup,
            action_type='memo', meeting_date=self.week_start, content='남의 메모',
        )
        self.client.force_login(self.user)

        response = self._update_activity('history', history.id, {'body': '건드리면 안 됨'})

        self.assertEqual(response.status_code, 403)
        history.refresh_from_db()
        self.assertEqual(history.content, '남의 메모')


# ─────────────────────────────────────────────────────────────────────────────
# 납품 → 파이프라인 '수주' 자동 반영 검증
# ─────────────────────────────────────────────────────────────────────────────

class DeliveryPipelineSyncTests(TestCase):
    """납품이 확인되면 카드가 자동으로 '수주'로 옮겨지는지, 여러 건이면 합산되는지 검증"""

    def setUp(self):
        self.client = Client()
        self.company = UserCompany.objects.create(name='납품동기화회사')
        self.user = make_user('delivery_sync_me', role='salesman', company=self.company)

    def _account(self, name, department_name=None, stage='quote'):
        from reporting.models import Company, Department, FollowUp
        customer_company = Company.objects.create(name=name + '업체', created_by=self.user)
        department = Department.objects.create(
            company=customer_company, name=department_name or (name + '연구실'),
            created_by=self.user,
        )
        return FollowUp.objects.create(
            user=self.user, user_company=self.company,
            customer_name=name + '담당자', company=customer_company,
            department=department, pipeline_stage=stage,
        )

    def test_standalone_delivery_history_advances_followup_to_won(self):
        from reporting.models import History
        followup = self._account('독립납품')

        History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='delivery_schedule', delivery_amount=1000000,
        )

        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'won')
        self.assertFalse(followup.pipeline_manually_set)

    def test_standalone_delivery_history_without_followup_advances_whole_department(self):
        from reporting.models import Company, Department, FollowUp, History
        customer_company = Company.objects.create(name='공용업체', created_by=self.user)
        department = Department.objects.create(
            company=customer_company, name='공용연구실', created_by=self.user,
        )
        followup_a = FollowUp.objects.create(
            user=self.user, user_company=self.company, customer_name='담당자A',
            company=customer_company, department=department, pipeline_stage='quote',
        )
        followup_b = FollowUp.objects.create(
            user=self.user, user_company=self.company, customer_name='담당자B',
            company=customer_company, department=department, pipeline_stage='quote',
        )

        History.objects.create(
            user=self.user, company=self.company, department=department,
            action_type='delivery_schedule', delivery_amount=500000,
        )

        followup_a.refresh_from_db()
        followup_b.refresh_from_db()
        self.assertEqual(followup_a.pipeline_stage, 'won')
        self.assertEqual(followup_b.pipeline_stage, 'won')

    def test_editing_existing_delivery_history_keeps_pipeline_at_won(self):
        """delivery_schedule 히스토리를 나중에 고쳐도(.save() 경유) '수주'를 다시 반영한다."""
        from reporting.models import History
        followup = self._account('납품수정')
        history = History.objects.create(
            user=self.user, company=self.company, followup=followup,
            action_type='delivery_schedule', delivery_amount=100000,
        )
        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'won')

        # 사람이 파이프라인을 다른 단계로 되돌렸다고 가정
        followup.pipeline_stage = 'negotiation'
        followup.pipeline_manually_set = True
        followup.save(update_fields=['pipeline_stage', 'pipeline_manually_set'])

        self.client.force_login(self.user)
        response = self.client.post(
            reverse('reporting:history_update_api', args=[history.id]),
            {'content': '납품 완료', 'delivery_amount': '200000', 'delivery_items': '시약 추가'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        followup.refresh_from_db()
        self.assertEqual(followup.pipeline_stage, 'won')

    def test_delivery_pipeline_value_sums_all_completed_deliveries_this_account(self):
        """계정이 여러 번 납품했으면 '수주' 카드 금액에 전부 합산돼야 한다(최근 1건만 세던 버그 수정)."""
        from datetime import time, timedelta
        from decimal import Decimal
        from reporting.models import DeliveryItem, Schedule

        followup = self._account('복수납품', stage='won')
        today = timezone.localdate()
        expected_total = Decimal('0')
        for offset, amount in ((60, 1000000), (10, 2000000)):
            schedule = Schedule.objects.create(
                user=self.user, company=self.company, followup=followup,
                visit_date=today - timedelta(days=offset), visit_time=time(10, 0),
                status='completed', activity_type='delivery',
            )
            # DeliveryItem.save()가 total_price를 단가*수량*1.1(부가세)로 재계산한다.
            item = DeliveryItem.objects.create(
                schedule=schedule, item_name='장비', quantity=1, unit_price=amount,
            )
            expected_total += item.total_price

        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:pipeline_command_center_api'))
        deal = next(item for item in response.json()['deals'] if item['id'] == followup.id)
        self.assertEqual(deal['value'], int(expected_total))


# ─────────────────────────────────────────────────────────────────────────────
# 파이프라인 "올해것만" — 연간 리셋 + 금액 연도 스코핑 검증
# ─────────────────────────────────────────────────────────────────────────────

class PipelineYearResetTests(TestCase):
    """매년 1월 1일 파이프라인 단계가 잠재로 리셋되고, 금액이 올해 것만 반영되는지 검증"""

    def setUp(self):
        from reporting.models import Company, Department, FollowUp
        self.client = Client()
        self.company = UserCompany.objects.create(name='연간리셋회사')
        self.user = make_user('year_reset_me', role='salesman', company=self.company)
        customer_company = Company.objects.create(name='연간리셋업체', created_by=self.user)
        self.department = Department.objects.create(
            company=customer_company, name='연간리셋연구실', created_by=self.user,
        )
        self.followup = FollowUp.objects.create(
            user=self.user, user_company=self.company, customer_name='연간리셋담당자',
            company=customer_company, department=self.department, pipeline_stage='won',
        )

    def test_ensure_pipeline_year_reset_resets_all_non_potential_stages(self):
        from datetime import date
        from reporting.funnel_views import _ensure_pipeline_year_reset
        from reporting.models import PipelineYearResetLog

        self.followup.pipeline_manually_set = True
        self.followup.save(update_fields=['pipeline_manually_set'])

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))

        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'potential')
        self.assertFalse(self.followup.pipeline_manually_set)
        self.assertTrue(PipelineYearResetLog.objects.filter(year=2099).exists())

    def test_ensure_pipeline_year_reset_is_idempotent_within_same_year(self):
        from datetime import date
        from reporting.funnel_views import _ensure_pipeline_year_reset

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))

        # 리셋 이후 올해 새 활동으로 다시 '수주'가 됐다고 가정
        self.followup.pipeline_stage = 'won'
        self.followup.save(update_fields=['pipeline_stage'])

        _ensure_pipeline_year_reset(today=date(2099, 6, 1))

        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'won')

    def test_ensure_pipeline_year_reset_fires_again_next_year(self):
        from datetime import date
        from reporting.funnel_views import _ensure_pipeline_year_reset

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))
        self.followup.pipeline_stage = 'quote'
        self.followup.save(update_fields=['pipeline_stage'])

        _ensure_pipeline_year_reset(today=date(2100, 1, 1))

        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'potential')

    def test_pipeline_command_center_api_triggers_year_reset_automatically(self):
        from datetime import date
        from unittest.mock import patch
        from reporting.models import PipelineYearResetLog

        PipelineYearResetLog.objects.all().delete()
        self.client.force_login(self.user)

        with patch('reporting.funnel_views.timezone.localdate', return_value=date(2099, 3, 1)):
            response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        self.assertEqual(response.status_code, 200)
        self.followup.refresh_from_db()
        self.assertEqual(self.followup.pipeline_stage, 'potential')
        self.assertTrue(PipelineYearResetLog.objects.filter(year=2099).exists())

    def test_pipeline_value_excludes_prior_year_delivery(self):
        """올해 이미 리셋을 거쳐 '수주'로 재확인된 계정이, 작년 납품 금액은 빼고
        올해 납품만 카드 금액에 반영하는지 검증한다."""
        from datetime import date, time
        from unittest.mock import patch
        from reporting.funnel_views import _ensure_pipeline_year_reset
        from reporting.models import DeliveryItem, Schedule

        # 올해 리셋이 이미 끝난 상태를 만든 뒤(먼저 리셋 → 그 다음에 다시 '수주'로
        # 확정) 리셋이 이번 API 호출에서 다시 값을 지우지 않도록 한다.
        _ensure_pipeline_year_reset(today=date(2099, 1, 1))
        self.followup.pipeline_stage = 'won'
        self.followup.save(update_fields=['pipeline_stage'])

        old_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=date(2020, 3, 1), visit_time=time(10, 0),
            status='completed', activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=old_schedule, item_name='작년 장비', quantity=1, unit_price=5000000,
        )
        current_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=date(2099, 2, 1), visit_time=time(10, 0),
            status='completed', activity_type='delivery',
        )
        current_item = DeliveryItem.objects.create(
            schedule=current_schedule, item_name='올해 장비', quantity=1, unit_price=1000000,
        )

        self.client.force_login(self.user)
        with patch('reporting.funnel_views.timezone.localdate', return_value=date(2099, 3, 1)):
            response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        deal = next(item for item in response.json()['deals'] if item['id'] == self.followup.id)
        self.assertEqual(deal['value'], int(current_item.total_price))

    def test_pipeline_value_excludes_prior_year_quote(self):
        """작년 견적은 카드 금액에서 빠지고, 올해 견적만 반영돼야 한다."""
        from datetime import date, time
        from unittest.mock import patch
        from reporting.funnel_views import _ensure_pipeline_year_reset
        from reporting.models import Quote, Schedule

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))
        self.followup.pipeline_stage = 'quote'
        self.followup.save(update_fields=['pipeline_stage'])

        old_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=date(2020, 5, 1), visit_time=time(10, 0),
            status='completed', activity_type='quote',
        )
        old_quote = Quote.objects.create(
            quote_number='OLD-1', schedule=old_schedule, followup=self.followup, user=self.user,
            valid_until=date(2020, 12, 31), subtotal=9000000, total_amount=9000000,
        )
        # quote_date는 auto_now_add라 생성 시점을 못 정하므로 저장 후 직접 되돌린다.
        Quote.objects.filter(pk=old_quote.pk).update(quote_date=date(2020, 5, 1))

        current_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=date(2099, 2, 1), visit_time=time(10, 0),
            status='completed', activity_type='quote',
        )
        current_quote = Quote.objects.create(
            quote_number='NEW-1', schedule=current_schedule, followup=self.followup, user=self.user,
            valid_until=date(2099, 12, 31), subtotal=2000000, total_amount=2000000,
        )
        Quote.objects.filter(pk=current_quote.pk).update(quote_date=date(2099, 2, 1))

        self.client.force_login(self.user)
        with patch('reporting.funnel_views.timezone.localdate', return_value=date(2099, 3, 1)):
            response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        deal = next(item for item in response.json()['deals'] if item['id'] == self.followup.id)
        current_quote.refresh_from_db()
        self.assertEqual(deal['value'], int(current_quote.total_amount))

    def test_pipeline_hides_advanced_stage_card_with_no_current_year_evidence(self):
        """근거 없는 진행 단계 카드는 0원으로 남기지 말고 보드에서 아예 뺀다."""
        from datetime import date, time
        from unittest.mock import patch
        from reporting.funnel_views import _ensure_pipeline_year_reset
        from reporting.models import Quote, Schedule

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))
        self.followup.pipeline_stage = 'quote'
        self.followup.save(update_fields=['pipeline_stage'])

        old_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=date(2020, 5, 1), visit_time=time(10, 0),
            status='completed', activity_type='quote',
        )
        old_quote = Quote.objects.create(
            quote_number='STALE-1', schedule=old_schedule, followup=self.followup, user=self.user,
            valid_until=date(2020, 12, 31), subtotal=9000000, total_amount=9000000,
        )
        Quote.objects.filter(pk=old_quote.pk).update(quote_date=date(2020, 5, 1))

        self.client.force_login(self.user)
        with patch('reporting.funnel_views.timezone.localdate', return_value=date(2099, 3, 1)):
            response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        payload = response.json()
        self.assertFalse(any(item['id'] == self.followup.id for item in payload['deals']))
        stages = {stage['id']: stage for stage in payload['stages']}
        self.assertEqual(stages['quote']['count'], 0)

    def test_pipeline_keeps_potential_card_with_no_evidence(self):
        """잠재 단계는 원래 근거가 없어도 정상 — 이 규칙에서 제외되어야 한다."""
        from datetime import date
        from unittest.mock import patch
        from reporting.funnel_views import _ensure_pipeline_year_reset

        _ensure_pipeline_year_reset(today=date(2099, 1, 1))
        self.followup.pipeline_stage = 'potential'
        self.followup.pipeline_manually_set = False
        self.followup.save(update_fields=['pipeline_stage', 'pipeline_manually_set'])

        self.client.force_login(self.user)
        with patch('reporting.funnel_views.timezone.localdate', return_value=date(2099, 3, 1)):
            response = self.client.get(reverse('reporting:pipeline_command_center_api'))

        payload = response.json()
        self.assertTrue(any(item['id'] == self.followup.id for item in payload['deals']))


# ─────────────────────────────────────────────────────────────────────────────
# 대시보드 매출 드릴다운 ('진짜 매출' 내역) 검증
# ─────────────────────────────────────────────────────────────────────────────

class RevenueDetailApiTests(TestCase):
    """대시보드 매출 카드 → 드릴다운 화면의 합계가 상단 숫자와 일치하는지 검증"""

    def setUp(self):
        from reporting.models import Company, Department, FollowUp
        self.client = Client()
        self.company = UserCompany.objects.create(name='매출드릴다운회사')
        self.user = make_user('revenue_detail_me', role='salesman', company=self.company)
        customer_company = Company.objects.create(name='매출드릴다운업체', created_by=self.user)
        self.department = Department.objects.create(
            company=customer_company, name='매출드릴다운연구실', created_by=self.user,
        )
        self.followup = FollowUp.objects.create(
            user=self.user, user_company=self.company, customer_name='매출드릴다운담당자',
            company=customer_company, department=self.department, pipeline_stage='won',
        )

    def test_requires_login(self):
        response = self.client.get(reverse('reporting:revenue_detail_api'))
        self.assertEqual(response.status_code, 401)

    def test_excludes_scheduled_delivery_and_includes_completed(self):
        from datetime import time
        from reporting.models import DeliveryItem, Schedule

        today = timezone.localdate()
        completed_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=today, visit_time=time(10, 0),
            status='completed', activity_type='delivery',
        )
        completed_item = DeliveryItem.objects.create(
            schedule=completed_schedule, item_name='완료납품', quantity=1, unit_price=1000000,
        )
        scheduled_schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=today, visit_time=time(11, 0),
            status='scheduled', activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=scheduled_schedule, item_name='예정납품', quantity=1, unit_price=9000000,
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:revenue_detail_api'), {'period': 'year'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        item_names = [item['itemName'] for item in payload['items'] if item['kind'] == 'delivery']
        self.assertIn('완료납품', item_names)
        self.assertNotIn('예정납품', item_names)
        completed_item.refresh_from_db()
        self.assertEqual(payload['summary']['deliveryTotal'], int(completed_item.total_price))

    def test_includes_prepayment_and_excludes_cancelled(self):
        from reporting.models import Prepayment

        today = timezone.localdate()
        active_prepayment = Prepayment.objects.create(
            customer=self.followup, company=self.followup.company, department=self.department,
            amount=500000, balance=500000, payment_date=today, created_by=self.user,
        )
        Prepayment.objects.create(
            customer=self.followup, company=self.followup.company, department=self.department,
            amount=700000, balance=700000, payment_date=today, created_by=self.user, status='cancelled',
        )

        self.client.force_login(self.user)
        response = self.client.get(reverse('reporting:revenue_detail_api'), {'period': 'year'})

        payload = response.json()
        self.assertEqual(payload['summary']['prepaymentTotal'], int(active_prepayment.amount))
        prepayment_hrefs = [item['href'] for item in payload['items'] if item['kind'] == 'prepayment']
        self.assertEqual(prepayment_hrefs, [f'/prepayments/{active_prepayment.id}/'])

    def test_total_matches_dashboard_summary_metric(self):
        from datetime import time
        from reporting.models import DeliveryItem, Prepayment, Schedule

        today = timezone.localdate()
        schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=today, visit_time=time(10, 0),
            status='completed', activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=schedule, item_name='매출검증납품', quantity=1, unit_price=1200000,
        )
        Prepayment.objects.create(
            customer=self.followup, company=self.followup.company, department=self.department,
            amount=300000, balance=300000, payment_date=today, created_by=self.user,
        )

        self.client.force_login(self.user)
        dashboard_response = self.client.get(reverse('reporting:dashboard_summary_api'))
        detail_response = self.client.get(reverse('reporting:revenue_detail_api'), {'period': 'year'})

        dashboard_year_revenue = dashboard_response.json()['metrics']['yearRevenue']
        detail_total = detail_response.json()['summary']['total']
        self.assertEqual(detail_total, dashboard_year_revenue)
        self.assertGreater(detail_total, 0)

    def test_month_period_total_matches_dashboard_monthly_metric(self):
        from datetime import time
        from reporting.models import DeliveryItem, Prepayment, Schedule

        today = timezone.localdate()
        schedule = Schedule.objects.create(
            user=self.user, company=self.company, followup=self.followup,
            visit_date=today, visit_time=time(10, 0),
            status='completed', activity_type='delivery',
        )
        DeliveryItem.objects.create(
            schedule=schedule, item_name='이번달매출검증', quantity=1, unit_price=400000,
        )
        Prepayment.objects.create(
            customer=self.followup, company=self.followup.company, department=self.department,
            amount=150000, balance=150000, payment_date=today, created_by=self.user,
        )

        self.client.force_login(self.user)
        dashboard_response = self.client.get(reverse('reporting:dashboard_summary_api'))
        detail_response = self.client.get(reverse('reporting:revenue_detail_api'), {'period': 'month'})

        dashboard_monthly_revenue = dashboard_response.json()['metrics']['monthlyRevenue']
        payload = detail_response.json()
        self.assertEqual(payload['period']['value'], 'month')
        self.assertEqual(payload['summary']['total'], dashboard_monthly_revenue)
        self.assertGreater(payload['summary']['total'], 0)
