from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required # 로그인 요구 데코레이터
from django.contrib.auth.models import User
from django.contrib import messages
from django import forms
from django.http import JsonResponse, HttpResponseForbidden, Http404, FileResponse
from django.db.models import Sum, Count, Q, Prefetch
from django.core.paginator import Paginator  # 페이지네이션 추가
from .models import FollowUp, Schedule, History, UserProfile, Company, Department, HistoryFile, DeliveryItem, UserCompany, Prepayment, PrepaymentUsage, EmailLog, CustomerCategory, WeeklyReport, OpportunityTracking, OpportunityLabel, Quote
from django.contrib.auth.views import LoginView, LogoutView
from django.urls import reverse_lazy, reverse
from functools import wraps
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST, require_http_methods
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from .decorators import hanagwahak_only, get_allowed_action_types, get_allowed_activity_types, filter_service_for_non_hanagwahak
import os
import mimetypes
import logging
import json

# 로거 설정
logger = logging.getLogger(__name__)

def save_delivery_items(request, instance_obj):
    """납품 품목 데이터를 저장하는 함수 (스케줄 또는 히스토리)"""
    from .models import DeliveryItem
    
    # 인스턴스 타입 확인
    from .models import Schedule, History
    is_schedule = isinstance(instance_obj, Schedule)
    is_history = isinstance(instance_obj, History)
    
    if not (is_schedule or is_history):
        logger.error(f"save_delivery_items: 지원되지 않는 객체 타입: {type(instance_obj)}")
        return
    
    # 기존 품목들 삭제 (수정 시)
    if is_schedule:
        existing_count = instance_obj.delivery_items_set.all().count()
        instance_obj.delivery_items_set.all().delete()
    else:  # is_history
        existing_count = instance_obj.delivery_items_set.all().count()
        instance_obj.delivery_items_set.all().delete()
    
    # delivery_items 관련 POST 데이터만 필터링
    delivery_items_data = {}
    
    for key, value in request.POST.items():
        if key.startswith('delivery_items[') and '][' in key:
            # delivery_items[0][name] -> index=0, field=name
            try:
                start = key.find('[') + 1
                end = key.find(']')
                index = int(key[start:end])
                
                field_start = key.rfind('[') + 1
                field_end = key.rfind(']')
                field = key[field_start:field_end]
                
                if index not in delivery_items_data:
                    delivery_items_data[index] = {}
                delivery_items_data[index][field] = value
            except (ValueError, IndexError) as e:
                logger.error(f"POST 데이터 파싱 실패: {key} = {value}, 오류: {e}")
                continue
    
    # 납품 품목 저장
    created_count = 0
    
    for index, item_data in delivery_items_data.items():
        item_name = item_data.get('name', '').strip()
        quantity = item_data.get('quantity', '').strip()
        unit_price = item_data.get('unit_price', '').strip()
        product_id = item_data.get('product_id', '').strip()
        
        if item_name and quantity:
            try:
                delivery_item = DeliveryItem(
                    item_name=item_name,
                    quantity=int(quantity),
                    unit='개',  # 기본값
                )
                
                # 제품 연결
                if product_id:
                    try:
                        from .models import Product
                        delivery_item.product = Product.objects.get(id=int(product_id))
                    except (Product.DoesNotExist, ValueError):
                        pass
                
                # 스케줄 또는 히스토리 연결
                if is_schedule:
                    delivery_item.schedule = instance_obj
                else:
                    delivery_item.history = instance_obj
                
                # unit_price 저장 (빈 문자열, None이 아니면 0 포함 모든 숫자 허용)
                if unit_price != '' and unit_price is not None:
                    from decimal import Decimal
                    try:
                        # "0", "0.0", 0 모두 Decimal로 변환
                        delivery_item.unit_price = Decimal(str(unit_price))
                    except (ValueError, decimal.InvalidOperation):
                        # 변환 실패 시 None으로 유지
                        pass
                # unit_price가 '' 또는 None이면 unit_price 필드는 None으로 유지
                
                delivery_item.save()
                created_count += 1
            except (ValueError, TypeError) as e:
                logger.error(f"납품 품목 저장 실패: {e}")
                continue  # 잘못된 데이터는 무시
    
    return created_count

# 권한 체크 데코레이터
def role_required(allowed_roles):
    """특정 역할을 가진 사용자만 접근할 수 있도록 하는 데코레이터"""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('reporting:login')
            
            try:
                user_profile = request.user.userprofile
                if user_profile.role in allowed_roles:
                    return view_func(request, *args, **kwargs)
                else:
                    messages.error(request, '이 페이지에 접근할 권한이 없습니다.')
                    return redirect('reporting:dashboard')
            except UserProfile.DoesNotExist:
                messages.error(request, '사용자 프로필이 설정되지 않았습니다. 관리자에게 문의하세요.')
                return redirect('reporting:dashboard')
        return wrapper
    return decorator

def get_user_profile(user):
    """사용자 프로필을 가져오는 헬퍼 함수"""
    try:
        return user.userprofile
    except UserProfile.DoesNotExist:
        # 프로필이 없는 경우 기본 salesman 권한으로 생성
        return UserProfile.objects.create(user=user, role='salesman')

def can_access_user_data(request_user, target_user):
    """현재 사용자가 대상 사용자의 데이터에 접근할 수 있는지 확인"""
    # 자기 자신의 데이터는 항상 접근 가능
    if request_user == target_user:
        return True
    
    user_profile = get_user_profile(request_user)
    target_profile = get_user_profile(target_user)
    
    # Admin은 모든 데이터 접근 가능
    if user_profile.is_admin():
        return True
    
    # 같은 회사 소속이면 조회 가능 (Salesman, Manager 모두)
    if user_profile.company and target_profile.company:
        if user_profile.company == target_profile.company:
            return True
    
    return False

# 파일 업로드 관련 헬퍼 함수들
def validate_file_upload(file):
    """업로드된 파일의 유효성을 검사"""
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    ALLOWED_EXTENSIONS = [
        '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', 
        '.txt', '.jpg', '.jpeg', '.png', '.gif', '.zip', '.rar',
        '.hwp', '.hwpx',
    ]

    # 파일 크기 검사
    if file.size > MAX_FILE_SIZE:
        return False, f"파일 크기가 너무 큽니다. 최대 {MAX_FILE_SIZE // (1024*1024)}MB까지 업로드 가능합니다."

    # 파일 확장자 검사
    file_extension = os.path.splitext(file.name)[1].lower()
    if file_extension not in ALLOWED_EXTENSIONS:
        return False, f"지원하지 않는 파일 형식입니다. 허용된 확장자: {', '.join(ALLOWED_EXTENSIONS)}"

    # MIME 매직 바이트 검사 (확장자 위장 방지)
    # 파일 포인터를 앞부분으로 이동 후 검사, 마지막에 원위치
    MIME_SIGNATURES = [
        (b'%PDF',           ['.pdf']),
        (b'PK\x03\x04',    ['.docx', '.xlsx', '.pptx', '.zip', '.hwpx']),  # ZIP 계열
        (b'\xff\xd8\xff',  ['.jpg', '.jpeg']),
        (b'\x89PNG\r\n',   ['.png']),
        (b'GIF87a',        ['.gif']),
        (b'GIF89a',        ['.gif']),
        (b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1',  # OLE2 구형 MS Office
                            ['.doc', '.xls', '.ppt', '.hwp']),
        (b'Rar!',          ['.rar']),
    ]
    SKIP_MAGIC_EXTENSIONS = ['.txt', '.hwp', '.hwpx', '.rar']  # 매직 검사 제외 (다양한 형식)

    if file_extension not in SKIP_MAGIC_EXTENSIONS:
        try:
            file.seek(0)
            header = file.read(16)
            file.seek(0)

            magic_ok = False
            for signature, allowed_exts in MIME_SIGNATURES:
                if header.startswith(signature):
                    if file_extension in allowed_exts:
                        magic_ok = True
                    else:
                        return False, "파일 형식이 확장자와 일치하지 않습니다."
                    break

            if not magic_ok and header:
                # 알 수 없는 시그니처 — txt나 단순 텍스트는 허용, 나머지는 차단
                if file_extension not in ['.txt']:
                    return False, "파일 형식을 확인할 수 없습니다."
        except Exception:
            # seek 불가 파일(InMemoryUploadedFile 오류 등)은 확장자 검사만으로 통과
            pass

    return True, "유효한 파일입니다."

def handle_file_uploads(files, history, user):
    """여러 파일 업로드를 처리"""
    MAX_FILES = 5
    uploaded_files = []
    errors = []
    
    if len(files) > MAX_FILES:
        return [], [f"최대 {MAX_FILES}개의 파일만 업로드할 수 있습니다."]
    
    for file in files:
        is_valid, message = validate_file_upload(file)
        if not is_valid:
            errors.append(f"{file.name}: {message}")
            continue
            
        try:
            history_file = HistoryFile.objects.create(
                history=history,
                file=file,
                original_filename=file.name,
                file_size=file.size,
                uploaded_by=user
            )
            uploaded_files.append(history_file)
        except Exception as e:
            errors.append(f"{file.name}: 업로드 중 오류가 발생했습니다.")
    
    return uploaded_files, errors

def can_modify_user_data(request_user, target_user):
    """현재 사용자가 대상 사용자의 데이터를 수정/추가/삭제할 수 있는지 확인"""
    user_profile = get_user_profile(request_user)
    
    # Admin은 모든 데이터 수정 가능
    if user_profile.is_admin():
        return True
    
    # Manager는 읽기만 가능하고 수정 불가
    if user_profile.is_manager():
        return False
    
    # Salesman은 자신의 데이터만 수정 가능
    return request_user == target_user


def can_access_followup(request_user, followup):
    """
    고객(FollowUp) 접근 권한 확인
    - 같은 회사(UserCompany) 소속이면 고객 정보 조회 가능
    - 단, 스케줄/히스토리 기록은 본인 것만 접근 가능
    """
    user_profile = get_user_profile(request_user)
    
    # Admin은 모든 데이터 접근 가능
    if user_profile.is_admin():
        return True
    
    # 같은 회사 소속인지 확인
    if user_profile.company and followup.user:
        target_profile = get_user_profile(followup.user)
        if target_profile.company and user_profile.company == target_profile.company:
            return True
    
    # 자신이 추가한 고객은 당연히 접근 가능
    return request_user == followup.user


def get_same_company_users(request_user):
    """같은 회사(UserCompany) 소속 사용자 목록 반환"""
    user_profile = get_user_profile(request_user)
    
    # Admin은 모든 사용자
    if user_profile.is_admin():
        return User.objects.filter(is_active=True)
    
    # 회사가 있으면 같은 회사 사용자들
    if user_profile.company:
        return User.objects.filter(
            is_active=True,
            userprofile__company=user_profile.company
        )
    
    # 회사가 없으면 자기 자신만
    return User.objects.filter(id=request_user.id)


def get_accessible_users(request_user, request=None):
    """
    현재 사용자가 접근할 수 있는 사용자 목록을 반환
    
    Args:
        request_user: 현재 로그인한 사용자
        request: HTTP request 객체 (관리자 필터 확인용)
        
    Returns:
        QuerySet: 접근 가능한 사용자 목록
    """
    user_profile = get_user_profile(request_user)
    
    if user_profile.is_admin():
        # 관리자: 필터링 적용
        if request and hasattr(request, 'admin_filter_user') and request.admin_filter_user:
            # 특정 사용자 선택됨
            return User.objects.filter(id=request.admin_filter_user.id)
        elif request and hasattr(request, 'admin_filter_company') and request.admin_filter_company:
            # 특정 회사 선택됨 - 해당 회사의 모든 실무자
            return User.objects.filter(
                userprofile__company=request.admin_filter_company,
                userprofile__role__in=['salesman', 'manager']
            )
        else:
            # 전체 접근
            return User.objects.all()
            
    elif user_profile.company:
        # Manager와 Salesman 모두 같은 회사의 모든 사용자에 접근 가능
        user_company = user_profile.company
        accessible_profiles = UserProfile.objects.filter(
            role__in=['salesman', 'manager'],
            company=user_company
        )
        return User.objects.filter(userprofile__in=accessible_profiles)
    else:
        # 회사 정보가 없는 경우 자기 자신만 접근 가능
        return User.objects.filter(id=request_user.id)

# 팔로우업 폼 클래스
class FollowUpForm(forms.ModelForm):
    # 자동완성을 위한 hidden 필드들
    company = forms.ModelChoiceField(
        queryset=Company.objects.all(),
        widget=forms.HiddenInput(),
        required=True,
        error_messages={'required': '업체/학교를 선택해주세요.'}
    )
    department = forms.ModelChoiceField(
        queryset=Department.objects.all(),
        widget=forms.HiddenInput(),
        required=True,
        error_messages={'required': '부서/연구실명은 필수 입력사항입니다.'}
    )
    
    category = forms.ModelChoiceField(
        queryset=CustomerCategory.objects.none(),
        required=False,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='카테고리'
    )
    
    class Meta:
        model = FollowUp
        fields = ['customer_name', 'company', 'department', 'manager', 'phone_number', 'email', 'address', 'notes', 'priority', 'category']
        widgets = {
            'customer_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '고객명을 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            'manager': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '책임자명을 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            'phone_number': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '010-0000-0000 (선택사항)', 'autocomplete': 'off'}),
            'email': forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'example@company.com (선택사항)', 'autocomplete': 'off'}),
            'address': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '상세주소를 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            'notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 4, 'placeholder': '상세 내용을 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            'priority': forms.Select(attrs={'class': 'form-control'}),
        }
        labels = {
            'customer_name': '고객명',
            'company': '업체/학교명',
            'department': '부서/연구실명',
            'manager': '책임자',
            'phone_number': '핸드폰 번호',
            'email': '메일 주소',
            'address': '상세주소',
            'notes': '상세 내용',
            'priority': '우선순위',
            'category': '카테고리',
        }
    
    def __init__(self, *args, **kwargs):
        user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        if user:
            # 카테고리를 계층 구조로 표시
            categories = CustomerCategory.objects.filter(user=user).select_related('parent').order_by('order', 'name')
            self.fields['category'].queryset = categories
            # 선택 옵션을 계층적으로 표시
            choices = [('', '---------')]
            for cat in categories:
                if cat.parent:
                    choices.append((cat.id, f'  └ {cat.name}'))
                else:
                    choices.append((cat.id, cat.name))
            self.fields['category'].choices = choices
        
    def clean_company(self):
        company = self.cleaned_data.get('company')
        if not company:
            raise forms.ValidationError('업체/학교를 선택해주세요.')
        return company

    def clean_department(self):
        department = self.cleaned_data.get('department')
        if not department:
            raise forms.ValidationError('부서/연구실명은 필수 입력사항입니다.')
        return department

# 일정 폼 클래스
class ScheduleForm(forms.ModelForm):
    followup = forms.ModelChoiceField(
        queryset=FollowUp.objects.none(),  # 초기에는 비어있음
        widget=forms.Select(attrs={
            'class': 'form-control followup-autocomplete',
            'data-placeholder': '팔로우업을 검색하세요...',
            'data-url': '',  # JavaScript에서 설정됨
        }),
        label='관련 팔로우업',
        help_text='고객명, 업체명 또는 부서명으로 검색할 수 있습니다.'
    )
    
    # 선결제 관련 필드
    prepayment = forms.ModelChoiceField(
        queryset=Prepayment.objects.none(),
        required=False,
        widget=forms.Select(attrs={
            'class': 'form-control',
            'id': 'id_prepayment',
        }),
        label='사용할 선결제',
        help_text='고객의 선결제 잔액에서 차감할 선결제를 선택하세요.'
    )
    
    # status 필드를 명시적으로 선언 (required=False)
    status = forms.ChoiceField(
        choices=Schedule.STATUS_CHOICES,
        required=False,
        initial='scheduled',
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='일정 상태'
    )
    
    
    class Meta:
        model = Schedule
        fields = ['followup', 'visit_date', 'visit_time', 'activity_type', 'location', 'status', 'notes', 
                  'use_prepayment', 'prepayment', 'prepayment_amount']
        widgets = {
            'visit_date': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'visit_time': forms.TimeInput(attrs={'class': 'form-control', 'type': 'time'}),
            'activity_type': forms.Select(attrs={'class': 'form-control'}),
            'location': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '방문 장소를 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            # status는 위에서 명시적으로 선언했으므로 여기서 제거
            'notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '메모를 입력하세요 (선택사항)', 'autocomplete': 'off'}),
            'use_prepayment': forms.CheckboxInput(attrs={'class': 'form-check-input', 'id': 'id_use_prepayment'}),
            'prepayment_amount': forms.NumberInput(attrs={'class': 'form-control', 'placeholder': '차감할 금액 (원)', 'min': '0', 'id': 'id_prepayment_amount'}),
        }
        labels = {
            'visit_date': '방문 날짜',
            'visit_time': '방문 시간',
            'activity_type': '일정 유형',
            'location': '장소',
            'status': '상태',
            'notes': '메모',
            'use_prepayment': '선결제 사용',
            'prepayment_amount': '선결제 차감 금액 (원)',
        }

    def __init__(self, *args, **kwargs):
        user = kwargs.pop('user', None)
        request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        
        # 새 일정 생성 시 기본값 설정
        if not self.instance.pk:
            self.initial['status'] = 'scheduled'
            self.fields['status'].initial = 'scheduled'
        
        if user:
            # 사용자 소속 회사(UserCompany)의 모든 팔로우업을 선택할 수 있도록 필터링
            if user.is_staff or user.is_superuser:
                base_queryset = FollowUp.objects.all()
            else:
                # 사용자의 UserCompany를 가져옴
                user_company = user.userprofile.company
                if user_company:
                    # 같은 UserCompany에 속한 모든 사용자의 팔로우업
                    base_queryset = FollowUp.objects.filter(user__userprofile__company=user_company)
                else:
                    # UserCompany가 없으면 본인 팔로우업만
                    base_queryset = FollowUp.objects.filter(user=user)
                
            # 기존 인스턴스가 있는 경우 해당 팔로우업도 포함
            if self.instance.pk and self.instance.followup:
                # Q 객체를 사용하여 기존 팔로우업과 사용자 팔로우업을 OR 조건으로 결합
                from django.db.models import Q
                if user.is_staff or user.is_superuser:
                    queryset_filter = Q()  # 모든 팔로우업
                else:
                    user_company = user.userprofile.company
                    if user_company:
                        queryset_filter = Q(user__userprofile__company=user_company) | Q(pk=self.instance.followup.pk)
                    else:
                        queryset_filter = Q(user=user) | Q(pk=self.instance.followup.pk)
                self.fields['followup'].queryset = FollowUp.objects.filter(queryset_filter).select_related('company', 'department').distinct()
            else:
                self.fields['followup'].queryset = base_queryset.select_related('company', 'department')
                
            # 자동완성 URL 설정
            from django.urls import reverse
            self.fields['followup'].widget.attrs['data-url'] = reverse('reporting:followup_autocomplete')
        
        # 선결제 필드 설정 (같은 부서 내 모든 고객의 선결제를 공유)
        if self.instance.pk and self.instance.followup:
            # 수정 시: 같은 부서의 모든 고객 선결제 목록
            department = self.instance.followup.department
            same_dept_followups = FollowUp.objects.filter(department=department).values_list('id', flat=True)
            self.fields['prepayment'].queryset = Prepayment.objects.filter(
                customer_id__in=same_dept_followups,
                status='active',
                balance__gt=0
            ).order_by('payment_date')
            # 선결제 옵션 라벨 설정 (고객명 포함)
            self.fields['prepayment'].label_from_instance = lambda obj: f"{obj.payment_date.strftime('%Y-%m-%d')} - {obj.customer.customer_name or ''} / {obj.payer_name or '미지정'} (잔액: {obj.balance:,}원)"
        elif 'followup' in self.data:
            # 생성 시 followup이 선택된 경우
            try:
                followup_id = int(self.data.get('followup'))
                followup = FollowUp.objects.get(pk=followup_id)
                # 생성 시: 같은 부서의 모든 고객 선결제 목록
                department = followup.department
                same_dept_followups = FollowUp.objects.filter(department=department).values_list('id', flat=True)
                self.fields['prepayment'].queryset = Prepayment.objects.filter(
                    customer_id__in=same_dept_followups,
                    status='active',
                    balance__gt=0
                ).order_by('payment_date')
                # 선결제 옵션 라벨 설정 (고객명 포함)
                self.fields['prepayment'].label_from_instance = lambda obj: f"{obj.payment_date.strftime('%Y-%m-%d')} - {obj.customer.customer_name or ''} / {obj.payer_name or '미지정'} (잔액: {obj.balance:,}원)"
            except (ValueError, TypeError, FollowUp.DoesNotExist):
                pass
            
        # 하나과학이 아닌 경우 activity_type에서 서비스 제거
        if request and not getattr(request, 'is_hanagwahak', False):
            self.fields['activity_type'].choices = [
                choice for choice in self.fields['activity_type'].choices 
                if choice[0] != 'service'
            ]
        
        # 견적 일정 수정 시 상태 선택지 제한 (예정→완료 불가, 취소만 가능)
        if self.instance.pk and self.instance.activity_type == 'quote':
            # 견적은 예정됨, 취소만 선택 가능 (완료 제외)
            self.fields['status'].choices = [
                ('scheduled', '예정됨'),
                ('cancelled', '취소됨'),
            ]
    
# 히스토리 폼 클래스
class HistoryForm(forms.ModelForm):
    # 파일 업로드 필드 (JavaScript로 여러 파일 처리)
    files = forms.FileField(
        required=False,
        widget=forms.FileInput(attrs={
            'class': 'form-control',
            'accept': '.pdf,.doc,.docx,.xls,.xlsx,.ppt,.pptx,.txt,.jpg,.jpeg,.png,.gif,.zip,.rar'
        }),
        label='첨부파일',
        help_text='최대 10MB, 최대 5개 파일까지 업로드 가능'
    )
    
    class Meta:
        model = History
        fields = ['followup', 'schedule', 'action_type', 'service_status', 'content', 'delivery_amount', 'delivery_items', 'delivery_date', 'meeting_date',
                  'meeting_situation', 'meeting_researcher_quote', 'meeting_confirmed_facts', 'meeting_obstacles', 'meeting_next_action',
                  'next_action', 'next_action_date']
        widgets = {
            'followup': forms.Select(attrs={'class': 'form-control'}),
            'schedule': forms.Select(attrs={'class': 'form-control'}),
            'action_type': forms.Select(attrs={'class': 'form-control'}),
            'service_status': forms.Select(attrs={'class': 'form-control'}),
            'content': forms.Textarea(attrs={'class': 'form-control', 'rows': 4, 'placeholder': '활동 내용을 입력하세요', 'autocomplete': 'off'}),
            'delivery_amount': forms.NumberInput(attrs={'class': 'form-control', 'placeholder': '납품 금액을 입력하세요 (원)', 'min': '0', 'autocomplete': 'off'}),
            'delivery_items': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '납품 품목을 입력하세요 (예: 제품A 10개, 제품B 5개)', 'autocomplete': 'off'}),
            'delivery_date': forms.DateInput(attrs={'class': 'form-control', 'type': 'date', 'placeholder': 'YYYY-MM-DD'}),
            'meeting_date': forms.DateInput(attrs={'class': 'form-control', 'type': 'date', 'placeholder': 'YYYY-MM-DD'}),
            'meeting_situation': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '오늘 미팅에서 파악한 전반적인 상황을 기록하세요', 'autocomplete': 'off'}),
            'meeting_researcher_quote': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '연구원이 직접 한 말을 인용하여 기록하세요 (예: "이번 프로젝트에 꼭 필요합니다")', 'autocomplete': 'off'}),
            'meeting_confirmed_facts': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '미팅에서 직접 확인한 사실을 기록하세요', 'autocomplete': 'off'}),
            'meeting_obstacles': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '영업 진행에 장애물이나 반대 의견을 기록하세요', 'autocomplete': 'off'}),
            'meeting_next_action': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '미팅 이후 수행할 다음 액션을 기록하세요', 'autocomplete': 'off'}),
            'next_action': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': '이번 활동 이후 수행할 다음 액션을 기록하세요', 'autocomplete': 'off'}),
            'next_action_date': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
        }
        labels = {
            'followup': '관련 고객 정보',
            'schedule': '관련 일정',
            'action_type': '활동 유형',
            'service_status': '서비스 상태',
            'content': '활동 내용',
            'delivery_amount': '납품 금액 (원)',
            'delivery_items': '납품 품목',
            'delivery_date': '납품 날짜',
            'meeting_date': '미팅 날짜',
            'meeting_situation': '오늘 상황',
            'meeting_researcher_quote': '연구원이 한 말(직접 인용)',
            'meeting_confirmed_facts': '내가 확인한 사실',
            'meeting_obstacles': '장애물/반대',
            'meeting_next_action': '다음 액션',
            'next_action': '다음 할 일',
            'next_action_date': '다음 예정일',
        }

    def __init__(self, *args, **kwargs):
        user = kwargs.pop('user', None)
        request = kwargs.pop('request', None)
        super().__init__(*args, **kwargs)
        if user:
            # 같은 회사 사용자들의 팔로우업도 선택 가능하도록 필터링
            if user.is_staff or user.is_superuser:
                self.fields['followup'].queryset = FollowUp.objects.all()
            else:
                # 같은 회사 소속 모든 사용자의 고객 선택 가능
                same_company_users = get_same_company_users(user)
                self.fields['followup'].queryset = FollowUp.objects.filter(user__in=same_company_users)
            
            # 일정 필드를 빈 상태로 초기화 (JavaScript에서 동적으로 로드)
            self.fields['schedule'].queryset = Schedule.objects.none()
            
            # 선택된 팔로우업이 있으면 해당 팔로우업의 일정만 표시
            if 'followup' in self.data:
                try:
                    followup_id = int(self.data.get('followup'))
                    self.fields['schedule'].queryset = Schedule.objects.filter(followup_id=followup_id, user=user)
                except (ValueError, TypeError):
                    pass  # 잘못된 입력인 경우 무시
            elif self.instance.pk:
                # 수정 시에는 해당 인스턴스의 팔로우업에 연결된 일정들을 표시
                if self.instance.followup:
                    self.fields['schedule'].queryset = self.instance.followup.schedules.all()
                else:
                    # 팔로우업이 없는 경우 (일반 메모 등) 빈 쿼리셋 유지
                    self.fields['schedule'].queryset = Schedule.objects.none()
                
            # 선택된 일정이 있으면 해당 일정의 activity_type에 맞게 action_type 매핑
            if 'schedule' in self.data and self.data.get('schedule'):
                try:
                    schedule_id = int(self.data.get('schedule'))
                    selected_schedule = Schedule.objects.get(id=schedule_id)
                    # 일정의 activity_type을 히스토리의 action_type으로 매핑
                    activity_mapping = {
                        'customer_meeting': 'customer_meeting',
                        'quote': 'quote',
                        'delivery': 'delivery_schedule',
                        'service': 'service',
                    }
                    # Schedule 모델의 activity_type을 확인하여 매핑
                    if hasattr(selected_schedule, 'activity_type'):
                        mapped_action = activity_mapping.get(selected_schedule.activity_type, 'customer_meeting')
                        self.fields['action_type'].initial = mapped_action
                except (ValueError, TypeError, Schedule.DoesNotExist):
                    pass
        
        # 일정은 선택사항으로 설정
        self.fields['schedule'].required = False
        self.fields['schedule'].empty_label = "관련 일정 없음"
        
        # 활동 유형에서 메모 제외 (메모는 별도 폼에서만 생성 가능)
        # 하나과학이 아닌 경우 서비스도 제외
        excluded_types = ['memo']
        if request and not getattr(request, 'is_hanagwahak', False):
            excluded_types.append('service')
            
        self.fields['action_type'].choices = [
            choice for choice in self.fields['action_type'].choices 
            if choice[0] not in excluded_types
        ]
        
        # 납품 금액은 선택사항으로 설정
        self.fields['delivery_amount'].required = False

@login_required # 이 뷰는 로그인이 필요함을 명시
def followup_list_view(request):
    """팔로우업 목록 보기 (같은 회사 소속 고객 전체 조회 가능)"""
    user_profile = get_user_profile(request.user)
    
    # 매니저용 실무자 필터 (세션 기반)
    view_all = request.GET.get('view_all') == 'true'
    
    # 전체 팀원 선택 시 세션 초기화
    if view_all and user_profile.can_view_all_users():
        if 'selected_user_id' in request.session:
            del request.session['selected_user_id']
        user_filter = None
    else:
        user_filter = request.GET.get('user')
        if not user_filter:
            user_filter = request.session.get('selected_user_id')
        
        if user_filter and user_profile.can_view_all_users():
            request.session['selected_user_id'] = str(user_filter)
    
    # 같은 회사 소속 사용자들의 고객 전체 조회 가능
    same_company_users = get_same_company_users(request.user)
    
    # 권한에 따른 데이터 필터링
    if user_profile.can_view_all_users():
        # Admin이나 Manager는 모든 또는 접근 가능한 사용자의 데이터 조회
        accessible_users = get_accessible_users(request.user, request)
        
        # 매니저가 특정 실무자를 선택한 경우
        if user_filter and not view_all:
            try:
                selected_user = accessible_users.get(id=user_filter)
                followups = FollowUp.objects.filter(user=selected_user).select_related('user', 'company', 'department').prefetch_related('schedules', 'histories')
            except (User.DoesNotExist, ValueError):
                followups = FollowUp.objects.filter(user__in=accessible_users).select_related('user', 'company', 'department').prefetch_related('schedules', 'histories')
        else:
            # 전체보기 또는 선택 안 함
            followups = FollowUp.objects.filter(user__in=accessible_users).select_related('user', 'company', 'department').prefetch_related('schedules', 'histories')
    else:
        # Salesman은 같은 회사 소속 사용자들의 고객 전체 조회 가능
        followups = FollowUp.objects.filter(user__in=same_company_users).select_related('user', 'company', 'department').prefetch_related('schedules', 'histories')
    
    # 고객명/업체명/책임자명 검색 기능 (다중 검색 지원: 쉼표로 구분)
    search_query = request.GET.get('search')
    if search_query:
        # 쉼표로 구분된 다중 검색어 처리
        search_terms = [term.strip() for term in search_query.split(',') if term.strip()]
        
        if search_terms:
            # 각 검색어에 대해 OR 조건으로 검색
            combined_q = Q()
            for term in search_terms:
                combined_q |= (
                    Q(customer_name__icontains=term) |
                    Q(company__name__icontains=term) |
                    Q(department__name__icontains=term) |
                    Q(manager__icontains=term) |
                    Q(notes__icontains=term)
                )
            followups = followups.filter(combined_q).distinct()
    
    # 우선순위 필터링
    priority_filter = request.GET.get('priority')
    if priority_filter:
        followups = followups.filter(priority=priority_filter)
    
    # 고객 등급 필터링
    grade_filter = request.GET.get('grade')
    if grade_filter:
        followups = followups.filter(customer_grade=grade_filter)
    
    # 파이프라인 단계 필터링
    pipeline_stage_filter = request.GET.get('pipeline_stage', '')
    if pipeline_stage_filter:
        followups = followups.filter(pipeline_stage=pipeline_stage_filter)

    # 종합 점수(우선순위 레벨) 필터링
    level_filter = request.GET.get('level')
    if level_filter:
        # 종합 점수 범위로 필터링 - queryset을 유지하기 위해 먼저 리스트로 변환
        followups_list = list(followups)
        if level_filter == 'critical':  # 최우선 85+
            followups_list = [f for f in followups_list if f.get_combined_score() >= 85]
        elif level_filter == 'high':  # 높음 70-84
            followups_list = [f for f in followups_list if 70 <= f.get_combined_score() < 85]
        elif level_filter == 'medium':  # 중간 50-69
            followups_list = [f for f in followups_list if 50 <= f.get_combined_score() < 70]
        elif level_filter == 'low':  # 낮음 30-49
            followups_list = [f for f in followups_list if 30 <= f.get_combined_score() < 50]
        elif level_filter == 'minimal':  # 최소 30-
            followups_list = [f for f in followups_list if f.get_combined_score() < 30]
    else:
        followups_list = None
    
    # 업체별 카운트 (업체 필터 적용 전 기준)
    from django.db.models import Count, Q as DbQ
    if level_filter:
        # 리스트로 변환되었으면 카운트 직접 계산
        stats = {
            'total_count': len(followups_list),
            'active_count': len([f for f in followups_list if f.status == 'active']),
            'completed_count': len([f for f in followups_list if f.status == 'completed']),
            'paused_count': len([f for f in followups_list if f.status == 'paused']),
        }
    else:
        stats = followups.aggregate(
            total_count=Count('id'),
            active_count=Count('id', filter=DbQ(status='active')),
            completed_count=Count('id', filter=DbQ(status='completed')),
            paused_count=Count('id', filter=DbQ(status='paused'))
        )
    
    # 업체 필터링 (카운트 계산 후에 적용)
    company_filter = request.GET.get('company')
    if company_filter:
        if level_filter:
            # 리스트인 경우
            followups_list = [f for f in followups_list if (f.company_id == int(company_filter) if company_filter.isdigit() else False) or (f.department and f.department.company_id == int(company_filter) if company_filter.isdigit() else False)]
        else:
            followups = followups.filter(
                Q(company_id=company_filter) | Q(department__company_id=company_filter)
            )
      
    # 정렬 (최신순)
    if level_filter:
        # 리스트인 경우 정렬
        followups_list = sorted(followups_list, key=lambda x: x.created_at, reverse=True)
        final_followups = followups_list
    else:
        followups = followups.order_by('-created_at')
        final_followups = followups
    
    # 우선순위 선택지 (필터용)
    priority_choices = FollowUp.PRIORITY_CHOICES
    
    # 고객 등급 선택지 (필터용)
    grade_choices = FollowUp.CUSTOMER_GRADE_CHOICES
    
    # 종합 점수 레벨 선택지 (필터용)
    level_choices = [
        ('critical', '🔥 최우선 (85점 이상)'),
        ('high', '⚡ 높음 (70-84점)'),
        ('medium', '⭐ 중간 (50-69점)'),
        ('low', '📋 낮음 (30-49점)'),
        ('minimal', '📌 최소 (30점 미만)'),
    ]
    
    # 업체 목록 (필터용) - 각 업체별 팔로우업 개수 계산
    accessible_users = get_accessible_users(request.user, request)
    companies = Company.objects.filter(
        Q(followup_companies__user__in=accessible_users) |
        Q(departments__followup_departments__user__in=accessible_users)
    ).distinct().order_by('name')
    
    # 각 업체별 팔로우업 개수 계산
    for company in companies:
        company.followup_count = FollowUp.objects.filter(
            Q(company=company) | Q(department__company=company),
            user__in=accessible_users
        ).count()
    
    # 선택된 우선순위 정보
    selected_priority = None
    selected_priority_display = None
    if priority_filter:
        # 유효한 우선순위 값인지 확인
        valid_priorities = [choice[0] for choice in FollowUp.PRIORITY_CHOICES]
        if priority_filter in valid_priorities:
            selected_priority = priority_filter
            # 우선순위 표시명 찾기
            for choice in FollowUp.PRIORITY_CHOICES:
                if choice[0] == priority_filter:
                    selected_priority_display = choice[1]
                    break
    
    # 선택된 업체 정보
    selected_company = None
    if company_filter:
        try:
            selected_company = Company.objects.get(id=company_filter)
        except (Company.DoesNotExist, ValueError):
            pass
    
    # 페이지네이션 처리
    paginator = Paginator(final_followups, 10) # 페이지당 10개 항목
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'followups': page_obj,
        'page_title': '팔로우업 목록', # 템플릿에 전달할 페이지 제목
        'search_query': search_query,
        'company_filter': company_filter,
        'priority_filter': priority_filter,
        'grade_filter': grade_filter,
        'level_filter': level_filter,
        'selected_priority': selected_priority,
        'selected_priority_display': selected_priority_display,
        'selected_company': selected_company,
        'total_count': stats['total_count'],
        'active_count': stats['active_count'],
        'completed_count': stats['completed_count'],
        'paused_count': stats['paused_count'],
        'priority_choices': priority_choices,
        'grade_choices': grade_choices,
        'level_choices': level_choices,
        'pipeline_stage_filter': pipeline_stage_filter,
        'pipeline_stage_choices': FollowUp.PIPELINE_STAGE_CHOICES,
        'companies': companies,
        'user_profile': user_profile,  # 사용자 프로필 추가
    }
    return render(request, 'reporting/followup_list.html', context)

# 여기에 앞으로 팔로우업 상세, 생성, 수정, 삭제 등의 뷰 함수를 추가할 예정입니다.

@login_required
def followup_detail_view(request, pk):
    """팔로우업 상세 보기 (부서 중심 - 같은 부서의 모든 고객 데이터 통합 조회)"""
    followup = get_object_or_404(FollowUp, pk=pk)
    
    # 권한 체크 (같은 회사 소속이면 고객 정보 조회 가능)
    if not can_access_followup(request.user, followup):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:followup_list')
    
    # 고객의 담당자 정보 (누가 추가했는지)
    followup_owner = followup.user
    is_own_customer = (request.user == followup_owner)
    user_profile = get_user_profile(request.user)
    
    # 부서 정보 가져오기
    department = followup.department
    company = followup.company
    
    # 같은 부서의 모든 팔로우업 (부서 중심 조회용)
    if department:
        department_followups = FollowUp.objects.filter(
            company=company,
            department=department
        ).select_related('user', 'company', 'department')
        same_department_followup_ids = list(department_followups.values_list('id', flat=True))
    else:
        department_followups = FollowUp.objects.filter(id=followup.id)
        same_department_followup_ids = [followup.id]
    
    # 같은 회사 사용자 목록 조회 (필터용) - 매니저 제외
    company_users = []
    if user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 데이터 필터 처리 (나, 전체, 특정 직원)
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')
    
    # 필터에 따른 사용자 목록 결정
    if data_filter == 'all':
        # 전체: 같은 회사 모든 사용자 (salesman만)
        if user_profile.company:
            filter_users = User.objects.filter(
                userprofile__company=user_profile.company,
                userprofile__role='salesman'
            )
        else:
            filter_users = User.objects.filter(id=request.user.id)
    elif data_filter == 'user' and filter_user_id:
        # 특정 직원
        filter_users = User.objects.filter(id=filter_user_id)
    else:
        # 나 (기본값)
        filter_users = User.objects.filter(id=request.user.id)
    
    # 히스토리 조회 (부서 기준 + 필터 적용)
    from django.db.models import Case, When, F
    related_histories = History.objects.filter(
        followup_id__in=same_department_followup_ids,
        user__in=filter_users
    ).select_related('followup', 'schedule', 'user').annotate(
        sort_date=Case(
            When(schedule__isnull=False, then=F('schedule__visit_date')),
            default=F('created_at__date')
        )
    ).order_by('-sort_date', '-created_at')[:20]
    
    # 서류 템플릿 조회 (견적서, 거래명세서 등)
    from reporting.models import DocumentTemplate
    user_company = request.user.userprofile.company
    quotation_templates = DocumentTemplate.objects.filter(
        company=user_company,
        document_type='quotation',
        is_active=True
    ).order_by('-is_default', '-created_at')
    
    transaction_templates = DocumentTemplate.objects.filter(
        company=user_company,
        document_type='transaction_statement',
        is_active=True
    ).order_by('-is_default', '-created_at')
    
    # 납품된 상품 목록 조회 (부서 기준 + 필터 적용)
    from reporting.models import DeliveryItem
    delivered_items = DeliveryItem.objects.filter(
        schedule__followup_id__in=same_department_followup_ids,  # 부서 기준
        schedule__activity_type='delivery',
        schedule__user__in=filter_users
    ).exclude(
        schedule__status='cancelled'
    ).select_related('product', 'schedule', 'schedule__user', 'schedule__followup').order_by('-schedule__visit_date', '-created_at')
    
    # 납품 품목 통계
    delivery_stats = {
        'total_items': delivered_items.count(),
        'total_revenue': delivered_items.aggregate(total=Sum('total_price'))['total'] or 0,
        'total_quantity': delivered_items.aggregate(total=Sum('quantity'))['total'] or 0,
    }
    
    # 필터에 사용된 사용자 정보
    selected_filter_user = None
    if data_filter == 'user' and filter_user_id:
        try:
            selected_filter_user = User.objects.get(id=filter_user_id)
        except User.DoesNotExist:
            pass
    
    # 연관 영업 기회 (OpportunityTracking)
    from datetime import date as date_cls
    opportunities = OpportunityTracking.objects.filter(
        followup_id__in=same_department_followup_ids
    ).select_related('followup', 'followup__company', 'followup__user').order_by(
        'current_stage', '-created_at'
    )

    # 예정 일정 (오늘 이후, 최대 5개)
    upcoming_schedules = Schedule.objects.filter(
        followup_id__in=same_department_followup_ids,
        status='scheduled',
        visit_date__gte=date_cls.today()
    ).select_related('followup', 'user', 'opportunity').order_by('visit_date', 'visit_time')[:5]

    # 최근 견적 (최대 5개)
    recent_quotes = Quote.objects.filter(
        followup_id__in=same_department_followup_ids
    ).select_related('user', 'schedule').order_by('-quote_date')[:5]

    # 페이지 제목 구성 (부서 중심)
    if department:
        page_title = f'{company.name if company else ""} - {department.name} 고객 상세'
    else:
        page_title = f'팔로우업 상세 - {followup.customer_name}'
    
    context = {
        'followup': followup,
        'department': department,
        'company': company,
        'department_followups': department_followups,  # 부서 내 모든 고객 목록
        'related_histories': related_histories,
        'quotation_templates': quotation_templates,
        'transaction_templates': transaction_templates,
        'delivered_items': delivered_items,
        'delivery_stats': delivery_stats,
        'followup_owner': followup_owner,  # 고객 담당자 (누가 추가했는지)
        'is_own_customer': is_own_customer,  # 본인이 추가한 고객인지
        'is_owner': is_own_customer,  # 템플릿 호환성을 위한 별칭
        'can_modify': can_modify_user_data(request.user, followup.user),  # 수정/삭제 권한 (관리자 포함)
        'can_view_history': True,  # 필터로 조회하므로 항상 True
        # 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'selected_filter_user': selected_filter_user,
        'company_users': company_users,
        'owner_info': {
            'username': followup_owner.username,
            'full_name': followup_owner.get_full_name() or followup_owner.username,
            'email': followup_owner.email,
        },
        'page_title': page_title,
        'opportunities': opportunities,
        'upcoming_schedules': upcoming_schedules,
        'recent_quotes': recent_quotes,
    }
    return render(request, 'reporting/followup_detail.html', context)


@login_required
def followup_create_view(request):
    """팔로우업 생성"""
    # Manager는 데이터 생성 불가 (뷰어 권한)
    user_profile = get_user_profile(request.user)
    if user_profile.is_manager():
        messages.error(request, '권한이 없습니다. Manager는 데이터를 생성할 수 없습니다.')
        return redirect('reporting:followup_list')

    if request.method == 'POST':
        form = FollowUpForm(request.POST, user=request.user)
        if form.is_valid():
            followup = form.save(commit=False)
            followup.user = request.user  # 현재 로그인한 사용자를 연결
            followup.save()
            
            messages.success(request, '고객 정보가 성공적으로 생성되었습니다.')
            return redirect('reporting:followup_detail', pk=followup.pk)
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        form = FollowUpForm(user=request.user)
    
    context = {
        'form': form,
        'page_title': '새 고객 정보 생성'
    }
    return render(request, 'reporting/followup_form.html', context)

@login_required
def followup_edit_view(request, pk):
    """팔로우업 수정"""
    followup = get_object_or_404(FollowUp, pk=pk)
    
    # 권한 체크: 수정 권한이 있는 경우만 수정 가능 (Manager는 읽기 전용)
    if not can_modify_user_data(request.user, followup.user):
        messages.error(request, '수정 권한이 없습니다. Manager는 읽기 전용입니다.')
        return redirect('reporting:followup_list')
    
    if request.method == 'POST':
        form = FollowUpForm(request.POST, instance=followup, user=request.user)
        if form.is_valid():
            updated_followup = form.save()
            
            messages.success(request, '고객 정보가 성공적으로 수정되었습니다.')
            return redirect('reporting:followup_detail', pk=followup.pk)
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        form = FollowUpForm(instance=followup, user=request.user)
    
    context = {
        'form': form,
        'followup': followup,
        'page_title': f'고객 정보 수정 - {followup.customer_name}'
    }
    return render(request, 'reporting/followup_form.html', context)

@login_required
def followup_delete_view(request, pk):
    """팔로우업 삭제"""
    followup = get_object_or_404(FollowUp, pk=pk)
    
    # 권한 체크: 삭제 권한이 있는 경우만 삭제 가능 (Manager는 읽기 전용)
    if not can_modify_user_data(request.user, followup.user):
        messages.error(request, '삭제 권한이 없습니다. Manager는 읽기 전용입니다.')
        return redirect('reporting:followup_list')
    
    if request.method == 'POST':
        customer_name = followup.customer_name or "고객명 미정"
        company_name = followup.company or "업체명 미정"
        followup.delete()
        messages.success(request, f'{customer_name} ({company_name}) 팔로우업이 삭제되었습니다.')
        return redirect('reporting:followup_list')
    
    context = {
        'followup': followup,
        'page_title': f'팔로우업 삭제 - {followup.customer_name or "고객명 미정"}'
    }
    return render(request, 'reporting/followup_delete.html', context)


OPPORTUNITY_STAGE_CHOICES = [
    ('lead', '리드'),
    ('contact', '컨택'),
    ('quote', '견적'),
    ('closing', '클로징'),
    ('won', '수주'),
    ('quote_lost', '견적실패'),
]


class OpportunityForm(forms.ModelForm):
    """영업 기회 생성/수정 폼"""

    class Meta:
        model = OpportunityTracking
        fields = [
            'title', 'label', 'current_stage',
            'expected_revenue', 'probability', 'expected_close_date',
            'lost_reason',
        ]
        widgets = {
            'title': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '예: 장비 A 구매, 소모품 정기 공급',
                'autocomplete': 'off',
            }),
            'label': forms.Select(attrs={'class': 'form-select'}),
            'current_stage': forms.Select(attrs={'class': 'form-select'}),
            'expected_revenue': forms.NumberInput(attrs={
                'class': 'form-control',
                'placeholder': '예상 매출 금액 (원)',
                'min': '0',
            }),
            'probability': forms.NumberInput(attrs={
                'class': 'form-control',
                'min': '0',
                'max': '100',
                'placeholder': '0 ~ 100',
            }),
            'expected_close_date': forms.DateInput(attrs={
                'class': 'form-control',
                'type': 'date',
            }),
            'lost_reason': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': 3,
                'placeholder': '실주 사유를 입력하세요 (선택사항)',
            }),
        }
        labels = {
            'title': '기회 제목',
            'label': '라벨',
            'current_stage': '영업 단계',
            'expected_revenue': '예상 매출 (원)',
            'probability': '수주 가능성 (%)',
            'expected_close_date': '예상 계약일',
            'lost_reason': '실주 사유',
        }

    def __init__(self, *args, user=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['label'].required = False
        self.fields['lost_reason'].required = False
        self.fields['expected_close_date'].required = False
        self.fields['title'].required = False
        # 사용자 소속 회사 라벨만 표시
        if user:
            try:
                user_company = user.userprofile.company
                self.fields['label'].queryset = OpportunityLabel.objects.filter(
                    user_company=user_company, is_active=True
                )
            except Exception:
                self.fields['label'].queryset = OpportunityLabel.objects.none()
        else:
            self.fields['label'].queryset = OpportunityLabel.objects.none()

    def clean_probability(self):
        prob = self.cleaned_data.get('probability')
        if prob is not None and not (0 <= prob <= 100):
            raise forms.ValidationError('수주 가능성은 0~100 사이의 값이어야 합니다.')
        return prob


@login_required
def opportunity_create_view(request, followup_pk):
    """영업 기회 생성 뷰"""
    followup = get_object_or_404(FollowUp, pk=followup_pk)
    if not can_access_followup(request.user, followup):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:followup_list')

    if request.method == 'POST':
        form = OpportunityForm(request.POST, user=request.user)
        if form.is_valid():
            opportunity = form.save(commit=False)
            opportunity.followup = followup
            opportunity.save()
            messages.success(request, '영업 기회가 등록되었습니다.')
            return redirect('reporting:opportunity_detail', pk=opportunity.pk)
    else:
        form = OpportunityForm(user=request.user)

    context = {
        'form': form,
        'followup': followup,
        'page_title': f'영업 기회 등록 — {followup.customer_name or "고객명 미정"}',
        'is_create': True,
        'stage_choices': OPPORTUNITY_STAGE_CHOICES,
    }
    return render(request, 'reporting/opportunity_form.html', context)


@login_required
def opportunity_edit_view(request, pk):
    """영업 기회 수정 뷰"""
    opportunity = get_object_or_404(
        OpportunityTracking.objects.select_related('followup'),
        pk=pk,
    )
    if not can_access_followup(request.user, opportunity.followup):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:opportunity_list')

    if request.method == 'POST':
        form = OpportunityForm(request.POST, instance=opportunity, user=request.user)
        if form.is_valid():
            updated = form.save(commit=False)
            # current_stage 변경 시 이력 기록
            if updated.current_stage != opportunity.current_stage:
                updated.update_stage(updated.current_stage)
            else:
                updated.save()
            messages.success(request, '영업 기회가 수정되었습니다.')
            return redirect('reporting:opportunity_detail', pk=opportunity.pk)
    else:
        form = OpportunityForm(instance=opportunity, user=request.user)

    context = {
        'form': form,
        'followup': opportunity.followup,
        'opportunity': opportunity,
        'page_title': f'영업 기회 수정 — {opportunity.title or "(제목 없음)"}',
        'is_create': False,
        'stage_choices': OPPORTUNITY_STAGE_CHOICES,
    }
    return render(request, 'reporting/opportunity_form.html', context)


@login_required
def opportunity_list_view(request):
    """영업 기회 목록 (자신 담당 고객 기준, 매니저는 팀 전체)"""
    from datetime import date as date_cls
    user_profile = get_user_profile(request.user)
    accessible_users = get_accessible_users(request.user, request)

    opportunities = OpportunityTracking.objects.filter(
        followup__user__in=accessible_users
    ).select_related(
        'followup', 'followup__company', 'followup__department', 'followup__user'
    ).order_by('-updated_at')

    # 단계 필터
    stage_filter = request.GET.get('stage', '')
    if stage_filter:
        opportunities = opportunities.filter(current_stage=stage_filter)

    # 예상 계약일 필터
    close_filter = request.GET.get('close', '')
    today = date_cls.today()
    if close_filter == 'this_month':
        opportunities = opportunities.filter(
            expected_close_date__year=today.year,
            expected_close_date__month=today.month
        )
    elif close_filter == 'overdue':
        opportunities = opportunities.filter(
            expected_close_date__lt=today,
            current_stage__in=['lead', 'contact', 'quote', 'closing']
        )

    # 통계
    stats = opportunities.aggregate(
        total=Count('id'),
        total_expected=Sum('expected_revenue'),
        total_weighted=Sum('weighted_revenue'),
    )

    # 단계별 개수 (필터 전 전체 기준)
    all_opps = OpportunityTracking.objects.filter(followup__user__in=accessible_users)
    stage_counts = {}
    for value, _label in OPPORTUNITY_STAGE_CHOICES:
        stage_counts[value] = all_opps.filter(current_stage=value).count()

    paginator = Paginator(opportunities, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'opportunities': page_obj,
        'stage_filter': stage_filter,
        'close_filter': close_filter,
        'stage_choices': OPPORTUNITY_STAGE_CHOICES,
        'stats': stats,
        'stage_counts': stage_counts,
        'today': today,
        'page_title': '영업 기회 목록',
        'user_profile': user_profile,
    }
    return render(request, 'reporting/opportunity_list.html', context)


@login_required
def opportunity_detail_view(request, pk):
    """영업 기회 상세 (조회 전용)"""
    opportunity = get_object_or_404(
        OpportunityTracking.objects.select_related(
            'followup', 'followup__company', 'followup__department', 'followup__user'
        ),
        pk=pk
    )
    # 권한 체크: 고객 접근 가능 여부로 판단
    if not can_access_followup(request.user, opportunity.followup):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:opportunity_list')

    # 연관 일정 (최신 10개)
    related_schedules = Schedule.objects.filter(
        opportunity=opportunity
    ).select_related('followup', 'user').order_by('-visit_date', '-visit_time')[:10]

    # 연관 견적 (최신 5개)
    related_quotes = Quote.objects.filter(
        followup=opportunity.followup
    ).select_related('user', 'schedule').order_by('-quote_date')[:5]

    # 연관 히스토리 (최신 5개)
    related_histories = History.objects.filter(
        followup=opportunity.followup,
        parent_history__isnull=True
    ).select_related('user').order_by('-created_at')[:5]

    context = {
        'opportunity': opportunity,
        'followup': opportunity.followup,
        'related_schedules': related_schedules,
        'related_quotes': related_quotes,
        'related_histories': related_histories,
        'stage_choices': OPPORTUNITY_STAGE_CHOICES,
        'page_title': f'영업 기회 - {opportunity}',
    }
    return render(request, 'reporting/opportunity_detail.html', context)


@login_required
@never_cache
def dashboard_view(request):
    from django.db.models import Count, Sum
    from django.utils import timezone
    from datetime import datetime, timedelta
    from reporting.models import DeliveryItem, PrepaymentUsage
    import calendar
    import logging
    
    logger = logging.getLogger(__name__)
    
    # 사용자 프로필 가져오기
    user_profile = get_user_profile(request.user)
    
    # URL 파라미터로 특정 사용자 필터링
    user_filter = request.GET.get('user')
    view_all = request.GET.get('view_all') == 'true'
    selected_user = None
    
    if user_profile.can_view_all_users():
        # 전체 팀원 선택 시 세션 초기화
        if view_all:
            if 'selected_user_id' in request.session:
                del request.session['selected_user_id']
            target_user = None  # 전체 팀원 데이터 표시
        else:
            # user_filter가 없으면 세션에서 가져오기
            if not user_filter:
                user_filter = request.session.get('selected_user_id')
            
            if user_filter:
                try:
                    selected_user = User.objects.get(id=user_filter)
                    target_user = selected_user
                    # 세션에 저장
                    request.session['selected_user_id'] = str(user_filter)
                except (User.DoesNotExist, ValueError):
                    target_user = None  # 전체 팀원 데이터 표시
                    # 잘못된 세션 값 제거
                    if 'selected_user_id' in request.session:
                        del request.session['selected_user_id']
            else:
                target_user = None  # 전체 팀원 데이터 표시
    else:
        target_user = request.user
    
    # 매니저용 팀원 목록
    salesman_users = []
    if user_profile.can_view_all_users():
        salesman_users = get_accessible_users(request.user, request)
    
    # 현재 연도와 월 가져오기
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # 권한에 따른 데이터 필터링
    if user_profile.is_admin() and not selected_user:
        # Admin은 필터링된 데이터 접근
        if hasattr(request, 'admin_filter_user') and request.admin_filter_user:
            # 특정 사용자 선택됨
            accessible_users = User.objects.filter(id=request.admin_filter_user.id)
            followup_count = FollowUp.objects.filter(user__in=accessible_users).count()
            schedule_count = Schedule.objects.filter(user__in=accessible_users, status='scheduled').count()
            sales_record_count = History.objects.filter(
                user__in=accessible_users,
                created_at__year=current_year, 
                action_type__in=['customer_meeting', 'delivery_schedule']
            ).count()
            histories = History.objects.filter(user__in=accessible_users)
            histories_current_year = History.objects.filter(user__in=accessible_users, created_at__year=current_year)
            schedules = Schedule.objects.filter(user__in=accessible_users)
            followups = FollowUp.objects.filter(user__in=accessible_users)
        elif hasattr(request, 'admin_filter_company') and request.admin_filter_company:
            # 특정 회사 선택됨
            accessible_users = User.objects.filter(
                userprofile__company=request.admin_filter_company,
                userprofile__role__in=['salesman', 'manager']
            )
            followup_count = FollowUp.objects.filter(user__in=accessible_users).count()
            schedule_count = Schedule.objects.filter(user__in=accessible_users, status='scheduled').count()
            sales_record_count = History.objects.filter(
                user__in=accessible_users,
                created_at__year=current_year, 
                action_type__in=['customer_meeting', 'delivery_schedule']
            ).count()
            histories = History.objects.filter(user__in=accessible_users)
            histories_current_year = History.objects.filter(user__in=accessible_users, created_at__year=current_year)
            schedules = Schedule.objects.filter(user__in=accessible_users)
            followups = FollowUp.objects.filter(user__in=accessible_users)
        else:
            # 전체 데이터
            followup_count = FollowUp.objects.count()
            schedule_count = Schedule.objects.filter(status='scheduled').count()
            sales_record_count = History.objects.filter(
                created_at__year=current_year, 
                action_type__in=['customer_meeting', 'delivery_schedule']
            ).count()
            histories = History.objects.all()
            histories_current_year = History.objects.filter(created_at__year=current_year)
            schedules = Schedule.objects.all()
            followups = FollowUp.objects.all()
    elif user_profile.can_view_all_users() and target_user is None:
        # Manager가 전체 팀원을 선택한 경우 - 접근 가능한 모든 사용자의 데이터
        accessible_users = get_accessible_users(request.user, request)
        followup_count = FollowUp.objects.filter(user__in=accessible_users).count()
        schedule_count = Schedule.objects.filter(user__in=accessible_users, status='scheduled').count()
        sales_record_count = History.objects.filter(
            user__in=accessible_users,
            created_at__year=current_year, 
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
        histories = History.objects.filter(user__in=accessible_users)
        histories_current_year = History.objects.filter(user__in=accessible_users, created_at__year=current_year)
        schedules = Schedule.objects.filter(user__in=accessible_users)
        followups = FollowUp.objects.filter(user__in=accessible_users)
    else:
        # 특정 사용자 또는 본인의 데이터만 접근
        followup_count = FollowUp.objects.filter(user=target_user).count()
        schedule_count = Schedule.objects.filter(user=target_user, status='scheduled').count()
        # 영업 기록 (미팅, 납품만 카운팅 - 서비스 제외)
        sales_record_count = History.objects.filter(
            user=target_user, 
            created_at__year=current_year, 
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
        histories = History.objects.filter(user=target_user)
        histories_current_year = History.objects.filter(user=target_user, created_at__year=current_year)
        schedules = Schedule.objects.filter(user=target_user)
        followups = FollowUp.objects.filter(user=target_user)

    # 올해 매출 통계 - 고객 리포트와 동일한 로직 사용
    # 1. Schedule에 연결 안된 History - History 금액 사용
    # 2. Schedule에 연결된 History - DeliveryItem이 있으면 DeliveryItem, 없으면 History 금액
    # 3. History에 연결 안된 Schedule - DeliveryItem 금액
    from decimal import Decimal
    
    if user_profile.is_admin() and not selected_user:
        if hasattr(request, 'admin_filter_user') and request.admin_filter_user:
            user_filter_for_dashboard = {'user__in': User.objects.filter(id=request.admin_filter_user.id)}
        elif hasattr(request, 'admin_filter_company') and request.admin_filter_company:
            user_filter_for_dashboard = {'user__in': User.objects.filter(
                userprofile__company=request.admin_filter_company,
                userprofile__role__in=['salesman', 'manager']
            )}
        else:
            user_filter_for_dashboard = {}  # 전체
    elif user_profile.can_view_all_users() and target_user is None:
        accessible_users = get_accessible_users(request.user, request)
        user_filter_for_dashboard = {'user__in': accessible_users}
    else:
        user_filter_for_dashboard = {'user': target_user}
    
    # 납품 일정 조회 (올해, 완료된 것만 - 일정 페이지와 동일한 조건)
    today = timezone.now().date()
    delivery_schedules = Schedule.objects.filter(
        visit_date__year=current_year,
        visit_date__lte=today,  # 오늘까지만 포함 (미래 일정 제외)
        activity_type='delivery',
        status='completed',  # 완료된 납품만 집계
        **user_filter_for_dashboard
    ).prefetch_related('delivery_items_set')
    
    # 납품 History 조회 (올해)
    delivery_histories = History.objects.filter(
        action_type='delivery_schedule',
        created_at__year=current_year,
        **user_filter_for_dashboard
    )
    
    # 매출 계산 - Schedule 기준 (중복 방지)
    total_delivery_amount = Decimal('0')
    delivery_count = 0
    processed_schedule_ids = set()
    
    # 1단계: 완료된 Schedule의 DeliveryItem 금액 집계
    for schedule in delivery_schedules:
        items = list(schedule.delivery_items_set.all())
        schedule_amount = sum(item.total_price or Decimal('0') for item in items)
        
        if schedule_amount > 0:
            # DeliveryItem이 있으면 DeliveryItem 금액 사용
            total_delivery_amount += schedule_amount
        else:
            # DeliveryItem이 없으면 연결된 History 중 가장 최근 것의 금액 사용
            related_history = delivery_histories.filter(schedule_id=schedule.id).order_by('-created_at').first()
            if related_history:
                total_delivery_amount += related_history.delivery_amount or Decimal('0')
        
        processed_schedule_ids.add(schedule.id)
        delivery_count += 1
    
    # 2단계: Schedule에 연결되지 않은 History의 금액 추가 (중복 방지)
    standalone_histories = delivery_histories.filter(schedule_id__isnull=True)
    for h in standalone_histories:
        total_delivery_amount += h.delivery_amount or Decimal('0')
        delivery_count += 1
    
    # 활동 유형별 통계 (Schedule 기준, 현재 연도)
    schedules_current_year_filter = schedules.filter(visit_date__year=current_year)
    
    if getattr(request, 'is_admin', False) or getattr(request, 'is_hanagwahak', False):
        activity_stats = schedules_current_year_filter.values('activity_type').annotate(
            count=Count('id')
        ).order_by('activity_type')
    else:
        # Admin이 아니고 하나과학이 아닌 경우 서비스 항목 제외
        activity_stats = schedules_current_year_filter.exclude(activity_type='service').values('activity_type').annotate(
            count=Count('id')
        ).order_by('activity_type')
    
    # 서비스 통계 추가 (완료된 서비스만 카운팅) - Admin이나 하나과학만
    if getattr(request, 'is_admin', False) or getattr(request, 'is_hanagwahak', False):
        service_count = schedules_current_year_filter.filter(activity_type='service', status='completed').count()
        
        # 이번 달 서비스 수 (완료된 것만)
        this_month_service_count = schedules.filter(
            activity_type='service',
            status='completed',
            visit_date__month=current_month,
            visit_date__year=current_year
        ).count()
    else:
        service_count = 0
        this_month_service_count = 0
    
    # 최근 활동 (현재 연도, 최근 5개, 메모 제외)
    recent_activities_queryset = histories_current_year.exclude(action_type='memo')
    if not getattr(request, 'is_admin', False) and not getattr(request, 'is_hanagwahak', False):
        # Admin이 아니고 하나과학이 아닌 경우 서비스도 제외
        recent_activities_queryset = recent_activities_queryset.exclude(action_type='service')
        
    recent_activities = recent_activities_queryset.order_by('-created_at')[:5]
    
    # 월별 고객 추가 현황 (최근 6개월)
    now = timezone.now()
    monthly_customers = []
    for i in range(6):
        month_start = (now.replace(day=1) - timedelta(days=32*i)).replace(day=1)
        month_end = (month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1]))
        
        count = followups.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        ).count()
        
        monthly_customers.append({
            'month': month_start.strftime('%Y-%m'),
            'month_name': f"{month_start.year}년 {month_start.month}월",
            'count': count
        })
    
    monthly_customers.reverse()  # 시간순 정렬
    
    # 일정 완료율 통계
    schedule_stats = schedules.aggregate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        cancelled=Count('id', filter=Q(status='cancelled')),
        scheduled=Count('id', filter=Q(status='scheduled'))
    )
    
    completion_rate = 0
    if schedule_stats['total'] > 0:
        completion_rate = round((schedule_stats['completed'] / schedule_stats['total']) * 100, 1)
      # 영업 기록 추이 (최근 14일, 미팅/납품만 - 서비스 제외)
    fourteen_days_ago = now - timedelta(days=14)
    daily_activities = []
    for i in range(14):
        day = fourteen_days_ago + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        # 영업 활동만 카운팅 (미팅, 납품 - 서비스 제외)
        activity_count = histories_current_year.filter(
            created_at__gte=day_start,
            created_at__lt=day_end,
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
        
        daily_activities.append({
            'date': day.strftime('%m/%d'),
            'full_date': day.strftime('%Y-%m-%d'),
            'count': activity_count
        })
    
    # 오늘 일정
    today = now.date()
    today_schedules = schedules.filter(
        visit_date=today,
        status='scheduled'
    ).order_by('visit_time')[:5]

    # 오늘 완료된 일정 중 히스토리 미작성 건수 (담당자용)
    from django.db.models import Exists, OuterRef
    today_unwritten_count = schedules.filter(
        visit_date=today,
        status='completed',
    ).exclude(
        Exists(History.objects.filter(schedule_id=OuterRef('pk'), parent_history__isnull=True))
    ).count()

    # 관리자용: 최근 30일 미검토 보고서 수
    pending_review_count = 0
    if user_profile.can_view_all_users():
        thirty_days_ago = now - timedelta(days=30)
        review_users = get_accessible_users(request.user, request)
        pending_review_count = History.objects.filter(
            user__in=review_users,
            action_type__in=['customer_meeting', 'delivery_schedule', 'quote', 'service'],
            parent_history__isnull=True,
            reviewed_at__isnull=True,
            created_at__gte=thirty_days_ago,
        ).count()

    # 최근 고객 (최근 7일)
    week_ago = now - timedelta(days=7)
    recent_customers = followups.filter(
        created_at__gte=week_ago
    ).order_by('-created_at')[:5]

    # 새로운 성과 지표 계산
    from django.db.models import Avg
    current_month = now.month
    current_year = now.year
    
    # ========================================
    # 이번 달 활동 현황 (일정 기준)
    # ========================================
    # 이번 달 시작/끝 날짜
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if current_month == 12:
        month_end = now.replace(year=current_year+1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        month_end = now.replace(month=current_month+1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # 이번 달 마지막 날 계산 (URL 표시용)
    from calendar import monthrange
    last_day = monthrange(current_year, current_month)[1]
    month_last_date = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)
    
    # 이번 달 미팅 일정 수 (일정 기준)
    monthly_meetings = schedules.filter(
        activity_type='customer_meeting',
        visit_date__gte=month_start.date(),
        visit_date__lt=month_end.date()
    ).count()
    
    # 처리해야 할 견적 (예정 상태의 견적만)
    # 견적은 납품 생성 시 자동으로 완료 처리되므로, 예정 상태인 것만 처리 대상
    quote_count = schedules.filter(
        activity_type='quote',
        status='scheduled'
    ).count()
    
    # 이번 달 견적 횟수 (이번 달의 모든 견적 일정)
    monthly_quote_count = schedules.filter(
        activity_type='quote',
        visit_date__gte=month_start.date(),
        visit_date__lt=month_end.date()
    ).count()
    
    # 이번 달 납품 일정 수 (취소된 일정 제외)
    monthly_delivery_count = schedules.filter(
        activity_type='delivery',
        status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        visit_date__gte=month_start.date(),
        visit_date__lt=month_end.date()
    ).count()
    
    # 이번 달 매출 (납품 일정의 DeliveryItem 총액)
    monthly_delivery_schedules = schedules.filter(
        activity_type='delivery',
        status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        visit_date__gte=month_start.date(),
        visit_date__lt=month_end.date()
    )
    
    # 납품 일정의 DeliveryItem 총액 (선결제 여부 상관없이 전체)
    monthly_revenue = DeliveryItem.objects.filter(
        schedule__in=monthly_delivery_schedules
    ).aggregate(total=Sum('total_price'))['total'] or 0
    
    from .models import Prepayment, PrepaymentUsage
    
    # 이번 달 선결제 건수 (결제일 기준) - 권한에 따라 필터링
    if user_profile.is_admin() and not selected_user:
        # Admin은 모든 선결제
        monthly_prepayment_count = Prepayment.objects.filter(
            payment_date__year=current_year,
            payment_date__month=current_month
        ).count()
    elif user_profile.can_view_all_users() and target_user is None:
        # Manager가 전체 팀원을 선택한 경우
        accessible_users = get_accessible_users(request.user, request)
        monthly_prepayment_count = Prepayment.objects.filter(
            created_by__in=accessible_users,
            payment_date__year=current_year,
            payment_date__month=current_month
        ).count()
    else:
        # 특정 사용자 또는 본인
        monthly_prepayment_count = Prepayment.objects.filter(
            created_by=target_user,
            payment_date__year=current_year,
            payment_date__month=current_month
        ).count()
    
    # 이번 달 서비스 수 (완료된 것만)
    monthly_services = histories.filter(
        action_type='service',
        service_status='completed',
        created_at__month=current_month,
        created_at__year=current_year
    ).count()
      # 전환율 계산 (현재 연도 기준)
    # Schedule 테이블 기반으로 계산 (현재 활성 데이터)
    schedules_current_year = schedules.filter(visit_date__year=current_year)
    
    total_meetings = schedules_current_year.filter(activity_type='customer_meeting').count()
    total_quotes = schedules_current_year.filter(activity_type='quote').count()
    total_deliveries = schedules_current_year.filter(activity_type='delivery', status__in=['scheduled', 'completed']).count()  # 취소된 일정 제외
    
    # 견적 → 납품 전환율: 같은 고객에 견적과 납품이 모두 있는 비율
    quote_schedules = schedules_current_year.filter(activity_type='quote')
    quotes_with_delivery = 0
    
    for quote in quote_schedules:
        # 같은 고객(followup)에 납품 일정이 있는지 확인
        has_delivery = schedules.filter(
            followup=quote.followup,
            activity_type='delivery'
        ).exists()
        if has_delivery:
            quotes_with_delivery += 1
    
    # 미팅 → 납품 전환율 (일정 기준: 일정 중 미팅 건수 대비 납품 완료 건수)
    schedule_meetings = schedules_current_year.filter(activity_type='customer_meeting').count()
    schedule_deliveries_completed = schedules_current_year.filter(activity_type='delivery', status='completed').count()
    meeting_to_delivery_rate = (schedule_deliveries_completed / schedule_meetings * 100) if schedule_meetings > 0 else 0
    
    # 견적 → 납품 전환율 (견적을 낸 것 중 납품으로 전환된 비율)
    quote_to_delivery_rate = (quotes_with_delivery / total_quotes * 100) if total_quotes > 0 else 0
    
    # 평균 거래 규모 (현재 연도 기준)
    avg_deal_size = histories_current_year.filter(
        action_type='delivery_schedule',
        delivery_amount__isnull=False
    ).aggregate(avg=Avg('delivery_amount'))['avg'] or 0
    
    # 월별 납품 금액 데이터 (최근 6개월)
    monthly_revenue_data = []
    monthly_revenue_labels = []
    
    for i in range(5, -1, -1):
        target_date = now - timedelta(days=30*i)
        month_revenue = histories.filter(
            action_type='delivery_schedule',
            created_at__month=target_date.month,
            created_at__year=target_date.year,
            delivery_amount__isnull=False
        ).aggregate(total=Sum('delivery_amount'))['total'] or 0
        
        monthly_revenue_data.append(float(month_revenue))
        monthly_revenue_labels.append(f"{target_date.year}년 {target_date.month}월")
      # 고객별 납품 현황 (현재 연도 기준, 상위 5개)
    customer_revenue_data = histories_current_year.filter(
        action_type='delivery_schedule',
        delivery_amount__isnull=False,
        followup__isnull=False
    ).values('followup__customer_name', 'followup__company').annotate(
        total_revenue=Sum('delivery_amount')
    ).order_by('-total_revenue')[:5]
    
    customer_labels = [f"{item['followup__customer_name'] or '미정'} ({item['followup__company'] or '미정'})" for item in customer_revenue_data]
    customer_amounts = [float(item['total_revenue']) for item in customer_revenue_data]

    # 월별 서비스 데이터 (최근 6개월, 완료된 서비스만) - Admin이나 하나과학만
    if getattr(request, 'is_admin', False) or getattr(request, 'is_hanagwahak', False):
        monthly_service_data = []
        monthly_service_labels = []
        for i in range(5, -1, -1):
            target_date = now - timedelta(days=30*i)
            service_count_monthly = histories.filter(
                action_type='service',
                service_status='completed',
                created_at__month=target_date.month,
                created_at__year=target_date.year
            ).count()
            
            monthly_service_data.append(service_count_monthly)
            monthly_service_labels.append(f"{target_date.year}년 {target_date.month}월")
    else:
        monthly_service_data = []
        monthly_service_labels = []

    # ============================================
    # 📊 새로운 7개 차트를 위한 데이터 준비 - Schedule 기준
    # ============================================
    
    # 1️⃣ 매출 및 납품 추이 (월별 납품 금액 + 건수) - Schedule 기준
    monthly_delivery_stats = {
        'labels': [],
        'amounts': [],
        'counts': []
    }
    
    for i in range(11, -1, -1):  # 최근 12개월
        target_date = now - timedelta(days=30*i)
        month_start = target_date.replace(day=1)
        if target_date.month == 12:
            month_end = target_date.replace(year=target_date.year+1, month=1, day=1)
        else:
            month_end = target_date.replace(month=target_date.month+1, day=1)
        
        month_schedules = schedules.filter(
            visit_date__gte=month_start.date(),
            visit_date__lt=month_end.date(),
            activity_type='delivery',
            status__in=['scheduled', 'completed']  # 취소된 일정 제외
        )
        
        # 납품 금액 합산
        month_amount = DeliveryItem.objects.filter(
            schedule__in=month_schedules
        ).aggregate(total=Sum('total_price'))['total'] or 0
        
        # 납품 건수 (일정 개수)
        month_count = month_schedules.count()
        
        monthly_delivery_stats['labels'].append(f"{target_date.month}월")
        monthly_delivery_stats['amounts'].append(float(month_amount))
        monthly_delivery_stats['counts'].append(month_count)
    
    # 2️⃣ 영업 퍼널 (미팅 → 견적 제출 → 발주 예정 → 납품 완료)
    # 기준: 모두 일정(Schedule) 기반으로 집계
    schedules_current_year = schedules.filter(visit_date__year=current_year)
    
    meeting_count = schedules_current_year.filter(activity_type='customer_meeting').count()
    quote_count_funnel = schedules_current_year.filter(activity_type='quote').count()
    scheduled_delivery_count = schedules_current_year.filter(activity_type='delivery', status='scheduled').count()
    completed_delivery_count = schedules_current_year.filter(activity_type='delivery', status='completed').count()
    
    sales_funnel = {
        'stages': ['미팅', '견적 제출', '발주 예정', '납품 완료'],
        'values': [
            meeting_count,
            quote_count_funnel,
            scheduled_delivery_count,
            completed_delivery_count
        ]
    }
    
    # 3️⃣ 고객사별 매출 비중 (Top 5 + 기타) - Schedule + History 기준
    # Schedule 기반 매출
    schedule_top_customers = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        schedule__followup__isnull=False,
        schedule__followup__company__isnull=False
    ).values('schedule__followup__company__name').annotate(
        total_revenue=Sum('total_price')
    )
    
    # History 기반 매출
    histories_current_year_with_company = histories_current_year.filter(
        followup__isnull=False,
        followup__company__isnull=False
    )
    
    history_top_customers = DeliveryItem.objects.filter(
        history__in=histories_current_year_with_company
    ).values('history__followup__company__name').annotate(
        total_revenue=Sum('total_price')
    )
    
    # 고객사별 매출 합산
    from collections import defaultdict
    company_revenue = defaultdict(float)
    
    for item in schedule_top_customers:
        company_name = item['schedule__followup__company__name'] or '미정'
        company_revenue[company_name] += float(item['total_revenue'])
    
    for item in history_top_customers:
        company_name = item['history__followup__company__name'] or '미정'
        company_revenue[company_name] += float(item['total_revenue'])
    
    # 상위 5개 추출
    sorted_companies = sorted(company_revenue.items(), key=lambda x: x[1], reverse=True)[:5]
    
    customer_distribution = {
        'labels': [],
        'data': []
    }
    
    total_top5_revenue = 0
    for company_name, revenue in sorted_companies:
        customer_distribution['labels'].append(company_name)
        customer_distribution['data'].append(revenue)
        total_top5_revenue += revenue
    
    # 기타 금액 계산 - Schedule + History 합산
    schedule_total = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__status__in=['scheduled', 'completed']
    ).aggregate(total=Sum('total_price'))['total'] or 0
    
    history_total = DeliveryItem.objects.filter(
        history__in=histories_current_year
    ).aggregate(total=Sum('total_price'))['total'] or 0
    
    total_all_revenue = float(schedule_total) + float(history_total)
    other_revenue = total_all_revenue - total_top5_revenue
    if other_revenue > 0:
        customer_distribution['labels'].append('기타')
        customer_distribution['data'].append(other_revenue)
    
    # 6️⃣ 고객 유형별 통계 (대학/기업/관공서) - Schedule + History 기준
    customer_type_stats = {
        'labels': ['대학', '기업', '관공서'],
        'revenue': [0, 0, 0],
        'count': [0, 0, 0]
    }
    
    # TODO: Company 모델에 customer_type 필드 추가 후 활성화
    # 현재는 company name으로 간단히 분류 (예: 대학교 포함 여부 등)
    
    # Schedule 기반 통계 (납품 일정만 카운트 - 견적 제외)
    schedule_company_stats = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__activity_type='delivery',  # 납품만 카운트 (견적 제외)
        schedule__status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        schedule__followup__isnull=False,
        schedule__followup__company__isnull=False
    ).values('schedule__followup__company__name').annotate(
        total_revenue=Sum('total_price'),
        count=Count('schedule', distinct=True)  # 일정 개수로 카운팅
    )
    
    # 간단한 키워드 기반 분류 (연구소 제외)
    for item in schedule_company_stats:
        company_name = item['schedule__followup__company__name'] or ''
        revenue = float(item['total_revenue']) / 1000000  # 백만원 단위
        cnt = item['count']
        
        if '대학' in company_name or '대학교' in company_name:
            customer_type_stats['revenue'][0] += revenue
            customer_type_stats['count'][0] += cnt
        elif '청' in company_name or '부' in company_name or '시' in company_name or '구' in company_name:
            customer_type_stats['revenue'][2] += revenue
            customer_type_stats['count'][2] += cnt
        else:
            # 연구소 포함 모든 기타 기업
            customer_type_stats['revenue'][1] += revenue
            customer_type_stats['count'][1] += cnt
    
    # History 기반 통계 (납품 이력만 카운트 - 견적 제외)
    history_company_stats = DeliveryItem.objects.filter(
        history__in=histories_current_year,
        history__action_type='delivery_schedule',  # 납품만 카운트 (History는 action_type 필드 사용)
        history__followup__isnull=False,
        history__followup__company__isnull=False
    ).values('history__followup__company__name').annotate(
        total_revenue=Sum('total_price'),
        count=Count('id')
    )
    
    for item in history_company_stats:
        company_name = item['history__followup__company__name'] or ''
        revenue = float(item['total_revenue']) / 1000000  # 백만원 단위
        cnt = item['count']
        
        if '대학' in company_name or '대학교' in company_name:
            customer_type_stats['revenue'][0] += revenue
            customer_type_stats['count'][0] += cnt
        elif '청' in company_name or '부' in company_name or '시' in company_name or '구' in company_name:
            customer_type_stats['revenue'][2] += revenue
            customer_type_stats['count'][2] += cnt
        else:
            # 연구소 포함 모든 기타 기업
            customer_type_stats['revenue'][1] += revenue
            customer_type_stats['count'][1] += cnt

    
    # 7️⃣ 활동 히트맵 (현재 달) - 현재 사용자의 일정만
    daily_activity_heatmap = []
    
    # 현재 달의 첫날과 마지막 날 계산
    current_month_start = now.replace(day=1).date()
    if now.month == 12:
        next_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        next_month = now.replace(month=now.month + 1, day=1)
    current_month_end = (next_month - timedelta(days=1)).date()
    
    # 현재 달의 각 날짜별 활동 카운트 (현재 사용자만)
    current_date = current_month_start
    while current_date <= current_month_end:
        day_activity_count = schedules.filter(
            visit_date=current_date
        ).count()
        
        daily_activity_heatmap.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_of_week': current_date.weekday(),  # 0=월, 6=일
            'intensity': day_activity_count
        })
        
        current_date += timedelta(days=1)

    # 종합 점수 분포 계산
    priority_level_distribution = {
        'critical': 0,  # 최우선 85+
        'high': 0,      # 높음 70-84
        'medium': 0,    # 중간 50-69
        'low': 0,       # 낮음 30-49
        'minimal': 0,   # 최소 30-
    }
    
    # 모든 팔로우업의 종합 점수 계산
    for followup in followups:
        combined_score = followup.get_combined_score()
        if combined_score >= 85:
            priority_level_distribution['critical'] += 1
        elif combined_score >= 70:
            priority_level_distribution['high'] += 1
        elif combined_score >= 50:
            priority_level_distribution['medium'] += 1
        elif combined_score >= 30:
            priority_level_distribution['low'] += 1
        else:
            priority_level_distribution['minimal'] += 1
    
    # 차트 데이터로 변환
    priority_level_chart = {
        'labels': ['🔥 최우선', '⚡ 높음', '⭐ 중간', '📋 낮음', '📌 최소'],
        'data': [
            priority_level_distribution['critical'],
            priority_level_distribution['high'],
            priority_level_distribution['medium'],
            priority_level_distribution['low'],
            priority_level_distribution['minimal'],
        ],
        'colors': ['#ef4444', '#f59e0b', '#3b82f6', '#10b981', '#6b7280'],
    }

    context = {        'page_title': '대시보드',
        'current_year': current_year,  # 현재 연도 정보 추가
        'selected_user': selected_user,  # 선택된 사용자 정보
        'target_user': target_user,  # 실제 대상 사용자
        'salesman_users': salesman_users,  # 매니저용 실무자 목록
        'view_all': False,  # 현재 전체보기 기능은 미사용
        'followup_count': followup_count,
        'schedule_count': schedule_count,
        'sales_record_count': sales_record_count,
        'total_delivery_amount': total_delivery_amount,
        'delivery_count': delivery_count,
        'activity_stats': activity_stats,
        'recent_activities': recent_activities,
        'monthly_customers': monthly_customers,
        'schedule_stats': schedule_stats,
        'completion_rate': completion_rate,
        'daily_activities': daily_activities,
        'today_schedules': today_schedules,        
        'recent_customers': recent_customers,
        'monthly_revenue': monthly_revenue,
        'monthly_meetings': monthly_meetings,
        'monthly_services': monthly_services,
        'service_count': service_count,
        'this_month_service_count': this_month_service_count,
        'meeting_to_delivery_rate': meeting_to_delivery_rate,
        'quote_to_delivery_rate': quote_to_delivery_rate,
        'avg_deal_size': avg_deal_size,
        # 템플릿에서 직접 사용할 수 있도록 추가
        'is_hanagwahak': getattr(request, 'is_hanagwahak', False),
        'is_admin': getattr(request, 'is_admin', False),
        'user_company_name': getattr(request, 'user_company_name', None),
        'monthly_revenue_data': monthly_revenue_data,
        'monthly_revenue_labels': monthly_revenue_labels,        'customer_revenue_labels': customer_labels,
        'customer_revenue_data': customer_amounts,
        'monthly_service_data': monthly_service_data,
        'monthly_service_labels': monthly_service_labels,
        # 새로운 차트 데이터
        'monthly_delivery_stats': json.dumps(monthly_delivery_stats, cls=DjangoJSONEncoder),
        'sales_funnel': json.dumps(sales_funnel, cls=DjangoJSONEncoder),
        'customer_distribution': json.dumps(customer_distribution, cls=DjangoJSONEncoder),
        'customer_type_stats': json.dumps(customer_type_stats, cls=DjangoJSONEncoder),
        'daily_activity_heatmap': json.dumps(daily_activity_heatmap, cls=DjangoJSONEncoder),
        'priority_level_chart': json.dumps(priority_level_chart, cls=DjangoJSONEncoder),
    }
    
    # 선결제 통계 추가
    from reporting.models import Prepayment
    from decimal import Decimal
    
    # 선결제 조회 - 등록자 본인만 (Manager도 자신이 등록한 것만)
    prepayments = Prepayment.objects.filter(created_by=target_user)
    
    # 선결제 통계 계산
    prepayment_total = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance')
    )
    
    prepayment_stats = {
        'total_amount': prepayment_total['total_amount'] or Decimal('0'),
        'total_balance': prepayment_total['total_balance'] or Decimal('0'),
        'total_used': (prepayment_total['total_amount'] or Decimal('0')) - (prepayment_total['total_balance'] or Decimal('0')),
        'active_count': prepayments.filter(status='active', balance__gt=0).count(),
        'depleted_count': prepayments.filter(status='depleted').count(),
        'total_count': prepayments.count(),
        'monthly_count': monthly_prepayment_count,  # 이번 달 선결제 등록 건수
    }
    
    context['prepayment_stats'] = prepayment_stats
    context['quote_count'] = quote_count  # 처리해야 할 견적 수 (납품 전환 안 된 것)
    context['monthly_quote_count'] = monthly_quote_count  # 이번 달 견적 횟수
    context['monthly_delivery_count'] = monthly_delivery_count  # 이번 달 납품 횟수
    context['current_month'] = current_month  # 현재 월
    context['month_start'] = month_start.date()  # 이번 달 시작일
    context['month_end'] = month_last_date.date()  # 이번 달 마지막 날 (URL 표시용)
    context['today_unwritten_count'] = today_unwritten_count  # 오늘 미작성 보고서 수
    context['pending_review_count'] = pending_review_count  # 관리자용 미검토 보고서 수
    context['today'] = today  # 오늘 날짜 (템플릿 URL 생성용)

    # Phase 4: 최근 영업 활동 (메모 제외, 최신 8개)
    recent_histories = histories.filter(
        parent_history__isnull=True,
    ).exclude(action_type='memo').select_related(
        'user', 'followup', 'followup__company', 'followup__department', 'schedule'
    ).order_by('-created_at')[:8]
    context['recent_histories'] = recent_histories

    # Phase 4: 지연된 후속 조치 (next_action_date가 오늘 이전, 최대 5개)
    overdue_next_actions = histories.filter(
        next_action_date__lt=today,
        next_action_date__isnull=False,
        parent_history__isnull=True,
    ).exclude(action_type='memo').select_related(
        'user', 'followup', 'followup__company'
    ).order_by('next_action_date')[:5]
    context['overdue_next_actions'] = overdue_next_actions

    # Phase 5: 오늘 예정 일정 (scheduled + completed 포함 — 노트 작성 가능하도록)
    today_schedules = schedules.filter(
        visit_date=today,
        status__in=['scheduled', 'completed']
    ).select_related('followup', 'followup__company', 'user').order_by('visit_time')[:5]
    context['today_schedules'] = today_schedules

    # Phase 5: 이번 주 예정 일정 (오늘 초과 ~ 7일 이내)
    week_later = today + timedelta(days=7)
    upcoming_schedules_dash = schedules.filter(
        visit_date__gt=today,
        visit_date__lte=week_later,
        status='scheduled'
    ).select_related('followup', 'followup__company', 'user').order_by('visit_date', 'visit_time')[:5]
    context['upcoming_schedules_dash'] = upcoming_schedules_dash

    # Phase 5+: 개인 일정 포함 (오늘 + 이번 주 예정)
    from .models import PersonalSchedule as _PersonalSchedule
    today_personal_schedules = _PersonalSchedule.objects.filter(
        schedule_date=today,
        **user_filter_for_dashboard
    ).select_related('user').order_by('schedule_time')[:5]
    context['today_personal_schedules'] = today_personal_schedules

    upcoming_personal_schedules_dash = _PersonalSchedule.objects.filter(
        schedule_date__gt=today,
        schedule_date__lte=week_later,
        **user_filter_for_dashboard
    ).select_related('user').order_by('schedule_date', 'schedule_time')[:5]
    context['upcoming_personal_schedules_dash'] = upcoming_personal_schedules_dash

    # Phase 5: 파이프라인 단계별 현황 (FollowUp 기준)
    pipeline_stage_order = ['potential', 'contact', 'quote', 'negotiation', 'won', 'lost']
    pipeline_stage_labels = dict(FollowUp.PIPELINE_STAGE_CHOICES)
    pipeline_raw = followups.values('pipeline_stage').annotate(cnt=Count('id'))
    pipeline_dict = {item['pipeline_stage']: item['cnt'] for item in pipeline_raw}
    pipeline_summary = [
        {
            'stage': stage,
            'label': pipeline_stage_labels.get(stage, stage),
            'count': pipeline_dict.get(stage, 0),
        }
        for stage in pipeline_stage_order
        if pipeline_dict.get(stage, 0) > 0
    ]
    context['pipeline_summary'] = pipeline_summary

    # Phase 5: 팀 활동 현황 (매니저/관리자 전용, 최근 30일)
    if user_profile.can_view_all_users() and salesman_users:
        thirty_days_ago = today - timedelta(days=30)
        team_activity = []
        for u in list(salesman_users)[:8]:
            recent_cnt = History.objects.filter(
                user=u,
                created_at__date__gte=thirty_days_ago,
                parent_history__isnull=True,
            ).exclude(action_type='memo').count()
            overdue_cnt = History.objects.filter(
                user=u,
                next_action_date__lt=today,
                next_action_date__isnull=False,
                parent_history__isnull=True,
            ).exclude(action_type='memo').count()
            team_activity.append({
                'user': u,
                'recent_count': recent_cnt,
                'overdue_count': overdue_cnt,
            })
        context['team_activity'] = team_activity
    else:
        context['team_activity'] = []

    return render(request, 'reporting/dashboard.html', context)

# ============ 일정(Schedule) 관련 뷰들 ============

@login_required
def schedule_list_view(request):
    """일정 목록 보기 (같은 회사 직원 데이터 필터 지원)"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    user_profile = get_user_profile(request.user)
    
    # 같은 회사 사용자 목록 조회 (필터용) - 매니저 제외
    company_users = []
    if user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 데이터 필터 처리 (나, 전체, 특정 직원)
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')
    
    # 필터에 따른 사용자 목록 결정
    if data_filter == 'all':
        # 전체: 같은 회사 모든 사용자 (salesman만)
        if user_profile.company:
            filter_users = User.objects.filter(
                userprofile__company=user_profile.company,
                userprofile__role='salesman'
            )
        else:
            filter_users = User.objects.filter(id=request.user.id)
    elif data_filter == 'user' and filter_user_id:
        # 특정 직원
        filter_users = User.objects.filter(id=filter_user_id)
    else:
        # 나 (기본값)
        filter_users = User.objects.filter(id=request.user.id)
    
    # 선택된 필터 사용자 정보
    selected_filter_user = None
    if data_filter == 'user' and filter_user_id:
        try:
            selected_filter_user = User.objects.get(id=filter_user_id)
        except User.DoesNotExist:
            pass
    
    # 기본 쿼리셋: 필터에 따른 사용자의 일정
    schedules = Schedule.objects.filter(user__in=filter_users).select_related('user')
    
    # 검색 기능
    search_query = request.GET.get('search')
    product_search = request.GET.get('product_search')  # 제품 검색 추가
    
    if search_query:
        schedules = schedules.filter(
            Q(followup__customer_name__icontains=search_query) |
            Q(followup__company__name__icontains=search_query) |
            Q(followup__department__name__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # 제품 검색 (별도 필드로 분리)
    if product_search:
        schedules = schedules.filter(
            Q(delivery_items_set__product__product_code__icontains=product_search) |
            Q(delivery_items_set__product__description__icontains=product_search)
        ).distinct()
    
    # 날짜 범위 필터링
    date_from = request.GET.get('date_from') or request.GET.get('start_date')
    date_to = request.GET.get('date_to') or request.GET.get('end_date')
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            schedules = schedules.filter(visit_date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            schedules = schedules.filter(visit_date__lte=to_date)
        except ValueError:
            pass
    
    # 필터 값 가져오기
    status_filter = request.GET.get('status')
    activity_type_filter = request.GET.get('activity_type')
    product_filter = request.GET.get('product')  # 제품 필터 추가
    email_filter = request.GET.get('email_sent')  # 메일 발송 필터 추가
    
    # 기본 쿼리셋 (검색, 담당자, 날짜 필터가 적용된 상태)
    base_queryset = schedules
    
    # 상태별 카운트 계산 (활동 유형 필터만 적용된 상태에서)
    if activity_type_filter:
        status_count_queryset = base_queryset.filter(activity_type=activity_type_filter)
    else:
        status_count_queryset = base_queryset
    
    total_count = status_count_queryset.count()
    scheduled_count = status_count_queryset.filter(status='scheduled').count()
    completed_count = status_count_queryset.filter(status='completed').count()
    cancelled_count = status_count_queryset.filter(status='cancelled').count()
    
    # 활동 유형별 카운트 계산 (상태 필터만 적용된 상태에서)
    if status_filter:
        activity_count_queryset = base_queryset.filter(status=status_filter)
    else:
        activity_count_queryset = base_queryset
    
    activity_total_count = activity_count_queryset.count()  # 활동 유형 필터용 전체 카운트
    meeting_count = activity_count_queryset.filter(activity_type='customer_meeting').count()
    quote_count = activity_count_queryset.filter(activity_type='quote').count()  # 견적 카운트 추가
    delivery_count = activity_count_queryset.filter(activity_type='delivery').count()
    service_count = activity_count_queryset.filter(activity_type='service').count()
    
    # 두 필터 모두 적용
    if status_filter:
        schedules = schedules.filter(status=status_filter)
    
    if activity_type_filter:
        schedules = schedules.filter(activity_type=activity_type_filter)
    
    # 제품 필터 적용
    if product_filter:
        schedules = schedules.filter(
            delivery_items_set__product__product_code__icontains=product_filter
        ).distinct()
    
    # 메일 발송 필터 적용
    if email_filter == 'true':
        # 메일을 한 번이라도 보낸 일정만 필터링
        schedules = schedules.filter(emails__isnull=False).distinct()
    elif email_filter == 'false':
        # 메일을 보내지 않은 일정만 필터링
        schedules = schedules.filter(emails__isnull=True)
    
    # 정렬 (예정됨 우선, 그 다음 최신 날짜순)
    # Django의 Case를 사용해서 상태별 우선순위 설정
    from django.db.models import Case, When, IntegerField
    schedules = schedules.annotate(
        status_priority=Case(
            When(status='scheduled', then=1),    # 예정됨: 최우선
            When(status='completed', then=2),    # 완료됨: 두번째
            When(status='cancelled', then=3),    # 취소됨: 마지막
            default=4,
            output_field=IntegerField()
        )
    ).order_by('status_priority', '-visit_date', '-visit_time')  # 상태 우선순위 → 최신 날짜순 → 최신 시간순
    
    # 페이지네이션 추가
    from django.core.paginator import Paginator
    paginator = Paginator(schedules, 30)  # 페이지당 30개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'schedules': page_obj,
        'page_obj': page_obj,
        'page_title': '일정 목록',
        'status_filter': status_filter,
        'activity_type_filter': activity_type_filter,
        'email_filter': email_filter,  # 메일 필터 추가
        'total_count': total_count,
        'scheduled_count': scheduled_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'activity_total_count': activity_total_count,
        'meeting_count': meeting_count,
        'quote_count': quote_count,  # 견적 카운트 추가
        'delivery_count': delivery_count,
        'service_count': service_count,
        'search_query': search_query,
        'product_search': product_search,  # 제품 검색 추가
        'date_from': date_from,
        'date_to': date_to,
        # 데이터 필터 관련
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'selected_filter_user': selected_filter_user,
        'company_users': company_users,
        'is_viewing_others': data_filter != 'me',  # 타인 데이터 조회 중인지
    }
    return render(request, 'reporting/schedule_list.html', context)

@login_required
def schedule_detail_view(request, pk):
    """일정 상세 보기 (같은 회사 직원 데이터 조회 가능, 수정은 본인만)"""
    schedule = get_object_or_404(Schedule, pk=pk)
    
    user_profile = get_user_profile(request.user)
    
    # 권한 체크: 같은 회사 소속이면 조회 가능
    can_view = False
    if schedule.user == request.user:
        can_view = True
    elif user_profile.company:
        # 같은 회사 소속인지 확인
        schedule_user_profile = get_user_profile(schedule.user)
        if schedule_user_profile.company == user_profile.company:
            can_view = True
    
    if not can_view:
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:schedule_list')
    
    # 본인 일정인지 여부 (수정/삭제 권한)
    is_own_schedule = (schedule.user == request.user)
    
    # 관련 히스토리 조회 (최신순)
    related_histories_all = History.objects.filter(schedule=schedule).order_by('-created_at')
    related_histories = related_histories_all[:10]  # 표시용 최신 10개
    
    # 납품 품목 조회 (DeliveryItem 모델)
    delivery_items = DeliveryItem.objects.filter(schedule=schedule)
    
    # 관련 이메일 스레드 조회
    from collections import defaultdict
    email_logs = EmailLog.objects.filter(schedule=schedule).order_by('gmail_thread_id', 'sent_at')
    email_threads = defaultdict(list)
    for email in email_logs:
        if email.gmail_thread_id:
            email_threads[email.gmail_thread_id].append(email)
    
    # 관련 히스토리에서 납품 품목 텍스트 찾기 (대체 방법)
    delivery_text = None
    delivery_amount = 0
    delivery_histories = related_histories_all.filter(
        action_type='delivery_schedule'
    ).exclude(
        delivery_items__isnull=True
    ).exclude(
        delivery_items=''
    )
    
    if delivery_histories.exists():
        latest_delivery = delivery_histories.first()
        raw_delivery_text = latest_delivery.delivery_items
        # \n을 실제 줄바꿈으로 변환
        if raw_delivery_text:
            delivery_text = raw_delivery_text.replace('\\n', '\n')
            delivery_text = raw_delivery_text.replace('\\n', '\n').replace('\\r\\n', '\n')
        
        # 납품 금액도 가져오기
        if latest_delivery.delivery_amount:
            delivery_amount = latest_delivery.delivery_amount
    
    # 이전 페이지 정보 (캘린더에서 온 경우)
    from_page = request.GET.get('from', 'list')  # 기본값은 'list'
    
    context = {
        'schedule': schedule,
        'related_histories': related_histories,
        'delivery_items': delivery_items,
        'delivery_text': delivery_text,  # 히스토리에서 가져온 납품 품목 텍스트
        'delivery_amount': delivery_amount,  # 납품 금액
        'email_threads': dict(email_threads),  # 이메일 스레드
        'from_page': from_page,
        'is_owner': is_own_schedule,  # 본인 일정 여부
        'can_modify': is_own_schedule,  # 수정/삭제 권한 (본인만)
        'schedule_owner': schedule.user,  # 일정 담당자
        'page_title': f'일정 상세 - {schedule.followup.customer_name}'
    }
    return render(request, 'reporting/schedule_detail.html', context)

@login_required
def schedule_create_view(request):
    """일정 생성 (캘린더에서 선택된 날짜 지원)"""
    # Manager는 데이터 생성 불가 (뷰어 권한)
    user_profile = get_user_profile(request.user)
    if user_profile.is_manager():
        messages.error(request, '권한이 없습니다. Manager는 일정을 생성할 수 없습니다.')
        return redirect('reporting:schedule_list')

    if request.method == 'POST':
        form = ScheduleForm(request.POST, user=request.user, request=request)
        if form.is_valid():
            schedule = form.save(commit=False)
            schedule.user = request.user
            schedule.save()
            
            # 복수 선결제 처리 로직
            selected_prepayments_json = request.POST.get('selected_prepayments')
            prepayment_amounts_json = request.POST.get('prepayment_amounts')
            
            if selected_prepayments_json and prepayment_amounts_json:
                import json
                from reporting.models import Prepayment, PrepaymentUsage
                from decimal import Decimal
                
                try:
                    selected_prepayments = json.loads(selected_prepayments_json)
                    prepayment_amounts = json.loads(prepayment_amounts_json)
                    
                    total_prepayment_used = Decimal('0')
                    
                    # 각 선결제에 대해 차감 처리
                    for prepayment_id in selected_prepayments:
                        prepayment_id = str(prepayment_id)
                        if prepayment_id not in prepayment_amounts:
                            continue
                        
                        amount = Decimal(str(prepayment_amounts[prepayment_id]))
                        if amount <= 0:
                            continue
                        
                        try:
                            prepayment = Prepayment.objects.get(id=int(prepayment_id))
                            
                            # 선결제 잔액 확인
                            if prepayment.balance >= amount:
                                # 선결제 잔액 차감
                                prepayment.balance -= amount
                                
                                # 잔액이 0이 되면 상태를 'depleted'로 변경
                                if prepayment.balance <= 0:
                                    prepayment.status = 'depleted'
                                
                                prepayment.save()
                                
                                # PrepaymentUsage 생성
                                PrepaymentUsage.objects.create(
                                    prepayment=prepayment,
                                    schedule=schedule,
                                    product_name=f"{schedule.get_activity_type_display()}",
                                    quantity=1,
                                    amount=amount,
                                    remaining_balance=prepayment.balance,
                                    memo=f"{schedule.get_activity_type_display()} - {schedule.followup.customer_name}"
                                )
                                
                                total_prepayment_used += amount
                                messages.success(request, f'선결제 {prepayment.payer_name or "미지정"} - {amount:,}원이 차감되었습니다. (남은 잔액: {prepayment.balance:,}원)')
                            else:
                                messages.warning(request, f'선결제 {prepayment.payer_name or "미지정"}의 잔액({prepayment.balance:,}원)이 부족합니다.')
                        
                        except Prepayment.DoesNotExist:
                            messages.error(request, f'선결제 ID {prepayment_id}를 찾을 수 없습니다.')
                    
                    if total_prepayment_used > 0:
                        # Schedule의 use_prepayment 플래그 설정
                        schedule.use_prepayment = True
                        # 첫 번째 선결제를 대표로 저장 (기존 필드 호환성)
                        if selected_prepayments:
                            first_prepayment = Prepayment.objects.filter(id=int(selected_prepayments[0])).first()
                            if first_prepayment:
                                schedule.prepayment = first_prepayment
                        schedule.prepayment_amount = total_prepayment_used
                        schedule.save()
                        
                        messages.info(request, f'총 선결제 사용 금액: {total_prepayment_used:,}원')
                
                except json.JSONDecodeError:
                    messages.error(request, '선결제 데이터 형식이 올바르지 않습니다.')
                except Exception as e:
                    messages.error(request, f'선결제 처리 중 오류 발생: {str(e)}')
            
            # 품목 데이터 처리 (견적 또는 납품)
            if schedule.activity_type in ['quote', 'delivery']:
                # 품목 데이터가 있으면 저장
                has_delivery_items = any(key.startswith('delivery_items[') for key in request.POST.keys())
                if has_delivery_items:
                    created_count = save_delivery_items(request, schedule)
                    if created_count > 0:
                        messages.success(request, f'{created_count}개의 품목이 저장되었습니다.')
                
                # 품목 저장 후 펀넬 예상 수주액 업데이트 (견적과 납품 모두)
                if has_delivery_items:
                    from decimal import Decimal
                    delivery_items = schedule.delivery_items_set.all()
                    if delivery_items.exists():
                        delivery_total = sum(Decimal(str(item.total_price or 0)) for item in delivery_items)
                        if delivery_total > 0:
                            # 일정의 예상 수주액 업데이트
                            if not schedule.expected_revenue or schedule.expected_revenue == 0:
                                schedule.expected_revenue = delivery_total
                                schedule.save()
                
                # 선결제 사용 시 PrepaymentUsage에 품목 정보 업데이트
                if schedule.use_prepayment:
                    from reporting.models import PrepaymentUsage, DeliveryItem
                    usages = PrepaymentUsage.objects.filter(schedule=schedule)
                    delivery_items = DeliveryItem.objects.filter(schedule=schedule).order_by('id')
                    
                    if usages.exists() and delivery_items.exists():
                        # 첫 번째 품목 정보를 첫 번째 usage에 저장
                        first_item = delivery_items.first()
                        for usage in usages:
                            usage.product_name = first_item.item_name
                            usage.quantity = first_item.quantity
                            usage.save()
                            break  # 첫 번째 usage만 업데이트
            
            # 견적에서 불러온 납품인 경우, 원본 견적 일정을 완료 처리
            from_quote_schedule_id = request.POST.get('from_quote_schedule_id')
            if from_quote_schedule_id and schedule.activity_type == 'delivery':
                try:
                    quote_schedule = Schedule.objects.get(
                        id=int(from_quote_schedule_id),
                        activity_type='quote'
                    )
                    quote_schedule.status = 'completed'
                    quote_schedule.save()
                    messages.info(request, f'견적 일정(ID: {from_quote_schedule_id})이 완료 처리되었습니다.')
                except (Schedule.DoesNotExist, ValueError):
                    pass
            
            messages.success(request, '일정이 성공적으로 생성되었습니다.')
            
            # 일정 캘린더로 리다이렉트 (모달이 자동으로 열리도록 schedule_id 전달)
            return redirect(f"{reverse('reporting:schedule_calendar')}?schedule_id={schedule.pk}")
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        # URL 파라미터에서 날짜 가져오기
        selected_date = request.GET.get('date')
        followup_id = request.GET.get('followup')
        initial_data = {}
        
        if selected_date:
            try:
                # 날짜 형식 검증 (YYYY-MM-DD)
                from datetime import datetime
                parsed_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
                initial_data['visit_date'] = parsed_date
            except ValueError:
                messages.warning(request, '잘못된 날짜 형식입니다.')
        
        # 팔로우업 ID가 있으면 초기 데이터에 설정 (예상 매출 등은 설정하지 않음)
        if followup_id:
            try:
                followup = FollowUp.objects.get(pk=followup_id)
                initial_data['followup'] = followup
            except FollowUp.DoesNotExist:
                pass
        
        form = ScheduleForm(user=request.user, request=request, initial=initial_data)
    
    context = {
        'form': form,
        'page_title': '새 일정 생성',
        'selected_date': request.GET.get('date')  # 템플릿에서 사용할 수 있도록
    }
    return render(request, 'reporting/schedule_form.html', context)

@login_required
def schedule_edit_view(request, pk):
    """일정 수정 (본인 일정만 수정 가능)"""
    
    schedule = get_object_or_404(Schedule, pk=pk)
    
    # 권한 체크: 본인 일정만 수정 가능
    if schedule.user != request.user:
        messages.error(request, '본인의 일정만 수정할 수 있습니다.')
        return redirect('reporting:schedule_list')
    
    if request.method == 'POST':
        form = ScheduleForm(request.POST, instance=schedule, user=request.user, request=request)
        if form.is_valid():
            updated_schedule = form.save()
            
            # 복수 선결제 처리 로직 (수정 시에도 적용)
            import json
            from reporting.models import Prepayment, PrepaymentUsage
            from decimal import Decimal
            
            selected_prepayments_json = request.POST.get('selected_prepayments')
            prepayment_amounts_json = request.POST.get('prepayment_amounts')
            use_prepayment_checkbox = request.POST.get('use_prepayment')
            
            # 기존 선결제 사용 내역 확인
            existing_usages = PrepaymentUsage.objects.filter(schedule=updated_schedule)
            had_prepayment = existing_usages.exists()
            
            # 선결제 사용 체크가 해제되었거나, 선결제 데이터가 없는 경우 → 기존 내역 복구
            if had_prepayment and (not use_prepayment_checkbox or not selected_prepayments_json or not prepayment_amounts_json):
                # 기존 선결제 사용 내역 복구 (잔액 되돌리기)
                restored_amount = Decimal('0')
                for usage in existing_usages:
                    prepayment = usage.prepayment
                    prepayment.balance += usage.amount
                    if prepayment.status == 'depleted' and prepayment.balance > 0:
                        prepayment.status = 'active'
                    prepayment.save()
                    restored_amount += usage.amount
                
                # 기존 사용 내역 삭제
                existing_usages.delete()
                
                # Schedule의 선결제 관련 필드 초기화
                updated_schedule.use_prepayment = False
                updated_schedule.prepayment = None
                updated_schedule.prepayment_amount = 0
                updated_schedule.save()
                
                if restored_amount > 0:
                    messages.info(request, f'선결제 사용이 취소되어 {restored_amount:,}원이 잔액으로 복구되었습니다.')
            
            # 선결제 사용 체크되고 데이터가 있는 경우 → 새로 적용
            elif use_prepayment_checkbox and selected_prepayments_json and prepayment_amounts_json:
                try:
                    # 기존 선결제 사용 내역이 있으면 먼저 복구
                    if had_prepayment:
                        for usage in existing_usages:
                            prepayment = usage.prepayment
                            prepayment.balance += usage.amount
                            if prepayment.status == 'depleted' and prepayment.balance > 0:
                                prepayment.status = 'active'
                            prepayment.save()
                        existing_usages.delete()
                    
                    # 새로운 선결제 적용
                    selected_prepayments = json.loads(selected_prepayments_json)
                    prepayment_amounts = json.loads(prepayment_amounts_json)
                    
                    total_prepayment_used = Decimal('0')
                    
                    for prepayment_id in selected_prepayments:
                        prepayment_id = str(prepayment_id)
                        if prepayment_id not in prepayment_amounts:
                            continue
                        
                        amount = Decimal(str(prepayment_amounts[prepayment_id]))
                        if amount <= 0:
                            continue
                        
                        try:
                            prepayment = Prepayment.objects.get(id=int(prepayment_id))
                            
                            if prepayment.balance >= amount:
                                prepayment.balance -= amount
                                
                                if prepayment.balance <= 0:
                                    prepayment.status = 'depleted'
                                
                                prepayment.save()
                                
                                PrepaymentUsage.objects.create(
                                    prepayment=prepayment,
                                    schedule=updated_schedule,
                                    product_name=f"{updated_schedule.get_activity_type_display()}",
                                    quantity=1,
                                    amount=amount,
                                    remaining_balance=prepayment.balance,
                                    memo=f"{updated_schedule.get_activity_type_display()} - {updated_schedule.followup.customer_name}"
                                )
                                
                                total_prepayment_used += amount
                                messages.success(request, f'선결제 {prepayment.payer_name or "미지정"} - {amount:,}원이 차감되었습니다.')
                            else:
                                messages.warning(request, f'선결제 {prepayment.payer_name or "미지정"}의 잔액({prepayment.balance:,}원)이 부족합니다.')
                        
                        except Prepayment.DoesNotExist:
                            messages.error(request, f'선결제 ID {prepayment_id}를 찾을 수 없습니다.')
                    
                    if total_prepayment_used > 0:
                        updated_schedule.use_prepayment = True
                        if selected_prepayments:
                            first_prepayment = Prepayment.objects.filter(id=int(selected_prepayments[0])).first()
                            if first_prepayment:
                                updated_schedule.prepayment = first_prepayment
                        updated_schedule.prepayment_amount = total_prepayment_used
                        updated_schedule.save()
                        
                        messages.info(request, f'총 선결제 사용 금액: {total_prepayment_used:,}원')
                    else:
                        # 선결제 금액이 0이면 플래그 해제
                        updated_schedule.use_prepayment = False
                        updated_schedule.prepayment = None
                        updated_schedule.prepayment_amount = 0
                        updated_schedule.save()
                
                except json.JSONDecodeError:
                    messages.error(request, '선결제 데이터 형식이 올바르지 않습니다.')
                except Exception as e:
                    messages.error(request, f'선결제 처리 중 오류 발생: {str(e)}')
            
            # 납품 품목 데이터가 있으면 저장
            has_delivery_items = any(key.startswith('delivery_items[') for key in request.POST.keys())
            if has_delivery_items:
                created_count = save_delivery_items(request, updated_schedule)
                if created_count > 0:
                    messages.success(request, f'{created_count}개의 품목이 저장되었습니다.')
            
            # 선결제 사용 시 품목 금액 변경에 따라 PrepaymentUsage 금액도 업데이트
            if updated_schedule.use_prepayment:
                from reporting.models import PrepaymentUsage, DeliveryItem
                from decimal import Decimal
                
                usages = PrepaymentUsage.objects.filter(schedule=updated_schedule).order_by('id')
                delivery_items = DeliveryItem.objects.filter(schedule=updated_schedule).order_by('id')
                
                if usages.exists() and delivery_items.exists():
                    # 새로운 품목 총액 계산
                    new_total = sum(item.total_price or Decimal('0') for item in delivery_items)
                    
                    # 기존 선결제 사용 총액
                    old_total = sum(usage.amount for usage in usages)
                    
                    # 금액 차이 계산
                    diff = new_total - old_total
                    
                    if diff != 0:
                        # 선결제 사용 금액 조정 (비례 배분)
                        # 단일 선결제인 경우 해당 선결제에 전체 차액 적용
                        # 복수 선결제인 경우 비례 배분
                        
                        for usage in usages:
                            if old_total > 0:
                                # 비례 배분: 기존 비율에 맞게 새 금액 계산
                                ratio = usage.amount / old_total
                                new_usage_amount = new_total * ratio
                            else:
                                # 기존 총액이 0이면 균등 배분
                                new_usage_amount = new_total / usages.count()
                            
                            # 금액 차이 계산
                            usage_diff = new_usage_amount - usage.amount
                            
                            # 선결제 잔액 조정
                            prepayment = usage.prepayment
                            prepayment.balance -= usage_diff  # 증가분은 잔액에서 차감, 감소분은 잔액에 추가
                            
                            # 잔액이 음수가 되지 않도록
                            if prepayment.balance < 0:
                                # 잔액 부족 시 가능한 만큼만 차감
                                possible_amount = usage.amount + prepayment.balance  # 현재 잔액까지만
                                prepayment.balance = Decimal('0')
                                new_usage_amount = possible_amount
                                messages.warning(request, f'선결제 {prepayment.payer_name or "미지정"}의 잔액이 부족하여 일부만 적용되었습니다.')
                            
                            # 상태 업데이트
                            if prepayment.balance <= 0:
                                prepayment.status = 'depleted'
                            elif prepayment.status == 'depleted':
                                prepayment.status = 'active'
                            
                            prepayment.save()
                            
                            # Usage 금액 업데이트
                            usage.amount = new_usage_amount
                            usage.remaining_balance = prepayment.balance
                            usage.save()
                        
                        # Schedule의 선결제 금액도 업데이트
                        updated_schedule.prepayment_amount = sum(u.amount for u in usages)
                        updated_schedule.save()
                        
                        if diff > 0:
                            messages.info(request, f'품목 금액 증가로 선결제 {abs(diff):,.0f}원이 추가 차감되었습니다.')
                        else:
                            messages.info(request, f'품목 금액 감소로 선결제 {abs(diff):,.0f}원이 잔액으로 복구되었습니다.')
                    
                    # 첫 번째 품목 정보를 usage에 저장 (품목명 업데이트)
                    first_item = delivery_items.first()
                    for usage in usages:
                        usage.product_name = first_item.item_name
                        usage.quantity = first_item.quantity
                        usage.save()
                        break  # 첫 번째 usage만 업데이트

            
            messages.success(request, '일정이 성공적으로 수정되었습니다.')
            return redirect('reporting:schedule_detail', pk=schedule.pk)
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        form = ScheduleForm(instance=schedule, user=request.user, request=request)
    
    # DeliveryItem 모델에서 납품 품목 데이터 가져오기 (우선순위 1)
    delivery_text = None
    delivery_amount = 0
    
    # 1차: DeliveryItem 모델에서 최신 데이터 확인
    delivery_items = schedule.delivery_items_set.all().order_by('id')
    
    if delivery_items.exists():
        from decimal import Decimal
        delivery_text_parts = []
        total_amount = Decimal('0')
        
        for item in delivery_items:
            # VAT 포함 총액 계산 (DeliveryItem의 save()에서 자동 계산됨)
            item_total = item.total_price or (item.quantity * item.unit_price * Decimal('1.1'))
            total_amount += item_total
            
            # 텍스트 형태로 변환
            text_part = f"{item.item_name}: {item.quantity}개 ({int(item_total):,}원)"
            delivery_text_parts.append(text_part)
        
        delivery_text = '\n'.join(delivery_text_parts)
        delivery_amount = int(total_amount)
    
    # 2차: DeliveryItem이 없으면 History에서 fallback
    if not delivery_text:
        related_histories = History.objects.filter(schedule=schedule, action_type='delivery_schedule').order_by('-created_at')
        
        if related_histories.exists():
            latest_delivery = related_histories.first()
            if latest_delivery.delivery_items:
                raw_text = latest_delivery.delivery_items
                delivery_text = raw_text.replace('\\n', '\n')
            if latest_delivery.delivery_amount:
                delivery_amount = latest_delivery.delivery_amount
    
    context = {
        'form': form,
        'schedule': schedule,
        'delivery_text': delivery_text,
        'delivery_amount': delivery_amount,
        'page_title': f'일정 수정 - {schedule.followup.customer_name}'
    }
    return render(request, 'reporting/schedule_form.html', context)

@login_required
def schedule_delete_view(request, pk):
    """일정 삭제 (본인 일정만 삭제 가능)"""
    import traceback
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        schedule = get_object_or_404(Schedule, pk=pk)
        
        # 권한 체크: 본인 일정만 삭제 가능
        if schedule.user != request.user:
            # AJAX 요청 감지 - X-Requested-With 헤더 확인
            is_ajax = request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
            if is_ajax:
                return JsonResponse({'success': False, 'error': '본인의 일정만 삭제할 수 있습니다.'}, status=403)
            messages.error(request, '본인의 일정만 삭제할 수 있습니다.')
            return redirect('reporting:schedule_list')
        
        if request.method == 'POST':
            customer_name = schedule.followup.customer_name or "고객명 미정"
            schedule_date = schedule.visit_date.strftime("%Y년 %m월 %d일")
            
            # 선결제 사용 내역 롤백 (선결제 잔액 복구)
            prepayment_usages = PrepaymentUsage.objects.filter(schedule=schedule)
            if prepayment_usages.exists():
                for usage in prepayment_usages:
                    # 선결제 잔액 복구
                    prepayment = usage.prepayment
                    old_balance = prepayment.balance
                    prepayment.balance += usage.amount
                    
                    # 잔액이 0원에서 복구되면 상태를 'active'로 변경
                    if old_balance == 0 and prepayment.balance > 0:
                        prepayment.status = 'active'
                    
                    prepayment.save()
                
                # 사용 내역 삭제
                prepayment_usages.delete()
            
            # 관련 히스토리 확인
            related_histories = schedule.histories.all()
            history_count = related_histories.count()
            
            # 일정과 관련 히스토리 삭제
            if history_count > 0:
                related_histories.delete()
            
            schedule.delete()
            
            # AJAX 요청 감지 - X-Requested-With 헤더 확인
            is_ajax = request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
            if is_ajax:
                success_message = f'{customer_name} ({schedule_date}) 일정이 삭제되었습니다.'
                if history_count > 0:
                    success_message += f' (관련 활동 기록 {history_count}개도 함께 삭제되었습니다.)'
                
                return JsonResponse({
                    'success': True, 
                    'message': success_message
                })
            
            # 일반 폼 요청인 경우 리다이렉트
            success_message = f'{customer_name} ({schedule_date}) 일정이 삭제되었습니다.'
            if history_count > 0:
                success_message += f' (관련 활동 기록 {history_count}개도 함께 삭제되었습니다.)'
            
            messages.success(request, success_message)
            
            # 이전 페이지 정보 확인 (캘린더에서 온 경우)
            from_page = request.GET.get('from', 'list')
            user_filter = request.GET.get('user', '')  # user 파라미터로 수정
            
            # 캘린더에서 온 경우 캘린더로 돌아가기
            if from_page == 'calendar':
                if user_filter:
                    return redirect(f"{reverse('reporting:schedule_calendar')}?user={user_filter}")
                else:
                    return redirect('reporting:schedule_calendar')
            else:
                return redirect('reporting:schedule_list')
        
        # GET 요청인 경우
        context = {
            'schedule': schedule,
            'page_title': f'일정 삭제 - {schedule.followup.customer_name or "고객명 미정"}'
        }
        return render(request, 'reporting/schedule_delete.html', context)
        
    except Exception as e:
        # 예외 발생 시 상세 로깅
        error_msg = str(e)
        error_traceback = traceback.format_exc()
        
        logger.error(f"일정 삭제 중 예외 발생:")
        logger.error(f"오류 메시지: {error_msg}")
        logger.error(f"스택 트레이스:\n{error_traceback}")
        
        # AJAX 요청인 경우 JSON 에러 응답
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False, 
                'error': f'일정 삭제 중 오류가 발생했습니다: {error_msg}',
                'detail': error_traceback if settings.DEBUG else None  # DEBUG 모드에서만 상세 정보 포함
            }, status=500)
        
        # 일반 요청인 경우 에러 메시지와 함께 리다이렉트
        messages.error(request, f'일정 삭제 중 오류가 발생했습니다: {error_msg}')
        return redirect('reporting:schedule_list')

@login_required
def schedule_update_delivery_items(request, pk):
    """일정의 납품 품목 업데이트"""
    schedule = get_object_or_404(Schedule, pk=pk)
    
    # 권한 체크: 수정 권한이 있는 경우만 수정 가능
    if not can_modify_user_data(request.user, schedule.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': '수정 권한이 없습니다.'}, status=403)
        messages.error(request, '수정 권한이 없습니다.')
        return redirect('reporting:schedule_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            # 납품 품목 저장
            created_count = save_delivery_items(request, schedule)
            
            if created_count == 0:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': '저장된 품목이 없습니다.'}, status=400)
                messages.warning(request, '저장된 품목이 없습니다. 품목명과 수량을 모두 입력했는지 확인해주세요.')
                return redirect('reporting:schedule_detail', pk=pk)
            
            # 관련된 History들의 delivery_items 텍스트도 업데이트
            related_histories = schedule.histories.filter(action_type='delivery_schedule')
            
            # 새로 저장된 DeliveryItem들을 텍스트로 변환
            delivery_items = schedule.delivery_items_set.all()
            
            delivery_text = ''
            total_delivery_amount = 0
            delivery_items_list = []
            
            if delivery_items.exists():
                delivery_lines = []
                
                for item in delivery_items:
                    # delivery_items_list에 JSON 직렬화 가능한 형태로 추가
                    # unit_price가 None이 아니면 실제 값 사용 (0 포함)
                    item_unit_price = None
                    if item.unit_price is not None:
                        item_unit_price = float(item.unit_price)
                    
                    delivery_items_list.append({
                        'id': item.id,
                        'item_name': item.item_name,
                        'quantity': float(item.quantity),
                        'unit_price': item_unit_price,
                        'product_id': item.product_id
                    })
                    
                    # unit_price가 None이 아니고 0보다 클 때만 금액 표시
                    if item.unit_price is not None and item.unit_price > 0:
                        # 부가세 포함 총액 계산 (단가 * 수량 * 1.1)
                        total_amount = int(float(item.unit_price) * item.quantity * 1.1)
                        total_delivery_amount += total_amount
                        delivery_lines.append(f"{item.item_name}: {item.quantity}개 ({total_amount:,}원)")
                    elif item.unit_price is not None and item.unit_price == 0:
                        # 0원인 경우
                        delivery_lines.append(f"{item.item_name}: {item.quantity}개 (0원)")
                    else:
                        # unit_price가 None인 경우
                        delivery_lines.append(f"{item.item_name}: {item.quantity}개")
                
                delivery_text = '\n'.join(delivery_lines)
                
                # 관련 History가 있으면 업데이트, 없으면 새로 생성
                if related_histories.exists():
                    # 기존 History 업데이트
                    for history in related_histories:
                        history.delivery_items = delivery_text
                        if total_delivery_amount > 0:
                            history.delivery_amount = total_delivery_amount
                        history.save(update_fields=['delivery_items', 'delivery_amount'])
                else:
                    # 새로운 History 생성
                    from .models import History
                    history = History.objects.create(
                        schedule=schedule,
                        user=request.user,
                        action_type='delivery_schedule',
                        delivery_items=delivery_text,
                        delivery_amount=total_delivery_amount if total_delivery_amount > 0 else None,
                        content=f'납품 품목 {created_count}개 추가'
                    )
            
            # AJAX 요청인 경우 JSON 응답 반환
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': '납품 품목이 성공적으로 업데이트되었습니다.',
                    'schedule': {
                        'id': schedule.id,
                        'delivery_items': delivery_text,
                        'delivery_items_list': delivery_items_list,
                        'delivery_amount': total_delivery_amount
                    }
                })
            
            messages.success(request, '납품 품목이 성공적으로 업데이트되었습니다.')
        except Exception as e:
            logger.error(f'납품 품목 업데이트 중 오류: {str(e)}', exc_info=True)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f'납품 품목 업데이트 중 오류가 발생했습니다: {str(e)}')
        
        return redirect('reporting:schedule_detail', pk=pk)
    
    # GET 요청은 허용하지 않음
    return redirect('reporting:schedule_detail', pk=pk)

@login_required
def schedule_calendar_view(request):
    """일정 캘린더 뷰 (같은 회사 직원 데이터 필터링)"""
    user_profile = get_user_profile(request.user)
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용) - 매니저 제외
    company_users = []
    if user_profile and user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    is_viewing_others = False
    
    if data_filter == 'all':
        # 같은 회사 전체
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id:
        # 특정 직원
        try:
            selected_filter_user = company_users.get(id=filter_user_id)
            is_viewing_others = True
        except User.DoesNotExist:
            data_filter = 'me'  # 유효하지 않은 경우 기본값
    # else: 'me' - 본인 데이터만
    
    context = {
        'page_title': '일정 캘린더',
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
    }
    return render(request, 'reporting/schedule_calendar.html', context)

@login_required
def schedule_api_view(request):
    """일정 데이터 API (JSON 응답) - 같은 회사 직원 데이터 필터링 + PersonalSchedule 포함"""
    try:
        from .models import PersonalSchedule
        from django.db.models import Prefetch
        from decimal import Decimal
        
        user_profile = get_user_profile(request.user)
        
        # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
        data_filter = request.GET.get('data_filter', 'me')
        filter_user_id = request.GET.get('filter_user')
        
        # 필터에 따른 대상 사용자 결정
        if data_filter == 'all' and user_profile and user_profile.company:
            # 같은 회사 전체
            filter_users = User.objects.filter(
                userprofile__company=user_profile.company,
                is_active=True
            )
        elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
            # 특정 직원 (같은 회사 확인)
            try:
                target_user = User.objects.get(
                    id=filter_user_id,
                    userprofile__company=user_profile.company,
                    is_active=True
                )
                filter_users = User.objects.filter(id=target_user.id)
            except User.DoesNotExist:
                filter_users = User.objects.filter(id=request.user.id)
        else:
            # 'me' - 본인만
            filter_users = User.objects.filter(id=request.user.id)
        
        # 스케줄 쿼리
        schedules = Schedule.objects.filter(user__in=filter_users)
        
        # 🔥 최적화: 관련 데이터를 한 번에 가져오기
        schedules = schedules.select_related(
            'followup',
            'followup__company',
            'followup__department',
            'user'
        ).prefetch_related(
            'delivery_items_set',  # DeliveryItem 미리 로드
            Prefetch(
                'histories',
                queryset=History.objects.filter(action_type='delivery_schedule').only(
                    'id', 'action_type', 'delivery_items', 'delivery_amount'
                ),
                to_attr='delivery_histories'
            )
        ).only(
            # 필요한 필드만 SELECT (메모리 절약)
            'id', 'visit_date', 'visit_time', 'location', 'status', 
            'activity_type', 'notes', 'expected_revenue', 'probability', 
            'expected_close_date',
            'followup__id', 'followup__customer_name', 'followup__company__id',
            'followup__company__name', 'followup__department__id', 
            'followup__department__name', 'followup__manager', 
            'followup__address', 'followup__priority',
            'user__id', 'user__username'
        )
        
        schedule_data = []
        for schedule in schedules:
            schedule_item = {
                'id': schedule.id,
                'followup_id': schedule.followup.id,  # 팔로우업 ID 추가
                'visit_date': schedule.visit_date.strftime('%Y-%m-%d'),
                'time': schedule.visit_time.strftime('%H:%M'),
                'customer': schedule.followup.customer_name or '고객명 미정',
                'company': str(schedule.followup.company) if schedule.followup.company else '업체명 미정',
                'department': str(schedule.followup.department) if schedule.followup.department else '부서명 미정',
                'department_id': schedule.followup.department_id if schedule.followup.department else None,
                'manager': schedule.followup.manager or '',
                'address': schedule.followup.address or '',
                'location': schedule.location or '',
                'status': schedule.status,
                'status_display': schedule.get_status_display(),
                'activity_type': schedule.activity_type,
                'activity_type_display': schedule.get_activity_type_display(),
                'notes': schedule.notes or '',
                'user_name': schedule.user.username,
                'priority': schedule.followup.priority,  # 고객 우선순위 추가
                # 펀넬 관련 필드 추가
                'expected_revenue': float(schedule.expected_revenue) if schedule.expected_revenue else 0,
                'probability': schedule.probability if schedule.probability is not None else 0,
                'expected_close_date': schedule.expected_close_date.strftime('%Y-%m-%d') if schedule.expected_close_date else '',
            }
            
            # 견적 또는 납품 일정인 경우 품목 정보 추가
            if schedule.activity_type in ['delivery', 'quote']:
                delivery_items_text = ''
                delivery_amount = 0
                has_schedule_items = False  # 스케줄에 직접 등록된 품목이 있는지 여부
                delivery_items_list = []  # 실제 품목 데이터 배열
                
                # 🔥 최적화: prefetch된 데이터 사용 (추가 쿼리 없음)
                schedule_delivery_items = list(schedule.delivery_items_set.all())
                if schedule_delivery_items:
                    has_schedule_items = True
                    delivery_text_parts = []
                    total_amount = 0
                    
                    for item in schedule_delivery_items:
                        # unit_price가 None이면 0으로 처리 (Decimal 타입 유지)
                        unit_price = item.unit_price if item.unit_price is not None else Decimal('0')
                        item_total = item.total_price or (item.quantity * unit_price * Decimal('1.1'))
                        total_amount += item_total
                        
                        # 실제 품목 데이터 추가
                        delivery_items_list.append({
                            'id': item.id,
                            'item_name': item.item_name,
                            'quantity': float(item.quantity),
                            'unit_price': float(unit_price),
                            'product_id': item.product.id if item.product else None,
                        })
                        
                        # 단가가 0이면 금액 표시 생략
                        if unit_price > 0:
                            text_part = f"{item.item_name}: {item.quantity}개 ({int(item_total):,}원)"
                        else:
                            text_part = f"{item.item_name}: {item.quantity}개"
                        delivery_text_parts.append(text_part)
                    
                    delivery_items_text = '\n'.join(delivery_text_parts)
                    delivery_amount = int(total_amount)
                else:
                    # 🔥 최적화: prefetch된 delivery_histories 사용
                    if hasattr(schedule, 'delivery_histories') and schedule.delivery_histories:
                        delivery_history = schedule.delivery_histories[0]
                        if delivery_history.delivery_items:
                            delivery_items_text = delivery_history.delivery_items.strip()
                            delivery_amount = int(delivery_history.delivery_amount) if delivery_history.delivery_amount else 0
                
                schedule_item.update({
                    'delivery_items': delivery_items_text,
                    'delivery_amount': delivery_amount,
                    'has_schedule_items': has_schedule_items,  # 품목 관리 제한용
                    'delivery_items_list': delivery_items_list,  # 실제 품목 데이터
                })
            
            schedule_data.append(schedule_item)
        
        # ====== PersonalSchedule 데이터 추가 ======
        # 같은 필터 적용
        personal_schedules = PersonalSchedule.objects.filter(user__in=filter_users)
        
        personal_schedules = personal_schedules.select_related('user', 'company').only(
            'id', 'title', 'content', 'schedule_date', 'schedule_time',
            'user__id', 'user__username', 'company__id', 'company__name'
        )
        
        # PersonalSchedule을 schedule_data에 추가 (type='personal' 구분자 추가)
        for ps in personal_schedules:
            personal_item = {
                'id': ps.id,
                'type': 'personal',  # 개인 일정 구분자
                'visit_date': ps.schedule_date.strftime('%Y-%m-%d'),
                'time': ps.schedule_time.strftime('%H:%M'),
                'title': ps.title,
                'content': ps.content or '',
                'user_name': ps.user.username,
                'company': str(ps.company) if ps.company else '',
                # 캘린더 표시용 기본값
                'customer': ps.title,  # 제목을 customer로 사용
                'status': 'personal',
                'status_display': '개인 일정',
                'activity_type': 'personal',
                'activity_type_display': '개인 일정',
                'location': '',
                'notes': ps.content or '',
                'priority': 'medium',
            }
            schedule_data.append(personal_item)
        
        return JsonResponse(schedule_data, safe=False)
    
    except Exception as e:
        # 에러 디버깅을 위한 JSON 응답
        import traceback
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc(),
            'message': 'schedule_api_view에서 오류가 발생했습니다.'
        }, status=500)# ============ 히스토리(History) 관련 뷰들 ============

@login_required
def history_list_view(request):
    """히스토리 목록 보기 (같은 회사 직원 데이터 필터링)"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.db.models import Q
    
    user_profile = get_user_profile(request.user)
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용) - 매니저 제외
    company_users = []
    if user_profile and user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    is_viewing_others = False
    
    if data_filter == 'all' and user_profile and user_profile.company:
        # 같은 회사 전체 (salesman만)
        filter_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        )
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        # 특정 직원 (같은 회사 확인)
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True
            )
            filter_users = User.objects.filter(id=selected_filter_user.id)
            is_viewing_others = True
        except User.DoesNotExist:
            filter_users = User.objects.filter(id=request.user.id)
            data_filter = 'me'
    else:
        # 'me' - 본인만
        filter_users = User.objects.filter(id=request.user.id)
    
    # 히스토리 쿼리 (매니저 메모 제외)
    histories = History.objects.filter(user__in=filter_users, parent_history__isnull=True)
    
    # 관련 객체들을 미리 로드하여 성능 최적화 (답글 메모도 포함)
    histories = histories.select_related(
        'user', 'followup', 'followup__company', 'followup__department', 'schedule', 'personal_schedule'
    ).prefetch_related('reply_memos__created_by')  # 답글 메모들을 미리 로드
    
    # 검색 기능 (책임자 검색 추가)
    search_query = request.GET.get('search')
    if search_query:
        histories = histories.filter(
            Q(content__icontains=search_query) |
            Q(followup__customer_name__icontains=search_query) |
            Q(followup__company__name__icontains=search_query) |
            Q(followup__manager__icontains=search_query)
        )
    
    # 팔로우업 필터링 (특정 팔로우업의 모든 히스토리 보기)
    followup_filter = request.GET.get('followup')
    if followup_filter:
        histories = histories.filter(followup_id=followup_filter)

    # Phase 4: 업체 필터링
    company_filter = request.GET.get('company_filter', '')
    if company_filter:
        histories = histories.filter(followup__company_id=company_filter)

    # Phase 5: 다음 액션 날짜 필터 (overdue / upcoming / has_date)
    from django.utils import timezone as tz_util
    history_list_today = tz_util.now().date()
    next_action_filter = request.GET.get('next_action_filter', '')
    if next_action_filter == 'overdue':
        histories = histories.filter(
            next_action_date__lt=history_list_today,
            next_action_date__isnull=False,
        )
    elif next_action_filter == 'upcoming':
        from datetime import timedelta as td
        history_list_week = history_list_today + td(days=7)
        histories = histories.filter(
            next_action_date__gte=history_list_today,
            next_action_date__lte=history_list_week,
            next_action_date__isnull=False,
        )
    elif next_action_filter == 'has_date':
        histories = histories.filter(next_action_date__isnull=False)

    # Phase 4: 접근 가능한 업체 목록 (검색 폼용)
    accessible_companies = Company.objects.filter(
        followup_companies__user__in=filter_users
    ).distinct().order_by('name')

    # 활동 유형별 카운트 계산
    base_queryset_for_counts = histories
    total_count = base_queryset_for_counts.count()
    meeting_count = base_queryset_for_counts.filter(action_type='customer_meeting').count()
    quote_count = base_queryset_for_counts.filter(action_type='quote').count()  # 견적 카운트 추가
    delivery_count = base_queryset_for_counts.filter(action_type='delivery_schedule').count()
    service_count = base_queryset_for_counts.filter(action_type='service', service_status='completed').count()
    memo_count = base_queryset_for_counts.filter(action_type='memo', personal_schedule__isnull=True).count()  # 개인 일정 제외
    personal_schedule_count = base_queryset_for_counts.filter(personal_schedule__isnull=False).count()  # 개인 일정 카운트
    
    # 활동 유형 필터링
    action_type_filter = request.GET.get('action_type')
    if action_type_filter:
        if action_type_filter == 'personal_schedule':
            # 개인 일정만 필터링
            histories = histories.filter(personal_schedule__isnull=False)
        elif action_type_filter == 'memo':
            # 메모 필터링 (개인 일정 제외)
            histories = histories.filter(action_type='memo', personal_schedule__isnull=True)
        elif action_type_filter == 'delivery_schedule':
            # 납품 일정 필터링 - 중복 제거 (같은 Schedule의 History 중 최신 1개만)
            histories = histories.filter(action_type=action_type_filter)
            
            # 중복 제거: 같은 schedule_id를 가진 History 중 가장 최근 것만 선택
            from django.db.models import OuterRef, Subquery
            
            # 각 schedule_id의 최신 History ID를 찾음
            latest_history_per_schedule = History.objects.filter(
                schedule_id=OuterRef('schedule_id'),
                action_type='delivery_schedule',
                user__in=filter_users,
                parent_history__isnull=True
            ).order_by('-created_at').values('id')[:1]
            
            # Schedule이 있는 History: 최신 것만
            histories_with_schedule = histories.filter(
                schedule__isnull=False,
                id=Subquery(latest_history_per_schedule)
            )
            
            # Schedule이 없는 History: 날짜+금액으로 중복 제거
            histories_without_schedule = histories.filter(schedule__isnull=True)
            
            # 두 쿼리셋을 합침 (union 대신 | 사용)
            histories = histories_with_schedule | histories_without_schedule
        else:
            histories = histories.filter(action_type=action_type_filter)
    
    # 월별 필터링 추가
    months_filter = request.GET.get('months')
    month_filter = None
    if months_filter:
        try:
            selected_months = [int(month.strip()) for month in months_filter.split(',') if month.strip().isdigit()]
            if selected_months:
                month_filter = months_filter  # 템플릿에서 사용할 원본 문자열
                # 월별 필터링 적용 순서:
                # 1. 관련 일정이 있는 경우 → 일정의 visit_date 월로 필터링
                # 2. 관련 일정이 없는 경우 → 히스토리 생성일자(created_at) 월로 필터링
                from django.db.models import Q, Case, When, IntegerField
                from django.db.models.functions import Extract
                histories = histories.annotate(
                    filter_month=Case(
                        When(schedule__isnull=False, then=Extract('schedule__visit_date', 'month')),
                        default=Extract('created_at', 'month'),
                        output_field=IntegerField()
                    )
                ).filter(filter_month__in=selected_months)
        except (ValueError, TypeError):
            pass

    # 날짜 범위 필터
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    if date_from_str:
        try:
            from datetime import datetime as dt
            date_from_val = dt.strptime(date_from_str, '%Y-%m-%d').date()
            histories = histories.filter(created_at__date__gte=date_from_val)
        except ValueError:
            date_from_str = ''
    if date_to_str:
        try:
            from datetime import datetime as dt
            date_to_val = dt.strptime(date_to_str, '%Y-%m-%d').date()
            histories = histories.filter(created_at__date__lte=date_to_val)
        except ValueError:
            date_to_str = ''

    # 미검토 필터 (관리자/매니저용)
    review_filter = request.GET.get('review_filter', '')
    unreviewed_count = 0
    if user_profile and user_profile.can_view_all_users():
        unreviewed_count = histories.filter(
            action_type__in=['customer_meeting', 'delivery_schedule', 'quote', 'service'],
            parent_history__isnull=True,
            reviewed_at__isnull=True,
        ).count()
        if review_filter == 'unreviewed':
            histories = histories.filter(reviewed_at__isnull=True)

    # 정렬 (일정이 있는 경우 일정 날짜 기준, 개인 일정이 있는 경우 개인 일정 날짜 기준, 없는 경우 작성일 기준으로 최신순)
    from django.db.models import Case, When, F
    histories = histories.annotate(
        sort_date=Case(
            When(schedule__isnull=False, then=F('schedule__visit_date')),
            When(personal_schedule__isnull=False, then=F('personal_schedule__schedule_date')),
            default=F('created_at__date')
        )
    ).order_by('-sort_date', '-created_at')
    
    # 선택된 팔로우업 정보
    selected_followup = None
    if followup_filter:
        try:
            candidate_followup = FollowUp.objects.get(id=followup_filter)
            # 같은 회사인지 확인
            if user_profile and user_profile.company:
                followup_profile = get_user_profile(candidate_followup.user)
                if followup_profile and followup_profile.company == user_profile.company:
                    selected_followup = candidate_followup
        except (FollowUp.DoesNotExist, ValueError):
            pass
    
    # 페이지 제목 동적 설정
    if selected_followup:
        page_title = f'{selected_followup.customer_name or "고객명 미정"} 활동 히스토리'
    else:
        page_title = '활동 히스토리'
    
    # 페이지네이션 추가
    from django.core.paginator import Paginator
    paginator = Paginator(histories, 30)  # 페이지당 30개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'histories': page_obj,
        'page_obj': page_obj,
        'page_title': page_title,
        'action_type_filter': action_type_filter,
        'month_filter': month_filter,
        'total_count': total_count,
        'meeting_count': meeting_count,
        'quote_count': quote_count,  # 견적 카운트 추가
        'delivery_count': delivery_count,
        'service_count': service_count,
        'memo_count': memo_count,
        'personal_schedule_count': personal_schedule_count,  # 개인 일정 카운트 추가
        'search_query': search_query,
        'followup_filter': followup_filter,
        'selected_followup': selected_followup,
        # 새로운 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
        # 날짜 범위 / 미검토 필터
        'date_from': date_from_str,
        'date_to': date_to_str,
        'review_filter': review_filter,
        'unreviewed_count': unreviewed_count,
        # Phase 4: 업체 필터
        'company_filter': company_filter,
        'accessible_companies': accessible_companies,
        # Phase 5: 다음 액션 날짜 필터 & 오늘 날짜
        'next_action_filter': next_action_filter,
        'today': history_list_today,
    }
    return render(request, 'reporting/history_list.html', context)

@login_required
def history_detail_view(request, pk):
    """히스토리 상세 보기 (Manager 권한 포함)"""
    from django.utils import timezone
    history = get_object_or_404(History, pk=pk)
    today = timezone.now().date()
    
    # 권한 체크 (Manager도 Salesman 데이터 접근 가능)
    if not can_access_user_data(request.user, history.user):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:history_list')
    
    # 매니저 메모 추가 처리
    user_profile = get_user_profile(request.user)
    if request.method == 'POST' and user_profile.is_manager():
        manager_memo = request.POST.get('manager_memo', '').strip()
        if manager_memo:
            # 매니저 메모를 부모 히스토리에 연결된 자식 히스토리로 생성
            memo_history = History.objects.create(
                followup=history.followup,
                user=history.user,  # 원래 실무자를 유지
                parent_history=history,  # 부모 히스토리 설정
                action_type='memo',
                content=manager_memo,  # 매니저 메모 표시 제거 (parent_history로 구분)
                created_by=request.user,  # 실제 작성자는 매니저
                schedule=history.schedule if history.schedule else None
            )
            messages.success(request, '매니저 메모가 추가되었습니다.')
            return redirect('reporting:history_detail', pk=pk)
    
    # 사용자 필터 정보 추가 (Manager가 특정 사용자의 활동을 보고 있는 경우)
    user_filter = request.GET.get('user_filter', '')
    if not user_filter and request.user != history.user:
        # Manager가 다른 사용자의 활동을 보고 있다면 해당 사용자 필터 설정
        user_profile = get_user_profile(request.user)
        if user_profile.can_view_all_users():
            user_filter = history.user.id
    
    # 동일한 팔로우업의 최근 히스토리들 가져오기 (메모 포함)
    related_histories = History.objects.filter(
        followup=history.followup
    ).select_related('user', 'created_by', 'schedule').order_by('-created_at')[:10]
    
    # 답글 메모(댓글) 조회
    history = History.objects.select_related(
        'user', 'created_by', 'schedule', 'followup', 'personal_schedule'
    ).prefetch_related(
        'reply_memos__user'
    ).get(pk=pk)
    
    # 본인 히스토리인지 여부
    is_owner = (request.user == history.user)
    
    context = {
        'history': history,
        'related_histories': related_histories,
        'user_filter': user_filter,
        'can_add_memo': user_profile.is_manager(),
        'can_modify': can_modify_user_data(request.user, history.user),  # 수정/삭제 권한 (관리자 포함)
        'is_owner': is_owner,  # 본인 데이터 여부
        'page_title': f'활동 상세 - {history.followup.customer_name if history.followup else "일반 메모"}',
        'today': today,  # Phase 4: 다음 액션 만료 여부 표시용
    }
    return render(request, 'reporting/history_detail.html', context)

@login_required
def history_create_view(request):
    """히스토리 생성"""
    # Manager는 데이터 생성 불가 (뷰어 권한)
    user_profile = get_user_profile(request.user)
    if user_profile.is_manager():
        messages.error(request, '권한이 없습니다. Manager는 활동을 기록할 수 없습니다.')
        return redirect('reporting:history_list')

    if request.method == 'POST':
        form = HistoryForm(request.POST, request.FILES, user=request.user, request=request)
        if form.is_valid():
            history = form.save(commit=False)
            history.user = request.user
            history.save()

            # 파이프라인 자동 진행 (고객 미팅 → contact 이상으로 앞으로만 이동)
            try:
                if history.action_type == 'customer_meeting' and history.followup:
                    from .funnel_views import _try_advance_pipeline
                    _try_advance_pipeline(history.followup, 'contact')
            except Exception:
                pass  # 파이프라인 업데이트 실패해도 히스토리 저장은 유지

            # 납품 품목 저장
            save_delivery_items(request, history)
            
            # 파일 업로드 처리
            uploaded_files = request.FILES.getlist('files')
            if uploaded_files:
                uploaded_file_objects, file_errors = handle_file_uploads(uploaded_files, history, request.user)
                
                if uploaded_file_objects:
                    file_count = len(uploaded_file_objects)
                    messages.success(request, f'활동 히스토리가 성공적으로 기록되었습니다. ({file_count}개 파일 업로드됨)')
                else:
                    messages.success(request, '활동 히스토리가 성공적으로 기록되었습니다.')
                
                # 파일 업로드 오류가 있는 경우 경고 메시지 표시
                for error in file_errors:
                    messages.warning(request, f'파일 업로드 오류: {error}')
            else:
                messages.success(request, '활동 히스토리가 성공적으로 기록되었습니다.')
            
            return redirect('reporting:history_detail', pk=history.pk)
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        form = HistoryForm(user=request.user, request=request)
    
    context = {
        'form': form,
        'page_title': '새 활동 기록'
    }
    return render(request, 'reporting/history_form.html', context)

@login_required
def history_edit_view(request, pk):
    """히스토리 수정"""
    history = get_object_or_404(History, pk=pk)
    
    # 권한 체크: 수정 권한이 있는 경우만 수정 가능 (Manager는 읽기 전용)
    if not can_modify_user_data(request.user, history.user):
        messages.error(request, '수정 권한이 없습니다. Manager는 읽기 전용입니다.')
        return redirect('reporting:history_list')
    
    if request.method == 'POST':
        form = HistoryForm(request.POST, request.FILES, instance=history, user=request.user, request=request)
        if form.is_valid():
            form.save()
            
            # 납품 품목 저장
            save_delivery_items(request, history)
            
            # History에 연결된 Schedule이 있다면 Schedule의 DeliveryItem도 동기화
            if history.schedule:
                # Schedule의 기존 DeliveryItem들 삭제
                history.schedule.delivery_items_set.all().delete()
                
                # History의 새로운 DeliveryItem들을 Schedule에도 복사
                history_delivery_items = history.delivery_items_set.all()
                for history_item in history_delivery_items:
                    DeliveryItem.objects.create(
                        schedule=history.schedule,
                        item_name=history_item.item_name,
                        quantity=history_item.quantity,
                        unit=history_item.unit,
                        unit_price=history_item.unit_price
                    )
            
            # 새로운 파일 업로드 처리
            uploaded_files = request.FILES.getlist('files')
            if uploaded_files:
                uploaded_file_objects, file_errors = handle_file_uploads(uploaded_files, history, request.user)
                
                if uploaded_file_objects:
                    file_count = len(uploaded_file_objects)
                    messages.success(request, f'활동 히스토리가 성공적으로 수정되었습니다. ({file_count}개 파일 추가 업로드됨)')
                else:
                    messages.success(request, '활동 히스토리가 성공적으로 수정되었습니다.')
                
                # 파일 업로드 오류가 있는 경우 경고 메시지 표시
                for error in file_errors:
                    messages.warning(request, f'파일 업로드 오류: {error}')
            else:
                messages.success(request, '활동 히스토리가 성공적으로 수정되었습니다.')
            
            return redirect('reporting:history_detail', pk=history.pk)
        else:
            messages.error(request, '입력 정보를 확인해주세요.')
    else:
        form = HistoryForm(instance=history, user=request.user, request=request)
    
    # 관련 스케줄의 납품 품목 정보 가져오기
    schedule_delivery_items = []
    delivery_text = ""
    delivery_amount = 0
    
    if history.schedule:
        # 스케줄에 연결된 납품 품목들 가져오기
        schedule_delivery_items = history.schedule.delivery_items_set.all()
        
        # 납품 품목들을 텍스트 형태로 변환하고 총액 계산
        if schedule_delivery_items:
            delivery_lines = []
            total_delivery_amount = 0
            for item in schedule_delivery_items:
                if item.unit_price:
                    # 부가세 포함 총액 계산 (단가 * 수량 * 1.1)
                    from decimal import Decimal
                    item_total_amount = int(float(item.unit_price) * item.quantity * 1.1)
                    total_delivery_amount += item_total_amount
                    delivery_lines.append(f"{item.item_name}: {item.quantity}개 ({item_total_amount:,}원)")
                else:
                    delivery_lines.append(f"{item.item_name}: {item.quantity}개")
            delivery_text = '\n'.join(delivery_lines)
            delivery_amount = total_delivery_amount
    
    context = {
        'form': form,
        'history': history,
        'existing_delivery_items': history.delivery_items_set.all(),
        'schedule_delivery_items': schedule_delivery_items,
        'delivery_text': delivery_text,
        'delivery_amount': delivery_amount,
        'page_title': f'활동 수정 - {history.followup.customer_name if history.followup else "일반 메모"}'
    }
    return render(request, 'reporting/history_form.html', context)

@login_required
def history_delete_view(request, pk):
    """히스토리 삭제"""
    history = get_object_or_404(History, pk=pk)
    
    # 권한 체크: 삭제 권한이 있는 경우만 삭제 가능 (Manager는 읽기 전용)
    if not can_modify_user_data(request.user, history.user):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': '삭제 권한이 없습니다. Manager는 읽기 전용입니다.'
            })
        messages.error(request, '삭제 권한이 없습니다. Manager는 읽기 전용입니다.')
        return redirect('reporting:history_list')
    
    if request.method == 'POST':
        customer_name = history.followup.customer_name or "고객명 미정" if history.followup else "일반 메모"
        action_display = history.get_action_type_display()
        
        try:
            history.delete()
            success_message = f'{customer_name} ({action_display}) 활동 기록이 삭제되었습니다.'
            
            # AJAX 요청인 경우 JSON 응답
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_message
                })
            
            # 일반 요청인 경우 기존 방식
            messages.success(request, success_message)
            return redirect('reporting:history_list')
            
        except Exception as e:
            error_message = f'활동 기록 삭제 중 오류가 발생했습니다: {str(e)}'
            
            # AJAX 요청인 경우 JSON 응답
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_message
                })
            
            # 일반 요청인 경우 기존 방식
            messages.error(request, error_message)
            return redirect('reporting:history_list')
    
    context = {
        'history': history,
        'page_title': f'활동 삭제 - {history.followup.customer_name or "고객명 미정" if history.followup else "일반 메모"}'
    }
    return render(request, 'reporting/history_delete.html', context)

@login_required
@require_POST
def history_toggle_reviewed(request, pk):
    """보고서 관리자 검토 완료 토글 (관리자/매니저만)"""
    history = get_object_or_404(History, pk=pk)
    user_profile = get_user_profile(request.user)
    if not user_profile.can_view_all_users():
        return JsonResponse({'success': False, 'error': '검토 권한이 없습니다.'}, status=403)
    if not can_access_user_data(request.user, history.user):
        return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
    from django.utils import timezone as tz
    if history.reviewed_at:
        history.reviewed_at = None
        history.reviewer = None
        is_reviewed = False
    else:
        history.reviewed_at = tz.now()
        history.reviewer = request.user
        is_reviewed = True
    history.save(update_fields=['reviewed_at', 'reviewer'])
    return JsonResponse({
        'success': True,
        'is_reviewed': is_reviewed,
        'reviewed_at': history.reviewed_at.strftime('%Y년 %m월 %d일 %H:%M') if is_reviewed else None,
        'reviewer': request.user.get_full_name() or request.user.username if is_reviewed else None,
    })


@login_required
def history_by_followup_view(request, followup_pk):
    """특정 팔로우업의 히스토리 목록"""
    followup = get_object_or_404(FollowUp, pk=followup_pk)
    
    # 권한 체크
    if not (request.user.is_staff or request.user.is_superuser or followup.user == request.user):
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:followup_list')
    
    histories = History.objects.filter(followup=followup)
    
    context = {
        'histories': histories,
        'followup': followup,
        'page_title': f'{followup.customer_name} 활동 히스토리'
    }
    return render(request, 'reporting/history_list.html', context)

# ============ API 엔드포인트들 ============

@login_required
def api_followup_schedules(request, followup_pk):
    """특정 팔로우업의 일정 목록을 JSON으로 반환 (AJAX용)"""
    try:
        followup = get_object_or_404(FollowUp, pk=followup_pk)
        
        # 권한 체크 (같은 회사 소속이면 접근 가능)
        if not can_access_followup(request.user, followup):
            return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
        
        # 본인 일정만 조회 (동료 고객이어도 내가 만든 일정만)
        schedules = Schedule.objects.filter(followup=followup, user=request.user)
        schedule_list = []
        
        for schedule in schedules:
            schedule_list.append({
                'id': schedule.id,
                'text': f"{schedule.visit_date} {schedule.visit_time} - {schedule.location or '장소 미정'}",
                'visit_date': schedule.visit_date.strftime('%Y-%m-%d')  # 납품 날짜 자동 설정용
            })
        
        return JsonResponse({'schedules': schedule_list})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def schedule_histories_api(request, schedule_id):
    """특정 일정의 관련 활동 기록을 JSON으로 반환"""
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        
        # 권한 체크
        if not can_access_user_data(request.user, schedule.user):
            return JsonResponse({'error': '권한이 없습니다.'}, status=403)
        
        # 해당 일정에 직접 연결된 활동 기록만 조회 (최신순) - 답글 메모 포함
        histories = History.objects.filter(schedule=schedule).prefetch_related(
            'reply_memos__user'
        ).order_by('-created_at')
        
        histories_data = []
        for history in histories:
            # 활동 타입에 따른 추가 정보 포함
            history_data = {
                'id': history.id,
                'action_type': history.action_type,
                'action_type_display': history.get_action_type_display(),
                'content': history.content or '',
                'created_at': history.created_at.strftime('%Y-%m-%d %H:%M'),
                'user': history.user.username,
                'created_by': history.created_by.username if history.created_by else history.user.username,
            }
            
            # 답글 메모(댓글) 정보 추가
            reply_memos_data = []
            for reply_memo in history.reply_memos.all():
                reply_memos_data.append({
                    'id': reply_memo.id,
                    'content': reply_memo.content,
                    'created_at': reply_memo.created_at.strftime('%Y-%m-%d %H:%M'),
                    'user': reply_memo.user.username,
                })
            history_data['reply_memos'] = reply_memos_data
            
            # 첨부파일 정보 추가
            files_data = []
            for file_obj in history.files.all():
                files_data.append({
                    'id': file_obj.id,
                    'filename': file_obj.original_filename,
                    'size': file_obj.get_file_size_display(),
                    'uploaded_at': file_obj.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                    'uploaded_by': file_obj.uploaded_by.username,
                })
            history_data['files'] = files_data
            
            # 납품 일정인 경우 추가 정보
            if history.action_type == 'delivery_schedule':
                # DeliveryItem 모델에서 최신 납품 품목 데이터 가져오기
                delivery_items_text = ''
                delivery_amount = history.delivery_amount or 0
                
                # 1차: History에 연결된 DeliveryItem 확인
                delivery_items = history.delivery_items_set.all().order_by('id')
                if delivery_items.exists():
                    delivery_text_parts = []
                    total_amount = 0
                    
                    for item in delivery_items:
                        item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
                        total_amount += item_total
                        text_part = f"{item.item_name}: {item.quantity}개 ({int(item_total):,}원)"
                        delivery_text_parts.append(text_part)
                    
                    delivery_items_text = '\n'.join(delivery_text_parts)
                    delivery_amount = int(total_amount)
                
                # 2차: History의 schedule에 연결된 DeliveryItem 확인 (fallback)
                elif history.schedule:
                    schedule_delivery_items = history.schedule.delivery_items_set.all().order_by('id')
                    if schedule_delivery_items.exists():
                        delivery_text_parts = []
                        total_amount = 0
                        
                        for item in schedule_delivery_items:
                            item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
                            total_amount += item_total
                            text_part = f"{item.item_name}: {item.quantity}개 ({int(item_total):,}원)"
                            delivery_text_parts.append(text_part)
                        
                        delivery_items_text = '\n'.join(delivery_text_parts)
                        delivery_amount = int(total_amount)
                
                # 3차: 기존 텍스트 필드 사용 (최종 fallback)
                if not delivery_items_text and history.delivery_items:
                    delivery_items_text = history.delivery_items
                
                history_data.update({
                    'delivery_amount': delivery_amount,
                    'delivery_items': delivery_items_text,
                    'delivery_date': history.delivery_date.strftime('%Y-%m-%d') if history.delivery_date else '',
                    'tax_invoice_issued': history.tax_invoice_issued,
                })
            
            # 고객 미팅인 경우 추가 정보
            elif history.action_type == 'customer_meeting':
                history_data.update({
                    'meeting_date': history.meeting_date.strftime('%Y-%m-%d') if history.meeting_date else '',
                    'meeting_situation': history.meeting_situation or '',
                    'meeting_researcher_quote': history.meeting_researcher_quote or '',
                    'meeting_confirmed_facts': history.meeting_confirmed_facts or '',
                    'meeting_obstacles': history.meeting_obstacles or '',
                    'meeting_next_action': history.meeting_next_action or '',
                })
            
            # 서비스인 경우 추가 정보
            elif history.action_type == 'service':
                history_data.update({
                    'service_status': history.service_status or '',
                    'service_status_display': history.get_service_status_display() if history.service_status else '',
                })
            
            histories_data.append(history_data)
        
        return JsonResponse({
            'success': True,
            'histories': histories_data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def followup_histories_api(request, followup_id):
    """특정 팔로우업의 모든 활동 기록을 JSON으로 반환"""
    try:
        followup = get_object_or_404(FollowUp, id=followup_id)
        
        # 권한 체크 - followup.user가 None인 경우를 처리
        if followup.user:
            if not can_access_user_data(request.user, followup.user):
                return JsonResponse({'error': '권한이 없습니다.'}, status=403)
        else:
            # followup.user가 None인 경우, 현재 사용자가 관리자이거나 매니저인 경우만 접근 허용
            user_profile = get_user_profile(request.user)
            if not (user_profile.is_admin() or user_profile.is_manager()):
                return JsonResponse({'error': '권한이 없습니다.'}, status=403)
        
        # 해당 팔로우업의 모든 활동 기록 조회 (최신순)
        histories = History.objects.filter(followup=followup).select_related('schedule').order_by('-created_at')
        
        histories_data = []
        for history in histories:
            try:
                # 활동 타입에 따른 추가 정보 포함
                history_data = {
                    'id': history.id,
                    'action_type': history.action_type,
                    'action_type_display': history.get_action_type_display(),
                    'content': history.content or '',
                    'created_at': history.created_at.strftime('%Y-%m-%d %H:%M'),
                    'user': history.user.username if history.user else '사용자 미정',
                    'created_by': history.created_by.username if history.created_by else (history.user.username if history.user else '사용자 미정'),
                }
                
                # 납품 일정인 경우 추가 정보 (품목 포함)
                if history.action_type == 'delivery_schedule':
                    # 납품 품목 조회 (DeliveryItem)
                    delivery_items_list = []
                    for item in history.delivery_items_set.all():
                        delivery_items_list.append({
                            'item_name': item.item_name,
                            'quantity': item.quantity,
                            'unit': item.unit or 'EA',
                            'unit_price': float(item.unit_price) if item.unit_price else 0,
                            'total_price': float(item.total_price) if item.total_price else 0,
                        })
                    
                    history_data.update({
                        'delivery_amount': history.delivery_amount,
                        'delivery_items': history.delivery_items or '',
                        'delivery_items_list': delivery_items_list,
                        'delivery_date': history.delivery_date.strftime('%Y-%m-%d') if history.delivery_date else '',
                        'tax_invoice_issued': history.tax_invoice_issued,
                    })
                
                # 견적 일정인 경우 추가 정보 (품목 포함)
                elif history.action_type == 'quote':
                    quote_items_list = []
                    total_quote_amount = 0
                    
                    # 디버깅: schedule 연결 상태
                    has_schedule = history.schedule is not None
                    schedule_id = history.schedule.id if history.schedule else None
                    
                    # 1. 연결된 스케줄의 견적 조회
                    if history.schedule:
                        # 스케줄에 연결된 DeliveryItem (견적용 품목)
                        for item in history.schedule.delivery_items_set.all():
                            quote_items_list.append({
                                'item_name': item.item_name,
                                'quantity': item.quantity,
                                'unit_price': float(item.unit_price) if item.unit_price else 0,
                                'subtotal': float(item.total_price) if item.total_price else 0,
                            })
                            total_quote_amount += float(item.total_price) if item.total_price else 0
                        
                        # 스케줄에 연결된 Quote 모델의 품목
                        quotes = history.schedule.quotes.all().prefetch_related('items__product')
                        for quote in quotes:
                            for item in quote.items.all():
                                quote_items_list.append({
                                    'item_name': item.product.product_code if item.product else '품목명 없음',
                                    'quantity': item.quantity,
                                    'unit_price': float(item.unit_price) if item.unit_price else 0,
                                    'subtotal': float(item.subtotal) if item.subtotal else 0,
                                })
                            total_quote_amount += float(quote.total_amount) if quote.total_amount else 0
                    
                    # 2. 히스토리에 직접 연결된 DeliveryItem도 확인
                    for item in history.delivery_items_set.all():
                        quote_items_list.append({
                            'item_name': item.item_name,
                            'quantity': item.quantity,
                            'unit_price': float(item.unit_price) if item.unit_price else 0,
                            'subtotal': float(item.total_price) if item.total_price else 0,
                        })
                        total_quote_amount += float(item.total_price) if item.total_price else 0
                    
                    history_data.update({
                        'quote_items_list': quote_items_list,
                        'total_quote_amount': total_quote_amount,
                        'has_schedule': has_schedule,  # 디버깅용
                        'schedule_id': schedule_id,  # 디버깅용
                    })
                
                # 고객 미팅인 경우 추가 정보
                elif history.action_type == 'customer_meeting':
                    history_data.update({
                        'meeting_date': history.meeting_date.strftime('%Y-%m-%d') if history.meeting_date else '',
                    })
                
                histories_data.append(history_data)
                
            except Exception as history_error:
                # 개별 히스토리 처리 중 에러가 발생해도 계속 진행
                continue
        
        # ===== 스케줄에서 직접 품목 정보 가져오기 (히스토리와 중복되지 않도록) =====
        # 히스토리에 연결되지 않은 스케줄의 품목도 표시
        processed_schedule_ids = set()
        for h in histories:
            if h.schedule:
                processed_schedule_ids.add(h.schedule.id)
        
        schedules_data = []
        schedules = Schedule.objects.filter(
            followup=followup,
            activity_type__in=['quote', 'delivery']
        ).prefetch_related('delivery_items_set').order_by('-visit_date')
        
        for schedule in schedules:
            # 이미 히스토리에서 처리된 스케줄은 건너뛰기
            if schedule.id in processed_schedule_ids:
                continue
            
            items_list = []
            total_amount = 0
            
            for item in schedule.delivery_items_set.all():
                items_list.append({
                    'item_name': item.item_name,
                    'quantity': item.quantity,
                    'unit': item.unit or 'EA',
                    'unit_price': float(item.unit_price) if item.unit_price else 0,
                    'total_price': float(item.total_price) if item.total_price else 0,
                })
                total_amount += float(item.total_price) if item.total_price else 0
            
            # 품목이 있는 경우만 추가
            if items_list:
                schedule_data = {
                    'id': f'schedule_{schedule.id}',
                    'type': 'schedule',
                    'action_type': 'delivery_schedule' if schedule.activity_type == 'delivery' else 'quote',
                    'action_type_display': '납품 일정' if schedule.activity_type == 'delivery' else '견적 제출',
                    'content': schedule.notes or '',
                    'created_at': schedule.visit_date.strftime('%Y-%m-%d'),
                    'user': schedule.user.username if schedule.user else '사용자 미정',
                    'items_list': items_list,
                    'total_amount': total_amount,
                    'schedule_id': schedule.id,
                    'status': schedule.get_status_display(),
                }
                schedules_data.append(schedule_data)
        
        return JsonResponse({
            'success': True,
            'customer_name': followup.customer_name or '고객명 미정',
            'company': followup.company.name if followup.company else '업체명 미정',
            'histories': histories_data,
            'schedules': schedules_data,  # 히스토리에 연결되지 않은 스케줄 품목
            'total_count': len(histories_data) + len(schedules_data)
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"followup_histories_api error for followup_id={followup_id}: {str(e)}", exc_info=True)
        return JsonResponse({'error': f'서버 에러: {str(e)}'}, status=500)

# ============ 인증 관련 뷰들 ============

class CustomLoginView(LoginView):
    """커스텀 로그인 뷰 (성공 메시지 추가)"""
    template_name = 'reporting/login.html'
    redirect_authenticated_user = True
    
    def form_valid(self, form):
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, '로그인에 실패했습니다. 아이디와 비밀번호를 확인해주세요.')
        return super().form_invalid(form)

class CustomLogoutView(LogoutView):
    """커스텀 로그아웃 뷰 (성공 메시지 추가)"""
    next_page = reverse_lazy('reporting:login')
    
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

# ============ 사용자 관리 뷰들 ============

# 사용자 생성 폼 (Admin 전용)
class UserCreationForm(forms.Form):
    username = forms.CharField(
        max_length=150,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'ID를 입력하세요', 'autocomplete': 'off'}),
        label='사용자 ID'
    )
    company = forms.CharField(
        max_length=100,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '회사명을 입력하세요 (예: 하나과학)'}),
        label='소속 회사',
        help_text='사용자가 속할 회사명을 직접 입력하세요'
    )
    password1 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='비밀번호'
    )
    password2 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='비밀번호 확인'
    )
    role = forms.ChoiceField(
        choices=[('manager', 'Manager (뷰어)'), ('salesman', 'SalesMan (실무자)')],
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='권한'
    )
    can_download_excel = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='엑셀 다운로드 권한',
        help_text='체크 시 팔로우업 엑셀 다운로드가 가능합니다'
    )
    can_use_ai = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='AI 분석 권한',
        help_text='체크 시 AI PainPoint 분석 기능을 사용할 수 있습니다'
    )
    first_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '성 (선택사항)', 'autocomplete': 'off'}),
        label='성'
    )
    last_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '이름 (선택사항)', 'autocomplete': 'off'}),
        label='이름'
    )
    
    def clean_password2(self):
        password1 = self.cleaned_data.get("password1")
        password2 = self.cleaned_data.get("password2")
        if password1 and password2 and password1 != password2:
            raise forms.ValidationError("비밀번호가 일치하지 않습니다.")
        return password2

# 매니저용 사용자 생성 폼 (Manager 전용)
class ManagerUserCreationForm(forms.Form):
    username = forms.CharField(
        max_length=150,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'ID를 입력하세요', 'autocomplete': 'off'}),
        label='사용자 ID'
    )
    password1 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='비밀번호'
    )
    password2 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='비밀번호 확인'
    )
    # 매니저는 실무자만 생성할 수 있음
    role = forms.CharField(
        initial='salesman',
        widget=forms.HiddenInput(),
        label='권한'
    )
    can_download_excel = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='엑셀 다운로드 권한',
        help_text='체크 시 팔로우업 엑셀 다운로드가 가능합니다'
    )
    first_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '성 (선택사항)', 'autocomplete': 'off'}),
        label='성'
    )
    last_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '이름 (선택사항)', 'autocomplete': 'off'}),
        label='이름'
    )
    
    def clean_password2(self):
        password1 = self.cleaned_data.get("password1")
        password2 = self.cleaned_data.get("password2")
        if password1 and password2 and password1 != password2:
            raise forms.ValidationError("비밀번호가 일치하지 않습니다.")
        return password2

# 사용자 편집 폼 클래스
class UserEditForm(forms.Form):
    username = forms.CharField(
        max_length=150,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'ID를 입력하세요', 'autocomplete': 'off'}),
        label='사용자 ID'
    )
    company = forms.CharField(
        max_length=100,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '회사명을 입력하세요 (예: 하나과학)'}),
        label='소속 회사',
        help_text='사용자가 속할 회사명을 직접 입력하세요'
    )
    role = forms.ChoiceField(
        choices=[('admin', 'Admin (최고권한자)'), ('manager', 'Manager (뷰어)'), ('salesman', 'SalesMan (실무자)')],
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='권한'
    )
    can_download_excel = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='엑셀 다운로드 권한',
        help_text='체크 시 팔로우업 엑셀 다운로드가 가능합니다 (관리자는 항상 가능)'
    )
    can_use_ai = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='AI 분석 권한',
        help_text='체크 시 AI PainPoint 분석 기능을 사용할 수 있습니다'
    )
    first_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '성 (선택사항)', 'autocomplete': 'off'}),
        label='성'
    )
    last_name = forms.CharField(
        max_length=30,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '이름 (선택사항)', 'autocomplete': 'off'}),
        label='이름'
    )
    change_password = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='비밀번호 변경'
    )
    password1 = forms.CharField(
        required=False,
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='새 비밀번호'
    )
    password2 = forms.CharField(
        required=False,
        widget=forms.PasswordInput(attrs={'class': 'form-control', 'autocomplete': 'new-password'}),
        label='새 비밀번호 확인'
    )
    
    def clean(self):
        cleaned_data = super().clean()
        change_password = cleaned_data.get('change_password')
        password1 = cleaned_data.get('password1')
        password2 = cleaned_data.get('password2')
        
        if change_password:
            if not password1:
                raise forms.ValidationError("비밀번호 변경을 선택했으면 새 비밀번호를 입력해야 합니다.")
            if password1 != password2:
                raise forms.ValidationError("비밀번호가 일치하지 않습니다.")
        
        return cleaned_data

@role_required(['admin'])
def user_list(request):
    """사용자 목록 (Admin 전용)"""
    users = User.objects.select_related('userprofile').all().order_by('username')
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # 역할별 필터
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(userprofile__role=role_filter)
    
    # 페이지네이션 (정렬된 쿼리셋 사용으로 경고 해결)
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)
    
    context = {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
        'role_choices': UserProfile.ROLE_CHOICES,
        'page_title': '사용자 관리'
    }
    return render(request, 'reporting/user_list.html', context)

@role_required(['admin'])
def user_create(request):
    """사용자 생성 (Admin 전용)"""
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            # 사용자 생성
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password1'],
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name']
            )
            
            # 회사 정보 가져오기 또는 생성
            company_name = form.cleaned_data['company']
            user_company, created = UserCompany.objects.get_or_create(name=company_name)
            
            # 사용자 프로필 생성
            UserProfile.objects.create(
                user=user,
                company=user_company,  # UserCompany 객체 사용
                role=form.cleaned_data['role'],
                can_download_excel=form.cleaned_data['can_download_excel'],
                can_use_ai=form.cleaned_data['can_use_ai'],
                created_by=request.user
            )
            
            messages.success(request, f'사용자 "{user.username}"이(가) 성공적으로 생성되었습니다.')
            return redirect('reporting:user_list')
    else:
        form = UserCreationForm()
    
    context = {
        'form': form,
        'page_title': '사용자 생성'
    }
    return render(request, 'reporting/user_create.html', context)

@role_required(['admin'])
def user_edit(request, user_id):
    """사용자 편집 (Admin 전용)"""
    user = get_object_or_404(User, id=user_id)
    user_profile = get_object_or_404(UserProfile, user=user)
    
    if request.method == 'POST':
        form = UserEditForm(request.POST)
        if form.is_valid():
            # 회사 정보 가져오기 또는 생성
            company_name = form.cleaned_data['company']
            user_company, created = UserCompany.objects.get_or_create(name=company_name)
            
            # 사용자 정보 수정
            user.username = form.cleaned_data['username']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            
            # 비밀번호 변경
            if form.cleaned_data['change_password'] and form.cleaned_data['password1']:
                user.set_password(form.cleaned_data['password1'])
            
            user.save()
            
            # 권한 및 회사 수정
            user_profile.company = user_company  # 회사 정보 업데이트
            user_profile.role = form.cleaned_data['role']
            user_profile.can_download_excel = form.cleaned_data['can_download_excel']
            user_profile.can_use_ai = form.cleaned_data['can_use_ai']
            user_profile.save()
            
            messages.success(request, f'사용자 "{user.username}"의 정보가 성공적으로 수정되었습니다.')
            return redirect('reporting:user_list')
    else:
        # 기존 데이터로 폼 초기화
        form = UserEditForm(initial={
            'username': user.username,
            'company': user_profile.company.name if user_profile.company else '',
            'first_name': user.first_name,
            'last_name': user.last_name,
            'role': user_profile.role,
            'can_download_excel': user_profile.can_download_excel,
            'can_use_ai': user_profile.can_use_ai,
        })
    
    context = {
        'form': form,
        'edit_user': user,
        'user_profile': user_profile,
        'page_title': f'사용자 편집 - {user.username}'
    }
    return render(request, 'reporting/user_edit.html', context)

@role_required(['admin'])
def user_delete(request, user_id):
    """사용자 삭제 (Admin 전용)"""
    import traceback
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        # 삭제할 사용자 가져오기
        user_to_delete = get_object_or_404(User, id=user_id)
        
        # 자기 자신은 삭제할 수 없음
        if user_to_delete.id == request.user.id:
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False, 
                    'error': '자신의 계정은 삭제할 수 없습니다.'
                }, status=400)
            messages.error(request, '자신의 계정은 삭제할 수 없습니다.')
            return redirect('reporting:user_list')
        
        if request.method == 'POST':
            username = user_to_delete.username
            user_profile = getattr(user_to_delete, 'userprofile', None)
            role_display = user_profile.get_role_display() if user_profile else '알 수 없음'
            
            # 관련 데이터 개수 확인
            followups_count = FollowUp.objects.filter(user=user_to_delete).count()
            schedules_count = Schedule.objects.filter(user=user_to_delete).count()
            histories_count = History.objects.filter(user=user_to_delete).count()
            
            # 사용자와 관련된 모든 데이터가 CASCADE로 삭제됨
            # (models.py에서 ForeignKey의 on_delete=models.CASCADE 설정에 의해)
            user_to_delete.delete()
            
            # AJAX 요청인 경우 JSON 응답
            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                success_message = f'사용자 "{username}"이(가) 성공적으로 삭제되었습니다.'
                if followups_count > 0 or schedules_count > 0 or histories_count > 0:
                    success_message += f' (관련 데이터 {followups_count + schedules_count + histories_count}개도 함께 삭제되었습니다.)'
                
                return JsonResponse({
                    'success': True,
                    'message': success_message
                })
            
            # 일반 요청인 경우 리다이렉트
            success_message = f'사용자 "{username}"이(가) 성공적으로 삭제되었습니다.'
            if followups_count > 0 or schedules_count > 0 or histories_count > 0:
                success_message += f' (관련 데이터 {followups_count + schedules_count + histories_count}개도 함께 삭제되었습니다.)'
            
            messages.success(request, success_message)
            return redirect('reporting:user_list')
        
        # GET 요청인 경우 (확인 페이지)
        context = {
            'user_to_delete': user_to_delete,
            'page_title': f'사용자 삭제 - {user_to_delete.username}'
        }
        return render(request, 'reporting/user_delete.html', context)
        
    except Exception as e:
        # 예외 발생 시 상세 로깅
        error_msg = str(e)
        error_traceback = traceback.format_exc()
        
        logger.error(f"사용자 삭제 중 예외 발생:")
        logger.error(f"오류 메시지: {error_msg}")
        logger.error(f"스택 트레이스:\n{error_traceback}")
        
        # AJAX 요청인 경우 JSON 에러 응답
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': f'사용자 삭제 중 오류가 발생했습니다: {error_msg}',
                'detail': error_traceback if settings.DEBUG else None
            }, status=500)
        
        # 일반 요청인 경우 에러 메시지와 함께 리다이렉트
        messages.error(request, f'사용자 삭제 중 오류가 발생했습니다: {error_msg}')
        return redirect('reporting:user_list')

# ============ 매니저용 사용자 관리 뷰들 ============

@role_required(['manager'])
def manager_user_list(request):
    """매니저가 자신의 회사 소속 사용자 목록을 볼 수 있는 뷰"""
    # 매니저의 회사 정보 확인
    if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
        messages.error(request, '회사 정보가 없어 사용자 관리를 할 수 없습니다.')
        return redirect('reporting:main')
    
    manager_company = request.user.userprofile.company
    
    # 같은 회사의 사용자들만 조회 (매니저와 실무자만)
    users = User.objects.select_related('userprofile').filter(
        userprofile__company=manager_company,
        userprofile__role__in=['manager', 'salesman']
    ).order_by('username')
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # 역할별 필터
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(userprofile__role=role_filter)
    
    # 페이지네이션
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)
    
    context = {
        'users': users,
        'search_query': search_query,
        'role_filter': role_filter,
        'role_choices': [('manager', 'Manager (뷰어)'), ('salesman', 'SalesMan (실무자)')],
        'page_title': f'사용자 관리 - {manager_company.name}',
        'company_name': manager_company.name
    }
    return render(request, 'reporting/manager_user_list.html', context)

@role_required(['manager'])
def manager_user_create(request):
    """매니저가 자신의 회사에 실무자 계정을 추가하는 뷰"""
    # 매니저의 회사 정보 확인
    if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
        messages.error(request, '회사 정보가 없어 사용자 생성을 할 수 없습니다.')
        return redirect('reporting:main')
    
    manager_company = request.user.userprofile.company
    
    if request.method == 'POST':
        form = ManagerUserCreationForm(request.POST)
        
        if form.is_valid():
            # 사용자명 중복 체크
            username = form.cleaned_data['username']
            if User.objects.filter(username=username).exists():
                form.add_error('username', f'사용자명 "{username}"은(는) 이미 사용 중입니다.')
                messages.error(request, f'사용자명 "{username}"은(는) 이미 사용 중입니다.')
                context = {
                    'form': form,
                    'page_title': f'실무자 계정 생성 - {manager_company.name}',
                    'company_name': manager_company.name
                }
                return render(request, 'reporting/manager_user_create.html', context)
            
            try:
                # 사용자 생성
                user = User.objects.create_user(
                    username=username,
                    password=form.cleaned_data['password1'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name']
                )
                
                # 사용자 프로필 생성 (매니저와 같은 회사로 자동 설정)
                UserProfile.objects.create(
                    user=user,
                    company=manager_company,  # 매니저와 같은 회사
                    role='salesman',  # 매니저는 실무자만 생성 가능
                    can_download_excel=form.cleaned_data['can_download_excel'],
                    created_by=request.user  # 생성자 기록
                )
                
                messages.success(request, f'실무자 계정 "{user.username}"이(가) {manager_company.name}에 성공적으로 생성되었습니다.')
                return redirect('reporting:manager_user_list')
            except Exception as e:
                messages.error(request, f'사용자 생성 중 오류가 발생했습니다: {str(e)}')
                context = {
                    'form': form,
                    'page_title': f'실무자 계정 생성 - {manager_company.name}',
                    'company_name': manager_company.name
                }
                return render(request, 'reporting/manager_user_create.html', context)
    else:
        form = ManagerUserCreationForm()
    
    context = {
        'form': form,
        'page_title': f'실무자 계정 생성 - {manager_company.name}',
        'company_name': manager_company.name
    }
    return render(request, 'reporting/manager_user_create.html', context)

@role_required(['manager'])
def manager_user_edit(request, user_id):
    """매니저가 자신의 회사 소속 사용자를 편집하는 뷰"""
    # 매니저의 회사 정보 확인
    if not hasattr(request.user, 'userprofile') or not request.user.userprofile.company:
        messages.error(request, '회사 정보가 없어 사용자 편집을 할 수 없습니다.')
        return redirect('reporting:main')
    
    manager_company = request.user.userprofile.company
    
    # 편집할 사용자 가져오기 (같은 회사의 매니저/실무자만)
    user = get_object_or_404(User, id=user_id, userprofile__company=manager_company, userprofile__role__in=['manager', 'salesman'])
    user_profile = get_object_or_404(UserProfile, user=user)
    
    # 자기 자신의 권한은 변경할 수 없음
    if user.id == request.user.id:
        messages.error(request, '자신의 계정 정보는 이 방법으로 수정할 수 없습니다.')
        return redirect('reporting:manager_user_list')
    
    if request.method == 'POST':
        # 매니저용 편집 폼 (역할 변경 불가)
        form_data = request.POST.copy()
        form_data['role'] = user_profile.role  # 기존 역할 유지
        
        form = UserEditForm(form_data)
        if form.is_valid():
            # 사용자명 중복 체크 (자기 자신 제외)
            username = form.cleaned_data['username']
            if User.objects.filter(username=username).exclude(id=user.id).exists():
                messages.error(request, f'사용자명 "{username}"은(는) 이미 사용 중입니다.')
            else:
                # 사용자 정보 수정
                user.username = username
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                
                # 비밀번호 변경
                if form.cleaned_data['change_password'] and form.cleaned_data['password1']:
                    user.set_password(form.cleaned_data['password1'])
                
                user.save()
                
                # 엑셀 다운로드 권한만 수정 가능 (회사와 역할은 변경 불가)
                user_profile.can_download_excel = form.cleaned_data['can_download_excel']
                user_profile.save()
                
                messages.success(request, f'사용자 "{user.username}"의 정보가 성공적으로 수정되었습니다.')
                return redirect('reporting:manager_user_list')
    else:
        # 기존 데이터로 폼 초기화
        form = UserEditForm(initial={
            'username': user.username,
            'company': user_profile.company.name if user_profile.company else '',
            'first_name': user.first_name,
            'last_name': user.last_name,
            'role': user_profile.role,
            'can_download_excel': user_profile.can_download_excel,
        })
    
    context = {
        'form': form,
        'edit_user': user,
        'user_profile': user_profile,
        'page_title': f'사용자 편집 - {user.username}',
        'company_name': manager_company.name,
        'is_manager_edit': True  # 매니저 편집 모드 표시
    }
    return render(request, 'reporting/manager_user_edit.html', context)

@role_required(['manager'])
@never_cache
def manager_dashboard(request):
    """Manager 전용 대시보드 - dashboard_view로 리다이렉트"""
    from django.shortcuts import redirect
    
    # user_id 파라미터가 있으면 그대로 전달, 없으면 기본 대시보드로
    user_id = request.GET.get('user_id')
    if user_id:
        return redirect(f"{reverse('reporting:dashboard')}?user={user_id}")
    else:
        return redirect('reporting:dashboard')
    
    # 현재 시간
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    
    # 데이터 필터링
    if view_all:
        # 전체보기: 모든 salesman의 데이터
        followups = FollowUp.objects.filter(user__in=target_users)
        schedules = Schedule.objects.filter(user__in=target_users)
        histories = History.objects.filter(user__in=target_users)
        histories_current_year = History.objects.filter(user__in=target_users, created_at__year=current_year)
        
        followup_count = followups.count()
        schedule_count = schedules.filter(status='scheduled').count()
        sales_record_count = histories_current_year.filter(
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
    else:
        # 특정 사용자 데이터
        followups = FollowUp.objects.filter(user=target_user)
        schedules = Schedule.objects.filter(user=target_user)
        histories = History.objects.filter(user=target_user)
        histories_current_year = History.objects.filter(user=target_user, created_at__year=current_year)
        
        followup_count = followups.count()
        schedule_count = schedules.filter(status='scheduled').count()
        sales_record_count = histories_current_year.filter(
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
    
    # 납품 금액 통계 (현재 연도)
    history_delivery_stats = histories_current_year.filter(
        action_type='delivery_schedule',
        delivery_amount__isnull=False
    ).aggregate(
        total_amount=Sum('delivery_amount'),
        delivery_count=Count('id')
    )
    
    if view_all:
        schedule_delivery_stats = DeliveryItem.objects.filter(
            schedule__user__in=target_users,
            schedule__created_at__year=current_year,
            schedule__activity_type='delivery'
        ).exclude(
            schedule__status='cancelled'
        ).aggregate(
            total_amount=Sum('total_price'),
            delivery_count=Count('schedule', distinct=True)
        )
    else:
        schedule_delivery_stats = DeliveryItem.objects.filter(
            schedule__user=target_user,
            schedule__created_at__year=current_year,
            schedule__activity_type='delivery'
        ).exclude(
            schedule__status='cancelled'
        ).aggregate(
            total_amount=Sum('total_price'),
            delivery_count=Count('schedule', distinct=True)
        )
    
    history_amount = history_delivery_stats['total_amount'] or 0
    schedule_amount = schedule_delivery_stats['total_amount'] or 0
    total_delivery_amount = history_amount + schedule_amount
    
    # 중복 제거된 납품 횟수
    history_with_schedule = histories_current_year.filter(
        action_type='delivery_schedule',
        schedule__isnull=False
    ).values_list('schedule_id', flat=True).distinct()
    
    history_without_schedule = histories_current_year.filter(
        action_type='delivery_schedule',
        schedule__isnull=True
    ).count()
    
    if view_all:
        schedules_with_delivery = DeliveryItem.objects.filter(
            schedule__user__in=target_users,
            schedule__created_at__year=current_year,
            schedule__activity_type='delivery'
        ).values_list('schedule_id', flat=True).distinct()
    else:
        schedules_with_delivery = DeliveryItem.objects.filter(
            schedule__user=target_user,
            schedule__created_at__year=current_year,
            schedule__activity_type='delivery'
        ).values_list('schedule_id', flat=True).distinct()
    
    unique_schedule_ids = set(history_with_schedule) | set(schedules_with_delivery)
    delivery_count = len(unique_schedule_ids) + history_without_schedule
    
    # 활동 유형별 통계 (현재 연도)
    activity_stats = histories_current_year.exclude(action_type='memo').values('action_type').annotate(
        count=Count('id')
    ).order_by('action_type')
    
    # 서비스 통계
    service_count = histories_current_year.filter(action_type='service', service_status='completed').count()
    this_month_service_count = histories.filter(
        action_type='service',
        service_status='completed',
        created_at__month=current_month,
        created_at__year=current_year
    ).count()
    
    # 최근 활동 (5개)
    recent_activities = histories_current_year.exclude(action_type='memo').order_by('-created_at')[:5]
    
    # 월별 고객 추가 현황 (최근 6개월)
    monthly_customers = []
    for i in range(6):
        month_start = (now.replace(day=1) - timedelta(days=32*i)).replace(day=1)
        month_end = (month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1]))
        
        count = followups.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        ).count()
        
        monthly_customers.append({
            'month': month_start.strftime('%Y-%m'),
            'month_name': f"{month_start.year}년 {month_start.month}월",
            'count': count
        })
    monthly_customers.reverse()
    
    # 일정 완료율
    schedule_stats = schedules.aggregate(
        total=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        cancelled=Count('id', filter=Q(status='cancelled')),
        scheduled=Count('id', filter=Q(status='scheduled'))
    )
    
    completion_rate = 0
    if schedule_stats['total'] > 0:
        completion_rate = round((schedule_stats['completed'] / schedule_stats['total']) * 100, 1)
    
    # 영업 기록 추이 (최근 14일)
    fourteen_days_ago = now - timedelta(days=14)
    daily_activities = []
    for i in range(14):
        day = fourteen_days_ago + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        activity_count = histories_current_year.filter(
            created_at__gte=day_start,
            created_at__lt=day_end,
            action_type__in=['customer_meeting', 'delivery_schedule']
        ).count()
        
        daily_activities.append({
            'date': day.strftime('%m/%d'),
            'full_date': day.strftime('%Y-%m-%d'),
            'count': activity_count
        })
    
    # 오늘 일정
    today = now.date()
    today_schedules = schedules.filter(
        visit_date=today,
        status='scheduled'
    ).order_by('visit_time')[:5]
    
    # 최근 고객 (최근 7일)
    week_ago = now - timedelta(days=7)
    recent_customers = followups.filter(
        created_at__gte=week_ago
    ).order_by('-created_at')[:5]
    
    # 성과 지표
    monthly_revenue = histories.filter(
        action_type='delivery_schedule',
        created_at__month=current_month,
        created_at__year=current_year,
        delivery_amount__isnull=False
    ).aggregate(total=Sum('delivery_amount'))['total'] or 0
    
    # 월별 통계 (Schedule 기준)
    schedules_current_month = schedules.filter(
        visit_date__month=current_month,
        visit_date__year=current_year
    )
    
    monthly_meetings = schedules_current_month.filter(activity_type='customer_meeting').count()
    monthly_services = schedules_current_month.filter(activity_type='service', status='completed').count()
    
    # 전환율 계산 (Schedule 기준)
    schedules_current_year = schedules.filter(visit_date__year=current_year)
    total_meetings = schedules_current_year.filter(activity_type='customer_meeting').count()
    total_deliveries = schedules_current_year.filter(activity_type='delivery', status='completed').count()
    
    # 미팅 → 납품 전환율
    conversion_rate = (total_deliveries / total_meetings * 100) if total_meetings > 0 else 0
    
    # 평균 거래 규모 (DeliveryItem 기준)
    avg_deal_size = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year.filter(activity_type='delivery')
    ).aggregate(avg=Avg('total_price'))['avg'] or 0
    
    # 월별 납품 금액 데이터 (최근 6개월)
    monthly_revenue_data = []
    monthly_revenue_labels = []
    
    for i in range(5, -1, -1):
        target_date = now - timedelta(days=30*i)
        month_revenue = histories.filter(
            action_type='delivery_schedule',
            created_at__month=target_date.month,
            created_at__year=target_date.year,
            delivery_amount__isnull=False
        ).aggregate(total=Sum('delivery_amount'))['total'] or 0
        
        monthly_revenue_data.append(float(month_revenue))
        monthly_revenue_labels.append(f"{target_date.year}년 {target_date.month}월")
    
    # 고객별 납품 현황 (상위 5개)
    customer_revenue_data = histories_current_year.filter(
        action_type='delivery_schedule',
        delivery_amount__isnull=False,
        followup__isnull=False
    ).values('followup__customer_name', 'followup__company').annotate(
        total_revenue=Sum('delivery_amount')
    ).order_by('-total_revenue')[:5]
    
    customer_labels = [f"{item['followup__customer_name'] or '미정'} ({item['followup__company'] or '미정'})" for item in customer_revenue_data]
    customer_amounts = [float(item['total_revenue']) for item in customer_revenue_data]
    
    # ============================================
    # 📊 새로운 7개 차트를 위한 데이터 준비 - Schedule 기준
    # ============================================
    
    # 1️⃣ 매출 및 납품 추이 (월별 납품 금액 + 건수) - Schedule 기준
    monthly_delivery_stats = {
        'labels': [],
        'amounts': [],
        'counts': []
    }
    
    for i in range(11, -1, -1):  # 최근 12개월
        target_date = now - timedelta(days=30*i)
        month_start = target_date.replace(day=1)
        if target_date.month == 12:
            month_end = target_date.replace(year=target_date.year+1, month=1, day=1)
        else:
            month_end = target_date.replace(month=target_date.month+1, day=1)
        
        month_schedules = schedules.filter(
            visit_date__gte=month_start.date(),
            visit_date__lt=month_end.date(),
            activity_type='delivery',
            status__in=['scheduled', 'completed']  # 취소된 일정 제외
        )
        
        # 납품 금액 합산
        month_amount = DeliveryItem.objects.filter(
            schedule__in=month_schedules
        ).aggregate(total=Sum('total_price'))['total'] or 0
        
        # 납품 건수 (일정 개수)
        month_count = month_schedules.count()
        
        monthly_delivery_stats['labels'].append(f"{target_date.month}월")
        monthly_delivery_stats['amounts'].append(float(month_amount))
        monthly_delivery_stats['counts'].append(month_count)
    
    # 2️⃣ 영업 퍼널 (미팅 → 견적 제출 → 발주 예정 → 납품 완료)
    # 기준: 모두 일정(Schedule) 기반으로 집계
    schedules_current_year = schedules.filter(visit_date__year=current_year)
    
    meeting_count_mgr = schedules_current_year.filter(activity_type='customer_meeting').count()
    quote_count_mgr = schedules_current_year.filter(activity_type='quote').count()
    scheduled_delivery_count_mgr = schedules_current_year.filter(activity_type='delivery', status='scheduled').count()
    completed_delivery_count_mgr = schedules_current_year.filter(activity_type='delivery', status='completed').count()
    
    # 전환율 계산
    meeting_to_delivery_rate = (completed_delivery_count_mgr / meeting_count_mgr * 100) if meeting_count_mgr > 0 else 0
    quote_to_delivery_rate = (completed_delivery_count_mgr / quote_count_mgr * 100) if quote_count_mgr > 0 else 0
    
    sales_funnel = {
        'stages': ['미팅', '견적 제출', '발주 예정', '납품 완료'],
        'values': [
            meeting_count_mgr,
            quote_count_mgr,
            scheduled_delivery_count_mgr,
            completed_delivery_count_mgr
        ]
    }
    
    # 3️⃣ 고객사별 매출 비중 (Top 5 + 기타) - Schedule + History 기준
    # Schedule 기반 매출
    schedule_top_customers = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        schedule__followup__isnull=False,
        schedule__followup__company__isnull=False
    ).values('schedule__followup__company__name').annotate(
        total_revenue=Sum('total_price')
    )
    
    # History 기반 매출
    histories_current_year_with_company = histories_current_year.filter(
        followup__isnull=False,
        followup__company__isnull=False
    )
    
    history_top_customers = DeliveryItem.objects.filter(
        history__in=histories_current_year_with_company
    ).values('history__followup__company__name').annotate(
        total_revenue=Sum('total_price')
    )
    
    # 고객사별 매출 합산
    from collections import defaultdict
    company_revenue = defaultdict(float)
    
    for item in schedule_top_customers:
        company_name = item['schedule__followup__company__name'] or '미정'
        company_revenue[company_name] += float(item['total_revenue'])
    
    for item in history_top_customers:
        company_name = item['history__followup__company__name'] or '미정'
        company_revenue[company_name] += float(item['total_revenue'])
    
    # 상위 5개 추출
    sorted_companies = sorted(company_revenue.items(), key=lambda x: x[1], reverse=True)[:5]
    
    customer_distribution = {
        'labels': [],
        'data': []
    }
    
    total_top5_revenue = 0
    for company_name, revenue in sorted_companies:
        customer_distribution['labels'].append(company_name)
        customer_distribution['data'].append(revenue)
        total_top5_revenue += revenue
    
    # 기타 금액 계산 - Schedule + History 합산
    schedule_total = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__status__in=['scheduled', 'completed']
    ).aggregate(total=Sum('total_price'))['total'] or 0
    
    history_total = DeliveryItem.objects.filter(
        history__in=histories_current_year
    ).aggregate(total=Sum('total_price'))['total'] or 0
    
    total_all_revenue = float(schedule_total) + float(history_total)
    other_revenue = total_all_revenue - total_top5_revenue
    if other_revenue > 0:
        customer_distribution['labels'].append('기타')
        customer_distribution['data'].append(other_revenue)
    
    # 4️⃣ 영업 활동 추이 (월별) - Schedule 기준
    monthly_activity_breakdown = {
        'labels': [],
        'sales': []
    }
    
    for i in range(11, -1, -1):  # 최근 12개월
        target_date = now - timedelta(days=30*i)
        month_start = target_date.replace(day=1)
        if target_date.month == 12:
            month_end = target_date.replace(year=target_date.year+1, month=1, day=1)
        else:
            month_end = target_date.replace(month=target_date.month+1, day=1)
        
        sales_count_month = schedules.filter(
            visit_date__gte=month_start.date(),
            visit_date__lt=month_end.date(),
            activity_type__in=['customer_meeting', 'delivery', 'quote']
        ).count()
        
        monthly_activity_breakdown['labels'].append(f"{target_date.month}월")
        monthly_activity_breakdown['sales'].append(sales_count_month)
    
    # 5️⃣ 개인 성과 지표 추세 (납품액, 전환율, 평균 거래 규모) - Schedule 기준
    performance_trends = {
        'labels': [],
        'delivery_amount': [],
        'conversion_rate': [],
        'avg_deal_size': []
    }
    
    for i in range(11, -1, -1):  # 최근 12개월
        target_date = now - timedelta(days=30*i)
        month_start = target_date.replace(day=1)
        if target_date.month == 12:
            month_end = target_date.replace(year=target_date.year+1, month=1, day=1)
        else:
            month_end = target_date.replace(month=target_date.month+1, day=1)
        
        # 해당 월의 일정 데이터
        month_schedules = schedules.filter(
            visit_date__gte=month_start.date(),
            visit_date__lt=month_end.date()
        )
        
        month_meetings = month_schedules.filter(activity_type='customer_meeting').count()
        month_deliveries = month_schedules.filter(activity_type='delivery', status='completed').count()
        
        # 해당 월의 DeliveryItem 통계
        month_delivery_stats = DeliveryItem.objects.filter(
            schedule__in=month_schedules.filter(activity_type='delivery')
        ).aggregate(
            total=Sum('total_price'),
            avg=Avg('total_price')
        )
        
        month_conversion = (month_deliveries / month_meetings * 100) if month_meetings > 0 else 0
        
        performance_trends['labels'].append(f"{target_date.month}월")
        performance_trends['delivery_amount'].append(float(month_delivery_stats['total'] or 0) / 1000000)  # 백만원 단위
        performance_trends['conversion_rate'].append(round(month_conversion, 1))
        performance_trends['avg_deal_size'].append(float(month_delivery_stats['avg'] or 0) / 1000000)  # 백만원 단위
    
    # 6️⃣ 고객 유형별 통계 (대학/기업/관공서) - Schedule + History 기준
    customer_type_stats = {
        'labels': ['대학', '기업', '관공서'],
        'revenue': [0, 0, 0],
        'count': [0, 0, 0]
    }
    
    # Schedule 기반 통계 (납품 일정만 카운트 - 견적 제외)
    schedule_company_stats = DeliveryItem.objects.filter(
        schedule__in=schedules_current_year,
        schedule__activity_type='delivery',  # 납품만 카운트 (견적 제외)
        schedule__status__in=['scheduled', 'completed'],  # 취소된 일정 제외
        schedule__followup__isnull=False,
        schedule__followup__company__isnull=False
    ).values('schedule__followup__company__name').annotate(
        total_revenue=Sum('total_price'),
        count=Count('schedule', distinct=True)  # 일정 개수로 카운팅
    )
    
    for item in schedule_company_stats:
        company_name = item['schedule__followup__company__name'] or ''
        revenue = float(item['total_revenue']) / 1000000  # 백만원 단위
        cnt = item['count']
        
        if '대학' in company_name or '대학교' in company_name:
            customer_type_stats['revenue'][0] += revenue
            customer_type_stats['count'][0] += cnt
        elif '청' in company_name or '부' in company_name or '시' in company_name or '구' in company_name:
            customer_type_stats['revenue'][2] += revenue
            customer_type_stats['count'][2] += cnt
        else:
            customer_type_stats['revenue'][1] += revenue
            customer_type_stats['count'][1] += cnt
    
    # History 기반 통계 (납품 이력만 카운트 - 견적 제외)
    history_company_stats = DeliveryItem.objects.filter(
        history__in=histories_current_year,
        history__action_type='delivery_schedule',  # 납품만 카운트 (History는 action_type 필드 사용)
        history__followup__isnull=False,
        history__followup__company__isnull=False
    ).values('history__followup__company__name').annotate(
        total_revenue=Sum('total_price'),
        count=Count('history', distinct=True)  # 히스토리 개수로 카운팅
    )
    
    for item in history_company_stats:
        company_name = item['history__followup__company__name'] or ''
        revenue = float(item['total_revenue']) / 1000000  # 백만원 단위
        cnt = item['count']
        
        if '대학' in company_name or '대학교' in company_name:
            customer_type_stats['revenue'][0] += revenue
            customer_type_stats['count'][0] += cnt
        elif '청' in company_name or '부' in company_name or '시' in company_name or '구' in company_name:
            customer_type_stats['revenue'][2] += revenue
            customer_type_stats['count'][2] += cnt
        else:
            customer_type_stats['revenue'][1] += revenue
            customer_type_stats['count'][1] += cnt
    
    # 7️⃣ 활동 히트맵 (현재 달) - 선택된 사용자(들)의 일정만
    daily_activity_heatmap = []
    
    # 현재 달의 첫날과 마지막 날 계산
    current_month_start = now.replace(day=1).date()
    if now.month == 12:
        next_month = now.replace(year=now.year + 1, month=1, day=1)
    else:
        next_month = now.replace(month=now.month + 1, day=1)
    current_month_end = (next_month - timedelta(days=1)).date()
    
    # 현재 달의 각 날짜별 활동 카운트 (선택된 사용자만)
    current_date = current_month_start
    while current_date <= current_month_end:
        day_activity_count = schedules.filter(
            visit_date=current_date
        ).count()
        
        daily_activity_heatmap.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'day_of_week': current_date.weekday(),  # 0=월, 6=일
            'intensity': day_activity_count
        })
        
        current_date += timedelta(days=1)
    
    # 선결제 통계 - 등록자 본인만 (Manager도 자신이 등록한 것만)
    from decimal import Decimal
    
    prepayments = Prepayment.objects.filter(created_by=target_user)
    
    prepayment_aggregate = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance')
    )
    
    prepayment_stats = {
        'total_amount': prepayment_aggregate['total_amount'] or Decimal('0'),
        'total_balance': prepayment_aggregate['total_balance'] or Decimal('0'),
        'total_used': (prepayment_aggregate['total_amount'] or Decimal('0')) - (prepayment_aggregate['total_balance'] or Decimal('0')),
        'active_count': prepayments.filter(status='active', balance__gt=0).count(),
        'depleted_count': prepayments.filter(status='depleted').count(),
        'total_count': prepayments.count(),
    }
    
    # Context 구성
    context = {
        'page_title': f"Manager 대시보드 - {'전체보기' if view_all else target_user.username}",
        'current_year': current_year,
        'followup_count': followup_count,
        'schedule_count': schedule_count,
        'sales_record_count': sales_record_count,
        'total_delivery_amount': total_delivery_amount,
        'delivery_count': delivery_count,
        'service_count': service_count,
        'this_month_service_count': this_month_service_count,
        'activity_stats': activity_stats,
        'recent_activities': recent_activities,
        'monthly_customers': monthly_customers,
        'schedule_stats': schedule_stats,
        'completion_rate': completion_rate,
        'daily_activities': daily_activities,
        'today_schedules': today_schedules,
        'recent_customers': recent_customers,
        'monthly_revenue': monthly_revenue,
        'monthly_meetings': monthly_meetings,
        'monthly_services': monthly_services,
        'conversion_rate': conversion_rate,
        'meeting_to_delivery_rate': meeting_to_delivery_rate,
        'quote_to_delivery_rate': quote_to_delivery_rate,
        'avg_deal_size': avg_deal_size,
        'monthly_revenue_data': json.dumps(monthly_revenue_data, cls=DjangoJSONEncoder),
        'monthly_revenue_labels': json.dumps(monthly_revenue_labels, cls=DjangoJSONEncoder),
        'customer_labels': json.dumps(customer_labels, cls=DjangoJSONEncoder),
        'customer_amounts': json.dumps(customer_amounts, cls=DjangoJSONEncoder),
        # 새로운 7개 차트 데이터
        'monthly_delivery_stats': json.dumps(monthly_delivery_stats, cls=DjangoJSONEncoder),
        'sales_funnel': json.dumps(sales_funnel, cls=DjangoJSONEncoder),
        'customer_distribution': json.dumps(customer_distribution, cls=DjangoJSONEncoder),
        'customer_type_stats': json.dumps(customer_type_stats, cls=DjangoJSONEncoder),
        'daily_activity_heatmap': json.dumps(daily_activity_heatmap, cls=DjangoJSONEncoder),
        # 선결제 통계 (실무자 대시보드와 동일한 형식)
        'prepayment_stats': prepayment_stats,
        # 관리자용 추가 정보
        'salesman_users': salesman_users,
        'selected_user': target_user,
        'view_all': view_all,
    }
    
    return render(request, 'reporting/dashboard.html', context)


@role_required(['manager'])
def salesman_detail(request, user_id):
    """특정 Salesman의 상세 정보 조회 (Manager, Admin 전용)"""
    try:
        # 먼저 user_id가 유효한 정수인지 확인
        try:
            user_id = int(user_id)
        except (ValueError, TypeError):
            messages.error(request, '잘못된 사용자 ID입니다.')
            return redirect('reporting:manager_dashboard')
        
        # 접근 권한 확인
        accessible_users = get_accessible_users(request.user, request)
        
        # 해당 사용자가 존재하고 접근 가능한지 확인
        try:
            selected_user = accessible_users.get(id=user_id)
        except User.DoesNotExist:
            messages.error(request, '해당 사용자를 찾을 수 없거나 접근 권한이 없습니다.')
            return redirect('reporting:manager_dashboard')
        
        # 사용자 프로필 확인
        try:
            user_profile = selected_user.userprofile
        except UserProfile.DoesNotExist:
            messages.error(request, '사용자 프로필이 설정되지 않았습니다.')
            return redirect('reporting:manager_dashboard')
        
        # 해당 사용자의 데이터만 필터링 (select_related로 성능 최적화)
        followups = FollowUp.objects.filter(user=selected_user).select_related('user')
        schedules = Schedule.objects.filter(user=selected_user).select_related('user', 'followup')
        histories = History.objects.filter(user=selected_user).select_related('user', 'followup', 'schedule')
        
        # 검색 및 필터링
        search_query = request.GET.get('search', '').strip()
        status_filter = request.GET.get('status', '').strip()
        
        if search_query:
            followups = followups.filter(
                Q(customer_name__icontains=search_query) |
                Q(company__icontains=search_query)
            )
        
        if status_filter and status_filter in ['initial', 'in_progress', 'visited', 'closed']:
            followups = followups.filter(status=status_filter)
        
        # 정렬 추가
        followups = followups.order_by('-updated_at')
        
        # 페이지네이션 전에 총 개수 계산 (안전하게)
        try:
            total_followups = followups.count()
        except Exception as e:
            total_followups = 0
        
        # 페이지네이션 처리 (단순화)
        try:
            from django.core.paginator import Paginator
            paginator = Paginator(followups, 10)
            page_number = request.GET.get('page', 1)
            try:
                page_number = int(page_number)
            except (ValueError, TypeError):
                page_number = 1
            followups = paginator.get_page(page_number)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"페이지네이션 처리 중 오류: {str(e)}")
            # 페이지네이션 실패 시 원본 쿼리셋 유지하되 첫 10개만
            followups = followups[:10]
        
        # 집계 값들을 안전하게 계산
        try:
            total_schedules = schedules.count()
            total_histories = histories.count()
        except Exception as e:
            total_schedules = 0
            total_histories = 0
        
        context = {
            'selected_user': selected_user,
            'followups': followups,
            'search_query': search_query,
            'status_filter': status_filter,
            'total_followups': total_followups,
            'total_schedules': total_schedules,
            'total_histories': total_histories,
            'page_title': f'{selected_user.username} 상세 정보'
        }
        return render(request, 'reporting/salesman_detail.html', context)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"salesman_detail 오류 - user_id: {user_id}, user: {request.user}, error: {str(e)}")
        messages.error(request, f'사용자 상세 정보를 불러오는 중 오류가 발생했습니다. 관리자에게 문의하세요.')
        return redirect('reporting:manager_dashboard')

@role_required(['admin'])
@require_POST
def user_toggle_active(request, user_id):
    """사용자 활성화/비활성화 토글 (Admin 전용)"""
    try:
        user = get_object_or_404(User, id=user_id)
        
        # 자기 자신은 비활성화할 수 없음
        if user == request.user:
            return JsonResponse({'error': '자기 자신의 계정은 비활성화할 수 없습니다.'}, status=400)
        
        # 상태 토글
        user.is_active = not user.is_active
        user.save()
        
        status_text = "활성화" if user.is_active else "비활성화"
        messages.success(request, f'사용자 "{user.username}"이(가) {status_text}되었습니다.')
        
        return JsonResponse({
            'success': True,
            'is_active': user.is_active,
            'message': f'사용자가 {status_text}되었습니다.'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@role_required(['admin'])
@require_POST
def user_toggle_ai(request, user_id):
    """사용자 AI 권한 토글 (Admin 전용)"""
    try:
        user = get_object_or_404(User, id=user_id)
        profile = get_object_or_404(UserProfile, user=user)

        profile.can_use_ai = not profile.can_use_ai
        profile.save(update_fields=['can_use_ai'])

        status_text = "부여" if profile.can_use_ai else "해제"
        return JsonResponse({
            'success': True,
            'can_use_ai': profile.can_use_ai,
            'message': f'사용자 "{user.username}"의 AI 권한이 {status_text}되었습니다.'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
@never_cache
def toggle_tax_invoice(request, history_id):
    """세금계산서 발행 여부 토글 (AJAX)"""
    try:
        history = get_object_or_404(History, id=history_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 토글 가능 (Manager는 읽기 전용)
        try:
            if not can_modify_user_data(request.user, history.user):
                return JsonResponse({
                    'success': False, 
                    'error': '수정 권한이 없습니다. Manager는 읽기 전용입니다.'
                }, status=403)
        except Exception as e:
            # 권한 체크 실패 시 자신의 히스토리인지만 확인
            if request.user != history.user:
                return JsonResponse({
                    'success': False, 
                    'error': '수정 권한이 없습니다.'
                }, status=403)
        
        # 납품 일정 히스토리인지 확인
        if history.action_type != 'delivery_schedule':
            return JsonResponse({
                'success': False,
                'error': '납품 일정 히스토리만 세금계산서 토글이 가능합니다.'
            }, status=400)
        
        # 토글 실행
        history.tax_invoice_issued = not history.tax_invoice_issued
        history.save()
        
        # 연결된 Schedule의 DeliveryItem들도 동기화
        if history.schedule:
            from reporting.models import DeliveryItem
            DeliveryItem.objects.filter(schedule=history.schedule).update(
                tax_invoice_issued=history.tax_invoice_issued
            )
        
        return JsonResponse({
            'success': True,
            'tax_invoice_issued': history.tax_invoice_issued,
            'message': f'세금계산서 발행 여부가 {"발행완료" if history.tax_invoice_issued else "미발행"}로 변경되었습니다.'
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'오류가 발생했습니다: {str(e)}'
        }, status=500)

@login_required
def toggle_all_tax_invoices(request, followup_id):
    """고객의 모든 미발행 납품을 발행완료로 일괄 변경 (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '잘못된 요청입니다.'}, status=400)
    
    try:
        # FollowUp 조회 및 권한 확인
        followup = get_object_or_404(FollowUp, id=followup_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 가능
        if not can_modify_user_data(request.user, followup.user):
            return JsonResponse({
                'success': False, 
                'error': '수정 권한이 없습니다. Manager는 읽기 전용입니다.'
            }, status=403)
        
        # 미발행 납품 History 조회
        pending_histories = History.objects.filter(
            followup=followup,
            action_type='delivery_schedule',
            tax_invoice_issued=False
        )
        
        # 일괄 업데이트
        updated_count = pending_histories.update(tax_invoice_issued=True)
        
        # 연결된 Schedule의 DeliveryItem들도 동기화
        from reporting.models import DeliveryItem, Schedule
        schedule_ids = pending_histories.filter(
            schedule__isnull=False
        ).values_list('schedule_id', flat=True)
        
        if schedule_ids:
            DeliveryItem.objects.filter(
                schedule_id__in=schedule_ids
            ).update(tax_invoice_issued=True)
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'message': f'{updated_count}건의 납품이 발행완료로 변경되었습니다.'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'오류가 발생했습니다: {str(e)}'
        }, status=500)

@login_required
def history_create_from_schedule(request, schedule_id):
    """일정에서 히스토리 생성 또는 기존 히스토리로 이동"""
    try:
        # Manager는 히스토리 생성 불가 (뷰어 권한)
        _profile = get_user_profile(request.user)
        if _profile.is_manager():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': '권한이 없습니다. Manager는 활동을 기록할 수 없습니다.'}, status=403)
            messages.error(request, '권한이 없습니다. Manager는 활동을 기록할 수 없습니다.')
            return redirect('reporting:schedule_list')

        schedule = get_object_or_404(Schedule, pk=schedule_id)
        
        # 권한 체크: 자신의 일정이거나 관리 권한이 있는 경우만 허용
        try:
            if not can_access_user_data(request.user, schedule.user):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': '이 일정에 대한 히스토리를 생성할 권한이 없습니다.'})
                messages.error(request, '이 일정에 대한 히스토리를 생성할 권한이 없습니다.')
                return redirect('reporting:schedule_list')
        except Exception as e:
            # 권한 체크 실패 시 자신의 일정인지만 확인
            if request.user != schedule.user:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': '이 일정에 대한 히스토리를 생성할 권한이 없습니다.'})
                messages.error(request, '이 일정에 대한 히스토리를 생성할 권한이 없습니다.')
                return redirect('reporting:schedule_list')
        
        # 스케줄의 activity_type에 따른 action_type 매핑
        action_type_mapping = {
            'customer_meeting': 'customer_meeting',
            'delivery': 'delivery_schedule', 
            'service': 'service_visit',
            'quote': 'quote_submission',  # 견적 일정 -> 견적 제출 히스토리
        }
        expected_action_type = action_type_mapping.get(schedule.activity_type, 'customer_meeting')
        
        # AJAX 요청으로 기존 히스토리 확인하는 경우
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.method == 'GET':
            existing_history = History.objects.filter(
                schedule=schedule,
                action_type=expected_action_type
            ).first()
            
            if existing_history:
                return JsonResponse({
                    'success': True,
                    'has_existing': True,
                    'message': f'이미 "{schedule.followup.customer_name}" 일정에 대한 활동 기록이 존재합니다.',
                    'history_id': existing_history.pk,
                    'history_url': f'/reporting/histories/{existing_history.pk}/',
                    'customer_name': schedule.followup.customer_name,
                    'visit_date': schedule.visit_date.strftime('%Y년 %m월 %d일')
                })
            else:
                return JsonResponse({
                    'success': True,
                    'has_existing': False,
                    'message': '새로운 활동 기록을 생성할 수 있습니다.',
                    'create_url': f'/reporting/histories/create-from-schedule/{schedule.pk}/',
                    'customer_name': schedule.followup.customer_name,
                    'visit_date': schedule.visit_date.strftime('%Y년 %m월 %d일')
                })
        
        # 기존 히스토리가 있는지 확인 (일반 GET 요청일 때만)
        if request.method == 'GET':
            existing_history = History.objects.filter(
                schedule=schedule,
                action_type=expected_action_type
            ).first()
            
            if existing_history:
                messages.info(request, f'이미 "{schedule.followup.customer_name}" 일정에 대한 활동 기록이 존재합니다. 기존 기록으로 이동합니다.')
                return redirect('reporting:history_detail', pk=existing_history.pk)
        
        
        if request.method == 'POST':
            # 먼저 기존 History가 있는지 확인 (중복 생성 방지)
            existing_history = History.objects.filter(
                schedule=schedule,
                action_type=expected_action_type
            ).first()
            
            if existing_history:
                # 이미 History가 있으면 생성하지 않고 기존 것으로 리다이렉트
                is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'error': f'이미 "{schedule.followup.customer_name}" 일정에 대한 활동 기록이 존재합니다.',
                        'history_id': existing_history.pk,
                        'history_url': f'/reporting/histories/{existing_history.pk}/'
                    })
                else:
                    messages.warning(request, f'이미 "{schedule.followup.customer_name}" 일정에 대한 활동 기록이 존재합니다. 기존 기록으로 이동합니다.')
                    return redirect('reporting:history_detail', pk=existing_history.pk)
            
            # AJAX 요청인지 확인
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            # 인라인 폼에서 온 데이터를 위해 followup과 schedule을 자동 설정
            post_data = request.POST.copy()
            post_data['followup'] = schedule.followup.id
            post_data['schedule'] = schedule.id
            
            # action_type이 POST 데이터에 없으면 일정의 activity_type에 따라 자동 설정
            if 'action_type' not in post_data or not post_data['action_type']:
                post_data['action_type'] = expected_action_type
            
            form = HistoryForm(post_data, user=request.user, request=request)
            if form.is_valid():
                history = form.save(commit=False)
                history.user = request.user
                history.followup = schedule.followup  # 일정의 팔로우업으로 강제 설정
                history.schedule = schedule  # 일정 연결
                  
                # 납품 일정인 경우 delivery_date가 설정되지 않았다면 일정 날짜로 설정
                if history.action_type == 'delivery_schedule' and not history.delivery_date:
                    history.delivery_date = schedule.visit_date
                
                # 고객 미팅인 경우 meeting_date가 설정되지 않았다면 일정 날짜로 설정
                if history.action_type == 'customer_meeting' and not history.meeting_date:
                    history.meeting_date = schedule.visit_date
                
                # 견적/납품 유형은 delivery_amount=0으로 강제 (더블 매출 방지)
                if history.action_type in ['delivery_schedule', 'quote_submission']:
                    history.delivery_amount = 0
                    
                history.save()

                # 파이프라인 자동 진행 (고객 미팅 → contact 이상으로 앞으로만 이동)
                try:
                    if history.action_type == 'customer_meeting' and history.followup:
                        from .funnel_views import _try_advance_pipeline
                        _try_advance_pipeline(history.followup, 'contact')
                except Exception:
                    pass  # 파이프라인 업데이트 실패해도 히스토리 저장은 유지

                # 파일 업로드 처리 (공통 validate_file_upload 사용)
                uploaded_files = request.FILES.getlist('files')
                for uploaded_file in uploaded_files:
                    if not uploaded_file:
                        continue
                    is_valid, _ = validate_file_upload(uploaded_file)
                    if not is_valid:
                        continue  # 검증 실패 파일은 건너뛰기
                    HistoryFile.objects.create(
                        history=history,
                        file=uploaded_file,
                        original_filename=uploaded_file.name,
                        file_size=uploaded_file.size,
                        uploaded_by=request.user
                    )
                
                if is_ajax:
                    # AJAX 요청인 경우 JSON 응답
                    return JsonResponse({
                        'success': True,
                        'message': f'"{schedule.followup.customer_name}" 일정에 대한 활동 히스토리가 성공적으로 기록되었습니다.',
                        'history_id': history.id
                    })
                else:
                    # 일반 폼 제출인 경우 일정 상세 페이지로 리다이렉트
                    messages.success(request, f'"{schedule.followup.customer_name}" 일정에 대한 활동 히스토리가 성공적으로 기록되었습니다.')
                    return redirect('reporting:schedule_detail', pk=schedule.pk)
            else:
                # 폼 검증 실패
                if is_ajax:
                    # AJAX 요청인 경우 오류 응답
                    errors = []
                    for field, field_errors in form.errors.items():
                        for error in field_errors:
                            field_label = form.fields.get(field, {}).get('label', field) if hasattr(form.fields, 'get') else field
                            errors.append(f"{field_label}: {error}")
                    return JsonResponse({
                        'success': False,
                        'error': '입력 정보를 확인해주세요: ' + ', '.join(errors),
                        'form_errors': form.errors
                    })
                else:
                    # 일반 폼 제출인 경우 기존 동작
                    messages.error(request, '입력 정보를 확인해주세요.')
                    # POST 실패 시에도 일정 정보를 유지하기 위해 폼 필드를 다시 설정
                    form.fields['followup'].queryset = FollowUp.objects.filter(id=schedule.followup.id)
                    form.fields['schedule'].queryset = Schedule.objects.filter(id=schedule.id)
                    form.fields['followup'].initial = schedule.followup
                    form.fields['schedule'].initial = schedule
                    # 필드를 읽기 전용으로 설정
                    form.fields['followup'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
                    form.fields['schedule'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
        else:
            # GET 요청 시 폼 초기화
            # 스케줄의 활동 유형에 따라 히스토리 action_type 설정
            action_type_mapping = {
                'customer_meeting': 'customer_meeting',
                'delivery': 'delivery_schedule', 
                'service': 'service_visit',
                'quote': 'quote_submission',  # 견적 일정 -> 견적 제출 히스토리
            }
            initial_action_type = action_type_mapping.get(schedule.activity_type, 'customer_meeting')
            
            initial_data = {
                'followup': schedule.followup.id,
                'schedule': schedule.id,
                'action_type': initial_action_type,  # 스케줄의 activity_type에 맞춰 설정
                'delivery_date': schedule.visit_date,  # 납품 날짜를 일정 날짜로 설정
                'meeting_date': schedule.visit_date,  # 미팅 날짜를 일정 날짜로 설정
            }
            form = HistoryForm(user=request.user, request=request, initial=initial_data)
            
            # 팔로우업과 일정 필드를 해당 일정으로 고정
            form.fields['followup'].queryset = FollowUp.objects.filter(id=schedule.followup.id)
            form.fields['schedule'].queryset = Schedule.objects.filter(id=schedule.id)
            
            # 필드를 읽기 전용으로 설정 (disabled 대신 시각적으로 비활성화)
            form.fields['followup'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
            form.fields['schedule'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
        
        # 스케줄의 납품 품목 정보 가져오기
        delivery_text = None
        delivery_amount = 0
        
        if schedule.activity_type == 'delivery':
            # 1차: DeliveryItem 모델에서 최신 데이터 확인
            delivery_items = schedule.delivery_items_set.all().order_by('id')
            
            if delivery_items.exists():
                delivery_text_parts = []
                total_amount = 0
                
                for item in delivery_items:
                    # VAT 포함 총액 계산
                    item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
                    total_amount += item_total
                    
                    # 텍스트 형태로 변환
                    text_part = f"{item.item_name}: {item.quantity}개 ({int(item_total):,}원)"
                    delivery_text_parts.append(text_part)
                
                delivery_text = '\n'.join(delivery_text_parts)
                delivery_amount = int(total_amount)
            
            # 2차: DeliveryItem이 없으면 History에서 fallback
            if not delivery_text:
                related_histories = History.objects.filter(schedule=schedule, action_type='delivery_schedule').order_by('-created_at')
                
                if related_histories.exists():
                    latest_delivery = related_histories.first()
                    if latest_delivery.delivery_items:
                        delivery_text = latest_delivery.delivery_items.replace('\\n', '\n')
                    if latest_delivery.delivery_amount:
                        delivery_amount = latest_delivery.delivery_amount
        
        context = {
            'form': form,
            'schedule': schedule,
            'delivery_text': delivery_text,
            'delivery_amount': delivery_amount,
            'page_title': f'활동 기록 추가 - {schedule.followup.customer_name} (일정: {schedule.visit_date})'
        }
        return render(request, 'reporting/history_form.html', context)
        
    except Exception as e:
        # 전체적인 오류 처리
        import traceback
        error_msg = f"오류 발생: {str(e)}"
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg})
        
        messages.error(request, error_msg)
        return redirect('reporting:schedule_list')
        initial_data = {
            'followup': schedule.followup.id,
            'schedule': schedule.id,
            'action_type': 'customer_meeting',  # 기본값으로 고객 미팅 설정 (올바른 값)
            'delivery_date': schedule.visit_date,  # 납품 날짜를 일정 날짜로 설정
            'meeting_date': schedule.visit_date,  # 미팅 날짜를 일정 날짜로 설정
        }
        form = HistoryForm(user=request.user, request=request, initial=initial_data)
        
        # 팔로우업과 일정 필드를 해당 일정으로 고정
        form.fields['followup'].queryset = FollowUp.objects.filter(id=schedule.followup.id)
        form.fields['schedule'].queryset = Schedule.objects.filter(id=schedule.id)
        
        # 필드를 읽기 전용으로 설정 (disabled 대신 시각적으로 비활성화)
        form.fields['followup'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
        form.fields['schedule'].widget.attrs.update({'readonly': True, 'style': 'pointer-events: none; background-color: #e9ecef;'})
    
    context = {
        'form': form,
        'schedule': schedule,
        'page_title': f'활동 기록 추가 - {schedule.followup.customer_name} (일정: {schedule.visit_date})'
    }
    return render(request, 'reporting/history_form.html', context)

@login_required
@require_POST
def schedule_move_api(request, pk):
    """일정을 다른 날짜로 이동하는 API"""
    schedule = get_object_or_404(Schedule, pk=pk)
    
    # 권한 체크: 수정 권한이 있는 경우만 이동 가능 (Manager는 읽기 전용)
    if not can_modify_user_data(request.user, schedule.user):
        return JsonResponse({'success': False, 'error': '이 일정을 이동할 권한이 없습니다. Manager는 읽기 전용입니다.'}, status=403)
    
    try:
        # POST 데이터에서 새로운 날짜 정보 가져오기 (FormData 형식)
        new_date = request.POST.get('new_date')  # 'YYYY-MM-DD' 형식
        
        if not new_date:
            return JsonResponse({'success': False, 'error': '새로운 날짜가 제공되지 않았습니다.'}, status=400)
        
        # 날짜 형식 검증 및 변환
        from datetime import datetime
        try:
            new_visit_date = datetime.strptime(new_date, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': '잘못된 날짜 형식입니다.'}, status=400)
        
        # 일정 날짜 업데이트
        old_date = schedule.visit_date
        schedule.visit_date = new_visit_date
        schedule.save()
        
        return JsonResponse({
            'success': True, 
            'message': f'일정이 {old_date.strftime("%Y년 %m월 %d일")}에서 {new_visit_date.strftime("%Y년 %m월 %d일")}로 이동되었습니다.',
            'old_date': old_date.strftime('%Y-%m-%d'),
            'new_date': new_visit_date.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'일정 이동 중 오류가 발생했습니다: {str(e)}'}, status=500)

@login_required
@require_POST
def schedule_status_update_api(request, schedule_id):
    """일정 상태 업데이트 API"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 상태 변경 가능 (Manager는 읽기 전용)
        if not can_modify_user_data(request.user, schedule.user):
            return JsonResponse({'error': '수정 권한이 없습니다. Manager는 읽기 전용입니다.'}, status=403)
        
        new_status = request.POST.get('status')
        
        if new_status not in ['scheduled', 'completed', 'cancelled']:
            return JsonResponse({'error': '잘못된 상태값입니다.'}, status=400)
        
        # 견적 일정은 완료로 변경 불가 (취소만 가능)
        if schedule.activity_type == 'quote' and new_status == 'completed':
            return JsonResponse({
                'error': '견적 일정은 완료 상태로 변경할 수 없습니다. 견적은 취소만 가능합니다.'
            }, status=400)
        
        old_status = schedule.status
        
        # 취소 처리 시 추가 작업
        if new_status == 'cancelled' and old_status != 'cancelled':
            from datetime import date
            from reporting.models import DeliveryItem, History
            
            # 1. 납품 품목 기록은 유지 (삭제하지 않음 - 카운팅에서만 제외)
            delivery_items = DeliveryItem.objects.filter(schedule=schedule)
            
            # 2. 관련 납품 히스토리 삭제 (납품 활동 기록)
            delivery_histories = History.objects.filter(schedule=schedule, action_type='delivery')
            delivery_histories_count = delivery_histories.count()
            
            if delivery_histories_count > 0:
                delivery_histories.delete()
        
        schedule.status = new_status
        schedule.save()
        
        status_display = {
            'scheduled': '예정됨',
            'completed': '완료됨', 
            'cancelled': '취소됨'
        }
        
        response_data = {
            'success': True,
            'new_status': new_status,
            'status_display': status_display[new_status],
            'message': f'일정 상태가 "{status_display[new_status]}"로 변경되었습니다.',
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"❌ 일정 상태 업데이트 오류: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_POST
def schedule_add_memo_api(request, schedule_id):
    """일정에 매니저 메모 추가 API"""
    try:
        schedule = get_object_or_404(Schedule, id=schedule_id)
        user_profile = get_user_profile(request.user)
        
        # 권한 체크: 매니저만 메모 추가 가능하고, 해당 실무자의 일정에만 접근 가능
        if not user_profile.is_manager():
            return JsonResponse({'error': '매니저만 메모를 추가할 수 있습니다.'}, status=403)
        
        if not can_access_user_data(request.user, schedule.user):
            return JsonResponse({'error': '이 일정에 접근할 권한이 없습니다.'}, status=403)
        
        memo_content = request.POST.get('memo', '').strip()
        if not memo_content:
            return JsonResponse({'error': '메모 내용을 입력해주세요.'}, status=400)
        
        # 매니저 메모를 히스토리로 생성
        memo_history = History.objects.create(
            followup=schedule.followup,
            user=schedule.user,  # 일정의 원래 담당자를 유지
            action_type='memo',
            content=f"[매니저 메모 - {request.user.username}] {memo_content}",  # 매니저 메모 표시 추가
            created_by=request.user,  # 실제 작성자는 매니저
            schedule=schedule
        )
        
        return JsonResponse({
            'success': True,
            'message': '매니저 메모가 추가되었습니다.',
            'memo': {
                'id': memo_history.id,
                'content': memo_history.content,
                'created_at': memo_history.created_at.strftime('%Y-%m-%d %H:%M'),
                'created_by': request.user.username
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': f'메모 추가 중 오류가 발생했습니다: {str(e)}'}, status=500)

# 자동완성 API 뷰들
@login_required
def company_autocomplete(request):
    """업체/학교명 자동완성 API"""
    import logging
    logger = logging.getLogger(__name__)
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 1:
        return JsonResponse({'results': []})
    
    # Admin 사용자는 모든 업체 검색 가능
    if getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin'):
        companies = Company.objects.filter(name__icontains=query).order_by('name')[:10]
    else:
        # 일반 사용자: 같은 회사 소속 사용자들이 생성한 업체만 검색 가능
        user_company = getattr(request, 'user_company', None)
        user_profile = getattr(request.user, 'userprofile', None)
        
        if user_company:
            # 미들웨어에서 설정한 user_company 사용
            same_company_users = User.objects.filter(userprofile__company=user_company)
            
            companies = Company.objects.filter(
                name__icontains=query,
                created_by__in=same_company_users
            ).order_by('name')[:10]
            
        elif user_profile and user_profile.company:
            # 백업: UserProfile에서 직접 가져오기
            same_company_users = User.objects.filter(userprofile__company=user_profile.company)
            
            companies = Company.objects.filter(
                name__icontains=query,
                created_by__in=same_company_users
            ).order_by('name')[:10]
        else:
            companies = Company.objects.none()
    
    results = []
    for company in companies:
        results.append({
            'id': company.id,
            'text': company.name
        })
    
    return JsonResponse({'results': results})

@login_required
def department_autocomplete(request):
    """부서/연구실명 자동완성 API"""
    import logging
    logger = logging.getLogger(__name__)
    
    query = request.GET.get('q', '').strip()
    company_id = request.GET.get('company') or request.GET.get('company_id')  # 둘 다 지원
    
    if len(query) < 1:
        return JsonResponse({'results': []})
    
    # Admin 사용자는 모든 부서 검색 가능
    if getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin'):
        departments = Department.objects.filter(name__icontains=query)
    else:
        # 일반 사용자: 같은 회사 사용자들이 생성한 업체의 부서만 검색
        user_company = getattr(request, 'user_company', None)
        user_profile = getattr(request.user, 'userprofile', None)
        
        if user_company:
            same_company_users = User.objects.filter(userprofile__company=user_company)
            # 같은 회사 사용자들이 생성한 업체의 부서만 필터링
            departments = Department.objects.filter(
                name__icontains=query,
                company__created_by__in=same_company_users
            )
        elif user_profile and user_profile.company:
            same_company_users = User.objects.filter(userprofile__company=user_profile.company)
            departments = Department.objects.filter(
                name__icontains=query,
                company__created_by__in=same_company_users
            )
        else:
            departments = Department.objects.none()
    
    # 회사가 선택된 경우 해당 회사의 부서만 필터링
    if company_id:
        departments = departments.filter(company_id=company_id)
    
    departments = departments.select_related('company').order_by('company__name', 'name')[:10]
    
    results = []
    for dept in departments:
        results.append({
            'id': dept.id,
            'text': f"{dept.company.name} - {dept.name}",
            'company_id': dept.company.id,
            'company_name': dept.company.name,
            'department_name': dept.name
        })
    
    return JsonResponse({'results': results})

@login_required
def followup_autocomplete(request):
    """팔로우업 자동완성 API (일정 생성용 + 고객별 기록 조회용)"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 1:
        return JsonResponse({'results': []})
    
    # 같은 회사 사용자들의 고객 모두 검색 가능
    same_company_users = get_same_company_users(request.user)
    
    # 검색어로 필터링 (고객명, 업체명, 부서명, 책임자명으로 검색)
    followups = FollowUp.objects.filter(user__in=same_company_users).filter(
        Q(customer_name__icontains=query) |
        Q(company__name__icontains=query) |
        Q(department__name__icontains=query) |
        Q(manager__icontains=query)
    ).select_related('company', 'department', 'user').order_by('company__name', 'customer_name')[:15]
    
    results = []
    for followup in followups:
        # 표시 텍스트 구성
        company_name = str(followup.company) if followup.company else '업체명 미정'
        department_name = str(followup.department) if followup.department else '부서명 미정'
        customer_name = followup.customer_name or '고객명 미정'
        manager_name = followup.user.get_full_name() or followup.user.username
        
        display_text = f"{company_name} - {department_name} | {customer_name}"
        
        # 동료 고객인 경우 담당자 표시
        if followup.user != request.user:
            display_text += f" ({manager_name})"
        
        results.append({
            'id': followup.id,
            'text': display_text,
            'customer_name': customer_name,
            'company': company_name,
            'email': followup.email or '',
            'manager': manager_name
        })
    
    return JsonResponse({'results': results})

@login_required
def schedule_activity_type(request):
    """일정의 activity_type을 반환하는 API"""
    schedule_id = request.GET.get('schedule_id')
    if schedule_id:
        try:
            schedule = Schedule.objects.get(id=schedule_id, user=request.user)
            # 일정의 activity_type을 히스토리의 action_type으로 매핑
            activity_mapping = {
                'customer_meeting': 'customer_meeting',
                'delivery': 'delivery_schedule',
                'service': 'service',
            }
            mapped_action = activity_mapping.get(schedule.activity_type, 'customer_meeting')
            return JsonResponse({
                'success': True,
                'activity_type': schedule.activity_type,
                'mapped_action_type': mapped_action
            })
        except Schedule.DoesNotExist:
            return JsonResponse({'success': False, 'error': '일정을 찾을 수 없습니다.'})
    
    return JsonResponse({'success': False, 'error': '일정 ID가 필요합니다.'})

@login_required
@require_POST
def company_create_api(request):
    """새 업체/학교 생성 API"""
    try:
        name = request.POST.get('name', '').strip()
        
        if not name:
            return JsonResponse({'error': '업체/학교명을 입력해주세요.'}, status=400)
        
        # 중복 체크 - 같은 회사 내에서만
        user_profile_obj = getattr(request.user, 'userprofile', None)
        if user_profile_obj and user_profile_obj.company:
            same_company_users = User.objects.filter(userprofile__company=user_profile_obj.company)
            if Company.objects.filter(name=name, created_by__in=same_company_users).exists():
                return JsonResponse({'error': '이미 존재하는 업체/학교명입니다.'}, status=400)
        
        company = Company.objects.create(name=name, created_by=request.user)
        
        return JsonResponse({
            'success': True,
            'company': {
                'id': company.id,
                'name': company.name
            },
            'message': f'"{name}" 업체/학교가 추가되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"업체 생성 오류: {str(e)} (user={request.user.username})", exc_info=True)
        return JsonResponse({'error': f'업체/학교 생성 중 오류가 발생했습니다: {str(e)}'}, status=500)

@login_required
@require_POST
def department_create_api(request):
    """새 부서/연구실 생성 API"""
    name = request.POST.get('name', '').strip()
    company_id = request.POST.get('company_id')
    
    if not name:
        return JsonResponse({'error': '부서/연구실명을 입력해주세요.'}, status=400)
    
    if not company_id:
        return JsonResponse({'error': '업체/학교를 먼저 선택해주세요.'}, status=400)
    
    try:
        company = Company.objects.get(id=company_id)
        
        # 중복 체크
        if Department.objects.filter(company=company, name=name).exists():
            return JsonResponse({'error': f'{company.name}에 이미 존재하는 부서/연구실명입니다.'}, status=400)
        
        department = Department.objects.create(company=company, name=name, created_by=request.user)
        return JsonResponse({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'company_id': company.id,
                'company_name': company.name
            },
            'message': f'"{company.name} - {name}" 부서/연구실이 추가되었습니다.'
        })
        
    except Company.DoesNotExist:
        return JsonResponse({'error': '존재하지 않는 업체/학교입니다.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'부서/연구실 생성 중 오류가 발생했습니다: {str(e)}'}, status=500)


# ============ 업체/부서 관리 뷰들 ============

@role_required(['admin', 'salesman'])
def company_list_view(request):
    """업체/학교 목록 (Admin, Salesman 전용)"""
    
    # Admin 사용자는 필터링된 업체를 볼 수 있음
    if getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin'):
        # Admin은 필터링된 사용자의 업체 조회
        accessible_users = get_accessible_users(request.user, request)
        if accessible_users.count() == User.objects.count():
            # 전체 사용자면 모든 업체
            companies = Company.objects.all().annotate(
                department_count=Count('departments', distinct=True),
                followup_count=Count('followup_companies', distinct=True)
            ).order_by('name')
        else:
            # 필터링된 사용자의 업체만
            companies = Company.objects.filter(created_by__in=accessible_users).annotate(
                department_count=Count('departments', distinct=True),
                followup_count=Count('followup_companies', distinct=True)
            ).order_by('name')
    else:
        # 일반 사용자: 같은 회사 소속 사용자들이 생성한 업체만 조회
        user_company = getattr(request.user, 'userprofile', None)
        if user_company and user_company.company:
            # 같은 회사 소속 사용자들이 생성한 업체만 조회
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            companies = Company.objects.filter(created_by__in=same_company_users).annotate(
                department_count=Count('departments', distinct=True),
                followup_count=Count('followup_companies', distinct=True)
            ).order_by('name')
        else:
            # 회사 정보가 없는 경우 빈 쿼리셋
            companies = Company.objects.none()
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    if search_query:
        companies = companies.filter(name__icontains=search_query)
    
    # 페이지네이션
    paginator = Paginator(companies, 10)
    page_number = request.GET.get('page')
    companies = paginator.get_page(page_number)
    
    context = {
        'companies': companies,
        'search_query': search_query,
        'page_title': '업체/학교 관리'
    }
    return render(request, 'reporting/company_list.html', context)

@role_required(['admin', 'salesman'])
def company_create_view(request):
    """업체/학교 생성 (Admin, Salesman 전용)"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, '업체/학교명을 입력해주세요.')
        else:
            # Admin은 전체 업체를 기준으로, 일반 사용자는 같은 회사 사용자들이 만든 업체를 기준으로 중복 확인
            is_admin = getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin')
            
            if is_admin:
                # Admin은 전체 업체 중 중복 확인
                existing_company = Company.objects.filter(name=name).exists()
            else:
                # 같은 회사 사용자들이 만든 업체 중에서 중복 확인
                user_company = getattr(request.user, 'userprofile', None)
                if user_company and user_company.company:
                    same_company_users = User.objects.filter(userprofile__company=user_company.company)
                    existing_company = Company.objects.filter(name=name, created_by__in=same_company_users).exists()
                else:
                    existing_company = Company.objects.filter(name=name, created_by=request.user).exists()
                
            if existing_company:
                messages.error(request, '이미 존재하는 업체/학교명입니다.')
            else:
                Company.objects.create(name=name, created_by=request.user)
                messages.success(request, f'"{name}" 업체/학교가 추가되었습니다.')
                
                return redirect('reporting:company_list')
    
    context = {
        'page_title': '새 업체/학교 추가'
    }
    return render(request, 'reporting/company_form.html', context)

@role_required(['admin', 'salesman'])
def company_edit_view(request, pk):
    """업체/학교 수정 (Admin, 생성자 전용)"""
    company = get_object_or_404(Company, pk=pk)
    
    # 권한 체크: 관리자이거나 생성자만 수정 가능
    user_profile = get_user_profile(request.user)
    is_admin = getattr(request, 'is_admin', False) or user_profile.role == 'admin'
    
    if not (is_admin or company.created_by == request.user):
        # Admin이 아닌 경우 같은 회사 사용자가 생성한 업체인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        if user_company and user_company.company:
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if not Company.objects.filter(pk=pk, created_by__in=same_company_users).exists():
                messages.error(request, '이 업체/학교를 수정할 권한이 없습니다.')
                return redirect('reporting:company_list')
        else:
            messages.error(request, '이 업체/학교를 수정할 권한이 없습니다. (생성자 또는 관리자만 가능)')
            return redirect('reporting:company_list')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, '업체/학교명을 입력해주세요.')
        elif Company.objects.filter(name=name).exclude(pk=company.pk).exists():
            messages.error(request, '이미 존재하는 업체/학교명입니다.')
        else:
            company.name = name
            company.save()
            messages.success(request, f'"{name}" 업체/학교 정보가 수정되었습니다.')
            
            return redirect('reporting:company_list')
    
    context = {
        'company': company,
        'page_title': f'업체/학교 수정 - {company.name}'
    }
    return render(request, 'reporting/company_form.html', context)

@role_required(['admin', 'salesman'])
def company_delete_view(request, pk):
    """업체/학교 삭제 (Admin, 생성자 전용)"""
    company = get_object_or_404(Company, pk=pk)
    
    # 권한 체크: 관리자이거나 생성자만 삭제 가능
    user_profile = get_user_profile(request.user)
    is_admin = getattr(request, 'is_admin', False) or user_profile.role == 'admin'
    
    if not (is_admin or company.created_by == request.user):
        # Admin이 아닌 경우 같은 회사 사용자가 생성한 업체인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        if user_company and user_company.company:
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if not Company.objects.filter(pk=pk, created_by__in=same_company_users).exists():
                messages.error(request, '이 업체/학교를 삭제할 권한이 없습니다.')
                return redirect('reporting:company_list')
        else:
            messages.error(request, '이 업체/학교를 삭제할 권한이 없습니다. (생성자 또는 관리자만 가능)')
            return redirect('reporting:company_list')
    
    # 관련 데이터 개수 확인
    department_count = company.departments.count()
    followup_count = company.followup_companies.count()
    
    if request.method == 'POST':
        company_name = company.name
        
        if followup_count > 0:
            messages.error(request, f'이 업체/학교를 사용하는 고객 정보가 {followup_count}개 있어 삭제할 수 없습니다.')
            return redirect('reporting:company_list')
        
        company.delete()
        messages.success(request, f'"{company_name}" 업체/학교가 삭제되었습니다.')
        
        return redirect('reporting:company_list')
    
    context = {
        'company': company,
        'department_count': department_count,
        'followup_count': followup_count,
        'page_title': f'업체/학교 삭제 - {company.name}'
    }
    return render(request, 'reporting/company_delete.html', context)

@role_required(['admin', 'salesman'])
def company_detail_view(request, pk):
    """업체/학교 상세 (부서 목록 포함) (Admin, Salesman 전용)"""
    company = get_object_or_404(Company, pk=pk)
    
    # Admin이 아닌 경우 권한 확인
    if not (getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin')):
        # 자신의 회사 소속 사용자들이 생성한 업체인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        if user_company and user_company.company:
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if not Company.objects.filter(pk=pk, created_by__in=same_company_users).exists():
                messages.error(request, '해당 업체/학교에 접근할 권한이 없습니다.')
                return redirect('reporting:company_list')
        else:
            messages.error(request, '회사 정보가 없어 접근할 수 없습니다.')
            return redirect('reporting:company_list')
    
    # 수정/삭제 권한 확인
    is_admin = getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin')
    can_edit_company = is_admin or company.created_by == request.user
    
    # 해당 업체의 부서 목록
    departments = company.departments.annotate(
        followup_count=Count('followup_departments')
    ).order_by('name')
    
    # 검색 기능 (부서명)
    dept_search = request.GET.get('dept_search', '')
    if dept_search:
        departments = departments.filter(name__icontains=dept_search)
    
    context = {
        'company': company,
        'departments': departments,
        'dept_search': dept_search,
        'can_edit_company': can_edit_company,  # 수정/삭제 권한 정보
        'page_title': f'{company.name} - 부서/연구실 관리'
    }
    return render(request, 'reporting/company_detail.html', context)

@role_required(['admin', 'salesman'])
def department_create_view(request, company_pk):
    """부서/연구실 생성 (Admin, Salesman 전용)"""
    company = get_object_or_404(Company, pk=company_pk)
    
    import logging
    logger = logging.getLogger(__name__)
    
    # Admin이 아닌 경우 권한 확인
    is_admin = getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin')
    
    if not is_admin:
        # 자신의 회사 소속 사용자들이 생성한 업체인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        if user_company and user_company.company:
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if not Company.objects.filter(pk=company_pk, created_by__in=same_company_users).exists():
                messages.error(request, '해당 업체/학교에 부서를 추가할 권한이 없습니다.')
                return redirect('reporting:company_detail', pk=company_pk)
        else:
            messages.error(request, '회사 정보가 없어 부서를 추가할 수 없습니다.')
            return redirect('reporting:company_detail', pk=company_pk)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, '부서/연구실명을 입력해주세요.')
        elif Department.objects.filter(company=company, name=name).exists():
            messages.error(request, f'{company.name}에 이미 존재하는 부서/연구실명입니다.')
        else:
            Department.objects.create(company=company, name=name, created_by=request.user)
            messages.success(request, f'"{company.name} - {name}" 부서/연구실이 추가되었습니다.')
            
            return redirect('reporting:company_detail', pk=company.pk)
    
    context = {
        'company': company,
        'page_title': f'{company.name} - 새 부서/연구실 추가'
    }
    return render(request, 'reporting/department_form.html', context)

@role_required(['admin', 'salesman'])
def department_edit_view(request, pk):
    """부서/연구실 수정 (Admin, 생성자 전용)"""
    department = get_object_or_404(Department, pk=pk)
    
    # 권한 체크: 관리자이거나 같은 회사 사용자만 수정 가능
    user_profile = get_user_profile(request.user)
    is_admin = getattr(request, 'is_admin', False) or user_profile.role == 'admin'
    
    if not is_admin:
        # Admin이 아닌 경우 같은 회사 사용자인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        has_edit_permission = False
        
        if user_company and user_company.company:
            # 같은 회사 사용자들이 생성한 업체의 부서인지 확인
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if Company.objects.filter(pk=department.company.pk, created_by__in=same_company_users).exists():
                has_edit_permission = True
        
        if not has_edit_permission:
            messages.error(request, '이 부서/연구실을 수정할 권한이 없습니다.')
            return redirect('reporting:company_detail', pk=department.company.pk)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, '부서/연구실명을 입력해주세요.')
        elif Department.objects.filter(company=department.company, name=name).exclude(pk=department.pk).exists():
            messages.error(request, f'{department.company.name}에 이미 존재하는 부서/연구실명입니다.')
        else:
            department.name = name
            department.save()
            messages.success(request, f'"{department.company.name} - {name}" 부서/연구실 정보가 수정되었습니다.')
            return redirect('reporting:company_detail', pk=department.company.pk)
    
    context = {
        'department': department,
        'page_title': f'{department.company.name} - 부서/연구실 수정',
        'customers': department.followup_departments.all().select_related('user')  # 소속 연구원/고객 목록
    }
    return render(request, 'reporting/department_form.html', context)

@role_required(['admin', 'salesman'])
def department_delete_view(request, pk):
    """부서/연구실 삭제 (Admin, 생성자 전용)"""
    department = get_object_or_404(Department, pk=pk)
    
    # 권한 체크: 관리자이거나 같은 회사 사용자만 삭제 가능
    user_profile = get_user_profile(request.user)
    is_admin = getattr(request, 'is_admin', False) or user_profile.role == 'admin'
    
    import logging
    logger = logging.getLogger(__name__)
    
    if not is_admin:
        # Admin이 아닌 경우 같은 회사 사용자인지 확인
        user_company = getattr(request.user, 'userprofile', None)
        has_delete_permission = False
        
        if user_company and user_company.company:
            # 같은 회사 사용자들이 생성한 업체의 부서인지 확인
            same_company_users = User.objects.filter(userprofile__company=user_company.company)
            if Company.objects.filter(pk=department.company.pk, created_by__in=same_company_users).exists():
                has_delete_permission = True
        
        if not has_delete_permission:
            messages.error(request, '이 부서/연구실을 삭제할 권한이 없습니다.')
            return redirect('reporting:company_detail', pk=department.company.pk)
    
    # 관련 데이터 개수 확인
    followup_count = department.followup_departments.count()
    
    if request.method == 'POST':
        department_name = department.name
        company_name = department.company.name
        company_pk = department.company.pk
        
        if followup_count > 0:
            messages.error(request, f'이 부서/연구실을 사용하는 고객 정보가 {followup_count}개 있어 삭제할 수 없습니다.')
            return redirect('reporting:company_detail', pk=company_pk)
        
        department.delete()
        messages.success(request, f'"{company_name} - {department_name}" 부서/연구실이 삭제되었습니다.')
        
        return redirect('reporting:company_detail', pk=company_pk)
    
    context = {
        'department': department,
        'followup_count': followup_count,
        'page_title': f'{department.company.name} - 부서/연구실 삭제'
    }
    return render(request, 'reporting/department_delete.html', context)

# ============ 매니저용 읽기 전용 업체/부서 뷰들 ============

@role_required(['manager'])
def manager_company_list_view(request):
    """매니저용 업체/학교 목록 (읽기 전용)"""
    
    # 현재 사용자 회사와 같은 회사의 사용자들이 생성한 업체만 조회
    user_profile_obj = getattr(request.user, 'userprofile', None)
    if user_profile_obj and user_profile_obj.company:
        # 같은 회사 소속 사용자들이 생성한 업체만 조회
        same_company_users = User.objects.filter(userprofile__company=user_profile_obj.company)
        companies = Company.objects.filter(created_by__in=same_company_users).annotate(
            department_count=Count('departments', distinct=True),
            followup_count=Count('followup_companies', distinct=True)
        ).order_by('name')
    else:
        # 회사 정보가 없는 경우 빈 쿼리셋
        companies = Company.objects.none()
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    if search_query:
        companies = companies.filter(name__icontains=search_query)
    
    # 각 업체별 담당자 정보 추가 (같은 회사 사용자들만)
    companies_with_salesmen = []
    if user_profile_obj and user_profile_obj.company:
        same_company_users_list = User.objects.filter(userprofile__company=user_profile_obj.company)
    else:
        same_company_users_list = []
        
    for company in companies:
        # 기본 담당자: 업체를 생성한 사람
        salesmen = []
        if company.created_by:
            salesmen.append(company.created_by)
        
        # 추가 담당자: 해당 업체의 FollowUp을 담당하는 실무자들 (같은 회사 사용자들만)
        followups_in_company = FollowUp.objects.filter(
            Q(company=company) | Q(department__company=company),
            user__in=same_company_users_list
        ).select_related('user').distinct()
        
        # FollowUp 담당자들을 추가 (중복 제거)
        for followup in followups_in_company:
            if followup.user not in salesmen:
                salesmen.append(followup.user)
        
        # 회사 객체에 담당자 정보 추가
        company.salesmen = salesmen
        companies_with_salesmen.append(company)
    
    # 페이지네이션
    paginator = Paginator(companies_with_salesmen, 10)
    page_number = request.GET.get('page')
    companies = paginator.get_page(page_number)
    
    context = {
        'companies': companies,
        'search_query': search_query,
        'page_title': '업체/학교 목록 (조회)',
        'is_readonly': True
    }
    return render(request, 'reporting/company_list.html', context)

@role_required(['manager'])
def manager_company_detail_view(request, pk):
    """매니저용 업체/학교 상세 (읽기 전용)"""
    
    # 현재 사용자 회사와 같은 회사에서 생성된 업체인지 확인
    user_profile_obj = getattr(request.user, 'userprofile', None)
    if user_profile_obj and user_profile_obj.company:
        same_company_users = User.objects.filter(userprofile__company=user_profile_obj.company)
        company = get_object_or_404(Company, pk=pk, created_by__in=same_company_users)
    else:
        # 회사 정보가 없는 경우 접근 거부
        messages.error(request, '접근 권한이 없습니다.')
        return redirect('reporting:manager_company_list')
    
    # 해당 업체의 부서 목록
    departments = company.departments.annotate(
        followup_count=Count('followup_departments')
    ).order_by('name')
    
    # 검색 기능 (부서명)
    dept_search = request.GET.get('dept_search', '')
    if dept_search:
        departments = departments.filter(name__icontains=dept_search)
    
    # 해당 업체를 관리하는 실무자들 조회 (같은 회사 사용자들만)
    followups_in_company = FollowUp.objects.filter(
        Q(company=company) | Q(department__company=company),
        user__in=same_company_users
    ).select_related('user', 'user__userprofile')
    
    # 실무자별 담당 고객 수 집계
    salesmen_stats = {}
    for followup in followups_in_company:
        user = followup.user
        if user not in salesmen_stats:
            salesmen_stats[user] = {
                'user': user,
                'followup_count': 0,
                'recent_activity': None
            }
        salesmen_stats[user]['followup_count'] += 1
        
        # 가장 최근 활동 히스토리 찾기 (메모 제외)
        recent_history = History.objects.filter(
            followup=followup
        ).exclude(action_type='memo').order_by('-created_at').first()
        
        if recent_history and (
            not salesmen_stats[user]['recent_activity'] or
            recent_history.created_at > salesmen_stats[user]['recent_activity'].created_at
        ):
            salesmen_stats[user]['recent_activity'] = recent_history
    
    # 실무자 목록 (담당 고객 수 순으로 정렬)
    salesmen_list = sorted(
        salesmen_stats.values(), 
        key=lambda x: x['followup_count'], 
        reverse=True
    )
    
    # 페이지네이션 (부서)
    paginator = Paginator(departments, 10)
    page_number = request.GET.get('page')
    departments = paginator.get_page(page_number)
    
    context = {
        'company': company,
        'departments': departments,
        'dept_search': dept_search,
        'salesmen_list': salesmen_list,
        'page_title': f'{company.name} - 상세 정보 (조회)',
        'is_readonly': True
    }
    return render(request, 'reporting/company_detail.html', context)

# ============ 추가 API 엔드포인트들 ============

@login_required
def api_company_detail(request, pk):
    """개별 회사 정보 조회 API"""
    try:
        company = get_object_or_404(Company, pk=pk)
        return JsonResponse({
            'success': True,
            'company': {
                'id': company.id,
                'name': company.name
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required  
def api_department_detail(request, pk):
    """개별 부서 정보 조회 API"""
    try:
        department = get_object_or_404(Department, pk=pk)
        return JsonResponse({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'company_id': department.company.id,
                'company_name': department.company.name
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def history_detail_api(request, history_id):
    """히스토리 상세 정보를 JSON으로 반환하는 API"""
    try:
        history = get_object_or_404(History, id=history_id)
        
        # 조회 권한 확인 (같은 회사면 조회 가능)
        if history.followup and not can_access_followup(request.user, history.followup):
            return JsonResponse({
                'success': False,
                'error': '이 기록에 접근할 권한이 없습니다.'
            }, status=403)
        
        # 수정 권한 여부도 함께 전달
        can_modify = can_modify_user_data(request.user, history.user)
        
        # 히스토리 데이터 직렬화
        history_data = {
            'id': history.id,
            'content': history.content or '',
            'action_type': history.action_type,
            'action_type_display': history.get_action_type_display(),
            'created_at': history.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'user': history.user.get_full_name() or history.user.username,
            'created_by': history.created_by.username if history.created_by else '',
            'can_modify': can_modify,  # 수정 권한 정보 추가
            'followup_id': history.followup.id if history.followup else None,
            'schedule_id': history.schedule.id if history.schedule else None,
            
            # 납품 일정 관련 필드
            'delivery_date': history.delivery_date.strftime('%Y-%m-%d') if history.delivery_date else '',
            'delivery_amount': history.delivery_amount,
            'delivery_items': history.delivery_items or '',
            'tax_invoice_issued': history.tax_invoice_issued,
            
            # 고객 미팅 관련 필드
            'meeting_date': history.meeting_date.strftime('%Y-%m-%d') if history.meeting_date else '',
            'meeting_situation': history.meeting_situation or '',
            'meeting_researcher_quote': history.meeting_researcher_quote or '',
            'meeting_confirmed_facts': history.meeting_confirmed_facts or '',
            'meeting_obstacles': history.meeting_obstacles or '',
            'meeting_next_action': history.meeting_next_action or '',
            
            # 서비스 관련 필드
            'service_status': history.service_status or '',
            'service_status_display': history.get_service_status_display() if history.service_status else '',
            
            # 첨부파일 정보
            'files': []
        }
        
        # 첨부파일 정보 추가
        for file in history.files.all():
            # 파일 크기를 읽기 쉬운 형태로 변환
            file_size = file.file_size
            if file_size < 1024:
                size_str = f"{file_size} bytes"
            elif file_size < 1024 * 1024:
                size_str = f"{file_size/1024:.1f} KB"
            else:
                size_str = f"{file_size/(1024*1024):.1f} MB"
            
            history_data['files'].append({
                'id': file.id,
                'filename': file.original_filename,
                'size': size_str,
                'uploaded_at': file.uploaded_at.strftime('%Y-%m-%d %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'history': history_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'히스토리 정보를 불러올 수 없습니다: {str(e)}'
        })

@login_required
@require_POST
def history_update_api(request, history_id):
    """히스토리 업데이트 API"""
    try:
        history = get_object_or_404(History, id=history_id)
        
        # 수정 권한 확인 (본인 데이터만 수정 가능, Manager는 읽기 전용)
        if not can_modify_user_data(request.user, history.user):
            return JsonResponse({
                'success': False,
                'error': '이 기록을 수정할 권한이 없습니다. 본인의 기록만 수정할 수 있습니다.'
            }, status=403)
        
        # 폼 데이터 처리
        content = request.POST.get('content', '').strip()
        
        # 고객 미팅이 아닌 경우 content 필수
        if history.action_type != 'customer_meeting' and not content:
            return JsonResponse({
                'success': False,
                'error': '활동 내용을 입력해주세요.'
            })
        
        # 기본 필드 업데이트
        history.content = content
        
        # 활동 유형별 추가 필드 처리
        if history.action_type == 'delivery_schedule':
            delivery_date = request.POST.get('delivery_date')
            if delivery_date:
                try:
                    from datetime import datetime
                    history.delivery_date = datetime.strptime(delivery_date, '%Y-%m-%d').date()
                except ValueError:
                    history.delivery_date = None
            else:
                history.delivery_date = None
            
            delivery_amount = request.POST.get('delivery_amount')
            if delivery_amount:
                try:
                    history.delivery_amount = int(delivery_amount)
                except (ValueError, TypeError):
                    history.delivery_amount = None
            else:
                history.delivery_amount = None
            
            history.delivery_items = request.POST.get('delivery_items', '').strip()
            history.tax_invoice_issued = request.POST.get('tax_invoice_issued') == 'on'
            
        elif history.action_type == 'customer_meeting':
            meeting_date = request.POST.get('meeting_date')
            if meeting_date:
                try:
                    from datetime import datetime
                    history.meeting_date = datetime.strptime(meeting_date, '%Y-%m-%d').date()
                except ValueError:
                    history.meeting_date = None
            else:
                history.meeting_date = None
            
            # 구조화된 미팅 노트 필드 저장
            history.meeting_situation = request.POST.get('meeting_situation', '').strip()
            history.meeting_researcher_quote = request.POST.get('meeting_researcher_quote', '').strip()
            history.meeting_confirmed_facts = request.POST.get('meeting_confirmed_facts', '').strip()
            history.meeting_obstacles = request.POST.get('meeting_obstacles', '').strip()
            history.meeting_next_action = request.POST.get('meeting_next_action', '').strip()
                
        elif history.action_type == 'service':
            history.service_status = request.POST.get('service_status', '')
        
        # 변경사항 저장
        history.save()
        
        # 기존 파일 삭제 처리
        delete_file_ids = request.POST.getlist('delete_files')
        if delete_file_ids:
            try:
                # 삭제할 파일들 조회 및 삭제
                files_to_delete = HistoryFile.objects.filter(
                    id__in=delete_file_ids,
                    history=history
                )
                for file_obj in files_to_delete:
                    # 실제 파일 삭제
                    if file_obj.file and os.path.exists(file_obj.file.path):
                        os.remove(file_obj.file.path)
                    # DB에서 삭제
                    file_obj.delete()
                    
            except Exception as delete_error:
                return JsonResponse({
                    'success': False,
                    'error': f'파일 삭제 중 오류가 발생했습니다: {str(delete_error)}'
                })
        
        # 새로운 파일 첨부 처리
        uploaded_files = request.FILES.getlist('files')
        if uploaded_files:
            try:
                # 현재 파일 개수 확인 (삭제 후)
                current_file_count = history.files.count()
                total_files_after_upload = current_file_count + len(uploaded_files)
                
                # 파일 개수 제한 (최대 5개)
                if total_files_after_upload > 5:
                    return JsonResponse({
                        'success': False,
                        'error': f'파일은 최대 5개까지만 첨부할 수 있습니다. (현재 {current_file_count}개, 추가 {len(uploaded_files)}개)'
                    })
                
                # 각 파일 유효성 검사 및 저장
                for file in uploaded_files:
                    # 파일 크기 검사 (10MB)
                    if file.size > 10 * 1024 * 1024:
                        return JsonResponse({
                            'success': False,
                            'error': f'파일 {file.name}이 10MB를 초과합니다.'
                        })
                    
                    # 파일 확장자 검사
                    allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.jpg', '.jpeg', '.png', '.gif', '.zip', '.rar']
                    file_extension = os.path.splitext(file.name)[1].lower()
                    if file_extension not in allowed_extensions:
                        return JsonResponse({
                            'success': False,
                            'error': f'파일 {file.name}은 지원되지 않는 형식입니다.'
                        })
                    
                    # 파일 저장
                    history_file = HistoryFile.objects.create(
                        history=history,
                        file=file,
                        original_filename=file.name,
                        file_size=file.size,
                        uploaded_by=request.user
                    )
                    
            except Exception as file_error:
                return JsonResponse({
                    'success': False,
                    'error': f'파일 업로드 중 오류가 발생했습니다: {str(file_error)}'
                })
        
        return JsonResponse({
            'success': True,
            'message': '활동 기록이 성공적으로 수정되었습니다.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'활동 기록 수정 중 오류가 발생했습니다: {str(e)}'
        })

# ============ 메모 관련 뷰들 ============

@login_required
def memo_create_view(request):
    """메모 생성 (팔로우업 연결 선택사항)"""
    followup_id = request.GET.get('followup') or request.POST.get('followup')
    followup = None
    
    if followup_id:
        try:
            followup = get_object_or_404(FollowUp, pk=followup_id)
            # 권한 체크 (팔로우업이 있는 경우만)
            if not can_modify_user_data(request.user, followup.user):
                if request.method == 'POST':
                    return JsonResponse({
                        'success': False,
                        'error': '메모 작성 권한이 없습니다.'
                    }, status=403)
                messages.error(request, '메모 작성 권한이 없습니다.')
                return redirect('reporting:followup_detail', pk=followup.pk)
        except FollowUp.DoesNotExist:
            followup = None
    
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        
        if not content:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/x-www-form-urlencoded':
                return JsonResponse({
                    'success': False,
                    'error': '메모 내용을 입력해주세요.'
                }, status=400)
            messages.error(request, '메모 내용을 입력해주세요.')
        else:
            # 메모 히스토리 생성
            history = History.objects.create(
                user=request.user,
                company=request.user.userprofile.company,
                followup=followup,  # followup이 None일 수도 있음
                action_type='memo',
                content=content,
                schedule=None  # 메모는 일정과 연결되지 않음
            )
            
            # AJAX 요청인 경우 JSON 응답
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/x-www-form-urlencoded':
                return JsonResponse({
                    'success': True,
                    'message': '메모가 성공적으로 추가되었습니다.',
                    'history_id': history.id
                })
            
            messages.success(request, '메모가 성공적으로 추가되었습니다.')
            
            # 팔로우업이 있으면 팔로우업 상세로, 없으면 히스토리 목록으로
            if followup:
                return redirect('reporting:followup_detail', pk=followup.pk)
            else:
                return redirect('reporting:history_list')
    
    context = {
        'followup': followup,
        'page_title': f'메모 추가{" - " + followup.customer_name if followup else ""}',
    }
    return render(request, 'reporting/memo_form.html', context)


# ============ 인라인 메모 편집 API ============

@login_required
@require_POST
def history_update_tax_invoice(request, pk):
    """세금계산서 발행 상태 업데이트"""
    try:
        history = get_object_or_404(History, pk=pk)
        
        # 권한 체크
        if not can_modify_user_data(request.user, history.user):
            return JsonResponse({
                'success': False,
                'error': '수정 권한이 없습니다.'
            })
        
        # 납품 일정 타입만 세금계산서 상태 변경 가능
        if history.action_type != 'delivery_schedule':
            return JsonResponse({
                'success': False,
                'error': '납품 일정만 세금계산서 상태를 변경할 수 있습니다.'
            })
        
        # 세금계산서 상태 업데이트
        tax_invoice_issued = request.POST.get('tax_invoice_issued') == 'true'
        history.tax_invoice_issued = tax_invoice_issued
        history.save()
        
        # 연결된 스케줄이 있는 경우 스케줄의 납품 품목과 동기화
        if history.schedule:
            try:
                # 연결된 스케줄의 모든 납품 품목 업데이트
                delivery_items = history.schedule.delivery_items_set.all()
                if delivery_items.exists():
                    delivery_items.update(tax_invoice_issued=tax_invoice_issued)
                
                # 연결된 스케줄에 속한 다른 히스토리들도 함께 업데이트
                related_histories = History.objects.filter(
                    schedule=history.schedule, 
                    action_type='delivery_schedule'
                ).exclude(id=history.id)
                
                if related_histories.exists():
                    related_histories.update(tax_invoice_issued=tax_invoice_issued)
                    
            except Exception as sync_error:
                logger.error(f"연결된 스케줄/히스토리 세금계산서 동기화 실패: {sync_error}")
        
        # sync_schedule 파라미터는 향후 확장용으로 유지
        sync_schedule = request.POST.get('sync_schedule') == 'true'
        
        # silent 모드인 경우 (자동 동기화) 별도 메시지
        is_silent = request.POST.get('silent') == 'true'
        if is_silent:
            message = '세금계산서 상태가 자동으로 동기화되었습니다.'
        else:
            message = '세금계산서 상태가 성공적으로 변경되었습니다.'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'세금계산서 상태 변경 중 오류가 발생했습니다: {str(e)}'
        })

@login_required
@require_POST
def history_update_memo(request, pk):
    """AJAX 요청으로 메모 내용 업데이트"""
    import json
    
    try:
        history = get_object_or_404(History, pk=pk, action_type='memo')
        
        # 권한 체크
        if not can_modify_user_data(request.user, history.user):
            return JsonResponse({
                'success': False,
                'error': '메모 수정 권한이 없습니다.'
            })
        
        # 요청 데이터 파싱
        data = json.loads(request.body)
        new_content = data.get('content', '').strip()
        
        if not new_content:
            return JsonResponse({
                'success': False,
                'error': '메모 내용을 입력해주세요.'
            })
        
        # 메모 내용 업데이트
        history.content = new_content
        history.save()
        
        return JsonResponse({
            'success': True,
            'message': '메모가 성공적으로 수정되었습니다.'
        })
        
    except History.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '메모를 찾을 수 없습니다.'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': '잘못된 요청 형식입니다.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'메모 수정 중 오류가 발생했습니다: {str(e)}'
        })


@login_required
def followup_excel_download(request):
    """팔로우업 전체 정보 엑셀 다운로드 (부서별 그룹화)"""
    user_profile = get_user_profile(request.user)
    
    # 엑셀 다운로드 권한 체크
    if not user_profile.can_excel_download():
        messages.error(request, '엑셀 다운로드 권한이 없습니다. 관리자에게 문의해주세요.')
        return redirect('reporting:followup_list')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from decimal import Decimal
    from collections import defaultdict
    import io
    from datetime import datetime
    
    user_profile = get_user_profile(request.user)
    
    # 권한에 따른 데이터 필터링 (기존 로직과 동일)
    if user_profile.can_view_all_users():
        accessible_users = get_accessible_users(request.user, request)
        followups = FollowUp.objects.filter(user__in=accessible_users).select_related(
            'user', 'company', 'department'
        ).prefetch_related('schedules', 'histories')
    else:
        followups = FollowUp.objects.filter(user=request.user).select_related(
            'user', 'company', 'department'
        ).prefetch_related('schedules', 'histories')
    
    # 검색 필터 적용
    search_query = request.GET.get('search')
    if search_query:
        followups = followups.filter(
            Q(customer_name__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # 우선순위 필터 적용
    priority_filter = request.GET.get('priority')
    if priority_filter:
        followups = followups.filter(priority=priority_filter)
    
    # 우선순위 정렬을 위한 순서 정의 (낮을수록 높은 우선순위)
    PRIORITY_ORDER = {
        'urgent': 1,      # 긴급 - 가장 높음
        'followup': 2,    # 팔로업
        'scheduled': 3,   # 예정
        'long_term': 4,   # 장기 - 가장 낮음
    }
    
    # 부서별로 그룹화
    departments_data = defaultdict(lambda: {
        'company_name': '',
        'department_name': '',
        'followups': [],
        'highest_priority': 99  # 부서 내 가장 높은 우선순위 (낮은 숫자가 높은 우선순위)
    })
    
    for followup in followups.order_by('company__name', 'department__name', 'customer_name'):
        company_name = followup.company.name if followup.company else '업체 미지정'
        department_name = followup.department.name if followup.department else '부서 미지정'
        dept_key = f"{company_name}||{department_name}"
        
        departments_data[dept_key]['company_name'] = company_name
        departments_data[dept_key]['department_name'] = department_name
        departments_data[dept_key]['followups'].append(followup)
        
        # 해당 부서 내 가장 높은 우선순위 업데이트
        priority_value = PRIORITY_ORDER.get(followup.priority, 99)
        if priority_value < departments_data[dept_key]['highest_priority']:
            departments_data[dept_key]['highest_priority'] = priority_value
    
    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "팔로우업 전체 정보"
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
    dept_header_font = Font(bold=True, color="FFFFFF", size=12)
    dept_header_fill = PatternFill(start_color="4A7C4E", end_color="4A7C4E", fill_type="solid")  # 녹색 계열
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 최대 히스토리 개수 계산
    max_histories = 0
    for followup in followups:
        history_count = followup.histories.count()
        max_histories = max(max_histories, history_count)
    
    # 헤더 정의
    headers = [
        '고객명', '책임자', '핸드폰 번호', 
        '메일 주소', '상세 주소', '고객 등급', '납품 품목', '총 납품 금액', '상세 내용'
    ]
    
    # 히스토리 컬럼 추가
    for i in range(1, max_histories + 1):
        headers.append(f'관련 활동 히스토리 {i}')
    
    current_row = 1
    
    # 부서별로 데이터 작성 (우선순위 높은 부서가 먼저, 그 다음 회사명/부서명 순)
    sorted_dept_keys = sorted(
        departments_data.keys(),
        key=lambda k: (departments_data[k]['highest_priority'], departments_data[k]['company_name'], departments_data[k]['department_name'])
    )
    
    for dept_key in sorted_dept_keys:
        dept_info = departments_data[dept_key]
        company_name = dept_info['company_name']
        department_name = dept_info['department_name']
        dept_followups = dept_info['followups']
        
        # 부서 구분 행 (회사명 - 부서명)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        dept_cell = ws.cell(row=current_row, column=1, value=f"📁 {company_name} - {department_name} ({len(dept_followups)}명)")
        dept_cell.font = dept_header_font
        dept_cell.fill = dept_header_fill
        dept_cell.alignment = center_alignment
        dept_cell.border = border
        current_row += 1
        
        # 헤더 행
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # 해당 부서의 고객 데이터 입력
        for followup in dept_followups:
            # 책임자 정보 가져오기 (FollowUp 모델의 manager 필드)
            manager_name = followup.manager or ''
            
            # 고객 등급 (우선순위)
            priority_display = followup.get_priority_display() or '보통'
            
            # 납품 관련 정보 집계
            delivery_histories = followup.histories.filter(action_type='delivery_schedule')
            
            # 납품 품목별 수량 집계용 딕셔너리
            item_quantities = {}
            total_delivery_amount = 0
            
            # 중복 방지를 위해 처리된 Schedule ID들을 추적
            processed_schedule_ids = set()
            
            for history in delivery_histories:
                # 납품 금액 집계 - History 우선
                if history.delivery_amount:
                    total_delivery_amount += history.delivery_amount
                
                # History에 실제 납품 품목 정보가 있는 경우만 Schedule ID 기록
                # (품목 정보가 없으면 Schedule에서 가져와야 함)
                if history.schedule_id and history.delivery_items:
                    processed_schedule_ids.add(history.schedule_id)
                
                # 납품 품목 집계 - History 텍스트에서만 처리 (Schedule DeliveryItem은 나중에 별도 처리)
                if history.delivery_items:
                    # 다양한 줄바꿈 문자 처리
                    processed_items = history.delivery_items
                    processed_items = processed_items.replace('\\n', '\n')
                    processed_items = processed_items.replace('\\r\\n', '\n')
                    processed_items = processed_items.replace('\\r', '\n')
                    processed_items = processed_items.replace('\r\n', '\n')
                    processed_items = processed_items.replace('\r', '\n')
                    processed_items = processed_items.strip()
                    
                    # 다양한 구분자로 분할 시도
                    lines = []
                    # 먼저 줄바꿈으로 분할
                    for line in processed_items.split('\n'):
                        line = line.strip()
                        if line:
                            # 쉼표로도 분할해보기
                            if ',' in line and ':' in line:
                                sub_lines = [sub.strip() for sub in line.split(',') if sub.strip()]
                                lines.extend(sub_lines)
                            else:
                                lines.append(line)
                
                    for line in lines:
                        # 다양한 패턴 시도
                        import re
                        
                        # 패턴 1: "품목명: 수량개 금액원 횟수회" 또는 "품목명 수량개 금액원 횟수회"
                        pattern1 = r'(.+?)[\s:]*([\d,]+)개[\s,]*([\d,]+)원[\s,]*([\d]+)회'
                        match1 = re.search(pattern1, line)
                        
                        if match1:
                            item_name = match1.group(1).replace(':', '').strip()
                            quantity = float(match1.group(2).replace(',', ''))
                            
                            if item_name in item_quantities:
                                item_quantities[item_name] += quantity
                            else:
                                item_quantities[item_name] = quantity
                            continue
                        
                        # 패턴 2: "품목명: 수량개" 또는 "품목명 수량개"
                        pattern2 = r'(.+?)[\s:]*([\d,]+(?:\.\d+)?)개'
                        match2 = re.search(pattern2, line)
                        
                        if match2:
                            item_name = match2.group(1).replace(':', '').strip()
                            quantity = float(match2.group(2).replace(',', ''))
                            
                            if item_name in item_quantities:
                                item_quantities[item_name] += quantity
                            else:
                                item_quantities[item_name] = quantity
                            continue
                        
                        # 패턴 3: 단순 품목명만 있는 경우
                        if line and not any(char in line for char in [':', '개', '원', '회']):
                            item_name = line.strip()
                            
                            if item_name in item_quantities:
                                item_quantities[item_name] += 1
                            else:
                                item_quantities[item_name] = 1
            
            # Schedule 기반 DeliveryItem도 포함 (모든 Schedule 처리)
            all_schedule_deliveries = followup.schedules.filter(
                delivery_items_set__isnull=False
            ).distinct()
            
            # 모든 Schedule 처리 (History에 품목 정보가 없으면 Schedule에서 가져옴)
            for schedule in all_schedule_deliveries:
                # History에서 이미 품목 정보를 처리한 Schedule은 금액만 확인
                if schedule.id in processed_schedule_ids:
                    # 금액만 추가 확인 (History에 없었을 수 있음)
                    schedule_total = 0
                    for item in schedule.delivery_items_set.all():
                        if item.total_price:
                            schedule_total += Decimal(str(item.total_price))
                    
                    if schedule_total > 0:
                        total_delivery_amount += schedule_total
                    continue
                
                # Schedule별 총액 계산 및 품목 집계
                schedule_total = 0
                schedule_items = []
                
                for item in schedule.delivery_items_set.all():
                    # Schedule 기반 품목의 금액 포함
                    if item.total_price:
                        schedule_total += Decimal(str(item.total_price))
                    
                    # 품목 정보 저장
                    schedule_items.append({
                        'name': item.item_name,
                        'quantity': float(item.quantity)
                    })
                
                # Schedule 총액을 전체 납품 금액에 추가 (이미 processed된 경우 위에서 처리됨)
                if schedule.id not in processed_schedule_ids and schedule_total > 0:
                    total_delivery_amount += schedule_total
                
                # Schedule 품목 집계 (모든 Schedule에서)
                for item_info in schedule_items:
                    item_name = item_info['name']
                    quantity = item_info['quantity']
                    
                    # 품목별 수량 누적 (원본 이름 그대로 사용)
                    if item_name in item_quantities:
                        item_quantities[item_name] += quantity
                    else:
                        item_quantities[item_name] = quantity

            
            # 품목 텍스트 생성 (품목명과 총 수량 표시)
            if item_quantities:
                items_list = []
                for item_name, total_qty in sorted(item_quantities.items()):
                    # 소수점이 있으면 그대로, 정수면 정수로 표시
                    if total_qty == int(total_qty):
                        qty_str = str(int(total_qty))
                    else:
                        qty_str = str(total_qty)
                    items_list.append(f"{item_name}: {qty_str}개")
                
                # 모든 품목 표시 (제한 제거)
                items_text = ', '.join(items_list)
            else:
                items_text = '납품 기록 없음'
            
            # 기본 정보 (부서별 그룹화이므로 업체/부서 컬럼 제외)
            data = [
                followup.customer_name or '',
                manager_name,  # FollowUp의 책임자 필드에서 가져오기
                followup.phone_number or '',
                followup.email or '',
                followup.address or '',
                priority_display,  # 고객 등급
                items_text,  # 납품 품목
                f"{total_delivery_amount:,}원" if total_delivery_amount > 0 else '납품 기록 없음',  # 총 납품 금액
                followup.notes or ''
            ]
            
            # 히스토리 정보 추가
            histories = list(followup.histories.all().order_by('-created_at'))
            for i in range(max_histories):
                if i < len(histories):
                    history = histories[i]
                    history_text = f"[{history.created_at.strftime('%Y-%m-%d')}] {history.get_action_type_display()}: {history.content or ''}"
                    data.append(history_text)
                else:
                    data.append('')
            
            # 데이터 셀에 값 입력 및 스타일 적용
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = border
                cell.alignment = wrap_alignment
            current_row += 1
        
        # 부서 사이에 빈 행 추가
        current_row += 1
    
    # 컬럼 너비 자동 조정 (부서별 그룹화에 맞게 수정)
    column_widths = {
        1: 15,   # 고객명
        2: 12,   # 책임자
        3: 15,   # 핸드폰 번호
        4: 25,   # 메일 주소
        5: 30,   # 상세 주소
        6: 10,   # 고객 등급
        7: 60,   # 납품 품목 (더 넓게 - 모든 품목 표시를 위해)
        8: 15,   # 총 납품 금액
        9: 30,   # 상세 내용
    }
    
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        col_num = column[0].column
        
        # 미리 정의된 너비가 있으면 사용
        if col_num in column_widths:
            ws.column_dimensions[column_letter].width = column_widths[col_num]
        else:
            # 히스토리 컬럼들에 대한 자동 조정
            max_length = 0
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ''
                    # 한글과 영문의 너비 차이를 고려
                    korean_chars = len([c for c in cell_value if ord(c) > 127])
                    english_chars = len(cell_value) - korean_chars
                    adjusted_length = korean_chars * 2 + english_chars
                    
                    if adjusted_length > max_length:
                        max_length = adjusted_length
                except:
                    pass
            
            # 히스토리 컬럼은 최소 20, 최대 50으로 제한
            adjusted_width = min(max(max_length + 3, 20), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 응답 생성
    today = datetime.now().strftime('%Y%m%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"팔로우업_전체정보_부서별_{today}.xlsx"
    
    # 한글 파일명을 올바르게 인코딩
    from urllib.parse import quote
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    
    # 엑셀 파일을 메모리에서 저장
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    response.write(excel_file.getvalue())
    
    return response


@login_required
def followup_basic_excel_download(request):
    """팔로우업 기본 정보 엑셀 다운로드 (권한 체크)"""
    user_profile = get_user_profile(request.user)
    
    # 엑셀 다운로드 권한 체크
    if not user_profile.can_excel_download():
        messages.error(request, '엑셀 다운로드 권한이 없습니다. 관리자에게 문의해주세요.')
        return redirect('reporting:followup_list')
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io
    from datetime import datetime
    
    user_profile = get_user_profile(request.user)
    
    # 권한에 따른 데이터 필터링 (기존 로직과 동일)
    if user_profile.can_view_all_users():
        accessible_users = get_accessible_users(request.user, request)
        followups = FollowUp.objects.filter(user__in=accessible_users).select_related(
            'user', 'company', 'department'
        )
    else:
        followups = FollowUp.objects.filter(user=request.user).select_related(
            'user', 'company', 'department'
        )
    
    # 검색 필터 적용
    search_query = request.GET.get('search')
    if search_query:
        followups = followups.filter(
            Q(customer_name__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(department__name__icontains=search_query) |
            Q(manager__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # 우선순위 필터 적용
    priority_filter = request.GET.get('priority')
    if priority_filter:
        followups = followups.filter(priority=priority_filter)
    
    # 엑셀 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "팔로우업 기본 정보"
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
    department_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # 헤더 생성 (업체/부서는 그룹 구분행에 표시)
    headers = ['고객명', '책임자', '핸드폰 번호', '메일 주소']
    
    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # 우선순위 정렬을 위한 순서 정의 (낮을수록 높은 우선순위)
    PRIORITY_ORDER = {
        'urgent': 1,      # 긴급 - 가장 높음
        'followup': 2,    # 팔로업
        'scheduled': 3,   # 예정
        'long_term': 4,   # 장기 - 가장 낮음
    }
    
    # 부서별로 그룹화
    from collections import defaultdict
    department_groups = defaultdict(lambda: {'followups': [], 'highest_priority': 99})
    for followup in followups:
        company_name = followup.company.name if followup.company else '미지정 업체'
        department_name = followup.department.name if followup.department else '미지정 부서'
        group_key = f"{company_name} / {department_name}"
        department_groups[group_key]['followups'].append(followup)
        
        # 해당 부서 내 가장 높은 우선순위 업데이트
        priority_value = PRIORITY_ORDER.get(followup.priority, 99)
        if priority_value < department_groups[group_key]['highest_priority']:
            department_groups[group_key]['highest_priority'] = priority_value
    
    # 데이터 입력 (우선순위 높은 부서가 먼저, 그 다음 회사명/부서명 순)
    sorted_groups = sorted(
        department_groups.items(),
        key=lambda x: (x[1]['highest_priority'], x[0])
    )
    
    current_row = 2
    for group_key, group_data in sorted_groups:
        group_followups = group_data['followups']
        # 부서 구분 행 추가
        cell = ws.cell(row=current_row, column=1, value=group_key)
        cell.font = Font(bold=True)
        cell.fill = department_fill
        cell.border = border
        cell.alignment = left_alignment
        
        # 나머지 열도 스타일 적용 및 병합
        for col in range(2, 5):
            cell = ws.cell(row=current_row, column=col, value='')
            cell.fill = department_fill
            cell.border = border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        # 해당 부서의 팔로우업 데이터 입력
        for followup in group_followups:
            manager_name = followup.manager or ''
            
            data = [
                followup.customer_name or '',
                manager_name,
                followup.phone_number or '',
                followup.email or ''
            ]
            
            # 데이터 셀에 값 입력 및 스타일 적용
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = border
                cell.alignment = left_alignment
            
            current_row += 1
    
    # 컬럼 너비 자동 조정 (개선된 버전)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ''
                # 한글과 영문의 너비 차이를 고려
                korean_chars = len([c for c in cell_value if ord(c) > 127])
                english_chars = len(cell_value) - korean_chars
                adjusted_length = korean_chars * 2 + english_chars
                
                if adjusted_length > max_length:
                    max_length = adjusted_length
            except:
                pass
        
        # 최소 8, 최대 50 문자로 제한하고, 여유분 추가
        adjusted_width = min(max(max_length + 3, 8), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 응답 생성
    today = datetime.now().strftime('%Y%m%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"기본정보_{today}.xlsx"
    
    # 한글 파일명을 올바르게 인코딩
    from urllib.parse import quote
    encoded_filename = quote(filename.encode('utf-8'))
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    
    # 엑셀 파일을 메모리에서 저장
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    response.write(excel_file.getvalue())
    
    return response

# 파일 관리 뷰들을 별도 모듈에서 import
from .file_views import (
    file_download_view, file_delete_view, history_files_api,
    schedule_file_upload, schedule_file_download, schedule_file_delete, schedule_files_api
) 

# ============ 고객 리포트 관련 뷰들 ============

@login_required
def customer_report_view(request):
    """고객별 활동 요약 리포트 목록 - Schedule DeliveryItem도 포함 (같은 회사 직원 데이터 필터링)"""
    from django.db.models import Count, Sum, Max, Q
    from django.contrib.auth.models import User
    from decimal import Decimal
    
    user_profile = get_user_profile(request.user)
    
    # === 년도 필터 ===
    from django.utils import timezone
    current_year = timezone.now().year
    selected_year = request.GET.get('year')
    if selected_year:
        try:
            selected_year = int(selected_year)
        except ValueError:
            selected_year = current_year
    else:
        selected_year = current_year
    
    # 년도 범위 생성 (최근 5년 + 내년)
    year_range = list(range(current_year - 4, current_year + 2))
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용) - 매니저 제외
    company_users = []
    if user_profile and user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    is_viewing_others = False
    target_user = request.user  # 기본값
    
    if data_filter == 'all' and user_profile and user_profile.company:
        # 같은 회사 전체 (salesman만)
        filter_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        )
        target_user = None  # 전체
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        # 특정 직원 (같은 회사 확인)
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True
            )
            filter_users = User.objects.filter(id=selected_filter_user.id)
            target_user = selected_filter_user
            is_viewing_others = True
        except User.DoesNotExist:
            filter_users = User.objects.filter(id=request.user.id)
            target_user = request.user
            data_filter = 'me'
    else:
        # 'me' - 본인만
        filter_users = User.objects.filter(id=request.user.id)
        target_user = request.user
    
    # 카테고리 필터링
    category_id = request.GET.get('category')
    # 모든 카테고리 가져오기 (계층 구조 포함)
    all_categories = CustomerCategory.objects.filter(user=request.user).select_related('parent').order_by('order', 'name')
    # 계층 구조로 정리
    parent_categories = all_categories.filter(parent__isnull=True)
    categories_dict = {}
    for cat in parent_categories:
        children = all_categories.filter(parent=cat)
        categories_dict[cat] = list(children)
    
    selected_category = None
    
    # 모든 고객 조회
    followups = FollowUp.objects.all()
    
    # 카테고리 필터 적용 (부서 기준)
    if category_id:
        try:
            selected_category = CustomerCategory.objects.get(id=category_id, user=request.user)
            # 선택한 카테고리가 상위 카테고리인 경우, 하위 카테고리도 포함
            if selected_category.is_parent():
                child_ids = selected_category.children.values_list('id', flat=True)
                followups = followups.filter(Q(department__category=selected_category) | Q(department__category_id__in=child_ids))
            else:
                followups = followups.filter(department__category=selected_category)
        except CustomerCategory.DoesNotExist:
            pass
    
    # 검색 기능
    search_query = request.GET.get('search')
    if search_query:
        followups = followups.filter(
            Q(customer_name__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(manager__icontains=search_query)
        )
    
    # ✅ 성능 최적화: Prefetch로 N+1 쿼리 방지
    from django.db.models import Prefetch
    import datetime
    
    # 선택된 년도의 시작일과 종료일
    year_start = datetime.datetime(selected_year, 1, 1)
    year_end = datetime.datetime(selected_year, 12, 31, 23, 59, 59)
    
    # 사용자 필터 설정
    if target_user is None:
        user_filter_q = Q(user__in=filter_users)
        prepayment_filter_q = Q(created_by__in=filter_users)
    else:
        user_filter_q = Q(user=target_user)
        prepayment_filter_q = Q(created_by=target_user)
    
    # 년도 필터 (각 모델에 맞는 날짜 필드 사용)
    history_year_filter_q = Q(created_at__year=selected_year)
    schedule_year_filter_q = Q(visit_date__year=selected_year)  # Schedule은 visit_date 사용
    prepayment_year_filter_q = Q(created_at__year=selected_year)
    
    # ✅ 핵심 최적화: 활동이 있는 고객만 먼저 필터링
    # 1. 대상 사용자의 History가 있는 FollowUp ID (선택된 년도)
    history_followup_ids = History.objects.filter(user_filter_q & history_year_filter_q).values_list('followup_id', flat=True).distinct()
    # 2. 대상 사용자의 Schedule이 있는 FollowUp ID (선택된 년도, visit_date 기준)
    schedule_followup_ids = Schedule.objects.filter(user_filter_q & schedule_year_filter_q).values_list('followup_id', flat=True).distinct()
    # 3. 선결제 잔액이 있는 FollowUp ID (년도 무관, 잔액만 체크)
    prepayment_followup_ids = Prepayment.objects.filter(prepayment_filter_q & Q(balance__gt=0)).values_list('customer_id', flat=True).distinct()
    
    # 활동이 있는 FollowUp ID 합집합
    active_followup_ids = set(history_followup_ids) | set(schedule_followup_ids) | set(prepayment_followup_ids)
    
    # ✅ 활동이 있는 고객만 조회 (대폭 감소)
    followups = followups.filter(id__in=active_followup_ids)
    
    # ✅ select_related로 FK 조인 최적화
    followups = followups.select_related('company', 'department', 'department__category', 'user')
    
    # ✅ 모든 관련 데이터를 한 번에 가져오기 (Prefetch)
    # 선결제는 년도 필터 없이 모두 가져옴 (잔액 확인용)
    followups = followups.prefetch_related(
        Prefetch('histories', queryset=History.objects.filter(user_filter_q & history_year_filter_q).select_related('user')),
        Prefetch('schedules', queryset=Schedule.objects.filter(user_filter_q & schedule_year_filter_q).select_related('user').prefetch_related('delivery_items_set')),
        Prefetch('prepayments', queryset=Prepayment.objects.filter(prepayment_filter_q).select_related('created_by'))
    )
    
    # 각 고객별 통계 계산
    followups_with_stats = []
    total_amount_sum = Decimal('0')
    total_meetings_sum = 0
    total_deliveries_sum = 0
    total_unpaid_sum = 0
    prepayment_customers = set()
    
    for followup in followups:
        # ✅ Prefetch된 데이터 사용 (추가 쿼리 없음!)
        all_histories = list(followup.histories.all())
        all_schedules = list(followup.schedules.all())
        
        # History 통계
        meetings = sum(1 for h in all_histories if h.action_type == 'customer_meeting')
        delivery_histories = [h for h in all_histories if h.action_type == 'delivery_schedule']
        deliveries = len(delivery_histories)
        history_amount = sum(h.delivery_amount or Decimal('0') for h in delivery_histories)
        unpaid = sum(1 for h in delivery_histories if not h.tax_invoice_issued)
        
        # Schedule 통계
        schedule_deliveries = [s for s in all_schedules if s.activity_type == 'delivery']
        schedule_delivery_count = len(schedule_deliveries)
        schedule_amount = sum(
            item.total_price or Decimal('0')
            for schedule in schedule_deliveries
            for item in schedule.delivery_items_set.all()
        )
        
        # 세금계산서 현황 계산 (History + Schedule DeliveryItem 통합 - 중복 제거)
        # 1. History와 Schedule 연결 관계 분석
        history_with_schedule_ids = set(
            h.schedule_id for h in delivery_histories if h.schedule_id is not None
        )
        
        # 2. History 기반 세금계산서 현황 (Schedule 연결 여부로 구분)
        history_with_schedule_issued = [
            h.schedule_id for h in delivery_histories 
            if h.schedule_id is not None and h.tax_invoice_issued
        ]
        history_with_schedule_pending = [
            h.schedule_id for h in delivery_histories 
            if h.schedule_id is not None and not h.tax_invoice_issued
        ]
        
        history_without_schedule_issued = sum(
            1 for h in delivery_histories 
            if h.schedule_id is None and h.tax_invoice_issued
        )
        history_without_schedule_pending = sum(
            1 for h in delivery_histories 
            if h.schedule_id is None and not h.tax_invoice_issued
        )
        
        # 3. Schedule DeliveryItem 세금계산서 현황 (Prefetch된 데이터 사용)
        # Schedule만 있는 경우 (History에 연결되지 않은 Schedule)
        schedule_only_deliveries = [
            s for s in schedule_deliveries 
            if s.id not in history_with_schedule_ids
        ]
        
        schedule_only_issued = 0
        schedule_only_pending = 0
        
        for schedule in schedule_only_deliveries:
            items = list(schedule.delivery_items_set.all())
            if items:
                # 해당 Schedule에 하나라도 발행된 품목이 있으면 발행으로 카운팅
                if any(item.tax_invoice_issued for item in items):
                    schedule_only_issued += 1
                else:
                    schedule_only_pending += 1
        
        # 4. 중복 제거된 최종 세금계산서 현황
        # History 우선 원칙: History와 Schedule이 모두 있는 경우 History 상태를 사용
        history_schedule_issued_set = set(history_with_schedule_issued)
        history_schedule_pending_set = set(history_with_schedule_pending)
        
        total_tax_issued = len(history_schedule_issued_set) + history_without_schedule_issued + schedule_only_issued
        total_tax_pending = len(history_schedule_pending_set) + history_without_schedule_pending + schedule_only_pending
        
        # 중복 제거된 납품 횟수 계산 (Prefetch된 데이터 사용)
        # History에 기록된 일정 ID들
        history_schedule_ids = set(
            h.schedule_id for h in delivery_histories if h.schedule_id is not None
        )
        # Schedule에 DeliveryItem이 있는 일정 ID들  
        schedule_ids = set(s.id for s in schedule_deliveries)
        
        # History만 있는 납품 + Schedule만 있는 납품 + 둘 다 있는 경우는 1개로 카운팅
        history_only_deliveries = sum(1 for h in delivery_histories if h.schedule_id is None)
        unique_deliveries_count = len(history_schedule_ids | schedule_ids) + history_only_deliveries
        
        # 통합 통계 (중복 제거) - Schedule 기준으로 집계
        # 1. Schedule별로 가장 최근 History 1개만 선택 (중복 방지)
        schedule_to_history = {}  # schedule_id -> 가장 최근 History
        standalone_histories = []  # Schedule 없는 History
        
        for h in delivery_histories:
            if h.schedule_id:
                if h.schedule_id not in schedule_to_history:
                    # 이 Schedule의 첫 번째(가장 최근) History만 저장
                    schedule_to_history[h.schedule_id] = h
            else:
                standalone_histories.append(h)
        
        # 2. Schedule ID별 DeliveryItem 금액 맵 생성
        schedule_item_amounts = {}
        for schedule in schedule_deliveries:
            items = list(schedule.delivery_items_set.all())
            schedule_item_amounts[schedule.id] = sum(item.total_price or Decimal('0') for item in items)
        
        # 3. 금액 계산 (Schedule 기준)
        total_amount = Decimal('0')
        processed_schedule_ids = set()
        
        # Schedule에 연결된 History (중복 제거됨)
        for schedule_id, history in schedule_to_history.items():
            processed_schedule_ids.add(schedule_id)
            schedule_item_amount = schedule_item_amounts.get(schedule_id, Decimal('0'))
            if schedule_item_amount > 0:
                # DeliveryItem이 있으면 DeliveryItem 금액 사용
                total_amount += schedule_item_amount
            else:
                # DeliveryItem이 없으면 History 금액 사용
                total_amount += history.delivery_amount or Decimal('0')
        
        # Schedule 없는 독립 History
        for h in standalone_histories:
            total_amount += h.delivery_amount or Decimal('0')
        
        # History 없는 Schedule의 DeliveryItem 금액 추가
        for schedule in schedule_deliveries:
            if schedule.id not in processed_schedule_ids:
                total_amount += schedule_item_amounts.get(schedule.id, Decimal('0'))
        
        total_meetings_count = meetings
        total_deliveries_count = unique_deliveries_count
        last_contact = max((h.created_at for h in all_histories), default=None)  # Prefetch된 데이터 사용
        
        # ✅ 선결제 통계 계산 - Prefetch된 데이터 사용 (추가 쿼리 없음!)
        all_prepayments = list(followup.prepayments.all())
        
        # 전체 선결제 잔액 (년도 무관 - 모든 년도의 잔액 합계)
        prepayment_balance_total = sum(p.balance or Decimal('0') for p in all_prepayments)
        prepayment_total_all = sum(p.amount or Decimal('0') for p in all_prepayments)
        prepayment_count_all = len([p for p in all_prepayments if p.balance > 0])  # 잔액이 있는 선결제만 카운트
        
        # 선결제 등록자 정보 (잔액이 있는 것만)
        prepayments_with_balance = [p for p in all_prepayments if p.balance > 0]
        prepayment_creators = list(set([p.created_by.get_full_name() or p.created_by.username for p in prepayments_with_balance])) if prepayments_with_balance else []
        
        # 객체에 통계 추가
        followup.total_meetings = total_meetings_count
        followup.total_deliveries = total_deliveries_count
        followup.total_amount = total_amount
        followup.tax_invoices_issued = total_tax_issued  # 세금계산서 발행 건수
        followup.tax_invoices_pending = total_tax_pending  # 세금계산서 미발행 건수
        followup.unpaid_count = total_tax_pending  # 미발행 건수를 unpaid_count로 사용
        followup.last_contact = last_contact
        followup.prepayment_total = prepayment_total_all  # 전체 선결제 총액
        followup.prepayment_balance = prepayment_balance_total  # 전체 선결제 잔액 (년도 무관)
        followup.prepayment_count = prepayment_count_all  # 잔액이 있는 선결제 건수
        followup.prepayment_creators = ', '.join(prepayment_creators) if prepayment_creators else ''  # 선결제 등록자
        
        # target_user의 활동이 하나라도 있는 경우 추가 (미팅, 납품, 선결제 잔액)
        # 선결제는 잔액이 남아있으면 년도 상관없이 표시
        if total_meetings_count > 0 or total_deliveries_count > 0 or prepayment_balance_total > 0:
            followups_with_stats.append(followup)
            
            # 전체 통계 누적 (해당 년도 데이터만)
            total_amount_sum += total_amount
            total_meetings_sum += total_meetings_count
            total_deliveries_sum += total_deliveries_count
            total_unpaid_sum += total_tax_pending  # 세금계산서 미발행 건수로 변경
            if prepayment_balance_total > 0:  # 전체 잔액이 있으면 카운트
                prepayment_customers.add(followup.id)
    
    # === 부서별 그룹화 ===
    from collections import defaultdict
    
    departments_data = defaultdict(lambda: {
        'company': None,
        'department': None,
        'customers': [],
        'total_amount': Decimal('0'),
        'total_meetings': 0,
        'total_deliveries': 0,
        'tax_invoices_issued': 0,
        'tax_invoices_pending': 0,
        'prepayment_total': Decimal('0'),
        'prepayment_balance': Decimal('0'),
        'last_contact': None,
    })
    
    for followup in followups_with_stats:
        # 부서 키 생성 (company_id-department_id)
        dept_key = f"{followup.company_id}-{followup.department_id}"
        
        dept_data = departments_data[dept_key]
        dept_data['company'] = followup.company
        dept_data['department'] = followup.department
        dept_data['customers'].append(followup)
        dept_data['total_amount'] += followup.total_amount
        dept_data['total_meetings'] += followup.total_meetings
        dept_data['total_deliveries'] += followup.total_deliveries
        dept_data['tax_invoices_issued'] += followup.tax_invoices_issued
        dept_data['tax_invoices_pending'] += followup.tax_invoices_pending
        dept_data['prepayment_total'] += followup.prepayment_total
        dept_data['prepayment_balance'] += followup.prepayment_balance
        
        # 최근 연락일 갱신
        if followup.last_contact:
            if dept_data['last_contact'] is None or followup.last_contact > dept_data['last_contact']:
                dept_data['last_contact'] = followup.last_contact
    
    # 부서 데이터를 리스트로 변환
    departments_list = list(departments_data.values())
    
    # 정렬 처리 - 기본값: 총 납품 금액 내림차순
    sort_by = request.GET.get('sort', 'amount')
    sort_order = request.GET.get('order', 'desc')
    
    from django.utils import timezone
    
    # 부서별 정렬 키 매핑
    sort_key_map = {
        'company': lambda x: (x['company'].name if x['company'] else '').lower(),
        'department': lambda x: (x['department'].name if x['department'] else '').lower(),
        'meetings': lambda x: x['total_meetings'],
        'deliveries': lambda x: x['total_deliveries'],
        'amount': lambda x: x['total_amount'],
        'prepayment': lambda x: x['prepayment_balance'],
        'unpaid': lambda x: x['tax_invoices_pending'],
        'last_contact': lambda x: x['last_contact'] or timezone.now().replace(year=1900),
    }
    
    # 정렬 키가 유효한지 확인
    if sort_by in sort_key_map:
        departments_list.sort(
            key=sort_key_map[sort_by],
            reverse=(sort_order == 'desc')
        )
    else:
        # 기본 정렬: 총 납품 금액 기준
        departments_list.sort(key=lambda x: x['total_amount'], reverse=True)
    
    # 페이지네이션 추가
    from django.core.paginator import Paginator
    paginator = Paginator(departments_list, 30)  # 페이지당 30개 부서
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'departments': page_obj,  # 부서별 그룹화된 데이터
        'page_obj': page_obj,
        'total_departments': len(departments_list),  # 총 부서 수
        'total_customers': len(followups_with_stats),  # 총 고객 수
        'total_amount_sum': total_amount_sum,
        'total_meetings_sum': total_meetings_sum,
        'total_deliveries_sum': total_deliveries_sum,
        'total_unpaid_sum': total_unpaid_sum,
        'total_prepayment_customers': len(prepayment_customers),  # 선결제 고객 수
        'sort_by': sort_by,
        'sort_order': sort_order,
        'search_query': search_query,
        'page_title': '부서별 리포트',  # 제목 변경
        # 새로운 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
        # 카테고리 필터
        'all_categories': all_categories,
        'categories_dict': categories_dict,
        'selected_category': selected_category,
        'category_id': category_id,
        # 년도 필터
        'selected_year': selected_year,
        'year_range': year_range,
    }
    
    return render(request, 'reporting/customer_report_list.html', context)

@login_required
def customer_detail_report_view(request, followup_id):
    """특정 고객의 상세 활동 리포트"""
    from django.db.models import Count, Sum, Q
    from django.core import serializers
    from django.core.serializers.json import DjangoJSONEncoder
    from datetime import datetime, timedelta
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    # 권한 확인 및 팔로우업 조회
    try:
        followup = FollowUp.objects.select_related('user', 'company', 'department').get(id=followup_id)
        
        # 권한 체크 (같은 회사 소속이면 접근 가능)
        if not can_access_followup(request.user, followup):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:customer_report')
            
    except FollowUp.DoesNotExist:
        messages.error(request, '해당 고객 정보를 찾을 수 없습니다.')
        return redirect('reporting:customer_report')
    
    # 본인 고객인지 확인
    is_own_customer = (request.user == followup.user)
    user_profile = get_user_profile(request.user)
    # 관리자/매니저만 전체 데이터 조회 가능, 일반 사용자는 본인 작성 데이터만
    can_view_all = user_profile.is_admin() or user_profile.is_manager()
    
    # 해당 고객의 활동 히스토리 (권한에 따라 필터링)
    if can_view_all:
        histories = History.objects.filter(followup=followup).order_by('-created_at')
    else:
        # 동료 고객: 본인이 작성한 히스토리만
        histories = History.objects.filter(followup=followup, user=request.user).order_by('-created_at')
    
    # 기본 통계 계산
    total_meetings = histories.filter(action_type='customer_meeting').count()
    total_deliveries = histories.filter(action_type='delivery_schedule').count()
    
    # Schedule DeliveryItem 총액 계산 (권한에 따라 필터링)
    if can_view_all:
        schedule_deliveries = Schedule.objects.filter(
            followup=followup, 
            activity_type='delivery'
        ).prefetch_related('delivery_items_set')
    else:
        # 동료 고객: 본인이 작성한 스케줄만
        schedule_deliveries = Schedule.objects.filter(
            followup=followup, 
            activity_type='delivery',
            user=request.user
        ).prefetch_related('delivery_items_set')
    
    # 총 금액 계산 (중복 제거 - Schedule 기준)
    total_amount = 0
    processed_schedule_ids = set()
    
    # 1. Schedule의 DeliveryItem 금액 집계
    for schedule in schedule_deliveries:
        schedule_total = 0
        has_items = False
        
        for item in schedule.delivery_items_set.all():
            has_items = True
            if item.total_price:
                schedule_total += float(item.total_price)
            elif item.unit_price:
                schedule_total += float(item.unit_price) * item.quantity * 1.1
        
        if has_items:
            # DeliveryItem이 있으면 그 금액 사용
            total_amount += schedule_total
        else:
            # DeliveryItem이 없으면 연결된 History 중 가장 최근 것의 금액 사용
            related_history = histories.filter(
                action_type='delivery_schedule',
                schedule_id=schedule.id
            ).order_by('-created_at').first()
            if related_history and related_history.delivery_amount:
                total_amount += float(related_history.delivery_amount)
        
        processed_schedule_ids.add(schedule.id)
    
    # 2. Schedule에 연결되지 않은 History의 금액 추가
    standalone_histories = histories.filter(
        action_type='delivery_schedule',
        schedule__isnull=True
    )
    for history in standalone_histories:
        if history.delivery_amount:
            total_amount += float(history.delivery_amount)
    
    # 세금계산서 현황 계산 (History + Schedule 통합, 중복 제거)
    delivery_histories = histories.filter(action_type='delivery_schedule')
    
    # 1. History와 Schedule 연결 관계 분석
    history_with_schedule_ids = set(
        delivery_histories.filter(schedule__isnull=False).values_list('schedule_id', flat=True)
    )
    
    # 2. History 기반 세금계산서 현황
    history_tax_issued = 0
    history_tax_pending = 0
    
    # History와 Schedule이 연결된 경우
    history_with_schedule_issued = delivery_histories.filter(
        schedule__isnull=False, tax_invoice_issued=True
    ).values_list('schedule_id', flat=True)
    history_with_schedule_pending = delivery_histories.filter(
        schedule__isnull=False, tax_invoice_issued=False
    ).values_list('schedule_id', flat=True)
    
    # History만 있는 경우 (Schedule에 연결되지 않은 History)
    history_without_schedule_issued = delivery_histories.filter(
        schedule__isnull=True, tax_invoice_issued=True
    ).count()
    history_without_schedule_pending = delivery_histories.filter(
        schedule__isnull=True, tax_invoice_issued=False
    ).count()
    
    # 3. Schedule DeliveryItem 세금계산서 현황 (History에 연결되지 않은 Schedule만)
    schedule_tax_issued = 0
    schedule_tax_pending = 0
    
    schedule_only_deliveries = schedule_deliveries.exclude(
        id__in=history_with_schedule_ids
    )
    
    for schedule in schedule_only_deliveries:
        items = schedule.delivery_items_set.all()
        if items.exists():
            # 해당 Schedule에 하나라도 발행된 품목이 있으면 발행으로 카운팅
            if items.filter(tax_invoice_issued=True).exists():
                schedule_tax_issued += 1
            else:
                schedule_tax_pending += 1
    
    # 4. 중복 제거된 최종 세금계산서 현황
    # History 우선 원칙: History와 Schedule이 모두 있는 경우 History 상태를 사용
    history_schedule_issued_set = set(history_with_schedule_issued)
    history_schedule_pending_set = set(history_with_schedule_pending)
    
    tax_invoices_issued = len(history_schedule_issued_set) + history_without_schedule_issued + schedule_tax_issued
    tax_invoices_pending = len(history_schedule_pending_set) + history_without_schedule_pending + schedule_tax_pending
    
    # 월별 활동 통계 (최근 12개월)
    from django.db.models.functions import TruncMonth
    monthly_stats = histories.filter(
        created_at__gte=datetime.now() - timedelta(days=365)
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        meetings=Count('id', filter=Q(action_type='customer_meeting')),
        deliveries=Count('id', filter=Q(action_type='delivery_schedule')),
        amount=Sum('delivery_amount')
    ).order_by('month')
    
    # Chart.js용 데이터 준비
    chart_labels = []
    chart_meetings = []
    chart_deliveries = []
    chart_amounts = []
    
    for stat in monthly_stats:
        chart_labels.append(stat['month'].strftime('%Y-%m'))
        chart_meetings.append(stat['meetings'])
        chart_deliveries.append(stat['deliveries'])
        chart_amounts.append(float(stat['amount'] or 0))
    
    # 납품 내역 상세 - 중복 완전 제거를 위해 정렬 순서 변경
    # created_at 기준으로 정렬하여 가장 최근에 작성된 History가 먼저 오도록
    delivery_histories = histories.filter(
        action_type='delivery_schedule'
    ).order_by('-created_at')  # created_at 기준으로만 정렬 (가장 최근 작성 우선)
    
    # Schedule 납품 일정에 총액 정보 추가
    for schedule in schedule_deliveries:
        schedule_total_amount = 0
        tax_invoice_issued_count = 0
        total_items_count = 0
        
        for item in schedule.delivery_items_set.all():
            if item.total_price:
                item_total = float(item.total_price)
            elif item.unit_price:
                item_total = float(item.unit_price) * item.quantity * 1.1
            else:
                item_total = 0
            schedule_total_amount += item_total
            total_items_count += 1
            if item.tax_invoice_issued:
                tax_invoice_issued_count += 1
        
        schedule.calculated_total_amount = schedule_total_amount
        schedule.tax_invoice_issued_count = tax_invoice_issued_count
        schedule.total_items_count = total_items_count
    
    # 통합 납품 내역 생성 - 완전히 새로운 방식 (중복 완전 제거)
    integrated_deliveries = []
    
    # delivery_histories를 리스트로 변환 (쿼리셋이 여러 번 평가되는 것 방지)
    delivery_histories_list = list(delivery_histories)
    
    # 1단계: Schedule별로 그룹화하여 가장 최근 History 1개만 선택
    schedule_to_history = {}  # schedule_id -> 가장 최근 History
    standalone_histories = []  # Schedule 없는 History
    
    for history in delivery_histories_list:
        if history.schedule_id:
            # Schedule에 연결된 History
            if history.schedule_id not in schedule_to_history:
                # 이 Schedule의 첫 번째(가장 최근 작성) History만 저장
                schedule_to_history[history.schedule_id] = history
        else:
            # Schedule 없는 독립 History
            standalone_histories.append(history)
    
    # 2단계: 선택된 Schedule 연결 History를 integrated_deliveries에 추가
    for schedule_id, history in schedule_to_history.items():
        delivery_data = {
            'type': 'history',
            'id': history.id,
            'date': (history.delivery_date or history.created_at.date()).strftime('%Y-%m-%d'),
            'schedule_id': schedule_id,
            'items_display': history.delivery_items or None,
            'amount': history.delivery_amount,
            'tax_invoice_issued': history.tax_invoice_issued,
            'content': history.content,
            'user': history.user.username,
            'has_schedule_items': False,
        }
        
        # Schedule의 DeliveryItem 확인
        schedule = next((s for s in schedule_deliveries if s.id == schedule_id), None)
        if schedule:
            schedule_items = schedule.delivery_items_set.all()
            if schedule_items.exists():
                delivery_data['has_schedule_items'] = True
                delivery_data['schedule_items'] = schedule_items
        
        integrated_deliveries.append(delivery_data)
    
    # 3단계: Schedule 없는 독립 History 중복 제거 (같은 날짜+금액은 1개만)
    standalone_unique = {}  # (date, amount) -> History
    for history in standalone_histories:
        delivery_date = (history.delivery_date or history.created_at.date()).strftime('%Y-%m-%d')
        delivery_amount = float(history.delivery_amount) if history.delivery_amount else 0
        key = (delivery_date, delivery_amount)
        
        if key not in standalone_unique:
            standalone_unique[key] = history
    
    # 독립 History 추가
    for history in standalone_unique.values():
        delivery_data = {
            'type': 'history',
            'id': history.id,
            'date': (history.delivery_date or history.created_at.date()).strftime('%Y-%m-%d'),
            'schedule_id': None,
            'items_display': history.delivery_items or None,
            'amount': history.delivery_amount,
            'tax_invoice_issued': history.tax_invoice_issued,
            'content': history.content,
            'user': history.user.username,
            'has_schedule_items': False,
        }
        integrated_deliveries.append(delivery_data)
    
    # 4단계: History 없는 Schedule 추가
    processed_schedule_ids = set(schedule_to_history.keys())
    for schedule in schedule_deliveries:
        if schedule.id not in processed_schedule_ids:
            delivery_data = {
                'type': 'schedule_only',
                'id': schedule.id,
                'date': schedule.visit_date.strftime('%Y-%m-%d') if schedule.visit_date else '',
                'schedule_id': schedule.id,
                'items_display': None,
                'amount': 0,
                'tax_invoice_issued': schedule.tax_invoice_issued_count == schedule.total_items_count if schedule.total_items_count > 0 else False,
                'content': schedule.notes or '일정 기반 납품',
                'user': schedule.user.username,
                'has_schedule_items': True,
                'schedule_amount': schedule.calculated_total_amount,
            }
            integrated_deliveries.append(delivery_data)
    
    # 날짜순 정렬
    integrated_deliveries.sort(key=lambda x: x['date'], reverse=True)
    
    # 미팅 기록
    meeting_histories = histories.filter(
        action_type='customer_meeting'
    ).order_by('-meeting_date', '-created_at')
    
    # 세금계산서 수정 권한: 항상 True (모든 사용자가 토글 가능)
    can_modify_tax_invoice = True
    
    context = {
        'followup': followup,
        'histories': histories,
        'total_meetings': total_meetings,
        'total_deliveries': total_deliveries,
        'total_amount': total_amount,
        'tax_invoices_issued': tax_invoices_issued,
        'tax_invoices_pending': tax_invoices_pending,
        'delivery_histories': delivery_histories,
        'schedule_deliveries': schedule_deliveries,
        'integrated_deliveries': integrated_deliveries,
        'integrated_deliveries_json': json.dumps(integrated_deliveries, ensure_ascii=False, cls=DjangoJSONEncoder),
        'meeting_histories': meeting_histories,
        'can_modify_tax_invoice': can_modify_tax_invoice,  # 세금계산서 수정 권한
        'chart_data': {
            'labels': json.dumps(chart_labels),
            'meetings': json.dumps(chart_meetings),
            'deliveries': json.dumps(chart_deliveries),
            'amounts': json.dumps(chart_amounts),
        },
        'page_title': f'{followup.customer_name or "고객명 미정"} 상세 리포트',
    }
    
    return render(request, 'reporting/customer_detail_report.html', context)


@require_http_methods(["POST"])
@login_required
def toggle_schedule_delivery_tax_invoice(request, schedule_id):
    """Schedule의 DeliveryItem 세금계산서 발행여부 일괄 토글 API"""
    try:
        schedule = get_object_or_404(Schedule, pk=schedule_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 가능
        if not can_modify_user_data(request.user, schedule.user):
            return JsonResponse({
                'success': False,
                'error': '세금계산서 상태를 변경할 권한이 없습니다.'
            }, status=403)
        
        # Schedule에 연결된 DeliveryItem들 조회
        delivery_items = schedule.delivery_items_set.all()
        
        if not delivery_items.exists():
            return JsonResponse({
                'success': False,
                'error': '해당 일정에 납품 품목이 없습니다.'
            })
        
        # 현재 상태 확인 (하나라도 미발행이면 모두 발행으로, 모두 발행이면 모두 미발행으로)
        any_not_issued = delivery_items.filter(tax_invoice_issued=False).exists()
        new_status = any_not_issued  # 미발행이 있으면 True(발행)로, 없으면 False(미발행)로
        
        # 일괄 업데이트
        updated_count = delivery_items.update(tax_invoice_issued=new_status)
        
        # 연결된 히스토리들도 함께 업데이트
        try:
            related_histories = History.objects.filter(
                schedule=schedule, 
                action_type='delivery_schedule'
            )
            
            if related_histories.exists():
                history_updated_count = related_histories.update(tax_invoice_issued=new_status)
                
        except Exception as sync_error:
            logger.error(f"스케줄 납품 품목 토글 시 히스토리 동기화 실패: {sync_error}")
        
        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'updated_count': updated_count,
            'status_text': '발행됨' if new_status else '미발행'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'세금계산서 상태 변경 중 오류가 발생했습니다: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def toggle_history_delivery_tax_invoice(request, history_id):
    """History의 DeliveryItem 세금계산서 발행여부 일괄 토글 API"""
    try:
        history = get_object_or_404(History, pk=history_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 가능
        if not can_modify_user_data(request.user, history.user):
            return JsonResponse({
                'success': False,
                'error': '세금계산서 상태를 변경할 권한이 없습니다.'
            }, status=403)
        
        # History의 Schedule에 연결된 DeliveryItem들 조회
        if not history.schedule:
            return JsonResponse({
                'success': False,
                'error': '일정이 연결되지 않은 히스토리입니다.'
            })
        
        delivery_items = history.schedule.delivery_items_set.all()
        
        if not delivery_items.exists():
            return JsonResponse({
                'success': False,
                'error': '해당 히스토리에 납품 품목이 없습니다.'
            })
        
        # 현재 상태 확인 (하나라도 미발행이면 모두 발행으로, 모두 발행이면 모두 미발행으로)
        any_not_issued = delivery_items.filter(tax_invoice_issued=False).exists()
        new_status = any_not_issued  # 미발행이 있으면 True(발행)로, 없으면 False(미발행)로
        
        # 일괄 업데이트
        updated_count = delivery_items.update(tax_invoice_issued=new_status)
        
        # History도 함께 업데이트
        history.tax_invoice_issued = new_status
        history.save()
        
        return JsonResponse({
            'success': True,
            'new_status': new_status,
            'updated_count': updated_count,
            'status_text': '발행완료' if new_status else '미발행'
        })
        
    except Exception as e:
        logger.error(f"히스토리 세금계산서 토글 오류: {e}")
        return JsonResponse({
            'success': False,
            'error': f'세금계산서 상태 변경 중 오류가 발생했습니다: {str(e)}'
        })


@require_http_methods(["DELETE"])
@login_required
def delete_manager_memo_api(request, history_id):
    """댓글 메모 삭제 API - 매니저는 모든 댓글, 실무자는 자신의 댓글만 삭제 가능"""
    try:
        user_profile = get_user_profile(request.user)
        
        # 히스토리 조회
        history = get_object_or_404(History, pk=history_id)
        
        # 메모 타입이고 부모 히스토리가 있는 댓글인지 확인
        if history.action_type != 'memo' or not history.parent_history:
            return JsonResponse({
                'success': False,
                'error': '댓글 메모만 삭제할 수 있습니다.'
            }, status=400)
        
        # 권한 확인 - 모든 사용자는 본인이 작성한 댓글만 삭제 가능
        can_delete = False
        
        # 매니저가 작성한 댓글인지 확인 (created_by가 있는 경우)
        if history.created_by:
            # 매니저 댓글은 작성한 매니저만 삭제 가능
            if history.created_by == request.user:
                can_delete = True
        else:
            # 일반 실무자 댓글은 작성한 실무자만 삭제 가능
            if history.user == request.user:
                can_delete = True
        
        if not can_delete:
            return JsonResponse({
                'success': False,
                'error': '이 댓글을 삭제할 권한이 없습니다.'
            }, status=403)
        
        # 삭제 실행
        history.delete()
        
        return JsonResponse({
            'success': True,
            'message': '댓글이 성공적으로 삭제되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"Manager memo deletion error: {e}")
        return JsonResponse({
            'success': False,
            'error': '메모 삭제 중 오류가 발생했습니다.'
        }, status=500)


@require_http_methods(["POST"])
@login_required
def add_manager_memo_to_history_api(request, history_id):
    """히스토리에 댓글 메모 추가 API - 매니저와 해당 실무자가 추가 가능"""
    try:
        user_profile = get_user_profile(request.user)
        
        # 부모 히스토리 조회
        parent_history = get_object_or_404(History, pk=history_id)
        
        # 권한 확인 - 매니저이거나 해당 히스토리의 담당자인 경우 허용
        if not (user_profile.is_manager() or request.user == parent_history.user):
            return JsonResponse({
                'success': False,
                'error': '이 히스토리에 댓글을 추가할 권한이 없습니다.'
            }, status=403)
        
        # 매니저가 다른 사용자 데이터에 접근하는 경우 권한 체크
        if user_profile.is_manager() and request.user != parent_history.user:
            if not can_access_user_data(request.user, parent_history.user):
                return JsonResponse({
                    'success': False,
                    'error': '이 히스토리에 접근할 권한이 없습니다.'
                }, status=403)
        
        # 메모 내용 가져오기
        memo_content = request.POST.get('memo', '').strip()
        if not memo_content:
            return JsonResponse({
                'success': False,
                'error': '메모 내용을 입력해주세요.'
            }, status=400)
        
        # 댓글 메모를 부모 히스토리에 연결된 자식 히스토리로 생성
        # 매니저의 경우: created_by에 매니저 정보 저장, user는 원래 실무자 유지
        # 실무자의 경우: created_by는 None, user는 본인
        memo_history = History.objects.create(
            followup=parent_history.followup,
            user=parent_history.user,  # 원래 실무자를 유지
            parent_history=parent_history,  # 부모 히스토리 설정
            action_type='memo',
            content=memo_content,
            created_by=request.user if user_profile.is_manager() else None,  # 매니저인 경우만 created_by 설정
            schedule=parent_history.schedule if parent_history.schedule else None
        )
        
        # 메시지도 역할에 따라 구분
        message = '댓글이 성공적으로 추가되었습니다.'
        if user_profile.is_manager():
            message = '매니저 메모가 성공적으로 추가되었습니다.'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'memo': {
                'id': memo_history.id,
                'content': memo_history.content,
                'created_at': memo_history.created_at.strftime('%Y-%m-%d %H:%M'),
                'created_by': request.user.username
            }
        })
        
    except Exception as e:
        logger.error(f"Add manager memo to history error: {e}")
        return JsonResponse({
            'success': False,
            'error': '메모 추가 중 오류가 발생했습니다.'
        }, status=500)

@require_http_methods(["POST"])
@login_required
def customer_priority_update(request, followup_id):
    """고객 우선순위 업데이트 API"""
    try:
        followup = get_object_or_404(FollowUp, pk=followup_id)
        
        # 권한 체크: 수정 권한이 있는 경우만 가능
        if not can_modify_user_data(request.user, followup.user):
            return JsonResponse({
                'success': False,
                'error': '고객 정보를 수정할 권한이 없습니다.'
            }, status=403)
        
        new_priority = request.POST.get('priority')
        
        # 유효한 우선순위인지 확인 (모델의 PRIORITY_CHOICES 사용)
        valid_priorities = [choice[0] for choice in FollowUp.PRIORITY_CHOICES]
        if new_priority not in valid_priorities:
            return JsonResponse({
                'success': False,
                'error': '유효하지 않은 우선순위입니다.'
            }, status=400)
        
        # 우선순위 업데이트
        followup.priority = new_priority
        followup.save()
        
        # 응답에 포함할 우선순위 표시명
        priority_display = dict(FollowUp.PRIORITY_CHOICES).get(new_priority, new_priority)
        
        return JsonResponse({
            'success': True,
            'message': f'고객 우선순위가 {priority_display}로 변경되었습니다.',
            'priority': new_priority,
            'priority_display': priority_display
        })
        
    except Exception as e:
        logger.error(f"Customer priority update error: {e}")
        return JsonResponse({
            'success': False,
            'error': '우선순위 업데이트 중 오류가 발생했습니다.'
        }, status=500)


@login_required
def schedule_delivery_items_api(request, schedule_id):
    """Schedule의 DeliveryItem 정보를 가져오는 API"""
    try:
        schedule = get_object_or_404(Schedule, pk=schedule_id)
        
        # 권한 체크: Schedule의 followup을 통해 사용자 확인
        if schedule.followup and schedule.followup.user:
            if not can_access_user_data(request.user, schedule.followup.user):
                return JsonResponse({
                    'success': False,
                    'error': '접근 권한이 없습니다.'
                }, status=403)
        
        # 연결된 History가 있는지 확인 (History 기준 세금계산서 상태 적용을 위해)
        related_history = None
        try:
            related_history = History.objects.filter(schedule=schedule).first()  # 이 Schedule에 연결된 첫 번째 History
        except:
            pass
        
        # DeliveryItem 정보 가져오기
        delivery_items = schedule.delivery_items_set.all().order_by('id')
        
        items_data = []
        for item in delivery_items:
            item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
            
            # History가 있으면 History 기준, 없으면 Schedule DeliveryItem 기준
            tax_invoice_status = related_history.tax_invoice_issued if related_history else item.tax_invoice_issued
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit_price': float(item.unit_price),
                'total_price': float(item_total),
                'tax_invoice_issued': tax_invoice_status,
            })
        
        return JsonResponse({
            'success': True,
            'items': items_data,
            'schedule_id': schedule.id,
            'visit_date': schedule.visit_date.strftime('%Y-%m-%d'),
            'has_related_history': related_history is not None,  # History 연결 여부
            'history_tax_status': related_history.tax_invoice_issued if related_history else None,  # History 세금계산서 상태
        })
        
    except Exception as e:
        logger.error(f"Schedule delivery items API error: {e}")
        return JsonResponse({
            'success': False,
            'error': 'DeliveryItem 정보를 가져오는 중 오류가 발생했습니다.'
        }, status=500)


@login_required
def history_delivery_items_api(request, history_id):
    """History의 DeliveryItem 정보를 가져오는 API"""
    import re
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        history = get_object_or_404(History, pk=history_id)
        
        # 권한 체크: 해당 활동기록을 볼 수 있는 권한이 있는지 확인
        if not can_access_user_data(request.user, history.user):
            return JsonResponse({
                'success': False,
                'error': '접근 권한이 없습니다.'
            }, status=403)
        
        # DeliveryItem 정보 가져오기
        delivery_items = history.delivery_items_set.all().order_by('id')
        
        items_data = []
        has_history_items = False
        has_schedule_items = False
        
        # 1. History DeliveryItem 모델이 있는 경우
        if delivery_items.exists():
            has_history_items = True
            for item in delivery_items:
                item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
                # History의 세금계산서 상태를 기준으로 함 (동기화)
                items_data.append({
                    'id': item.id,
                    'item_name': item.item_name,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price),
                    'total_price': float(item_total),
                    'tax_invoice_issued': history.tax_invoice_issued,  # History 기준으로 강제 설정
                    'source': 'history'  # 출처 표시
                })
        else:
            pass
        
        # 2. History DeliveryItem이 없지만 기존 텍스트 데이터가 있는 경우 (fallback)
        if not has_history_items and history.delivery_items and history.delivery_items.strip():
            has_history_items = True
            # 기존 텍스트 데이터 파싱
            delivery_text = history.delivery_items.strip()
            
            # 줄바꿈으로 분리하여 각 라인 처리
            lines = delivery_text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
            
            # 만약 줄바꿈이 문자열로 저장되어 있다면 \\n으로도 분리 시도
            if len(lines) == 1 and '\\n' in delivery_text:
                lines = delivery_text.split('\\n')
            
            # 그래도 하나의 라인이면, 품목 패턴을 찾아서 분리
            if len(lines) == 1:
                # 정규식으로 품목 패턴을 모두 찾기
                pattern = r'([A-Z0-9.]+:\s*\d+(?:\.\d+)?개\s*\([0-9,]+원\))'
                matches = re.findall(pattern, delivery_text)
                if len(matches) > 1:
                    lines = matches
            
            for i, line in enumerate(lines):
                line = line.strip()
                
                if not line:
                    continue
                
                # "품목명: 수량개 (금액원)" 패턴 파싱
                match = re.match(r'^(.+?):\s*(\d+(?:\.\d+)?)개\s*\((.+?)원\)$', line)
                if match:
                    item_name = match.group(1).strip()
                    quantity = float(match.group(2))
                    amount_str = match.group(3).replace(',', '').replace(' ', '')
                    
                    try:
                        total_amount = float(amount_str)
                        # 부가세 포함 금액에서 단가 역산 (부가세 포함 / 수량)
                        unit_price = total_amount / quantity if quantity > 0 else 0
                    except ValueError as e:
                        logger.error(f"[HISTORY_DELIVERY_API] 금액 파싱 실패: {e}")
                        total_amount = 0
                        unit_price = 0
                    
                    items_data.append({
                        'id': f'text_{len(items_data)}',  # 임시 ID
                        'item_name': item_name,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'total_price': total_amount,
                        'tax_invoice_issued': history.tax_invoice_issued,  # History 기준
                        'source': 'history_text'  # 출처 표시
                    })
                else:
                    # 패턴에 맞지 않는 경우, 전체를 품목명으로 처리
                    items_data.append({
                        'id': f'text_{len(items_data)}',  # 임시 ID
                        'item_name': line,
                        'quantity': 1,
                        'unit_price': 0,
                        'total_price': 0,
                        'tax_invoice_issued': history.tax_invoice_issued,  # History 기준
                        'source': 'history_text'  # 출처 표시
                    })
        
        # 3. 연결된 Schedule의 DeliveryItem도 항상 확인 (History 기준 세금계산서 상태 적용)
        # History에 DeliveryItem이나 텍스트가 없어도 Schedule DeliveryItem은 항상 확인
        if history.schedule:
            schedule_items = history.schedule.delivery_items_set.all().order_by('id')
            
            if schedule_items.exists():
                has_schedule_items = True
                for item in schedule_items:
                    item_total = item.total_price or (item.quantity * item.unit_price * 1.1)
                    
                    items_data.append({
                        'id': f'schedule_{item.id}',
                        'item_name': item.item_name,
                        'quantity': item.quantity,
                        'unit_price': float(item.unit_price),
                        'total_price': float(item_total),
                        'tax_invoice_issued': history.tax_invoice_issued,  # History 기준으로 강제 설정
                        'source': 'schedule'  # 출처 표시
                    })
        else:
            pass
        
        return JsonResponse({
            'success': True,
            'items': items_data,
            'history_id': history.id,
            'delivery_date': history.delivery_date.strftime('%Y-%m-%d') if history.delivery_date else '',
            'is_legacy_data': not has_history_items and bool(history.delivery_items),  # 레거시 데이터 여부
            'has_history_items': has_history_items,
            'has_schedule_items': has_schedule_items,
            'tax_invoice_status': history.tax_invoice_issued,  # History 기준 세금계산서 상태
        })
        
    except Exception as e:
        logger.error(f"[HISTORY_DELIVERY_API] API 오류: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'DeliveryItem 정보를 가져오는 중 오류가 발생했습니다.'
        }, status=500)

@login_required
def customer_detail_report_view_simple(request, followup_id):
    """부서 기준 상세 활동 리포트 (고객 클릭 시 해당 부서 전체 기록 표시)"""
    from django.db.models import Count, Sum, Q
    from datetime import datetime, timedelta
    from django.contrib.auth.models import User
    import json
    
    # 권한 확인 및 팔로우업 조회
    try:
        followup = FollowUp.objects.select_related('user', 'company', 'department').get(id=followup_id)
        
        # 권한 체크 (같은 회사 소속이면 접근 가능)
        if not can_access_followup(request.user, followup):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:customer_report')
            
    except FollowUp.DoesNotExist:
        messages.error(request, '해당 고객 정보를 찾을 수 없습니다.')
        return redirect('reporting:customer_report')
    
    # 부서 정보 가져오기 (이제 부서 기준으로 조회)
    department = followup.department
    company = followup.company
    
    # 해당 부서의 모든 고객(팔로우업) 목록
    department_followups = FollowUp.objects.filter(department=department)
    department_customers = list(department_followups.values_list('id', flat=True))
    
    # 본인 고객인지 확인
    is_own_customer = (request.user == followup.user)
    user_profile = get_user_profile(request.user)
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용) - 매니저 제외
    company_users = []
    if user_profile and user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    target_user = request.user  # 기본값
    
    if data_filter == 'all' and user_profile and user_profile.company:
        # 같은 회사 전체 (salesman만)
        filter_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        )
        target_user = None  # 전체
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        # 특정 직원 (같은 회사 확인)
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True
            )
            filter_users = User.objects.filter(id=selected_filter_user.id)
            target_user = selected_filter_user
        except User.DoesNotExist:
            filter_users = User.objects.filter(id=request.user.id)
            target_user = request.user
            data_filter = 'me'
    else:
        # 'me' - 본인만
        filter_users = User.objects.filter(id=request.user.id)
        target_user = request.user
    
    # 부서 전체 History 데이터 조회 (필터 적용)
    if target_user is None:
        # 전체: 같은 회사 모든 salesman
        histories = History.objects.filter(
            followup__department=department,
            user__in=filter_users
        ).select_related('followup', 'user').order_by('-created_at')
    else:
        # 특정 사용자 (본인 또는 선택된 직원)
        histories = History.objects.filter(
            followup__department=department,
            user=target_user
        ).select_related('followup', 'user').order_by('-created_at')
    
    # 기본 통계 계산
    delivery_histories = histories.filter(action_type='delivery_schedule')
    meeting_histories = histories.filter(action_type='customer_meeting')
    
    total_amount = 0
    for history in delivery_histories:
        if history.delivery_amount:
            total_amount += float(history.delivery_amount)
    
    # 부서 전체 Schedule 기반 납품 일정 (필터 적용)
    if target_user is None:
        # 전체: 같은 회사 모든 salesman
        schedule_deliveries = Schedule.objects.filter(
            followup__department=department,
            activity_type='delivery',
            user__in=filter_users
        ).select_related('followup', 'user').order_by('-visit_date')
    else:
        # 특정 사용자 (본인 또는 선택된 직원)
        schedule_deliveries = Schedule.objects.filter(
            followup__department=department,
            activity_type='delivery',
            user=target_user
        ).select_related('followup', 'user').order_by('-visit_date')
    
    # 디버깅: 권한 및 데이터 확인
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[DEPT_REPORT] User: {request.user.username}, Department: {department.name if department else 'None'}")
    logger.info(f"[DEPT_REPORT] can_view_all: {user_profile.can_view_all_users()}, customers in dept: {len(department_customers)}")
    logger.info(f"[DEPT_REPORT] histories count: {histories.count()}, schedule_deliveries count: {schedule_deliveries.count()}")
    
    # 통합 납품 내역 생성 (템플릿 호환성을 위해)
    integrated_deliveries = []
    
    # History 기반 납품 내역
    for history in delivery_histories:
        # 이 History와 연결된 Schedule이 있고, 그 Schedule에 DeliveryItem이 있는지 확인
        has_schedule_items = False
        schedule_amount = 0
        if history.schedule:
            schedule_delivery_items = history.schedule.delivery_items_set.all()
            if schedule_delivery_items.exists():
                has_schedule_items = True
                # Schedule의 DeliveryItem 금액 계산
                for item in schedule_delivery_items:
                    if item.total_price:
                        schedule_amount += float(item.total_price)
        
        # History의 세금계산서 상태 정보 계산
        history_delivery_items = history.delivery_items_set.all()
        history_tax_status = {
            'has_items': history_delivery_items.exists(),
            'total_count': history_delivery_items.count(),
            'issued_count': history_delivery_items.filter(tax_invoice_issued=True).count(),
            'pending_count': history_delivery_items.filter(tax_invoice_issued=False).count(),
        }
        history_tax_status['all_issued'] = (history_tax_status['total_count'] > 0 and 
                                          history_tax_status['issued_count'] == history_tax_status['total_count'])
        history_tax_status['none_issued'] = (history_tax_status['total_count'] > 0 and 
                                            history_tax_status['issued_count'] == 0)
        
        # has_schedule_items인 경우 Schedule 금액만 사용, 아닌 경우 History 금액 사용
        if has_schedule_items:
            final_amount = schedule_amount
        else:
            final_amount = float(history.delivery_amount) if history.delivery_amount else 0
        
        # 고객 정보 추가 (부서 기준 조회이므로 어떤 고객인지 표시)
        customer_name = history.followup.customer_name if history.followup else '미정'
        
        delivery_data = {
            'type': 'history',
            'id': history.id,
            'date': (history.delivery_date or history.created_at.date()).strftime('%Y-%m-%d'),
            'schedule_id': history.schedule_id if history.schedule else None,
            'scheduleId': history.schedule_id if history.schedule else None,  # JavaScript 호환
            'items_display': history.delivery_items or None,
            'items': history.delivery_items or '',  # JavaScript에서 사용
            'amount': final_amount,
            'tax_invoice_issued': history.tax_invoice_issued,
            'content': history.content or '',
            'user': history.user.username,
            'customer_name': customer_name,  # 고객명 추가
            'followup_id': history.followup.id if history.followup else None,  # 고객 ID 추가
            'has_schedule_items': has_schedule_items,
            'history_tax_status': history_tax_status,  # 세금계산서 상태 정보 추가
        }
        integrated_deliveries.append(delivery_data)
    
    # Schedule 기반 납품 일정 추가
    for schedule in schedule_deliveries:
        # 해당 Schedule과 연결된 History 찾기
        related_history = delivery_histories.filter(schedule=schedule).first()
        
        if related_history:
            # Schedule과 연결된 History가 있으면 History 데이터 우선 (이미 위에서 처리됨)
            continue
        else:
            # History에 연결되지 않은 Schedule - 일정 기반으로 표시
            # 하지만 혹시 다른 History에서 이 Schedule을 참조하는지 확인
            schedule_amount = 0
            schedule_items = '일정 기반 (품목 미확정)'
            
            # 이 Schedule을 참조하는 History 검색 (권한에 따라 필터링)
            if user_profile.can_view_all_users():
                all_related_histories = History.objects.filter(schedule=schedule)
            else:
                # 동료 고객: 본인이 작성한 히스토리만
                all_related_histories = History.objects.filter(schedule=schedule, user=request.user)
            
            for hist in all_related_histories:
                if hist.delivery_amount:
                    schedule_amount += float(hist.delivery_amount)
                if hist.delivery_items:
                    schedule_items = hist.delivery_items
            
            # Schedule에 직접 연결된 DeliveryItem들에서 금액과 품목 정보 가져오기
            schedule_delivery_items = schedule.delivery_items_set.all()
            
            if schedule_delivery_items.exists():
                item_names = []
                delivery_item_amount = 0
                
                for item in schedule_delivery_items:
                    if item.total_price:
                        delivery_item_amount += float(item.total_price)
                    item_names.append(f"{item.item_name}: {item.quantity}개")
                
                if delivery_item_amount > 0:
                    schedule_amount += delivery_item_amount
                
                if item_names:
                    schedule_items = ' / '.join(item_names)
            
            # Schedule의 세금계산서 상태 정보 계산
            schedule_tax_status = {
                'has_items': schedule_delivery_items.exists(),
                'total_count': schedule_delivery_items.count(),
                'issued_count': schedule_delivery_items.filter(tax_invoice_issued=True).count(),
                'pending_count': schedule_delivery_items.filter(tax_invoice_issued=False).count(),
            }
            schedule_tax_status['all_issued'] = (schedule_tax_status['total_count'] > 0 and 
                                              schedule_tax_status['issued_count'] == schedule_tax_status['total_count'])
            schedule_tax_status['none_issued'] = (schedule_tax_status['total_count'] > 0 and 
                                                schedule_tax_status['issued_count'] == 0)
            
            # 고객 정보 추가 (부서 기준 조회이므로 어떤 고객인지 표시)
            customer_name = schedule.followup.customer_name if schedule.followup else '미정'
            
            delivery_data = {
                'type': 'schedule_only',
                'id': schedule.id,
                'date': schedule.visit_date.strftime('%Y-%m-%d') if schedule.visit_date else '',
                'schedule_id': schedule.id,
                'scheduleId': schedule.id,  # JavaScript 호환
                'items_display': schedule_items,
                'items': schedule_items if schedule_items != '일정 기반 (품목 미확정)' else '',  # JavaScript에서 사용
                'amount': schedule_amount,
                'scheduleAmount': schedule_amount,  # JavaScript 호환
                'tax_invoice_issued': False,
                'content': schedule.notes or '예정된 납품 일정',
                'user': schedule.user.username,
                'customer_name': customer_name,  # 고객명 추가
                'followup_id': schedule.followup.id if schedule.followup else None,  # 고객 ID 추가
                'has_schedule_items': True,
                'schedule_tax_status': schedule_tax_status,  # 세금계산서 상태 정보 추가
            }
            integrated_deliveries.append(delivery_data)
    
    # 날짜순 정렬
    integrated_deliveries.sort(key=lambda x: x['date'], reverse=True)

    # JSON 직렬화
    try:
        integrated_deliveries_json = json.dumps(integrated_deliveries, ensure_ascii=False, cls=DjangoJSONEncoder)
    except Exception as e:
        integrated_deliveries_json = "[]"

    # 통합 데이터 기반 통계 계산
    integrated_total_amount = 0
    integrated_tax_issued = 0
    integrated_tax_pending = 0
    
    for delivery in integrated_deliveries:
        # 금액 계산
        if delivery['type'] == 'schedule_only':
            amount = delivery.get('scheduleAmount', 0)
        else:  # history
            amount = delivery.get('amount', 0)
        integrated_total_amount += amount
        
        # 세금계산서 상태 계산 (건별로 계산)
        delivery_has_issued_items = False
        delivery_has_pending_items = False
        
        if delivery['type'] == 'history':
            tax_status = delivery.get('history_tax_status', {})
            if tax_status.get('has_items', False):
                # DeliveryItem이 있는 경우 - 하나라도 발행된 것이 있으면 발행건으로 간주
                if tax_status.get('all_issued', False):
                    delivery_has_issued_items = True
                elif tax_status.get('none_issued', False):
                    delivery_has_pending_items = True
                else:
                    # 일부만 발행된 경우 - 혼재 상태이므로 발행건으로 간주
                    delivery_has_issued_items = True
            else:
                # 단순 History인 경우
                if delivery.get('tax_invoice_issued', False):
                    delivery_has_issued_items = True
                else:
                    delivery_has_pending_items = True
        else:  # schedule_only
            tax_status = delivery.get('schedule_tax_status', {})
            if tax_status.get('total_count', 0) > 0:
                if tax_status.get('all_issued', False):
                    delivery_has_issued_items = True
                elif tax_status.get('none_issued', False):
                    delivery_has_pending_items = True
                else:
                    # 일부만 발행된 경우 - 혼재 상태이므로 발행건으로 간주
                    delivery_has_issued_items = True
            # 품목이 없는 경우는 미발행으로 간주
            else:
                delivery_has_pending_items = True
        
        # 건별로 카운트
        if delivery_has_issued_items:
            integrated_tax_issued += 1
        elif delivery_has_pending_items:
            integrated_tax_pending += 1

    # 선결제 통계 계산 - 해당 부서의 모든 고객에 등록된 선결제
    prepayments = Prepayment.objects.filter(
        customer__department=department
    )
    
    prepayment_total = prepayments.aggregate(total=Sum('amount'))['total'] or 0
    prepayment_balance = prepayments.aggregate(total=Sum('balance'))['total'] or 0
    prepayment_count = prepayments.count()
    
    # 월별 활동 트렌드 데이터 계산 (최근 12개월) - 부서 기준 + 필터 적용
    from dateutil.relativedelta import relativedelta
    from django.utils import timezone
    
    chart_labels = []
    chart_meetings = []
    chart_deliveries = []
    chart_amounts = []
    
    today = timezone.now().date()
    
    for i in range(11, -1, -1):
        # 해당 월의 시작일과 종료일 계산
        target_date = today - relativedelta(months=i)
        month_start = target_date.replace(day=1)
        if i == 0:
            month_end = today
        else:
            month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
        
        chart_labels.append(f"{target_date.month}월")
        
        # 해당 월의 미팅 횟수 (History 기반) - 부서 전체 + 필터 적용
        if target_user is None:
            month_meetings = History.objects.filter(
                followup__department=department,
                action_type='customer_meeting',
                user__in=filter_users,
                created_at__date__gte=month_start,
                created_at__date__lte=month_end
            ).count()
        else:
            month_meetings = History.objects.filter(
                followup__department=department,
                action_type='customer_meeting',
                user=target_user,
                created_at__date__gte=month_start,
                created_at__date__lte=month_end
            ).count()
        chart_meetings.append(month_meetings)
        
        # 해당 월의 납품 건수와 금액
        month_delivery_count = 0
        month_delivery_amount = 0
        
        for delivery in integrated_deliveries:
            try:
                delivery_date = datetime.strptime(delivery['date'], '%Y-%m-%d').date()
                if month_start <= delivery_date <= month_end:
                    month_delivery_count += 1
                    month_delivery_amount += delivery.get('amount', 0) or 0
            except (ValueError, TypeError):
                continue
        
        chart_deliveries.append(month_delivery_count)
        chart_amounts.append(int(month_delivery_amount))

    context = {
        'followup': followup,  # 클릭한 고객 (참조용)
        'department': department,  # 부서 정보
        'company': company,  # 회사 정보
        'department_followups': department_followups,  # 해당 부서의 모든 고객 목록
        'histories': histories,
        'total_amount': integrated_total_amount,
        'total_meetings': meeting_histories.count(),
        'total_deliveries': len(integrated_deliveries),
        'tax_invoices_issued': integrated_tax_issued,
        'tax_invoices_pending': integrated_tax_pending,
        'prepayment_total': prepayment_total,  # 선결제 총액
        'prepayment_balance': prepayment_balance,  # 선결제 잔액
        'prepayment_count': prepayment_count,  # 선결제 건수
        'chart_labels': json.dumps(chart_labels, ensure_ascii=False),
        'chart_meetings': json.dumps(chart_meetings, ensure_ascii=False),
        'chart_deliveries': json.dumps(chart_deliveries, ensure_ascii=False),
        'chart_amounts': json.dumps(chart_amounts, ensure_ascii=False),
        'delivery_histories': delivery_histories,
        'schedule_deliveries': schedule_deliveries,
        'integrated_deliveries': integrated_deliveries,
        'meeting_histories': meeting_histories,
        'chart_data': {
            'labels': json.dumps(chart_labels, ensure_ascii=False),
            'meetings': json.dumps(chart_meetings, ensure_ascii=False),
            'deliveries': json.dumps(chart_deliveries, ensure_ascii=False),
            'amounts': json.dumps(chart_amounts, ensure_ascii=False),
        },
        # 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'selected_filter_user': selected_filter_user,
        'company_users': company_users,
        'page_title': f'{company.name} - {department.name}'  # 회사명 - 부서명으로 변경
    }
    
    return render(request, 'reporting/customer_detail_report.html', context)

@login_required
@require_POST
def followup_create_ajax(request):
    """AJAX로 팔로우업을 생성하는 뷰"""
    try:
        # Manager는 데이터 생성 불가 (뷰어 권한)
        _ajax_profile = get_user_profile(request.user)
        if _ajax_profile.is_manager():
            return JsonResponse({'success': False, 'error': '권한이 없습니다. Manager는 데이터를 생성할 수 없습니다.'}, status=403)

        # 필수 필드 검증
        customer_name = request.POST.get('customer_name', '').strip()
        company_id = request.POST.get('company', '').strip()
        department_id = request.POST.get('department', '').strip()
        priority = request.POST.get('priority', '').strip()
        
        if not customer_name:
            return JsonResponse({
                'success': False,
                'error': '고객명을 입력해주세요.'
            })
        
        if not company_id:
            return JsonResponse({
                'success': False,
                'error': '업체/학교를 선택해주세요.'
            })
            
        if not department_id:
            return JsonResponse({
                'success': False,
                'error': '부서/연구실을 선택해주세요.'
            })
        
        if not priority:
            return JsonResponse({
                'success': False,
                'error': '우선순위를 선택해주세요.'
            })
        
        # Company와 Department 객체 가져오기
        try:
            company = Company.objects.get(id=company_id)
            department = Department.objects.get(id=department_id, company=company)
        except (Company.DoesNotExist, Department.DoesNotExist):
            return JsonResponse({
                'success': False,
                'error': '선택한 업체 또는 부서가 존재하지 않습니다.'
            })
        
        # 권한 체크: 같은 회사에서 생성된 업체인지 확인
        user_profile_obj = getattr(request.user, 'userprofile', None)
        is_admin = getattr(request, 'is_admin', False) or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'admin')
        
        import logging
        logger = logging.getLogger(__name__)
        
        if not is_admin:
            if user_profile_obj and user_profile_obj.company:
                same_company_users = User.objects.filter(userprofile__company=user_profile_obj.company)
                if company.created_by not in same_company_users:
                    return JsonResponse({
                        'success': False,
                        'error': '접근 권한이 없는 업체입니다.'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'error': '회사 정보가 없어 팔로우업을 생성할 수 없습니다.'
                })
        
        # 중복 체크 (같은 고객명, 회사, 부서)
        existing_followup = FollowUp.objects.filter(
            customer_name=customer_name,
            company=company,
            department=department,
            user=request.user
        ).first()
        
        if existing_followup:
            return JsonResponse({
                'success': False,
                'error': '이미 동일한 팔로우업이 존재합니다.'
            })
        
        # 팔로우업 생성
        followup = FollowUp.objects.create(
            user=request.user,
            customer_name=customer_name,
            company=company,
            department=department,
            manager=request.POST.get('manager', '').strip(),
            phone_number=request.POST.get('phone_number', '').strip(),
            email=request.POST.get('email', '').strip(),
            priority=priority,  # 요청에서 받은 우선순위 사용
            address=request.POST.get('address', '').strip(),     # 상세주소 추가
            notes=request.POST.get('notes', '').strip(),
            status='active'
        )
        
        # 사용자 회사 정보 설정
        if user_profile_obj and user_profile_obj.company:
            followup.user_company = user_profile_obj.company
            followup.save(update_fields=['user_company'])
        
        # 성공 응답
        followup_text = f"{followup.customer_name} ({followup.company.name} - {followup.department.name})"
        
        return JsonResponse({
            'success': True,
            'followup_id': followup.id,
            'followup_text': followup_text,
            'message': '팔로우업이 성공적으로 생성되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"AJAX 팔로우업 생성 오류: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': '서버 오류가 발생했습니다. 다시 시도해주세요.'
        }, status=500)

@login_required
def department_list_ajax(request, company_id):
    """특정 업체의 부서 목록을 AJAX로 반환"""
    try:
        company = get_object_or_404(Company, id=company_id)
        
        # 권한 체크: 같은 회사에서 생성된 업체인지 확인
        user_profile_obj = getattr(request.user, 'userprofile', None)
        if user_profile_obj and user_profile_obj.company:
            same_company_users = User.objects.filter(userprofile__company=user_profile_obj.company)
            if company.created_by not in same_company_users:
                return JsonResponse({
                    'error': '접근 권한이 없는 업체입니다.'
                }, status=403)
        
        departments = Department.objects.filter(company=company).values('id', 'name')
        departments_list = [{'id': dept['id'], 'name': dept['name']} for dept in departments]
        
        return JsonResponse(departments_list, safe=False)
        
    except Exception as e:
        logger.error(f"부서 목록 AJAX 조회 오류: {str(e)}")
        return JsonResponse({
            'error': '부서 목록을 가져오는 중 오류가 발생했습니다.'
        }, status=500)

@login_required
@role_required(['admin', 'salesman'])
@require_POST
def update_tax_invoice_status(request):
    """세금계산서 상태 업데이트 API"""
    try:
        data = json.loads(request.body)
        delivery_type = data.get('type')
        delivery_id = data.get('id')
        tax_invoice_issued = data.get('tax_invoice_issued')
        
        if delivery_type == 'history':
            # History 레코드의 세금계산서 상태 업데이트
            try:
                history = History.objects.get(id=delivery_id)
                
                # 권한 체크
                if not can_access_user_data(request.user, history.user):
                    return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                # 하나과학이 아닌 경우 같은 회사 체크
                if not getattr(request, 'is_hanagwahak', False):
                    user_profile_obj = getattr(request.user, 'userprofile', None)
                    history_user_profile = getattr(history.user, 'userprofile', None)
                    if (user_profile_obj and user_profile_obj.company and 
                        history_user_profile and history_user_profile.company and
                        user_profile_obj.company != history_user_profile.company):
                        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                history.tax_invoice_issued = tax_invoice_issued
                history.save()
                
                # History에 직접 연결된 DeliveryItem들도 함께 업데이트
                history_delivery_items = history.delivery_items_set.all()
                if history_delivery_items.exists():
                    updated_count = history_delivery_items.update(tax_invoice_issued=tax_invoice_issued)
                
                # History와 연결된 Schedule의 모든 DeliveryItem도 함께 업데이트
                if history.schedule:
                    schedule_updated_count = DeliveryItem.objects.filter(schedule=history.schedule).update(
                        tax_invoice_issued=tax_invoice_issued
                    )
                
                return JsonResponse({
                    'success': True,
                    'message': '세금계산서 상태가 업데이트되었습니다.'
                })
                
            except History.DoesNotExist:
                return JsonResponse({'error': '해당 납품 기록을 찾을 수 없습니다.'}, status=404)
                
        elif delivery_type == 'delivery_item':
            # DeliveryItem의 세금계산서 상태 업데이트
            try:
                delivery_item = DeliveryItem.objects.get(id=delivery_id)
                
                # 권한 체크
                if not can_access_user_data(request.user, delivery_item.schedule.user):
                    return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                # 하나과학이 아닌 경우 같은 회사 체크
                if not getattr(request, 'is_hanagwahak', False):
                    user_profile_obj = getattr(request.user, 'userprofile', None)
                    item_user_profile = getattr(delivery_item.schedule.user, 'userprofile', None)
                    if (user_profile_obj and user_profile_obj.company and 
                        item_user_profile and item_user_profile.company and
                        user_profile_obj.company != item_user_profile.company):
                        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                delivery_item.tax_invoice_issued = tax_invoice_issued
                delivery_item.save()
                
                return JsonResponse({
                    'success': True,
                    'message': '세금계산서 상태가 업데이트되었습니다.'
                })
                
            except DeliveryItem.DoesNotExist:
                return JsonResponse({'error': '해당 납품 품목을 찾을 수 없습니다.'}, status=404)
                
        elif delivery_type == 'schedule_bulk':
            # Schedule의 모든 DeliveryItem 일괄 업데이트
            try:
                schedule = Schedule.objects.get(id=delivery_id)
                
                # 권한 체크
                if not can_access_user_data(request.user, schedule.user):
                    return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                # 하나과학이 아닌 경우 같은 회사 체크
                if not getattr(request, 'is_hanagwahak', False):
                    user_profile_obj = getattr(request.user, 'userprofile', None)
                    schedule_user_profile = getattr(schedule.user, 'userprofile', None)
                    if (user_profile_obj and user_profile_obj.company and 
                        schedule_user_profile and schedule_user_profile.company and
                        user_profile_obj.company != schedule_user_profile.company):
                        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
                
                # Schedule의 모든 DeliveryItem 업데이트
                updated_count = DeliveryItem.objects.filter(schedule=schedule).update(
                    tax_invoice_issued=tax_invoice_issued
                )
                
                # 연결된 History도 함께 업데이트
                History.objects.filter(schedule=schedule).update(
                    tax_invoice_issued=tax_invoice_issued
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'세금계산서 상태가 업데이트되었습니다. (품목 {updated_count}개)',
                    'updated_count': updated_count
                })
                
            except Schedule.DoesNotExist:
                return JsonResponse({'error': '해당 일정을 찾을 수 없습니다.'}, status=404)
        
        else:
            return JsonResponse({'error': '잘못된 요청입니다.'}, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': '잘못된 JSON 형식입니다.'}, status=400)
    except Exception as e:
        logger.error(f"세금계산서 상태 업데이트 오류: {str(e)}")
        return JsonResponse({'error': '서버 오류가 발생했습니다.'}, status=500)

@login_required
@role_required(['admin', 'salesman', 'manager'])
def schedule_delivery_items_api(request, schedule_id):
    """Schedule의 DeliveryItem 정보를 반환하는 API"""
    try:
        schedule = Schedule.objects.get(id=schedule_id)
        
        # 권한 체크: Schedule의 followup을 통해 사용자 확인
        if schedule.followup and schedule.followup.user:
            if not can_access_user_data(request.user, schedule.followup.user):
                return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
            
            # Manager는 회사 체크를 건너뜀 (모든 데이터 조회 가능)
            user_profile = get_user_profile(request.user)
            if not user_profile.is_manager():
                # 하나과학이 아닌 경우 같은 회사 체크
                if not getattr(request, 'is_hanagwahak', False):
                    user_profile_obj = getattr(request.user, 'userprofile', None)
                    schedule_user_profile = getattr(schedule.followup.user, 'userprofile', None)
                    if (user_profile_obj and user_profile_obj.company and 
                        schedule_user_profile and schedule_user_profile.company and
                        user_profile_obj.company != schedule_user_profile.company):
                        return JsonResponse({'error': '접근 권한이 없습니다.'}, status=403)
        
        # DeliveryItem 정보 가져오기
        items = []
        for item in schedule.delivery_items_set.all():
            items.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price),
                'tax_invoice_issued': item.tax_invoice_issued
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'schedule_id': schedule.id,
            'visit_date': schedule.visit_date.strftime('%Y-%m-%d')
        })
        
    except Schedule.DoesNotExist:
        return JsonResponse({'error': '해당 일정을 찾을 수 없습니다.'}, status=404)
    except Exception as e:
        logger.error(f"Schedule 납품 품목 API 오류: {str(e)}")
        return JsonResponse({'error': '서버 오류가 발생했습니다.'}, status=500)


# ============ Admin 전용 API 뷰들 ============

@role_required(['admin'])
@require_http_methods(["GET"])
def api_users_list(request):
    """사용자 목록 API (Admin 전용)"""
    try:
        users = User.objects.select_related('userprofile', 'userprofile__company').all()
        
        users_data = []
        for user in users:
            user_data = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'company': None
            }
            
            if hasattr(user, 'userprofile') and user.userprofile.company:
                user_data['company'] = user.userprofile.company.name
            
            users_data.append(user_data)
        
        return JsonResponse({
            'success': True,
            'users': users_data
        })
        
    except Exception as e:
        logger.error(f"사용자 목록 API 에러: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@role_required(['admin'])
@require_http_methods(["POST"])
def api_change_company_creator(request):
    """업체 생성자 변경 API (Admin 전용)"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        company_id = request.POST.get('company_id')
        new_creator_id = request.POST.get('new_creator_id')
        
        if not company_id or not new_creator_id:
            return JsonResponse({
                'success': False,
                'message': '필수 파라미터가 누락되었습니다.'
            }, status=400)
        
        # 업체 조회
        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '존재하지 않는 업체입니다.'
            }, status=404)
        
        # 새 생성자 조회
        try:
            new_creator = User.objects.select_related('userprofile', 'userprofile__company').get(id=new_creator_id)
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '존재하지 않는 사용자입니다.'
            }, status=404)
        
        # 기존 생성자 정보 로깅
        old_creator = company.created_by
        old_creator_info = 'None'
        if old_creator:
            old_creator_company = 'Unknown'
            if hasattr(old_creator, 'userprofile') and old_creator.userprofile.company:
                old_creator_company = old_creator.userprofile.company.name
            old_creator_info = f"{old_creator.username} ({old_creator_company})"
        
        # 새 생성자 정보
        new_creator_company = 'Unknown'
        if hasattr(new_creator, 'userprofile') and new_creator.userprofile.company:
            new_creator_company = new_creator.userprofile.company.name
        new_creator_info = f"{new_creator.username} ({new_creator_company})"
        
        # 업체 생성자 변경
        company.created_by = new_creator
        company.save()
        
        # 응답 데이터
        response_data = {
            'success': True,
            'message': '업체 생성자가 성공적으로 변경되었습니다.',
            'new_creator': {
                'id': new_creator.id,
                'username': new_creator.username,
                'company': new_creator_company
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"업체 생성자 변경 API 에러: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'서버 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_company_departments(request, company_id):
    """업체별 부서 목록 API"""
    try:
        # 업체 조회 및 접근 권한 확인
        company = get_object_or_404(Company, pk=company_id)
        
        # 접근 권한 확인
        same_company_users = User.objects.filter(
            userprofile__company=request.user.userprofile.company
        )
        
        # 업체가 사용자의 회사와 연결되어 있는지 확인
        if company.created_by not in same_company_users:
            return JsonResponse({'success': False, 'message': '접근 권한이 없습니다.'}, status=403)
        
        # 해당 업체의 부서 목록 조회
        departments = Department.objects.filter(
            company=company
        ).order_by('name')
        
        departments_data = []
        for dept in departments:
            # 각 부서별 고객 수 직접 계산
            followup_count = FollowUp.objects.filter(department=dept).count()
            
            dept_data = {
                'id': dept.id,
                'name': dept.name,
                'followup_count': followup_count,
                'created_date': dept.created_at.strftime('%Y-%m-%d'),
                'created_by': dept.created_by.username if dept.created_by else '정보 없음'
            }
            departments_data.append(dept_data)
        
        return JsonResponse({
            'success': True,
            'company_name': company.name,
            'departments': departments_data,
            'total_count': len(departments_data)
        })
        
    except Exception as e:
        logger.error(f"업체별 부서 목록 API 에러: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_company_customers(request, company_id):
    """업체별 고객 정보 목록 API"""
    try:
        # 업체 조회 및 접근 권한 확인
        company = get_object_or_404(Company, pk=company_id)
        
        # 접근 권한 확인
        same_company_users = User.objects.filter(
            userprofile__company=request.user.userprofile.company
        )
        
        # 업체가 사용자의 회사와 연결되어 있는지 확인
        if company.created_by not in same_company_users:
            return JsonResponse({'success': False, 'message': '접근 권한이 없습니다.'}, status=403)
        
        # 해당 업체의 고객 정보들 조회
        followups = FollowUp.objects.filter(
            company=company
        ).select_related('department')
        
        customers_data = []
        for followup in followups:
            customer_data = {
                'id': followup.id,
                'customer_name': followup.customer_name or '고객명 미정',
                'phone': followup.phone_number or '-',
                'email': followup.email or '-',
                'position': followup.manager or '-',
                'department_name': followup.department.name if followup.department else '-',
                'created_date': followup.created_at.strftime('%Y-%m-%d'),
                'last_contact': followup.updated_at.strftime('%Y-%m-%d') if followup.updated_at else '연락 없음'
            }
            customers_data.append(customer_data)
        
        return JsonResponse({
            'success': True,
            'company_name': company.name,
            'customers': customers_data,
            'total_count': len(customers_data)
        })
        
    except Exception as e:
        logger.error(f"업체별 고객 정보 API 에러: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============ 프로필 관리 뷰들 ============

@login_required
def profile_view(request):
    """사용자 프로필 조회"""
    context = {
        'page_title': '프로필 정보',
        'user': request.user,
    }
    return render(request, 'reporting/profile.html', context)


@login_required
def profile_edit_view(request):
    """사용자 프로필 수정"""
    from django.contrib.auth import update_session_auth_hash
    from django.contrib.auth.forms import PasswordChangeForm
    from django import forms
    
    class ProfileEditForm(forms.Form):
        username = forms.CharField(
            max_length=150,
            label="사용자명",
            widget=forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '사용자명을 입력하세요'
            })
        )
        first_name = forms.CharField(
            max_length=30,
            label="이름",
            required=False,
            widget=forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '이름을 입력하세요'
            })
        )
        last_name = forms.CharField(
            max_length=30,
            label="성",
            required=False,
            widget=forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '성을 입력하세요'
            })
        )
        email = forms.EmailField(
            label="이메일",
            required=False,
            widget=forms.EmailInput(attrs={
                'class': 'form-control',
                'placeholder': '이메일을 입력하세요'
            })
        )
        
        def __init__(self, user, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.user = user
            # 현재 값으로 초기화
            self.fields['username'].initial = user.username
            self.fields['first_name'].initial = user.first_name
            self.fields['last_name'].initial = user.last_name
            self.fields['email'].initial = user.email
            
        def clean_username(self):
            username = self.cleaned_data['username']
            # 현재 사용자가 아닌 다른 사용자가 같은 사용자명을 사용하는지 확인
            if User.objects.exclude(pk=self.user.pk).filter(username=username).exists():
                raise forms.ValidationError("이미 사용 중인 사용자명입니다.")
            return username
    
    if request.method == 'POST':
        if 'profile_submit' in request.POST:
            # 프로필 정보 수정
            profile_form = ProfileEditForm(request.user, request.POST)
            password_form = PasswordChangeForm(request.user)
            
            if profile_form.is_valid():
                user = request.user
                user.username = profile_form.cleaned_data['username']
                user.first_name = profile_form.cleaned_data['first_name']
                user.last_name = profile_form.cleaned_data['last_name']
                user.email = profile_form.cleaned_data['email']
                user.save()
                
                messages.success(request, '프로필 정보가 성공적으로 수정되었습니다.')
                return redirect('reporting:profile')
            else:
                messages.error(request, '프로필 정보 수정 중 오류가 발생했습니다.')
                
        elif 'password_submit' in request.POST:
            # 비밀번호 변경
            profile_form = ProfileEditForm(request.user)
            password_form = PasswordChangeForm(request.user, request.POST)
            
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # 세션 유지
                messages.success(request, '비밀번호가 성공적으로 변경되었습니다.')
                return redirect('reporting:profile')
            else:
                messages.error(request, '비밀번호 변경 중 오류가 발생했습니다.')
    else:
        profile_form = ProfileEditForm(request.user)
        password_form = PasswordChangeForm(request.user)
    
    context = {
        'page_title': '프로필 수정',
        'profile_form': profile_form,
        'password_form': password_form,
    }
    return render(request, 'reporting/profile_edit.html', context)


@login_required
def followup_quote_items_api(request, followup_id):
    """
    특정 팔로우업의 견적 품목을 가져오는 API
    납품 일정 생성 시 견적에서 품목을 불러오기 위해 사용
    같은 회사 소속이면 동료 고객의 견적도 불러올 수 있음
    """
    try:
        followup = get_object_or_404(FollowUp, pk=followup_id)
        
        # 권한 체크 (같은 회사 소속이면 접근 가능)
        if not can_access_followup(request.user, followup):
            return JsonResponse({
                'error': '접근 권한이 없습니다.'
            }, status=403)
        
        # 해당 팔로우업의 본인 견적 일정 조회 (납품되지 않은 것만)
        # 동료 고객이라도 본인이 작성한 견적만 불러올 수 있음
        quote_schedules = Schedule.objects.filter(
            followup=followup,
            activity_type='quote',
            user=request.user  # 본인이 작성한 견적만
        ).order_by('-visit_date', '-visit_time')
        
        if not quote_schedules.exists():
            return JsonResponse({
                'error': '이 고객에 대한 본인 작성 견적이 없습니다.'
            })
        
        # 모든 견적 정보 수집 (납품되지 않은 것만)
        from reporting.models import DeliveryItem
        quotes_data = []
        
        for quote_schedule in quote_schedules:
            logger.info(f"[QUOTE_ITEMS_API] Schedule ID: {quote_schedule.id}, visit_date: {quote_schedule.visit_date}")
            
            # 이미 납품된 견적인지 확인
            # 이 견적(Schedule)에서 직접 복사된 납품 일정이 있는지 확인
            has_delivery = Schedule.objects.filter(
                followup=followup,
                activity_type='delivery',
                notes__icontains=f'견적 ID {quote_schedule.id}'  # 납품 메모에 견적 ID가 포함되어 있는지
            ).exists()
            
            # 또는 DeliveryItem이 완전히 동일한 납품이 있는지 확인
            if not has_delivery:
                # 견적 품목 가져오기
                quote_items = DeliveryItem.objects.filter(schedule=quote_schedule)
                if quote_items.exists():
                    # 같은 품목 구성의 완료된 납품이 있는지 확인
                    for delivery_schedule in Schedule.objects.filter(
                        followup=followup,
                        activity_type='delivery',
                        status='completed'
                    ):
                        delivery_items = DeliveryItem.objects.filter(schedule=delivery_schedule)
                        # 품목 개수와 품목명이 모두 일치하면 이미 납품된 것으로 간주
                        if (delivery_items.count() == quote_items.count() and
                            set(delivery_items.values_list('item_name', flat=True)) == 
                            set(quote_items.values_list('item_name', flat=True))):
                            has_delivery = True
                            break
            
            if has_delivery:
                continue
            
            items = DeliveryItem.objects.filter(schedule=quote_schedule)
            
            if items.exists():
                items_data = [{
                    'item_name': item.item_name,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price) if item.unit_price else 0,
                } for item in items]
                
                quote_data = {
                    'schedule_id': quote_schedule.id,
                    'quote_date': quote_schedule.visit_date.strftime('%Y-%m-%d'),
                    'expected_revenue': float(quote_schedule.expected_revenue) if quote_schedule.expected_revenue else 0,
                    'items': items_data,
                }
                quotes_data.append(quote_data)
        
        if not quotes_data:
            return JsonResponse({
                'error': '견적 품목이 없습니다.'
            })
        
        return JsonResponse({
            'success': True,
            'quotes': quotes_data,  # 모든 견적 반환
            'count': len(quotes_data),
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'오류가 발생했습니다: {str(e)}'
        }, status=500)


# ============================================
# 선결제 관리
# ============================================

@login_required
def prepayment_list_view(request):
    """선결제 목록 뷰 (같은 회사 직원 데이터 필터링)"""
    from reporting.models import Prepayment
    from django.db.models import Q, Sum
    from django.contrib.auth.models import User
    
    user_profile = get_user_profile(request.user)
    base_queryset = Prepayment.objects.select_related('customer', 'company', 'created_by')
    
    # === 데이터 필터: 나 / 전체(같은 회사) / 특정 직원 ===
    data_filter = request.GET.get('data_filter', 'me')  # 기본값: 나
    filter_user_id = request.GET.get('filter_user')  # 특정 직원 ID
    
    # 같은 회사 사용자 목록 가져오기 (드롭다운용) - 매니저 제외
    company_users = []
    if user_profile and user_profile.company:
        company_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        ).exclude(id=request.user.id).select_related('userprofile').order_by('username')
    
    # 필터에 따른 대상 사용자 결정
    selected_filter_user = None
    is_viewing_others = False
    
    if data_filter == 'all' and user_profile and user_profile.company:
        # 같은 회사 전체 (salesman만)
        filter_users = User.objects.filter(
            userprofile__company=user_profile.company,
            userprofile__role='salesman',
            is_active=True
        )
        base_queryset = base_queryset.filter(created_by__in=filter_users)
        is_viewing_others = True
    elif data_filter == 'user' and filter_user_id and user_profile and user_profile.company:
        # 특정 직원 (같은 회사 확인)
        try:
            selected_filter_user = User.objects.get(
                id=filter_user_id,
                userprofile__company=user_profile.company,
                is_active=True
            )
            base_queryset = base_queryset.filter(created_by=selected_filter_user)
            is_viewing_others = True
        except User.DoesNotExist:
            base_queryset = base_queryset.filter(created_by=request.user)
            data_filter = 'me'
    else:
        # 'me' - 본인만
        base_queryset = base_queryset.filter(created_by=request.user)
    
    # 검색 필터
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        base_queryset = base_queryset.filter(
            Q(customer__customer_name__icontains=search_query) |
            Q(payer_name__icontains=search_query) |
            Q(memo__icontains=search_query)
        )
    
    if status_filter:
        base_queryset = base_queryset.filter(status=status_filter)
    
    # 정렬
    prepayments = base_queryset.order_by('-payment_date', '-created_at')
    
    # 통계
    stats = base_queryset.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance')
    )
    
    # 사용 금액 계산
    total_amount = stats['total_amount'] or 0
    total_balance = stats['total_balance'] or 0
    stats['total_used'] = total_amount - total_balance
    
    # 페이지네이션 추가
    from django.core.paginator import Paginator
    paginator = Paginator(prepayments, 30)  # 페이지당 30개
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': '선결제 현황',
        'prepayments': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'stats': stats,
        # 새로운 필터 관련 컨텍스트
        'data_filter': data_filter,
        'filter_user_id': filter_user_id,
        'company_users': company_users,
        'selected_filter_user': selected_filter_user,
        'is_viewing_others': is_viewing_others,
    }
    
    return render(request, 'reporting/prepayment/list.html', context)


@login_required
def prepayment_create_view(request):
    """선결제 등록 뷰"""
    from reporting.models import Prepayment, FollowUp
    from django import forms
    
    # Tailwind CSS 클래스
    input_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    select_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    textarea_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    
    class PrepaymentForm(forms.ModelForm):
        customer = forms.ModelChoiceField(
            queryset=FollowUp.objects.all(),
            label='고객',
            widget=forms.Select(attrs={'class': select_class})
        )
        
        class Meta:
            model = Prepayment
            fields = ['customer', 'amount', 'payment_date', 'payment_method', 'payer_name', 'memo']
            widgets = {
                'amount': forms.NumberInput(attrs={'class': input_class, 'placeholder': '금액 입력'}),
                'payment_date': forms.DateInput(attrs={'class': input_class, 'type': 'date'}),
                'payment_method': forms.Select(attrs={'class': select_class}),
                'payer_name': forms.TextInput(attrs={'class': input_class, 'placeholder': '입금자명 (선택)'}),
                'memo': forms.Textarea(attrs={'class': textarea_class, 'rows': 3, 'placeholder': '메모 (선택)'}),
            }
    
    if request.method == 'POST':
        form = PrepaymentForm(request.POST)
        if form.is_valid():
            prepayment = form.save(commit=False)
            prepayment.balance = prepayment.amount  # 초기 잔액 = 입금액
            prepayment.company = prepayment.customer.company
            prepayment.created_by = request.user
            prepayment.save()
            
            messages.success(request, f'{prepayment.customer.customer_name}의 선결제 {prepayment.amount:,}원이 등록되었습니다.')
            return redirect('reporting:prepayment_detail', pk=prepayment.pk)
    else:
        # 한국 시간대의 오늘 날짜를 기본값으로 설정
        from django.utils import timezone
        import pytz
        korea_tz = pytz.timezone('Asia/Seoul')
        today_korea = timezone.now().astimezone(korea_tz).date()
        
        form = PrepaymentForm(initial={'payment_date': today_korea})
    
    # 고객 목록 필터링 (회사별)
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.role != 'admin':
        # 같은 UserCompany 소속 사용자들이 등록한 고객만 표시
        from reporting.models import UserProfile
        same_company_users = UserProfile.objects.filter(
            company=user_profile.company
        ).values_list('user_id', flat=True)
        form.fields['customer'].queryset = FollowUp.objects.filter(user_id__in=same_company_users)
    
    context = {
        'page_title': '선결제 등록',
        'form': form,
    }
    
    return render(request, 'reporting/prepayment/form.html', context)


@login_required
def prepayment_detail_view(request, pk):
    """선결제 상세 뷰"""
    from reporting.models import Prepayment, UserProfile
    
    prepayment = get_object_or_404(Prepayment, pk=pk)
    
    # 권한 체크 - 같은 회사 사용자만 조회 가능
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.company:
        same_company_users = UserProfile.objects.filter(
            company=user_profile.company
        ).values_list('user_id', flat=True)
        
        if prepayment.created_by_id not in same_company_users:
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 본인 데이터 여부
    is_owner = (prepayment.created_by == request.user)
    
    # 사용 내역
    usages = prepayment.usages.select_related(
        'schedule', 
        'schedule__followup'
    ).prefetch_related(
        'schedule__delivery_items_set'
    ).order_by('-used_at')
    
    # 각 usage에 delivery_items 첨부
    for usage in usages:
        if usage.schedule:
            usage.delivery_items = usage.schedule.delivery_items_set.all()
        else:
            usage.delivery_items = []
    
    # 금액 계산
    total_used = prepayment.amount - prepayment.balance
    usage_percent = 0
    if prepayment.amount > 0:
        from decimal import Decimal
        usage_percent = round(float(total_used / prepayment.amount) * 100, 1)
        balance_percent = round(float(prepayment.balance / prepayment.amount) * 100, 1)
    else:
        balance_percent = 0
    
    context = {
        'page_title': f'선결제 상세 - {prepayment.customer.customer_name}',
        'prepayment': prepayment,
        'usages': usages,
        'total_used': total_used,
        'usage_percent': usage_percent,
        'balance_percent': balance_percent,
        'is_owner': is_owner,
    }
    
    return render(request, 'reporting/prepayment/detail.html', context)


@login_required
def prepayment_edit_view(request, pk):
    """선결제 수정 뷰"""
    from reporting.models import Prepayment, FollowUp
    from django import forms
    
    prepayment = get_object_or_404(Prepayment, pk=pk)
    
    # 권한 체크 - 본인 데이터만 수정 가능
    if prepayment.created_by != request.user:
        messages.error(request, '본인이 등록한 선결제만 수정할 수 있습니다.')
        return redirect('reporting:prepayment_list')
    
    # Tailwind CSS 클래스
    input_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    select_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    textarea_class = 'w-full px-4 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'
    
    class PrepaymentEditForm(forms.ModelForm):
        customer = forms.ModelChoiceField(
            queryset=FollowUp.objects.all(),
            label='고객',
            widget=forms.Select(attrs={'class': select_class})
        )
        
        class Meta:
            model = Prepayment
            fields = ['customer', 'amount', 'balance', 'payment_date', 'payment_method', 'payer_name', 'status', 'memo']
            widgets = {
                'amount': forms.NumberInput(attrs={'class': input_class, 'placeholder': '금액 입력'}),
                'balance': forms.NumberInput(attrs={'class': input_class, 'placeholder': '잔액'}),
                'payment_date': forms.DateInput(attrs={'class': input_class, 'type': 'date'}),
                'payment_method': forms.Select(attrs={'class': select_class}),
                'payer_name': forms.TextInput(attrs={'class': input_class, 'placeholder': '입금자명 (선택)'}),
                'status': forms.Select(attrs={'class': select_class}),
                'memo': forms.Textarea(attrs={'class': textarea_class, 'rows': 3, 'placeholder': '메모 (선택)'}),
            }
    
    if request.method == 'POST':
        form = PrepaymentEditForm(request.POST, instance=prepayment)
        if form.is_valid():
            prepayment = form.save(commit=False)
            prepayment.company = prepayment.customer.company
            prepayment.save()
            
            messages.success(request, '선결제 정보가 수정되었습니다.')
            return redirect('reporting:prepayment_detail', pk=prepayment.pk)
    else:
        form = PrepaymentEditForm(instance=prepayment)
    
    # 고객 목록 필터링 (회사별)
    user_profile = get_user_profile(request.user)
    if user_profile and user_profile.company:
        accessible_users = get_accessible_users(request.user, request)
        form.fields['customer'].queryset = FollowUp.objects.filter(user__in=accessible_users)
    
    context = {
        'page_title': '선결제 수정',
        'form': form,
        'prepayment': prepayment,
    }
    
    return render(request, 'reporting/prepayment/form.html', context)


@login_required
@login_required
def prepayment_transfer_view(request, pk):
    """선결제 이관 뷰 - 같은 회사 내 다른 영업사원에게 이관"""
    from reporting.models import Prepayment, UserProfile
    from django.contrib.auth.models import User

    prepayment = get_object_or_404(Prepayment, pk=pk)

    # 본인 선결제만 이관 가능
    if prepayment.created_by != request.user:
        messages.error(request, '본인이 등록한 선결제만 이관할 수 있습니다.')
        return redirect('reporting:prepayment_detail', pk=pk)

    # 같은 회사 내 다른 salesman 목록
    try:
        my_profile = UserProfile.objects.get(user=request.user)
        colleagues = UserProfile.objects.filter(
            company=my_profile.company,
            role='salesman'
        ).exclude(user=request.user).select_related('user')
    except UserProfile.DoesNotExist:
        colleagues = []

    if request.method == 'POST':
        target_user_id = request.POST.get('target_user')
        reason = request.POST.get('reason', '').strip()

        if not target_user_id:
            messages.error(request, '이관 대상을 선택해주세요.')
        else:
            target_user = get_object_or_404(User, pk=target_user_id)
            # 같은 회사 소속 확인
            if not UserProfile.objects.filter(user=target_user, company=my_profile.company).exists():
                messages.error(request, '같은 회사 소속이 아닙니다.')
            else:
                from_name = request.user.get_full_name() or request.user.username
                to_name = target_user.get_full_name() or target_user.username

                # 이관 메모 기록
                transfer_note = f"[이관] {from_name} → {to_name} ({prepayment.created_at.strftime('%Y-%m-%d')})"
                if reason:
                    transfer_note += f"\n사유: {reason}"

                prepayment.created_by = target_user
                if prepayment.memo:
                    prepayment.memo = prepayment.memo + '\n' + transfer_note
                else:
                    prepayment.memo = transfer_note
                prepayment.save()

                messages.success(request, f'선결제가 {to_name}님께 이관되었습니다.')
                return redirect('reporting:prepayment_detail', pk=pk)

    context = {
        'prepayment': prepayment,
        'colleagues': colleagues,
        'page_title': f'선결제 이관 - {prepayment.customer.customer_name}',
    }
    return render(request, 'reporting/prepayment/transfer.html', context)


@login_required
def prepayment_delete_view(request, pk):
    """선결제 삭제 뷰"""
    from reporting.models import Prepayment
    from django.utils import timezone
    
    prepayment = get_object_or_404(Prepayment, pk=pk)
    
    # 권한 체크 - 본인 데이터만 삭제 가능
    if prepayment.created_by != request.user:
        messages.error(request, '본인이 등록한 선결제만 삭제할 수 있습니다.')
        return redirect('reporting:prepayment_list')
    
    if request.method == 'POST':
        # 사용 내역 개수 확인
        usage_count = prepayment.usages.count()
        
        # 취소 요청인 경우
        if request.POST.get('action') == 'cancel':
            prepayment.status = 'cancelled'
            prepayment.cancelled_at = timezone.now()
            prepayment.cancel_reason = request.POST.get('cancel_reason', '사용자 요청으로 취소')
            prepayment.save()
            messages.success(request, f'{prepayment.customer.customer_name}의 선결제가 취소되었습니다.')
            return redirect('reporting:prepayment_detail', pk=pk)
        
        # 삭제 요청인 경우
        if usage_count > 0:
            messages.error(request, f'이미 {usage_count}개의 사용 내역이 있는 선결제는 삭제할 수 없습니다.')
            return redirect('reporting:prepayment_detail', pk=pk)
        
        customer_name = prepayment.customer.customer_name
        prepayment.delete()
        messages.success(request, f'{customer_name}의 선결제가 삭제되었습니다.')
        return redirect('reporting:prepayment_list')
    
    context = {
        'page_title': '선결제 삭제',
        'prepayment': prepayment,
    }
    
    return render(request, 'reporting/prepayment/delete_confirm.html', context)


@login_required
def prepayment_customer_view(request, customer_id):
    """부서별 선결제 관리 뷰 (부서 중심 - 동일 부서 내 모든 고객의 선결제 표시)"""
    from reporting.models import Prepayment, FollowUp, Department
    from django.db.models import Sum, Q, Count
    
    # 고객 정보 가져오기 (URL에서 전달받은 customer_id 기준)
    customer = get_object_or_404(FollowUp, pk=customer_id)
    
    # 부서 정보 가져오기
    department = customer.department
    company = customer.company
    
    # 권한 체크 - 고객의 담당자 또는 해당 고객에게 선결제를 등록한 사용자가 접근 가능
    user_profile = get_user_profile(request.user)
    
    # Admin과 Manager는 모든 고객에 접근 가능
    if not (user_profile.is_admin() or user_profile.is_manager()):
        # Salesman인 경우
        # 1. 고객의 담당자이거나
        # 2. 해당 고객에게 선결제를 등록한 적이 있는 경우 접근 가능
        is_customer_owner = (customer.user == request.user)
        has_prepayment = Prepayment.objects.filter(
            customer=customer,
            created_by=request.user
        ).exists()
        
        if not (is_customer_owner or has_prepayment):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 현재 보고 있는 사용자 결정 (세션에서 선택된 사용자 또는 본인)
    target_user = request.user
    
    if user_profile.can_view_all_users():
        user_filter = request.session.get('selected_user_id')
        if user_filter:
            from django.contrib.auth.models import User
            accessible_users = get_accessible_users(request.user, request)
            try:
                target_user = accessible_users.get(id=user_filter)
            except User.DoesNotExist:
                target_user = request.user
    
    # 부서 기준 조회: 동일 부서 내 모든 고객의 선결제 조회
    if department:
        # 부서 내 모든 고객(FollowUp) 조회
        department_followups = FollowUp.objects.filter(department=department).select_related('company', 'department')
        
        # 부서 내 모든 고객의 선결제 조회
        prepayments = Prepayment.objects.filter(
            customer__department=department,
            created_by=target_user
        ).select_related('company', 'customer', 'created_by').prefetch_related('usages').order_by('payment_date', 'id')
    else:
        # 부서 정보가 없는 경우 기존 고객 기준 조회
        department_followups = [customer]
        prepayments = Prepayment.objects.filter(
            customer=customer,
            created_by=target_user
        ).select_related('company', 'customer', 'created_by').prefetch_related('usages').order_by('payment_date', 'id')
    
    # 각 선결제의 사용금액 계산
    for prepayment in prepayments:
        prepayment.used_amount = prepayment.amount - prepayment.balance
    
    # 통계 계산
    stats = prepayments.aggregate(
        total_amount=Sum('amount'),
        total_balance=Sum('balance'),
        count=Count('id')
    )
    
    total_amount = stats['total_amount'] or 0
    total_balance = stats['total_balance'] or 0
    stats['total_used'] = total_amount - total_balance
    
    # 상태별 개수
    active_count = prepayments.filter(status='active').count()
    depleted_count = prepayments.filter(status='depleted').count()
    cancelled_count = prepayments.filter(status='cancelled').count()
    
    # 페이지 제목 구성
    if department and company:
        page_title = f'{company.name} - {department.name} 선결제 관리'
    elif company:
        page_title = f'{company.name} - 선결제 관리'
    else:
        page_title = f'{customer.customer_name} - 선결제 관리'
    
    context = {
        'page_title': page_title,
        'customer': customer,
        'company': company,
        'department': department,
        'department_followups': department_followups,
        'prepayments': prepayments,
        'stats': stats,
        'active_count': active_count,
        'depleted_count': depleted_count,
        'cancelled_count': cancelled_count,
    }
    
    return render(request, 'reporting/prepayment/customer.html', context)


@login_required
def prepayment_customer_excel(request, customer_id):
    """부서별 선결제 엑셀 다운로드 (부서 중심 - 동일 부서 내 모든 고객의 선결제 포함)"""
    from reporting.models import Prepayment, FollowUp, PrepaymentUsage, Department
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime
    
    # 고객 정보 가져오기
    customer = get_object_or_404(FollowUp, pk=customer_id)
    
    # 부서 정보 가져오기
    department = customer.department
    company = customer.company
    
    # 권한 체크 - 고객의 담당자 또는 해당 고객에게 선결제를 등록한 사용자가 접근 가능
    user_profile = get_user_profile(request.user)
    
    # Admin과 Manager는 모든 고객에 접근 가능
    if not (user_profile.is_admin() or user_profile.is_manager()):
        # Salesman인 경우
        is_customer_owner = (customer.user == request.user)
        has_prepayment = Prepayment.objects.filter(
            customer=customer,
            created_by=request.user
        ).exists()
        
        if not (is_customer_owner or has_prepayment):
            messages.error(request, '접근 권한이 없습니다.')
            return redirect('reporting:prepayment_list')
    
    # 현재 보고 있는 사용자 결정 (세션에서 선택된 사용자 또는 본인)
    target_user = request.user
    
    if user_profile.can_view_all_users():
        user_filter = request.session.get('selected_user_id')
        if user_filter:
            from django.contrib.auth.models import User
            accessible_users = get_accessible_users(request.user, request)
            try:
                target_user = accessible_users.get(id=user_filter)
            except User.DoesNotExist:
                target_user = request.user
    
    # 부서 기준 조회: 동일 부서 내 모든 고객의 선결제 조회
    if department:
        prepayments = Prepayment.objects.filter(
            customer__department=department,
            created_by=target_user
        ).select_related('company', 'customer', 'created_by').prefetch_related(
            'usages__schedule__delivery_items_set'
        ).order_by('payment_date', 'id')
    else:
        # 부서 정보가 없는 경우 기존 고객 기준 조회
        prepayments = Prepayment.objects.filter(
            customer=customer,
            created_by=target_user
        ).select_related('company', 'customer', 'created_by').prefetch_related(
            'usages__schedule__delivery_items_set'
        ).order_by('payment_date', 'id')
    
    # 엑셀 제목 구성
    if department and company:
        excel_title = f"{company.name} - {department.name} 선결제 요약"
        filename_prefix = f"{company.name}_{department.name}"
    elif company:
        excel_title = f"{company.name} 선결제 요약"
        filename_prefix = company.name
    else:
        excel_title = f"{customer.customer_name} 선결제 요약"
        filename_prefix = customer.customer_name
    
    # 엑셀 생성
    wb = Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    # ========================================
    # 시트 1: 선결제 요약
    # ========================================
    ws1 = wb.active
    ws1.title = "선결제 요약"
    
    # 제목 (부서 중심으로 변경)
    ws1.merge_cells('A1:I1')
    title_cell = ws1['A1']
    title_cell.value = excel_title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    ws1.row_dimensions[1].height = 30
    
    # 헤더 (고객명 컬럼 추가)
    headers1 = ['번호', '결제일', '고객명', '지불자', '결제방법', '선결제금액', '사용금액', '남은잔액', '상태']
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15   # 고객명
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 12
    
    # 데이터 행
    total_amount = 0
    total_used = 0
    total_balance = 0
    
    for idx, prepayment in enumerate(prepayments, 1):
        row = idx + 3
        used_amount = prepayment.amount - prepayment.balance
        
        total_amount += prepayment.amount
        total_used += used_amount
        total_balance += prepayment.balance
        
        # 데이터 (고객명 컬럼 추가)
        data = [
            idx,
            prepayment.payment_date.strftime('%Y-%m-%d'),
            prepayment.customer.customer_name if prepayment.customer else '-',  # 고객명 추가
            prepayment.payer_name or '-',
            prepayment.get_payment_method_display(),
            float(prepayment.amount),  # Decimal을 float으로 변환
            float(used_amount),
            float(prepayment.balance),
            prepayment.get_status_display(),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col_num)
            cell.value = value
            cell.border = border
            
            # 정렬 및 서식
            if col_num == 1 or col_num == 9:  # No, 상태
                cell.alignment = center_alignment
            elif col_num >= 6 and col_num <= 8:  # 금액 (컬럼 위치 조정)
                cell.alignment = right_alignment
                cell.number_format = '#,##0'
            
            # 상태별 배경색
            if col_num == 9:  # 상태 (컬럼 위치 조정)
                if prepayment.status == 'active':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif prepayment.status == 'depleted':
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
                elif prepayment.status == 'cancelled':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # 잔액에 따른 배경색
            if col_num == 8:  # 남은잔액 (컬럼 위치 조정)
                if prepayment.balance > 0:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    
    # 합계 행 (컬럼 위치 조정)
    summary_row = len(prepayments) + 4
    ws1.merge_cells(f'A{summary_row}:E{summary_row}')  # 5컬럼으로 확장
    summary_cell = ws1.cell(row=summary_row, column=1)
    summary_cell.value = "합계"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = center_alignment
    summary_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    summary_cell.border = border
    
    for col in range(2, 6):  # 2~5열까지
        ws1.cell(row=summary_row, column=col).border = border
        ws1.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 합계 금액 (컬럼 위치 조정: 6, 7, 8열)
    for col_num, value in [(6, float(total_amount)), (7, float(total_used)), (8, float(total_balance))]:
        cell = ws1.cell(row=summary_row, column=col_num)
        cell.value = value
        cell.font = Font(bold=True, size=11)
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
    
    ws1.cell(row=summary_row, column=9).border = border
    ws1.cell(row=summary_row, column=9).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ========================================
    # 시트 2: 품목별 집계
    # ========================================
    ws2 = wb.create_sheet(title="품목별 집계")
    
    # 제목
    ws2.merge_cells('A1:D1')
    title_cell2 = ws2['A1']
    title_cell2.value = "품목별 사용 집계"
    title_cell2.font = Font(bold=True, size=14)
    title_cell2.alignment = center_alignment
    ws2.row_dimensions[1].height = 30
    
    # 헤더
    headers2 = ['품목명', '총 수량', '단가', '총 사용금액']
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 18
    
    # 품목별 집계 데이터 수집
    # 중복 방지: 동일 schedule의 품목이 여러 선결제 usage에서 중복 집계되지 않도록 처리
    from collections import defaultdict
    item_stats = defaultdict(lambda: {'quantity': 0, 'amount': 0, 'count': 0, 'unit_prices': []})
    processed_schedules = set()  # 이미 처리한 schedule ID 추적
    
    for prepayment in prepayments:
        usages = prepayment.usages.all()
        for usage in usages:
            if usage.schedule and usage.schedule.id not in processed_schedules:
                processed_schedules.add(usage.schedule.id)  # 중복 방지
                delivery_items = usage.schedule.delivery_items_set.all()
                if delivery_items.exists():
                    for item in delivery_items:
                        # 금액 계산
                        item_amount = item.total_price if item.total_price else (float(item.quantity) * float(item.unit_price) * 1.1)
                        
                        # 통계 업데이트
                        item_stats[item.item_name]['quantity'] += float(item.quantity)
                        item_stats[item.item_name]['amount'] += float(item_amount)
                        item_stats[item.item_name]['count'] += 1
                        item_stats[item.item_name]['unit_prices'].append(float(item.unit_price))
    
    # 품목별 데이터 작성 (사용금액 기준 내림차순)
    sorted_items = sorted(item_stats.items(), key=lambda x: x[1]['amount'], reverse=True)
    row_num = 4
    total_summary_quantity = 0
    total_summary_amount = 0
    total_summary_unit_price = 0
    
    for item_name, stats in sorted_items:
        # 평균 단가 계산
        avg_unit_price = sum(stats['unit_prices']) / len(stats['unit_prices']) if stats['unit_prices'] else 0
        
        data = [
            item_name,
            stats['quantity'],
            avg_unit_price,
            stats['amount'],
        ]
        
        total_summary_quantity += stats['quantity']
        total_summary_amount += stats['amount']
        total_summary_unit_price += avg_unit_price
        
        for col_num, value in enumerate(data, 1):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = right_alignment
                if col_num in [2, 3, 4]:  # 수량, 단가, 금액
                    cell.number_format = '#,##0'
        
        row_num += 1
    
    # 품목별 합계 행 추가
    if sorted_items:
        summary_row = row_num
        ws2.cell(row=summary_row, column=1).value = "합계"
        ws2.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=1).alignment = center_alignment
        ws2.cell(row=summary_row, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=1).border = border
        
        ws2.cell(row=summary_row, column=2).value = total_summary_quantity
        ws2.cell(row=summary_row, column=2).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=2).alignment = right_alignment
        ws2.cell(row=summary_row, column=2).number_format = '#,##0'
        ws2.cell(row=summary_row, column=2).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=2).border = border
        
        ws2.cell(row=summary_row, column=3).value = total_summary_unit_price
        ws2.cell(row=summary_row, column=3).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=3).alignment = right_alignment
        ws2.cell(row=summary_row, column=3).number_format = '#,##0'
        ws2.cell(row=summary_row, column=3).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=3).border = border
        
        ws2.cell(row=summary_row, column=4).value = total_summary_amount
        ws2.cell(row=summary_row, column=4).font = Font(bold=True, size=11)
        ws2.cell(row=summary_row, column=4).alignment = right_alignment
        ws2.cell(row=summary_row, column=4).number_format = '#,##0'
        ws2.cell(row=summary_row, column=4).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws2.cell(row=summary_row, column=4).border = border
        
        # 총 남은 선결제 잔액 정보 추가
        row_num += 2
        balance_row = row_num
        ws2.merge_cells(f'A{balance_row}:B{balance_row}')
        balance_label_cell = ws2.cell(row=balance_row, column=1)
        balance_label_cell.value = "총 남은 선결제 잔액"
        balance_label_cell.font = Font(bold=True, size=12, color="FFFFFF")
        balance_label_cell.alignment = center_alignment
        balance_label_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        balance_label_cell.border = border
        ws2.cell(row=balance_row, column=2).border = border
        ws2.cell(row=balance_row, column=2).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        ws2.merge_cells(f'C{balance_row}:D{balance_row}')
        balance_value_cell = ws2.cell(row=balance_row, column=3)
        balance_value_cell.value = float(total_balance)
        balance_value_cell.font = Font(bold=True, size=12)
        balance_value_cell.alignment = right_alignment
        balance_value_cell.number_format = '#,##0 "원"'
        balance_value_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        balance_value_cell.border = border
        ws2.cell(row=balance_row, column=4).border = border
        ws2.cell(row=balance_row, column=4).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # 품목 데이터가 없는 경우
    if not sorted_items:
        ws2.merge_cells('A4:D4')
        no_data_cell = ws2['A4']
        no_data_cell.value = "품목별 사용 내역이 없습니다."
        no_data_cell.alignment = center_alignment
        no_data_cell.font = Font(italic=True, color="999999")
    
    # HTTP 응답 (부서 중심 파일명 사용)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{filename_prefix}_선결제내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    
    wb.save(response)
    return response


@login_required
def prepayment_list_excel(request):
    """전체 선결제 엑셀 다운로드"""
    from reporting.models import Prepayment
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime
    from django.db.models import Q
    
    # 권한 체크 및 데이터 필터링
    user_profile = get_user_profile(request.user)
    
    if user_profile.is_admin() or user_profile.is_manager():
        # Admin과 Manager는 모든 선결제 접근 가능
        prepayments = Prepayment.objects.all()
    else:
        # Salesman은 본인이 등록한 선결제 또는 본인이 등록한 고객의 선결제
        prepayments = Prepayment.objects.filter(
            Q(created_by=request.user) | Q(customer__user=request.user)
        )
    
    prepayments = prepayments.select_related(
        'customer', 'company', 'created_by'
    ).order_by('-payment_date', '-created_at')
    
    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "전체 선결제"
    
    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    # 제목
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"선결제 전체 내역 ({datetime.now().strftime('%Y-%m-%d')})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # 헤더
    headers = ['No', '고객명', '결제일', '지불자', '결제방법', '선결제금액', '사용금액', '남은잔액', '상태', '등록자', '등록일']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 컬럼 너비 설정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 16
    
    # 데이터 행
    total_amount = 0
    total_used = 0
    total_balance = 0
    
    for idx, prepayment in enumerate(prepayments, 1):
        row = idx + 3
        used_amount = prepayment.amount - prepayment.balance
        
        total_amount += prepayment.amount
        total_used += used_amount
        total_balance += prepayment.balance
        
        # 데이터
        data = [
            idx,
            prepayment.customer.customer_name if prepayment.customer else '-',
            prepayment.payment_date.strftime('%Y-%m-%d'),
            prepayment.payer_name or '-',
            prepayment.get_payment_method_display(),
            prepayment.amount,
            used_amount,
            prepayment.balance,
            prepayment.get_status_display(),
            prepayment.created_by.get_full_name() or prepayment.created_by.username,
            prepayment.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = border
            
            # 정렬
            if col_num == 1 or col_num == 9:  # No, 상태
                cell.alignment = center_alignment
            elif col_num >= 6 and col_num <= 8:  # 금액
                cell.alignment = right_alignment
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
            
            # 상태별 배경색
            if col_num == 9:
                if prepayment.status == 'active':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif prepayment.status == 'depleted':
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
                elif prepayment.status == 'cancelled':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            # 잔액에 따른 배경색
            if col_num == 8:  # 남은잔액
                if prepayment.balance > 0:
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    
    # 합계 행
    summary_row = len(prepayments) + 4
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "합계"
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = center_alignment
    summary_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    summary_cell.border = border
    
    for col in range(2, 6):
        ws.cell(row=summary_row, column=col).border = border
        ws.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 합계 금액
    for col_num, value in [(6, total_amount), (7, total_used), (8, total_balance)]:
        cell = ws.cell(row=summary_row, column=col_num)
        cell.value = value
        cell.font = Font(bold=True, size=11)
        cell.alignment = right_alignment
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = border
    
    for col in range(9, 12):
        ws.cell(row=summary_row, column=col).border = border
        ws.cell(row=summary_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # HTTP 응답
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"선결제전체내역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    
    wb.save(response)
    return response


@login_required
def prepayment_api_list(request):
    """고객별 선결제 목록 API (AJAX용)"""
    from reporting.models import Prepayment
    from django.http import JsonResponse
    
    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse({'prepayments': []})
    
    try:
        # 해당 고객의 부서 기준으로 같은 부서 내 모든 고객의 선결제를 불러옴
        followup = FollowUp.objects.select_related('department').get(id=customer_id)
        department = followup.department
        
        # 같은 부서의 모든 고객 ID 조회
        same_dept_followup_ids = FollowUp.objects.filter(
            department=department
        ).values_list('id', flat=True)
        
        prepayments = Prepayment.objects.filter(
            customer_id__in=same_dept_followup_ids,
            status='active',
            balance__gt=0
        ).select_related('customer').order_by('payment_date')
        
        prepayments_data = [{
            'id': p.id,
            'payment_date': p.payment_date.strftime('%Y-%m-%d'),
            'payer_name': p.payer_name or '미지정',
            'customer_name': p.customer.customer_name or str(p.customer),
            'amount': float(p.amount),
            'balance': float(p.balance),
        } for p in prepayments]
        
        return JsonResponse({'prepayments': prepayments_data})
    except Exception as e:
        return JsonResponse({'prepayments': [], 'error': str(e)})


# ============================================
# 제품 관리 뷰
# ============================================

@login_required
def product_list(request):
    """제품 목록"""
    from reporting.models import Product, DeliveryItem, QuoteItem
    from django.db.models import Q, Count
    
    user_profile = get_user_profile(request.user)
    
    # 검색 기능
    search_query = request.GET.get('search', '')
    is_active = request.GET.get('is_active', '')
    
    # 회사별 필터링
    if user_profile.is_admin():
        # 관리자는 필터링된 사용자의 제품만
        accessible_users = get_accessible_users(request.user, request)
        if accessible_users.count() == User.objects.count():
            # 전체 사용자면 모두 표시
            products = Product.objects.all()
        else:
            # 필터링된 사용자의 제품만
            products = Product.objects.filter(created_by__in=accessible_users)
    elif user_profile.company:
        # 같은 회사의 사용자가 생성한 제품만
        accessible_users = get_accessible_users(request.user, request)
        products = Product.objects.filter(created_by__in=accessible_users)
    else:
        # 본인이 생성한 제품만
        products = Product.objects.filter(created_by=request.user)
    
    if search_query:
        products = products.filter(
            Q(product_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if is_active:
        products = products.filter(is_active=(is_active == 'true'))
    
    # 견적 횟수와 판매 횟수를 미리 계산 (정렬을 위해)
    from django.db.models import Count, Case, When, IntegerField, Value
    from django.db.models.functions import Coalesce
    
    products = products.annotate(
        quote_count=Count(
            'delivery_items__schedule',
            distinct=True,
            filter=Q(
                delivery_items__schedule__isnull=False,
                delivery_items__schedule__activity_type='quote'
            )
        ),
        completed_schedule_count=Count(
            'delivery_items__schedule',
            distinct=True, 
            filter=Q(
                delivery_items__schedule__status='completed',
                delivery_items__schedule__activity_type='delivery'
            )
        ),
        history_count=Count(
            'delivery_items__history',
            distinct=True,
            filter=Q(delivery_items__history__isnull=False)
        )
    ).annotate(
        delivery_count=Coalesce('completed_schedule_count', Value(0)) + Coalesce('history_count', Value(0))
    )
    
    # 정렬
    sort_by = request.GET.get('sort', 'code')
    sort_order = request.GET.get('order', 'asc')  # asc 또는 desc
    
    # 정렬 필드 매핑
    sort_fields = {
        'code': 'product_code',
        'description': 'description',
        'price': 'standard_price',
        'status': 'is_active',
        'quote_count': 'quote_count',
        'delivery_count': 'delivery_count',
    }
    
    # 기본 정렬 필드
    order_field = sort_fields.get(sort_by, 'product_code')
    
    # 내림차순인 경우 '-' 추가
    if sort_order == 'desc':
        order_field = '-' + order_field
    
    # 현재가(프로모션 가격 포함) 정렬은 따로 처리
    if sort_by == 'promo_price':
        from django.db.models import Case, When, F
        if sort_order == 'desc':
            products = products.order_by(
                Case(
                    When(is_promo=True, then=F('promotion_price')),
                    default=F('standard_price')
                ).desc()
            )
        else:
            products = products.order_by(
                Case(
                    When(is_promo=True, then=F('promotion_price')),
                    default=F('standard_price')
                )
            )
    else:
        products = products.order_by(order_field)
    
    # 각 제품의 견적/판매 횟수는 이미 annotate로 계산됨 - 아래 루프 제거
    
    # 페이지네이션
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'is_active': is_active,
        'sort_by': sort_by,
        'sort_order': sort_order,
    }
    
    return render(request, 'reporting/product_list.html', context)


@login_required
def product_create(request):
    """제품 등록"""
    from reporting.models import Product
    from decimal import Decimal
    from django.db import IntegrityError
    
    if request.method == 'POST':
        # AJAX 요청 처리
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                product_code = request.POST.get('product_code', '').strip()
                
                # 품번 중복 체크
                if Product.objects.filter(product_code=product_code).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'품번 "{product_code}"은(는) 이미 등록되어 있습니다.'
                    }, status=400)
                
                product = Product(
                    product_code=product_code,
                    standard_price=Decimal(request.POST.get('standard_price', 0)),
                    is_active=True,  # 기본값
                    created_by=request.user,  # 생성자 저장
                )
                
                # 선택 필드들
                if request.POST.get('description'):
                    product.description = request.POST.get('description')
                if request.POST.get('specification'):
                    product.specification = request.POST.get('specification')
                if request.POST.get('unit'):
                    product.unit = request.POST.get('unit')
                
                # 프로모션 설정
                if request.POST.get('is_promo') == 'on':
                    product.is_promo = True
                    if request.POST.get('promo_price'):
                        product.promo_price = Decimal(request.POST.get('promo_price'))
                    if request.POST.get('promo_start'):
                        product.promo_start = request.POST.get('promo_start')
                    if request.POST.get('promo_end'):
                        product.promo_end = request.POST.get('promo_end')
                
                product.save()
                return JsonResponse({
                    'success': True,
                    'product': {
                        'id': product.id,
                        'product_code': product.product_code,
                        'name': product.product_code,  # 품번을 이름으로 사용
                        'unit_price': str(product.standard_price),
                        'description': product.description or '',
                    }
                })
                
            except IntegrityError as e:
                logger.error(f"제품 등록 실패 (중복): {e}")
                return JsonResponse({
                    'success': False,
                    'error': '이미 등록된 품번입니다.'
                }, status=400)
            except Exception as e:
                logger.error(f"제품 등록 실패: {e}")
                error_msg = str(e)
                if 'UNIQUE constraint' in error_msg:
                    error_msg = '이미 등록된 품번입니다.'
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)
        
        # 일반 폼 제출 처리
        try:
            product_code = request.POST.get('product_code', '').strip()
            
            # 품번 중복 체크
            if Product.objects.filter(product_code=product_code).exists():
                messages.error(request, f'품번 "{product_code}"은(는) 이미 등록되어 있습니다.')
                return render(request, 'reporting/product_form.html', {})
            
            product = Product(
                product_code=product_code,
                standard_price=Decimal(request.POST.get('standard_price', 0)),
                is_active=request.POST.get('is_active') == 'on',
                created_by=request.user,  # 생성자 저장
            )
            
            # 선택 필드들
            if request.POST.get('description'):
                product.description = request.POST.get('description')
            
            # 프로모션 설정
            if request.POST.get('is_promo') == 'on':
                product.is_promo = True
                if request.POST.get('promo_price'):
                    product.promo_price = Decimal(request.POST.get('promo_price'))
                if request.POST.get('promo_start'):
                    product.promo_start = request.POST.get('promo_start')
                if request.POST.get('promo_end'):
                    product.promo_end = request.POST.get('promo_end')
            
            product.save()
            messages.success(request, f'제품 "{product.product_code}"이(가) 등록되었습니다.')
            return redirect('reporting:product_list')
            
        except IntegrityError as e:
            logger.error(f"제품 등록 실패 (중복): {e}")
            messages.error(request, '이미 등록된 품번입니다.')
        except Exception as e:
            logger.error(f"제품 등록 실패: {e}")
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg:
                error_msg = '이미 등록된 품번입니다.'
            messages.error(request, f'제품 등록에 실패했습니다: {error_msg}')
    
    return render(request, 'reporting/product_form.html', {})


@login_required
@require_POST
def product_bulk_create(request):
    """엑셀 데이터 일괄 제품 등록 (AJAX) - 중복 시 업데이트"""
    from reporting.models import Product
    from decimal import Decimal
    from django.db import IntegrityError
    import json
    
    try:
        data = json.loads(request.body)
        products_data = data.get('products', [])
        
        if not products_data:
            return JsonResponse({
                'success': False,
                'error': '등록할 제품 데이터가 없습니다.'
            }, status=400)
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        for item in products_data:
            try:
                product_code = item.get('product_code', '').strip()
                product_name = item.get('product_name', '').strip()
                specification = item.get('specification', '').strip()
                unit = item.get('unit', 'EA').strip()
                standard_price = Decimal(str(item.get('standard_price', 0)))
                
                # 기존 제품 체크
                existing = Product.objects.filter(product_code=product_code).first()
                
                if existing:
                    # 데이터가 다르면 업데이트
                    needs_update = False
                    if product_name and existing.description != product_name:
                        existing.description = product_name
                        needs_update = True
                    if specification and existing.specification != specification:
                        existing.specification = specification
                        needs_update = True
                    if unit and existing.unit != unit:
                        existing.unit = unit
                        needs_update = True
                    if standard_price and existing.standard_price != standard_price:
                        existing.standard_price = standard_price
                        needs_update = True
                    
                    if needs_update:
                        existing.save()
                        updated_count += 1
                    else:
                        skipped_count += 1
                        errors.append(f'{product_code}: 동일한 데이터 (변경 없음)')
                else:
                    # 신규 등록
                    product = Product(
                        product_code=product_code,
                        description=product_name or product_code,
                        specification=specification,
                        unit=unit,
                        standard_price=standard_price,
                        is_active=True,
                        created_by=request.user
                    )
                    product.save()
                    created_count += 1
                
            except IntegrityError as e:
                skipped_count += 1
                errors.append(f'{product_code}: 데이터베이스 오류')
                logger.error(f"제품 등록 실패 ({product_code}): {e}")
            except Exception as e:
                errors.append(f'{product_code}: {str(e)}')
                logger.error(f"제품 등록 실패 ({product_code}): {e}")
        
        return JsonResponse({
            'success': True,
            'created_count': created_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count,
            'errors': errors
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': '잘못된 JSON 형식입니다.'
        }, status=400)
    except Exception as e:
        logger.error(f"일괄 제품 등록 실패: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def product_edit(request, product_id):
    """제품 수정"""
    from reporting.models import Product
    from decimal import Decimal
    
    product = get_object_or_404(Product, id=product_id)
    
    # 권한 체크: 같은 회사 사용자만 수정 가능
    user_profile = get_user_profile(request.user)
    if not user_profile.is_admin():
        if product.created_by:
            # 제품 생성자가 있는 경우
            accessible_users = get_accessible_users(request.user, request)
            if product.created_by not in accessible_users:
                messages.error(request, '이 제품을 수정할 권한이 없습니다.')
                return redirect('reporting:product_list')
    
    if request.method == 'POST':
        try:
            product.product_code = request.POST.get('product_code')
            product.standard_price = Decimal(request.POST.get('standard_price', 0))
            product.is_active = request.POST.get('is_active') == 'on'
            
            # 선택 필드들
            product.description = request.POST.get('description', '')
            
            # 프로모션 설정
            product.is_promo = request.POST.get('is_promo') == 'on'
            if product.is_promo:
                if request.POST.get('promo_price'):
                    product.promo_price = Decimal(request.POST.get('promo_price'))
                if request.POST.get('promo_start'):
                    product.promo_start = request.POST.get('promo_start')
                if request.POST.get('promo_end'):
                    product.promo_end = request.POST.get('promo_end')
            else:
                product.promo_price = None
                product.promo_start = None
                product.promo_end = None
            
            product.save()
            messages.success(request, f'제품 "{product.product_code}"이(가) 수정되었습니다.')
            return redirect('reporting:product_list')
            
        except Exception as e:
            logger.error(f"제품 수정 실패: {e}")
            messages.error(request, f'제품 수정에 실패했습니다: {str(e)}')
    
    context = {
        'product': product,
    }
    
    return render(request, 'reporting/product_form.html', context)


@login_required
@require_POST
def product_delete(request, product_id):
    """제품 삭제"""
    from reporting.models import Product
    
    product = get_object_or_404(Product, id=product_id)
    
    # 권한 체크: 같은 회사 사용자만 삭제 가능
    user_profile = get_user_profile(request.user)
    if not user_profile.is_admin():
        if product.created_by:
            # 제품 생성자가 있는 경우
            accessible_users = get_accessible_users(request.user, request)
            if product.created_by not in accessible_users:
                messages.error(request, '이 제품을 삭제할 권한이 없습니다.')
                return redirect('reporting:product_list')
    
    # 사용 중인 제품인지 확인
    if product.delivery_items.exists() or product.quoteitems.exists():
        messages.warning(request, '이미 견적 또는 납품에 사용된 제품은 삭제할 수 없습니다. 비활성화를 권장합니다.')
        return redirect('reporting:product_list')
    
    product_code = product.product_code
    product.delete()
    messages.success(request, f'제품 "{product_code}"이(가) 삭제되었습니다.')
    return redirect('reporting:product_list')


@login_required
def product_api_list(request):
    """제품 목록 API (AJAX용) - 견적/납품 작성 시 제품 선택"""
    from reporting.models import Product
    
    user_profile = get_user_profile(request.user)
    search = request.GET.get('search', '')
    
    # 회사별 필터링
    if user_profile.is_admin():
        products = Product.objects.filter(is_active=True)
    elif user_profile.company:
        # 같은 회사의 사용자가 생성한 제품만
        accessible_users = get_accessible_users(request.user, request)
        products = Product.objects.filter(
            is_active=True
        ).filter(
            Q(created_by__in=accessible_users) | Q(created_by__isnull=True)
        )
    else:
        # 본인이 생성한 제품 + 생성자가 없는 제품
        products = Product.objects.filter(
            is_active=True
        ).filter(
            Q(created_by=request.user) | Q(created_by__isnull=True)
        )
    
    if search:
        products = products.filter(
            Q(product_code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # 제한 제거 - 모든 제품을 가져와서 클라이언트에서 검색
    # 성능을 위해 필요시 제한 추가 가능: products = products.order_by('product_code')[:1000]
    products = products.order_by('product_code')
    
    products_data = [{
        'id': p.id,
        'product_code': p.product_code,
        'name': p.product_code,
        'description': p.description,  # description 필드 추가
        'standard_price': float(p.standard_price),
        'current_price': float(p.get_current_price()),
        'is_promo': p.is_promo,
    } for p in products]
    
    return JsonResponse({'products': products_data})


# ============================================================
# 서류 템플릿 관리 뷰
# ============================================================

@login_required
def document_template_list(request):
    """서류 템플릿 목록"""
    from reporting.models import DocumentTemplate
    
    # 슈퍼유저는 모든 서류 조회, 일반 사용자는 자신의 회사 서류만
    if request.user.is_superuser:
        templates = DocumentTemplate.objects.filter(
            is_active=True
        ).select_related('created_by', 'company')
    else:
        user_company = request.user.userprofile.company
        templates = DocumentTemplate.objects.filter(
            company=user_company,
            is_active=True
        ).select_related('created_by')
    
    # 서류 종류별 필터
    document_type = request.GET.get('type')
    if document_type:
        templates = templates.filter(document_type=document_type)
    
    context = {
        'templates': templates,
        'selected_type': document_type,
        'document_types': DocumentTemplate.DOCUMENT_TYPE_CHOICES,
        'page_title': '서류 관리'
    }
    return render(request, 'reporting/document_template_list.html', context)


@login_required
@role_required(['admin', 'manager'])
def document_template_create(request):
    """서류 템플릿 생성 (Admin, Manager 전용)"""
    from reporting.models import DocumentTemplate, UserCompany
    import os
    
    # 관리자(superuser)는 모든 회사 선택 가능
    if request.user.is_superuser:
        companies = UserCompany.objects.all().order_by('name')
    else:
        companies = None
    
    if request.method == 'POST':
        try:
            # 관리자는 회사 선택, 일반 사용자는 자신의 회사
            if request.user.is_superuser:
                company_id = request.POST.get('company')
                if company_id:
                    user_company = UserCompany.objects.get(pk=company_id)
                else:
                    messages.error(request, '회사를 선택해주세요.')
                    return redirect('reporting:document_template_create')
            else:
                user_company = request.user.userprofile.company
            
            document_type = request.POST.get('document_type')
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            is_default = request.POST.get('is_default') == 'on'
            file = request.FILES.get('file')
            
            if not file:
                messages.error(request, '파일을 선택해주세요.')
                return redirect('reporting:document_template_create')
            
            # 파일 확장자 확인 (엑셀만 허용)
            file_ext = os.path.splitext(file.name)[1].lower()
            if file_ext == '.xlsx' or file_ext == '.xls':
                file_type = 'xlsx'
            else:
                messages.error(request, '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.')
                return redirect('reporting:document_template_create')
            
            template = DocumentTemplate.objects.create(
                company=user_company,
                document_type=document_type,
                name=name,
                file=file,
                file_type=file_type,
                description=description,
                is_default=is_default,
                created_by=request.user
            )
            
            # 업로드 로그
            logger.info(f"[서류 업로드] 템플릿 생성 완료 - ID: {template.id}")
            logger.info(f"[서류 업로드] 파일 public_id: {template.file.public_id if hasattr(template.file, 'public_id') else 'N/A'}")
            logger.info(f"[서류 업로드] 파일 URL: {template.file.url}")
            if hasattr(template.file, 'storage'):
                logger.info(f"[서류 업로드] 파일 스토리지: {template.file.storage.__class__.__name__}")
            else:
                logger.info(f"[서류 업로드] 파일 타입: CloudinaryField")
            
            # Cloudinary 설정 확인
            from django.conf import settings
            logger.info(f"[서류 업로드] DEFAULT_FILE_STORAGE: {settings.DEFAULT_FILE_STORAGE}")
            if hasattr(settings, 'CLOUDINARY_STORAGE'):
                logger.info(f"[서류 업로드] CLOUDINARY_STORAGE 설정됨")
                logger.info(f"[서류 업로드] CLOUD_NAME: {settings.CLOUDINARY_STORAGE.get('CLOUD_NAME', 'NOT_SET')}")
            
            if hasattr(template.file, 'public_id'):
                logger.info(f"[서류 업로드] Cloudinary public_id: {template.file.public_id}")
            else:
                logger.warning(f"[서류 업로드] public_id 속성 없음 - Cloudinary에 저장되지 않았을 수 있음")
            
            # Cloudinary에 실제 저장된 파일 목록 조회
            try:
                import cloudinary
                import cloudinary.api
                # 최근 업로드된 raw 파일 조회
                recent_resources = cloudinary.api.resources(
                    resource_type='raw',
                    type='upload',
                    prefix='document_templates/',
                    max_results=5
                )
                logger.info(f"[서류 업로드] Cloudinary 최근 파일 목록:")
                for resource in recent_resources.get('resources', []):
                    logger.info(f"  - public_id: {resource.get('public_id')}")
                    logger.info(f"    URL: {resource.get('secure_url')}")
            except Exception as e:
                logger.warning(f"[서류 업로드] Cloudinary 목록 조회 실패: {e}")
                import traceback
                logger.error(traceback.format_exc())
            
            messages.success(request, f'서류 "{name}"이(가) 등록되었습니다.')
            return redirect('reporting:document_template_list')
            
        except Exception as e:
            logger.error(f"서류 템플릿 등록 실패: {e}")
            messages.error(request, f'서류 등록에 실패했습니다: {str(e)}')
    
    from reporting.models import DocumentTemplate
    context = {
        'document_types': DocumentTemplate.DOCUMENT_TYPE_CHOICES,
        'companies': companies,  # 관리자만 사용
        'is_superuser': request.user.is_superuser,
        'page_title': '서류 등록'
    }
    return render(request, 'reporting/document_template_form.html', context)


@login_required
@role_required(['admin', 'manager'])
def document_template_edit(request, pk):
    """서류 템플릿 수정 (Admin, Manager 전용)"""
    from reporting.models import DocumentTemplate
    import os
    
    template = get_object_or_404(DocumentTemplate, pk=pk)
    
    # 권한 체크: 자신의 회사 서류만 수정 가능
    if template.company != request.user.userprofile.company:
        messages.error(request, '다른 회사의 서류는 수정할 수 없습니다.')
        return redirect('reporting:document_template_list')
    
    if request.method == 'POST':
        try:
            template.document_type = request.POST.get('document_type')
            template.name = request.POST.get('name')
            template.description = request.POST.get('description', '')
            template.is_default = request.POST.get('is_default') == 'on'
            
            # 파일 변경 시 (엑셀만 허용)
            if 'file' in request.FILES:
                file = request.FILES['file']
                file_ext = os.path.splitext(file.name)[1].lower()
                
                if file_ext == '.xlsx' or file_ext == '.xls':
                    template.file_type = 'xlsx'
                else:
                    messages.error(request, '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.')
                    return redirect('reporting:document_template_edit', pk=pk)
                
                template.file = file
            
            template.save()
            
            # 업로드 로그
            logger.info(f"[서류 업데이트] 템플릿 수정 완료 - ID: {template.id}")
            logger.info(f"[서류 업데이트] 파일 public_id: {template.file.public_id if hasattr(template.file, 'public_id') else 'N/A'}")
            logger.info(f"[서류 업데이트] 파일 URL: {template.file.url}")
            if hasattr(template.file, 'public_id'):
                logger.info(f"[서류 업데이트] Cloudinary public_id: {template.file.public_id}")
            
            messages.success(request, f'서류 "{template.name}"이(가) 수정되었습니다.')
            return redirect('reporting:document_template_list')
            
        except Exception as e:
            logger.error(f"서류 템플릿 수정 실패: {e}")
            messages.error(request, f'서류 수정에 실패했습니다: {str(e)}')
    
    context = {
        'template': template,
        'document_types': DocumentTemplate.DOCUMENT_TYPE_CHOICES,
        'page_title': '서류 수정'
    }
    return render(request, 'reporting/document_template_form.html', context)


@login_required
@role_required(['admin', 'manager'])
@require_POST
def document_template_delete(request, pk):
    """서류 템플릿 삭제 (Admin, Manager 전용)"""
    from reporting.models import DocumentTemplate
    
    template = get_object_or_404(DocumentTemplate, pk=pk)
    
    # 권한 체크
    if template.company != request.user.userprofile.company:
        messages.error(request, '다른 회사의 서류는 삭제할 수 없습니다.')
        return redirect('reporting:document_template_list')
    
    template_name = template.name
    template.is_active = False
    template.save()
    
    messages.success(request, f'서류 "{template_name}"이(가) 삭제되었습니다.')
    return redirect('reporting:document_template_list')


@login_required
def document_template_download(request, pk):
    """서류 템플릿 다운로드"""
    from reporting.models import DocumentTemplate
    from django.http import FileResponse
    from django.shortcuts import redirect as django_redirect
    import os
    
    template = get_object_or_404(DocumentTemplate, pk=pk)
    
    # 권한 체크 (관리자는 모든 서류 다운로드 가능)
    if not request.user.is_superuser:
        if template.company != request.user.userprofile.company:
            messages.error(request, '다른 회사의 서류는 다운로드할 수 없습니다.')
            return django_redirect('reporting:document_template_list')
    
    if not template.file:
        messages.error(request, '파일이 존재하지 않습니다.')
        return django_redirect('reporting:document_template_list')
    
    try:
        # CloudinaryField는 URL로 리다이렉트, FileField는 파일 다운로드
        if hasattr(template.file, 'public_id'):
            # CloudinaryField - URL로 리다이렉트
            return django_redirect(template.file.url)
        else:
            # FileField - 로컬 파일 다운로드
            file_path = template.file.path
            file_name = os.path.basename(file_path)
            
            response = FileResponse(
                open(file_path, 'rb'),
                as_attachment=True,
                filename=file_name
            )
            return response
    except Exception as e:
        logger.error(f"파일 다운로드 실패: {e}")
        messages.error(request, '파일 다운로드에 실패했습니다.')
        return django_redirect('reporting:document_template_list')


@login_required
@require_POST
def document_template_toggle_default(request, pk):
    """기본 템플릿 설정/해제 (AJAX)"""
    from reporting.models import DocumentTemplate
    
    template = get_object_or_404(DocumentTemplate, pk=pk)
    
    # 권한 체크
    if template.company != request.user.userprofile.company:
        return JsonResponse({'success': False, 'error': '권한이 없습니다.'}, status=403)
    
    template.is_default = not template.is_default
    template.save()  # save() 메서드에서 자동으로 다른 기본 템플릿 해제
    
    return JsonResponse({
        'success': True,
        'is_default': template.is_default
    })


@login_required
def get_document_template_data(request, document_type, schedule_id):
    """
    클라이언트에서 xlwings로 처리하기 위한 템플릿 URL과 변수 데이터 반환
    
    Returns:
        JSON {
            template_url: Cloudinary URL,
            template_filename: 원본 파일명,
            variables: {변수명: 값} 딕셔너리,
            file_info: {파일명, 회사명, 고객명 등}
        }
    """
    from reporting.models import Schedule, DeliveryItem, DocumentTemplate
    from decimal import Decimal
    import pytz
    from datetime import timedelta
    
    try:
        schedule = get_object_or_404(Schedule, pk=schedule_id)
        
        # 권한 체크
        if not can_access_user_data(request.user, schedule.user):
            return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
        
        # 해당 회사의 기본 서류 템플릿 찾기
        company = request.user.userprofile.company
        
        document_template = DocumentTemplate.objects.filter(
            company=company,
            document_type=document_type,
            is_active=True,
            is_default=True
        ).first()
        
        if not document_template:
            document_template = DocumentTemplate.objects.filter(
                company=company,
                document_type=document_type,
                is_active=True
            ).first()
        
        if not document_template:
            doc_type_names = {
                'quotation': '견적서',
                'transaction_statement': '거래명세서',
                'delivery_note': '납품서',
            }
            doc_type_name = doc_type_names.get(document_type, '서류')
            
            return JsonResponse({
                'success': False,
                'error': f'{doc_type_name} 템플릿이 등록되어 있지 않습니다.'
            }, status=404)
        
        if not document_template.file:
            return JsonResponse({
                'success': False,
                'error': '서류 템플릿 파일을 찾을 수 없습니다.'
            }, status=404)
        
        # 납품 품목 조회
        delivery_items = DeliveryItem.objects.filter(schedule=schedule).select_related('product')
        
        # 총액 계산
        subtotal = sum([item.unit_price * item.quantity for item in delivery_items], Decimal('0'))
        tax = subtotal * Decimal('0.1')
        total = subtotal + tax
        
        # 총액을 한글로 변환
        def number_to_korean(number):
            num = int(number)
            if num == 0:
                return '영'
            
            units = ['', '만', '억', '조']
            digits = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
            
            result = []
            unit_idx = 0
            
            while num > 0:
                segment = num % 10000
                if segment > 0:
                    segment_str = []
                    
                    if segment >= 1000:
                        d = segment // 1000
                        if d > 1:
                            segment_str.append(digits[d])
                        segment_str.append('천')
                        segment %= 1000
                    
                    if segment >= 100:
                        d = segment // 100
                        if d > 1:
                            segment_str.append(digits[d])
                        segment_str.append('백')
                        segment %= 100
                    
                    if segment >= 10:
                        d = segment // 10
                        if d > 1:
                            segment_str.append(digits[d])
                        segment_str.append('십')
                        segment %= 10
                    
                    if segment > 0:
                        segment_str.append(digits[segment])
                    
                    if unit_idx > 0:
                        segment_str.append(units[unit_idx])
                    
                    result.insert(0, ''.join(segment_str))
                
                num //= 10000
                unit_idx += 1
            
            return ''.join(result)
        
        total_korean = number_to_korean(total)
        
        # 거래번호 생성
        korea_tz = pytz.timezone('Asia/Seoul')
        today = timezone.now().astimezone(korea_tz)
        today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        from reporting.models import DocumentGenerationLog
        today_count = DocumentGenerationLog.objects.filter(
            company=company,
            created_at__gte=today_start,
            created_at__lte=today_end
        ).count() + 1
        
        transaction_number = f"{today.strftime('%Y')}-{today.strftime('%m')}-{today.strftime('%d')}-{today_count:03d}"
        
        # 담당자 정보
        salesman_name = f"{schedule.user.last_name}{schedule.user.first_name}" if schedule.user.last_name and schedule.user.first_name else schedule.user.username
        
        # 연결된 견적번호 자동 채움
        _linked_quote = schedule.quotes.order_by('-created_at').first()
        _quote_number = _linked_quote.quote_number if _linked_quote else ''

        # 데이터 매핑
        data_map = {
            '년': today.strftime('%Y'),
            '월': today.strftime('%m'),
            '일': today.strftime('%d'),
            '거래번호': transaction_number,
            
            '고객명': schedule.followup.customer_name,
            '업체명': str(schedule.followup.company) if schedule.followup.company else '',
            '학교명': str(schedule.followup.company) if schedule.followup.company else '',
            '부서명': str(schedule.followup.department) if schedule.followup.department else '',
            '연구실': str(schedule.followup.department) if schedule.followup.department else '',
            '담당자': schedule.followup.customer_name,
            '이메일': schedule.followup.email or '',
            '담당자이메일': schedule.followup.email or '',
            '연락처': schedule.followup.phone_number or '',
            '전화번호': schedule.followup.phone_number or '',
            
            '실무자': salesman_name,
            '영업담당자': salesman_name,
            '담당영업': salesman_name,
            '영업담당자이메일': schedule.user.email or '',
            
            '일정날짜': schedule.visit_date.strftime('%Y년 %m월 %d일'),
            '날짜': schedule.visit_date.strftime('%Y년 %m월 %d일'),
            '발행일': today.strftime('%Y년 %m월 %d일'),

            # 견적 정보 (자동 채움)
            '견적번호': _quote_number,
            '메모': schedule.notes or '',
            
            '회사명': company.name,
            
            '공급가액': f"{int(subtotal):,}",
            '소계': f"{int(subtotal):,}",
            '부가세액': f"{int(tax):,}",
            '부가세': f"{int(tax):,}",
            '총액': f"{int(total):,}",
            '합계': f"{int(total):,}",
            '총액한글': f"금 {total_korean}원정",
            '한글금액': f"금 {total_korean}원정",
        }
        
        # 품목 데이터 추가
        items_data = []
        for idx, item in enumerate(delivery_items, 1):
            item_subtotal = item.unit_price * item.quantity
            item_unit = item.unit if item.unit else (item.product.unit if item.product and item.product.unit else 'EA')
            
            item_data = {
                f'품목{idx}_이름': item.item_name,
                f'품목{idx}_품목명': item.item_name,
                f'품목{idx}_수량': str(item.quantity),
                f'품목{idx}_단위': item_unit,
                f'품목{idx}_규격': item.product.specification if item.product and item.product.specification else '',
                f'품목{idx}_설명': item.product.description if item.product and item.product.description else '',
                f'품목{idx}_공급가액': f"{int(item_subtotal):,}",
                f'품목{idx}_단가': f"{int(item.unit_price):,}",
                f'품목{idx}_부가세액': f"{int(item.unit_price * item.quantity * Decimal('0.1')):,}",
                f'품목{idx}_금액': f"{int(item_subtotal):,}",
                f'품목{idx}_총액': f"{int(item_subtotal * Decimal('1.1')):,}",
            }
            data_map.update(item_data)
            items_data.append({
                'index': idx,
                'name': item.item_name,
                'quantity': item.quantity,
                'unit': item_unit,
                'unit_price': int(item.unit_price),
                'subtotal': int(item_subtotal)
            })
        
        # 유효일 계산용 기준일
        base_date = schedule.visit_date.strftime('%Y-%m-%d')
        
        # 파일명 정보
        doc_type_names = {
            'quotation': '견적서',
            'transaction_statement': '거래명세서',
            'delivery_note': '납품서',
        }
        doc_name = doc_type_names.get(document_type, document_template.name)
        customer_company = schedule.followup.company.name if schedule.followup.company else schedule.followup.customer_name
        
        return JsonResponse({
            'success': True,
            'template_url': document_template.file.url,
            'template_filename': document_template.file.name,
            'variables': data_map,
            'file_info': {
                'company_name': company.name,
                'customer_company': customer_company,
                'doc_name': doc_name,
                'today_str': today.strftime('%Y%m%d'),
                'base_date': base_date,
                'transaction_number': transaction_number
            },
            'items': items_data,
            'item_count': len(delivery_items)
        })
        
    except Exception as e:
        logger.error(f"서류 데이터 조회 실패: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'서류 데이터 조회 중 오류가 발생했습니다: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def generate_document_pdf(request, document_type, schedule_id, output_format='xlsx'):
    """
    일정 기반 서류 생성 및 다운로드
    - 업로드된 서류 템플릿에 실제 데이터를 채워서 다운로드
    - 견적서 (quotation)
    - 거래명세서 (transaction_statement)
    
    output_format:
    - 'xlsx': 엑셀 파일로 다운로드 (기본값)
    - 'pdf': PDF 파일로 다운로드
    
    템플릿 파일 형식:
    - 엑셀: {{고객명}}, {{업체명}}, {{품목1_이름}}, {{품목1_수량}}, {{품목1_단가}}, {{품목1_금액}} 등의 변수 사용
    """
    
    from reporting.models import Schedule, DeliveryItem, DocumentTemplate
    from django.http import HttpResponse, FileResponse
    from urllib.parse import quote
    from openpyxl import load_workbook
    from io import BytesIO
    from decimal import Decimal
    import os
    import re
    
    try:
        schedule = get_object_or_404(Schedule, pk=schedule_id)
        
        # 권한 체크
        if not can_access_user_data(request.user, schedule.user):
            logger.warning(f"[서류생성] 권한 없음 - 사용자: {request.user.username}")
            return JsonResponse({'success': False, 'error': '접근 권한이 없습니다.'}, status=403)
        
        # 해당 회사의 기본 서류 템플릿 찾기
        company = request.user.userprofile.company
        
        document_template = DocumentTemplate.objects.filter(
            company=company,
            document_type=document_type,
            is_active=True,
            is_default=True
        ).first()
        
        if not document_template:
            # 기본이 없으면 활성화된 첫 번째 템플릿
            document_template = DocumentTemplate.objects.filter(
                company=company,
                document_type=document_type,
                is_active=True
            ).first()
        
        if not document_template:
            logger.error(f"[서류생성] 템플릿 없음 - 회사: {company.name}, 타입: {document_type}")
            # 서류 종류 이름 매핑
            doc_type_names = {
                'quotation': '견적서',
                'transaction_statement': '거래명세서',
                'delivery_note': '납품서',
            }
            doc_type_name = doc_type_names.get(document_type, '서류')
            
            return JsonResponse({
                'success': False,
                'error': f'{doc_type_name} 템플릿이 등록되어 있지 않습니다. 서류 관리 메뉴에서 먼저 템플릿을 등록해주세요.'
            }, status=404)
        
        
        # 파일이 존재하는지 확인 (Cloudinary 지원)
        if not document_template.file:
            logger.error(f"[서류생성] 파일 없음 - 템플릿 ID: {document_template.id}")
            return JsonResponse({
                'success': False,
                'error': '서류 템플릿 파일을 찾을 수 없습니다.'
            }, status=404)
        
        # Cloudinary 또는 로컬 파일 시스템에서 파일 가져오기
        import tempfile
        import requests
        
        # CloudinaryField인 경우 직접 URL 사용
        if hasattr(document_template.file, 'public_id'):
            # CloudinaryField - public_id와 URL 직접 사용
            public_id = document_template.file.public_id
            file_url = document_template.file.url
            
            # URL이 이미 http(s)로 시작하면 바로 다운로드
            if file_url.startswith('http://') or file_url.startswith('https://'):
                response = requests.get(file_url)
                if response.status_code != 200:
                    logger.error(f"[서류생성] 다운로드 실패 - 상태코드: {response.status_code}")
                    return JsonResponse({
                        'success': False,
                        'error': '서류 템플릿 파일을 다운로드할 수 없습니다.'
                    }, status=500)
                
                # 임시 파일에 저장
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp_file:
                    tmp_file.write(response.content)
                    template_file_path = tmp_file.name
            else:
                logger.error(f"[서류생성] Cloudinary URL이 상대 경로: {file_url}")
                return JsonResponse({
                    'success': False,
                    'error': '서류 템플릿 파일 URL이 올바르지 않습니다.'
                }, status=500)
        else:
            # FileField - 로컬 파일 경로 사용
            template_file_path = document_template.file.path
        
        # 납품 품목 조회
        delivery_items = DeliveryItem.objects.filter(schedule=schedule).select_related('product')
        
        # 원본 파일 확장자 확인 (CloudinaryField는 public_id 사용)
        if hasattr(document_template.file, 'public_id'):
            # CloudinaryField - URL에서 확장자 추출 (public_id는 확장자가 없을 수 있음)
            file_url = document_template.file.url
            original_ext = os.path.splitext(file_url.split('?')[0])[1].lower()  # 쿼리 스트링 제거 후 확장자 추출
        else:
            # FileField - name 속성 사용
            original_ext = os.path.splitext(document_template.file.name)[1].lower()
        
        
        # 엑셀 파일인 경우 데이터 채우기
        if original_ext in ['.xlsx', '.xls', '.xlsm']:
            try:
                # ZIP 레벨에서 직접 처리 (한글 완벽 보존 + 이미지 보존)
                import shutil
                import zipfile
                
                # template_file_path를 사용 (이미 Cloudinary에서 다운로드되었거나 로컬 경로)
                # 원본 파일을 임시 위치에 복사
                with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp_file:
                    shutil.copy2(template_file_path, tmp_file.name)
                    temp_path = tmp_file.name
                
                
                # 총액 계산
                subtotal = sum([item.unit_price * item.quantity for item in delivery_items], Decimal('0'))
                tax = subtotal * Decimal('0.1')
                total = subtotal + tax
                
                # 총액을 한글로 변환
                def number_to_korean(number):
                    """숫자를 한글로 변환 (예: 11633600 -> 일천백육십삼만삼천육백)"""
                    num = int(number)
                    if num == 0:
                        return '영'
                    
                    units = ['', '만', '억', '조']
                    digits = ['', '일', '이', '삼', '사', '오', '육', '칠', '팔', '구']
                    sub_units = ['', '십', '백', '천']
                    
                    result = []
                    unit_idx = 0
                    
                    while num > 0:
                        segment = num % 10000  # 4자리씩 끊기
                        if segment > 0:
                            segment_str = []
                            
                            # 천의 자리
                            if segment >= 1000:
                                d = segment // 1000
                                if d > 1:
                                    segment_str.append(digits[d])
                                segment_str.append('천')
                                segment %= 1000
                            
                            # 백의 자리
                            if segment >= 100:
                                d = segment // 100
                                if d > 1:
                                    segment_str.append(digits[d])
                                segment_str.append('백')
                                segment %= 100
                            
                            # 십의 자리
                            if segment >= 10:
                                d = segment // 10
                                if d > 1:
                                    segment_str.append(digits[d])
                                segment_str.append('십')
                                segment %= 10
                            
                            # 일의 자리
                            if segment > 0:
                                segment_str.append(digits[segment])
                            
                            # 만/억/조 단위 추가
                            if unit_idx > 0:
                                segment_str.append(units[unit_idx])
                            
                            result.insert(0, ''.join(segment_str))
                        
                        num //= 10000
                        unit_idx += 1
                    
                    return ''.join(result)
                
                total_korean = number_to_korean(total)
                
                # 거래번호 생성 (년-월-일-순번) - 한국 시간대 사용
                import pytz
                korea_tz = pytz.timezone('Asia/Seoul')
                today = timezone.now().astimezone(korea_tz)
                today_str = today.strftime('%Y%m%d')
                
                # 오늘 생성된 모든 서류(견적서 + 거래명세서) 개수 + 1
                from reporting.models import DocumentGenerationLog
                from django.db.models import Count
                
                # 오늘 자정 시각 계산
                today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
                today_end = today.replace(hour=23, minute=59, second=59, microsecond=999999)
                
                # 같은 날짜에 생성된 서류 개수 (견적서 + 거래명세서 모두 포함)
                today_count = DocumentGenerationLog.objects.filter(
                    company=company,
                    created_at__gte=today_start,
                    created_at__lte=today_end
                ).count() + 1
                
                transaction_number = f"{today.strftime('%Y')}-{today.strftime('%m')}-{today.strftime('%d')}-{today_count:03d}"
                
                # 데이터 매핑
                from datetime import timedelta
                
                # 담당자(실무자) 정보
                salesman_name = f"{schedule.user.last_name}{schedule.user.first_name}" if schedule.user.last_name and schedule.user.first_name else schedule.user.username
                
                # 연결된 견적번호 자동 채움
                _linked_quote = schedule.quotes.order_by('-created_at').first()
                _quote_number = _linked_quote.quote_number if _linked_quote else ''

                data_map = {
                    # 기본 정보
                    '년': today.strftime('%Y'),
                    '월': today.strftime('%m'),
                    '일': today.strftime('%d'),
                    '거래번호': transaction_number,
                    
                    # 고객 정보
                    '고객명': schedule.followup.customer_name,
                    '업체명': str(schedule.followup.company) if schedule.followup.company else '',
                    '학교명': str(schedule.followup.company) if schedule.followup.company else '',
                    '부서명': str(schedule.followup.department) if schedule.followup.department else '',
                    '연구실': str(schedule.followup.department) if schedule.followup.department else '',
                    '담당자': schedule.followup.customer_name,
                    '이메일': schedule.followup.email or '',
                    '담당자이메일': schedule.followup.email or '',
                    '연락처': schedule.followup.phone_number or '',
                    '전화번호': schedule.followup.phone_number or '',
                    
                    # 실무자(영업담당자) 정보
                    '실무자': salesman_name,
                    '영업담당자': salesman_name,
                    '담당영업': salesman_name,
                    '영업담당자이메일': schedule.user.email or '',
                    # 날짜 정보
                    '일정날짜': schedule.visit_date.strftime('%Y년 %m월 %d일'),
                    '날짜': schedule.visit_date.strftime('%Y년 %m월 %d일'),
                    '발행일': today.strftime('%Y년 %m월 %d일'),
                    
                    # 회사 정보
                    '회사명': company.name,

                    # 견적 정보 (자동 채움)
                    '견적번호': _quote_number,
                    '메모': schedule.notes or '',
                    
                    # 금액 정보
                    '공급가액': f"{int(subtotal):,}",
                    '소계': f"{int(subtotal):,}",
                    '부가세액': f"{int(tax):,}",
                    '부가세': f"{int(tax):,}",
                    '총액': f"{int(total):,}",
                    '합계': f"{int(total):,}",
                    '총액한글': f"금 {total_korean}원정",
                    '한글금액': f"금 {total_korean}원정",
                }
                
                # 품목 데이터 추가
                for idx, item in enumerate(delivery_items, 1):
                    item_subtotal = item.unit_price * item.quantity
                    # 단위 결정: DeliveryItem에 있으면 사용, 없으면 Product에서, 그것도 없으면 'EA'
                    item_unit = item.unit if item.unit else (item.product.unit if item.product and item.product.unit else 'EA')
                    
                    data_map[f'품목{idx}_이름'] = item.item_name
                    data_map[f'품목{idx}_품목명'] = item.item_name
                    data_map[f'품목{idx}_수량'] = str(item.quantity)
                    data_map[f'품목{idx}_단위'] = item_unit
                    data_map[f'품목{idx}_규격'] = item.product.specification if item.product and item.product.specification else ''
                    data_map[f'품목{idx}_설명'] = item.product.description if item.product and item.product.description else ''
                    data_map[f'품목{idx}_공급가액'] = f"{int(item_subtotal):,}"
                    data_map[f'품목{idx}_단가'] = f"{int(item.unit_price):,}"
                    data_map[f'품목{idx}_부가세액'] = f"{int(item.unit_price * item.quantity * Decimal('0.1')):,}"
                    data_map[f'품목{idx}_금액'] = f"{int(item_subtotal):,}"
                    data_map[f'품목{idx}_총액'] = f"{int(item_subtotal * Decimal('1.1')):,}"
                
                # 1단계: ZIP에서 이미지/차트/미디어 파일 백업
                media_files = {}  # {filename: (ZipInfo, data)}
                
                try:
                    with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                        for file_info in zip_ref.infolist():
                            # 이미지, 차트, 미디어 파일 백업
                            if (file_info.filename.startswith('xl/media/') or 
                                file_info.filename.startswith('xl/drawings/') or 
                                file_info.filename.startswith('xl/charts/')):
                                media_files[file_info.filename] = (file_info, zip_ref.read(file_info.filename))
                except Exception as zip_error:
                    logger.error(f"[서류생성] ZIP 파일 열기 실패: {zip_error}")
                    raise
                
                
                # 2단계: ZIP 레벨에서 직접 sharedStrings.xml 수정 (한글 보존)
                import re
                from datetime import timedelta
                from xml.etree import ElementTree as ET
                
                temp_modified = temp_path.replace('.xlsx', '_modified.xlsx')
                replaced_count = 0
                
                with zipfile.ZipFile(temp_path, 'r') as zip_in:
                    with zipfile.ZipFile(temp_modified, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                        for item in zip_in.infolist():
                            data = zip_in.read(item.filename)
                            
                            # sharedStrings.xml 처리 (한글 변수 치환)
                            if item.filename == 'xl/sharedStrings.xml':
                                try:
                                    xml_str = data.decode('utf-8')
                                    
                                    # 원본 일부 로그 (처음 500자)
                                    
                                    # 변수 치환 (한글 그대로 UTF-8 유지)
                                    for key, value in data_map.items():
                                        pattern = f'{{{{{key}}}}}'
                                        if pattern in xml_str:
                                            # 그대로 치환 (XML은 이미 올바른 형식이므로)
                                            xml_str = xml_str.replace(pattern, str(value))
                                            replaced_count += 1
                                    
                                    # {{유효일+숫자}} 패턴 처리
                                    valid_date_pattern = r'\{\{유효일\+(\d+)\}\}'
                                    valid_matches = re.findall(valid_date_pattern, xml_str)
                                    for days_str in set(valid_matches):
                                        days = int(days_str)
                                        valid_date = schedule.visit_date + timedelta(days=days)
                                        pattern = f'{{{{유효일+{days_str}}}}}'
                                        formatted_date = valid_date.strftime('%Y년 %m월 %d일')
                                        xml_str = xml_str.replace(pattern, formatted_date)
                                        replaced_count += 1
                                    
                                    # {{품목N_xxx}} 패턴 - 품목 없으면 빈칸
                                    item_patterns = re.findall(r'\{\{품목(\d+)_\w+\}\}', xml_str)
                                    for item_pattern in set(item_patterns):
                                        item_num = int(item_pattern)
                                        if item_num > len(delivery_items):
                                            # 해당 품목 변수를 빈칸으로
                                            pattern = r'\{\{품목' + str(item_num) + r'_\w+\}\}'
                                            xml_str = re.sub(pattern, '', xml_str)
                                    
                                    # UTF-8로 인코딩 (한글 그대로)
                                    data = xml_str.encode('utf-8')
                                    
                                    # 수정된 내용 일부 로그
                                except Exception as xml_error:
                                    logger.warning(f"[서류생성] sharedStrings.xml 처리 오류: {xml_error}")
                                    import traceback
                                    logger.error(traceback.format_exc())
                            
                            zip_out.writestr(item, data)
                
                # 원본 삭제하고 수정본으로 교체
                os.unlink(temp_path)
                shutil.move(temp_modified, temp_path)
                
                
                # 3단계: ZIP으로 미디어 파일 복원
                if media_files:
                    temp_output = temp_path.replace('.xlsx', '_with_media.xlsx')
                    
                    with zipfile.ZipFile(temp_path, 'r') as zip_in:
                        with zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                            # 기존 파일 모두 복사
                            for item in zip_in.infolist():
                                # 미디어 파일은 나중에 덮어쓰기
                                if item.filename not in media_files:
                                    data = zip_in.read(item.filename)
                                    zip_out.writestr(item, data)
                            
                            # 백업한 미디어 파일 복원 (원본 ZipInfo 사용)
                            for filename, (file_info, data) in media_files.items():
                                zip_out.writestr(file_info, data)
                    
                    # 원본 삭제하고 새 파일로 교체
                    os.unlink(temp_path)
                    shutil.move(temp_output, temp_path)
                
                
                # 파일명에 사용할 정보 준비
                import pytz
                korea_tz = pytz.timezone('Asia/Seoul')
                today_for_filename = timezone.now().astimezone(korea_tz)
                today_str = today_for_filename.strftime('%Y%m%d')
                company_name = company.name
                customer_company = schedule.followup.company.name if schedule.followup.company else schedule.followup.customer_name
                # 서류 종류 한글명
                doc_type_names = {
                    'quotation': '견적서',
                    'transaction_statement': '거래명세서',
                    'delivery_note': '납품서',
                }
                doc_name = doc_type_names.get(document_type, document_template.name)
                
                # 출력 형식에 따라 저장
                if output_format.lower() == 'pdf':
                    # PDF로 변환 (unoconv 사용)
                    pdf_path = temp_path.replace('.xlsx', '.pdf')
                    
                    try:
                        import subprocess
                        result = subprocess.run([
                            'unoconv',
                            '-f', 'pdf',
                            '-o', pdf_path,
                            temp_path
                        ], capture_output=True, timeout=30, check=True)
                        
                        
                        # PDF 파일 읽기
                        with open(pdf_path, 'rb') as f:
                            output_data = f.read()
                        
                        # 임시 파일 삭제
                        try:
                            os.unlink(temp_path)
                            os.unlink(pdf_path)
                        except:
                            pass
                        
                        # 파일명 및 Content-Type
                        file_name = f"[{company_name}] {customer_company}_{doc_name}({today_str}).pdf"
                        content_type = 'application/pdf'
                        
                    except Exception as pdf_error:
                        logger.warning(f"[서류생성] PDF 변환 실패: {pdf_error}. Excel 파일로 반환합니다.")
                        # PDF 변환 실패 시 Excel 파일 반환
                        with open(temp_path, 'rb') as f:
                            output_data = f.read()
                        
                        try:
                            os.unlink(temp_path)
                        except:
                            pass
                        
                        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        file_name = f"[{company_name}] {customer_company}_{doc_name}({today_str}).xlsx"
                    
                else:
                    # Excel 파일로 저장
                    
                    # 저장된 파일을 읽어서 반환
                    with open(temp_path, 'rb') as f:
                        output_data = f.read()
                    
                    # 임시 파일 삭제
                    try:
                        os.unlink(temp_path)
                    except:
                        pass
                    
                    # Content-Type 결정
                    if original_ext == '.xlsm':
                        content_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
                    elif original_ext == '.xls':
                        content_type = 'application/vnd.ms-excel'
                    else:  # .xlsx
                        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    
                    # 파일명 설정
                    file_name = f"[{company_name}] {customer_company}_{doc_name}({today_str}).xlsx"
                
                encoded_filename = quote(file_name)
                
                # 서류 생성 로그 저장
                from reporting.models import DocumentGenerationLog
                DocumentGenerationLog.objects.create(
                    company=company,
                    document_type=document_type,
                    schedule=schedule,
                    user=request.user,
                    transaction_number=transaction_number,
                    output_format=output_format
                )
                
                # 응답
                response = HttpResponse(
                    output_data,
                    content_type=content_type
                )
                response['Content-Disposition'] = 'attachment'
                response['X-Filename'] = encoded_filename
                response['Access-Control-Expose-Headers'] = 'X-Filename'
                
                return response
                
            except Exception as excel_error:
                logger.error(f"엑셀 처리 오류: {excel_error}")
                import traceback
                logger.error(traceback.format_exc())
                
                # 정리
                try:
                    if 'temp_path' in locals() and os.path.exists(temp_path):
                        os.unlink(temp_path)
                    # Cloudinary에서 다운로드한 임시 파일 정리
                    if 'cloudinary' in file_url or file_url.startswith('http'):
                        if os.path.exists(template_file_path):
                            os.unlink(template_file_path)
                except Exception as cleanup_error:
                    logger.error(f"정리 오류: {cleanup_error}")
                
                return JsonResponse({
                    'success': False,
                    'error': f'엑셀 파일 처리 중 오류가 발생했습니다: {str(excel_error)}'
                }, status=500)
            
            # 성공 시에도 Cloudinary 임시 파일 정리
            finally:
                try:
                    if 'cloudinary' in file_url or file_url.startswith('http'):
                        if 'template_file_path' in locals() and os.path.exists(template_file_path):
                            os.unlink(template_file_path)
                except Exception as cleanup_error:
                    logger.error(f"최종 정리 오류: {cleanup_error}")
            
        else:
            # PDF 등 다른 형식은 그대로 다운로드 (향후 PDF 편집 기능 추가 예정)
            # CloudinaryField는 path 속성이 없으므로 template_file_path 사용
            file_path = template_file_path
            file_name = f"{document_template.name}_{schedule.followup.customer_name}_{timezone.now().strftime('%Y%m%d')}{original_ext}"
            
            content_type_map = {
                '.pdf': 'application/pdf',
            }
            content_type = content_type_map.get(original_ext, 'application/octet-stream')
            
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            encoded_filename = quote(file_name)
            response['Content-Disposition'] = f'attachment; filename="{file_name}"; filename*=UTF-8\'\'{encoded_filename}'
            
            return response
        
    except Exception as e:
        logger.error(f"서류 생성 실패: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== 관리자 필터 API ====================

@login_required
@require_POST
def set_admin_filter(request):
    """
    관리자 필터 설정 (회사/사용자 선택)
    
    POST /reporting/set-admin-filter/
    Body: {
        "company_id": "1",
        "user_id": "2"
    }
    """
    # 관리자만 접근 가능
    if not request.is_admin:
        return JsonResponse({'success': False, 'error': '관리자만 접근 가능합니다.'}, status=403)
    
    try:
        data = json.loads(request.body)
        company_id = data.get('company_id', '')
        user_id = data.get('user_id', '')
        
        # 세션에 저장
        if company_id:
            request.session['admin_selected_company'] = int(company_id)
        else:
            request.session.pop('admin_selected_company', None)
        
        if user_id:
            request.session['admin_selected_user'] = int(user_id)
        else:
            request.session.pop('admin_selected_user', None)
        
        request.session.modified = True
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        logger.error(f"관리자 필터 설정 오류: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_company_users(request, company_id):
    """
    특정 회사의 사용자 목록 반환 (관리자용)
    
    GET /reporting/get-company-users/<company_id>/
    """
    # 관리자만 접근 가능
    if not request.is_admin:
        return JsonResponse({'success': False, 'error': '관리자만 접근 가능합니다.'}, status=403)
    
    try:
        users = User.objects.filter(
            userprofile__company_id=company_id
        ).select_related('userprofile')
        
        user_list = [{
            'id': user.id,
            'name': user.get_full_name() or user.username,
            'username': user.username,
            'role': user.userprofile.get_role_display() if hasattr(user, 'userprofile') else ''
        } for user in users]
        
        return JsonResponse({'success': True, 'users': user_list})
        
    except Exception as e:
        logger.error(f"사용자 목록 조회 오류: {e}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# 법적 문서 뷰
# ============================================

def privacy_policy_view(request):
    """개인정보처리방침 페이지"""
    return render(request, 'reporting/privacy_policy.html')


def terms_of_service_view(request):
    """서비스 이용약관 페이지"""
    return render(request, 'reporting/terms_of_service.html')


@login_required
def customer_records_api(request, followup_id):
    """고객 부서의 전체 납품기록 및 견적기록을 가져오는 API (본인 기록만, 부서 기준)"""
    from decimal import Decimal
    
    try:
        followup = get_object_or_404(FollowUp, pk=followup_id)
        
        # 권한 체크 (같은 회사 고객도 접근 가능)
        if not can_access_followup(request.user, followup):
            return JsonResponse({
                'success': False,
                'error': '접근 권한이 없습니다.'
            }, status=403)
        
        # 본인 고객인지 확인
        is_own_customer = (request.user == followup.user)
        
        # 부서가 있는 경우 부서 내 모든 고객, 없으면 해당 고객만
        if followup.department:
            # 같은 부서의 모든 고객
            department_followups = FollowUp.objects.filter(department=followup.department)
            # 납품 기록 조회 (본인 기록만 - user 필터 + 부서 기준)
            delivery_schedules = Schedule.objects.filter(
                followup__in=department_followups,
                user=request.user,  # 본인 기록만
                activity_type='delivery'
            ).select_related('followup').prefetch_related('delivery_items_set').order_by('-visit_date')
        else:
            # 부서 정보가 없으면 해당 고객만
            delivery_schedules = Schedule.objects.filter(
                followup=followup,
                user=request.user,  # 본인 기록만
                activity_type='delivery'
            ).select_related('followup').prefetch_related('delivery_items_set').order_by('-visit_date')
        
        deliveries = []
        total_delivery_amount = Decimal('0')
        for schedule in delivery_schedules:
            items = []
            schedule_total = Decimal('0')
            
            # 1. 먼저 Schedule에 직접 연결된 delivery_items 확인
            delivery_items = list(schedule.delivery_items_set.all())
            
            # 2. Schedule에 없으면 연결된 History에서 가져오기
            if not delivery_items:
                from reporting.models import History
                related_history = History.objects.filter(schedule=schedule).first()
                if related_history:
                    delivery_items = list(related_history.delivery_items_set.all())
            
            for item in delivery_items:
                if item.total_price:
                    item_total = Decimal(str(item.total_price))
                else:
                    unit_price = item.unit_price if item.unit_price else Decimal('0')
                    item_total = Decimal(str(item.quantity)) * Decimal(str(unit_price)) * Decimal('1.1')
                items.append({
                    'item_name': item.item_name,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price) if item.unit_price else 0,
                    'total_price': float(item_total),
                })
                schedule_total += item_total
            
            total_delivery_amount += schedule_total
            deliveries.append({
                'id': schedule.id,
                'visit_date': schedule.visit_date.strftime('%Y-%m-%d'),
                'customer_name': schedule.followup.customer_name if schedule.followup else '-',  # 고객명 추가
                'status': schedule.get_status_display(),
                'status_code': schedule.status,
                'items': items,
                'total_amount': float(schedule_total),
                'notes': schedule.notes or '',
            })
        
        # 견적 기록 조회 (본인 기록만 - user 필터 + 부서 기준)
        from reporting.models import Quote, QuoteItem
        if followup.department:
            # 같은 부서의 모든 고객
            quote_records = Quote.objects.filter(
                followup__in=department_followups,
                user=request.user  # 본인 기록만
            ).select_related('schedule', 'followup').prefetch_related('items__product').order_by('-created_at')
        else:
            # 부서 정보가 없으면 해당 고객만
            quote_records = Quote.objects.filter(
                followup=followup,
                user=request.user  # 본인 기록만
            ).select_related('schedule', 'followup').prefetch_related('items__product').order_by('-created_at')
        
        quotes = []
        total_quote_amount = Decimal('0')
        for quote in quote_records:
            items = []
            quote_total = Decimal('0')
            for item in quote.items.all():
                if item.subtotal:
                    item_total = Decimal(str(item.subtotal))
                else:
                    item_total = Decimal(str(item.quantity)) * Decimal(str(item.unit_price))
                
                # 제품명 가져오기
                item_name = item.product.name if item.product else '제품명 없음'
                
                items.append({
                    'item_name': item_name,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price),
                    'total_price': float(item_total),
                })
                quote_total += item_total
            
            total_quote_amount += quote_total
            
            # 견적 상태 결정
            status_display = quote.get_status_display() if hasattr(quote, 'get_status_display') else '견적'
            status_code = quote.status if hasattr(quote, 'status') else 'quote'
            
            # 날짜 결정 (schedule이 있으면 schedule의 날짜, 없으면 생성일)
            visit_date = quote.schedule.visit_date.strftime('%Y-%m-%d') if quote.schedule else quote.created_at.strftime('%Y-%m-%d')
            
            quotes.append({
                'id': quote.schedule.id if quote.schedule else quote.id,
                'visit_date': visit_date,
                'customer_name': quote.followup.customer_name if quote.followup else '-',  # 고객명 추가
                'status': status_display,
                'status_code': status_code,
                'items': items,
                'total_amount': float(quote_total),
                'notes': quote.notes or '',
            })
        
        # 응답 데이터 구성
        response_data = {
            'success': True,
            'customer': {
                'id': followup.id,
                'customer_name': followup.customer_name or '-',
                'company_name': followup.company.name if followup.company else '-',
                'department_name': followup.department.name if followup.department else '-',
            },
            'deliveries': deliveries,
            'delivery_count': len(deliveries),
            'total_delivery_amount': float(total_delivery_amount),
            'quotes': quotes,
            'quote_count': len(quotes),
            'total_quote_amount': float(total_quote_amount),
            'is_own_customer': is_own_customer,
        }
        
        # 동료 고객이고 기록이 없는 경우 안내 메시지 추가
        if not is_own_customer and len(deliveries) == 0 and len(quotes) == 0:
            response_data['message'] = '이 고객에 대한 본인의 납품/견적 기록이 없습니다.'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"Customer records API error: {e}")
        return JsonResponse({
            'success': False,
            'error': '고객 기록을 가져오는 중 오류가 발생했습니다.'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def quick_add_customer(request):
    """빠른 고객 등록 API (이메일 발송용)"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        
        customer_name = data.get('customer_name', '').strip()
        email = data.get('email', '').strip()
        company_id = data.get('company_id', '').strip()
        company_name = data.get('company_name', '').strip()
        department_id = data.get('department_id', '').strip()
        department_name = data.get('department_name', '').strip()
        manager = data.get('manager', '').strip()
        phone_number = data.get('phone_number', '').strip()
        priority = data.get('priority', 'scheduled').strip()
        address = data.get('address', '').strip()
        notes = data.get('notes', '').strip()
        
        # 필수 필드 검증
        if not email:
            return JsonResponse({
                'success': False,
                'error': '이메일은 필수입니다.'
            })
        
        if not company_name:
            return JsonResponse({
                'success': False,
                'error': '업체/학교명은 필수입니다.'
            })
        
        # 이메일 유효성 검사
        import re
        email_pattern = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_pattern, email):
            return JsonResponse({
                'success': False,
                'error': '유효한 이메일 주소를 입력해주세요.'
            })
        
        # 중복 체크 (같은 사용자가 같은 이메일로 이미 등록했는지)
        existing = FollowUp.objects.filter(
            user=request.user,
            email=email
        ).first()
        
        if existing:
            return JsonResponse({
                'success': False,
                'error': f'이미 등록된 이메일입니다. (고객: {existing.customer_name})'
            })
        
        # 업체 처리
        company = None
        user_company = request.user.userprofile.company
        
        if company_id:
            # 기존 업체 선택
            try:
                company = Company.objects.get(
                    id=company_id,
                    created_by__userprofile__company=user_company
                )
            except Company.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': '선택한 업체를 찾을 수 없습니다.'
                })
        else:
            # 새 업체 생성
            company = Company.objects.create(
                name=company_name,
                created_by=request.user
            )
            logger.info(f"새 업체 생성: {company_name} by {request.user.username}")
        
        # 부서 처리
        department = None
        if department_name:
            if department_id:
                # 기존 부서 선택
                try:
                    department = Department.objects.get(
                        id=department_id,
                        company=company
                    )
                except Department.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': '선택한 부서를 찾을 수 없습니다.'
                    })
            else:
                # 새 부서 생성
                department = Department.objects.create(
                    name=department_name,
                    company=company,
                    created_by=request.user
                )
                logger.info(f"새 부서 생성: {department_name} by {request.user.username}")
        
        # 팔로우업 생성
        followup = FollowUp.objects.create(
            user=request.user,
            customer_name=customer_name or email.split('@')[0],  # 고객명 없으면 이메일 ID 사용
            email=email,
            company=company,
            department=department,
            manager=manager or customer_name,  # 담당자 없으면 고객명 사용
            phone_number=phone_number,
            priority=priority,
            address=address,
            notes=notes,
            status='active',
            user_company=user_company
        )
        
        logger.info(f"빠른 고객 등록 성공: {followup.customer_name} ({email}) by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'followup_id': followup.id,
            'message': '고객이 등록되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"Quick add customer error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': '고객 등록 중 오류가 발생했습니다.'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def quick_add_company(request):
    """업체 즉시 생성 API"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        company_name = data.get('company_name', '').strip()
        
        if not company_name:
            return JsonResponse({
                'success': False,
                'error': '업체/학교명은 필수입니다.'
            })
        
        # 중복 체크 (같은 회사 내)
        user_company = request.user.userprofile.company
        existing = Company.objects.filter(
            name=company_name,
            created_by__userprofile__company=user_company
        ).first()
        
        if existing:
            return JsonResponse({
                'success': True,
                'company_id': existing.id,
                'company_name': existing.name,
                'message': '이미 등록된 업체입니다.'
            })
        
        # 새 업체 생성
        company = Company.objects.create(
            name=company_name,
            created_by=request.user
        )
        
        logger.info(f"새 업체 생성: {company_name} (ID: {company.id}) by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'company_id': company.id,
            'company_name': company.name,
            'message': '업체가 등록되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"Quick add company error: {e}")
        return JsonResponse({
            'success': False,
            'error': '업체 등록 중 오류가 발생했습니다.'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def quick_add_department(request):
    """부서 즉시 생성 API"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        department_name = data.get('department_name', '').strip()
        company_id = data.get('company_id', '').strip()
        
        if not department_name:
            return JsonResponse({
                'success': False,
                'error': '부서/연구실명은 필수입니다.'
            })
        
        if not company_id:
            return JsonResponse({
                'success': False,
                'error': '업체를 먼저 선택하거나 등록해주세요.'
            })
        
        # 업체 확인
        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': '선택한 업체를 찾을 수 없습니다.'
            })
        
        # 중복 체크
        existing = Department.objects.filter(
            name=department_name,
            company=company
        ).first()
        
        if existing:
            return JsonResponse({
                'success': True,
                'department_id': existing.id,
                'department_name': existing.name,
                'message': '이미 등록된 부서입니다.'
            })
        
        # 새 부서 생성
        department = Department.objects.create(
            name=department_name,
            company=company,
            created_by=request.user
        )
        
        logger.info(f"새 부서 생성: {department_name} (ID: {department.id}) by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'department_id': department.id,
            'department_name': department.name,
            'message': '부서가 등록되었습니다.'
        })
        
    except Exception as e:
        logger.error(f"Quick add department error: {e}")
        return JsonResponse({
            'success': False,
            'error': '부서 등록 중 오류가 발생했습니다.'
        }, status=500)


@login_required
def category_create(request):
    """카테고리 생성"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            parent_id = request.POST.get('parent_id', '').strip()
            color = request.POST.get('color', '#007bff')
            description = request.POST.get('description', '').strip()
            order = int(request.POST.get('order', 0))
            
            if not name:
                return JsonResponse({'success': False, 'error': '카테고리명을 입력하세요.'})
            
            # 상위 카테고리 처리
            parent = None
            if parent_id and parent_id != '':
                try:
                    parent = CustomerCategory.objects.get(id=parent_id, user=request.user)
                except CustomerCategory.DoesNotExist:
                    return JsonResponse({'success': False, 'error': '상위 카테고리를 찾을 수 없습니다.'})
            
            # 중복 체크
            if CustomerCategory.objects.filter(user=request.user, name=name, parent=parent).exists():
                return JsonResponse({'success': False, 'error': '이미 존재하는 카테고리명입니다.'})
            
            # 생성
            category = CustomerCategory.objects.create(
                user=request.user,
                name=name,
                parent=parent,
                color=color,
                description=description,
                order=order
            )
            
            return JsonResponse({
                'success': True,
                'category_id': category.id,
                'message': '카테고리가 생성되었습니다.'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def category_update(request, category_id):
    """카테고리 수정"""
    if request.method == 'POST':
        try:
            category = CustomerCategory.objects.get(id=category_id, user=request.user)
            
            name = request.POST.get('name', '').strip()
            parent_id = request.POST.get('parent_id', '').strip()
            color = request.POST.get('color', '#007bff')
            description = request.POST.get('description', '').strip()
            order = int(request.POST.get('order', 0))
            
            if not name:
                return JsonResponse({'success': False, 'error': '카테고리명을 입력하세요.'})
            
            # 상위 카테고리 처리
            parent = None
            if parent_id and parent_id != '':
                try:
                    parent = CustomerCategory.objects.get(id=parent_id, user=request.user)
                    # 자기 자신을 부모로 설정할 수 없음
                    if parent.id == category_id:
                        return JsonResponse({'success': False, 'error': '자기 자신을 상위 카테고리로 설정할 수 없습니다.'})
                    # 자신의 하위 카테고리를 부모로 설정할 수 없음 (순환 참조 방지)
                    if parent.parent_id == category_id:
                        return JsonResponse({'success': False, 'error': '하위 카테고리를 상위 카테고리로 설정할 수 없습니다.'})
                except CustomerCategory.DoesNotExist:
                    return JsonResponse({'success': False, 'error': '상위 카테고리를 찾을 수 없습니다.'})
            
            # 중복 체크 (본인 제외)
            if CustomerCategory.objects.filter(user=request.user, name=name, parent=parent).exclude(id=category_id).exists():
                return JsonResponse({'success': False, 'error': '이미 존재하는 카테고리명입니다.'})
            
            # 수정
            category.name = name
            category.parent = parent
            category.color = color
            category.description = description
            category.order = order
            category.save()
            
            return JsonResponse({
                'success': True,
                'message': '카테고리가 수정되었습니다.'
            })
            
        except CustomerCategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': '카테고리를 찾을 수 없습니다.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def department_assign_category(request, department_id):
    """부서에 카테고리 빠르게 할당"""
    if request.method == 'POST':
        try:
            department = Department.objects.get(id=department_id)
            
            # 권한 체크 - 부서에 속한 고객들의 담당자 확인
            followups = FollowUp.objects.filter(department=department)
            if followups.exists():
                # 첫 번째 고객의 담당자 권한으로 체크
                first_followup = followups.first()
                if not can_modify_user_data(request.user, first_followup.user):
                    return JsonResponse({'success': False, 'error': '수정 권한이 없습니다.'})
            
            category_id = request.POST.get('category_id', '').strip()
            
            if category_id and category_id != '':
                try:
                    category = CustomerCategory.objects.get(id=category_id, user=request.user)
                    department.category = category
                except CustomerCategory.DoesNotExist:
                    return JsonResponse({'success': False, 'error': '카테고리를 찾을 수 없습니다.'})
            else:
                department.category = None
            
            department.save()
            
            return JsonResponse({
                'success': True,
                'message': '부서에 카테고리가 할당되었습니다.',
                'category_name': department.category.get_full_path() if department.category else '없음',
                'category_color': department.category.color if department.category else '#6c757d'
            })
            
        except Department.DoesNotExist:
            return JsonResponse({'success': False, 'error': '부서를 찾을 수 없습니다.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def category_delete(request, category_id):
    """카테고리 삭제"""
    if request.method == 'POST':
        try:
            category = CustomerCategory.objects.get(id=category_id, user=request.user)
            category_name = category.name
            
            # 이 카테고리를 사용하는 부서들의 카테고리를 None으로 변경
            Department.objects.filter(category=category).update(category=None)
            
            # 카테고리 삭제
            category.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'"{category_name}" 카테고리가 삭제되었습니다.'
            })
            
        except CustomerCategory.DoesNotExist:
            return JsonResponse({'success': False, 'error': '카테고리를 찾을 수 없습니다.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# ============================================
# 부서 메모 API
# ============================================

@login_required
@require_http_methods(["GET", "POST"])
def department_memo_api(request, department_id):
    """부서 메모 조회/저장 API"""
    from reporting.models import DepartmentMemo, Department
    
    department = get_object_or_404(Department, pk=department_id)
    
    if request.method == 'GET':
        memo = department.memos.order_by('-updated_at').first()
        if memo:
            return JsonResponse({
                'success': True,
                'content': memo.content,
                'updated_at': memo.updated_at.strftime('%Y-%m-%d %H:%M'),
                'updated_by': memo.created_by.get_full_name() or memo.created_by.username,
            })
        return JsonResponse({'success': True, 'content': '', 'updated_at': None, 'updated_by': None})
    
    # POST: 저장 (기존 메모 업데이트 또는 신규 생성)
    content = request.POST.get('content', '').strip()
    
    memo = department.memos.order_by('-updated_at').first()
    if memo:
        memo.content = content
        memo.created_by = request.user
        memo.save()
    else:
        memo = DepartmentMemo.objects.create(
            department=department,
            content=content,
            created_by=request.user,
        )
    
    return JsonResponse({
        'success': True,
        'content': memo.content,
        'updated_at': memo.updated_at.strftime('%Y-%m-%d %H:%M'),
        'updated_by': request.user.get_full_name() or request.user.username,
    })


# ============================================
# 주간보고
# ============================================

@login_required
def weekly_report_list(request):
    """주간보고 목록"""
    user = request.user
    profile = getattr(user, 'userprofile', None)
    company = profile.company if profile else None

    # 같은 회사 사람들 (관리자는 전체, 영업사원은 자기 것만)
    if profile and profile.role in ['admin', 'superadmin', 'manager']:
        if company:
            users = User.objects.filter(userprofile__company=company, is_active=True)
        else:
            users = User.objects.filter(is_active=True)
        reports = WeeklyReport.objects.filter(user__in=users).select_related('user')
    else:
        reports = WeeklyReport.objects.filter(user=user)

    # 연도/월 필터
    year = request.GET.get('year')
    month = request.GET.get('month')
    target_user = request.GET.get('user_id')
    if year:
        reports = reports.filter(week_start__year=year)
    if month:
        reports = reports.filter(week_start__month=month)
    if target_user:
        reports = reports.filter(user_id=target_user)

    reports = reports.order_by('-week_start')[:60]

    # 소속 동료 목록 (필터용)
    if company:
        colleagues = User.objects.filter(userprofile__company=company, is_active=True)
    else:
        colleagues = User.objects.none()

    return render(request, 'reporting/weekly_report/list.html', {
        'reports': reports,
        'colleagues': colleagues,
        'selected_year': year,
        'selected_month': month,
        'selected_user': target_user,
        'years': [2024, 2025, 2026, 2027],
        'months': list(range(1, 13)),
    })


# ─── 주간보고 리치 텍스트 헬퍼 ──────────────────────────────────────────────
def _render_report_field(text: str) -> str:
    """주간보고 필드를 안전한 HTML로 변환 (뷰 내부 전용)."""
    from reporting.utils_html import render_report_field
    return render_report_field(text)


@login_required
def weekly_report_create(request):
    """주간보고 작성"""
    import datetime
    from reporting.utils_html import sanitize_html
    today = datetime.date.today()
    # 이번 주 월~금 기본값
    monday = today - datetime.timedelta(days=today.weekday())
    friday = monday + datetime.timedelta(days=4)

    if request.method == 'POST':
        week_start_str = request.POST.get('week_start')
        week_end_str = request.POST.get('week_end')
        title = request.POST.get('title', '').strip()
        # 리치 텍스트 에디터가 HTML을 전송하므로 서버 사이드 정화 적용
        activity_notes = sanitize_html(request.POST.get('activity_notes', '').strip())
        quote_delivery_notes = sanitize_html(request.POST.get('quote_delivery_notes', '').strip())
        other_notes = sanitize_html(request.POST.get('other_notes', '').strip())

        try:
            week_start = datetime.date.fromisoformat(week_start_str)
            week_end = datetime.date.fromisoformat(week_end_str)
        except (TypeError, ValueError):
            messages.error(request, '날짜 형식이 올바르지 않습니다.')
            return redirect('reporting:weekly_report_create')

        if not title:
            title = f"{week_start.strftime('%Y년 %m월 %d일')} 주간보고"

        report, created = WeeklyReport.objects.update_or_create(
            user=request.user,
            week_start=week_start,
            defaults={
                'week_end': week_end,
                'title': title,
                'activity_notes': activity_notes,
                'quote_delivery_notes': quote_delivery_notes,
                'other_notes': other_notes,
            }
        )
        messages.success(request, '주간보고가 저장되었습니다.')
        return redirect('reporting:weekly_report_detail', pk=report.pk)

    # GET: 이번 주 기존 보고서가 있으면 수정 폼으로 이동
    existing = WeeklyReport.objects.filter(user=request.user, week_start=monday).first()
    if existing:
        return redirect('reporting:weekly_report_edit', pk=existing.pk)

    # 이번 주 일정 자동 로드 (참고용)
    schedules = Schedule.objects.filter(
        user=request.user,
        visit_date__gte=monday,
        visit_date__lte=friday,
    ).select_related('followup', 'followup__company', 'followup__department').order_by('visit_date')

    profile = getattr(request.user, 'userprofile', None)
    return render(request, 'reporting/weekly_report/form.html', {
        'week_start': monday,
        'week_end': friday,
        'schedules': schedules,
        'is_edit': False,
        'can_use_ai': profile and profile.can_use_ai,
    })


@login_required
def weekly_report_edit(request, pk):
    """주간보고 수정"""
    import datetime
    report = get_object_or_404(WeeklyReport, pk=pk, user=request.user)

    if request.method == 'POST':
        week_start_str = request.POST.get('week_start')
        week_end_str = request.POST.get('week_end')
        try:
            week_start = datetime.date.fromisoformat(week_start_str)
            week_end = datetime.date.fromisoformat(week_end_str)
        except (TypeError, ValueError):
            messages.error(request, '날짜 형식이 올바르지 않습니다.')
            return redirect('reporting:weekly_report_edit', pk=pk)

        title = request.POST.get('title', '').strip() or f"{week_start.strftime('%Y년 %m월 %d일')} 주간보고"
        from reporting.utils_html import sanitize_html
        report.week_start = week_start
        report.week_end = week_end
        report.title = title
        # 리치 텍스트 에디터가 HTML을 전송하므로 서버 사이드 정화 적용
        report.activity_notes = sanitize_html(request.POST.get('activity_notes', '').strip())
        report.quote_delivery_notes = sanitize_html(request.POST.get('quote_delivery_notes', '').strip())
        report.other_notes = sanitize_html(request.POST.get('other_notes', '').strip())
        report.save()
        messages.success(request, '주간보고가 수정되었습니다.')
        return redirect('reporting:weekly_report_detail', pk=report.pk)

    schedules = Schedule.objects.filter(
        user=request.user,
        visit_date__gte=report.week_start,
        visit_date__lte=report.week_end,
    ).select_related('followup', 'followup__company', 'followup__department').order_by('visit_date')

    edit_profile = getattr(request.user, 'userprofile', None)
    return render(request, 'reporting/weekly_report/form.html', {
        'report': report,
        'week_start': report.week_start,
        'week_end': report.week_end,
        'schedules': schedules,
        'is_edit': True,
        'can_use_ai': edit_profile and edit_profile.can_use_ai,
    })


@login_required
def weekly_report_detail(request, pk):
    """주간보고 상세/출력"""
    report = get_object_or_404(WeeklyReport, pk=pk)
    profile = getattr(request.user, 'userprofile', None)
    company = profile.company if profile else None

    # 열람 권한: 본인 또는 같은 회사 관리자
    if report.user != request.user:
        if not (profile and profile.role in ['admin', 'superadmin', 'manager'] and 
                getattr(getattr(report.user, 'userprofile', None), 'company', None) == company):
            raise Http404

    return render(request, 'reporting/weekly_report/detail.html', {
        'report': report,
        'is_manager': profile and profile.role in ['admin', 'superadmin', 'manager'],
        # 리치 텍스트 / 레거시 플레인텍스트 모두 안전하게 렌더링된 HTML
        'activity_notes_html': _render_report_field(report.activity_notes),
        'quote_delivery_notes_html': _render_report_field(report.quote_delivery_notes),
        'other_notes_html': _render_report_field(report.other_notes),
    })


@login_required
def weekly_report_delete(request, pk):
    """주간보고 삭제"""
    report = get_object_or_404(WeeklyReport, pk=pk, user=request.user)
    if request.method == 'POST':
        report.delete()
        messages.success(request, '주간보고가 삭제되었습니다.')
        return redirect('reporting:weekly_report_list')
    return redirect('reporting:weekly_report_detail', pk=pk)


@login_required
def weekly_report_load_schedules(request):
    """AJAX: 선택 기간 일정 반환 (주간보고 폼에서 참고용) — 카테고리 분류 + 연결 History 포함"""
    import datetime
    week_start_str = request.GET.get('week_start')
    week_end_str = request.GET.get('week_end')
    try:
        week_start = datetime.date.fromisoformat(week_start_str)
        week_end = datetime.date.fromisoformat(week_end_str)
    except (TypeError, ValueError):
        return JsonResponse({'error': '날짜 오류'}, status=400)

    # Schedule.activity_type → 카테고리 매핑
    QUOTE_DELIVERY_TYPES = {'quote', 'delivery'}
    ACTIVITY_TYPES = {'customer_meeting', 'service'}

    from django.db.models import Prefetch

    schedules = Schedule.objects.filter(
        user=request.user,
        visit_date__gte=week_start,
        visit_date__lte=week_end,
    ).select_related(
        'followup', 'followup__company', 'followup__department'
    ).prefetch_related(
        # 이 일정에 직접 연결된 History (History.schedule FK)
        Prefetch(
            'histories',
            queryset=History.objects.filter(
                user=request.user,
                parent_history__isnull=True,
            ).order_by('-created_at'),
            to_attr='linked_histories',
        ),
        # 이 일정에 연결된 Quote
        Prefetch(
            'quotes',
            queryset=Quote.objects.order_by('-created_at'),
            to_attr='linked_quotes',
        ),
    ).order_by('visit_date')

    # 카테고리별 목록
    activity_list = []
    quote_delivery_list = []

    def _history_snippet(h):
        """히스토리에서 핵심 내용 추출"""
        parts = []
        if h.meeting_situation:
            parts.append(h.meeting_situation)
        if h.meeting_confirmed_facts:
            parts.append(h.meeting_confirmed_facts)
        if h.content and not parts:
            parts.append(h.content)
        if h.next_action:
            nd = f" ({h.next_action_date.strftime('%m/%d')})" if h.next_action_date else ''
            parts.append(f"다음 액션: {h.next_action}{nd}")
        return ' / '.join(parts)[:200] if parts else ''

    for s in schedules:
        fu = s.followup
        base = {
            'id': s.pk,
            'date': s.visit_date.strftime('%m/%d'),
            'weekday': ['월', '화', '수', '목', '금', '토', '일'][s.visit_date.weekday()],
            'customer': fu.customer_name or '-',
            'company': str(fu.company) if fu.company else '',
            'department': str(fu.department) if fu.department else '',
            'manager': fu.manager or '',
            'activity_type': s.activity_type,
            'activity_type_display': s.get_activity_type_display(),
            'notes': s.notes or '',
            'status': s.status,
            # 연결된 히스토리 (직접 FK)
            'histories': [
                {
                    'id': h.pk,
                    'type': h.get_action_type_display(),
                    'snippet': _history_snippet(h),
                    'next_action': h.next_action or '',
                    'next_action_date': h.next_action_date.strftime('%m/%d') if h.next_action_date else '',
                }
                for h in (s.linked_histories or [])
            ],
            # 연결된 견적
            'quotes': [
                {
                    'number': q.quote_number,
                    'stage': q.get_stage_display(),
                    'amount': f'{int(q.total_amount):,}원' if q.total_amount else '',
                    'probability': q.probability,
                }
                for q in (s.linked_quotes or [])
            ],
        }

        if s.activity_type in QUOTE_DELIVERY_TYPES:
            quote_delivery_list.append(base)
        else:
            # customer_meeting, service, 기타 → 영업활동
            activity_list.append(base)

    # 기존 flat 형식도 유지 (backward compatibility)
    flat_data = []
    for s_obj in list(schedules):
        fu = s_obj.followup
        flat_data.append({
            'date': s_obj.visit_date.strftime('%m/%d'),
            'customer': fu.customer_name or '-',
            'company': str(fu.company) if fu.company else '',
            'department': str(fu.department) if fu.department else '',
            'activity_type': s_obj.get_activity_type_display(),
            'notes': s_obj.notes or '',
        })

    return JsonResponse({
        'schedules': flat_data,           # 기존 형식 유지
        'categorized': {
            'activity': activity_list,     # 영업활동
            'quote_delivery': quote_delivery_list,  # 견적/납품
        },
    })



@login_required
def weekly_report_ai_draft(request):
    """AJAX GET: 해당 주 활동 기반 AI 주간보고 초안 생성"""
    import datetime as _dt
    profile = getattr(request.user, 'userprofile', None)
    if not (profile and profile.can_use_ai):
        return JsonResponse({'error': 'AI 기능 사용 권한이 없습니다.'}, status=403)

    week_start_str = request.GET.get('week_start')
    week_end_str = request.GET.get('week_end')
    try:
        week_start = _dt.date.fromisoformat(week_start_str)
        week_end = _dt.date.fromisoformat(week_end_str)
    except (TypeError, ValueError):
        return JsonResponse({'error': '날짜 형식 오류'}, status=400)

    try:
        from ai_chat.services import generate_weekly_report_draft
        result = generate_weekly_report_draft(request.user, week_start, week_end)
        if result is None:
            return JsonResponse({'error': '해당 기간에 활동 기록이 없어 초안을 생성할 수 없습니다.'}, status=404)
        return JsonResponse({'draft': result})
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        logger.error(f"weekly_report_ai_draft error: {e}")
        return JsonResponse({'error': 'AI 초안 생성 중 오류가 발생했습니다.'}, status=500)


@login_required
def weekly_report_manager_comment(request, pk):
    """POST: 관리자 코멘트 저장"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    profile = getattr(request.user, 'userprofile', None)
    if not (profile and profile.role in ['admin', 'superadmin', 'manager']):
        return JsonResponse({'error': '권한 없음'}, status=403)

    report = get_object_or_404(WeeklyReport, pk=pk)
    # 같은 회사 소속인지 확인
    company = profile.company if profile else None
    if company:
        report_company = getattr(getattr(report.user, 'userprofile', None), 'company', None)
        if report_company != company:
            return JsonResponse({'error': '권한 없음'}, status=403)

    comment = request.POST.get('manager_comment', '').strip()
    report.manager_comment = comment
    report.reviewed_by = request.user
    report.reviewed_at = timezone.now()
    report.save(update_fields=['manager_comment', 'reviewed_by', 'reviewed_at'])
    return JsonResponse({
        'ok': True,
        'reviewer': request.user.get_full_name() or request.user.username,
        'reviewed_at': report.reviewed_at.strftime('%Y-%m-%d %H:%M'),
        'comment': comment,
    })


# ============ Phase 6: 분석 보고서 뷰들 ============

@login_required
def analytics_dashboard_view(request):
    """
    영업 분석 보고서 대시보드
    - admin/manager: 전체 또는 특정 영업사원 필터 가능
    - salesman: 본인 데이터만 조회
    """
    from django.db.models import Max, Min
    import datetime

    user_profile = get_user_profile(request.user)
    is_manager_or_admin = user_profile.role in ('admin', 'manager')

    today = timezone.now().date()

    # ─── 날짜 범위 필터 ───
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    try:
        date_from = datetime.date.fromisoformat(date_from_str) if date_from_str else (today - datetime.timedelta(days=30))
    except ValueError:
        date_from = today - datetime.timedelta(days=30)
    try:
        date_to = datetime.date.fromisoformat(date_to_str) if date_to_str else today
    except ValueError:
        date_to = today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    # ─── 영업사원 필터 ───
    # admin/manager만 다른 사용자 데이터 조회 가능
    selected_user_id = request.GET.get('user_id', '')
    filter_users = None  # None = 자기 자신

    salesperson_list = []
    if is_manager_or_admin:
        # 같은 회사의 영업사원 목록
        if user_profile.company:
            salesperson_list = User.objects.filter(
                userprofile__company=user_profile.company,
                userprofile__role='salesman',
                is_active=True,
            ).select_related('userprofile').order_by('last_name', 'first_name', 'username')
        else:
            salesperson_list = User.objects.filter(
                userprofile__role='salesman',
                is_active=True,
            ).select_related('userprofile').order_by('last_name', 'first_name', 'username')

        if selected_user_id:
            try:
                uid = int(selected_user_id)
                filter_users = User.objects.filter(pk=uid)
            except (ValueError, TypeError):
                filter_users = None
        # filter_users=None → 전체 조회 (admin/manager)
        if filter_users is None:
            # 전체 영업사원
            if user_profile.company:
                filter_users = User.objects.filter(
                    userprofile__company=user_profile.company,
                    userprofile__role='salesman',
                    is_active=True,
                )
            else:
                filter_users = User.objects.filter(
                    userprofile__role='salesman',
                    is_active=True,
                )
    else:
        # salesman → 본인만
        filter_users = User.objects.filter(pk=request.user.pk)

    # ─── 기본 쿼리셋 ───
    histories_qs = History.objects.filter(
        user__in=filter_users,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
        parent_history__isnull=True,  # 댓글 메모 제외
    )
    followups_qs = FollowUp.objects.filter(user__in=filter_users)

    # ─── 요약 카드 ───
    total_histories = histories_qs.count()

    # 완료된 follow-up (next_action_date 있고 reviewed_at 있는 경우)
    completed_followups = histories_qs.filter(
        next_action_date__isnull=False,
        reviewed_at__isnull=False,
    ).count()

    # 지연된 follow-up (next_action_date가 오늘 이전이고 reviewed_at 없음)
    overdue_followups = History.objects.filter(
        user__in=filter_users,
        next_action_date__lt=today,
        next_action_date__isnull=False,
        reviewed_at__isnull=True,
        parent_history__isnull=True,
    ).count()

    # 예정된 follow-up (next_action_date가 오늘 이후)
    upcoming_followups = History.objects.filter(
        user__in=filter_users,
        next_action_date__gte=today,
        next_action_date__isnull=False,
        reviewed_at__isnull=True,
        parent_history__isnull=True,
    ).count()

    # 활성 파이프라인 항목
    active_pipeline = followups_qs.filter(
        status='active',
        pipeline_stage__in=['potential', 'contact', 'quote', 'negotiation'],
    ).count()

    # ─── 영업사원별 활동 보고서 ───
    activity_report = []
    report_users = filter_users.select_related('userprofile') if filter_users else []
    for u in report_users:
        user_histories = histories_qs.filter(user=u)
        user_hist_all = History.objects.filter(
            user=u,
            parent_history__isnull=True,
        )
        last_activity = user_histories.aggregate(last=Max('created_at'))['last']
        overdue_count = user_hist_all.filter(
            next_action_date__lt=today,
            next_action_date__isnull=False,
            reviewed_at__isnull=True,
        ).count()
        followup_count = FollowUp.objects.filter(user=u, status='active').count()
        activity_report.append({
            'user': u,
            'history_count': user_histories.count(),
            'followup_count': followup_count,
            'overdue_count': overdue_count,
            'last_activity': last_activity,
        })
    # 활동 수 내림차순 정렬
    activity_report.sort(key=lambda x: x['history_count'], reverse=True)

    # ─── 고객 활동 보고서 ───
    # 기간 내 활동이 있는 거래처 (최대 50개)
    active_followup_ids = histories_qs.values_list('followup_id', flat=True).distinct()
    customer_report = []
    for fup in FollowUp.objects.filter(
        pk__in=active_followup_ids,
    ).select_related('company', 'user').order_by('-updated_at')[:50]:
        last_hist = History.objects.filter(
            followup=fup,
            parent_history__isnull=True,
        ).aggregate(last=Max('created_at'), next_date=Max('next_action_date'))
        customer_report.append({
            'followup': fup,
            'last_activity': last_hist['last'],
            'next_action_date': last_hist['next_date'],
        })

    # ─── 파이프라인 단계별 현황 ───
    pipeline_stage_labels = dict(FollowUp.PIPELINE_STAGE_CHOICES)
    stage_order = ['potential', 'contact', 'quote', 'negotiation', 'won', 'lost']
    pipeline_summary = []
    for stage in stage_order:
        cnt = followups_qs.filter(pipeline_stage=stage).count()
        pipeline_summary.append({
            'stage': stage,
            'label': pipeline_stage_labels.get(stage, stage),
            'count': cnt,
        })

    # 파이프라인 바 차트 최대값 (0 나눗셈 방지)
    max_pipeline_count = max((item['count'] for item in pipeline_summary), default=1) or 1

    context = {
        'user_profile': user_profile,
        'is_manager_or_admin': is_manager_or_admin,
        'salesperson_list': salesperson_list,
        'selected_user_id': selected_user_id,
        'date_from': date_from,
        'date_to': date_to,
        'today': today,
        # 요약 카드
        'total_histories': total_histories,
        'completed_followups': completed_followups,
        'overdue_followups': overdue_followups,
        'upcoming_followups': upcoming_followups,
        'active_pipeline': active_pipeline,
        # 보고서
        'activity_report': activity_report,
        'customer_report': customer_report,
        'pipeline_summary': pipeline_summary,
        'max_pipeline_count': max_pipeline_count,
    }
    return render(request, 'reporting/analytics_dashboard.html', context)


def _get_activity_export_date_range(request, today):
    """날짜 파라미터 파싱 헬퍼 — activity/pipeline export 공통 사용"""
    import datetime as dt
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    try:
        date_from = dt.date.fromisoformat(date_from_str) if date_from_str else (today - dt.timedelta(days=30))
    except ValueError:
        date_from = today - dt.timedelta(days=30)
    try:
        date_to = dt.date.fromisoformat(date_to_str) if date_to_str else today
    except ValueError:
        date_to = today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _build_activity_rows(user_profile, date_from, date_to, today):
    """
    activity export 공통 데이터 빌드.
    반환: (headers, rows)  — rows는 list-of-list
    """
    from django.db.models import Prefetch, OuterRef, Subquery, DecimalField
    from django.db.models.functions import Coalesce

    # 회사 범위 필터
    if user_profile.company:
        histories_qs = History.objects.filter(
            user__userprofile__company=user_profile.company,
        )
    else:
        histories_qs = History.objects.all()

    histories_qs = (
        histories_qs
        .filter(
            parent_history__isnull=True,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        )
        .select_related(
            'user', 'followup', 'followup__company', 'followup__department',
            'reviewer',
        )
        .prefetch_related(
            Prefetch(
                'followup__quotes',
                queryset=Quote.objects.order_by('-quote_date'),
                to_attr='_latest_quotes',
            ),
            Prefetch(
                'followup__prepayments',
                queryset=Prepayment.objects.filter(status='active').order_by('-payment_date'),
                to_attr='_active_prepayments',
            ),
        )
        .order_by('user__last_name', 'user__first_name', '-created_at')
    )

    HEADERS = [
        '활동일', '영업사원', '거래처', '부서/연구실', '담당자',
        '활동유형', '내용요약', '미팅상황', '다음액션', '다음예정일',
        '지연여부', '지연일수', '파이프라인단계',
        '견적제출여부', '최근견적금액(원)', '납품금액(원)', '납품품목',
        '선결제잔액(원)', '선결제최근입금일',
        '관리자검토', '검토관리자',
    ]

    rows = []
    for h in histories_qs:
        fu = h.followup

        # 활동일: meeting_date 우선, delivery_date, created_at 순
        activity_date = (
            h.meeting_date or h.delivery_date
            or h.created_at.date()
        )

        salesperson   = h.user.get_full_name() or h.user.username
        customer_name = fu.customer_name if fu else '-'
        company_name  = (fu.company.name if fu and fu.company else '-')
        dept_name     = (fu.department.name if fu and fu.department else '-')
        manager_name  = (fu.manager or '-') if fu else '-'

        # 내용 요약 (미팅 구조화 필드 or content)
        summary = (h.meeting_situation or h.content or '').strip()
        if len(summary) > 120:
            summary = summary[:120] + '…'

        situation = (h.meeting_situation or '').strip()
        if len(situation) > 120:
            situation = situation[:120] + '…'

        next_action      = (h.next_action or h.meeting_next_action or '').strip()
        next_action_date = h.next_action_date.strftime('%Y-%m-%d') if h.next_action_date else '-'

        # 지연 여부
        is_overdue   = bool(h.next_action_date and h.next_action_date < today and not h.reviewed_at)
        overdue_days = (today - h.next_action_date).days if is_overdue else 0
        overdue_str  = '지연' if is_overdue else ''

        # 파이프라인
        pipeline_stage = '-'
        if fu:
            stage_labels = dict(FollowUp.PIPELINE_STAGE_CHOICES)
            pipeline_stage = stage_labels.get(fu.pipeline_stage, fu.pipeline_stage)

        # 견적 정보 (최신 1건)
        quote_submitted = '-'
        quote_amount    = '-'
        if fu and hasattr(fu, '_latest_quotes') and fu._latest_quotes:
            lq = fu._latest_quotes[0]
            stage_label_map = dict(Quote.STAGE_CHOICES)
            quote_submitted = stage_label_map.get(lq.stage, lq.stage)
            quote_amount    = int(lq.total_amount) if lq.total_amount else 0

        # 납품 정보
        delivery_amount = int(h.delivery_amount) if h.delivery_amount is not None else '-'
        delivery_items  = (h.delivery_items or '').replace('\n', ' / ').strip() or '-'

        # 선결제 정보 (활성 잔액 합산)
        prepay_balance  = '-'
        prepay_last_date = '-'
        if fu and hasattr(fu, '_active_prepayments') and fu._active_prepayments:
            total_bal = sum(int(p.balance) for p in fu._active_prepayments)
            prepay_balance   = total_bal
            prepay_last_date = fu._active_prepayments[0].payment_date.strftime('%Y-%m-%d')

        # 관리자 검토
        reviewed_str = '검토완료' if h.reviewed_at else '미검토'
        reviewer_str = (h.reviewer.get_full_name() or h.reviewer.username) if h.reviewer else '-'

        rows.append([
            activity_date.strftime('%Y-%m-%d'),
            salesperson,
            customer_name,
            company_name,
            dept_name,
            manager_name,
            h.get_action_type_display(),
            summary,
            situation,
            next_action,
            next_action_date,
            overdue_str,
            overdue_days if is_overdue else '',
            pipeline_stage,
            quote_submitted,
            quote_amount,
            delivery_amount,
            delivery_items,
            prepay_balance,
            prepay_last_date,
            reviewed_str,
            reviewer_str,
        ])

    return HEADERS, rows


def _build_pipeline_rows(user_profile, today):
    """
    pipeline export 공통 데이터 빌드.
    반환: (headers, rows)
    """
    from django.db.models import Max, Sum, Prefetch

    if user_profile.company:
        followups_qs = FollowUp.objects.filter(
            user__userprofile__company=user_profile.company,
        )
    else:
        followups_qs = FollowUp.objects.all()

    followups_qs = (
        followups_qs
        .select_related('user', 'company', 'department')
        .prefetch_related(
            Prefetch(
                'histories',
                queryset=History.objects.filter(parent_history__isnull=True).order_by('-created_at'),
                to_attr='_all_histories',
            ),
            Prefetch(
                'quotes',
                queryset=Quote.objects.order_by('-quote_date'),
                to_attr='_latest_quotes',
            ),
            Prefetch(
                'prepayments',
                queryset=Prepayment.objects.filter(status='active').order_by('-payment_date'),
                to_attr='_active_prepayments',
            ),
        )
        .order_by('pipeline_stage', 'company__name', 'customer_name')
    )

    HEADERS = [
        '거래처', '부서/연구실', '담당자(고객)', '영업사원',
        '파이프라인단계', '고객상태', '고객등급', '우선순위',
        '최근활동일', '다음액션', '다음예정일', '지연여부', '지연일수(일)',
        '최근견적상태', '최근견적금액(원)', '견적성공확률(%)',
        '선결제잔액(원)', '선결제최근입금일',
        '총납품금액(원)',
    ]

    stage_labels   = dict(FollowUp.PIPELINE_STAGE_CHOICES)
    status_labels  = dict(FollowUp.STATUS_CHOICES)
    priority_labels = dict(FollowUp.PRIORITY_CHOICES)
    quote_stage_labels = dict(Quote.STAGE_CHOICES)

    rows = []
    for fu in followups_qs:
        customer_name  = fu.customer_name or '-'
        company_name   = fu.company.name if fu.company else '-'
        dept_name      = fu.department.name if fu.department else '-'
        manager_name   = fu.manager or '-'
        salesperson    = fu.user.get_full_name() or fu.user.username

        pipeline_stage = stage_labels.get(fu.pipeline_stage, fu.pipeline_stage)
        status_str     = status_labels.get(fu.status, fu.status)
        grade_str      = fu.customer_grade or '-'
        priority_str   = priority_labels.get(fu.priority, fu.priority)

        # 최근 활동일
        last_hist = fu._all_histories[0] if (hasattr(fu, '_all_histories') and fu._all_histories) else None
        last_activity_str = '-'
        next_action_str   = '-'
        next_action_date_str = '-'
        overdue_str  = ''
        overdue_days = ''

        if last_hist:
            act_date = last_hist.meeting_date or last_hist.delivery_date or last_hist.created_at.date()
            last_activity_str = act_date.strftime('%Y-%m-%d')
            next_action_str  = (last_hist.next_action or last_hist.meeting_next_action or '').strip()[:100]
            if last_hist.next_action_date:
                next_action_date_str = last_hist.next_action_date.strftime('%Y-%m-%d')
                if last_hist.next_action_date < today and not last_hist.reviewed_at:
                    overdue_str  = '지연'
                    overdue_days = (today - last_hist.next_action_date).days

        # 최근 견적 정보
        quote_stage_str  = '-'
        quote_amount_str = '-'
        quote_prob_str   = '-'
        if hasattr(fu, '_latest_quotes') and fu._latest_quotes:
            lq = fu._latest_quotes[0]
            quote_stage_str  = quote_stage_labels.get(lq.stage, lq.stage)
            quote_amount_str = int(lq.total_amount) if lq.total_amount else 0
            quote_prob_str   = lq.probability

        # 선결제 잔액
        prepay_balance   = '-'
        prepay_last_date = '-'
        if hasattr(fu, '_active_prepayments') and fu._active_prepayments:
            total_bal = sum(int(p.balance) for p in fu._active_prepayments)
            prepay_balance   = total_bal
            prepay_last_date = fu._active_prepayments[0].payment_date.strftime('%Y-%m-%d')

        # 총 납품 금액
        total_delivery = '-'
        if hasattr(fu, '_all_histories'):
            delivery_total = sum(
                int(h.delivery_amount)
                for h in fu._all_histories
                if h.delivery_amount is not None and h.action_type == 'delivery_schedule'
            )
            total_delivery = delivery_total if delivery_total else 0

        rows.append([
            customer_name, dept_name, manager_name, salesperson,
            pipeline_stage, status_str, grade_str, priority_str,
            last_activity_str, next_action_str, next_action_date_str, overdue_str, overdue_days,
            quote_stage_str, quote_amount_str, quote_prob_str,
            prepay_balance, prepay_last_date,
            total_delivery,
        ])

    return HEADERS, rows


# ── Phase 6.5-6: 개선된 activity CSV export ──────────────────────────────────
@login_required
def analytics_activity_csv_export(request):
    """
    활동 내역 상세 CSV 내보내기 (admin/manager 전용)
    활동 단위 행 출력, UTF-8 BOM — 한국어 Excel 호환
    """
    import csv

    user_profile = get_user_profile(request.user)
    if not (user_profile.is_admin() or user_profile.is_manager()):
        return HttpResponseForbidden('접근 권한이 없습니다.')

    today     = timezone.now().date()
    date_from, date_to = _get_activity_export_date_range(request, today)
    headers, rows = _build_activity_rows(user_profile, date_from, date_to, today)

    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="activity_detail_{date_from}_{date_to}.csv"'
    )
    response.write('\ufeff')  # BOM

    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return response


# ── Phase 6.5-6: 개선된 pipeline CSV export ──────────────────────────────────
@login_required
def analytics_pipeline_csv_export(request):
    """
    파이프라인 거래처별 상세 CSV 내보내기 (admin/manager 전용)
    UTF-8 BOM — 한국어 Excel 호환
    """
    import csv

    user_profile = get_user_profile(request.user)
    if not (user_profile.is_admin() or user_profile.is_manager()):
        return HttpResponseForbidden('접근 권한이 없습니다.')

    today = timezone.now().date()
    headers, rows = _build_pipeline_rows(user_profile, today)

    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = (
        f'attachment; filename="pipeline_detail_{today}.csv"'
    )
    response.write('\ufeff')  # BOM

    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return response


# ── Phase 6.5-6: activity XLSX export ────────────────────────────────────────
@login_required
def analytics_activity_xlsx_export(request):
    """
    활동 내역 상세 XLSX 내보내기 (admin/manager 전용)
    헤더 스타일, 첫 행 고정, 지연 행 강조
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io

    user_profile = get_user_profile(request.user)
    if not (user_profile.is_admin() or user_profile.is_manager()):
        return HttpResponseForbidden('접근 권한이 없습니다.')

    today     = timezone.now().date()
    date_from, date_to = _get_activity_export_date_range(request, today)
    headers, rows = _build_activity_rows(user_profile, date_from, date_to, today)

    wb = Workbook()
    ws = wb.active
    ws.title = '활동 내역'

    # 스타일 정의
    header_font   = Font(bold=True, color='FFFFFF', size=10)
    header_fill   = PatternFill(fill_type='solid', fgColor='4F46E5')  # 인디고
    overdue_fill  = PatternFill(fill_type='solid', fgColor='FEE2E2')  # 연한 빨강
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_side     = Side(style='thin', color='D1D5DB')
    border        = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 헤더 행
    ws.row_dimensions[1].height = 22
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border

    # 첫 행 고정
    ws.freeze_panes = 'A2'

    # 지연 열 인덱스 (0-based in rows list: index 11 = '지연여부')
    OVERDUE_IDX = 11  # headers 기준 인덱스 (0-based)

    # 데이터 행
    for row_num, row_data in enumerate(rows, 2):
        is_overdue = str(row_data[OVERDUE_IDX]) == '지연'
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = left_align
            if is_overdue:
                cell.fill = overdue_fill

    # 열 너비 자동 조정
    COL_WIDTHS = [
        12, 10, 16, 18, 14, 12, 12, 30, 30, 30,
        12, 8, 8, 14,
        14, 14, 14, 20,
        14, 12, 8, 12,
    ]
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 제목 정보 시트
    ws_info = wb.create_sheet(title='내보내기 정보')
    ws_info['A1'] = '보고서 기간'
    ws_info['B1'] = f'{date_from} ~ {date_to}'
    ws_info['A2'] = '내보낸 행수'
    ws_info['B2'] = len(rows)
    ws_info['A3'] = '생성일시'
    ws_info['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M')
    ws_info['A4'] = '생성자'
    ws_info['B4'] = request.user.get_full_name() or request.user.username

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="activity_detail_{date_from}_{date_to}.xlsx"'
    )
    return response


# ── Phase 6.5-6: pipeline XLSX export ────────────────────────────────────────
@login_required
def analytics_pipeline_xlsx_export(request):
    """
    파이프라인 거래처별 상세 XLSX 내보내기 (admin/manager 전용)
    헤더 스타일, 첫 행 고정, 지연 행 강조, 단계별 시트 분리
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io

    user_profile = get_user_profile(request.user)
    if not (user_profile.is_admin() or user_profile.is_manager()):
        return HttpResponseForbidden('접근 권한이 없습니다.')

    today = timezone.now().date()
    headers, rows = _build_pipeline_rows(user_profile, today)

    wb = Workbook()

    # 스타일 정의
    header_font   = Font(bold=True, color='FFFFFF', size=10)
    header_fill   = PatternFill(fill_type='solid', fgColor='16A34A')  # 녹색
    overdue_fill  = PatternFill(fill_type='solid', fgColor='FEE2E2')  # 연한 빨강
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin_side     = Side(style='thin', color='D1D5DB')
    border        = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    COL_WIDTHS = [
        16, 16, 12, 12,
        14, 10, 8, 10,
        12, 28, 12, 8, 10,
        14, 14, 12,
        14, 12,
        14,
    ]

    OVERDUE_IDX = 11  # headers 기준 인덱스 (0-based): '지연여부'

    # 단계 순서 (pipeline_stage 값 기준 — rows의 컬럼 4 = '파이프라인단계' 표시값)
    STAGE_ORDER = ['잠재', '접촉/미팅', '견적 제출', '협상', '수주', '실주']
    STAGE_COLORS = {
        '잠재':    '6B7280',
        '접촉/미팅': '3B82F6',
        '견적 제출': 'F59E0B',
        '협상':    'EF4444',
        '수주':    '10B981',
        '실주':    '9CA3AF',
    }

    def _write_sheet(ws, sheet_rows):
        ws.row_dimensions[1].height = 22
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = border
        ws.freeze_panes = 'A2'
        for row_num, row_data in enumerate(sheet_rows, 2):
            is_overdue = str(row_data[OVERDUE_IDX]) == '지연'
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border    = border
                cell.alignment = left_align
                if is_overdue:
                    cell.fill = overdue_fill
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 전체 시트
    ws_all = wb.active
    ws_all.title = '전체'
    _write_sheet(ws_all, rows)

    # 단계별 시트
    for stage_name in STAGE_ORDER:
        stage_rows = [r for r in rows if r[4] == stage_name]
        if not stage_rows:
            continue
        ws_s = wb.create_sheet(title=stage_name)
        fill_color = STAGE_COLORS.get(stage_name, '4F46E5')
        # 단계별로 헤더 색상 변경
        orig_fill = header_fill
        stage_header_fill = PatternFill(fill_type='solid', fgColor=fill_color)
        # 임시로 교체하지 않고 직접 작성
        ws_s.row_dimensions[1].height = 22
        for col_idx, h in enumerate(headers, 1):
            cell = ws_s.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = stage_header_fill
            cell.alignment = center_align
            cell.border    = border
        ws_s.freeze_panes = 'A2'
        for row_num, row_data in enumerate(stage_rows, 2):
            is_overdue = str(row_data[OVERDUE_IDX]) == '지연'
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_s.cell(row=row_num, column=col_idx, value=value)
                cell.border    = border
                cell.alignment = left_align
                if is_overdue:
                    cell.fill = overdue_fill
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws_s.column_dimensions[get_column_letter(col_idx)].width = width

    # 내보내기 정보 시트
    ws_info = wb.create_sheet(title='내보내기 정보')
    ws_info['A1'] = '생성일시'
    ws_info['B1'] = timezone.now().strftime('%Y-%m-%d %H:%M')
    ws_info['A2'] = '생성자'
    ws_info['B2'] = request.user.get_full_name() or request.user.username
    ws_info['A3'] = '전체 건수'
    ws_info['B3'] = len(rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="pipeline_detail_{today}.xlsx"'
    )
    return response
